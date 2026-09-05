# 실사용 문장 배터리 — 사용자가 GUI에서 실제로 친 문장 그대로 돌린다.
#
# 2026-08-18 사용자 제안: "내가 테스트할 때 넣은 문구로, 내 질문 스타일대로 테스트".
# 출처는 chat_log의 GUI 세션(테스트 세션 제외). 문장마다 정답 액션을 정하지 않는다 —
# 실사용 문장은 문맥 의존이 커서 되묻기도 정답일 수 있다. 대신 **절대 어겨선 안 되는
# 것**만 판정한다:
#   ① 조용한 오염 없음(머리글 A1이 문장 텍스트로 덮이지 않음, 예외 없음)
#   ② 결정성(REAL_REPEAT=N이면 N번 같은 결과)
#   ③ 되묻기라면 문구가 비어 있지 않고 무엇으로 이해했는지 드러남
# GUI 조건(workbook_id 없음, approval_id 재개)으로 돈다.
#
#   $env:REAL_REPEAT="3"; $env:EXCEL_LIVE_ENGINE="file"; & $PY scripts\run_real_usage_phrases.py
import asyncio
import json
import os
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.routers.excel_live import (
    ApprovalResponse,
    ExcelLiveCommandRequest,
    post_approval,
    post_command,
)
from office_claw_sidecar.services.excel_live_service import (
    get_excel_live_service,
    invalidate_excel_engine_cache,
)
from office_claw_sidecar.services.llm_service import get_llm_service

def _workspace_root():
    # 이 개발기 절대경로가 박혀 있어 다른 머신에서 전멸했다(2026-09-06 감사 D).
    from office_claw_sidecar.config import get_workspace_root
    return Path(get_workspace_root())


WS = _workspace_root()
WB = WS / "실사용문장_probe.xlsx"
PHRASES = json.loads((Path(__file__).with_name("real_usage_phrases.json")).read_text(encoding="utf-8"))

SEED = [
    ["지역", "주문건수", "출고건수", "정시배송률", "지연건수", "클레임"],
    ["수도권", 10452, 10158, 97.1, 145, 12],
    ["충청권", 3892, 3773, 95.2, 89, 6],
    ["호남권", 3214, 3086, 94.7, 112, 5],
    ["영남권", 6789, 6512, 95.8, 174, 5],
    ["강원제주", 2495, 2383, 92.6, 145, 0],
]


def _fresh_book():
    if WB.exists():
        WB.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "지역성과"
    for row in SEED:
        ws.append(row)
    wb.create_sheet("매출").append(["날짜", "지역", "담당자", "금액"])
    wb.create_sheet("Sales_Data").append(["Order_ID", "Region", "Amount"])
    wb.save(WB)
    wb.close()


async def run_one(idx, text, ctx, llm, rep=0):
    _fresh_book()
    invalidate_excel_engine_cache()
    get_excel_live_service().select_workbook(str(WB))
    t0 = time.time()
    try:
        req = ExcelLiveCommandRequest(message=text, session_id=f"test-real-{idx}-r{rep}", workbook_id=None, approve=False)
        if ctx:
            req = req.model_copy(update={"context_range": ctx})
        resp = await post_command(req, llm)
        if getattr(resp, "approval_required", False):
            resp = await post_approval(
                ApprovalResponse(approval_id=resp.pending_approval.approval_id, approved=True), llm
            )
    except Exception as exc:
        return {"text": text, "ok": False, "why": f"예외 {type(exc).__name__}", "action": "", "reply": str(exc)[:120]}
    r = getattr(resp, "result", None) or {}
    action = str(getattr(resp, "action", ""))
    asked = bool(r.get("ask_follow_up"))
    reply = str(r.get("follow_up_question") or r.get("execution_report") or getattr(resp, "reason", ""))
    why = ""
    if asked and len(reply.strip()) < 6:
        why = "빈 되묻기"
    try:
        wb = load_workbook(WB)
        a1 = wb["지역성과"]["A1"].value
        wb.close()
        if isinstance(a1, str) and a1 != "지역" and len(a1) > 10:
            why = f"A1 오염: {a1[:30]}"
    except Exception as exc:
        why = f"파일 손상 {type(exc).__name__}"
    return {"text": text, "ok": not why, "why": why, "action": action, "asked": asked,
            "reply": reply[:160], "secs": round(time.time() - t0, 1)}


async def main():
    repeat = max(1, int(os.environ.get("REAL_REPEAT", "1")))
    llm = get_llm_service()
    out = []
    for i, ph in enumerate(PHRASES, 1):
        e = await run_one(i, ph["text"], ph.get("context_range"), llm)
        for rep in range(1, repeat):
            # 회차마다 세션을 갈아야 한다 — 같으면 1회차의 되묻기 슬롯이 2회차 답변으로
            # 소비돼 비결정으로 오판된다(2026-08-18 실측: 표 인터뷰 4건).
            again = await run_one(i, ph["text"], ph.get("context_range"), llm, rep=rep)
            if (again["ok"], again["action"], again.get("asked")) != (e["ok"], e["action"], e.get("asked")):
                e = dict(e, ok=False, why=f"비결정: {e['action']} vs {again['action']}")
                break
        flag = "OK  " if e["ok"] else "FAIL"
        tag = " (질문)" if e.get("asked") else ""
        print(f"[{i:3d}] {flag}{tag} {ph['text'][:44]:46s} {e['action'].split('.')[-1][:22]:24s} {e['why'][:36]}")
        out.append(e)
    ok = sum(1 for e in out if e["ok"])
    asked = sum(1 for e in out if e.get("asked"))
    print(f"\n합계 {ok}/{len(out)} · 되묻기 {asked}건")
    Path(__file__).with_name("real_usage_log.json").write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    if WB.exists():
        WB.unlink()


from office_claw_sidecar.services import decision_trace as _dt

# 배터리 턴을 사람이 친 명령과 가를 출처 태그(2026-08-19 로그 감사: 실사용 로그의 source가 전부 비어 있었다).
_dt.source(kind="script", name="real_usage").__enter__()
asyncio.run(main())
