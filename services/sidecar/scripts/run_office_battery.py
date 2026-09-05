"""웬만한 엑셀 업무 전 카테고리 배터리 — 실제 HTTP + 매 케이스 파일 검증.

2026-08-17, 사용자 요구: "웬만한 엑셀 업무, 함수 반영이 다 잘 되도록". 표현 하나를
GUI에서 발견할 때마다 고치는 대신, 카테고리 단위로 쓸어서 구멍을 한꺼번에 찾는다.

원칙:
  - 판정은 응답이 아니라 **파일**로 한다 (CLAUDE.md §3.7 — 자기보고를 믿지 않는다).
  - 케이스마다 새 워크북. 멀티턴은 같은 세션으로 이어 보낸다.
  - 승인 카드는 앱과 같은 경로(/approval + approval_id)로 통과한다.
"""

from __future__ import annotations

import json
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook

BASE = "http://127.0.0.1:19533"
H = {
    "Content-Type": "application/json; charset=utf-8",
    "X-Auth-Token": "dev-token",
    "Authorization": "Bearer dev-token",
}
def _workspace_root():
    # 이 개발기 절대경로가 박혀 있어 다른 머신에서 전멸했다(2026-09-06 감사 D).
    from office_claw_sidecar.config import get_workspace_root
    return Path(get_workspace_root())


ROOT = _workspace_root()


def post(path: str, pl: dict) -> dict:
    req = urllib.request.Request(
        BASE + path, data=json.dumps(pl, ensure_ascii=False).encode(), headers=H
    )
    try:
        with urllib.request.urlopen(req, timeout=240) as r:
            return json.load(r)
    except urllib.error.HTTPError as e:
        try:
            return json.loads(e.read() or b"{}")
        except Exception:
            return {"ok": False, "http_error": e.code}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def turn(msg: str, wb_path: Path, session: str) -> dict:
    pl = {"message": msg, "session_id": session, "workbook_id": str(wb_path), "approve": False}
    b = post("/excel-live/command", pl)
    aid = (b.get("pending_approval") or {}).get("approval_id")
    if b.get("approval_required") and aid:
        b = post("/excel-live/approval", {"approval_id": aid, "approved": True})
    elif b.get("approval_required"):
        b = post("/excel-live/command", {**pl, "approve": True})
    return b


# ── 시드 ──────────────────────────────────────────────────────────────
BASE_ROWS = [
    ["2026-01-01", "서울", "김철수", 120000],
    ["2026-01-02", "경기", "이영희", 85000],
    ["2026-01-03", "부산", "박민수", 143000],
    ["2026-01-04", "서울", "정수진", 98000],
    ["2026-01-05", "대구", "한지원", 67000],
    ["2026-01-06", "서울", "김철수", 155000],
    ["2026-01-07", "경기", "이영희", 72000],
    ["2026-01-08", "부산", "박민수", 110000],
]


def seed_sales(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출"
    ws.append(["날짜", "지역", "담당자", "금액"])
    for r in BASE_ROWS:
        ws.append(r)
    wb.save(path)
    wb.close()


def seed_empty(path: Path) -> None:
    wb = Workbook()
    wb.active.title = "매출"
    wb.save(path)
    wb.close()


def seed_dupes(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출"
    ws.append(["날짜", "지역", "담당자", "금액"])
    for r in [*BASE_ROWS[:4], BASE_ROWS[1], BASE_ROWS[2]]:  # 2건 중복
        ws.append(r)
    wb.save(path)
    wb.close()


# ── 검사 도우미 ────────────────────────────────────────────────────────
def cell(ws, ref):
    return ws[ref].value


def fill_of(ws, ref):
    c = ws[ref]
    if c.fill is None or c.fill.patternType != "solid":
        return ""
    return str(getattr(c.fill.fgColor, "rgb", "") or "")


def has_border(ws, ref):
    b = ws[ref].border
    return any(getattr(b, s) and getattr(b, s).style for s in ("left", "right", "top", "bottom"))


def font_color(ws, ref):
    f = ws[ref].font
    return str(getattr(getattr(f, "color", None), "rgb", "") or "")


def numfmt(ws, ref):
    return str(ws[ref].number_format or "")


# ── 케이스 ────────────────────────────────────────────────────────────
# (카테고리, 라벨, 시드, [턴들], 검사함수(ws, wb) -> (ok, detail))
CASES = [
    # A. 서식
    ("서식", "배경색", seed_sales, ["A1:D1 배경 노란색으로 칠해줘"],
     lambda ws, wb: ("FFFF00" in fill_of(ws, "A1"), f"A1 fill={fill_of(ws,'A1')}")),
    ("서식", "글자색(배경 보존)", seed_sales, ["A1:D1 글자 빨간색으로 바꿔줘"],
     lambda ws, wb: ("FF0000" in font_color(ws, "A1") and "FFFF00" not in fill_of(ws, "A1"),
                     f"font={font_color(ws,'A1')} fill={fill_of(ws,'A1')}")),
    ("서식", "굵게", seed_sales, ["A1:D1 굵게 해줘"],
     lambda ws, wb: (bool(ws["A1"].font and ws["A1"].font.bold), f"bold={ws['A1'].font.bold}")),
    ("서식", "테두리", seed_sales, ["A1:D9 테두리 넣어줘"],
     lambda ws, wb: (has_border(ws, "C5"), f"C5 border={has_border(ws,'C5')}")),
    ("서식", "천단위 콤마", seed_sales, ["금액에 천 단위 콤마 넣어줘"],
     lambda ws, wb: ("#,##0" in numfmt(ws, "D2") and cell(ws, "D2") == 120000,
                     f"fmt={numfmt(ws,'D2')} val={cell(ws,'D2')!r}")),
    ("서식", "소수점 자릿수", seed_sales, ["D2:D9 소수점 둘째 자리까지 보이게 해줘"],
     lambda ws, wb: ("0.00" in numfmt(ws, "D2"), f"fmt={numfmt(ws,'D2')}")),
    # B. 수식
    # D:D도 정답이다 — SUM/AVERAGE는 머리글 텍스트를 무시하므로 값이 같다.
    ("수식", "합계", seed_sales, ["F2에 금액 합계 수식 넣어줘"],
     lambda ws, wb: (bool(__import__("re").match(r"=SUM\(\$?D", str(cell(ws, "F2")))), f"F2={cell(ws,'F2')!r}")),
    ("수식", "평균", seed_sales, ["F3에 금액 평균 구하는 수식 넣어줘"],
     lambda ws, wb: (bool(__import__("re").match(r"=AVERAGE\(\$?D", str(cell(ws, "F3")))), f"F3={cell(ws,'F3')!r}")),
    ("수식", "최댓값", seed_sales, ["F4에 가장 큰 금액 넣어줘"],
     lambda ws, wb: (str(cell(ws, "F4")).startswith("=MAX(D2:D9"), f"F4={cell(ws,'F4')!r}")),
    ("수식", "건수", seed_sales, ["F5에 거래 건수 세는 수식 넣어줘"],
     lambda ws, wb: (str(cell(ws, "F5")).upper().startswith("=COUNT"), f"F5={cell(ws,'F5')!r}")),
    # C. 정렬·필터·중복
    ("정렬필터", "내림차순 정렬", seed_sales, ["금액 기준으로 내림차순 정렬해줘"],
     lambda ws, wb: (cell(ws, "D2") == 155000, f"D2={cell(ws,'D2')!r} (기대 155000)")),
    ("정렬필터", "오름차순 정렬", seed_sales, ["금액 낮은 순으로 정렬해줘"],
     lambda ws, wb: (cell(ws, "D2") == 67000, f"D2={cell(ws,'D2')!r} (기대 67000)")),
    ("정렬필터", "값 필터", seed_sales, ["지역이 서울인 행만 남겨줘"],
     lambda ws, wb: (
         all(v == "서울" for v in [c[0].value for c in ws["B2:B10"]] if v is not None)
         and cell(ws, "B2") == "서울",
         f"B열={[c[0].value for c in ws['B2:B6']]}")),
    ("정렬필터", "중복 제거", seed_dupes, ["중복된 행 지워줘"],
     lambda ws, wb: (
         sum(1 for row in ws.iter_rows(min_row=2, max_col=4) if row[0].value is not None) == 4,
         f"데이터행={sum(1 for row in ws.iter_rows(min_row=2, max_col=4) if row[0].value is not None)} (기대 4)")),
    # D. 데이터 조작
    ("데이터", "셀 입력", seed_sales, ["A12에 합계 라고 입력해줘"],
     lambda ws, wb: (cell(ws, "A12") == "합계", f"A12={cell(ws,'A12')!r}")),
    ("데이터", "찾아 바꾸기", seed_sales, ["서울을 전부 SEOUL로 바꿔줘"],
     lambda ws, wb: (cell(ws, "B2") == "SEOUL" and cell(ws, "B5") == "SEOUL",
                     f"B2={cell(ws,'B2')!r} B5={cell(ws,'B5')!r}")),
    ("데이터", "행 추가 입력", seed_sales, ["A10:D10에 2026-01-09, 인천, 최영호, 88000 입력해줘"],
     lambda ws, wb: (cell(ws, "B10") == "인천" and cell(ws, "D10") in (88000, "88000"),
                     f"B10={cell(ws,'B10')!r} D10={cell(ws,'D10')!r}")),
    # E. 표·템플릿 (멀티턴 포함)
    ("표", "빈 표", seed_empty, ["A1부터 3행 4열 표 만들어줘"],
     lambda ws, wb: (has_border(ws, "A1") or has_border(ws, "B2"), f"A1 border={has_border(ws,'A1')}")),
    ("표", "회의록 템플릿", seed_empty, ["회의록 표 만들어줘", "응 그렇게 해줘"],
     lambda ws, wb: (cell(ws, "A1") == "날짜" and cell(ws, "B1") == "참석자",
                     f"1행={[cell(ws,c+'1') for c in 'ABC']}")),
    ("표", "출석부 일별", seed_empty, ["A1:D13 여기에 출석부 만들어줘", "일별로"],
     lambda ws, wb: (cell(ws, "A1") == "날짜" and cell(ws, "B1") == "이름",
                     f"1행={[cell(ws,c+'1') for c in 'ABCD']}")),
    # F. 피벗·집계
    ("피벗", "지역별 집계", seed_sales, ["지역별 금액 합계 집계표 만들어줘"],
     lambda ws, wb: (
         any("집계" in n or "피벗" in n or "요약" in n for n in wb.sheetnames if n != "매출")
         or cell(ws, "F1") is not None,
         f"sheets={wb.sheetnames}")),
    # G. 멀티턴 연속 서식
    ("멀티턴", "배경→글자", seed_sales, ["A1:D1 배경 남색으로 칠해줘", "글자도 흰색으로"],
     lambda ws, wb: (fill_of(ws, "A1") not in ("", "00000000") and "FFFFFF" in font_color(ws, "A1")
                     and fill_of(ws, "A1")[-6:] != "FFFFFF",
                     f"fill={fill_of(ws,'A1')} font={font_color(ws,'A1')}")),
    ("멀티턴", "합계→평균", seed_sales, ["F2에 금액 합계 수식 넣어줘", "그 아래 칸에는 평균 넣어줘"],
     lambda ws, wb: (str(cell(ws, "F3")).upper().startswith("=AVERAGE"), f"F3={cell(ws,'F3')!r}")),
    ("멀티턴", "정정", seed_sales, ["A12에 서울 입력", "아니 부산으로 바꿔줘"],
     lambda ws, wb: (cell(ws, "A12") == "부산" and cell(ws, "B2") == "서울",
                     f"A12={cell(ws,'A12')!r} B2={cell(ws,'B2')!r}")),
]


def main() -> None:
    only = sys.argv[1] if len(sys.argv) > 1 else ""
    results = []
    for i, (cat, label, seeder, turns, check) in enumerate(CASES):
        if only and only not in f"{cat}/{label}":
            continue
        path = ROOT / f"배터리_{i:02d}.xlsx"
        path.unlink(missing_ok=True)
        seeder(path)
        session = f"bat-{i}"
        bodies = [turn(t, path, session) for t in turns]
        last = bodies[-1]
        asked = bool((last.get("result") or {}).get("ask_follow_up"))
        wb = load_workbook(path)
        ws = wb["매출"] if "매출" in wb.sheetnames else wb.active
        try:
            ok, detail = check(ws, wb)
        except Exception as exc:
            ok, detail = False, f"검사 예외: {exc}"
        wb.close()
        results.append((cat, label, ok, detail, last, asked))
        mark = "OK  " if ok else "FAIL"
        extra = f"  [action={last.get('action')} ask={asked} reason={str(last.get('reason'))[:60]}]" if not ok else ""
        print(f"{mark} {cat}/{label:14s} {detail[:90]}{extra}", flush=True)
        path.unlink(missing_ok=True)

    total = len(results)
    passed = sum(1 for r in results if r[2])
    print(f"\n{'='*80}\n{passed}/{total} 통과")
    fails = [(c, label, d, b) for c, label, ok, d, b, _ in results if not ok]
    if fails:
        print("\n실패 상세:")
        for c, label, d, b in fails:
            print(f"  {c}/{label}: {d}")
            print(f"    action={b.get('action')} reason={str(b.get('reason'))[:100]}")


main()
