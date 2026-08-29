r"""라우터 **최종 결정**과 의도 해석의 일치율·예측력 — 라운드 6 재계측.

    & $PY scripts\measure_final_decision_agreement.py                # 보고서 전체
    & $PY scripts\measure_final_decision_agreement.py 120            # 앞 120건만(층화)

왜 다시 재는가(감사 라운드 6): 기존 measure_rule_intent_agreement.py는
`_build_quick_action_plan`의 **중간 산출**을 쟀다. 라우터는 그 뒤에서 quick 계획을
버리기도(_quick_plan_underfits_message), 근거 검사로 깎기도 한다 — 34.4%는 그래서
폐기됐다. 이 스크립트는 게이트 보고서에 남은 **실제 실행 액션**(라우터 최종 결정)과
의도 해석의 종류를 견주고, 그 불일치가 오답(WRONG)을 예고하는지까지 잰다.

- 최종 결정: `datasets/eval/blind_paraphrases_v1_report.json`의 `action` (게이트 실측)
- 의도 해석: normalize_intent를 그 문장이 게이트에서 봤던 것과 **같은 씨앗 워크북**의
  다이제스트로 호출(게이트 모듈의 task→seed를 그대로 씀). 다이제스트가 다르면
  해석이 다른 관측을 보고 판단한 셈이라 비교가 성립하지 않는다.
- 예측력: P(WRONG | 불일치) vs P(WRONG | 일치), 그리고 되묻기 비용(불일치∧정답 수).

주의: LLM(일반 모델) 호출이 문장 수만큼 나간다 — 게이트·배터리와 동시에 돌리지 않는다.
"""
from __future__ import annotations

import asyncio
import collections
import json
import os
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("EXCEL_LIVE_ENGINE", "file")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

from office_claw_sidecar.services.excel_intent_normalizer import normalize_intent
from office_claw_sidecar.services.excel_live_agent import normalize_common_typos
from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
from office_claw_sidecar.services.excel_workbook_digest import build_workbook_digest
from office_claw_sidecar.services.llm_service import get_llm_service

ROOT = Path(__file__).resolve().parent.parent.parent
REPORT = ROOT / "datasets/eval/blind_paraphrases_v1_report.json"
GATE = Path(__file__).resolve().parent / "run_blind_paraphrase_gate.py"

#: 해석 종류 → 그 종류가 낼 수 있는 최종 액션들. 종류로 견준다 — 계획끼리 견주면
#: 매핑이 없어 물러난 종류의 불일치가 안 보인다(measure_rule_intent_agreement.py 참조).
TASK_ACTIONS: dict[str, set[str]] = {
    "fill_color": {"excel_live.fill_range"},
    "font": {"excel_live.set_font"},
    "highlight": {"excel_live.highlight_by_condition", "excel_live.apply_formula_cf"},
    "number_format": {"excel_live.set_number_format"},
    "formula": {"excel_live.set_formula", "excel_live.calculate_column_stat"},
    "sort": {"excel_live.sort_range", "excel_live.sort_rows"},
    "filter": {"excel_live.filter_rows"},
    "clear_values": {"excel_live.clear_range"},
    "reset_all": {"excel_live.clear_range", "excel_live.apply_border"},
    "write_value": {"excel_live.write_range"},
    "find_replace": {"excel_live.find_replace"},
    "create_table": {"excel_live.create_table", "excel_live.convert_to_excel_table"},
    "pivot": {"excel_live.pivot_table"},
    "chart": {"excel_live.create_chart"},
    "dedupe": {"excel_live.dedupe_rows", "excel_live.find_duplicates"},
    "read": {"excel_live.read_range", "excel_live.calculate_column_stat"},
    "border": {"excel_live.apply_border", "excel_live.set_border"},
    "merge": {"excel_live.merge_cells"},
    "create_sheet": {"excel_live.create_sheet"},
    "delete_charts": {"excel_live.delete_charts"},
    "freeze": {"excel_live.freeze_panes"},
    "autofit": {"excel_live.autofit_columns"},
    "rename_sheet": {"excel_live.rename_sheet"},
}


def _load_gate_tasks() -> dict[str, dict]:
    """게이트 모듈의 TASKS(task→씨앗·ctx)를 실행 없이 빌려 온다."""
    ns: dict = {"__name__": "blind_gate_borrowed", "__file__": str(GATE)}
    source = GATE.read_text(encoding="utf-8")
    # 실행부를 잘라낸다 — 게이트는 `if __name__` 없이 모듈 수준 `asyncio.run(main())`으로
    # 돈다. 통째로 exec하면 **게이트가 실제로 시작된다**(2026-08-25 실측: argv가 없어
    # IndexError로 멈춘 덕에 무사했다). 우리는 씨앗과 ctx만 필요하다.
    cuts = [i for i in (source.find("\nasyncio.run("), source.find("if __name__ ==")) if i > 0]
    cut = min(cuts) if cuts else len(source)
    body = source[:cut]
    if "asyncio.run(" in body:
        raise RuntimeError("게이트 실행부를 잘라내지 못했다 — 여기서 멈춘다")
    exec(compile(body, str(GATE), "exec"), ns)
    return ns["TASKS"]


def _digest_for_seed(seed_fn, cache: dict, work: Path) -> dict:
    """씨앗 함수별로 임시 워크북을 만들어 **생산 코드의 다이제스트**를 얻는다."""
    key = getattr(seed_fn, "__name__", str(seed_fn))
    if key in cache:
        return cache[key]
    path = work / f"{key}.xlsx"
    wb = Workbook()
    seed_fn(wb)
    wb.save(path)
    svc = FileExcelLiveService()
    svc.select_workbook(str(path))
    digest = build_workbook_digest(svc, workbook_id=str(path), use_cache=False)
    cache[key] = digest
    return digest


async def main() -> None:
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 0
    run_id = datetime.now().strftime("%m%d-%H%M%S-final-agreement")
    rows = json.loads(REPORT.read_text(encoding="utf-8"))
    if limit:
        stride = max(1, len(rows) // limit)
        rows = rows[::stride][:limit]
    tasks = _load_gate_tasks()
    work = Path(tempfile.mkdtemp(prefix="agree_"))
    digest_cache: dict = {}
    llm = get_llm_service()

    tally = collections.Counter()
    #: (일치 여부, outcome) 교차표 — 예측력의 원자료
    cross = collections.Counter()
    examples: list[str] = []
    out_rows: list[dict] = []
    t0 = time.time()

    for i, row in enumerate(rows, 1):
        text = str(row.get("text") or "")
        action = str(row.get("action") or "")
        outcome = str(row.get("outcome") or "")
        spec = tasks.get(str(row.get("task") or "")) or {}
        seed_fn = spec.get("seed")
        digest = _digest_for_seed(seed_fn, digest_cache, work) if seed_fn else {}
        try:
            intent = await normalize_intent(normalize_common_typos(text), digest, llm) or {}
        except Exception:
            intent = {}
        task = str(intent.get("task") or "")
        expected = TASK_ACTIONS.get(task, set())

        if action and expected:
            key = "일치" if action in expected else "불일치"
            cross[(key, outcome)] += 1
            if key == "불일치" and len(examples) < 20:
                examples.append(f"{outcome:9} {text[:36]:38} 최종={action[11:]:22} 해석={task}")
        elif action:
            key = "해석없음"  # 해석이 종류를 못 정했거나 매핑 밖(other 등)
        else:
            key = "실행없음"  # ASK/ERROR — 최종 결정이 실행이 아니었다
        tally[key] += 1
        out_rows.append({"text": text, "action": action, "outcome": outcome,
                         "intent_task": task, "agree": key})
        print(f"[{i:3d}/{len(rows)}] {key:5} {outcome:9} {text[:40]}", flush=True)

    total = sum(tally.values())
    both = tally["일치"] + tally["불일치"]
    wrong_dis = sum(n for (k, o), n in cross.items() if k == "불일치" and o == "WRONG")
    wrong_agr = sum(n for (k, o), n in cross.items() if k == "일치" and o == "WRONG")
    pass_dis = sum(
        n for (k, o), n in cross.items() if k == "불일치" and o in {"PASS_RULE", "PASS_CARD"}
    )
    lines = [
        f"문장 {total} · 둘 다 결정 {both} (일치 {tally['일치']} · 불일치 {tally['불일치']}) · "
        f"해석없음 {tally['해석없음']} · 실행없음 {tally['실행없음']}",
    ]
    if tally["불일치"]:
        lines.append(
            f"P(WRONG|불일치) = {wrong_dis}/{tally['불일치']} = {100 * wrong_dis / tally['불일치']:.1f}%"
        )
    if tally["일치"]:
        lines.append(
            f"P(WRONG|일치)   = {wrong_agr}/{tally['일치']} = {100 * wrong_agr / tally['일치']:.1f}%"
        )
    lines.append(f"되묻기 비용: 불일치인데 정답이었던 문장 {pass_dis}건 — 이만큼이 공연히 멈춘다")
    report = {
        "run_id": run_id,
        "요약": lines,
        "tally": dict(tally),
        "교차표": {f"{k}|{o}": n for (k, o), n in sorted(cross.items())},
        "불일치예시": examples,
        "rows": out_rows,
        "초": round(time.time() - t0),
    }
    out_path = ROOT / "logs" / f"final_agreement_{run_id}.json"
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print("\n" + "=" * 78)
    for line in lines:
        print(line)
    if examples:
        print("\n불일치 예시:")
        for line in examples:
            print(f"  {line}")
    print(f"\nrun_id={run_id} → {out_path}")


asyncio.run(main())
