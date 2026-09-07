"""배터리 결과 워크북을 전수 감사한다 — API의 자기보고 대신 파일을 연다.

    & $PY python-sidecar/scripts/audit_result_workbooks.py [워크스페이스 경로]

기본 경로는 `%LOCALAPPDATA%/office_claw/Workspace`다. 발견이 0건이어야 정상이다.

한 가지 예외: `AI_Excel_Automation_Demo.xlsx`의 `AI_Command_Center` 시트는
**예시 명령을 일부러 적어 둔 데모**라 9건이 늘 잡힌다 — 오염이 아니다.

2026-08-19 감사에서 배운 것: 배터리가 99.7%라고 한 상태에서 결과 워크북에는
명령문이 박힌 칸 4개, 원본 시트에 잘못 쓰인 수식 6건, **지워진 학생 이름 1개**가 있었다.

찾는 것:
  1) 셀에 **명령문**이 박힌 것 — "…해줘", "…아래에", "…라고" 로 끝나는 긴 글자
  2) 셀에 **수식 요청문**이 글자로 박힌 것 — 셀 주소 + 계산 낱말
  3) `0.1` 같은 **자릿수 오해 표시 형식**
  4) 순환 참조 수식(`A2 = =SUM(A2:A2)`)
  5) 표 전체가 한 칸으로 병합된 것
"""
import glob
import os
import re
import sys

from openpyxl import load_workbook

# 호출자가 PYTHONUTF8=1 을 안 붙여도 한국어 출력이 죽지 않게 한다(2026-09-07 실측:
# cp949 콘솔에서 첫 print 가 UnicodeEncodeError 로 죽어 '로그가 잘린다'로 보였다).
from _console import force_utf8

force_utf8()

def _default_workspace():
    from office_claw_sidecar.config import get_workspace_root
    return str(get_workspace_root())


ROOT = sys.argv[1] if len(sys.argv) > 1 else _default_workspace()

COMMAND_TAIL = re.compile(r"(해줘|해 줘|주세요|줄래|넣어|입력해|바꿔|칠해|만들어|지워|보여줘|부탁)\s*[.!~…]*$")
LOCATIVE = re.compile(r"(아래에|밑에|위에|옆에|다음 줄에|여기에|거기에)\s*$")
FORMULA_TEXT = re.compile(r"(?<![A-Za-z0-9])[A-Za-z]{1,3}\d{1,7}(?![A-Za-z0-9])")
OPERATION = re.compile(r"(빼기|더하기|나누기|곱하기|뺀|더한|나눈|곱한|합계|평균|차이)")
BAD_FORMAT = re.compile(r"^(#,##0|0)\.[1-9]$")
CELL = re.compile(r"^([A-Za-z]{1,3})(\d{1,7})$")


def cell_index(ref: str):
    m = CELL.fullmatch(str(ref or "").strip())
    if not m:
        return None
    col = 0
    for ch in m.group(1).upper():
        col = col * 26 + (ord(ch) - 64)
    return int(m.group(2)), col


def circular(coord: str, formula: str) -> bool:
    target = cell_index(coord)
    if target is None or not formula.startswith("="):
        return False
    for m in re.finditer(r"(?<![A-Za-z0-9_!])([A-Za-z]{1,3}\d{1,7})\s*:\s*([A-Za-z]{1,3}\d{1,7})", formula):
        if formula[: m.start()].rstrip().endswith("!"):
            continue
        a, b = cell_index(m.group(1)), cell_index(m.group(2))
        if not a or not b:
            continue
        if min(a[0], b[0]) <= target[0] <= max(a[0], b[0]) and min(a[1], b[1]) <= target[1] <= max(a[1], b[1]):
            return True
    return False


def main() -> int:
    findings = 0
    # `~$…`는 엑셀이 파일을 열어 둔 동안 만드는 잠금 파일이다 — 감사 대상이 아니다.
    books = [
        p
        for p in sorted(glob.glob(os.path.join(ROOT, "*.xlsx")))
        if not os.path.basename(p).startswith("~$")
    ]
    for path in books:
        name = os.path.basename(path)
        try:
            wb = load_workbook(path)
        except Exception as exc:
            print(f"[{name}] 열기 실패: {exc}")
            findings += 1
            continue
        hits = []
        for ws in wb.worksheets:
            for merged in ws.merged_cells.ranges:
                bounds = str(merged)
                if ws.max_row >= 4 and re.fullmatch(r"A1:[A-Z]{1,3}(\d+)", bounds):
                    rows = int(re.fullmatch(r"A1:[A-Z]{1,3}(\d+)", bounds).group(1))
                    if rows >= 4:
                        hits.append(f"{ws.title}!{bounds} 표 전체가 한 칸으로 병합")
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    fmt = str(c.number_format or "")
                    if BAD_FORMAT.match(fmt):
                        hits.append(f"{ws.title}!{c.coordinate} 표시형식={fmt} (자릿수 오해)")
                    if not isinstance(v, str):
                        continue
                    text = v.strip()
                    if text.startswith("="):
                        if circular(c.coordinate, text):
                            hits.append(f"{ws.title}!{c.coordinate} 순환 수식 {text[:40]}")
                        continue
                    if len(text) < 6:
                        continue
                    if COMMAND_TAIL.search(text) or LOCATIVE.search(text):
                        hits.append(f"{ws.title}!{c.coordinate} 명령문이 값으로: {text[:50]!r}")
                    elif FORMULA_TEXT.search(text) and OPERATION.search(text):
                        hits.append(f"{ws.title}!{c.coordinate} 수식 요청문이 값으로: {text[:50]!r}")
        if hits:
            findings += len(hits)
            print(f"[{name}] {len(hits)}건")
            for h in hits[:12]:
                print(f"    {h}")
    print(f"\n워크북 {len(books)}개 · 발견 {findings}건")
    return 0


if __name__ == "__main__":
    sys.exit(main())
