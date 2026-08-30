"""대화형 재현 러너 — 사람 질문 ↔ 모델 답변을 전부 기록하며 끝까지 완주한다.

각본 JSON: {"workbook_name": ..., "turns": [{"zone", "command", "expect"?, "paste"?, "ui"?}]}
  expect: "ok"(기본) — 실행돼야 함 / "ask" — 되묻기·해석 카드가 나와야 함(사람의
          실수·모호한 질문 턴; 다음 턴은 사람의 답변) / "noop" — 실행되면 안 됨.
  paste : "A1:F6" — 사람이 Excel에서 그 범위를 잡고 Ctrl+C 한 뒤 채팅창에 Ctrl+V 했다.
          프론트(`excelPaste.js`)는 붙여넣은 표를 `[[EXCEL_RANGE:A1:F6]]` + 안내 문구로
          바꾸고 lastExcelRangeRef를 A1:F6으로 둔다. `command`는 그 뒤에 사람이 이어
          친 문장이다("지역,주문건수,…; 수도권,… 넣어줘"). 2026-08-18 GUI 실측 화면
          그대로 — 사람은 "시트 A1:F6에 …" 같은 좌표 문장을 치지 않는다.
  ui    : {"activate_sheet": "지역성과"} — 사람이 Excel에서 시트 탭을 클릭했다(채팅 아님).

GUI(`WorkspacePage.jsx handleSend`)와 **같은 순서·같은 규칙**으로 요청을 만든다:
  마크업 제거 → 복합문 분리(splitExcelCompositeCommand) → "여기/이 범위" 지시어면
  붙여넣기 범위를 접두(applyRangeContextToCommand) → 각 부분: 문장에 범위가 있으면
  context_range 없음, 아니면 lastExcelRangeRef → 성공 응답의 result.address가 다음
  lastExcelRangeRef. workbook_id 없음("선택된 통합문서"), 승인은 approval_id 재개.
판정은 파일 상태(공통 불변식: 첫 시트 A1이 문장 텍스트로 오염되면 실패)와 기대 일치다.
"""

from __future__ import annotations

import asyncio
import json
import os
import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.routers.excel_live import (
    ApprovalResponse,
    ExcelLiveCommandRequest,
    post_approval,
    post_command,
)
from office_claw_sidecar.services.excel_live_service import (
    get_excel_live_service,
    invalidate_excel_engine_cache,
)
from office_claw_sidecar.services.llm_service import get_llm_service

ROUTER = Path(__file__).resolve().parents[1] / "office_claw_sidecar" / "routers" / "excel_live.py"
_router_src = ROUTER.read_text(encoding="utf-8")
EXECUTABLE = set(re.findall(r'action == "(excel_live\.[a-z_]+)"', _router_src))
EXECUTABLE |= set(re.findall(r'"(excel_live\.[a-z_]+)"', _router_src))

SCENARIO = Path(sys.argv[1])
SC = json.loads(SCENARIO.read_text(encoding="utf-8"))
WB = Path(r"C:/Users/asdjj/AppData/Local/office_claw/Workspace") / SC["workbook_name"]
SESSION_BASE = "test-dialogue-" + SCENARIO.stem
TURNS = SC["turns"]

# ---- 프론트 규칙 이식 (src/lib/excelCommandUtils.js · WorkspacePage.jsx) ----------
_SPLIT_SEPS = [
    r"\s+그리고\s+", r"\s+그리고나서\s+", r"\s*하고나서\s*", r"\s*한\s*다음(?:에)?\s*",
    r"\s*한\s*뒤(?:에)?\s*", r"\s*후에\s*",
    r"\s+그\s*다음(?:으로)?\s+(?!(?:줄|행|칸|열|시트|탭|표|페이지|단계))",
    r"\s+다음(?:으로)?\s+(?!(?:줄|행|칸|열|시트|탭|표|페이지|단계))",
    r"\s+이후\s+", r"\s*하고\s+", r"\s+then\s+", r"\s+and then\s+",
]


_FILLER_ONLY_PART = re.compile(
    r"^(?:(?:아|어|음|응|네|넵|옙|예|ㅇㅇ|ㅇㅋ|ㅋㅋ+|ㅎㅎ+|흠|오|아하|좋아|그래|ok|okay|그럼|자|그|이|저|일단|먼저|우선|그냥|아니|근데|그런데|그리고|그리구|또)[\s,.!~]*)+$",
    re.I,
)


def split_composite(raw: str) -> list[str]:
    # 프론트와 같다: 줄바꿈은 살리고 가로 공백만 하나로(여러 줄 값 붙여넣기 = 행).
    text = re.sub(r"\r\n?", "\n", str(raw or ""))
    text = re.sub(r"[^\S\n\t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text).strip()
    if not text:
        return []
    if looks_like_value_list_write(text):
        return [text]
    parts = [text]
    for sep in _SPLIT_SEPS:
        parts = [p for part in parts for p in re.split(sep, part, flags=re.I)]
    out = []
    for p in parts:
        p = re.sub(r"^[,\s]+|[,\s]+$", "", p.strip())
        if p and not _FILLER_ONLY_PART.match(p):
            out.append(p)
    return out


_CELL_ASCII = re.compile(
    r"\b([A-Z]{1,3}\d{1,7}:[A-Z]{1,3}\d{1,7}|[A-Z]{1,3}:[A-Z]{1,3}|[A-Z]{1,3}\d{1,7})\b",
    re.I | re.ASCII,
)
_TARGET_REF = re.compile(
    r"\b[A-Z]{1,3}\d{1,7}:[A-Z]{1,3}\d{1,7}\b|\b[A-Z]{1,3}\d{1,7}\s*(?:에|에다|부터|까지)",
    re.I | re.ASCII,
)


def looks_like_value_list_write(text: str) -> bool:
    return len(re.findall(r"[,;\t\n]", text)) >= 3 and re.search(
        r"(입력|기록|넣어|채워|써)\s*(?:해)?\s*(?:줘요|줘|주세요|주라|줄래|놔|둬|봐|조)?\s*[~.!?…]*\s*$",
        text,
    ) is not None


def has_explicit_range(cmd: str) -> bool:
    text = str(cmd or "")
    if looks_like_value_list_write(text):
        return _TARGET_REF.search(text) is not None
    return _CELL_ASCII.search(text) is not None


def apply_range_context(cmd: str, range_ref: str | None) -> str:
    text = str(cmd or "").strip()
    if not range_ref or not text or has_explicit_range(text):
        return text
    if not re.search(r"(이\s*범위|해당\s*범위|복사한\s*범위|선택한\s*범위|여기)", text, re.I):
        return text
    return f"{range_ref} {text}"


def range_shape(ref: str) -> tuple[int, int]:
    m = re.match(r"^([A-Z]{1,3})(\d{1,7})(?::([A-Z]{1,3})(\d{1,7}))?$", ref.upper())
    if not m:
        return (0, 0)
    if not m.group(3):
        return (1, 1)

    def col(s: str) -> int:
        n = 0
        for ch in s:
            n = n * 26 + (ord(ch) - 64)
        return n

    return (abs(int(m.group(4)) - int(m.group(2))) + 1, abs(col(m.group(3)) - col(m.group(1))) + 1)


def paste_note(ref: str) -> str:
    rows, cols = range_shape(ref)
    return f"📋 엑셀에서 붙여넣은 {rows}행 × {cols}열 — {ref.upper()} 범위로 인식했습니다"


# ---------------------------------------------------------------------------------


def _reply_text(resp, result, approval_summary: str) -> str:
    """사용자가 화면에서 실제로 읽는 문구 하나."""
    if result.get("ask_follow_up"):
        return str(result.get("follow_up_question") or getattr(resp, "reason", ""))
    if result.get("execution_report"):
        return str(result["execution_report"])
    if approval_summary:
        return approval_summary
    return str(getattr(resp, "reason", ""))


async def run_once(round_no: int) -> list[dict]:
    if WB.exists():
        WB.unlink()
    Workbook().save(WB)
    print(f"[라운드 {round_no}] {WB.name} · {len(TURNS)}턴")
    llm = get_llm_service()
    session = f"{SESSION_BASE}-r{round_no}"
    log: list[dict] = []
    last_ref: str | None = None  # 프론트의 lastExcelRangeRef

    for i, turn in enumerate(TURNS, 1):
        zone = turn.get("zone", "")
        expect = turn.get("expect", "ok")
        ui = turn.get("ui") or {}
        paste = str(turn.get("paste") or "").strip().upper()
        raw = str(turn.get("command") or "")
        invalidate_excel_engine_cache()
        svc = get_excel_live_service()
        svc.select_workbook(str(WB))
        if ui.get("activate_sheet"):
            # 사람이 Excel에서 시트 탭을 클릭했다 — 채팅이 아니라 UI 동작이다.
            try:
                svc.select_sheet(None, str(ui["activate_sheet"]))
            except Exception as exc:
                print(f"[{i:3d}] 탭 클릭 실패 {ui['activate_sheet']}: {exc}")
        if not raw.strip():
            log.append({"turn": i, "zone": zone, "text": "", "ui": ui, "expect": expect,
                        "ok": True, "asked": False, "action": "", "reply": "", "why": ""})
            print(f"[{i:3d}] (UI) 탭 클릭 → {ui.get('activate_sheet', '')}")
            continue
        display = raw
        range_tag = None
        tsv = str(turn.get("tsv") or "")
        if paste:
            last_ref = paste  # handlePaste: setLastExcelRangeRef(address)
            range_tag = paste
            if tsv:
                # 다른 앱·통합문서에서 복사한 표: 선택은 비어 있고 값은 표에 있다 →
                # 프론트가 값을 살려 보낸다(buildPasteBlock keepValues).
                rows = [ln for ln in tsv.split("\n") if ln.strip()]
                cols = len(rows[0].split("\t")) if rows else 0
                display = f"📋 밖에서 가져온 표 {len(rows)}행 × {cols}열 — {paste}부터 넣습니다\n{tsv}\n{raw}"
                raw = tsv + "\n" + raw
            else:
                display = paste_note(paste) + "\n" + raw
        parts = [apply_range_context(c, range_tag) for c in split_composite(raw)]
        t0 = time.time()
        approval_summary = ""
        interpretation = False
        results: list[tuple[str, object, dict, str | None]] = []
        error = None
        for cmd in parts:
            ctx = None if has_explicit_range(cmd) else last_ref
            try:
                resp = await post_command(
                    ExcelLiveCommandRequest(
                        message=cmd, session_id=session, workbook_id=None,
                        approve=False, context_range=ctx,
                    ),
                    llm,
                )
                if getattr(resp, "approval_required", False):
                    pending = getattr(resp, "pending_approval", None)
                    approval_summary = str(getattr(pending, "summary", "") or "")
                    interpretation = bool(getattr(pending, "interpretation", False))
                    resp = await post_approval(
                        ApprovalResponse(approval_id=pending.approval_id, approved=True), llm
                    )
            except Exception as exc:
                error = exc
                break
            result = getattr(resp, "result", None) or {}
            results.append((cmd, resp, result, ctx))
            if result.get("ask_follow_up") or getattr(resp, "ok", False) is False:
                break  # GUI도 여기서 멈춘다
            addr = str(result.get("address") or "").strip().upper()
            if addr:
                last_ref = addr
        if error is not None:
            log.append({"turn": i, "zone": zone, "text": display, "paste": paste, "expect": expect,
                        "ok": False, "why": f"예외 {type(error).__name__}", "reply": str(error)[:200],
                        "asked": False, "action": ""})
            print(f"[{i:3d}] 예외  {raw[:44]}  {type(error).__name__}")
            continue

        cmd, resp, result, ctx = results[-1]
        action = str(getattr(resp, "action", ""))
        asked = bool(result.get("ask_follow_up")) or "clarify" in action
        all_executed = len(results) == len(parts) and all(
            bool(getattr(r, "ok", False)) and not (rs.get("ask_follow_up")) and str(getattr(r, "action", "")) in EXECUTABLE
            for _, r, rs, _ in results
        )
        if expect == "ask":
            good = asked
            why = "" if good else "되물었어야 함"
        elif expect == "noop":
            good = asked or action in {"excel_live.noop", "excel_live.clarify", "excel_live.not_excel"} or bool(result.get("noop"))
            why = "" if good else "실행돼 버림"
        else:
            good = all_executed
            why = "" if good else ("되묻기" if asked else "미지원")
            # 미검출 오실행 감지 — "합계 줄 넣어줘, 이 표 아래에"가 값 쓰기로 처리돼 머리글이 덮였는데 '성공'으로
            # 집계됐다(2026-08-19 ex11 v2 실측). 집계·서식 문장이 값 몇 개짜리 write_range로 끝나면 실패로 센다.
            if good and action == "excel_live.write_range" and not paste:
                _plain = re.sub(r"\s+", " ", str(raw or ""))
                _report = str(result.get("execution_report") or getattr(resp, "execution_report", "") or "")
                if (
                    re.search(r"(합계|총합|평균|소계|개수)\s*(?:줄|행|한\s*줄|하나)|(?:아래|밑)에\s*(?:합계|평균)", _plain)
                    and not re.search(r"[;\n]", _plain)
                    and "수식" not in _report
                ):
                    good = False
                    why = "오실행(집계를 값으로 씀)"
        reply = _reply_text(resp, result, approval_summary)
        entry = {
            "turn": i, "zone": zone, "text": display, "paste": paste, "ui": ui, "expect": expect,
            "action": action, "ok": good, "why": why, "asked": asked, "interpretation": interpretation,
            "reply": reply[:600], "secs": round(time.time() - t0, 1),
            "parts": len(parts), "context_range": ctx or "",
        }
        log.append(entry)
        flag = "OK  " if good else "FAIL"
        tag = " (질문)" if asked else (" (해석카드)" if interpretation else "")
        ptag = f" [📋{paste}]" if paste else ""
        print(f"[{i:3d}] {flag}{tag}{ptag} {raw[:46]}")
        if not good:
            print(f"          action={action} {why} ctx={ctx} · {reply[:80]}")

    ok = sum(1 for t in log if t["ok"])
    print(f"성공 {ok} / {len(TURNS)}")
    # 공통 불변식: 첫 데이터 시트 A1이 문장 텍스트로 오염되면 라운드 실패다.
    wb = load_workbook(WB)
    for name in wb.sheetnames:
        a1 = wb[name]["A1"].value
        if isinstance(a1, str) and len(a1) > 14 and ("줘" in a1 or "해" in a1[-2:]):
            print(f"  !! {name}!A1 오염: {a1[:40]}")
            for t in log:
                t["ok"] = False if t["turn"] == log[-1]["turn"] else t["ok"]
    wb.close()
    out_dir = Path(os.environ.get("DIALOGUE_LOG_DIR") or SCENARIO.parent)
    (out_dir / (SCENARIO.stem + "_log.json")).write_text(
        json.dumps(log, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return log


def _acquire_battery_lock():
    """게이트·다른 배터리와의 동시 실행을 자물쇠로 막는다(pid 생존 검사 포함).

    2026-08-30 실측: 파이프에 물린 러너가 고아로 살아남아 두 프로세스가 같은
    워크북을 두드렸고, 팔A 측정 전체가 오염됐다. 규율(CLAUDE.md)이 아니라
    코드가 막아야 한다. 자물쇠 구현은 루트 scripts/_run_lock.py 한 곳뿐이다.
    """
    import sys
    from pathlib import Path

    root = Path(__file__).resolve().parents[3]
    sys.path.insert(0, str(root / "scripts"))
    from _run_lock import RunLock

    lock = RunLock("dialogue-battery").__enter__()
    if not lock.acquired:
        print(f"!! 동시 실행 금지: 자물쇠를 쥔 쪽 → {lock.held_by}", flush=True)
        print("   (게이트나 다른 배터리가 도는 중이다. 끝난 뒤 다시 실행하라.)", flush=True)
        raise SystemExit(3)
    return lock


async def main() -> None:
    repeat = max(1, int(os.environ.get("HUMAN_REPEAT", "1")))
    logs = [await run_once(r) for r in range(1, repeat + 1)]
    if repeat > 1:
        base = [(t["ok"], t.get("action"), t.get("asked")) for t in logs[0]]
        flaky = [
            (r, t["turn"], t["text"][:36])
            for r, lg in enumerate(logs[1:], start=2)
            for i, t in enumerate(lg)
            if (t["ok"], t.get("action"), t.get("asked")) != base[i]
        ]
        print(f"결정성: {repeat}회 중 흔들린 턴 {len(flaky)}건")
        for f in flaky:
            print("   ", f)


from office_claw_sidecar.services import decision_trace as _dt

# 배터리 턴을 사람이 친 명령과 가를 출처 태그(2026-08-19 로그 감사: 실사용 로그의 source가 전부 비어 있었다).
_dt.source(kind="script", name="run_dialogue", scenario=SCENARIO.stem).__enter__()
_battery_lock = _acquire_battery_lock()
try:
    asyncio.run(main())
finally:
    _battery_lock.__exit__(None, None, None)
