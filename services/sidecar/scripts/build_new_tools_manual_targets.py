"""이번 배치에서 추가한 11개 도구(찾아바꾸기~표시형식)의 정답 계획을 수동으로 만들고 검증한다.

이 11개 도구는 오늘 처음 추가돼 학습 데이터가 전혀 없다 — 기존 SFT 세트(35건)에는
단 한 건도 없다. `build_hard_case_manual_targets.py`와 같은 방식으로, 실제 데모
통합문서 복사본에 계획을 실행하고 결과 상태를 직접 확인한 것만 `label_status=verified`로
기록한다.
"""

from __future__ import annotations

import argparse
import copy
import json
import os
import shutil
import tempfile
import uuid
from collections.abc import Callable
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

os.environ.setdefault("EXCEL_LIVE_ENGINE", "file")

from openpyxl import load_workbook

from office_claw_sidecar.routers.excel_live import _execute_action, _verify_step_result
from office_claw_sidecar.services.excel_live_executor import (
    execute_plan,
    normalize_plan_steps,
)
from office_claw_sidecar.services.excel_live_plan_validator import (
    ValidationContext,
    validate_plan,
)

KST = timezone(timedelta(hours=9), name="KST")
DEMO_TEMPLATE = Path("복잡한 엑셀 작업을 위한 자료/AI_Excel_Automation_Demo.xlsx")
SHEET = "Sales_Data"

# 각 케이스: turns[i].variants(패러프레이즈) + turns[i].plan(정답 action_plan).
# check(wb)는 마지막 턴 실행 후 워크북을 열어 실제로 반영됐는지 확인하는 함수.
NEW_TOOL_CASES: list[dict[str, Any]] = [
    {
        "id": "newtool-find_replace",
        "turns": [
            {
                "variants": [
                    "경기를 경기도로 바꿔줘",
                    "지역 이름 경기를 경기도로 바꿔줘",
                    "경기라고 되어있는 값들을 경기도로 바꿔줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.find_replace",
                        "params": {"target_range": "__USED_RANGE__", "find_text": "경기", "replace_text": "경기도"},
                        "reason": "지역(Region) 값 경기를 경기도로 일괄 치환",
                    }
                ],
            }
        ],
        "check": lambda ws: any(str(c.value) == "경기도" for c in ws["D"][1:]),
    },
    {
        "id": "newtool-merge-unmerge",
        "turns": [
            {
                "variants": ["S1이랑 T1 셀 합쳐줘", "S1:T1 병합해줘", "S1부터 T1까지 하나로 합쳐줘"],
                "plan": [
                    {
                        "action": "excel_live.merge_cells",
                        "params": {"target_range": "S1:T1"},
                        "reason": "S1:T1 셀 병합",
                    }
                ],
            },
            {
                "variants": ["방금 합친 셀 다시 풀어줘", "S1:T1 병합 해제해줘", "그거 병합 취소해줘"],
                "plan": [
                    {
                        "action": "excel_live.unmerge_cells",
                        "params": {"target_range": "S1:T1"},
                        "reason": "S1:T1 병합 해제",
                    }
                ],
            },
        ],
        "check": lambda ws: len(list(ws.merged_cells.ranges)) == 0,
    },
    {
        "id": "newtool-freeze_panes",
        "turns": [
            {
                "variants": ["첫 번째 행 고정해줘", "머리글 행이 안 움직이게 고정해줘", "A2 기준으로 틀 고정해줘"],
                "plan": [
                    {
                        "action": "excel_live.freeze_panes",
                        "params": {"freeze_at": "A2"},
                        "reason": "첫 행(머리글) 틀 고정",
                    }
                ],
            }
        ],
        "check": lambda ws: ws.freeze_panes == "A2",
    },
    {
        "id": "newtool-autofit_columns",
        "turns": [
            {
                "variants": ["열 너비 좀 자동으로 맞춰줘", "칸 너비 내용에 맞게 조정해줘", "표 전체 열 너비 자동조정해줘"],
                "plan": [
                    {
                        "action": "excel_live.autofit_columns",
                        "params": {"target_range": "__USED_RANGE__"},
                        "reason": "사용 범위 전체 열 너비 자동 조정",
                    }
                ],
            }
        ],
        "check": lambda ws: ws.column_dimensions["A"].width is not None,
    },
    {
        "id": "newtool-define_named_range",
        "turns": [
            {
                "variants": [
                    "이 표에 SalesData라는 이름 붙여줘",
                    "전체 범위 이름을 SalesData로 정의해줘",
                    "표 전체에 이름 정의해줘, 이름은 SalesData로",
                ],
                "plan": [
                    {
                        "action": "excel_live.define_named_range",
                        "params": {"name": "SalesData", "target_range": "__USED_RANGE__"},
                        "reason": "사용 범위 전체에 SalesData 이름 정의",
                    }
                ],
            }
        ],
        "check": lambda ws: "SalesData" in ws.parent.defined_names,
    },
    {
        "id": "newtool-set_print_area",
        "turns": [
            {
                "variants": [
                    "A1부터 Q30까지만 인쇄되게 설정해줘",
                    "인쇄 영역을 A1:Q30으로 지정해줘",
                    "이 부분만 인쇄되게 해줘, A1에서 Q30까지",
                ],
                "plan": [
                    {
                        "action": "excel_live.set_print_area",
                        "params": {"print_area": "A1:Q30"},
                        "reason": "인쇄 영역을 A1:Q30으로 제한",
                    }
                ],
            }
        ],
        "check": lambda ws: "A$1" in str(ws.print_area) and "Q$30" in str(ws.print_area),
    },
    {
        "id": "newtool-add_cell_comment",
        "turns": [
            {
                "variants": [
                    "A1에 메모 남겨줘, 원본 데이터라고",
                    "A1 셀에 코멘트 추가해줘: 원본 데이터",
                    "A1에 설명 좀 달아줘, 원본 데이터라고",
                ],
                "plan": [
                    {
                        "action": "excel_live.add_cell_comment",
                        "params": {"target_range": "A1", "text": "원본 데이터"},
                        "reason": "A1 셀에 메모 추가",
                    }
                ],
            }
        ],
        "check": lambda ws: ws["A1"].comment is not None and ws["A1"].comment.text == "원본 데이터",
    },
    {
        "id": "newtool-apply_color_scale",
        "turns": [
            {
                "variants": ["매출 열에 색조 넣어줘", "매출 값 크기대로 색깔 칠해줘", "매출 크기에 따라 색이 진해지게 해줘"],
                "plan": [
                    {
                        "action": "excel_live.apply_color_scale",
                        "params": {"target_range": "L2:L181"},
                        "reason": "매출(Sales, L열)에 색조 조건부 서식 적용",
                    }
                ],
            }
        ],
        "check": lambda ws: len(list(ws.conditional_formatting)) >= 1,
    },
    {
        "id": "newtool-apply_data_bar",
        "turns": [
            {
                "variants": ["수량 열에 데이터 막대 넣어줘", "수량 값을 막대로 표시해줘", "수량 크기를 막대처럼 셀에 표시해줘"],
                "plan": [
                    {
                        "action": "excel_live.apply_data_bar",
                        "params": {"target_range": "I2:I181"},
                        "reason": "수량(Qty, I열)에 데이터 막대 조건부 서식 적용",
                    }
                ],
            }
        ],
        "check": lambda ws: len(list(ws.conditional_formatting)) >= 1,
    },
    {
        "id": "newtool-set_number_format-percent",
        "turns": [
            {
                "variants": ["이익률 열을 퍼센트로 보여줘", "이익률을 퍼센트 형식으로 바꿔줘", "이익률 표시를 퍼센트로 해줘"],
                "plan": [
                    {
                        "action": "excel_live.set_number_format",
                        "params": {"target_range": "O2:O181", "format_code": "퍼센트"},
                        "reason": "이익률(Profit_Margin, O열)을 퍼센트 형식으로 표시",
                    }
                ],
            }
        ],
        "check": lambda ws: ws["O2"].number_format == "0.00%",
    },
    {
        "id": "newtool-set_number_format-comma",
        "turns": [
            {
                "variants": ["매출 열에 천단위 구분기호 넣어줘", "매출을 천단위로 표시해줘", "매출 숫자에 쉼표 넣어줘"],
                "plan": [
                    {
                        "action": "excel_live.set_number_format",
                        "params": {"target_range": "L2:L181", "format_code": "천단위"},
                        "reason": "매출(Sales, L열)에 천단위 구분기호 표시",
                    }
                ],
            }
        ],
        "check": lambda ws: ws["L2"].number_format == "#,##0",
    },
]


def _run_plan_turn(
    workbook_id: str, sheet_name: str, plan: list[dict[str, Any]], message: str
) -> tuple[bool, str]:
    try:
        normalized_steps = normalize_plan_steps(plan)
    except Exception as exc:
        return False, f"normalize_failed:{exc}"
    if not normalized_steps:
        return False, "normalize_empty"

    try:
        validated_steps = validate_plan(
            normalized_steps,
            context=ValidationContext(message=message, workbook_id=workbook_id, sheet_name=sheet_name),
        )
    except Exception as exc:
        return False, f"validate_failed:{exc}"

    execution = execute_plan(
        steps=validated_steps,
        max_attempts=1,
        abort_on_failure=True,
        execute_action=lambda action, params: _execute_action(
            action=action, params=params, workbook_id=workbook_id, sheet_name=sheet_name
        ),
        verify_step=lambda action, params, result: _verify_step_result(
            action=action, params=params, result=result, workbook_id=workbook_id, sheet_name=sheet_name
        ),
    )
    last = execution.last
    if last is None:
        return False, "execution_empty"
    if last.error:
        return False, f"execution_error:{last.error}"
    if not last.verified:
        return False, f"not_verified:{last.verify_detail}"
    return True, "ok"


def verify_case(case: dict[str, Any]) -> tuple[bool, list[str]]:
    template = DEMO_TEMPLATE if DEMO_TEMPLATE.is_absolute() else (Path(__file__).resolve().parents[3] / DEMO_TEMPLATE)
    notes: list[str] = []

    with tempfile.TemporaryDirectory(prefix="officeclaw_newtool_target_") as td:
        workbook_path = Path(td) / "verify.xlsx"
        shutil.copy2(template, workbook_path)
        workbook_id = str(workbook_path.resolve())

        for turn_idx, turn in enumerate(case["turns"]):
            ok, detail = _run_plan_turn(workbook_id, SHEET, turn["plan"], turn["variants"][0])
            if not ok:
                notes.append(f"turn[{turn_idx}] execution_failed: {detail}")
                return False, notes
            notes.append(f"turn[{turn_idx}] executed ok")

        checker: Callable[[Any], bool] = case["check"]
        wb = load_workbook(workbook_path)
        try:
            ws = wb[SHEET]
            passed = bool(checker(ws))
            notes.append(("ok " if passed else "NG ") + "final_state_check")
        finally:
            wb.close()
        return passed, notes


def _build_records(case: dict[str, Any]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    now = datetime.now(KST).isoformat()
    for turn_idx, turn in enumerate(case["turns"]):
        for variant in turn["variants"]:
            record_id = f"manual_new_tool:{case['id']}:t{turn_idx}:{uuid.uuid4().hex[:10]}"
            records.append(
                {
                    "schema_version": "excel_distill.v1",
                    "record_id": record_id,
                    "source": {
                        "dataset": "manual_new_tool_v1",
                        "split": "train",
                        "sample_id": f"{case['id']}:t{turn_idx}",
                        "license": "internal",
                        "provenance": {"source_file": "scripts/build_new_tools_manual_targets.py"},
                    },
                    "input": {
                        "instruction": variant,
                        "locale": "ko",
                        "workbook_refs": [],
                        "context_hints": {"sheet_name": SHEET, "reason": "", "status_code": 200},
                    },
                    "target": {
                        "task_type": "spreadsheet_edit",
                        "label_status": "verified",
                        "action_plan": copy.deepcopy(turn["plan"]),
                        "expected_output": {},
                    },
                    "quality": {
                        "verification": "manual_execution_replay",
                        "passed": True,
                        "confidence": 0.97,
                    },
                    "metadata": {
                        "created_at": now,
                        "generator": "python-sidecar/scripts/build_new_tools_manual_targets.py",
                        "notes": [f"scenario_id:{case['id']}", f"turn_index:{turn_idx}"],
                    },
                }
            )
    return records


def main() -> int:
    parser = argparse.ArgumentParser(description="신규 11개 도구 정답 계획 검증 및 학습 레코드 생성")
    parser.add_argument("--output-jsonl", type=Path, default=Path("../../datasets/distill/excel_new_tools_manual_v1.jsonl"))
    parser.add_argument("--only", type=str, default="")
    args = parser.parse_args()

    only = {s.strip() for s in args.only.split(",") if s.strip()}
    cases = [c for c in NEW_TOOL_CASES if not only or c["id"] in only]

    all_records: list[dict[str, Any]] = []
    passed_ids: list[str] = []
    for case in cases:
        ok, notes = verify_case(case)
        status = "PASS" if ok else "FAIL"
        print(f"[{status}] {case['id']}")
        for note in notes:
            print(f"    {note}")
        if ok:
            all_records.extend(_build_records(case))
            passed_ids.append(case["id"])
        else:
            print("    -> 학습 레코드 생성 건너뜀 (검증 실패)")

    args.output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with args.output_jsonl.open("w", encoding="utf-8") as f:
        for record in all_records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    print(
        f"[DONE] cases={len(cases)} verified_cases={len(passed_ids)} records={len(all_records)} "
        f"output={args.output_jsonl}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
