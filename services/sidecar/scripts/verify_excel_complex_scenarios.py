from __future__ import annotations

import argparse
import asyncio
import contextlib
import json
import shutil
import tempfile
import time
import uuid
from collections.abc import Iterator
from datetime import datetime, timedelta, timezone
from itertools import pairwise
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.cell_range import CellRange

from office_claw_sidecar.routers.excel_live import ExcelLiveCommandRequest, post_command
from office_claw_sidecar.services.llm_service import (
    get_llm_service,
    load_llm_config,
    reload_llm_service,
    save_llm_config,
)

KST = timezone(timedelta(hours=9), name="KST")


def _now_iso() -> str:
    return datetime.now(KST).isoformat()


def _safe_float(value: Any) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _safe_int(value: Any) -> int:
    try:
        return int(value)
    except Exception:
        return 0


def _normalize_action(value: Any) -> str:
    return str(value or "").strip().lower()


def _load_pack(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"시나리오 파일이 없습니다: {path}")
    parsed = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(parsed, dict):
        raise ValueError("시나리오 파일 루트는 JSON 객체여야 합니다.")
    scenarios = parsed.get("scenarios")
    if not isinstance(scenarios, list) or not scenarios:
        raise ValueError("scenarios 배열이 비어 있습니다.")
    return parsed


def _build_seed_workbook(path: Path, profile: str) -> None:
    wb = Workbook()

    sales = wb.active
    sales.title = "매출"
    sales.append(["월", "카테고리", "수량", "단가", "금액", "상태", "담당자", "코드"])
    sales_rows = [
        ["1월", "A", 10, 100, 1000, "완료", "김민수", "A001"],
        ["1월", "B", 8, 110, 880, "진행중", "이지은", "A002"],
        ["2월", "A", 7, 120, 840, "완료", "박준호", "A003"],
        ["2월", "B", 11, 90, 990, "지연", "최서연", "A004"],
        ["3월", "A", 9, 130, 1170, "완료", "김민수", "A005"],
        ["3월", "B", 12, 120, 1440, "진행중", "이지은", "A006"],
        ["4월", "C", 5, 200, 1000, "완료", "박준호", "A007"],
        ["4월", "C", 5, 200, 1000, "완료", "박준호", "A007"],
    ]
    for row in sales_rows:
        sales.append(row)

    prev = wb.create_sheet("전월")
    prev.append(["월", "카테고리", "수량", "단가", "금액", "상태", "담당자", "코드"])
    prev_rows = [
        ["1월", "A", 9, 100, 900, "완료", "김민수", "A001"],
        ["1월", "B", 8, 105, 840, "진행중", "이지은", "A002"],
        ["2월", "A", 7, 110, 770, "완료", "박준호", "A003"],
        ["2월", "B", 10, 90, 900, "지연", "최서연", "A004"],
        ["3월", "A", 9, 120, 1080, "완료", "김민수", "A005"],
        ["3월", "B", 11, 115, 1265, "진행중", "이지은", "A006"],
        ["4월", "C", 5, 195, 975, "완료", "박준호", "A007"],
        ["4월", "C", 5, 195, 975, "완료", "박준호", "A007"],
    ]
    for row in prev_rows:
        prev.append(row)

    comp = wb.create_sheet("비교")
    comp.append(["월", "카테고리", "수량", "단가", "금액", "상태", "담당자", "코드"])
    comp_rows = [
        ["1월", "A", 10, 100, 1000, "완료", "김민수", "A001"],
        ["1월", "B", 8, 111, 888, "진행중", "이지은", "A002"],
        ["2월", "A", 7, 120, 840, "완료", "박준호", "A003"],
        ["2월", "B", 11, 90, 990, "지연", "최서연", "A004"],
        ["3월", "A", 9, 130, 1170, "완료", "김민수", "A005"],
        ["3월", "B", 12, 121, 1452, "진행중", "이지은", "A006"],
        ["4월", "C", 5, 200, 1000, "완료", "박준호", "A007"],
        ["4월", "C", 6, 200, 1200, "완료", "박준호", "A007"],
    ]
    for row in comp_rows:
        comp.append(row)

    lookup = wb.create_sheet("조회표")
    lookup.append(["코드", "단가"])
    lookup_rows = [
        ["A001", 100],
        ["A002", 110],
        ["A003", 120],
        ["A004", 90],
        ["A005", 130],
        ["A006", 120],
        ["A007", 200],
    ]
    for row in lookup_rows:
        lookup.append(row)

    order_sheet = wb.create_sheet("주문")
    order_sheet.append(["코드", "단가"])
    for code in ("A001", "A002", "A003", "A004", "A005", "A006", "A007"):
        order_sheet.append([code, None])

    eval_sheet = wb.create_sheet("평가")
    eval_sheet.append(["이름", "점수", "결과"])
    eval_rows = [
        ["민수", 85, None],
        ["지은", 72, None],
        ["준호", 65, None],
        ["서연", 91, None],
        ["현우", 58, None],
        ["소민", 77, None],
        ["다은", 69, None],
    ]
    for row in eval_rows:
        eval_sheet.append(row)

    input_sheet = wb.create_sheet("입력")
    input_sheet.append(["일자", "항목", "수량", "단가", "금액", "상태"])
    for i in range(2, 15):
        input_sheet[f"A{i}"] = f"2026-07-{i:02d}"
        input_sheet[f"B{i}"] = f"항목{i - 1}"
        input_sheet[f"C{i}"] = i
        input_sheet[f"D{i}"] = 100 + i
        input_sheet[f"E{i}"] = (100 + i) * i
        input_sheet[f"F{i}"] = "진행중"

    for quarter, month in (("1분기", "1월"), ("2분기", "4월"), ("3분기", "7월")):
        q = wb.create_sheet(quarter)
        q.append(["월", "카테고리", "수량", "단가", "금액"])
        q.append([month, "A", 10, 100, 1000])
        q.append([month, "B", 8, 110, 880])
        q.append([month, "C", 7, 120, 840])

    if str(profile or "").strip().lower() == "sales_core":
        pass

    # 차트/요약용 기본 컬러 샘플(검증용)
    sales["A1"].fill = PatternFill(fill_type="solid", start_color="FFFDE7", end_color="FFFDE7")

    wb.save(path)
    wb.close()


def _seed_from_template(path: Path, template: Path) -> None:
    """실제 업무 파일(데모 워크북)을 그대로 복사해 시드로 쓴다.

    합성 시드는 머리글이 한국어라 "Region/Sales" 같은 실제 영문 표를 못 잡는 문제를 감춘다.
    """
    resolved = template if template.is_absolute() else (Path(__file__).resolve().parents[3] / template)
    if not resolved.exists():
        raise FileNotFoundError(f"시드 템플릿을 찾을 수 없습니다: {resolved}")
    shutil.copy2(resolved, path)


def _build_merge_folder(root: Path) -> Path:
    merge_dir = root / "merge_inputs"
    merge_dir.mkdir(parents=True, exist_ok=True)
    for idx in range(1, 3):
        wb = Workbook()
        ws = wb.active
        ws.title = "원본"
        ws.append(["월", "카테고리", "수량", "단가", "금액"])
        ws.append([f"{idx}월", "A", 10 + idx, 100, (10 + idx) * 100])
        ws.append([f"{idx}월", "B", 7 + idx, 120, (7 + idx) * 120])
        target = merge_dir / f"merge_{idx}.xlsx"
        wb.save(target)
        wb.close()
    return merge_dir


def _render_message(template: str, context: dict[str, str]) -> str:
    msg = str(template or "")
    for key, value in context.items():
        msg = msg.replace(f"{{{key}}}", str(value))
    return msg


def _extract_plan_actions(result_obj: dict[str, Any]) -> list[str]:
    if not isinstance(result_obj, dict):
        return []
    plan = result_obj.get("plan")
    if not isinstance(plan, list):
        return []
    rows: list[str] = []
    for step in plan:
        if not isinstance(step, dict):
            continue
        action = _normalize_action(step.get("action"))
        if action:
            rows.append(action)
    return rows


def _values_equal(actual: Any, expected: Any) -> bool:
    if expected is None:
        return actual is None or str(actual or "").strip() == ""
    try:
        af = float(actual)
        ef = float(expected)
        return abs(af - ef) <= 1e-9
    except Exception:
        pass
    return str(actual).strip() == str(expected).strip()


def _count_non_empty_cells(ws, range_ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is None:
                continue
            if str(cell.value).strip() == "":
                continue
            count += 1
    return count


def _ranges_intersect(left: CellRange, right: CellRange) -> bool:
    return not (
        left.max_col < right.min_col
        or left.min_col > right.max_col
        or left.max_row < right.min_row
        or left.min_row > right.max_row
    )


def _count_filled_cells(ws, range_ref: str) -> int:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    count = 0
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            pattern = str(getattr(cell.fill, "patternType", "") or "")
            if pattern not in {"", "none", "None"}:
                count += 1
    return count


def _column_values(ws, range_ref: str) -> list[Any]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    values: list[Any] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            values.append(cell.value)
    return values


def _check_assertion(wb, assertion: dict[str, Any], workbook_path: Path) -> tuple[bool, str]:
    atype = str(assertion.get("type", "")).strip()
    sheet_name = str(assertion.get("sheet", "")).strip()

    if atype == "sibling_file_exists":
        # PDF 내보내기처럼 결과물이 통합문서 밖에 생기는 작업. 파일이 실제로 생겼는지 본다.
        pattern = str(assertion.get("glob", "")).strip()
        matches = sorted(workbook_path.parent.glob(pattern))
        sized = [p for p in matches if p.stat().st_size > 0]
        ok = bool(sized)
        return ok, f"sibling_file_exists glob={pattern} matched={[p.name for p in matches]}"

    if atype == "sheet_exists":
        ok = sheet_name in wb.sheetnames
        return ok, f"sheet_exists sheet={sheet_name} actual={ok}"

    if sheet_name and sheet_name not in wb.sheetnames:
        return False, f"sheet_not_found sheet={sheet_name}"

    ws = wb[sheet_name] if sheet_name else wb.active

    if atype == "cell_equals":
        cell = str(assertion.get("cell", "")).strip()
        expected = assertion.get("expected")
        actual = ws[cell].value
        ok = _values_equal(actual, expected)
        return ok, f"cell_equals {sheet_name}!{cell} expected={expected} actual={actual}"

    if atype == "cell_formula_startswith":
        cell = str(assertion.get("cell", "")).strip()
        prefix = str(assertion.get("prefix", "")).strip()
        actual = str(ws[cell].value or "")
        ok = actual.startswith(prefix)
        return ok, f"cell_formula_startswith {sheet_name}!{cell} prefix={prefix} actual={actual}"

    if atype == "range_non_empty_at_least":
        range_ref = str(assertion.get("range", "")).strip()
        min_count = _safe_int(assertion.get("min_count", 0))
        actual = _count_non_empty_cells(ws, range_ref)
        ok = actual >= min_count
        return ok, f"range_non_empty_at_least {sheet_name}!{range_ref} min={min_count} actual={actual}"

    if atype == "sheet_protected":
        expected = bool(assertion.get("expected", True))
        actual = bool(getattr(ws.protection, "sheet", False))
        ok = actual == expected
        return ok, f"sheet_protected {sheet_name} expected={expected} actual={actual}"

    if atype == "data_validation_exists":
        target_range = str(assertion.get("target_range", "")).strip()
        target = CellRange(target_range)
        found = False
        for dv in list(getattr(ws.data_validations, "dataValidation", [])):
            try:
                ranges = list(dv.ranges.ranges)
            except Exception:
                ranges = []
            for row in ranges:
                if _ranges_intersect(target, row):
                    found = True
                    break
            if found:
                break
        return found, f"data_validation_exists {sheet_name}!{target_range} actual={found}"

    if atype == "chart_count_at_least":
        min_charts = _safe_int(assertion.get("min_charts", 1))
        actual = len(getattr(ws, "_charts", []))
        ok = actual >= min_charts
        return ok, f"chart_count_at_least {sheet_name} min={min_charts} actual={actual}"

    if atype == "cell_has_border":
        cell = str(assertion.get("cell", "")).strip()
        border = ws[cell].border
        styles = [
            getattr(border.left, "style", None),
            getattr(border.right, "style", None),
            getattr(border.top, "style", None),
            getattr(border.bottom, "style", None),
        ]
        ok = any(bool(style) for style in styles)
        return ok, f"cell_has_border {sheet_name}!{cell} styles={styles}"

    if atype == "cell_has_fill":
        cell = str(assertion.get("cell", "")).strip()
        fill = ws[cell].fill
        pattern = str(getattr(fill, "patternType", "") or "")
        ok = pattern not in {"", "none", "None"}
        return ok, f"cell_has_fill {sheet_name}!{cell} pattern={pattern}"

    if atype == "filled_cell_count_equals":
        # 조건부서식이 "정확히 몇 칸"을 칠했는지. 0칸(조건 미적용)과 전부 칠하기를 함께 걸러낸다.
        range_ref = str(assertion.get("range", "")).strip()
        expected = _safe_int(assertion.get("expected", 0))
        actual = _count_filled_cells(ws, range_ref)
        ok = actual == expected
        return ok, f"filled_cell_count_equals {sheet_name}!{range_ref} expected={expected} actual={actual}"

    if atype == "range_sum_equals":
        # 집계 결과가 원본 합계와 같은지. "시트가 생겼다"만 보면 값이 틀려도 통과한다.
        range_ref = str(assertion.get("range", "")).strip()
        expected = float(assertion.get("expected", 0) or 0)
        tolerance = float(assertion.get("tolerance", 1.0) or 0)
        actual = sum(
            float(value)
            for value in _column_values(ws, range_ref)
            if isinstance(value, (int, float)) and not isinstance(value, bool)
        )
        ok = abs(actual - expected) <= tolerance
        return ok, f"range_sum_equals {sheet_name}!{range_ref} expected={expected} actual={actual}"

    if atype == "row_count_between":
        # 피벗이 고유키로 묶여 "집계가 안 된" 결과를 걸러낸다.
        range_ref = str(assertion.get("range", "")).strip()
        low = _safe_int(assertion.get("min_rows", 0))
        high = _safe_int(assertion.get("max_rows", 10**9))
        actual = sum(
            1
            for value in _column_values(ws, range_ref)
            if value is not None and str(value).strip() != ""
        )
        ok = low <= actual <= high
        return ok, f"row_count_between {sheet_name}!{range_ref} range=[{low},{high}] actual={actual}"

    if atype == "column_values_all_in":
        # 필터가 "행이 줄었다"만으로는 부족하다. 남은 값이 전부 조건을 만족해야 한다.
        range_ref = str(assertion.get("range", "")).strip()
        allowed = {str(v).strip() for v in (assertion.get("allowed") or [])}
        actual = [
            str(value).strip()
            for value in _column_values(ws, range_ref)
            if value is not None and str(value).strip() != ""
        ]
        offenders = sorted({value for value in actual if value not in allowed})
        ok = not offenders and bool(actual)
        return ok, (
            f"column_values_all_in {sheet_name}!{range_ref} n={len(actual)} offenders={offenders[:5]}"
        )

    if atype == "header_equals":
        cell = str(assertion.get("cell", "")).strip()
        expected = str(assertion.get("expected", "")).strip()
        actual = str(ws[cell].value or "").strip()
        ok = actual == expected
        return ok, f"header_equals {sheet_name}!{cell} expected={expected} actual={actual}"

    if atype == "header_absent":
        # 열 삭제 확인. 머리글 행에 그 이름이 더는 없어야 한다.
        name = str(assertion.get("name", "")).strip()
        headers = [str(c.value or "").strip() for c in ws[1]]
        ok = name not in headers
        return ok, f"header_absent {sheet_name} name={name} headers={headers[:20]}"

    if atype == "column_sorted":
        range_ref = str(assertion.get("range", "")).strip()
        descending = bool(assertion.get("descending", False))
        values = [v for v in _column_values(ws, range_ref) if v is not None]
        try:
            pairs = list(pairwise(values))
            ok = all((b <= a) if descending else (a <= b) for a, b in pairs)
        except TypeError:
            ok = False
        return ok, f"column_sorted {sheet_name}!{range_ref} desc={descending} n={len(values)} ok={ok}"

    return False, f"unsupported_assertion_type type={atype}"


def _check_turn_expectation(turn_result: dict[str, Any], exp: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    status_code = _safe_int(turn_result.get("status_code"))
    action = _normalize_action(turn_result.get("action"))
    ask_follow_up = bool(turn_result.get("ask_follow_up", False))
    result_obj = turn_result.get("result") if isinstance(turn_result.get("result"), dict) else {}

    if "status_code" in exp:
        expected_status = _safe_int(exp.get("status_code"))
        if status_code != expected_status:
            errors.append(f"status_code expected={expected_status} actual={status_code}")
    if "status_in" in exp:
        allowed = {_safe_int(v) for v in (exp.get("status_in") or [])}
        if allowed and status_code not in allowed:
            errors.append(f"status_in expected={sorted(allowed)} actual={status_code}")
    if "action" in exp:
        expected_action = _normalize_action(exp.get("action"))
        if action != expected_action:
            errors.append(f"action expected={expected_action} actual={action}")
    if "action_in" in exp:
        allowed_actions = {_normalize_action(v) for v in (exp.get("action_in") or []) if str(v).strip()}
        if allowed_actions and action not in allowed_actions:
            errors.append(f"action_in expected={sorted(allowed_actions)} actual={action}")
    if "ask_follow_up" in exp:
        expected_follow_up = bool(exp.get("ask_follow_up"))
        if ask_follow_up != expected_follow_up:
            errors.append(f"ask_follow_up expected={expected_follow_up} actual={ask_follow_up}")

    numeric_gte = exp.get("result_numeric_gte")
    if isinstance(numeric_gte, dict):
        for key, expected_value in numeric_gte.items():
            actual_value = _safe_float(result_obj.get(str(key)))
            if actual_value < float(expected_value):
                errors.append(
                    f"result_numeric_gte key={key} expected>={expected_value} actual={actual_value}"
                )

    result_equals = exp.get("result_equals")
    if isinstance(result_equals, dict):
        for key, expected_value in result_equals.items():
            actual_value = result_obj.get(str(key))
            if not _values_equal(actual_value, expected_value):
                errors.append(f"result_equals key={key} expected={expected_value} actual={actual_value}")

    return errors


async def _run_turn(
    *,
    llm,
    message: str,
    workbook_id: str,
    sheet_name: str,
    session_id: str,
    approve: bool,
    timeout_seconds: float,
) -> dict[str, Any]:
    t0 = time.perf_counter()
    try:
        req = ExcelLiveCommandRequest(
            message=message,
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            session_id=session_id,
            approve=approve,
        )
        response = await asyncio.wait_for(
            post_command(req=req, llm=llm),
            timeout=timeout_seconds,
        )
        result_obj = response.result if isinstance(response.result, dict) else {}
        return {
            "message": message,
            "status_code": 200,
            "ok": bool(response.ok),
            "action": str(response.action or ""),
            "reason": str(response.reason or ""),
            "ask_follow_up": bool(result_obj.get("ask_follow_up", False)),
            "approval_required": bool(response.approval_required),
            "result": result_obj,
            "plan_actions": _extract_plan_actions(result_obj),
            "elapsed_ms": int((time.perf_counter() - t0) * 1000),
            "error": "",
        }
    except HTTPException as exc:
        return {
            "message": message,
            "status_code": int(exc.status_code),
            "ok": False,
            "action": "",
            "reason": "",
            "ask_follow_up": False,
            "approval_required": False,
            "result": {},
            "plan_actions": [],
            "elapsed_ms": int((time.perf_counter() - t0) * 1000),
            "error": str(exc.detail),
        }
    except Exception as exc:
        return {
            "message": message,
            "status_code": 0,
            "ok": False,
            "action": "",
            "reason": "",
            "ask_follow_up": False,
            "approval_required": False,
            "result": {},
            "plan_actions": [],
            "elapsed_ms": int((time.perf_counter() - t0) * 1000),
            "error": f"{type(exc).__name__}: {exc}",
        }


def _observed_actions(turns: list[dict[str, Any]]) -> list[str]:
    rows: list[str] = []
    for turn in turns:
        top = _normalize_action(turn.get("action"))
        if top:
            rows.append(top)
        for step_action in turn.get("plan_actions") or []:
            norm = _normalize_action(step_action)
            if norm:
                rows.append(norm)
    return sorted(set(rows))


async def _run_scenario(
    *,
    llm,
    scenario: dict[str, Any],
    defaults: dict[str, Any],
    root: Path,
    turn_timeout_seconds: float,
) -> dict[str, Any]:
    scenario_id = str(scenario.get("id", "unknown"))
    title = str(scenario.get("title", ""))
    severity = str(scenario.get("severity", "major")).strip().lower() or "major"
    sheet_name = str(scenario.get("sheet_name") or defaults.get("sheet_name") or "매출")
    seed_profile = str(scenario.get("seed_profile") or defaults.get("seed_profile") or "sales_core")
    approve = bool(scenario.get("approve", defaults.get("approve", True)))

    scenario_root = root / scenario_id
    scenario_root.mkdir(parents=True, exist_ok=True)
    workbook_path = scenario_root / "scenario.xlsx"
    seed_template = str(scenario.get("seed_template") or defaults.get("seed_template") or "").strip()
    if seed_template:
        _seed_from_template(workbook_path, Path(seed_template))
    else:
        _build_seed_workbook(workbook_path, seed_profile)
    merge_dir = _build_merge_folder(scenario_root)

    context = {
        "WORKBOOK_PATH": str(workbook_path),
        "MERGE_DIR": str(merge_dir).replace("\\", "/"),
    }
    session_id = f"complex-{uuid.uuid4().hex[:10]}"

    turns_raw = scenario.get("turns") if isinstance(scenario.get("turns"), list) else []
    turn_results: list[dict[str, Any]] = []
    for turn in turns_raw:
        if not isinstance(turn, dict):
            continue
        msg = _render_message(str(turn.get("message", "")), context)
        turn_result = await _run_turn(
            llm=llm,
            message=msg,
            workbook_id=str(workbook_path),
            sheet_name=sheet_name,
            session_id=session_id,
            approve=approve,
            timeout_seconds=turn_timeout_seconds,
        )
        turn_results.append(turn_result)

    oracle = scenario.get("oracle") if isinstance(scenario.get("oracle"), dict) else {}
    conversation = oracle.get("conversation") if isinstance(oracle.get("conversation"), dict) else {}
    execution = oracle.get("execution") if isinstance(oracle.get("execution"), dict) else {}
    result_oracle = oracle.get("result") if isinstance(oracle.get("result"), dict) else {}

    conversation_errors: list[str] = []
    expected_turns = (
        conversation.get("turn_expectations")
        if isinstance(conversation.get("turn_expectations"), list)
        else []
    )
    expected_indexes: set[int] = set()
    for exp in expected_turns:
        if not isinstance(exp, dict):
            continue
        idx = _safe_int(exp.get("turn_index"))
        expected_indexes.add(idx)
        if idx < 0 or idx >= len(turn_results):
            conversation_errors.append(f"turn_index_out_of_range index={idx}")
            continue
        errs = _check_turn_expectation(turn_results[idx], exp)
        for err in errs:
            conversation_errors.append(f"turn[{idx}] {err}")

    for idx, turn_result in enumerate(turn_results):
        if idx in expected_indexes:
            continue
        if _safe_int(turn_result.get("status_code")) != 200:
            conversation_errors.append(
                f"turn[{idx}] status_code default-expected=200 actual={turn_result.get('status_code')}"
            )

    observed_actions = _observed_actions(turn_results)
    execution_errors: list[str] = []
    must_include = execution.get("must_include_actions") if isinstance(execution.get("must_include_actions"), list) else []
    for required_action in must_include:
        normalized = _normalize_action(required_action)
        if normalized and normalized not in observed_actions:
            execution_errors.append(f"must_include_actions missing={normalized}")
    forbid_actions = execution.get("forbid_actions") if isinstance(execution.get("forbid_actions"), list) else []
    for forbid in forbid_actions:
        normalized = _normalize_action(forbid)
        if normalized and normalized in observed_actions:
            execution_errors.append(f"forbid_actions detected={normalized}")

    assertion_rows = result_oracle.get("assertions") if isinstance(result_oracle.get("assertions"), list) else []
    assertion_results: list[dict[str, Any]] = []
    result_errors: list[str] = []
    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception as exc:
        wb = None
        result_errors.append(f"workbook_open_failed: {exc}")
    if wb is not None:
        try:
            for idx, assertion in enumerate(assertion_rows):
                if not isinstance(assertion, dict):
                    continue
                ok, detail = _check_assertion(wb, assertion, workbook_path)
                assertion_results.append(
                    {
                        "index": idx,
                        "type": str(assertion.get("type", "")),
                        "ok": bool(ok),
                        "detail": detail,
                    }
                )
                if not ok:
                    result_errors.append(f"assertion[{idx}] {detail}")
        finally:
            wb.close()

    passed = not conversation_errors and not execution_errors and not result_errors
    critical_failure = bool((severity == "critical") and (not passed))
    return {
        "id": scenario_id,
        "title": title,
        "category": str(scenario.get("category", "")),
        "difficulty": str(scenario.get("difficulty", "")),
        "severity": severity,
        "sheet_name": sheet_name,
        "workbook_path": str(workbook_path),
        "session_id": session_id,
        "observed_actions": observed_actions,
        "passed": passed,
        "critical_failure": critical_failure,
        "turns": turn_results,
        "assertion_results": assertion_results,
        "errors": {
            "conversation": conversation_errors,
            "execution": execution_errors,
            "result": result_errors,
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="복잡 업무형 Excel 시나리오 자동 검증")
    parser.add_argument(
        "--scenario-pack",
        type=Path,
        default=Path(__file__).resolve().parents[3] / "datasets" / "excel_complex_scenarios_v1.json",
    )
    parser.add_argument(
        "--output-json",
        type=Path,
        default=Path(__file__).resolve().parents[3] / "logs" / "excel_complex_verify_report.json",
    )
    parser.add_argument("--model", type=str, default="", help="검증 시 사용할 planner 모델(선택)")
    parser.add_argument("--turn-timeout-seconds", type=float, default=45.0)
    parser.add_argument("--limit", type=int, default=0, help="앞에서부터 N개 시나리오만 실행 (0=전체)")
    parser.add_argument(
        "--scenario-id",
        type=str,
        default="",
        help="쉼표로 구분한 시나리오 id만 실행 (디버깅용)",
    )
    parser.add_argument(
        "--artifact-dir",
        type=Path,
        default=None,
        help="실행 후 워크북을 남길 디렉터리 (미지정 시 임시 폴더로 삭제)",
    )
    parser.add_argument("--stop-on-fail", action="store_true")
    return parser.parse_args()


@contextlib.contextmanager
def _artifact_root(artifact_dir: Path | None) -> Iterator[Path]:
    """시나리오 워크북을 둘 디렉터리. 지정이 없으면 끝나고 지운다."""
    if artifact_dir is None:
        with tempfile.TemporaryDirectory(prefix="officeclaw_complex_verify_") as td:
            yield Path(td)
        return
    artifact_dir.mkdir(parents=True, exist_ok=True)
    yield artifact_dir


async def _run(args: argparse.Namespace) -> dict[str, Any]:
    pack = _load_pack(args.scenario_pack)
    defaults = pack.get("defaults") if isinstance(pack.get("defaults"), dict) else {}
    scenarios_raw = pack.get("scenarios") if isinstance(pack.get("scenarios"), list) else []
    scenarios = [row for row in scenarios_raw if isinstance(row, dict)]
    wanted = {part.strip() for part in str(args.scenario_id or "").split(",") if part.strip()}
    if wanted:
        scenarios = [row for row in scenarios if str(row.get("id", "")) in wanted]
    if int(args.limit or 0) > 0:
        scenarios = scenarios[: int(args.limit)]

    llm = get_llm_service()
    results: list[dict[str, Any]] = []
    with _artifact_root(args.artifact_dir) as root:
        for idx, scenario in enumerate(scenarios, start=1):
            result = await _run_scenario(
                llm=llm,
                scenario=scenario,
                defaults=defaults,
                root=root,
                turn_timeout_seconds=float(args.turn_timeout_seconds),
            )
            results.append(result)
            print(
                f"[{idx}/{len(scenarios)}] {result['id']} passed={result['passed']} "
                f"critical_failure={result['critical_failure']}",
                flush=True,
            )
            if args.stop_on_fail and not result["passed"]:
                break

    total = len(results)
    passed = sum(1 for row in results if row.get("passed"))
    critical_failures = sum(1 for row in results if row.get("critical_failure"))
    elapsed_values = [
        _safe_int(turn.get("elapsed_ms"))
        for row in results
        for turn in (row.get("turns") or [])
        if isinstance(turn, dict)
    ]
    avg_turn_ms = int(sum(elapsed_values) / len(elapsed_values)) if elapsed_values else 0
    failed_ids = [str(row.get("id")) for row in results if not row.get("passed")]
    payload = {
        "schema_version": "excel_complex_verify_report.v1",
        "at": _now_iso(),
        "scenario_pack": str(args.scenario_pack),
        "model_for_parse": str(args.model or load_llm_config().get("model", "")),
        "total_scenarios": total,
        "passed_scenarios": passed,
        "pass_rate": round((passed / total) if total else 0.0, 4),
        "critical_failures": critical_failures,
        "average_turn_latency_ms": avg_turn_ms,
        "failed_scenarios": failed_ids[:50],
        "results": results,
    }
    return payload


async def main() -> None:
    args = parse_args()
    original_config = load_llm_config()
    try:
        model = str(args.model or "").strip()
        if model:
            save_llm_config({"provider": "ollama", "model": model})
            reload_llm_service()
        payload = await _run(args)
    finally:
        save_llm_config(original_config)
        reload_llm_service()

    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    print("")
    print("=== Excel 복잡 시나리오 자동 검증 ===")
    print(f"total_scenarios={payload['total_scenarios']}")
    print(f"passed_scenarios={payload['passed_scenarios']}")
    print(f"pass_rate={payload['pass_rate']}")
    print(f"critical_failures={payload['critical_failures']}")
    print(f"average_turn_latency_ms={payload['average_turn_latency_ms']}")
    print(f"output={args.output_json}")


if __name__ == "__main__":
    asyncio.run(main())
