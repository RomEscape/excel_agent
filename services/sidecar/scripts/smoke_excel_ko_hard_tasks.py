from __future__ import annotations

import argparse
import asyncio
import json
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from office_claw_sidecar.routers.excel_live import ExcelLiveCommandRequest, post_command
from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
from office_claw_sidecar.services.llm_service import (
    get_llm_service,
    load_llm_config,
    reload_llm_service,
    save_llm_config,
)


def _build_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출"
    ws.append(["월", "카테고리", "수량", "단가", "금액"])
    ws.append(["1월", "A", 10, 100, 1000])
    ws.append(["1월", "B", 8, 110, 880])
    ws.append(["2월", "A", 7, 120, 840])
    ws.append(["2월", "B", 11, 90, 990])
    ws.append(["3월", "A", 9, 130, 1170])
    ws.append(["3월", "A", 9, 130, 1170])  # dedupe 테스트용 중복

    ws2 = wb.create_sheet("비교")
    ws2.append(["월", "카테고리", "수량", "단가", "금액"])
    ws2.append(["1월", "A", 10, 100, 1000])
    ws2.append(["1월", "B", 8, 111, 888])  # diff 유도
    ws2.append(["2월", "A", 7, 120, 840])
    ws2.append(["2월", "B", 11, 90, 990])
    ws2.append(["3월", "A", 9, 130, 1170])
    ws2.append(["3월", "B", 9, 130, 1171])  # diff 유도

    ws3 = wb.create_sheet("Q2")
    ws3.append(["월", "카테고리", "수량", "단가", "금액"])
    ws3.append(["4월", "A", 9, 150, 1350])
    ws3.append(["4월", "B", 10, 155, 1550])

    wb.save(path)
    wb.close()


async def run_parse_ko_smoke(model: str, timeout_s: float, workbook_id: str) -> dict[str, Any]:
    original = load_llm_config()
    rows: list[dict[str, Any]] = []
    cases = [
        {
            "id": "create_table",
            "prompt": "매출 시트 B2부터 6행 4열 표 만들어줘",
            "expected_actions": {"excel_live.create_table"},
            "follow_ups": [],
        },
        {
            "id": "sort_range",
            "prompt": "매출 시트 A1:E8 범위를 금액 열 기준 내림차순으로 정렬해줘",
            "expected_actions": {"excel_live.sort_range"},
            "follow_ups": [],
        },
        {
            "id": "pivot_table",
            "prompt": "매출 시트 A1:E8에서 월을 행으로, 카테고리를 열로 두고 금액 합계 피벗을 피벗 시트 A1에 만들어줘",
            "expected_actions": {"excel_live.pivot_table"},
            "follow_ups": [
                "행 기준은 월, 값 열은 금액, 열 기준은 카테고리로 해줘",
                "원본 범위는 A1:E8이고 집계 방식은 합계야",
            ],
        },
        {
            "id": "compare_ranges",
            "prompt": "매출 시트 A1:E8이랑 비교 시트 A1:E8을 비교해서 비교결과 시트에 차이 써줘",
            "expected_actions": {"excel_live.compare_ranges"},
            "follow_ups": [],
        },
        {
            "id": "forecast",
            "prompt": "매출 시트 C2:C8 기준으로 앞으로 4개월 예측해서 예측 시트 A1에 써줘",
            "expected_actions": {"excel_live.forecast_linear"},
            "follow_ups": [],
        },
        {
            "id": "set_data_validation",
            "prompt": "E2:E200에는 0부터 100000 사이 숫자만 입력되게 제한해줘",
            "expected_actions": {"excel_live.set_data_validation"},
            "follow_ups": [],
        },
        {
            "id": "set_formula",
            "prompt": "매출 시트 F2:F8에 =C2*D2 수식 넣어줘",
            "expected_actions": {"excel_live.set_formula"},
            "follow_ups": [],
        },
    ]
    try:
        save_llm_config({"provider": "ollama", "model": model})
        reload_llm_service()
        llm = get_llm_service()
        for case in cases:
            t0 = time.perf_counter()
            action = ""
            err = ""
            ok = False
            follow_up_count = 0
            try:
                req = ExcelLiveCommandRequest(
                    message=case["prompt"],
                    workbook_id=workbook_id,
                    sheet_name="매출",
                    session_id=f"ko-hard-{case['id']}",
                )
                response = await asyncio.wait_for(
                    post_command(
                        req=req,
                        llm=llm,
                    ),
                    timeout=timeout_s,
                )
                action = str(response.action).strip()
                if isinstance(response.result, dict) and response.result.get("ask_follow_up"):
                    for follow in case.get("follow_ups", []):
                        follow_up_count += 1
                        req = ExcelLiveCommandRequest(
                            message=str(follow),
                            workbook_id=workbook_id,
                            sheet_name="매출",
                            session_id=f"ko-hard-{case['id']}",
                        )
                        response = await asyncio.wait_for(
                            post_command(
                                req=req,
                                llm=llm,
                            ),
                            timeout=timeout_s,
                        )
                        action = str(response.action).strip()
                        if not (isinstance(response.result, dict) and response.result.get("ask_follow_up")):
                            break
                ok = action in case["expected_actions"]
            except Exception as exc:
                err = str(exc)
            rows.append(
                {
                    "id": case["id"],
                    "prompt": case["prompt"],
                    "expected_actions": sorted(case["expected_actions"]),
                    "action": action,
                    "ok": ok,
                    "elapsed_ms": int((time.perf_counter() - t0) * 1000),
                    "follow_up_count": follow_up_count,
                    "error": err,
                }
            )
    finally:
        save_llm_config(original)
        reload_llm_service()

    success = sum(1 for r in rows if r["ok"])
    return {
        "model": model,
        "total": len(rows),
        "success": success,
        "accuracy": round((success / len(rows)) if rows else 0.0, 4),
        "rows": rows,
    }


def run_execution_hard_smoke(workbook_path: Path) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    service = FileExcelLiveService(workspace_root=workbook_path.parent)
    service.select_workbook(workbook_path.name)

    def record(task: str, ok: bool, detail: dict[str, Any] | None = None, error: str = "") -> None:
        rows.append(
            {
                "task": task,
                "ok": bool(ok),
                "detail": detail or {},
                "error": error,
            }
        )

    try:
        result = service.sort_range(
            workbook_id=None,
            sheet_name="매출",
            target_range="A1:E8",
            key_column="금액",
            order="desc",
            has_header=True,
        )
        record("sort_range", bool(result.get("sorted_rows", 0) >= 1), result)
    except Exception as exc:
        record("sort_range", False, error=str(exc))

    try:
        result = service.dedupe_rows(
            workbook_id=None,
            sheet_name="매출",
            target_range="A1:E8",
            key_columns=["월", "카테고리", "수량", "단가", "금액"],
            has_header=True,
        )
        record("dedupe_rows", bool(result.get("removed_rows", 0) >= 1), result)
    except Exception as exc:
        record("dedupe_rows", False, error=str(exc))

    try:
        result = service.filter_rows(
            workbook_id=None,
            sheet_name="매출",
            target_range="A1:E8",
            column="금액",
            operator=">=",
            value=1000,
            has_header=True,
        )
        record("filter_rows", bool(result.get("filtered_rows", 0) >= 1), result)
    except Exception as exc:
        record("filter_rows", False, error=str(exc))

    try:
        result = service.forecast_linear(
            workbook_id=None,
            sheet_name="매출",
            source_range="C2:C8",
            horizon=4,
            output_sheet="예측",
            output_start="A1",
        )
        record("forecast_linear", bool(result.get("created", False)), result)
    except Exception as exc:
        record("forecast_linear", False, error=str(exc))

    try:
        result = service.pivot_table(
            workbook_id=None,
            sheet_name="매출",
            source_range="A1:E8",
            row_field="월",
            value_field="금액",
            agg="sum",
            column_field="카테고리",
            output_sheet="피벗",
            output_start="A1",
            has_header=True,
        )
        record("pivot_table", bool(result.get("created", False)), result)
    except Exception as exc:
        record("pivot_table", False, error=str(exc))

    try:
        service.set_formula(
            workbook_id=None,
            sheet_name="매출",
            range_ref="E2:E8",
            formula_a1="=C2*D2",
        )
        result = service.verify_formula_result(
            workbook_id=None,
            sheet_name="매출",
            range_ref="E2:E8",
        )
        record("set_formula+verify_formula_result", bool(result.get("non_empty_cells", 0) >= 1), result)
    except Exception as exc:
        record("set_formula+verify_formula_result", False, error=str(exc))

    try:
        result = service.compare_ranges(
            workbook_id=None,
            left_sheet="매출",
            left_range="A1:E8",
            right_sheet="비교",
            right_range="A1:E8",
            output_sheet="비교결과",
        )
        record("compare_ranges", bool(result.get("diff_cells", 0) >= 1), result)
    except Exception as exc:
        record("compare_ranges", False, error=str(exc))

    try:
        result = service.consolidate_sheets(
            workbook_id=None,
            source_sheets=["매출", "Q2"],
            output_sheet="통합",
            include_header_once=True,
            add_source_sheet_col=True,
        )
        record("consolidate_sheets", bool(result.get("created", False)), result)
    except Exception as exc:
        record("consolidate_sheets", False, error=str(exc))

    try:
        result = service.set_data_validation(
            workbook_id=None,
            sheet_name="매출",
            target_range="E2:E200",
            validation_type="decimal",
            minimum=0,
            maximum=100000,
            allow_blank=True,
            show_error=True,
            error_message="0~100000 범위만 입력 가능합니다.",
        )
        record("set_data_validation", bool(result.get("applied", False)), result)
    except Exception as exc:
        record("set_data_validation", False, error=str(exc))

    success = sum(1 for r in rows if r["ok"])
    return {
        "total": len(rows),
        "success": success,
        "accuracy": round((success / len(rows)) if rows else 0.0, 4),
        "rows": rows,
    }


async def main() -> None:
    parser = argparse.ArgumentParser(description="한국어 고난도 엑셀 작업 파싱/실행 스모크")
    parser.add_argument("--model", type=str, default="skt/A.X-4.0-Light:latest")
    parser.add_argument("--parse-timeout", type=float, default=35.0)
    args = parser.parse_args()

    with tempfile.TemporaryDirectory(prefix="officeclaw_hard_ko_") as td:
        root = Path(td)
        workbook = root / "hard_tasks.xlsx"
        _build_workbook(workbook)

        parse_report = await run_parse_ko_smoke(
            model=args.model,
            timeout_s=float(args.parse_timeout),
            workbook_id=str(workbook),
        )
        exec_report = run_execution_hard_smoke(workbook)

        payload = {
            "at": datetime.now().isoformat(timespec="seconds"),
            "model_for_parse": args.model,
            "korean_command_e2e_hard_tasks": parse_report,
            "execution_hard_tasks": exec_report,
            "workbook_temp": str(workbook),
        }
        print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    asyncio.run(main())

