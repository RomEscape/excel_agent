r"""명령 커버리지 v2 — 게이트 코퍼스에 **없던** 작업 종류로 인식·계획 정확도를 잰다.

    & $PY scripts\run_coverage_recognition.py                       # 전체
    & $PY scripts\run_coverage_recognition.py --only dedupe filter  # 과제 지정

왜(2026-08-25, 사용자 요청): "다양하게 명령 커버리지를 넓혀서 질문을 집어넣었을 때
정확히 뭘 해야 되는지 인식하는지, 그에 맞는 계획도 정확하게 세우는지". 말투 624·파괴 72는
32개 과제뿐이라 편집 액션 43개 중 21개가 한 번도 최종 액션으로 나온 적이 없다. 이 묶음은
그 빈자리(중복 제거·필터·피벗·조회·시트 삭제·열 연산·병합 해제·색조·초기화·표 원샷·
유효성·보호·메모·엑셀 표·이름 정의) + 오늘 GUI 사고 문형(값 없는 쓰기·정정)이다.

판정 두 겹:
  인식(recognition): 라우터의 **최종 결정** 액션이 기대 액션군에 드는가 (되묻기 기대면 되묻기)
  계획(plan):        실행까지 갔다면 ok로 끝났고, 과제별 파일 오라클(있는 경우)이 맞는가
결과 셋: OK(인식·계획 모두) · MISREAD(다른 종류로 실행 — 위험) · ASK(되묻기, 기대 밖) ·
        PLAN_BAD(종류는 맞는데 실행 실패/오라클 불일치) · ERROR

주의: LLM(플래너·의도 해석)이 실제로 불린다 — 게이트·배터리와 동시에 돌리지 않는다.
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import tempfile
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

os.environ.setdefault("PYTHONUTF8", "1")
os.environ["EXCEL_LIVE_ENGINE"] = "file"
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.models.approval import ApprovalResponse
from office_claw_sidecar.routers import excel_live as router
from office_claw_sidecar.routers.excel_live import ExcelLiveCommandRequest
from office_claw_sidecar.services.llm_service import get_llm_service

ROOT = Path(__file__).resolve().parent.parent.parent
CASES = ROOT / "datasets/eval/coverage_v2.jsonl"
PASS = {"PASS_RULE", "PASS_CARD"}

SEED = [
    ["날짜", "지역", "담당자", "금액", "상태", "비고"],
    ["2026-01-02", "서울", "김", 1200, "완료", ""],
    ["2026-01-03", "부산", "이", 800, "대기", ""],
    ["2026-01-03", "부산", "이", 800, "대기", ""],  # 중복 행 하나
    ["2026-01-04", "서울", "박", 1500, "취소", ""],
    ["2026-01-05", "대전", "최", 300, "완료", ""],
    ["2026-01-06", "서울", "김", 2200, "대기", ""],
    ["2026-01-07", "부산", "정", 950, "완료", ""],
    ["2026-01-08", "대전", "최", 400, "완료", ""],
]


def _seed(path: Path, task: str = "") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "매출"
    for row in SEED:
        ws.append(row)
    if task == "unmerge" or task.endswith("_with_note"):
        # unmerge 과제용 — **사용 범위 안**에 심는다. 첫 판(0825-170320)은 H1:K1에 심어
        # `__USED_RANGE__`(A1:F9) 밖이라 4문장 전부 "병합 1건 남음"으로 나왔다 — 제품이 아니라
        # 하네스 결함이었다. 데이터 아래 11행에 메모 줄을 병합해 둔다(값 손실 없음).
        # **과제별로만** 심는다 — 모든 과제에 심었더니 dedupe가 MergedCell 예외로 죽었다
        # (두 번째 판). 그 결함 자체는 `*_with_note` 과제가 따로 잰다.
        ws["A11"] = "비고 메모"
        ws.merge_cells("A11:C11")
    wb.create_sheet("임시")  # delete_sheet 과제용
    wb.save(path)


# 과제별 파일 오라클 — 실행 뒤 워크북을 다시 읽어 "계획이 맞았는가"를 본다. 없으면 인식만 본다.
def _oracle(task: str, path: Path) -> str:
    wb = load_workbook(path)
    ws = wb["매출"] if "매출" in wb.sheetnames else wb.active
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if task == "delete_sheet":
        return "" if "임시" not in wb.sheetnames else "임시 시트가 남아 있음"
    if task == "drop_column":
        return "" if "비고" not in header else "비고 열이 남아 있음"
    if task == "add_column":
        return "" if "확인자" in header else f"확인자 열 없음 {header}"
    if task == "rename_column":
        return "" if ("매출액" in header and "금액" not in header) else f"머리글 {header}"
    if task == "unmerge":
        return "" if len(ws.merged_cells.ranges) == 0 else f"병합 {len(ws.merged_cells.ranges)}건 남음"
    if task.startswith("dedupe"):
        rows = [tuple(ws.cell(r, c).value for c in range(1, 7)) for r in range(2, ws.max_row + 1)]
        rows = [r for r in rows if any(v not in (None, "") for v in r)]
        return "" if len(rows) == len(set(rows)) else "중복 행이 남아 있음"
    if task == "create_table_oneshot":
        return "" if ws.cell(1, 1).value not in (None, "") else "A1이 비어 있음"
    return ""


async def _run_one(row: dict, work: Path, llm) -> dict:
    task = str(row["task"])
    xlsx = work / f"{task}.xlsx"
    _seed(xlsx, task)
    svc = router.get_excel_live_service()
    svc.select_workbook(str(xlsx))
    svc.select_sheet(None, "매출")
    router._pending_create_table_slots.clear()
    router._pending_operation_slots.clear()
    router._pending_clarifications.clear()
    session = f"test-coverage-{task}-{abs(hash(row['text'])) % 10_000}"
    t0 = time.time()
    try:
        resp = await router.post_command(
            ExcelLiveCommandRequest(message=row["text"], session_id=session, approve=False), llm
        )
        if getattr(resp, "approval_required", False) and getattr(resp, "pending_approval", None):
            resp = await router.post_approval(
                ApprovalResponse(approval_id=resp.pending_approval.approval_id, approved=True), llm
            )
    except Exception as exc:
        return {**row, "outcome": "ERROR", "action": "", "detail": f"{type(exc).__name__}: {exc}"[:160], "secs": round(time.time() - t0, 1)}
    result = resp.result if isinstance(resp.result, dict) else {}
    action = str(resp.action or "")
    asked = bool(result.get("ask_follow_up")) or action.endswith("clarify")
    expect = list(row.get("expect") or [])
    if row.get("expect_ask"):
        outcome = "OK" if asked else "MISREAD"
    elif asked:
        outcome = "ASK"
    elif action in expect:
        detail = _oracle(task, xlsx) if resp.ok else f"실행 실패: {str(resp.reason)[:80]}"
        outcome = "OK" if (resp.ok and not detail) else "PLAN_BAD"
        if detail:
            return {**row, "outcome": outcome, "action": action, "detail": detail, "secs": round(time.time() - t0, 1)}
    else:
        outcome = "MISREAD"
    return {**row, "outcome": outcome, "action": action, "detail": str(resp.reason or "")[:100], "secs": round(time.time() - t0, 1)}


async def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", nargs="*")
    args = ap.parse_args()
    rows = [json.loads(x) for x in CASES.read_text(encoding="utf-8").splitlines() if x.strip()]
    if args.only:
        rows = [r for r in rows if r["task"] in set(args.only)]
    run_id = datetime.now().strftime("%m%d-%H%M%S-coverage-v2")
    work = Path(tempfile.mkdtemp(prefix="coverage_v2_"))
    llm = get_llm_service()
    out: list[dict] = []
    for i, row in enumerate(rows, 1):
        res = await _run_one(row, work, llm)
        out.append(res)
        flag = {"OK": "OK ", "MISREAD": "BAD", "ASK": "ASK", "PLAN_BAD": "PLN", "ERROR": "ERR"}[res["outcome"]]
        print(f"[{i:3d}/{len(rows)}] {flag} {res['task']:20s} {res['text'][:36]:38s} → {res['action'].replace('excel_live.', ''):22s} {res['detail'][:50] if res['outcome'] != 'OK' else ''}", flush=True)

    total = Counter(r["outcome"] for r in out)
    per: dict[str, Counter] = defaultdict(Counter)
    for r in out:
        per[r["task"]][r["outcome"]] += 1
    recognized = sum(1 for r in out if r["outcome"] in ("OK", "PLAN_BAD"))
    lines = [
        f"# 명령 커버리지 v2 {run_id}",
        "",
        f"문장 {len(out)} · 과제 {len(per)}개 · **인식 정확 {recognized}/{len(out)} ({100 * recognized / max(len(out), 1):.1f}%)** · "
        f"인식+계획 OK {total['OK']} · 오인식 {total['MISREAD']} · 되묻기 {total['ASK']} · 계획 불량 {total['PLAN_BAD']} · 오류 {total['ERROR']}",
        "",
        "| 과제 | n | OK | 오인식 | 되묻기 | 계획불량 | 오류 |",
        "|---|---|---|---|---|---|---|",
    ]
    for task, c in per.items():
        n = sum(c.values())
        lines.append(f"| {task} | {n} | {c['OK']} | {c['MISREAD']} | {c['ASK']} | {c['PLAN_BAD']} | {c['ERROR']} |")
    bad = [r for r in out if r["outcome"] != "OK"]
    if bad:
        lines += ["", "## 실패 상세", ""]
        for r in bad:
            lines.append(f"- [{r['outcome']}] {r['task']} · {r['text'][:50]} → `{r['action']}` {r['detail'][:90]}")
    report_dir = ROOT / "logs"
    (report_dir / f"coverage_v2_{run_id}.md").write_text("\n".join(lines), encoding="utf-8")
    (CASES.with_name("coverage_v2_report.json")).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print("\n" + "\n".join(lines[:4]))
    print(f"\nrun_id={run_id} → logs/coverage_v2_{run_id}.md")
    return 0 if total["MISREAD"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
