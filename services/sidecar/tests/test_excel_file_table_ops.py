"""
파일 엔진의 표 단위 작업 — 열 이름으로 통계·정렬·열 편집·집계를 한다.

이 작업들은 도구 목록에는 오래전부터 올라 있었지만 파일 엔진에는 구현이 없어서,
"매출 합계 얼마야" 같은 평범한 요청이 실행 단계에서 조용히 터졌다.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
from office_claw_sidecar.services.excel_live_service import ExcelLiveError


@pytest.fixture()
def sales(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    rows = [
        ["지역", "상품", "매출"],
        ["서울", "사과", 6000000],
        ["부산", "배", 3000000],
        ["서울", "감", 8000000],
        ["대구", "사과", 3000000],
        ["부산", "배", 3000000],
    ]
    for row in rows:
        ws.append(row)
    path = tmp_path / "sales.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture()
def service(tmp_path: Path) -> FileExcelLiveService:
    return FileExcelLiveService(workspace_root=tmp_path)


def test_column_stat_sums_by_header_name(service, sales):
    result = service.calculate_column_stat(str(sales), "Sheet1", "매출", "sum")
    assert result["value"] == 23000000.0
    assert result["column"] == "C"
    assert result["header"] == "매출"
    assert result["numeric_count"] == 5


def test_column_stat_accepts_column_letter(service, sales):
    result = service.calculate_column_stat(str(sales), "Sheet1", "C", "average")
    assert result["value"] == pytest.approx(23000000.0 / 5)
    # 열 문자로 부른 경우엔 머리글 이름을 지어내지 않는다.
    assert result["header"] is None


def test_column_stat_counts_only_numbers(service, sales):
    result = service.calculate_column_stat(str(sales), "Sheet1", "지역", "count")
    assert result["value"] == 0.0
    assert result["numeric_count"] == 0


def test_column_stat_refuses_unknown_column(service, sales):
    with pytest.raises(ExcelLiveError) as caught:
        service.calculate_column_stat(str(sales), "Sheet1", "재고수량", "sum")
    # 어떤 열이 있는지 알려 줘야 사용자가 다음 말을 할 수 있다.
    assert "매출" in str(caught.value)


def test_column_stat_refuses_sum_without_numbers(service, sales):
    with pytest.raises(ExcelLiveError):
        service.calculate_column_stat(str(sales), "Sheet1", "지역", "sum")


def test_sort_rows_orders_whole_sheet_and_keeps_header(service, sales):
    result = service.sort_rows(str(sales), "Sheet1", "매출", "desc")
    assert result["sorted_rows"] == 5
    assert result["order"] == "desc"

    wb = openpyxl.load_workbook(str(sales))
    ws = wb["Sheet1"]
    assert ws["C1"].value == "매출"
    assert ws["C2"].value == 8000000
    assert ws["C3"].value == 6000000
    wb.close()


def test_dedupe_without_range_uses_whole_sheet(service, sales):
    result = service.dedupe_rows(str(sales), "Sheet1")
    assert result["kept_rows"] == 4
    assert result["removed_duplicates"] == 1


def test_dedupe_by_subset_columns(service, sales):
    result = service.dedupe_rows(str(sales), "Sheet1", columns=["지역"])
    assert result["kept_rows"] == 3
    assert result["removed_duplicates"] == 2


def test_drop_column_shifts_remaining_left(service, sales):
    result = service.drop_column(str(sales), "Sheet1", "상품")
    assert result["dropped_column"] == "상품"
    assert result["remaining_columns"] == 2

    wb = openpyxl.load_workbook(str(sales))
    ws = wb["Sheet1"]
    assert [ws.cell(row=1, column=i).value for i in (1, 2, 3)] == ["지역", "매출", None]
    assert ws["B2"].value == 6000000
    wb.close()


def test_rename_column_touches_header_only(service, sales):
    result = service.rename_column(str(sales), "Sheet1", "매출", "revenue")
    assert result["old_name"] == "매출"

    wb = openpyxl.load_workbook(str(sales))
    ws = wb["Sheet1"]
    assert ws["C1"].value == "revenue"
    assert ws["C2"].value == 6000000
    wb.close()


def test_add_column_translates_formula_per_row(service, sales):
    result = service.add_column(str(sales), "Sheet1", "이익", formula_a1="=C2*0.1")
    assert result["column"] == "D"
    assert result["formula_filled_cells"] == 5

    wb = openpyxl.load_workbook(str(sales))
    ws = wb["Sheet1"]
    assert ws["D1"].value == "이익"
    # 모든 행이 2행을 가리키면 열 전체가 같은 값이 된다. 행마다 옮겨져야 한다.
    assert ws["D2"].value == "=C2*0.1"
    assert ws["D6"].value == "=C6*0.1"
    wb.close()


def test_group_by_aggregate_sorts_by_value(service, sales):
    result = service.group_by_aggregate(
        str(sales), "Sheet1", group_column="지역", agg="sum", value_column="매출"
    )
    assert result["agg"] == "sum"
    assert result["groups"][0] == {"key": "서울", "value": 14000000.0, "count": 2}
    assert [g["key"] for g in result["groups"]] == ["서울", "부산", "대구"]


def test_group_by_aggregate_counts_without_value_column(service, sales):
    result = service.group_by_aggregate(str(sales), "Sheet1", group_column="상품", agg="count")
    assert {g["key"]: g["count"] for g in result["groups"]} == {"사과": 2, "배": 2, "감": 1}


def test_group_by_aggregate_requires_value_column_for_sum(service, sales):
    with pytest.raises(ExcelLiveError):
        service.group_by_aggregate(str(sales), "Sheet1", group_column="지역", agg="sum")


def test_recalculate_works_without_calc_properties(service, sales):
    """calcPr이 없는 워크북에서도 재계산 표시가 돼야 한다.

    openpyxl로 저장한 파일에는 calcPr이 없어 wb.calculation이 None이다.
    여기서 터지면 "집계 후 새로고침" 같은 다단계 계획이 마지막 단계에서 실패한다.
    """
    wb = openpyxl.load_workbook(str(sales))
    wb.calculation = None
    wb.save(str(sales))
    wb.close()
    stripped = openpyxl.load_workbook(str(sales))
    assert stripped.calculation is None
    stripped.close()

    result = service.recalculate(str(sales), "Sheet1")

    assert result["recalculated"] is True
    reloaded = openpyxl.load_workbook(str(sales))
    assert reloaded.calculation.fullCalcOnLoad is True
    reloaded.close()
