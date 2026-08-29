"""승인 경로가 계획의 나머지를 버리는지 잰다.

## 무엇을 재는가

한때 `/excel-live/command`는 계획에 CONFIRM 단계가 하나라도 있으면 **그 한 단계만**
`_pending_approvals`에 담고 즉시 반환했다. `post_approval`은 그 한 단계를
`_execute_action`으로 바로 실행하고 끝냈다. 나머지 단계도, 검증도, 롤백도 없었다.
지금은 계획 전체를 보관했다가 같은 실행 루프로 이어 붙인다. 이 수트는 그 손실이
0으로 유지되는지 계속 감시한다.

같은 계획을 두 경로로 태워 결과를 나란히 놓는다.

- **direct** — `approve: true`로 한 번에 보낸다. 실행 루프가 전부 돌고
  `_verify_step_result`와 롤백이 살아 있다. 사용자가 승인 버튼을 눌렀을 때
  *일어나야 하는* 일이다.
- **gated** — `approve: false`로 보내 승인 요청을 받고 `/excel-live/approval`로
  승인한다. 프론트(`WorkspacePage.jsx`)가 실제로 타는 경로다.

## 왜 기존 지표로는 안 보였나

154건 승격 게이트는 액션 이름만 채점하고 실행을 하지 않는다. 검증기 변이 수트
(`verifier_mutants.py`)는 계획을 실행기에 직접 주입해 승인 게이트를 건너뛴다.
라우터의 승인 테스트는 전부 단일 액션 `/action` 경로다. 그래서 이 결함이
살아남았다.

## 세 가지 손실을 따로 센다

1. **계획 이행률** — 계획한 단계 중 실제로 실행된 비율.
2. **검증 도달률** — 실행된 변경이 `_verify_step_result`를 거쳤는가.
3. **롤백 보호** — 실행기가 거짓말할 때(성공을 보고하면서 다른 값을 씀)
   원래 값으로 되돌아가는가.

3번이 가장 날카롭다. 승인 경로가 실행 루프를 우회하면 false pass 0%로 만들어 둔
검증기가 호출조차 되지 않아, 틀린 값이 그대로 파일에 남는다.
"""

from __future__ import annotations

import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

HEADERS = {"Authorization": "Bearer dev-token"}


# ── 케이스 ────────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class ApprovalCase:
    """CONFIRM으로 시작하는 다단계 계획 하나.

    케이스는 두 종류다.

    - **계획 고정** — `plan`을 주면 플래너를 그 계획으로 갈아끼운다. 모델의
      변덕을 끼워 넣지 않고 승인 게이트만 잰다. `expected_cells`에 마지막
      단계까지 가야 채워지는 셀을 넣는다.
    - **라우터 자체 계획** — `plan`을 비우면 라우터가 스스로 계획을 세운다.
      실제 사용자 경로 그대로다. 다만 슬롯 파싱 결과까지 단언하면 무관한
      버그에 흔들리므로, `watch_cells`만 정하고 **두 경로의 결과가 같은지**만
      본다. 기대값은 대조 경로가 실제로 만든 상태로 잡는다.
    """

    case_id: str
    description: str
    sheet: str
    rows: list[list[Any]]
    plan: list[dict[str, Any]] = field(default_factory=list)
    expected_cells: dict[str, Any] = field(default_factory=dict)
    # 라우터가 직접 계획하는 케이스용.
    message: str = ""
    watch_cells: tuple[str, ...] = field(default_factory=tuple)
    expected_step_count: int = 0
    # 잘려나간 단계가 무엇을 잃게 하는가.
    #   data         — 셀 값이 사라진다. 값만 봐도 드러난다.
    #   formatting   — 값은 맞고 서식만 사라진다. 값만 보는 검사로는 못 잡는다.
    #   verification — 검증 단계라 파일이 아예 그대로다. 손실이 보이지 않는다.
    # 뒤로 갈수록 위험하다. 사용자도 테스트도 아무 이상을 못 느낀다.
    loss_kind: str = "data"
    # 서식 손실을 재는 케이스에서 배경색이 칠해져야 하는 셀.
    expected_fill_cell: str = ""
    # 실행기가 거짓말할 때 되돌아가야 하는 원래 값. 비우면 롤백을 재지 않는다.
    lie_on_value: Any = None
    lie_writes_instead: Any = None
    rollback_cell: str = ""
    tags: tuple[str, ...] = field(default_factory=tuple)

    @property
    def prompt(self) -> str:
        return self.message or self.description

    @property
    def step_count(self) -> int:
        return len(self.plan) or self.expected_step_count


def _sales_rows() -> list[list[Any]]:
    return [
        ["월", "지역", "금액"],
        ["1월", "서울", 111],
        ["2월", "부산", 999],
    ]


ALL_CASES: list[ApprovalCase] = [
    # 계획을 스텁하지 않는다. 라우터의 `_build_create_table_steps`가 스스로
    # `[create_table, write_range]`를 만들고, 승인 게이트가 그것을 자른다.
    # 결함이 실제 사용자 경로에서 재현된다는 증거라 이 케이스가 가장 중요하다.
    ApprovalCase(
        case_id="real_create_table_flow",
        description="표를 만들고 머리글을 채운다 — 라우터가 스스로 세우는 2단계 계획",
        message="2행 2열 표 만들어줘. 머리글은 이름, 점수",
        sheet="Sheet1",
        rows=[],
        watch_cells=("A1", "B1"),
        expected_step_count=2,
        tags=("create_table", "router_planned"),
    ),
    ApprovalCase(
        case_id="formula_then_verify",
        description="수식을 넣고 결과를 검증한다 — `_operation_action_plan`의 수식 계획",
        sheet="매출",
        rows=_sales_rows(),
        plan=[
            {
                "action": "excel_live.set_formula",
                "params": {"range_ref": "D2:D3", "formula_a1": "=C2*2"},
                "reason": "계산 수식 적용",
            },
            {
                "action": "excel_live.verify_formula_result",
                "params": {"range_ref": "D2:D3"},
                "reason": "수식 결과 검증",
            },
        ],
        expected_cells={"D2": "=C2*2"},
        loss_kind="verification",
        tags=("formula",),
    ),
    # 계획을 고정해도 라우터의 `_ACTION_EVIDENCE` 가드는 살아 있다. 문장에
    # 근거가 없는 단계는 실행 전에 떨어져 나가므로, 문장에 "배경색"을 넣어야
    # fill_range가 계획에 남는다.
    ApprovalCase(
        case_id="write_then_highlight",
        description="값을 쓰고 그 셀에 배경색을 칠한다",
        message="C3에 120 쓰고 그 셀 배경색 칠해줘",
        sheet="매출",
        rows=_sales_rows(),
        plan=[
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "C3", "values_2d": [[120]]},
                "reason": "값 입력",
            },
            {
                "action": "excel_live.fill_range",
                "params": {"target_range": "C3", "fill_color": "#FFFF00"},
                "reason": "강조",
            },
        ],
        expected_cells={"C3": 120},
        loss_kind="formatting",
        expected_fill_cell="C3",
        tags=("write", "format"),
    ),
    ApprovalCase(
        case_id="write_three_cells",
        description="세 셀을 순서대로 채운다 — 단계 수만큼 손실이 커지는지 본다",
        sheet="매출",
        rows=_sales_rows(),
        plan=[
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "E1", "values_2d": [["하나"]]},
                "reason": "1",
            },
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "E2", "values_2d": [["둘"]]},
                "reason": "2",
            },
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "E3", "values_2d": [["셋"]]},
                "reason": "3",
            },
        ],
        expected_cells={"E1": "하나", "E2": "둘", "E3": "셋"},
        tags=("write",),
    ),
    # 대조군 — 단일 CONFIRM이면 두 경로가 같아야 한다. 하네스가 무조건
    # 실패를 찍는 게 아니라는 증거다.
    ApprovalCase(
        case_id="single_step_control",
        description="대조군 — 단계가 하나뿐이면 잘릴 것이 없다",
        sheet="매출",
        rows=_sales_rows(),
        plan=[
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "C3", "values_2d": [[120]]},
                "reason": "값 입력",
            }
        ],
        expected_cells={"C3": 120},
        lie_on_value=120,
        lie_writes_instead=777,
        rollback_cell="C3",
        tags=("control", "rollback"),
    ),
]


# ── 실행 ──────────────────────────────────────────────────────────────────


@dataclass
class PathOutcome:
    """한 경로(direct 또는 gated)를 태운 결과."""

    ok: bool
    cells: dict[str, Any]
    matched_cells: int
    total_cells: int
    error: str = ""
    # 서식 케이스에서만 채워진다. None이면 재지 않았다는 뜻.
    fill_applied: bool | None = None
    # 응답이 보고한, 실제로 실행된 단계 수.
    executed_steps: int = 0

    @property
    def file_correct(self) -> bool:
        return self.matched_cells == self.total_cells


@dataclass
class ApprovalOutcome:
    case_id: str
    description: str
    planned_steps: int
    direct: PathOutcome
    gated: PathOutcome
    approval_required: bool
    loss_kind: str = "data"
    approved_action: str = ""
    rollback_direct: Any = None
    rollback_gated: Any = None
    rollback_measured: bool = False

    @property
    def lost_steps(self) -> int:
        """승인 경로가 실행하지 못한 단계 수.

        "승인 경로는 어차피 하나만 실행한다"고 가정하지 않는다. 응답이 보고한
        실행 단계 수를 대조 경로와 비교해서 센다 — 그래야 고친 뒤에 0이 되는지를
        측정으로 확인할 수 있다.
        """
        return max(0, self.direct.executed_steps - self.gated.executed_steps)

    @property
    def completion_rate(self) -> float:
        if not self.planned_steps:
            return 1.0
        return (self.planned_steps - self.lost_steps) / self.planned_steps

    @property
    def diverged(self) -> bool:
        """두 경로의 최종 파일 상태가 다른가. 값과 서식을 함께 본다."""
        return (
            self.direct.cells != self.gated.cells
            or self.direct.fill_applied != self.gated.fill_applied
        )

    @property
    def formatting_lost(self) -> bool:
        """대조 경로는 칠했는데 승인 경로는 못 칠했는가."""
        return bool(self.direct.fill_applied) and not self.gated.fill_applied

    @property
    def rollback_lost(self) -> bool:
        """direct는 되돌렸는데 gated는 안 되돌렸는가."""
        if not self.rollback_measured:
            return False
        return self.rollback_direct != self.rollback_gated


def _build_workbook(root: Path, case: ApprovalCase) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = case.sheet
    for row in case.rows:
        worksheet.append(row)
    path = root / f"{case.case_id}.xlsx"
    workbook.save(path)
    return path


def _read_cells(path: Path, sheet: str, refs: list[str]) -> dict[str, Any]:
    worksheet = load_workbook(path, data_only=False)[sheet]
    return {ref: worksheet[ref].value for ref in refs}


def _is_filled(path: Path, sheet: str, ref: str) -> bool:
    """배경색이 실제로 칠해졌는가. 색상 표기는 엔진마다 달라 존재만 본다."""
    fill = load_workbook(path, data_only=False)[sheet][ref].fill
    return str(getattr(fill, "fill_type", "") or "") not in ("", "none")


def install_fixed_plan(monkeypatch, router, plan: list[dict[str, Any]]) -> None:
    """플래너를 고정 계획으로 갈아끼운다.

    승인 게이트만 재는 것이 목적이므로 모델의 변덕을 끼워 넣지 않는다.
    빠른 규칙도 꺼야 문장에 따라 계획이 바뀌지 않는다.
    """

    async def _fake_parse(message, llm_service=None, context=None, **kwargs):
        return {"action_plan": plan, "intent": "edit", "reason": "승인 게이트 측정"}

    monkeypatch.setattr(router, "parse_excel_live_command", _fake_parse)
    monkeypatch.setattr(router, "_build_quick_action_plan", lambda *a, **kw: None)


def _install_lying_writer(monkeypatch, router, *, planned: Any, actual: Any) -> None:
    """계획한 값을 쓰라고 하면 다른 값을 쓰면서 성공을 보고한다.

    보호된 시트나 병합 셀에서 실제로 일어나는 일이다. 롤백도 write_range를
    쓰므로 계획한 값일 때만 거짓말해야 복구 경로가 살아 있다.
    """
    service = router.get_excel_live_service()
    original = service.write_range

    def _lying(workbook_id, sheet_name, start_cell, values_2d, **kwargs):
        flat = [c for row in values_2d or [] for c in (row or [])]
        if planned in flat:
            values_2d = [[actual]]
        return original(workbook_id, sheet_name, start_cell, values_2d, **kwargs)

    monkeypatch.setattr(service, "write_range", _lying)
    monkeypatch.setattr(router, "get_excel_live_service", lambda: service)


def _isolate(monkeypatch, root: Path) -> None:
    from office_claw_sidecar.services import excel_live_file_service as file_service
    from office_claw_sidecar.services import excel_live_service as live_service

    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
    monkeypatch.setattr(file_service, "WORKSPACE_ROOT", root)
    monkeypatch.setattr(live_service, "_excel_live_service", None)
    monkeypatch.setattr(live_service, "_excel_live_service_engine", None)


def isolated_workbook(monkeypatch, *, sheet: str = "Sheet1", rows: list[list[Any]] | None = None) -> str:
    """격리된 임시 워크스페이스에 워크북 하나를 만들고 그 id(경로)를 준다.

    승인 레코드의 내용처럼 파일 상태만으로는 볼 수 없는 것을 검사할 때 쓴다.
    """
    root = Path(tempfile.mkdtemp(prefix="oc-approval-"))
    _isolate(monkeypatch, root)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    for row in rows or []:
        worksheet.append(row)
    path = root / "contract.xlsx"
    workbook.save(path)
    return str(path)


def _run_one_path(
    client,
    monkeypatch,
    case: ApprovalCase,
    *,
    gated: bool,
    lie: bool,
) -> tuple[PathOutcome, bool, str]:
    """한 경로를 격리 워크스페이스에서 태우고 최종 파일 상태를 읽는다."""
    from office_claw_sidecar.routers import excel_live as router

    with monkeypatch.context() as patch:
        root = Path(tempfile.mkdtemp(prefix="oc-approval-"))
        _isolate(patch, root)
        path = _build_workbook(root, case)
        if case.plan:
            install_fixed_plan(patch, router, case.plan)
        if lie:
            _install_lying_writer(
                patch, router, planned=case.lie_on_value, actual=case.lie_writes_instead
            )

        refs = sorted(
            set(case.expected_cells)
            | set(case.watch_cells)
            | ({case.rollback_cell} if case.rollback_cell else set())
        )
        approval_required = False
        approved_action = ""
        try:
            body = client.post(
                "/excel-live/command",
                json={
                    "message": case.prompt,
                    "workbook_id": str(path),
                    "sheet_name": case.sheet,
                    "approve": not gated,
                },
                headers=HEADERS,
            ).json()

            approval_required = bool(body.get("approval_required"))
            if gated and approval_required:
                approval_id = (body.get("pending_approval") or {}).get("approval_id")
                approved = client.post(
                    "/excel-live/approval",
                    json={"approval_id": approval_id, "approved": True},
                    headers=HEADERS,
                ).json()
                approved_action = str(approved.get("action") or "")
                body = approved
        except Exception as exc:
            return (
                PathOutcome(False, {}, 0, len(case.expected_cells), error=str(exc)),
                approval_required,
                approved_action,
            )

        cells = _read_cells(path, case.sheet, refs)
        matched = sum(
            1 for ref, expected in case.expected_cells.items() if cells.get(ref) == expected
        )
        result = body.get("result") or {}
        # 단계가 하나뿐이면 응답에 executed_steps가 없다. 성공했으면 1로 센다.
        executed = int(result.get("executed_steps") or 0) or (1 if body.get("ok") else 0)
        return (
            PathOutcome(
                ok=bool(body.get("ok")),
                cells=cells,
                matched_cells=matched,
                total_cells=len(case.expected_cells),
                executed_steps=executed,
                fill_applied=(
                    _is_filled(path, case.sheet, case.expected_fill_cell)
                    if case.expected_fill_cell
                    else None
                ),
            ),
            approval_required,
            approved_action,
        )


def _score_against(outcome: PathOutcome, expected: dict[str, Any]) -> None:
    """기대값이 케이스에 없는 경우, 대조 경로가 만든 상태를 기준으로 채점한다."""
    outcome.total_cells = len(expected)
    outcome.matched_cells = sum(
        1 for ref, value in expected.items() if outcome.cells.get(ref) == value
    )


def run_case(client, monkeypatch, case: ApprovalCase) -> ApprovalOutcome:
    """한 케이스를 direct/gated 두 경로로 태워 나란히 놓는다."""
    direct, _, _ = _run_one_path(client, monkeypatch, case, gated=False, lie=False)
    gated, approval_required, approved_action = _run_one_path(
        client, monkeypatch, case, gated=True, lie=False
    )

    if not case.expected_cells:
        # 라우터가 스스로 계획한 케이스. 대조 경로의 결과를 정답으로 삼는다.
        reference = {ref: direct.cells.get(ref) for ref in case.watch_cells}
        _score_against(direct, reference)
        _score_against(gated, reference)

    outcome = ApprovalOutcome(
        case_id=case.case_id,
        description=case.description,
        planned_steps=case.step_count,
        direct=direct,
        gated=gated,
        approval_required=approval_required,
        loss_kind=case.loss_kind,
        approved_action=approved_action,
    )

    if case.rollback_cell:
        lie_direct, _, _ = _run_one_path(client, monkeypatch, case, gated=False, lie=True)
        lie_gated, _, _ = _run_one_path(client, monkeypatch, case, gated=True, lie=True)
        outcome.rollback_measured = True
        outcome.rollback_direct = lie_direct.cells.get(case.rollback_cell)
        outcome.rollback_gated = lie_gated.cells.get(case.rollback_cell)

    return outcome


def run_all(client, monkeypatch) -> list[ApprovalOutcome]:
    return [run_case(client, monkeypatch, case) for case in ALL_CASES]


# ── 집계 ──────────────────────────────────────────────────────────────────


def summarize(outcomes: list[ApprovalOutcome]) -> dict[str, Any]:
    multi = [o for o in outcomes if o.planned_steps > 1]
    planned = sum(o.planned_steps for o in outcomes)
    lost = sum(o.lost_steps for o in outcomes)
    rollback_cases = [o for o in outcomes if o.rollback_measured]
    silent = [o for o in multi if o.lost_steps and o.loss_kind == "verification"]
    return {
        "cases": len(outcomes),
        "multi_step_cases": len(multi),
        "planned_steps": planned,
        "lost_steps": lost,
        "completion_rate": round((planned - lost) / planned, 4) if planned else 1.0,
        "direct_file_correct": sum(1 for o in outcomes if o.direct.file_correct),
        "gated_file_correct": sum(1 for o in outcomes if o.gated.file_correct),
        "diverged": sum(1 for o in outcomes if o.diverged),
        "formatting_lost": sum(1 for o in outcomes if o.formatting_lost),
        # 파일은 멀쩡한데 검증 단계만 사라진 경우. 아무도 눈치채지 못한다.
        "silent_verification_loss": len(silent),
        "rollback_measured": len(rollback_cases),
        "rollback_lost": sum(1 for o in rollback_cases if o.rollback_lost),
    }


def to_report(outcomes: list[ApprovalOutcome]) -> dict[str, Any]:
    return {
        "summary": summarize(outcomes),
        "cases": [
            {
                "case_id": o.case_id,
                "description": o.description,
                "planned_steps": o.planned_steps,
                "lost_steps": o.lost_steps,
                "loss_kind": o.loss_kind,
                "completion_rate": round(o.completion_rate, 4),
                "approval_required": o.approval_required,
                "approved_action": o.approved_action,
                "direct": {
                    "ok": o.direct.ok,
                    "file_correct": o.direct.file_correct,
                    "cells": {k: _jsonable(v) for k, v in o.direct.cells.items()},
                    "fill_applied": o.direct.fill_applied,
                    "error": o.direct.error,
                },
                "gated": {
                    "ok": o.gated.ok,
                    "file_correct": o.gated.file_correct,
                    "cells": {k: _jsonable(v) for k, v in o.gated.cells.items()},
                    "fill_applied": o.gated.fill_applied,
                    "error": o.gated.error,
                },
                "diverged": o.diverged,
                "formatting_lost": o.formatting_lost,
                "rollback": (
                    {
                        "measured": True,
                        "direct": _jsonable(o.rollback_direct),
                        "gated": _jsonable(o.rollback_gated),
                        "lost": o.rollback_lost,
                    }
                    if o.rollback_measured
                    else {"measured": False}
                ),
            }
            for o in outcomes
        ],
    }


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
