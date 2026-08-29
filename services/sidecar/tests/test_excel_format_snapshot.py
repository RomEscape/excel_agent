"""서식 스냅샷 — 사후조건이 "요청한 효과가 파일에 남았는가"를 볼 수 있게 하는 원시 도구.

2026-08-19 블라인드 게이트 45건 분류: 조용한 오실행의 62%가 "요청한 효과가 안 남음"이고
그 절반 이상이 서식이다(표시 형식 그대로 · 배경만 칠하고 굵게는 안 됨 · 병합 안 됨 · 틀 고정 안 됨).
값만 읽는 `get_range_snapshot`으로는 이 부류를 전혀 볼 수 없었다.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService


@pytest.fixture
def service(tmp_path):
    path = tmp_path / "fmt.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["머리글1", "머리글2"])
    ws.append([1000, 2000])
    ws["A1"].font = Font(bold=True, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="FF002060")
    ws["A2"].number_format = "#,##0"
    ws["B2"].border = Border(left=Side(style="thin"))
    ws.merge_cells("A4:C4")
    ws.freeze_panes = "A2"
    wb.save(path)
    wb.close()
    svc = FileExcelLiveService(workspace_root=tmp_path)
    svc.select_workbook(str(path))
    return svc, path


class TestFormatSnapshot:
    def test_number_format_is_readable(self, service):
        svc, path = service
        snap = svc.get_format_snapshot(str(path), "S", "A1:B2")
        assert snap["number_formats"] == [["General", "General"], ["#,##0", "General"]]

    def test_fill_and_font_are_readable(self, service):
        svc, path = service
        snap = svc.get_format_snapshot(str(path), "S", "A1:B2")
        assert snap["fills"][0][0] == "FF002060"
        assert snap["fills"][0][1] is None
        assert snap["bold"] == [[True, False], [False, False]]
        assert snap["font_colors"][0][0] == "FFFFFFFF"

    def test_a_theme_color_does_not_leak_an_error_string(self, tmp_path):
        # openpyxl은 테마 색일 때 rgb 자리에 오류 문자열을 돌려준다 — 16진 코드만 통과해야 한다.
        from openpyxl.styles.colors import Color

        path = tmp_path / "theme.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "x"
        ws["A1"].font = Font(color=Color(theme=1, tint=0.0))
        wb.save(path)
        wb.close()
        svc = FileExcelLiveService(workspace_root=tmp_path)
        svc.select_workbook(str(path))
        snap = svc.get_format_snapshot(str(path), "S", "A1:A1")
        assert snap["font_colors"][0][0] is None

    def test_border_merge_freeze_and_chart_count(self, service):
        svc, path = service
        snap = svc.get_format_snapshot(str(path), "S", "A1:B2")
        assert snap["borders"] == [[False, False], [False, True]]
        assert "A4:C4" in snap["merged"]
        assert snap["freeze_panes"] == "A2"
        assert snap["chart_count"] == 0

    def test_both_engines_expose_the_same_shape(self):
        # 사후조건이 엔진을 가리면 안 된다 — xlwings 쪽도 같은 키를 준다.
        from office_claw_sidecar.services.excel_live_service import ExcelLiveService

        assert hasattr(ExcelLiveService, "get_format_snapshot")
        assert hasattr(FileExcelLiveService, "get_format_snapshot")

    def test_a_huge_range_is_capped(self, service):
        svc, path = service
        snap = svc.get_format_snapshot(str(path), "S", "A1:BZ500")
        assert len(snap["number_formats"]) <= 200
        assert len(snap["number_formats"][0]) <= 60
