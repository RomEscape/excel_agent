"""연습용 워크북 자동 배치 — 새 클론의 `엑셀 작업 폴더` 가 비어 있던 결함(2026-09-06 실클론 감사).

README 는 "연습용 AI_Excel_Automation_Demo.xlsx 가 들어 있다"고 약속하지만 그 폴더는
gitignore 다. 소스 트리에서 처음 켤 때 `config._seed_demo_workbook` 이 한 부 넣는다.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from office_claw_sidecar import config


def _make_repo(tmp_path: Path, *, with_source: bool = True) -> tuple[Path, Path]:
    repo = tmp_path / "repo"
    workspace = repo / config.DEV_WORKSPACE_DIRNAME
    workspace.mkdir(parents=True)
    if with_source:
        src = repo / config.DEMO_WORKBOOK_SOURCE
        src.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.active["A1"] = "데모"
        wb.save(src)
    return repo, workspace


def test_empty_workspace_gets_demo_copy(tmp_path):
    repo, workspace = _make_repo(tmp_path)
    out = config._seed_demo_workbook(repo, workspace)
    assert out == workspace / "AI_Excel_Automation_Demo.xlsx"
    assert out.is_file()
    assert openpyxl.load_workbook(out).active["A1"].value == "데모"
    # 원본은 그대로 — 복사지 이동이 아니다.
    assert (repo / config.DEMO_WORKBOOK_SOURCE).is_file()


def test_existing_user_file_is_left_alone(tmp_path):
    """사용자 파일이 하나라도 있으면 아무것도 넣지 않는다 — 사용자 폴더를 어지럽히지 않는다."""
    repo, workspace = _make_repo(tmp_path)
    openpyxl.Workbook().save(workspace / "내파일.xlsx")
    assert config._seed_demo_workbook(repo, workspace) is None
    assert sorted(p.name for p in workspace.glob("*.xlsx")) == ["내파일.xlsx"]


def test_missing_source_is_silent(tmp_path):
    repo, workspace = _make_repo(tmp_path, with_source=False)
    assert config._seed_demo_workbook(repo, workspace) is None
    assert list(workspace.glob("*.xlsx")) == []


def test_get_workspace_root_seeds_in_source_tree(tmp_path, monkeypatch):
    """실제 진입점 — 소스 트리 감지가 되면 첫 호출에서 채워진다."""
    repo, workspace = _make_repo(tmp_path)
    monkeypatch.delenv("OFFICE_CLAW_WORKSPACE_DIR", raising=False)
    monkeypatch.setattr(config, "_detect_workspace_root", lambda: repo)
    root = config.get_workspace_root()
    assert root == workspace
    assert (workspace / "AI_Excel_Automation_Demo.xlsx").is_file()


def test_env_override_never_seeds(tmp_path, monkeypatch):
    """OFFICE_CLAW_WORKSPACE_DIR 로 지정한 폴더는 사용자 것 — 손대지 않는다."""
    repo, _workspace = _make_repo(tmp_path)
    custom = tmp_path / "custom"
    monkeypatch.setenv("OFFICE_CLAW_WORKSPACE_DIR", str(custom))
    monkeypatch.setattr(config, "_detect_workspace_root", lambda: repo)
    root = config.get_workspace_root()
    assert root == custom
    assert list(custom.glob("*.xlsx")) == []
