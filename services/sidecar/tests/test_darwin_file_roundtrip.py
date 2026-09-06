"""macOS 파일 왕복 폴백 — 2026-09-06 사용자 "맥에서 Excel 켜진 상태에서도 동작돼야".

Mac용 Excel 자동화 사전에는 데이터 막대·색조·수식 조건부 서식·유효성 검사가 없다. 예전엔
`ExcelLiveError`로 거절했다. 지금은 저장→닫기→파일 엔진(openpyxl) 적용→다시 열기→시트·선택
복원 순서로 해낸다. 이 PC(Windows)에는 Mac이 없으므로 **순서와 결과 파일**만 고정한다 —
xlwings 객체는 가짜로 대신하고, 파일 엔진은 진짜 openpyxl 로 tmp 워크북에 적용한다.
"""

from __future__ import annotations

import sys

import pytest
from openpyxl import Workbook, load_workbook

from office_claw_sidecar.services import excel_live_service as live


class _FakeRange:
    def __init__(self, address: str) -> None:
        self.address = address
        self.selected: list[str] = []

    def select(self) -> None:
        self.selected.append(self.address)


class _FakeSheet:
    def __init__(self, name: str, log: list[str]) -> None:
        self.name = name
        self._log = log

    def activate(self) -> None:
        self._log.append(f"activate:{self.name}")

    def range(self, ref: str) -> _FakeRange:
        self._log.append(f"select:{ref}")
        return _FakeRange(ref)


class _FakeWorkbook:
    def __init__(self, path: str, log: list[str]) -> None:
        self.fullname = path
        self.name = path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        self._log = log
        self._sheets = {"Sheet1": _FakeSheet("Sheet1", log)}

    def save(self) -> None:
        self._log.append("save")

    def close(self) -> None:
        self._log.append("close")


@pytest.fixture
def darwin_service(tmp_path, monkeypatch):
    """darwin 로 위장한 xlwings 서비스 + 가짜 통합문서. 파일은 진짜 xlsx."""
    path = tmp_path / "mac.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["점수"])
    for v in (10, 40, 70, 100):
        ws.append([v])
    wb.save(path)

    log: list[str] = []
    fake_wb = _FakeWorkbook(str(path), log)

    svc = live.ExcelLiveService.__new__(live.ExcelLiveService)
    svc._selected_workbook_id = str(path)
    svc._find_workbook = lambda ident: fake_wb  # type: ignore[method-assign]
    svc._find_sheet = lambda wb_, name: fake_wb._sheets["Sheet1"]  # type: ignore[method-assign]
    svc._resolve_target_range = staticmethod(lambda sheet, ref: _FakeRange(ref))  # type: ignore[method-assign]
    svc.get_active_selection_ref = lambda wb_id, sheet: "A2:A5"  # type: ignore[method-assign]
    svc.open_workbook_in_excel = lambda p: (log.append(f"open:{p}") or True)  # type: ignore[method-assign]
    monkeypatch.setattr(sys, "platform", "darwin")
    return svc, path, log


class TestDarwinFileRoundtrip:
    def test_data_bar_is_written_by_the_file_engine_between_close_and_reopen(self, darwin_service) -> None:
        svc, path, log = darwin_service
        out = svc.apply_data_bar(None, "Sheet1", "A2:A5", color="#638EC6")

        assert out["applied_via"] == "file_roundtrip"
        assert "다시 열어" in out["note"]
        # 순서: 저장 → 닫기 → (파일 적용) → 다시 열기 → 시트 활성 → 선택 복원
        assert log[:2] == ["save", "close"]
        assert log[2] == f"open:{path}"
        assert "activate:Sheet1" in log and "select:A2:A5" in log
        # 파일에 실제로 규칙이 들어갔다
        ws = load_workbook(path)["Sheet1"]
        rules = [r for rng in ws.conditional_formatting for r in rng.rules]
        assert any(r.type == "dataBar" for r in rules), [r.type for r in rules]

    def test_formula_cf_and_validation_and_color_scale_take_the_same_path(self, darwin_service) -> None:
        svc, path, log = darwin_service
        svc.apply_formula_cf(None, "Sheet1", "A2:A5", "A2>50")
        svc.apply_color_scale(None, "Sheet1", "A2:A5")
        svc.set_data_validation(None, "Sheet1", target_range="A2:A5", validation_type="whole", minimum=0, maximum=100)

        assert log.count("save") == 3 and log.count("close") == 3
        ws = load_workbook(path)["Sheet1"]
        rule_types = sorted(r.type for rng in ws.conditional_formatting for r in rng.rules)
        assert "expression" in rule_types and "colorScale" in rule_types, rule_types
        assert ws.data_validations.dataValidation, "유효성 검사가 파일에 없다"

    def test_unsaved_workbook_is_refused_before_closing(self, darwin_service) -> None:
        svc, _path, log = darwin_service
        svc._find_workbook = lambda ident: type("Wb", (), {"fullname": "", "name": "Book1"})()  # type: ignore[method-assign]
        with pytest.raises(live.ExcelLiveError):
            svc.apply_data_bar(None, "Sheet1", "A2:A5")
        assert "close" not in log

    def test_windows_path_is_untouched(self, darwin_service, monkeypatch) -> None:
        """darwin 이 아니면 예전 COM 경로 그대로 — 가짜 range 엔 api 가 없어 AttributeError 로 드러난다."""
        svc, _path, log = darwin_service
        monkeypatch.setattr(sys, "platform", "win32")
        with pytest.raises(AttributeError):
            svc.apply_data_bar(None, "Sheet1", "A2:A5")
        assert "close" not in log
