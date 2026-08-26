"""
Excel Live Service — 실행 중인 Excel 실시간 제어용 서비스(xlwings).

MVP Day 1 범위:
  - Excel 연결 가능 여부 확인
  - 열린 통합문서 목록 조회
  - 통합문서 선택/조회 상태 관리
  - 기본 범위 읽기(read_range)

주의:
  - Windows/macOS + Excel Desktop 환경을 전제로 한다.
  - xlwings 의존성은 lazy import로 처리하여 비 Excel 테스트 환경에서도
    모듈 import 자체는 실패하지 않도록 한다.
"""

from __future__ import annotations

import importlib
import math
import os
import re
import shutil
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from office_claw_sidecar.services.excel_header_lexicon import resolve_header

#: 사람이 쓰는 맞춤 낱말 → 표준 값. 표에 없는 낱말은 None이 돼 맞춤을 건드리지 않는다.
#: 한국어 "정렬"은 줄 세우기도 뜻하므로, **방향 낱말이 있을 때만** 맞춤이다.
_ALIGN_WORDS = {
    "left": "left",
    "center": "center",
    "centre": "center",
    "right": "right",
    "justify": "justify",
    "왼쪽": "left",
    "좌": "left",
    "가운데": "center",
    "중앙": "center",
    "오른쪽": "right",
    "우": "right",
    "양쪽": "justify",
}



#: 사용자 작업물로 볼 수 없는 디렉터리 — 벤더 데모(xlwings quickstart 등)·백업이 산다.
#: 원래 파일 엔진에만 있었는데, **xlwings 경로에는 이 가드가 없어** 벤더 데모 파일이
#: 열려 있다는 이유로 auto가 xlwings를 골랐고 `books.active`가 그 파일에 실행했다
#: (2026-08-04 실측, 감사 B1). 목록은 여기 한 곳 — 파일 엔진이 이걸 임포트한다.
_SCAN_EXCLUDED_DIRS = frozenset(
    {
        ".git",
        ".venv",
        "venv",
        "env",
        "node_modules",
        "site-packages",
        "__pycache__",
        "officeclaw_backups",
        "dist",
        "build",
    }
)


def _is_user_workbook_path(fullname: str) -> bool:
    """열린 통합문서 경로가 **사람의 작업물**로 보이는가.

    벤더·가상환경·백업 경로면 False — 엔진 선택의 근거로도, `books.active` 폴백의
    대상으로도 삼지 않는다. 판정 불가(경로 없음·비정상)면 True 쪽으로 둔다:
    새 통합문서(저장 전, fullname이 "통합 문서1")를 거절하면 정상 사용이 막힌다.
    """
    text = str(fullname or "").strip()
    if not text or ("\\" not in text and "/" not in text):
        return True  # 저장 전 새 문서 — 경로가 아직 없다
    from pathlib import PurePath

    parts = PurePath(text).parts[:-1]
    return not any(part in _SCAN_EXCLUDED_DIRS or str(part).startswith(".") for part in parts if part)


class ExcelLiveError(Exception):
    """Excel Live 서비스 기본 예외."""


class ExcelDependencyError(ExcelLiveError):
    """xlwings 또는 OS별 자동화 의존성 누락/import 실패."""


class ExcelConnectionError(ExcelLiveError):
    """실행 중인 Excel 인스턴스 연결 실패."""


class WorkbookNotFoundError(ExcelLiveError):
    """요청한 통합문서를 찾지 못함."""


class WorksheetNotFoundError(ExcelLiveError):
    """요청한 시트를 찾지 못함."""


class AmbiguousWorkbookError(ExcelLiveError):
    """대상 통합문서를 특정하지 못함 — 후보가 여럿이라 되물어야 한다.

    대상을 지정하지 않았을 때 아무 파일이나 골라 편집하면, 사용자가 보고 있지도
    않은 통합문서가 조용히 바뀐다. 후보를 들고 되묻기 위해 별도 예외로 둔다.
    """

    def __init__(self, message: str, candidates: list[str] | None = None) -> None:
        super().__init__(message)
        self.candidates = candidates or []


@dataclass(frozen=True)
class WorkbookInfo:
    workbook_id: str
    name: str
    full_path: str
    active_sheet: str

    def as_dict(self) -> dict[str, str]:
        return {
            "workbook_id": self.workbook_id,
            "name": self.name,
            "full_path": self.full_path,
            "active_sheet": self.active_sheet,
        }


# VBA 식별자는 문자로 시작하고 문자·숫자·밑줄만 쓴다. `Module1.Macro`처럼 모듈로
# 한정하는 것까지 허용한다.
_MACRO_NAME_PATTERN = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*(\.[A-Za-z_][A-Za-z0-9_]*)?$")


def _validate_macro_name(macro_name: str | None) -> str:
    """실행 문자열에 끼어들 수 있는 이름을 막는다.

    매크로 이름은 `'파일명'!매크로` 문자열로 조립돼 `Application.Run`에 들어간다.
    이름에 `!`나 따옴표나 경로 구분자가 들어오면 우리가 못 박아 둔 대상 한정을
    빠져나가 다른 파일의 매크로를 부를 수 있다.
    """
    macro = str(macro_name or "").strip()
    if not macro:
        raise ExcelLiveError("macro_name이 필요합니다.")
    if not _MACRO_NAME_PATTERN.match(macro):
        raise ExcelLiveError(
            f"매크로 이름 '{macro}'을 쓸 수 없습니다. "
            "영문·숫자·밑줄과 모듈 한정(`모듈.매크로`)만 허용합니다."
        )
    return macro


# 서식 스냅샷 상한 — 검증용이라 표 하나를 덮을 정도면 충분하다.
_FORMAT_SNAPSHOT_MAX_ROWS = 200
_FORMAT_SNAPSHOT_MAX_COLS = 60

class ExcelLiveService:
    """실행 중인 Excel 제어 서비스(xlwings 기반)."""

    engine = "xlwings"

    def __init__(self, xw_module: Any | None = None) -> None:
        self._xw = xw_module
        self._selected_workbook_id: str | None = None

    def _xw_module(self) -> Any:
        if self._xw is not None:
            return self._xw
        try:
            import xlwings as xw  # type: ignore[import]
        except Exception as exc:  # pragma: no cover - 환경 의존
            raise ExcelDependencyError(
                "xlwings 모듈을 불러올 수 없습니다. "
                "Windows는 pywin32, macOS는 appscript 의존성을 함께 설치해 주세요."
            ) from exc
        self._xw = xw
        return xw

    def _app(self):
        xw = self._xw_module()
        try:
            app = xw.apps.active
        except Exception as exc:  # pragma: no cover - 환경 의존
            raise ExcelConnectionError(
                "실행 중인 Excel 인스턴스에 연결할 수 없습니다. Excel을 먼저 실행해 주세요."
            ) from exc
        if app is None:
            raise ExcelConnectionError(
                "실행 중인 Excel 인스턴스를 찾지 못했습니다. Excel을 먼저 실행해 주세요."
            )
        return app

    def is_available(self) -> bool:
        """Excel 연결 가능 여부를 bool로 반환한다 (예외 미전파)."""
        try:
            self._app()
            return True
        except ExcelLiveError:
            return False

    def list_workbooks(self) -> list[dict[str, str]]:
        """현재 열린 통합문서 목록을 반환한다."""
        app = self._app()
        rows: list[dict[str, str]] = []
        for wb in app.books:
            info = WorkbookInfo(
                workbook_id=self._workbook_id(wb),
                name=str(getattr(wb, "name", "")),
                full_path=str(getattr(wb, "fullname", "") or ""),
                active_sheet=self._active_sheet_name(wb),
            )
            rows.append(info.as_dict())
        return rows

    def select_workbook(self, workbook_id_or_name: str) -> dict[str, Any]:
        """작업 대상 통합문서를 선택하고 선택 결과를 반환한다."""
        wb = self._find_workbook(workbook_id_or_name)
        workbook_id = self._workbook_id(wb)
        self._selected_workbook_id = workbook_id
        return {"selected": True, "workbook_id": workbook_id}

    def get_selected_workbook_id(self) -> str | None:
        return self._selected_workbook_id

    def list_sheets(self, workbook_id: str | None) -> dict[str, Any]:
        """대상 통합문서의 시트 목록과 활성 시트를 반환한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            rows = self.list_workbooks()
            if not rows:
                raise WorkbookNotFoundError("열린 통합문서가 없습니다.")
            target_id = rows[0]["workbook_id"]

        wb = self._find_workbook(target_id)
        sheet_names = [str(getattr(sheet, "name", "") or "") for sheet in wb.sheets]
        return {
            "sheets": sheet_names,
            "count": len(sheet_names),
            "active_sheet": self._active_sheet_name(wb),
        }

    def select_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
    ) -> dict[str, Any]:
        """대상 통합문서에서 작업 시트를 전환한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            rows = self.list_workbooks()
            if not rows:
                raise WorkbookNotFoundError("열린 통합문서가 없습니다.")
            target_id = rows[0]["workbook_id"]

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        try:
            sheet.activate()
        except Exception:
            pass
        return {
            "selected": True,
            "sheet_name": str(getattr(sheet, "name", "") or sheet_name),
            "active_sheet": self._active_sheet_name(wb),
        }

    def create_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
        make_active: bool = True,
    ) -> dict[str, Any]:
        """시트를 생성하고(이미 있으면 재사용) 필요 시 활성화한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            rows = self.list_workbooks()
            if not rows:
                raise WorkbookNotFoundError("열린 통합문서가 없습니다.")
            target_id = rows[0]["workbook_id"]

        wb = self._find_workbook(target_id)
        target_name = self._sanitize_sheet_name(sheet_name)
        created = False
        try:
            sheet = self._find_sheet(wb, target_name)
        except WorksheetNotFoundError:
            sheet = wb.sheets.add(name=target_name)
            created = True

        if make_active:
            try:
                sheet.activate()
            except Exception:
                pass

        return {
            "created": created,
            "sheet_name": str(getattr(sheet, "name", "") or target_name),
            "active_sheet": self._active_sheet_name(wb),
        }

    def rename_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
        new_name: str,
    ) -> dict[str, Any]:
        """기존 시트 이름을 바꾼다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            rows = self.list_workbooks()
            if not rows:
                raise WorkbookNotFoundError("열린 통합문서가 없습니다.")
            target_id = rows[0]["workbook_id"]

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        target_name = self._sanitize_sheet_name(new_name)
        old_name = str(getattr(sheet, "name", "") or sheet_name)
        existing = {str(getattr(item, "name", "") or "") for item in wb.sheets}
        if target_name != old_name and target_name in existing:
            raise ExcelLiveError(f"이미 '{target_name}' 시트가 있습니다.")
        sheet.name = target_name
        return {
            "renamed": True,
            "old_name": old_name,
            "sheet_name": str(getattr(sheet, "name", "") or target_name),
            "sheets": [str(getattr(item, "name", "") or "") for item in wb.sheets],
            "active_sheet": self._active_sheet_name(wb),
        }

    def delete_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
    ) -> dict[str, Any]:
        """시트를 삭제한다. 마지막 시트는 남긴다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            rows = self.list_workbooks()
            if not rows:
                raise WorkbookNotFoundError("열린 통합문서가 없습니다.")
            target_id = rows[0]["workbook_id"]

        wb = self._find_workbook(target_id)
        names = [str(getattr(item, "name", "") or "") for item in wb.sheets]
        if len(names) <= 1:
            raise ExcelLiveError("마지막 시트는 삭제할 수 없습니다.")
        sheet = self._find_sheet(wb, sheet_name)
        deleted_name = str(getattr(sheet, "name", "") or sheet_name)
        sheet.delete()
        remaining = [str(getattr(item, "name", "") or "") for item in wb.sheets]
        return {
            "deleted": True,
            "sheet_name": deleted_name,
            "sheets": remaining,
            "active_sheet": self._active_sheet_name(wb),
        }

    def read_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        range_ref: str,
    ) -> dict[str, Any]:
        """
        지정 범위를 읽어 2차원 배열 형태로 반환한다.

        반환:
          {
            "values": [[...], ...],
            "address": "A1:C3",
            "row_count": 3,
            "col_count": 3
          }
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError(
                "선택된 통합문서가 없습니다. workbook_id를 지정하거나 select_workbook을 먼저 호출해 주세요."
            )

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = sheet.range(range_ref)
        raw = rng.options(ndim=2).value
        values = self._normalize_values(raw)
        if values == []:
            rows_obj = getattr(rng, "rows", None)
            cols_obj = getattr(rng, "columns", None)
            row_count = int(getattr(rows_obj, "count", 1) or 1)
            col_count = int(getattr(cols_obj, "count", 1) or 1)
            values = [[None for _ in range(col_count)] for _ in range(row_count)]

        row_count = len(values)
        col_count = len(values[0]) if values else 0
        return {
            "values": values,
            "address": str(range_ref),
            "row_count": row_count,
            "col_count": col_count,
        }

    def get_format_snapshot(
        self, workbook_id: str | None, sheet_name: str | None, range_ref: str
    ) -> dict[str, Any]:
        """범위의 서식 스냅샷(검증용). 파일 엔진과 같은 모양을 돌려준다 — 사후조건이 엔진을 가리지 않게."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name) if sheet_name else wb.sheets.active
        rng = sheet.range(range_ref)
        rows = min(int(rng.rows.count), _FORMAT_SNAPSHOT_MAX_ROWS)
        cols = min(int(rng.columns.count), _FORMAT_SNAPSHOT_MAX_COLS)
        number_formats: list[list[str]] = []
        fills: list[list[str | None]] = []
        bold: list[list[bool]] = []
        font_colors: list[list[str | None]] = []
        borders: list[list[bool]] = []

        def _hex(color: Any) -> str | None:
            if color is None:
                return None
            if isinstance(color, (tuple, list)) and len(color) >= 3:
                return "FF" + "".join(f"{int(v):02X}" for v in color[:3])
            return str(color)

        for r in range(1, rows + 1):
            nf_row, fill_row, bold_row, color_row, border_row = [], [], [], [], []
            for c in range(1, cols + 1):
                cell = rng[r - 1, c - 1]
                try:
                    nf_row.append(str(cell.number_format or "General"))
                except Exception:
                    nf_row.append("General")
                try:
                    fill_row.append(_hex(cell.color))
                except Exception:
                    fill_row.append(None)
                try:
                    bold_row.append(bool(cell.api.Font.Bold))
                except Exception:
                    bold_row.append(False)
                try:
                    color_row.append(_hex(cell.font.color))
                except Exception:
                    color_row.append(None)
                try:
                    # xlEdgeLeft=7 … xlEdgeRight=10, LineStyle -4142 == none
                    border_row.append(
                        any(cell.api.Borders(idx).LineStyle != -4142 for idx in (7, 8, 9, 10))
                    )
                except Exception:
                    border_row.append(False)
            number_formats.append(nf_row)
            fills.append(fill_row)
            bold.append(bold_row)
            font_colors.append(color_row)
            borders.append(border_row)
        try:
            merged = [str(a.address).replace("$", "") for a in rng.api.MergeArea] if rng.api.MergeCells else []
        except Exception:
            merged = []
        try:
            freeze = str(sheet.api.Application.ActiveWindow.SplitRow or 0)
            freeze = f"A{int(freeze) + 1}" if freeze and int(freeze) > 0 else ""
        except Exception:
            freeze = ""
        try:
            chart_count = len(sheet.charts)
        except Exception:
            chart_count = 0
        return {
            "address": str(rng.address).replace("$", ""),
            "sheet_name": str(getattr(sheet, "name", "") or ""),
            "number_formats": number_formats,
            "fills": fills,
            "bold": bold,
            "font_colors": font_colors,
            "borders": borders,
            "merged": merged,
            "freeze_panes": freeze,
            "chart_count": chart_count,
        }

    def get_range_snapshot(
        self,
        workbook_id: str | None,
        sheet_name: str | None,
        range_ref: str,
    ) -> dict[str, Any]:
        """
        범위 상태 스냅샷(검증용)을 반환한다.
        - row/col 크기
        - 값이 채워진 셀 개수
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        if sheet_name:
            resolved_sheet = self._find_sheet(wb, sheet_name)
            resolved_sheet_name = str(getattr(resolved_sheet, "name", "") or "")
        else:
            resolved_sheet = wb.sheets.active
            resolved_sheet_name = str(getattr(resolved_sheet, "name", "") or "")
        data = self.read_range(target_id, resolved_sheet_name, range_ref)
        values = data.get("values", [])
        filled = 0
        for row in values:
            for v in row:
                if v is None:
                    continue
                if isinstance(v, str) and not v.strip():
                    continue
                filled += 1
        return {
            "address": data.get("address", range_ref),
            "row_count": int(data.get("row_count", 0) or 0),
            "col_count": int(data.get("col_count", 0) or 0),
            "filled_cells": filled,
        }

    def write_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        start_cell: str,
        values_2d: list[list[Any]],
    ) -> dict[str, Any]:
        """지정 시작 셀 기준으로 2차원 값을 기록한다."""
        if not values_2d:
            return {"written_cells": 0, "address": start_cell}

        rows = len(values_2d)
        cols = max(len(r) for r in values_2d)
        normalized = [row + [None] * (cols - len(row)) for row in values_2d]

        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = sheet.range(start_cell).resize(rows, cols)
        rng.value = normalized
        return {"written_cells": rows * cols, "address": str(rng.address)}

    def highlight_by_condition(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        operator: str,
        threshold: float,
        fill_color: str = "#FFFF00",
        compare_column: str | None = None,
        value: Any = None,
    ) -> dict[str, Any]:
        """조건에 맞는 셀 배경색을 변경한다.

        compare_column을 주면 고정 기준값 대신 같은 행의 그 열 값과 비교한다
        ("현재고가 재주문점 이하").
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        values = self._normalize_values(rng.options(ndim=2).value)
        rgb = self._hex_to_rgb(fill_color)
        compare_letter = str(compare_column or "").strip().upper() or None

        start_row = int(getattr(rng, "row", 1))
        start_col = int(getattr(rng, "column", 1))

        matched = 0
        changed = 0
        # 조건을 실제로 몇 칸에 대 봤는가. 0건이 "조건에 맞는 게 없어서"인지
        # "범위가 비어서"인지를 이 값으로만 가를 수 있다.
        scanned = 0
        for r_idx, row in enumerate(values):
            for c_idx, cell_value in enumerate(row):
                scanned += 1
                absolute_row = start_row + r_idx
                limit = threshold
                if compare_letter:
                    other = sheet.range(f"{compare_letter}{absolute_row}").value
                    if not isinstance(other, (int, float)) or isinstance(other, bool):
                        continue
                    limit = float(other)
                if self._cell_matches_highlight(cell_value, operator, limit, value):
                    matched += 1
                    absolute_col = start_col + c_idx
                    cell_ref = f"{self._idx_to_col(absolute_col)}{absolute_row}"
                    cell = sheet.range(cell_ref)
                    cell.color = rgb
                    # 흰색/옅은색 채우기는 Excel 기본 격자선이 시각적으로 사라져 보일 수 있다.
                    # 변경 셀에 얇은 경계선을 유지해 "하얀 블록"처럼 보이지 않게 보정한다.
                    self._ensure_visual_gridline(cell, rgb)
                    changed += 1

        return {
            "matched_cells": matched,
            "changed_cells": changed,
            "scanned_cells": scanned,
            "address": str(rng.address),
        }

    def fill_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        fill_color: str = "#FFFF00",
    ) -> dict[str, Any]:
        """지정 범위 전체의 배경색을 변경한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        if str(fill_color or "").strip().lower() in {"none", "no_fill", "transparent", "무색", "없음"}:
            # "채우기 없음" — 흰색 칠은 기본 격자선을 가린다(파일 엔진과 같은 계약).
            rng.color = None
            rows_obj = getattr(rng, "rows", None)
            cols_obj = getattr(rng, "columns", None)
            return {
                "changed_cells": int(getattr(rows_obj, "count", 1) or 1)
                * int(getattr(cols_obj, "count", 1) or 1),
                "address": str(rng.address),
            }
        rgb = self._hex_to_rgb(fill_color)
        rng.color = rgb

        start_row = int(getattr(rng, "row", 1))
        start_col = int(getattr(rng, "column", 1))
        rows_obj = getattr(rng, "rows", None)
        cols_obj = getattr(rng, "columns", None)
        row_count = int(getattr(rows_obj, "count", 1) or 1)
        col_count = int(getattr(cols_obj, "count", 1) or 1)
        for r in range(row_count):
            for c in range(col_count):
                cell_ref = f"{self._idx_to_col(start_col + c)}{start_row + r}"
                self._ensure_visual_gridline(sheet.range(cell_ref), rgb)

        return {
            "changed_cells": row_count * col_count,
            "address": str(rng.address),
        }

    def clear_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
    ) -> dict[str, Any]:
        """지정 범위의 값/수식을 비운다(서식은 유지)."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)

        # 원래 값이 있던 칸 수. 0이면 "완료"가 사용자 눈에는 무동작이라, 응답이
        # 그 사실을 말할 수 있게 지우기 전에 센다. 읽기가 실패해도 지우기는 진행한다.
        emptied = None
        try:
            before = rng.value
            flat = (
                [c for row in before for c in (row if isinstance(row, (list, tuple)) else [row])]
                if isinstance(before, (list, tuple))
                else [before]
            )
            emptied = sum(1 for v in flat if v is not None and str(v).strip() != "")
        except Exception:
            pass

        api_range = getattr(rng, "api", None)
        if api_range is not None and hasattr(api_range, "ClearContents"):
            api_range.ClearContents()
        else:
            # API 객체를 직접 쓸 수 없는 환경에서는 값 대입으로 비운다.
            rng.value = None

        rows_obj = getattr(rng, "rows", None)
        cols_obj = getattr(rng, "columns", None)
        row_count = int(getattr(rows_obj, "count", 1) or 1)
        col_count = int(getattr(cols_obj, "count", 1) or 1)
        result: dict[str, Any] = {
            "cleared_cells": row_count * col_count,
            "address": str(rng.address),
        }
        if emptied is not None:
            result["emptied_values"] = emptied
        return result

    def apply_border(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        line_style: str = "continuous",
        weight: str = "medium",
        color: str = "#000000",
    ) -> dict[str, Any]:
        """지정 범위에 경계선을 적용한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        api_range = getattr(rng, "api", None)
        if api_range is None:
            raise ExcelLiveError("경계선을 적용할 수 없습니다. Excel API 객체를 찾지 못했습니다.")

        # Excel COM 상수 (late binding)
        # -4142: xlLineStyleNone (경계선 제거)
        style_map = {"continuous": 1, "none": -4142}
        weight_map = {"thin": 2, "medium": -4138, "thick": 4}
        line_style_value = style_map.get((line_style or "").strip().lower(), 1)
        weight_value = weight_map.get((weight or "").strip().lower(), 2)
        border_color = self._rgb_to_excel_color(self._hex_to_rgb(color))
        edges = (7, 8, 9, 10, 11, 12)  # left, top, bottom, right, inside_v, inside_h

        for edge in edges:
            border = api_range.Borders(edge)
            border.LineStyle = line_style_value
            if line_style_value != -4142:
                border.Weight = weight_value
                border.Color = border_color

        rows_obj = getattr(rng, "rows", None)
        cols_obj = getattr(rng, "columns", None)
        row_count = int(getattr(rows_obj, "count", 1) or 1)
        col_count = int(getattr(cols_obj, "count", 1) or 1)
        return {
            "changed_cells": row_count * col_count,
            "address": str(rng.address),
        }

    def set_formula(
        self,
        workbook_id: str | None,
        sheet_name: str,
        range_ref: str,
        formula_a1: str,
    ) -> dict[str, Any]:
        """지정 범위에 동일한 A1 수식을 설정한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = sheet.range(range_ref)
        try:
            # Formula2가 동적 배열(SEQUENCE·FILTER 등)의 스필을 보존한다.
            # 구형 Formula 속성은 암시적 교차(@)로 강등시킨다.
            rng.api.Formula2 = formula_a1
        except Exception:
            rng.formula = formula_a1

        values = self._normalize_values(rng.options(ndim=2).value)
        row_count = len(values)
        col_count = len(values[0]) if values else 0
        return {
            "formula_applied_cells": row_count * col_count,
            "address": str(rng.address),
        }

    def verify_formula_result(
        self,
        workbook_id: str | None,
        sheet_name: str,
        range_ref: str,
    ) -> dict[str, Any]:
        """
        수식 적용 후 결과 범위를 읽어 값 검증 요약을 반환한다.
        - 비어있지 않은 셀 수
        - 숫자 셀 수 / 합계 / 평균
        - 샘플 값(최대 10개)
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = sheet.range(range_ref)
        values = self._normalize_values(rng.options(ndim=2).value)
        non_empty = 0
        numeric_values: list[float] = []
        samples: list[Any] = []
        for row in values:
            for cell in row:
                if not self._is_empty(cell):
                    non_empty += 1
                    if len(samples) < 10:
                        samples.append(cell)
                num = self._as_float(cell)
                if num is not None:
                    numeric_values.append(num)
        total = sum(numeric_values) if numeric_values else 0.0
        avg = (total / len(numeric_values)) if numeric_values else 0.0
        return {
            "address": str(rng.address),
            "non_empty_cells": non_empty,
            "numeric_cells": len(numeric_values),
            "sum": total,
            "average": avg,
            "sample_values": samples,
        }

    def create_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        start_cell: str,
        rows: int,
        cols: int,
        with_border: bool = True,
    ) -> dict[str, Any]:
        """시작 셀 기준으로 지정 크기의 표 영역을 생성한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        row_count = max(1, min(100, int(rows)))
        col_count = max(1, min(50, int(cols)))

        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = sheet.range(start_cell).resize(row_count, col_count)
        rng.value = [["" for _ in range(col_count)] for _ in range(row_count)]

        if with_border:
            api_range = getattr(rng, "api", None)
            if api_range is not None:
                # left, top, bottom, right, inside_v, inside_h
                for edge in (7, 8, 9, 10, 11, 12):
                    border = api_range.Borders(edge)
                    border.LineStyle = 1
                    border.Weight = 2
                    border.Color = self._rgb_to_excel_color((0, 0, 0))

        return {
            "created": True,
            "address": str(rng.address),
            "rows": row_count,
            "cols": col_count,
        }

    def sort_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_column: str | int = 1,
        order: str = "asc",
        has_header: bool = True,
    ) -> dict[str, Any]:
        """범위를 지정 열 기준으로 정렬한다.

        2026-08-19 파일 엔진과 계약을 맞춤: 꼬리의 집계 줄(합계·평균 이름표거나 수식이 든
        마지막 줄들)은 데이터가 아니라 **고정**하고, 정렬은 Excel의 Range.Sort로 한다 —
        예전처럼 값을 읽어 되쓰면 수식이 값으로 굳고 합계 줄이 데이터 사이로 섞였다.
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)

        values = self._normalize_values(rng.options(ndim=2).value)
        if not values:
            return {"sorted_rows": 0, "address": str(rng.address)}
        col_count = max(len(row) for row in values)
        normalized = [row + [None] * (col_count - len(row)) for row in values]
        formulas_raw = getattr(rng, "formula", None)
        has_formula_grid = (
            isinstance(formulas_raw, (list, tuple)) and bool(formulas_raw) and isinstance(formulas_raw[0], (list, tuple))
        )
        if has_formula_grid:
            formulas = [list(r) + [None] * (col_count - len(r)) for r in formulas_raw]
        else:
            formulas = [list(r) for r in normalized]
        header_row = normalized[0] if has_header else None
        body_rows = normalized[1:] if has_header else normalized
        body_formulas = formulas[1:] if has_header else formulas
        if not body_rows:
            return {"sorted_rows": 0, "address": str(rng.address)}

        agg_labels = {"합계", "총계", "계", "총합", "평균", "최대", "최소", "개수", "total", "sum", "avg", "average"}
        pinned = 0
        while len(body_rows) - pinned > 1:
            tail_vals = body_rows[-1 - pinned]
            tail_f = body_formulas[-1 - pinned] if len(body_formulas) == len(body_rows) else tail_vals
            label = tail_vals[0] if tail_vals else None
            has_formula = any(isinstance(v, str) and v.startswith("=") for v in tail_f)
            is_agg = isinstance(label, str) and label.strip().lower() in agg_labels
            if not (has_formula or is_agg):
                break
            pinned += 1
        data_count = len(body_rows) - pinned
        if data_count <= 0:
            return {"sorted_rows": 0, "address": str(rng.address)}

        start_row = int(getattr(rng, "row", 1) or 1)
        start_col = int(getattr(rng, "column", 1) or 1)
        key_idx = self._resolve_column_selector(key_column, start_col, col_count, header_row)
        reverse = str(order or "asc").strip().lower() in {"desc", "descending", "내림차순"}
        first_data_row = start_row + (1 if has_header else 0)
        last_data_row = first_data_row + data_count - 1
        left = self._col_letter(start_col)
        right = self._col_letter(start_col + col_count - 1)
        block_top = start_row if has_header else first_data_row
        block = sheet.range(f"{left}{block_top}:{right}{last_data_row}")
        key_letter = self._col_letter(start_col + key_idx)
        key_rng = sheet.range(f"{key_letter}{first_data_row}:{key_letter}{last_data_row}")
        try:
            block.api.Sort(
                Key1=key_rng.api,
                Order1=2 if reverse else 1,  # xlDescending / xlAscending
                Header=1 if has_header else 2,  # xlYes / xlNo
                Orientation=1,  # xlSortColumns
            )
        except (AttributeError, TypeError):
            # COM Sort가 없는 환경(테스트 가짜 등): 데이터 줄만 파이썬으로 세워 수식째 되쓴다.
            data_f = body_formulas[:data_count] if len(body_formulas) == len(body_rows) else body_rows[:data_count]
            data_v = body_rows[:data_count]
            order_idx = sorted(
                range(data_count),
                key=lambda i: self._sortable_value(data_v[i][key_idx] if key_idx < len(data_v[i]) else None),
                reverse=reverse,
            )
            reordered = [data_f[i] for i in order_idx]
            data_block = sheet.range(f"{left}{first_data_row}:{right}{last_data_row}")
            if has_formula_grid:
                data_block.formula = reordered
            else:
                data_block.value = reordered
        return {
            "sorted_rows": data_count,
            "address": str(rng.address),
            "key_column_index": key_idx + 1,
            "order": "desc" if reverse else "asc",
            "pinned_tail_rows": pinned,
        }

    def filter_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        column: str | int = 1,
        operator: str = "==",
        value: Any = None,
        has_header: bool = True,
        mode: str = "keep",
    ) -> dict[str, Any]:
        """조건에 맞는 행만 남기고 나머지 행을 지운다.

        `mode="remove"`면 반대로 조건에 맞는 행을 지운다 — "취소된 주문은 빼줘"처럼
        제외를 요청한 문장용이다.

        자동필터로 숨기기만 하면 파일 엔진과 같은 액션 이름이 다른 결과를 내고,
        사용자가 저장해 남기면 숨겨진 행이 그대로 살아 있다. 두 엔진이 같은 계약을
        지키도록 실제로 행을 지운다.
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        values = self._normalize_values(rng.options(ndim=2).value)
        if not values:
            return {"filtered_rows": 0, "removed_rows": 0, "address": str(rng.address)}

        col_count = max(len(row) for row in values)
        normalized = [row + [None] * (col_count - len(row)) for row in values]
        start_col = int(getattr(rng, "column", 1) or 1)
        start_row = int(getattr(rng, "row", 1) or 1)
        header_row = normalized[0] if has_header else None
        # 기준 열을 못 찾으면 거르지 않는다(dedupe와 같은 부류, 2026-08-26).
        field_idx = self._resolve_column_selector(
            column, start_col, col_count, header_row, strict=True
        )
        if field_idx < 0:
            raise ExcelLiveError(self._unresolved_filter_column_error(column, header_row))
        op = str(operator or "==").strip()
        drop_matches = str(mode or "keep").strip().lower() == "remove"

        body_offset = 1 if has_header else 0
        kept = 0
        doomed_rows: list[int] = []
        for offset, row in enumerate(normalized[body_offset:]):
            matches = self._matches_generic_condition(row[field_idx], op, value)
            if matches != drop_matches:
                kept += 1
            else:
                doomed_rows.append(start_row + body_offset + offset)

        if not drop_matches and kept == 0 and doomed_rows:
            # 조건에 맞는 행이 하나도 없다 — 그대로 지우면 시트가 통째로 빈다.
            # 조건 값이 틀렸을 가능성이 압도적이므로 파일을 건드리지 않는다
            # (파일 엔진과 같은 계약, 2026-08-17 실측).
            return {
                "filtered_rows": 0,
                "matched_rows": 0,
                "removed_rows": 0,
                "no_change": True,
                "address": str(rng.address),
            }

        # 아래에서 위로 지워야 남은 행의 번호가 밀리지 않는다.
        for row_index in reversed(doomed_rows):
            try:
                sheet.api.Rows(row_index).Delete()
            except Exception as exc:  # pragma: no cover - COM 환경 의존
                raise ExcelLiveError(f"행 삭제 실패: {exc}") from exc

        return {
            "filtered_rows": kept,
            "removed_rows": len(doomed_rows),
            "address": str(rng.address),
            "column_index": field_idx + 1,
            "operator": op,
            "value": value,
            "mode": "remove" if drop_matches else "keep",
        }

    def read_computed_range(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, Any]:
        """계산된 값 읽기. xlwings는 Excel이 이미 계산한 값을 주므로 일반 읽기와 같다."""
        return self.read_range(workbook_id, sheet_name, range_ref)

    # Excel 셀 오류의 COM 원시값 코드. xlwings `.value`는 오류 셀을 **None으로**
    # 돌려주므로(2026-08-25 실측) 값 경로에서는 오류가 빈칸과 구분되지 않는다.
    _CELL_ERROR_CODES = {
        -2146826288: "#NULL!",
        -2146826281: "#DIV/0!",
        -2146826273: "#VALUE!",
        -2146826265: "#REF!",
        -2146826259: "#NAME?",
        -2146826252: "#NUM!",
        -2146826246: "#N/A",
    }

    def count_error_cells(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, list[str]]:
        """범위 안의 오류 셀을 {오류 종류: [셀 주소…]}로 보고한다.

        수식 사후 검증용(감사 B3): `formula_applied_cells>0`은 "넣었다"만 말할 뿐
        "#NAME?이 떴다"는 못 본다. COM 원시값의 오류 코드를 직접 읽는다.
        """
        from openpyxl.utils import get_column_letter

        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            return {}
        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = sheet.range(range_ref)
        raw = rng.api.Value
        if not isinstance(raw, tuple):
            raw = ((raw,),)
        top_row = int(getattr(rng, "row", 1) or 1)
        left_col = int(getattr(rng, "column", 1) or 1)
        found: dict[str, list[str]] = {}
        for r, row in enumerate(raw):
            cells = row if isinstance(row, tuple) else (row,)
            for c, value in enumerate(cells):
                kind = self._CELL_ERROR_CODES.get(value) if isinstance(value, int) else None
                if kind:
                    found.setdefault(kind, []).append(f"{get_column_letter(left_col + c)}{top_row + r}")
        return found

    def find_duplicates(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_columns: list[str | int] | None = None,
        has_header: bool = True,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        """중복을 제거하지 않고 어디에 몇 건 있는지만 보고한다."""
        payload = self.read_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", []) if isinstance(payload, dict) else []
        address = str(payload.get("address", target_range)) if isinstance(payload, dict) else target_range
        if not values:
            return {"duplicate_groups": 0, "duplicate_rows": 0, "address": address, "samples": []}
        col_count = max(len(row) for row in values)
        header_row = values[0] if has_header else None
        body = values[1:] if has_header else values
        if key_columns:
            indexes = [
                self._resolve_column_selector(col, 1, col_count, header_row) for col in key_columns
            ]
        else:
            indexes = list(range(col_count))
        seen: dict[tuple, list[int]] = {}
        for offset, row in enumerate(body):
            key = tuple(str(row[i]) if i < len(row) else "" for i in indexes)
            seen.setdefault(key, []).append(offset + (2 if has_header else 1))
        samples = [
            {"value": " / ".join(key), "count": len(rows), "rows": rows[:10]}
            for key, rows in seen.items()
            if len(rows) > 1
        ]
        return {
            "duplicate_groups": len(samples),
            "duplicate_rows": sum(int(s["count"]) for s in samples),
            "address": address,
            "samples": samples[:20],
        }

    def recalculate(self, workbook_id: str | None, sheet_name: str | None = None) -> dict[str, Any]:
        """Excel에 전체 재계산과 연결 새로고침을 요청한다."""
        target_id = workbook_id or self._selected_workbook_id
        wb = self._find_workbook(target_id) if target_id else None
        if wb is None:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        app = getattr(wb, "app", None)
        if app is not None:
            app.calculate()
        api = getattr(wb, "api", None)
        if api is not None:
            try:
                api.RefreshAll()
            except Exception:  # pragma: no cover - COM 환경 의존
                pass
        return {"recalculated": True, "workbook_id": str(target_id), "sheets": [sheet_name or ""]}

    def export_pdf(
        self,
        workbook_id: str | None,
        sheet_name: str | None = None,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """Excel 네이티브 기능으로 PDF를 내보낸다."""
        target_id = workbook_id or self._selected_workbook_id
        wb = self._find_workbook(target_id) if target_id else None
        if wb is None:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        source = self._find_sheet(wb, sheet_name) if sheet_name else wb
        target = str(output_path or "").strip() or None
        if target:
            # 플래너가 폴더 없는 이름을 지어내면 프로세스 작업 폴더에 떨어진다. 통합문서 옆에 둔다.
            requested = Path(target)
            if not requested.is_absolute():
                book_path = Path(str(getattr(wb, "fullname", "") or ""))
                if book_path.parent.exists():
                    target = str(book_path.parent / requested.name)
        try:
            if target:
                source.to_pdf(target)
            else:
                target = str(source.to_pdf())
        except Exception as exc:  # pragma: no cover - COM 환경 의존
            raise ExcelLiveError(f"PDF 내보내기에 실패했습니다: {exc}") from exc
        return {"exported": True, "pdf_path": target, "sheet_name": sheet_name or "(전체)"}

    def dedupe_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_columns: list[str | int] | None = None,
        has_header: bool = True,
    ) -> dict[str, Any]:
        """중복 행을 제거한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        sheet = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        values = self._normalize_values(rng.options(ndim=2).value)
        if not values:
            return {"removed_rows": 0, "remaining_rows": 0, "address": str(rng.address)}

        col_count = max(len(row) for row in values)
        start_col = int(getattr(rng, "column", 1) or 1)
        header_row = values[0] if has_header else None
        if key_columns:
            resolved = [
                self._resolve_column_selector(col, start_col, col_count, header_row, strict=True)
                for col in key_columns
            ]
            # 못 찾은 기준 열은 지우기 전에 멈춘다 — 1번 열 강등은 조용한 오실행이다.
            if any(idx < 0 for idx in resolved):
                missing = [c for c, idx in zip(key_columns, resolved) if idx < 0]
                raise ExcelLiveError(self._unresolved_key_columns_error(missing, header_row))
            cols = [idx + 1 for idx in resolved]
        else:
            cols = list(range(1, col_count + 1))

        before_rows = max(0, len(values) - 1 if has_header else len(values))
        api_range = getattr(rng, "api", None)
        if api_range is None:
            raise ExcelLiveError("중복 제거를 실행할 수 없습니다. Excel API 객체를 찾지 못했습니다.")
        try:
            # xlYes=1, xlNo=2
            api_range.RemoveDuplicates(Columns=cols, Header=1 if has_header else 2)
        except Exception as exc:  # pragma: no cover - COM 환경 의존
            raise ExcelLiveError(f"중복 제거 실패: {exc}") from exc

        after_values = self._normalize_values(rng.options(ndim=2).value)
        after_rows = max(0, len(after_values) - 1 if has_header else len(after_values))
        removed = max(0, before_rows - after_rows)
        return {
            "removed_rows": removed,
            "remaining_rows": after_rows,
            "address": str(rng.address),
            "key_columns": cols,
        }

    def pivot_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        source_range: str,
        row_field: str | int,
        value_field: str | int,
        agg: str = "sum",
        column_field: str | int | None = None,
        output_sheet: str | None = None,
        output_start: str = "A1",
        has_header: bool = True,
    ) -> dict[str, Any]:
        """
        범위를 집계해 피벗 형태 요약표를 생성한다.
        (Excel PivotTable 오브젝트 대신 결정론적 집계 결과 표를 작성)
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        source_ws = self._find_sheet(wb, sheet_name)
        source_rng = self._resolve_target_range(source_ws, source_range)
        values = self._normalize_values(source_rng.options(ndim=2).value)
        if not values:
            raise ExcelLiveError("피벗 대상 데이터가 비어 있습니다.")

        col_count = max(len(row) for row in values)
        normalized = [row + [None] * (col_count - len(row)) for row in values]
        headers = normalized[0] if has_header else [f"column_{i+1}" for i in range(col_count)]
        body_rows = normalized[1:] if has_header else normalized
        start_col = int(getattr(source_rng, "column", 1) or 1)

        row_idx = self._resolve_column_selector(row_field, start_col, col_count, headers)
        value_idx = self._resolve_column_selector(value_field, start_col, col_count, headers)
        col_idx = (
            self._resolve_column_selector(column_field, start_col, col_count, headers)
            if column_field is not None
            else None
        )
        agg_name = str(agg or "sum").strip().lower()
        if agg_name not in {"sum", "avg", "count"}:
            agg_name = "sum"

        output_rows = self._build_pivot_rows(
            rows=body_rows,
            row_idx=row_idx,
            value_idx=value_idx,
            agg=agg_name,
            col_idx=col_idx,
            row_header=str(headers[row_idx]),
            value_header=str(headers[value_idx]),
        )
        out_ws = self._find_or_create_sheet(wb, output_sheet or sheet_name)
        out_rng = out_ws.range(output_start).resize(len(output_rows), len(output_rows[0]))
        out_rng.value = output_rows
        return {
            "created": True,
            "address": str(out_rng.address),
            "rows": len(output_rows),
            "cols": len(output_rows[0]),
            "sheet_name": str(getattr(out_ws, "name", "") or ""),
        }

    def delete_charts(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        """시트의 차트를 전부 지운다 — 삭제 액션 부재로 생성 슬롯에 새던 문형."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        ws = self._find_sheet(wb, sheet_name)
        try:
            deleted = int(ws.api.ChartObjects().Count)
            if deleted:
                ws.api.ChartObjects().Delete()
        except Exception as exc:  # pragma: no cover - COM 환경 의존
            raise ExcelLiveError(f"차트 삭제 실패: {exc}") from exc
        return {"deleted": deleted, "no_change": deleted == 0, "sheet": str(ws.name)}

    def create_chart(
        self,
        workbook_id: str | None,
        sheet_name: str,
        source_range: str,
        chart_type: str = "line",
        title: str | None = None,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        """지정 범위 기반 차트를 생성한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        src_ws = self._find_sheet(wb, sheet_name)
        src_rng = self._resolve_target_range(src_ws, source_range)
        out_ws = self._find_or_create_sheet(wb, output_sheet or sheet_name)
        try:
            # xlwings Chart.api가 튜플인 버전이 있어 COM ChartObjects로 직접 만든다.
            left = float(src_rng.left) + float(src_rng.width) + 20
            top = float(src_rng.top)
            chart_object = out_ws.api.ChartObjects().Add(Left=left, Top=top, Width=520, Height=320)
            chart = chart_object.Chart
            chart.SetSourceData(src_rng.api)
            chart.ChartType = self._chart_type_to_excel(chart_type)
            chart.HasTitle = True
            chart.ChartTitle.Text = str(title or "데이터 차트")
            chart_name = str(chart_object.Name)
        except Exception as exc:  # pragma: no cover - COM 환경 의존
            raise ExcelLiveError(f"차트 생성 실패: {exc}") from exc
        return {
            "created": True,
            "chart_name": chart_name,
            "chart_type": str(chart_type or "line"),
            "source_address": str(src_rng.address),
            "sheet_name": str(getattr(out_ws, "name", "") or ""),
        }

    def set_font(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        *,
        bold: bool | None = None,
        name: str | None = None,
        size: float | None = None,
        color: str | None = None,
        align: str | None = None,
    ) -> dict[str, Any]:
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        sheet = self._find_sheet(self._find_workbook(target_id), sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        # 두 엔진의 공개 계약은 같아야 한다(2026-08-19 GUI 실측 교훈: 파일 엔진에만
        # 있던 메서드를 Excel을 띄운 채 부르면 AttributeError로 죽었다).
        _xl_align = {"left": -4131, "center": -4108, "right": -4152, "justify": -4130}.get(
            _ALIGN_WORDS.get(str(align or "").strip().lower()) or ""
        )
        if _xl_align is not None:
            rng.api.HorizontalAlignment = _xl_align
        font = rng.api.Font
        if bold is not None:
            font.Bold = bool(bold)
        if name:
            font.Name = str(name)
        if size is not None:
            font.Size = float(size)
        if color:
            red, green, blue = self._hex_to_rgb(color)
            font.Color = red + (green << 8) + (blue << 16)
        shape = getattr(rng, "shape", None)
        if isinstance(shape, tuple) and len(shape) >= 2:
            changed = max(1, int(shape[0] or 1)) * max(1, int(shape[1] or 1))
        else:
            changed = int(rng.api.Rows.Count) * int(rng.api.Columns.Count)
        return {"changed_cells": changed, "address": str(rng.address), "bold": bold}

    # ---- 파일 엔진에만 있던 작업들 — Excel 앱(xlwings) 경로에도 같은 계약으로 (2026-08-19) ----
    #
    # 2026-08-19 GUI 실측: "주문건수랑 출고건수는 콤마 찍어주라"가 Excel을 띄운 채(xlwings)에서
    # `'ExcelLiveService' object has no attribute 'set_number_format'`로 실패했다. 배터리는
    # 파일 엔진(openpyxl)으로 돌아 이 구멍을 못 봤다. 두 엔진의 공개 메서드는 같아야 한다 —
    # 파일 엔진에만 있던 16개를 여기서 COM으로 구현한다(반환 키는 파일 엔진과 동일).

    def _open_target(self, workbook_id: str | None, sheet_name: str):
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        return wb, self._find_sheet(wb, sheet_name)

    @staticmethod
    def _range_cell_count(rng) -> int:
        shape = getattr(rng, "shape", None)
        if isinstance(shape, tuple) and len(shape) >= 2:
            return max(1, int(shape[0] or 1)) * max(1, int(shape[1] or 1))
        try:
            return int(rng.api.Rows.Count) * int(rng.api.Columns.Count)
        except Exception:
            return int(getattr(rng.rows, "count", 1) or 1) * int(getattr(rng.columns, "count", 1) or 1)

    def set_number_format(
        self, workbook_id: str | None, sheet_name: str, target_range: str, format_code: str
    ) -> dict[str, Any]:
        code = str(format_code or "").strip()
        if not code:
            raise ExcelLiveError("set_number_format.format_code가 비어 있습니다.")
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        rng.number_format = code
        return {"formatted_cells": self._range_cell_count(rng), "address": str(rng.address), "format_code": code}

    def merge_cells(self, workbook_id: str | None, sheet_name: str, target_range: str) -> dict[str, Any]:
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        # 병합은 왼쪽 위 값만 남긴다 — Excel이 경고창을 띄우지 않게 DisplayAlerts를 잠깐 끈다.
        app_api = getattr(getattr(sheet, "book", None), "app", None)
        api = getattr(app_api, "api", None)
        prev = None
        try:
            if api is not None:
                prev = api.DisplayAlerts
                api.DisplayAlerts = False
            rng.merge()
        finally:
            if api is not None and prev is not None:
                api.DisplayAlerts = prev
        return {"merged": True, "address": str(rng.address)}

    def unmerge_cells(self, workbook_id: str | None, sheet_name: str, target_range: str) -> dict[str, Any]:
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        merged_before = 0
        try:
            merged_before = 1 if bool(rng.api.MergeCells) else 0
        except Exception:
            merged_before = 0
        rng.unmerge()
        return {"unmerged_ranges": merged_before, "address": str(rng.address)}

    def freeze_panes(self, workbook_id: str | None, sheet_name: str, freeze_at: str | None = None) -> dict[str, Any]:
        wb, sheet = self._open_target(workbook_id, sheet_name)
        cell_ref = str(freeze_at or "A2").strip().upper() or "A2"
        window = wb.app.api.ActiveWindow
        sheet.activate()
        if cell_ref in {"NONE", "해제", "OFF"}:
            window.FreezePanes = False
            window.SplitRow = 0
            window.SplitColumn = 0
            return {"frozen": False, "freeze_at": None}
        cell = sheet.range(cell_ref)
        row_no = int(getattr(cell, "row", 1) or 1)
        col_no = int(getattr(cell, "column", 1) or 1)
        window.FreezePanes = False
        window.SplitRow = max(0, row_no - 1)
        window.SplitColumn = max(0, col_no - 1)
        window.FreezePanes = True
        return {"frozen": True, "freeze_at": cell_ref}

    def autofit_columns(self, workbook_id: str | None, sheet_name: str, target_range: str | None = None) -> dict[str, Any]:
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        if target_range and target_range not in {"__USED_RANGE__", "__ACTIVE_SELECTION__"}:
            rng = self._resolve_target_range(sheet, target_range)
        else:
            rng = sheet.used_range
        rng.columns.autofit()
        cols = int(getattr(rng.columns, "count", 0) or 0) or (
            int(rng.shape[1]) if isinstance(getattr(rng, "shape", None), tuple) else 1
        )
        return {"adjusted_columns": cols, "address": str(rng.address)}

    def find_replace(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        find_text: str,
        replace_text: str,
        *,
        match_case: bool = False,
        whole_cell: bool = False,
    ) -> dict[str, Any]:
        if not find_text:
            raise ExcelLiveError("find_replace.find_text가 비어 있습니다.")
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        values = self._normalize_values(rng.options(ndim=2).value)
        start_row = int(getattr(rng, "row", 1) or 1)
        start_col = int(getattr(rng, "column", 1) or 1)
        needle = find_text if match_case else find_text.lower()
        replaced = 0
        for r_off, row in enumerate(values or []):
            for c_off, value in enumerate(row):
                if not isinstance(value, str):
                    continue
                haystack = value if match_case else value.lower()
                new_value: str | None = None
                if whole_cell:
                    if haystack == needle:
                        new_value = replace_text
                elif needle in haystack:
                    if match_case:
                        new_value = value.replace(find_text, replace_text)
                    else:
                        out: list[str] = []
                        cursor = 0
                        while True:
                            idx = haystack.find(needle, cursor)
                            if idx == -1:
                                out.append(value[cursor:])
                                break
                            out.append(value[cursor:idx])
                            out.append(replace_text)
                            cursor = idx + len(needle)
                        new_value = "".join(out)
                if new_value is None:
                    continue
                # 바뀐 칸만 쓴다 — 격자를 통째로 되쓰면 수식이 값으로 굳는다.
                sheet.range(f"{self._col_letter(start_col + c_off)}{start_row + r_off}").value = new_value
                replaced += 1
        return {"replaced_cells": replaced, "address": str(rng.address)}

    def define_named_range(
        self, workbook_id: str | None, sheet_name: str, name: str, target_range: str
    ) -> dict[str, Any]:
        clean_name = str(name or "").strip()
        # Excel은 한글 정의 이름을 허용한다 — 영문 전용 검사가 "매출표"를 반려해
        # 사용자가 부른 이름으로 정의할 수 없었다(2026-08-26 커버리지 0906 실측).
        # 셀 주소 꼴(A1)은 Excel 자체가 금지하므로 계속 막는다.
        if (
            not clean_name
            or not re.fullmatch(r"[A-Za-z_가-힣][A-Za-z0-9_.가-힣]*", clean_name)
            or re.fullmatch(r"[A-Za-z]{1,3}\d{1,7}", clean_name)
        ):
            raise ExcelLiveError("define_named_range.name은 문자/밑줄/한글로 시작해야 하고 셀 주소 꼴일 수 없습니다.")
        wb, sheet = self._open_target(workbook_id, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        sheet_title = str(getattr(sheet, "name", sheet_name) or sheet_name).replace("'", "''")
        refers_to = f"='{sheet_title}'!{str(rng.address).replace('$', '')}"
        try:
            existing = wb.names[clean_name]
            existing.refers_to = refers_to
        except Exception:
            wb.names.add(clean_name, refers_to)
        return {"name": clean_name, "refers_to": refers_to}

    def set_print_area(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        print_area: str | None = None,
        orientation: str | None = None,
        fit_to_page: bool | None = None,
    ) -> dict[str, Any]:
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        setup = sheet.api.PageSetup
        if print_area:
            rng = self._resolve_target_range(sheet, print_area)
            resolved_area = str(rng.address).replace("$", "")
            setup.PrintArea = resolved_area
        else:
            resolved_area = str(getattr(setup, "PrintArea", "") or "").replace("$", "") or None
        orient_out = None
        if orientation:
            normalized = str(orientation).strip().lower()
            if normalized in {"landscape", "가로"}:
                setup.Orientation = 2  # xlLandscape
                orient_out = "landscape"
            elif normalized in {"portrait", "세로"}:
                setup.Orientation = 1  # xlPortrait
                orient_out = "portrait"
        if orient_out is None:
            try:
                orient_out = "landscape" if int(setup.Orientation) == 2 else "portrait"
            except Exception:
                orient_out = None
        if fit_to_page:
            setup.Zoom = False
            setup.FitToPagesWide = 1
            setup.FitToPagesTall = 1
        return {"print_area": resolved_area, "orientation": orient_out, "fit_to_page": bool(fit_to_page)}

    def add_cell_comment(
        self, workbook_id: str | None, sheet_name: str, target_range: str, text: str, author: str = "OfficeClaw AI"
    ) -> dict[str, Any]:
        if not str(text or "").strip():
            raise ExcelLiveError("add_cell_comment.text가 비어 있습니다.")
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        cell = sheet.range(f"{self._col_letter(int(rng.column))}{int(rng.row)}")
        body = f"{author}:\n{text}" if author else str(text)
        try:
            if cell.api.Comment is not None:
                cell.api.Comment.Delete()
        except Exception:
            pass
        cell.api.AddComment(body)
        return {"address": str(cell.address).replace("$", ""), "author": author, "text": str(text)}

    def describe_sheet_layout(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        """서식·수식·병합까지 담은 시트 요약. Excel 앱 경로에서는 값·수식 격자만으로 요약한다."""
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        ref = self.get_used_range_ref(workbook_id, sheet_name)
        rng = sheet.range(ref) if ref else sheet.used_range
        values = self._normalize_values(rng.options(ndim=2).value) or []
        formulas = rng.formula
        if not isinstance(formulas, (list, tuple)):
            formulas = [[formulas]]
        formula_cells: list[str] = []
        start_row = int(getattr(rng, "row", 1) or 1)
        start_col = int(getattr(rng, "column", 1) or 1)
        for r_off, row in enumerate(formulas):
            row_list = row if isinstance(row, (list, tuple)) else [row]
            for c_off, f in enumerate(row_list):
                if isinstance(f, str) and f.startswith("="):
                    formula_cells.append(f"{self._col_letter(start_col + c_off)}{start_row + r_off}")
        header = [str(v) if v is not None else "" for v in (values[0] if values else [])]
        return {
            "sheet_name": str(getattr(sheet, "name", sheet_name) or sheet_name),
            "used_range": str(rng.address).replace("$", ""),
            "row_count": len(values),
            "col_count": max((len(r) for r in values), default=0),
            "header": header,
            "formula_cells": formula_cells[:200],
            "formula_count": len(formula_cells),
        }

    @staticmethod
    def _col_letter(index: int) -> str:
        out = ""
        n = max(1, int(index))
        while n:
            n, rem = divmod(n - 1, 26)
            out = chr(65 + rem) + out
        return out

    # ---- 열 이름 기반 표 작업(파일 엔진과 같은 계약) ----

    def _table_grid(self, workbook_id: str | None, sheet_name: str) -> tuple[list[Any], list[list[Any]], int]:
        """사용 범위를 머리글 한 줄과 본문으로 나눠 준다. 열 이름 기반 작업의 공통 진입점."""
        ref = self.get_used_range_ref(workbook_id, sheet_name)
        payload = self.read_computed_range(workbook_id, sheet_name, ref)
        rows = payload.get("values") or []
        if not rows:
            raise ExcelLiveError(f"'{sheet_name}' 시트에 읽을 데이터가 없습니다.")
        width = max(len(row) for row in rows)
        grid = [list(row) + [None] * (width - len(row)) for row in rows]
        return grid[0], grid[1:], width

    @staticmethod
    def _pick_column(selector: str | int, header: list[Any], width: int) -> tuple[int, str | None]:
        """머리글 이름·한국어 개념어·열 문자를 열 번호로 바꾼다. 못 찾으면 실패한다."""
        if isinstance(selector, int):
            index = selector - 1
            if not 0 <= index < width:
                raise ExcelLiveError(f"{selector}번째 열이 범위를 벗어났습니다. (열 개수 {width})")
            return index, None
        text = str(selector or "").strip()
        if not text:
            raise ExcelLiveError("대상 열을 지정해 주세요.")
        names = [str(cell or "").strip() for cell in header]
        for index, name in enumerate(names):
            if name and name.lower() == text.lower():
                return index, name
        mapped = resolve_header(text, [name for name in names if name])
        if mapped:
            for index, name in enumerate(names):
                if name == mapped:
                    return index, name
        if re.fullmatch(r"[A-Za-z]{1,3}", text):
            index = 0
            for ch in text.upper():
                index = index * 26 + (ord(ch) - 64)
            index -= 1
            if 0 <= index < width:
                return index, names[index] or None
        available = ", ".join(n for n in names if n) or "(머리글 없음)"
        raise ExcelLiveError(f"'{text}' 열을 찾지 못했습니다. 있는 열: {available}")

    @staticmethod
    def _grid_number(value: Any) -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, datetime):
            return value.timestamp()
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day).timestamp()
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        percent = text.endswith("%")
        if percent:
            text = text[:-1].strip()
        try:
            number = float(text)
        except ValueError:
            return None
        return number / 100.0 if percent else number

    def sort_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        column: str | int,
        order: str = "asc",
    ) -> dict[str, Any]:
        header, _body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        ref = self.get_used_range_ref(workbook_id, sheet_name)
        result = self.sort_range(workbook_id, sheet_name, ref, key_column=index + 1, order=order, has_header=True)
        return {
            "sorted_rows": result.get("sorted_rows", 0),
            "order": result.get("order", "asc"),
            "column": name or self._col_letter(index + 1),
            "address": result.get("address", ref),
        }

    def calculate_column_stat(
        self,
        workbook_id: str | None,
        sheet_name: str,
        column: str | int,
        stat: str = "sum",
    ) -> dict[str, Any]:
        header, body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        numbers = [n for n in (self._grid_number(row[index]) for row in body) if n is not None]
        kind = str(stat or "sum").strip().lower()
        kind = "average" if kind in {"average", "avg", "mean"} else kind
        if kind not in {"sum", "average", "count", "max", "min"}:
            raise ExcelLiveError(f"지원하지 않는 통계입니다: {stat}")
        if kind == "count":
            value = float(len(numbers))
        elif not numbers:
            raise ExcelLiveError(f"'{name or column}' 열에 숫자가 없어 {kind} 통계를 낼 수 없습니다.")
        elif kind == "sum":
            value = float(sum(numbers))
        elif kind == "average":
            value = float(sum(numbers)) / len(numbers)
        else:
            value = float(max(numbers) if kind == "max" else min(numbers))
        return {
            "value": value,
            "column": self._col_letter(index + 1),
            "header": name,
            "stat": kind,
            "numeric_count": len(numbers),
        }

    def drop_column(self, workbook_id: str | None, sheet_name: str, column: str | int) -> dict[str, Any]:
        header, _body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        letter = self._col_letter(index + 1)
        sheet.range(f"{letter}:{letter}").api.Delete()
        return {"dropped_column": name or letter, "remaining_columns": max(0, width - 1)}

    def rename_column(
        self,
        workbook_id: str | None,
        sheet_name: str,
        column: str | int,
        new_name: str,
    ) -> dict[str, Any]:
        target = str(new_name or "").strip()
        if not target:
            raise ExcelLiveError("새 열 이름을 알려주세요.")
        header, _body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        sheet.range(f"{self._col_letter(index + 1)}1").value = target
        return {"old_name": name or self._col_letter(index + 1), "new_name": target, "column": self._col_letter(index + 1)}

    def add_column(
        self,
        workbook_id: str | None,
        sheet_name: str,
        name: str,
        formula_a1: str | None = None,
    ) -> dict[str, Any]:
        label = str(name or "").strip()
        if not label:
            raise ExcelLiveError("추가할 열 이름을 알려주세요.")
        _header, body, width = self._table_grid(workbook_id, sheet_name)
        target_col = width + 1
        letter = self._col_letter(target_col)
        _wb, sheet = self._open_target(workbook_id, sheet_name)
        sheet.range(f"{letter}1").value = label
        filled = 0
        if formula_a1 and body:
            # 첫 행 수식을 채우고 아래로 늘리면 Excel이 상대 참조를 행마다 옮겨 준다.
            first = sheet.range(f"{letter}2")
            first.formula = formula_a1
            if len(body) > 1:
                block = sheet.range(f"{letter}2:{letter}{1 + len(body)}")
                first.api.AutoFill(block.api, 0)  # xlFillDefault
            filled = len(body)
        return {"column": letter, "name": label, "formula_filled_cells": filled}

    def group_by_aggregate(
        self,
        workbook_id: str | None,
        sheet_name: str,
        group_column: str | int,
        agg: str = "sum",
        value_column: str | int | None = None,
    ) -> dict[str, Any]:
        """열 하나로 묶어 집계한 결과를 값으로 돌려준다. 시트에 쓰지는 않는다."""
        header, body, width = self._table_grid(workbook_id, sheet_name)
        group_index, group_name = self._pick_column(group_column, header, width)
        kind = str(agg or "sum").strip().lower()
        kind = "average" if kind in {"average", "avg", "mean"} else kind
        if kind not in {"sum", "average", "count", "max", "min"}:
            raise ExcelLiveError(f"지원하지 않는 집계입니다: {agg}")
        value_index: int | None = None
        value_name: str | None = None
        if value_column is not None and str(value_column).strip():
            value_index, value_name = self._pick_column(value_column, header, width)
        elif kind != "count":
            raise ExcelLiveError(f"{kind} 집계에는 대상 값 열이 필요합니다. 예: 매출")
        buckets: dict[str, dict[str, Any]] = {}
        for row in body:
            raw = row[group_index] if group_index < len(row) else None
            key = "" if raw is None else str(raw).strip().lower()
            bucket = buckets.setdefault(key, {"label": raw, "numbers": [], "count": 0})
            bucket["count"] += 1
            if value_index is not None:
                number = self._grid_number(row[value_index]) if value_index < len(row) else None
                if number is not None:
                    bucket["numbers"].append(number)
        groups: list[dict[str, Any]] = []
        for bucket in buckets.values():
            numbers = bucket["numbers"]
            if kind == "count":
                value = float(bucket["count"])
            elif not numbers:
                value = 0.0
            elif kind == "sum":
                value = float(sum(numbers))
            elif kind == "average":
                value = float(sum(numbers)) / len(numbers)
            else:
                value = float(max(numbers) if kind == "max" else min(numbers))
            groups.append({"key": bucket["label"], "value": value, "count": bucket["count"]})
        groups.sort(key=lambda g: (-g["value"], str(g["key"])))
        return {
            "agg": kind,
            "group_column": group_name or self._col_letter(group_index + 1),
            "value_column": value_name,
            "groups": groups,
        }

    def convert_to_excel_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        table_name: str = "",
        has_header: bool = True,
    ) -> dict[str, Any]:
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        sheet = self._find_sheet(self._find_workbook(target_id), sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        existing = {str(item.Name) for item in sheet.api.ListObjects}
        cleaned = re.sub(r"[^A-Za-z0-9_]", "", str(table_name or "")) or f"{sheet_name}Table"
        if cleaned[0].isdigit():
            cleaned = f"T{cleaned}"
        display = cleaned
        index = 1
        while display in existing:
            index += 1
            display = f"{cleaned}{index}"
        listed = sheet.api.ListObjects.Add(1, rng.api, True, 1 if has_header else 2)
        listed.Name = display
        listed.TableStyle = "TableStyleMedium2"
        return {
            "created": True,
            "address": str(rng.address),
            "table_name": display,
            "has_header": bool(has_header),
        }

    def apply_formula_cf(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        formula: str,
        fill_color: str = "#FFC7CE",
        font_color: str | None = "#9C0006",
    ) -> dict[str, Any]:
        formula_text = str(formula or "").strip()
        if not formula_text.startswith("="):
            formula_text = f"={formula_text}"
        if formula_text == "=":
            raise ExcelLiveError("apply_formula_cf.formula가 비어 있습니다.")
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        sheet = self._find_sheet(self._find_workbook(target_id), sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        condition = rng.api.FormatConditions.Add(Type=2, Formula1=formula_text)
        red, green, blue = self._hex_to_rgb(fill_color)
        condition.Interior.Color = red + (green << 8) + (blue << 16)
        if font_color:
            fr, fg, fb = self._hex_to_rgb(font_color)
            condition.Font.Color = fr + (fg << 8) + (fb << 16)
        return {"address": str(rng.address), "applied": True, "rule": "formula", "formula": formula_text.lstrip("=")}

    def apply_color_scale(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        *,
        min_color: str = "#F8696B",
        mid_color: str = "#FFEB84",
        max_color: str = "#63BE7B",
    ) -> dict[str, Any]:
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        sheet = self._find_sheet(self._find_workbook(target_id), sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        scale = rng.api.FormatConditions.AddColorScale(3)
        for index, hex_code in ((1, min_color), (2, mid_color), (3, max_color)):
            red, green, blue = self._hex_to_rgb(hex_code)
            scale.ColorScaleCriteria(index).FormatColor.Color = red + (green << 8) + (blue << 16)
        return {"address": str(rng.address), "applied": True, "rule": "color_scale"}

    def apply_data_bar(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        *,
        color: str = "#638EC6",
    ) -> dict[str, Any]:
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        sheet = self._find_sheet(self._find_workbook(target_id), sheet_name)
        rng = self._resolve_target_range(sheet, target_range)
        bar = rng.api.FormatConditions.AddDatabar()
        red, green, blue = self._hex_to_rgb(color)
        bar.BarColor.Color = red + (green << 8) + (blue << 16)
        return {"address": str(rng.address), "applied": True, "rule": "data_bar"}

    def validate_data(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        checks: list[str] | None = None,
        has_header: bool = True,
        date_min: str | None = None,
        date_max: str | None = None,
    ) -> dict[str, Any]:
        """범위 데이터 품질을 점검하고 이슈 요약을 반환한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        ws = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(ws, target_range)
        values = self._normalize_values(rng.options(ndim=2).value)
        if not values:
            return {"address": str(rng.address), "issues": [], "total_issues": 0}

        col_count = max(len(row) for row in values)
        normalized = [row + [None] * (col_count - len(row)) for row in values]
        rows = normalized[1:] if has_header else normalized
        check_set = {str(c).strip().lower() for c in (checks or ["empty", "negative", "outlier"])}
        start_row = int(getattr(rng, "row", 1) or 1) + (1 if has_header else 0)
        start_col = int(getattr(rng, "column", 1) or 1)
        issues: list[dict[str, Any]] = []

        if "empty" in check_set:
            empty_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, v in enumerate(row):
                    if self._is_empty(v):
                        empty_cells.append(f"{self._idx_to_col(start_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "empty", "count": len(empty_cells), "samples": empty_cells[:20]})

        if "negative" in check_set:
            negative_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, v in enumerate(row):
                    num = self._as_float(v)
                    if num is not None and num < 0:
                        negative_cells.append(f"{self._idx_to_col(start_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "negative", "count": len(negative_cells), "samples": negative_cells[:20]})

        if "outlier" in check_set:
            outlier_cells: list[str] = []
            for c_idx in range(col_count):
                numeric_values: list[tuple[int, float]] = []
                for r_idx, row in enumerate(rows):
                    num = self._as_float(row[c_idx] if c_idx < len(row) else None)
                    if num is not None:
                        numeric_values.append((r_idx, num))
                if len(numeric_values) < 4:
                    continue
                mean = sum(v for _, v in numeric_values) / len(numeric_values)
                variance = sum((v - mean) ** 2 for _, v in numeric_values) / len(numeric_values)
                std = math.sqrt(variance)
                if std <= 0:
                    continue
                for r_idx, val in numeric_values:
                    z = abs((val - mean) / std)
                    if z >= 3.0:
                        outlier_cells.append(f"{self._idx_to_col(start_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "outlier", "count": len(outlier_cells), "samples": outlier_cells[:20]})

        if "date_range" in check_set and (date_min or date_max):
            min_dt = self._parse_datetime_value(date_min) if date_min else None
            max_dt = self._parse_datetime_value(date_max) if date_max else None
            invalid_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, v in enumerate(row):
                    dt = self._parse_datetime_value(v)
                    if dt is None:
                        continue
                    if min_dt and dt < min_dt:
                        invalid_cells.append(f"{self._idx_to_col(start_col + c_idx)}{start_row + r_idx}")
                        continue
                    if max_dt and dt > max_dt:
                        invalid_cells.append(f"{self._idx_to_col(start_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "date_range", "count": len(invalid_cells), "samples": invalid_cells[:20]})

        total = sum(int(it.get("count", 0) or 0) for it in issues)
        return {"address": str(rng.address), "issues": issues, "total_issues": total}

    def protect_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        password: str | None = None,
        lock_formula_cells: bool = True,
        unlock_range: str | None = None,
    ) -> dict[str, Any]:
        """시트 보호/잠금 설정을 적용한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        ws = self._find_sheet(wb, sheet_name)
        ws_api = getattr(ws, "api", None)
        if ws_api is None:
            raise ExcelLiveError("시트 보호를 적용할 수 없습니다. Excel API 객체를 찾지 못했습니다.")
        pwd = str(password or "")
        try:
            ws_api.Unprotect(Password=pwd)
        except Exception:
            pass

        used = getattr(ws, "used_range", None) or ws.range("A1")
        try:
            used.api.Locked = False
        except Exception:
            pass

        if lock_formula_cells:
            try:
                # xlCellTypeFormulas = -4123
                formula_cells = used.api.SpecialCells(-4123)
                formula_cells.Locked = True
            except Exception:
                # 수식 셀이 없으면 무시
                pass

        unlocked_address = ""
        if unlock_range:
            try:
                unlock_rng = self._resolve_target_range(ws, unlock_range)
                unlock_rng.api.Locked = False
                unlocked_address = str(unlock_rng.address)
            except Exception as exc:
                raise ExcelLiveError(f"잠금 해제 범위 적용 실패: {exc}") from exc

        try:
            ws_api.Protect(Password=pwd, UserInterfaceOnly=True)
        except Exception as exc:
            raise ExcelLiveError(f"시트 보호 적용 실패: {exc}") from exc
        return {
            "protected": True,
            "sheet_name": str(getattr(ws, "name", "") or ""),
            "lock_formula_cells": bool(lock_formula_cells),
            "unlock_range": unlocked_address,
        }

    def set_data_validation(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        target_range: str,
        validation_type: str = "list",
        source: str | None = None,
        minimum: float | None = None,
        maximum: float | None = None,
        allow_blank: bool = True,
        show_error: bool = True,
        error_message: str | None = None,
    ) -> dict[str, Any]:
        """입력 유효성(드롭다운/숫자/날짜 제한)을 설정한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        ws = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(ws, target_range)
        api_rng = getattr(rng, "api", None)
        if api_rng is None:
            raise ExcelLiveError("유효성 검사를 적용할 수 없습니다. Excel API 객체를 찾지 못했습니다.")

        vtype = str(validation_type or "list").strip().lower()
        type_map = {
            "list": 3,  # xlValidateList
            "whole": 1,  # xlValidateWholeNumber
            "decimal": 2,  # xlValidateDecimal
            "date": 4,  # xlValidateDate
        }
        if vtype not in type_map:
            raise ExcelLiveError(f"지원하지 않는 validation_type: {validation_type}")
        try:
            api_rng.Validation.Delete()
        except Exception:
            pass

        try:
            if vtype == "list":
                formula1 = str(source or "").strip()
                if not formula1:
                    raise ExcelLiveError("list 유효성은 source(예: 완료,진행중,지연)가 필요합니다.")
                api_rng.Validation.Add(
                    Type=type_map[vtype],
                    AlertStyle=1,  # xlValidAlertStop
                    Operator=1,  # xlBetween
                    Formula1=formula1,
                )
            elif vtype in {"whole", "decimal"}:
                if minimum is None or maximum is None:
                    raise ExcelLiveError("숫자 유효성은 minimum/maximum이 필요합니다.")
                api_rng.Validation.Add(
                    Type=type_map[vtype],
                    AlertStyle=1,
                    Operator=1,
                    Formula1=str(minimum),
                    Formula2=str(maximum),
                )
            else:  # date
                if minimum is None or maximum is None:
                    raise ExcelLiveError("날짜 유효성은 minimum/maximum(Excel serial 또는 YYYYMMDD 숫자)이 필요합니다.")
                api_rng.Validation.Add(
                    Type=type_map[vtype],
                    AlertStyle=1,
                    Operator=1,
                    Formula1=str(minimum),
                    Formula2=str(maximum),
                )
            api_rng.Validation.IgnoreBlank = bool(allow_blank)
            api_rng.Validation.ShowError = bool(show_error)
            if error_message:
                api_rng.Validation.ErrorMessage = str(error_message)
        except Exception as exc:
            raise ExcelLiveError(f"유효성 검사 설정 실패: {exc}") from exc
        return {
            "applied": True,
            "address": str(rng.address),
            "validation_type": vtype,
        }

    def consolidate_sheets(
        self,
        workbook_id: str | None,
        *,
        source_sheets: list[str],
        output_sheet: str = "통합결과",
        include_header_once: bool = True,
        add_source_sheet_col: bool = True,
    ) -> dict[str, Any]:
        """같은 통합문서 내 여러 시트를 하나로 합친다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        if not source_sheets:
            raise ExcelLiveError("source_sheets가 비어 있습니다.")
        wb = self._find_workbook(target_id)
        out_ws = self._find_or_create_sheet(wb, output_sheet)
        merged: list[list[Any]] = []
        header_written = False
        body_rows = 0
        for sname in source_sheets:
            ws = self._find_sheet(wb, sname)
            values = self._normalize_values(ws.used_range.options(ndim=2).value)
            if not values:
                continue
            header = values[0]
            body = values[1:] if len(values) > 1 else []
            if include_header_once and not header_written:
                out_header = (["source_sheet"] if add_source_sheet_col else []) + list(header)
                merged.append(out_header)
                header_written = True
            elif not include_header_once:
                out_header = (["source_sheet"] if add_source_sheet_col else []) + list(header)
                merged.append(out_header)
            for row in body:
                merged.append(( [sname] if add_source_sheet_col else [] ) + list(row))
                body_rows += 1
        if not merged:
            raise ExcelLiveError("통합할 데이터가 없습니다.")
        max_cols = max(len(r) for r in merged)
        normalized = [r + [None] * (max_cols - len(r)) for r in merged]
        out_rng = out_ws.range("A1").resize(len(normalized), max_cols)
        out_rng.value = normalized
        return {
            "created": True,
            "sheet_name": str(getattr(out_ws, "name", "") or ""),
            "address": str(out_rng.address),
            "rows": len(normalized),
            "cols": max_cols,
            "merged_rows": body_rows,
        }

    def consolidate_workbooks_from_folder(
        self,
        workbook_id: str | None,
        *,
        folder_path: str,
        pattern: str = "*.xlsx",
        source_sheet: str | None = None,
        output_sheet: str = "파일통합결과",
        include_header_once: bool = True,
        add_source_file_col: bool = True,
    ) -> dict[str, Any]:
        """폴더 내 여러 파일의 시트를 현재 통합문서로 합친다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        root = Path(folder_path).expanduser()
        if not root.exists() or not root.is_dir():
            raise ExcelLiveError(f"유효하지 않은 폴더 경로: {folder_path}")
        files = sorted(root.glob(pattern))
        if not files:
            raise ExcelLiveError("통합할 파일이 없습니다.")

        xw = self._xw_module()
        wb_out = self._find_workbook(target_id)
        out_ws = self._find_or_create_sheet(wb_out, output_sheet)
        merged: list[list[Any]] = []
        header_written = False
        merged_rows = 0
        opened_count = 0
        for fp in files:
            wb_in = None
            try:
                wb_in = xw.Book(str(fp))
                opened_count += 1
                if source_sheet:
                    ws_in = self._find_sheet(wb_in, source_sheet)
                else:
                    ws_in = wb_in.sheets[0]
                values = self._normalize_values(ws_in.used_range.options(ndim=2).value)
                if not values:
                    continue
                header = values[0]
                body = values[1:] if len(values) > 1 else []
                if include_header_once and not header_written:
                    out_header = (["source_file"] if add_source_file_col else []) + list(header)
                    merged.append(out_header)
                    header_written = True
                elif not include_header_once:
                    out_header = (["source_file"] if add_source_file_col else []) + list(header)
                    merged.append(out_header)
                for row in body:
                    merged.append(([fp.name] if add_source_file_col else []) + list(row))
                    merged_rows += 1
            finally:
                try:
                    if wb_in is not None:
                        wb_in.close()
                except Exception:
                    pass
        if not merged:
            raise ExcelLiveError("파일에서 읽은 데이터가 없습니다.")
        max_cols = max(len(r) for r in merged)
        normalized = [r + [None] * (max_cols - len(r)) for r in merged]
        out_rng = out_ws.range("A1").resize(len(normalized), max_cols)
        out_rng.value = normalized
        return {
            "created": True,
            "sheet_name": str(getattr(out_ws, "name", "") or ""),
            "address": str(out_rng.address),
            "rows": len(normalized),
            "cols": max_cols,
            "opened_files": opened_count,
            "merged_rows": merged_rows,
        }

    def refresh_power_query(self, workbook_id: str | None) -> dict[str, Any]:
        """통합문서의 연결/쿼리를 새로고침한다(RefreshAll)."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        api_wb = getattr(wb, "api", None)
        if api_wb is None:
            raise ExcelLiveError("Power Query 새로고침을 실행할 수 없습니다.")
        try:
            api_wb.RefreshAll()
        except Exception as exc:
            raise ExcelLiveError(f"Power Query 새로고침 실패: {exc}") from exc
        return {"refreshed": True, "workbook_id": self._workbook_id(wb)}

    def run_vba_macro(
        self,
        workbook_id: str | None,
        *,
        macro_name: str,
        args: list[Any] | None = None,
    ) -> dict[str, Any]:
        """VBA 매크로를 **지정한 통합문서 안에서** 실행한다.

        `Application.Run("매크로명")`은 이름을 ActiveWorkbook·PERSONAL.XLSB 등
        열려 있는 아무 통합문서에서 해석한다. 사용자가 "A 파일에서 실행"을
        승인했는데 B 파일의 동명 매크로가 도는 일이 가능했다 — 승인한 대상과
        실행된 대상이 다르면 승인 자체가 무의미하다. `'파일명'!매크로` 형태로
        대상을 못 박는다.
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        app = self._app()
        macro = _validate_macro_name(macro_name)
        wb_name = str(getattr(wb, "name", "") or "")
        if not wb_name:
            raise ExcelLiveError("대상 통합문서 이름을 확인할 수 없어 매크로를 실행하지 않았습니다.")
        if not self._workbook_has_macro(wb, macro):
            raise ExcelLiveError(
                f"'{wb_name}'에 매크로 '{macro}'가 없습니다. 대상 파일과 매크로 이름을 확인해 주세요."
            )
        # 작은따옴표로 감싸야 공백이 든 파일명도 한 토큰으로 읽힌다.
        qualified = f"'{wb_name}'!{macro}"
        macro_args = list(args or [])
        try:
            app.api.Run(qualified, *macro_args)
        except Exception as exc:
            raise ExcelLiveError(f"VBA 매크로 실행 실패: {exc}") from exc
        return {
            "executed": True,
            "macro_name": macro,
            "qualified_macro": qualified,
            "workbook_name": wb_name,
            "args_count": len(macro_args),
        }

    @staticmethod
    def _workbook_has_macro(wb: Any, macro: str) -> bool:
        """대상 통합문서의 VBA 프로젝트에 그 프로시저가 있는가.

        VBA 프로젝트 접근이 막힌 환경(신뢰 설정 off)에서는 확인할 방법이 없다.
        그때는 통과시킨다 — 여기서 막으면 정상 매크로도 못 돌린다. 대상 한정은
        호출 문자열이 이미 보장한다.
        """
        module_name = str(macro).split(".", 1)[0] if "." in macro else None
        procedure = str(macro).rsplit(".", 1)[-1]
        try:
            components = wb.api.VBProject.VBComponents
        except Exception:
            return True
        # `CodeModule.Find`는 인자가 ByRef라 COM 경유로 부르기 까다롭다.
        # 소스를 그대로 읽어 선언부를 찾는 편이 환경을 덜 탄다.
        declaration = re.compile(
            rf"^\s*(?:public\s+|private\s+|friend\s+)?(?:static\s+)?(?:sub|function)\s+{re.escape(procedure)}\b",
            re.IGNORECASE | re.MULTILINE,
        )
        try:
            for component in components:
                if module_name and str(getattr(component, "Name", "")) != module_name:
                    continue
                code = component.CodeModule
                line_count = int(getattr(code, "CountOfLines", 0) or 0)
                if line_count <= 0:
                    continue
                if declaration.search(str(code.Lines(1, line_count) or "")):
                    return True
            return False
        except Exception:
            return True

    def compare_ranges(
        self,
        workbook_id: str | None,
        *,
        left_sheet: str,
        left_range: str,
        right_sheet: str,
        right_range: str,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        """두 범위를 셀 단위 비교해 차이를 반환/기록한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        ws_l = self._find_sheet(wb, left_sheet)
        ws_r = self._find_sheet(wb, right_sheet)
        rng_l = self._resolve_target_range(ws_l, left_range)
        rng_r = self._resolve_target_range(ws_r, right_range)
        vals_l = self._normalize_values(rng_l.options(ndim=2).value)
        vals_r = self._normalize_values(rng_r.options(ndim=2).value)
        rows = max(len(vals_l), len(vals_r))
        cols = max(max((len(r) for r in vals_l), default=0), max((len(r) for r in vals_r), default=0))
        diffs: list[list[Any]] = [["row", "col", "left_value", "right_value"]]
        diff_count = 0
        for r in range(rows):
            for c in range(cols):
                lv = vals_l[r][c] if r < len(vals_l) and c < len(vals_l[r]) else None
                rv = vals_r[r][c] if r < len(vals_r) and c < len(vals_r[r]) else None
                if str(lv) != str(rv):
                    diff_count += 1
                    diffs.append([r + 1, c + 1, lv, rv])
        result: dict[str, Any] = {
            "left_address": str(rng_l.address),
            "right_address": str(rng_r.address),
            "diff_cells": diff_count,
            "sample_diffs": diffs[1:21],
        }
        if output_sheet:
            out_ws = self._find_or_create_sheet(wb, output_sheet)
            out_rng = out_ws.range("A1").resize(len(diffs), len(diffs[0]))
            out_rng.value = diffs
            result["output_sheet"] = str(getattr(out_ws, "name", "") or "")
            result["output_address"] = str(out_rng.address)
        return result

    def forecast_linear(
        self,
        workbook_id: str | None,
        *,
        sheet_name: str,
        source_range: str,
        horizon: int = 3,
        output_sheet: str | None = None,
        output_start: str = "A1",
    ) -> dict[str, Any]:
        """단순 선형회귀 기반 예측값을 생성한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("workbook_id가 필요합니다.")
        wb = self._find_workbook(target_id)
        ws = self._find_sheet(wb, sheet_name)
        rng = self._resolve_target_range(ws, source_range)
        values = self._normalize_values(rng.options(ndim=2).value)
        series: list[float] = []
        for row in values:
            if not row:
                continue
            num = self._as_float(row[-1])
            if num is not None:
                series.append(num)
        if len(series) < 2:
            raise ExcelLiveError("예측을 위해 최소 2개 이상의 숫자 데이터가 필요합니다.")
        n = len(series)
        xs = list(range(1, n + 1))
        sum_x = sum(xs)
        sum_y = sum(series)
        sum_xx = sum(x * x for x in xs)
        sum_xy = sum(x * y for x, y in zip(xs, series))
        denom = (n * sum_xx) - (sum_x * sum_x)
        slope = ((n * sum_xy) - (sum_x * sum_y)) / denom if denom else 0.0
        intercept = (sum_y - slope * sum_x) / n
        horizon_n = max(1, min(36, int(horizon)))
        out_rows: list[list[Any]] = [["step", "forecast"]]
        for step in range(1, horizon_n + 1):
            x = n + step
            out_rows.append([x, slope * x + intercept])
        out_ws = self._find_or_create_sheet(wb, output_sheet or sheet_name)
        out_rng = out_ws.range(output_start).resize(len(out_rows), 2)
        out_rng.value = out_rows
        return {
            "created": True,
            "sheet_name": str(getattr(out_ws, "name", "") or ""),
            "address": str(out_rng.address),
            "horizon": horizon_n,
            "slope": slope,
            "intercept": intercept,
        }

    def save_workbook(self, workbook_id: str | None) -> dict[str, Any]:
        """현재 통합문서를 디스크에 저장한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError(
                "저장할 통합문서를 찾지 못했습니다. workbook_id를 지정하거나 select_workbook을 먼저 호출해 주세요."
            )

        wb = self._find_workbook(target_id)
        full_path = str(getattr(wb, "fullname", "") or "").strip()
        if not full_path:
            raise ExcelLiveError(
                "아직 파일 경로가 없는 통합문서입니다. Excel에서 먼저 다른 이름으로 저장해 주세요."
            )
        try:
            wb.save()
        except Exception as exc:  # pragma: no cover - COM 환경 의존
            raise ExcelLiveError(f"통합문서 저장에 실패했습니다: {exc}") from exc

        return {
            "saved": True,
            "workbook_id": self._workbook_id(wb),
            "name": str(getattr(wb, "name", "") or ""),
            "full_path": full_path,
        }

    def create_workbook_backup(
        self,
        workbook_id: str | None,
        *,
        label: str = "auto",
    ) -> dict[str, Any]:
        """
        현재 통합문서의 복구용 백업 사본을 생성한다.

        - 기본 저장 위치: 원본 파일 폴더의 `officeclaw_backups/`
        - 저장 방식: Excel COM SaveCopyAs 우선, 실패 시 파일 복사 fallback
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("백업할 통합문서를 찾지 못했습니다. workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        full_path = str(getattr(wb, "fullname", "") or "").strip()
        if not full_path:
            raise ExcelLiveError(
                "아직 파일 경로가 없는 통합문서입니다. 자동 복구 백업을 위해 먼저 다른 이름으로 저장해 주세요."
            )
        src = Path(full_path)
        if not src.exists() or not src.is_file():
            raise ExcelLiveError(f"원본 파일 경로가 유효하지 않습니다: {full_path}")

        backup_dir = src.parent / "officeclaw_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_label = re.sub(r"[^A-Za-z0-9_-]+", "_", str(label or "auto")).strip("_") or "auto"
        backup_name = f"{src.stem}.{safe_label}.{stamp}{src.suffix}"
        backup_path = backup_dir / backup_name
        suffix = 1
        while backup_path.exists():
            backup_path = backup_dir / f"{src.stem}.{safe_label}.{stamp}_{suffix}{src.suffix}"
            suffix += 1

        try:
            api_wb = getattr(wb, "api", None)
            if api_wb is not None and hasattr(api_wb, "SaveCopyAs"):
                api_wb.SaveCopyAs(str(backup_path))
            else:
                shutil.copy2(src, backup_path)
        except Exception as exc:
            raise ExcelLiveError(f"자동 복구 백업 생성 실패: {exc}") from exc

        return {
            "backup_created": True,
            "backup_path": str(backup_path),
            "source_path": str(src),
            "workbook_id": self._workbook_id(wb),
        }

    def list_workbook_backups(
        self,
        workbook_id: str | None,
        *,
        limit: int = 20,
    ) -> dict[str, Any]:
        """대상 통합문서의 복구 백업 목록을 최신순으로 반환한다."""
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("백업 목록을 조회할 통합문서를 찾지 못했습니다. workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        full_path = str(getattr(wb, "fullname", "") or "").strip()
        if not full_path:
            raise ExcelLiveError("아직 파일 경로가 없는 통합문서입니다. 먼저 다른 이름으로 저장해 주세요.")

        src = Path(full_path).resolve()
        backup_dir = src.parent / "officeclaw_backups"
        max_items = max(1, min(200, int(limit)))
        rows: list[dict[str, Any]] = []

        if backup_dir.exists() and backup_dir.is_dir():
            pattern = f"{src.stem}.*{src.suffix}"
            for fp in backup_dir.glob(pattern):
                if not fp.is_file():
                    continue
                try:
                    stat = fp.stat()
                    rows.append(
                        {
                            "backup_path": str(fp.resolve()),
                            "backup_name": fp.name,
                            "size_bytes": int(stat.st_size),
                            "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                            "modified_ts": float(stat.st_mtime),
                        }
                    )
                except Exception:
                    continue

        rows.sort(key=lambda item: float(item.get("modified_ts", 0.0)), reverse=True)
        for row in rows:
            row.pop("modified_ts", None)

        return {
            "workbook_id": self._workbook_id(wb),
            "source_path": str(src),
            "backup_dir": str(backup_dir),
            "backups": rows[:max_items],
        }

    def restore_workbook_from_backup(
        self,
        workbook_id: str | None,
        *,
        backup_path: str | None = None,
    ) -> dict[str, Any]:
        """
        백업 파일로 통합문서를 복구한다.

        동작:
        1) 현재 파일을 pre_restore 백업으로 1회 보존
        2) 현재 통합문서를 저장하지 않고 닫음
        3) 선택한 백업 파일을 원본 파일 경로로 덮어씀
        4) 원본 파일을 다시 열어 작업 계속
        """
        target_id = workbook_id or self._selected_workbook_id
        if not target_id:
            raise WorkbookNotFoundError("복구할 통합문서를 찾지 못했습니다. workbook_id가 필요합니다.")

        wb = self._find_workbook(target_id)
        source_path = str(getattr(wb, "fullname", "") or "").strip()
        if not source_path:
            raise ExcelLiveError("파일 경로가 없는 통합문서는 복구할 수 없습니다.")
        src = Path(source_path).resolve()
        if not src.exists() or not src.is_file():
            raise ExcelLiveError(f"원본 통합문서 경로가 유효하지 않습니다: {source_path}")

        if backup_path:
            chosen_backup = Path(str(backup_path)).expanduser().resolve()
        else:
            listed = self.list_workbook_backups(target_id, limit=1)
            backups = listed.get("backups", [])
            if not backups:
                raise ExcelLiveError("복구 가능한 백업이 없습니다.")
            chosen_backup = Path(str(backups[0].get("backup_path", ""))).resolve()

        if not chosen_backup.exists() or not chosen_backup.is_file():
            raise ExcelLiveError(f"백업 파일을 찾을 수 없습니다: {chosen_backup}")

        pre_restore_info = self.create_workbook_backup(target_id, label="pre_restore")

        try:
            api_wb = getattr(wb, "api", None)
            if api_wb is not None and hasattr(api_wb, "Close"):
                api_wb.Close(SaveChanges=False)
            else:
                wb.close()
        except Exception as exc:
            raise ExcelLiveError(f"복구 전 통합문서 닫기 실패: {exc}") from exc

        try:
            shutil.copy2(chosen_backup, src)
        except Exception as exc:
            raise ExcelLiveError(f"백업 파일 복구 실패: {exc}") from exc

        try:
            reopened = self._xw_module().Book(str(src))
            self._selected_workbook_id = self._workbook_id(reopened)
        except Exception as exc:
            raise ExcelLiveError(f"복구 후 통합문서 다시 열기 실패: {exc}") from exc

        return {
            "restored": True,
            "workbook_id": self._workbook_id(reopened),
            "name": str(getattr(reopened, "name", "") or ""),
            "full_path": str(src),
            "restored_from_backup_path": str(chosen_backup),
            "pre_restore_backup_path": str(pre_restore_info.get("backup_path", "")),
            "active_sheet": self._active_sheet_name(reopened),
        }

    def _resolve_workbook(self, workbook_id: str | None):
        """대상 통합문서를 찾는다. 지목이 없으면 **활성 통합문서**로 떨어진다.

        2026-08-16 실측: 예전에는 `workbook_id`도 `_selected_workbook_id`도 없으면
        곧장 "A1"을 돌려줬다. 그런데 앱은 workbook_id를 안 보내고(요청이 null),
        사용자가 `/excel-live/select-workbook`을 부른 적도 없다 — 즉 **평소 경로가
        전부 그 폴백을 탔다.** Excel의 실제 선택이 `$A$1:$G$8`인데 `A1`이 반환돼
        드래그한 영역이 통째로 무시됐고, "표 크기를 알려주세요"로 되물었다.
        """
        target_id = workbook_id or self._selected_workbook_id
        if target_id:
            # 사용자가 지목했으면 가드를 거치지 않는다 — 그 파일을 정말 원한 것이다.
            return self._find_workbook(target_id)
        active = self._app().books.active
        if not _is_user_workbook_path(getattr(active, "fullname", "")):
            # 활성 통합문서가 벤더·가상환경 파일이다. 여기에 조용히 실행하면
            # 사용자가 보지도 않는 파일이 바뀐다(2026-08-04 실측, 감사 B1).
            raise WorkbookNotFoundError(
                "활성 통합문서가 작업 파일이 아니라 라이브러리 데모 파일입니다. "
                "작업할 파일을 먼저 선택해 주세요."
            )
        return active

    def get_active_selection_ref(
        self,
        workbook_id: str | None,
        sheet_name: str | None,
    ) -> str:
        """
        범위를 말하지 않은 명령이 대상으로 삼을 영역을 A1 표기로 반환한다.

        사용자가 여러 칸을 실제로 끌어서 선택했으면 그 선택을 존중한다. 하지만 커서가
        한 칸에 놓여 있을 뿐이면 "범위를 지정하지 않았다"는 뜻이므로, Excel이 정렬·필터에서
        하듯 그 셀을 둘러싼 데이터 영역으로 넓힌다. 한 칸을 그대로 돌려주면 정렬·집계·중복제거가
        한 칸에만 적용된 채 성공으로 보고돼, 화면은 그대로인데 "처리했습니다"만 남는다.
        """
        try:
            wb = self._resolve_workbook(workbook_id)
        except Exception:
            return "A1"
        sheet = self._find_sheet(wb, sheet_name) if sheet_name else wb.sheets.active

        selected = self._selection_ref_on_sheet(sheet)
        if ":" in selected:
            return selected

        return self._data_region_ref(sheet, selected) or selected or "A1"

    def _selection_ref_on_sheet(self, sheet: Any) -> str:
        """앱의 현재 선택이 이 시트의 것일 때만 A1 표기로 돌려준다.

        선택은 사용자가 마지막으로 클릭한 곳이라 다른 통합문서·다른 시트일 수 있다.
        그걸 그대로 쓰면 엉뚱한 시트의 주소로 편집하게 된다.
        """
        try:
            selection = getattr(self._app(), "selection", None)
            if selection is None:
                return ""
            sel_sheet = getattr(selection, "sheet", None)
            if sel_sheet is not None:
                if str(getattr(sel_sheet, "name", "")) != str(getattr(sheet, "name", "")):
                    return ""
                sel_book = getattr(sel_sheet, "book", None)
                target_book = getattr(sheet, "book", None)
                if (
                    sel_book is not None
                    and target_book is not None
                    and str(getattr(sel_book, "name", "")) != str(getattr(target_book, "name", ""))
                ):
                    return ""
            return self._normalize_address_ref(str(getattr(selection, "address", "") or ""))
        except Exception:
            return ""

    def _data_region_ref(self, sheet: Any, anchor: str) -> str:
        """기준 셀을 둘러싼 데이터 영역. 못 구하면 시트 사용 영역으로 떨어진다."""
        for ref in (anchor or "A1", "A1"):
            try:
                region = getattr(sheet.range(ref), "current_region", None)
                if region is None:
                    continue
                address = self._normalize_address_ref(str(getattr(region, "address", "") or ""))
                if ":" in address:
                    return address
            except Exception:
                continue
        try:
            used = getattr(sheet, "used_range", None)
            if used is None:
                return ""
            return self._normalize_address_ref(str(getattr(used, "address", "") or ""))
        except Exception:
            return ""

    def get_used_range_ref(
        self,
        workbook_id: str | None,
        sheet_name: str | None,
    ) -> str:
        """
        현재 시트의 사용 영역(used_range)을 A1 표기 문자열로 반환한다.

        - "전체 지우기/초기화"처럼 범위를 명시하지 않은 명령의 기본 타깃으로 사용.
        - used_range를 얻지 못하면 A1로 폴백.
        """
        try:
            wb = self._resolve_workbook(workbook_id)
        except Exception:
            return "A1"
        sheet = self._find_sheet(wb, sheet_name) if sheet_name else wb.sheets.active
        try:
            used = getattr(sheet, "used_range", None)
            if used is None:
                return "A1"
            address = str(getattr(used, "address", "") or "")
            cleaned = self._normalize_address_ref(address)
            if cleaned:
                return cleaned
        except Exception:
            pass
        return "A1"

    @staticmethod
    def _path_is_unwritable(wb_api: Any) -> bool:
        """통합문서 파일이 디스크에서 실제로 쓰기 불가인가.

        저장된 적 없는 새 통합문서(FullName이 경로가 아님)는 쓰기 가능으로 본다.
        """
        try:
            full = str(getattr(wb_api, "FullName", "") or "")
        except Exception:
            return False
        if not full or not os.path.isabs(full):
            return False
        try:
            return os.path.exists(full) and not os.access(full, os.W_OK)
        except Exception:
            return False

    def get_workbook_path(self, workbook_id: str | None = None) -> str:
        """대상 통합문서의 실제 파일 경로. 저장된 적 없으면 빈 문자열."""
        try:
            wb = self._resolve_workbook(workbook_id)
            full = str(getattr(wb.api, "FullName", "") or "")
        except Exception:
            return ""
        return full if os.path.isabs(full) else ""

    def close_workbook_without_saving(self, workbook_id: str | None = None) -> str:
        """Excel에서 통합문서를 닫는다(저장하지 않음). 닫은 파일 경로를 돌려준다.

        **읽기 전용 통합문서에만 쓴다.** 읽기 전용이면 저장되지 않은 변경이 있을 수
        없으므로 닫아도 사용자가 잃는 것이 없다. 편집 가능한 통합문서에 이걸 쓰면
        작업 중이던 내용이 사라진다 — 호출부가 반드시 상태를 확인해야 한다.

        왜 필요한가 (2026-08-16 실측): 이 PC의 Excel은 정품 인증이 안 돼 여는 파일이
        전부 읽기 전용이다. 그 상태에서 Excel은 파일에 **배타적 잠금**을 걸어
        openpyxl조차 못 쓴다(PermissionError). 즉 Excel이 파일을 붙들고 있는 한
        어떤 방법으로도 편집이 불가능하다. 닫아야 길이 열린다.
        """
        path = self.get_workbook_path(workbook_id)
        try:
            wb = self._resolve_workbook(workbook_id)
            wb.close()
        except Exception:
            return ""
        return path

    def open_workbook_in_excel(self, path: str) -> bool:
        """파일을 Excel에서 다시 연다. 편집 결과를 사용자가 바로 보게 하기 위함.

        마지막 통합문서를 닫으면 **Excel 앱 자체가 종료된다** — 그러면
        `xw.apps.active`가 None이라 `self._app()`이 실패한다(2026-08-16 실측:
        브리지가 편집은 성공했는데 창이 안 돌아왔다). 살아 있는 인스턴스가 없으면
        새로 띄운다.
        """
        try:
            xw = self._xw_module()
            apps = list(getattr(xw, "apps", []) or [])
            app = apps[0] if apps else xw.App(visible=True, add_book=False)
            app.visible = True
            app.books.open(str(path))
            return True
        except Exception:
            return False

    def get_write_protection(
        self, workbook_id: str | None, sheet_name: str | None = None
    ) -> dict[str, Any]:
        """편집을 막아야 하는 보호 상태를 읽는다(F-08).

        COM 속성은 Excel 버전에 따라 없을 수 있으므로 하나씩 getattr로 감싼다.
        하나가 없다고 전체를 포기하면 나머지 방어까지 사라진다.
        """
        flags: dict[str, Any] = {"sheet_name": sheet_name or ""}

        def _flag(obj: Any, name: str) -> bool:
            try:
                return bool(getattr(obj, name))
            except Exception:
                return False

        try:
            # 지목이 없으면 활성 통합문서로 떨어진다. 예전에는 빈 문자열로 조회해
            # 예외가 나고 플래그가 통째로 비었다 — 즉 보호 상태를 **한 번도 못 읽었다**.
            wb = self._resolve_workbook(workbook_id)
        except Exception:
            return flags
        wb_api = getattr(wb, "api", None)
        if wb_api is None:
            return flags
        # `wb.api.ReadOnly`를 그대로 믿는다 (2026-08-16, 두 번 재고 나서 확정).
        #
        # 처음엔 이걸 오탐으로 봤다 — 갓 만들어 저장한 파일도 ReadOnly=True인데
        # 디스크는 os.access(W_OK)=True였기 때문이다. 그래서 디스크와 교차 검증하게
        # 고쳤는데, **실제로 써 보지 않고 내린 판단이었다.**
        #
        # 다시 재 보니 그 상태에서 쓰기를 시도하면 정확히 이렇게 실패한다:
        #   (-2147352567, '예외가 발생했습니다.',
        #    (0, 'Microsoft Excel', '파일이 읽기 전용인 경우에는 이 작업을 수행할 수 없습니다.', …))
        # 즉 ReadOnly=True는 처음부터 정확했다. 이 PC의 Excel이 정품 인증이 안 된
        # 무료 버전이라 여는 통합문서가 전부 읽기 전용인 것이다.
        #
        # 교차 검증을 두면 이 경우를 통과시켜, 사용자는 위의 날 COM 예외 덤프를 본다.
        # 파일시스템 신호는 참고용으로만 남긴다.
        flags["workbook_read_only"] = _flag(wb_api, "ReadOnly")
        flags["path_unwritable"] = self._path_is_unwritable(wb_api)
        flags["structure_protected"] = _flag(wb_api, "ProtectStructure")
        flags["marked_final"] = _flag(wb_api, "Final")

        try:
            ws = self._find_sheet(wb, sheet_name) if sheet_name else wb.sheets.active
            ws_api = getattr(ws, "api", None)
            if ws_api is not None:
                flags["sheet_protected"] = _flag(ws_api, "ProtectContents")
                flags["sheet_name"] = str(getattr(ws, "name", sheet_name or ""))
        except Exception:
            pass
        return flags

    def _find_workbook(self, workbook_id_or_name: str):
        candidate = (workbook_id_or_name or "").strip().lower()
        for wb in self._app().books:
            wb_id = self._workbook_id(wb).lower()
            wb_name = str(getattr(wb, "name", "")).strip().lower()
            if candidate in {wb_id, wb_name}:
                return wb
        raise WorkbookNotFoundError(f"통합문서를 찾을 수 없습니다: {workbook_id_or_name}")

    @staticmethod
    def _workbook_id(workbook: Any) -> str:
        fullname = str(getattr(workbook, "fullname", "") or "").strip()
        name = str(getattr(workbook, "name", "") or "").strip()
        return fullname or name

    @staticmethod
    def _active_sheet_name(workbook: Any) -> str:
        try:
            sheet = workbook.sheets.active
            return str(getattr(sheet, "name", "") or "")
        except Exception:
            return ""

    @staticmethod
    def _find_sheet(workbook: Any, sheet_name: str):
        target = (sheet_name or "").strip().lower()
        for sheet in workbook.sheets:
            if str(getattr(sheet, "name", "")).strip().lower() == target:
                return sheet
        raise WorksheetNotFoundError(f"시트를 찾을 수 없습니다: {sheet_name}")

    @staticmethod
    def _normalize_values(raw: Any) -> list[list[Any]]:
        # xlwings range.value 특성:
        # - 단일 셀: scalar
        # - 단일 행/열: 1차원 list
        # - 다중 범위: 2차원 list
        if raw is None:
            return []
        if isinstance(raw, list):
            if not raw:
                return []
            if isinstance(raw[0], list):
                return raw
            return [raw]
        return [[raw]]

    @staticmethod
    def _normalize_address_ref(address: str) -> str:
        """
        Excel 주소 문자열을 A1/A1:C3 형태로 정규화한다.
        예: "$C$3" -> "C3", "'Sheet1'!$B$2:$D$3" -> "B2:D3"
        """
        text = str(address or "").strip()
        if not text:
            return ""
        text = text.lstrip("=")
        if "!" in text:
            text = text.split("!")[-1]
        text = text.replace("$", "").strip()
        # 다중 선택(콤마 구분)인 경우 첫 영역만 사용
        if "," in text:
            text = text.split(",")[0].strip()
        return text or ""

    @staticmethod
    def _sanitize_sheet_name(sheet_name: str) -> str:
        text = str(sheet_name or "").strip()
        if not text:
            raise ExcelLiveError("sheet_name이 비어 있습니다.")
        # Excel 시트명 금지 문자: : \ / ? * [ ]
        text = re.sub(r"[:\\/?*\[\]]", "_", text)
        text = text[:31].strip()
        if not text:
            raise ExcelLiveError("유효한 sheet_name이 필요합니다.")
        return text

    @staticmethod
    def _cell_matches_highlight(cell_value: Any, operator: str, threshold: float, value: Any = None) -> bool:
        text_value = None if value is None or str(value).strip() == "" else str(value).strip()
        op = str(operator or "").strip()
        if text_value is not None and op in {"==", "=", "!=", "<>"}:
            left = "" if cell_value is None else str(cell_value).strip()
            matched_text = left == text_value
            if op in {"!=", "<>"}:
                return not matched_text
            return matched_text
        return ExcelLiveService._matches_condition(cell_value, operator, threshold)

    @staticmethod
    def _matches_condition(value: Any, operator: str, threshold: float) -> bool:
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return False

        op = operator.strip()
        if op == ">":
            return numeric > threshold
        if op == ">=":
            return numeric >= threshold
        if op == "<":
            return numeric < threshold
        if op == "<=":
            return numeric <= threshold
        if op == "==":
            return numeric == threshold
        if op == "!=":
            return numeric != threshold
        raise ValueError(f"지원하지 않는 연산자: {operator}")

    @staticmethod
    def _hex_to_rgb(color: str) -> tuple[int, int, int]:
        value = color.strip().lstrip("#")
        if len(value) != 6:
            raise ValueError(f"색상 형식이 올바르지 않습니다: {color}")
        return (
            int(value[0:2], 16),
            int(value[2:4], 16),
            int(value[4:6], 16),
        )

    @staticmethod
    def _ensure_visual_gridline(cell: Any, rgb: tuple[int, int, int]) -> None:
        """셀 경계선이 비어 있으면 얇은 보더를 적용해 시인성을 높인다."""
        try:
            api_range = getattr(cell, "api", None)
            if api_range is None:
                return
            # Excel COM 상수 (late binding 용 숫자 상수 사용)
            xl_edge_left = 7
            xl_edge_top = 8
            xl_edge_bottom = 9
            xl_edge_right = 10
            xl_continuous = 1
            xl_thin = 2
            xl_line_style_none = -4142
            border_color = ExcelLiveService._rgb_to_excel_color((217, 217, 217))

            for edge in (xl_edge_left, xl_edge_top, xl_edge_bottom, xl_edge_right):
                border = api_range.Borders(edge)
                # 기존 보더가 이미 있으면 유지한다.
                line_style = getattr(border, "LineStyle", xl_line_style_none)
                if line_style not in (None, 0, xl_line_style_none):
                    continue
                border.LineStyle = xl_continuous
                border.Weight = xl_thin
                border.Color = border_color
        except Exception:
            # COM/테마 환경에 따라 보더 설정 실패 가능 — 비치명적
            return

    @staticmethod
    def _rgb_to_excel_color(rgb: tuple[int, int, int]) -> int:
        """(R,G,B) 튜플을 Excel COM Color 정수로 변환한다."""
        r, g, b = rgb
        return int(r) + (int(g) << 8) + (int(b) << 16)

    @classmethod
    def _col_to_idx(cls, col: str) -> int:
        n = 0
        for ch in str(col or "").strip().upper():
            if not ("A" <= ch <= "Z"):
                raise ValueError(f"유효하지 않은 열 식별자: {col}")
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n

    @classmethod
    def _resolve_column_selector(
        cls,
        selector: str | int,
        start_col_idx: int,
        col_count: int,
        header_row: list[Any] | None,
        *,
        strict: bool = False,
    ) -> int:
        """열 지목을 0-기반 색인으로. `strict`면 못 찾았을 때 **-1**을 돌려준다.

        기본값(strict=False)은 못 찾으면 1번 열로 강등한다 — 정렬·필터·피벗이 그 관용에
        기대고 있어 전역으로 바꾸면 지금 통과하는 경로가 한꺼번에 오류가 된다.
        다만 **행을 지우는** dedupe에서는 그 강등이 조용한 오실행이 된다: 머리글에 없는
        '이름'이 A열(날짜)로 강등돼 날짜가 같은 행이 지워지는데 사후조건은
        `removed_rows >= 0`이라 성공으로 보고된다(2026-08-26 감사).
        그래서 파괴 호출부만 strict로 부른다.
        """
        if isinstance(selector, int):
            raw = selector
        else:
            text = str(selector or "").strip()
            headers = [str(h or "").strip() for h in (header_row or [])]
            if not text:
                raw = 1
            elif re.fullmatch(r"\d+", text):
                raw = int(text)
            else:
                raw = 0
                lowered = text.lower()
                for idx, header in enumerate(headers, start=1):
                    if header.lower() == lowered:
                        raw = idx
                        break
                if raw == 0 and headers:
                    # "매출"→Sales 처럼 한국어 개념어로 부른 열을 잇는다.
                    mapped = resolve_header(text, [h for h in headers if h])
                    if mapped:
                        raw = headers.index(mapped) + 1
                if raw == 0 and re.fullmatch(r"[A-Z]{1,3}", text.upper()):
                    # 열 문자 해석은 마지막이다. 'Sales' 같은 영문 머리글을 열 문자로 읽으면
                    # 범위 끝 열로 잘려 엉뚱한 열이 집계된다.
                    abs_idx = cls._col_to_idx(text.upper())
                    raw = abs_idx - start_col_idx + 1
                if raw == 0:
                    if strict:
                        return -1
                    raw = 1
        raw = max(1, min(col_count, int(raw)))
        return raw - 1  # zero-based

    @staticmethod
    def _unresolved_key_columns_error(names: list[Any], header_row: list[Any] | None) -> str:
        """dedupe 기준 열을 못 찾았을 때 사람이 고칠 수 있는 문구로."""
        headers = [str(h or "").strip() for h in (header_row or []) if str(h or "").strip()]
        shown = ", ".join(str(n) for n in names)
        tail = f" 이 시트의 머리글: {', '.join(headers[:8])}" if headers else ""
        return f"중복 기준 열을 찾지 못했습니다: {shown}.{tail}"

    @staticmethod
    def _unresolved_filter_column_error(name: Any, header_row: list[Any] | None) -> str:
        """filter 기준 열을 못 찾았을 때 — 이 액션은 조건에 안 맞는 행을 지운다."""
        headers = [str(h or "").strip() for h in (header_row or []) if str(h or "").strip()]
        tail = f" 이 시트의 머리글: {', '.join(headers[:8])}" if headers else ""
        return f"필터 기준 열을 찾지 못했습니다: {name}.{tail}"

    @staticmethod
    def _sortable_value(value: Any) -> tuple[int, Any]:
        if value is None:
            return (3, "")
        if isinstance(value, (int, float)):
            return (0, float(value))
        text = str(value).strip()
        if text == "":
            return (3, "")
        try:
            return (1, float(text))
        except Exception:
            return (2, text.lower())

    @staticmethod
    def _build_filter_criteria(operator: str, value: Any) -> str:
        op = str(operator or "==").strip()
        if op in {"=", "=="}:
            return str(value)
        if op == "!=":
            return f"<>{value}"
        if op in {">", ">=", "<", "<="}:
            return f"{op}{value}"
        return str(value)

    @staticmethod
    def _matches_generic_condition(cell_value: Any, operator: str, target_value: Any) -> bool:
        op = str(operator or "==").strip()
        if op in {"=", "=="}:
            return str(cell_value) == str(target_value)
        if op == "!=":
            return str(cell_value) != str(target_value)
        left_num = ExcelLiveService._as_float(cell_value)
        right_num = ExcelLiveService._as_float(target_value)
        if left_num is None or right_num is None:
            return False
        if op == ">":
            return left_num > right_num
        if op == ">=":
            return left_num >= right_num
        if op == "<":
            return left_num < right_num
        if op == "<=":
            return left_num <= right_num
        return False

    @staticmethod
    def _build_pivot_rows(
        *,
        rows: list[list[Any]],
        row_idx: int,
        value_idx: int,
        agg: str,
        col_idx: int | None,
        row_header: str,
        value_header: str,
    ) -> list[list[Any]]:
        if col_idx is None:
            groups: dict[str, list[float]] = {}
            for row in rows:
                key = str(row[row_idx] if row_idx < len(row) else "")
                val = ExcelLiveService._as_float(row[value_idx] if value_idx < len(row) else None)
                if key not in groups:
                    groups[key] = []
                if val is not None:
                    groups[key].append(val)
            out: list[list[Any]] = [[row_header, f"{agg}_{value_header}"]]
            for key in sorted(groups.keys()):
                vals = groups[key]
                if agg == "count":
                    agg_val: Any = len(vals)
                elif agg == "avg":
                    agg_val = (sum(vals) / len(vals)) if vals else 0
                else:
                    agg_val = sum(vals)
                out.append([key, agg_val])
            return out

        row_keys: set[str] = set()
        col_keys: set[str] = set()
        groups2d: dict[tuple[str, str], list[float]] = {}
        for row in rows:
            r_key = str(row[row_idx] if row_idx < len(row) else "")
            c_key = str(row[col_idx] if col_idx < len(row) else "")
            val = ExcelLiveService._as_float(row[value_idx] if value_idx < len(row) else None)
            row_keys.add(r_key)
            col_keys.add(c_key)
            group_key = (r_key, c_key)
            if group_key not in groups2d:
                groups2d[group_key] = []
            if val is not None:
                groups2d[group_key].append(val)
        sorted_rows = sorted(row_keys)
        sorted_cols = sorted(col_keys)
        out = [[row_header, *sorted_cols]]
        for r_key in sorted_rows:
            row_out: list[Any] = [r_key]
            for c_key in sorted_cols:
                vals = groups2d.get((r_key, c_key), [])
                if agg == "count":
                    row_out.append(len(vals))
                elif agg == "avg":
                    row_out.append((sum(vals) / len(vals)) if vals else 0)
                else:
                    row_out.append(sum(vals))
            out.append(row_out)
        return out

    @staticmethod
    def _find_or_create_sheet(workbook: Any, sheet_name: str):
        target = str(sheet_name or "").strip()
        for sheet in workbook.sheets:
            if str(getattr(sheet, "name", "") or "").strip().lower() == target.lower():
                return sheet
        return workbook.sheets.add(name=target)

    @staticmethod
    def _chart_type_to_excel(chart_type: str) -> int:
        """
        Excel COM ChartType 상수 매핑.
        - line: 4 (xlLine)
        - bar: 51 (xlColumnClustered)
        - pie: 5 (xlPie)
        """
        normalized = str(chart_type or "line").strip().lower()
        if normalized in {"bar", "column"}:
            return 51
        if normalized in {"pie", "donut"}:
            return 5
        return 4

    @staticmethod
    def _is_empty(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        return False

    @staticmethod
    def _as_float(value: Any) -> float | None:
        try:
            if value is None:
                return None
            if isinstance(value, bool):
                return None
            return float(value)
        except Exception:
            return None

    @staticmethod
    def _parse_datetime_value(value: Any) -> datetime | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        text = str(value).strip()
        if not text:
            return None
        normalized = text.replace("/", "-")
        try:
            return datetime.fromisoformat(normalized)
        except Exception:
            return None

    @classmethod
    def _idx_to_col(cls, idx: int) -> str:
        if idx < 1:
            raise ValueError(f"유효하지 않은 열 인덱스: {idx}")
        result = ""
        n = idx
        while n > 0:
            n, rem = divmod(n - 1, 26)
            result = chr(ord("A") + rem) + result
        return result

    @staticmethod
    def _resolve_target_range(sheet: Any, target_range: str):
        """A:A 같은 전체 열 범위를 used_range 기준으로 축소한다."""
        text = str(target_range or "").strip().upper()
        col_match = re.fullmatch(r"([A-Z]+):([A-Z]+)", text)
        if not col_match:
            return sheet.range(text)

        used = getattr(sheet, "used_range", None)
        if used is None:
            return sheet.range(text)

        used_start_row = int(getattr(used, "row", 1) or 1)
        used_rows_obj = getattr(used, "rows", None)
        used_row_count = int(getattr(used_rows_obj, "count", 1) or 1)
        used_end_row = max(used_start_row, used_start_row + used_row_count - 1)
        left_col, right_col = col_match.groups()
        return sheet.range(f"{left_col}{used_start_row}:{right_col}{used_end_row}")


_excel_live_service: ExcelLiveService | None = None
_excel_live_service_engine: str | None = None

# 실행 중인 Excel 탐지 결과 캐시. COM 왕복이 비싸서 짧게 재사용한다.
_EXCEL_PROBE_TTL_SEC = 5.0
_excel_probe_cache: tuple[float, bool] | None = None


def _excel_app_has_open_workbook() -> bool:
    """실행 중인 Excel에 열린 통합문서가 있는지 본다.

    열려 있다면 그 파일은 OS가 잠근 상태라 openpyxl로는 저장이 불가능하다.
    이 경우 Excel을 직접 제어하는 xlwings만이 편집을 반영할 수 있다.
    """
    global _excel_probe_cache
    now = time.monotonic()
    if _excel_probe_cache is not None and now - _excel_probe_cache[0] < _EXCEL_PROBE_TTL_SEC:
        return _excel_probe_cache[1]

    try:
        xw = importlib.import_module("xlwings")
        # 벤더 데모 파일(.venv의 xlwings quickstart 등)이 열려 있다는 이유로
        # xlwings를 고르면, 뒤의 books.active 폴백이 그 파일에 실행한다(2026-08-04).
        # **사람의 작업물이 열려 있을 때만** xlwings 엔진의 근거가 된다.
        found = any(
            _is_user_workbook_path(getattr(book, "fullname", ""))
            for app in xw.apps
            for book in app.books
        )
    except Exception:
        found = False

    _excel_probe_cache = (now, found)
    return found


def invalidate_excel_engine_cache() -> None:
    """엔진 선택 캐시를 버린다.

    엔진은 "Excel에 열린 통합문서가 있는가"로 골라지고 그 판정에 5초 TTL 캐시가
    붙어 있다. 읽기 전용 브리지가 통합문서를 닫은 직후에는 그 캐시가 거짓이 되므로
    (여전히 xlwings를 고른다) 반드시 버려야 file 엔진으로 넘어간다.
    """
    global _excel_probe_cache, _excel_live_service, _excel_live_service_engine
    _excel_probe_cache = None
    _excel_live_service = None
    _excel_live_service_engine = None


def get_excel_live_service() -> ExcelLiveService:
    """
    Excel Live 서비스 싱글톤 반환.

    환경변수:
    - EXCEL_LIVE_ENGINE=auto (기본): Excel에 열린 문서가 있으면 xlwings, 없으면 file.
      사용자가 엑셀을 띄워 놓고 명령하는 게 정상 사용 흐름이라 이때 file 엔진을 쓰면
      OS 파일 잠금 때문에 저장이 통째로 실패한다. 그래서 실행 환경을 보고 고른다.
    - EXCEL_LIVE_ENGINE=file: openpyxl로 파일을 직접 편집. Excel 앱이 없어도 된다.
    - EXCEL_LIVE_ENGINE=xlwings: 항상 실행 중인 Excel 앱을 제어.

    예전 설정 파일이 쓰던 "pandas"는 "file"과 같은 뜻으로 받는다.
    """
    global _excel_live_service, _excel_live_service_engine
    engine = str(os.getenv("EXCEL_LIVE_ENGINE", "auto") or "auto").strip().lower()
    if engine == "pandas":
        engine = "file"
    if engine not in {"xlwings", "file", "auto"}:
        engine = "auto"
    if engine == "auto":
        engine = "xlwings" if _excel_app_has_open_workbook() else "file"

    if _excel_live_service is None or _excel_live_service_engine != engine:
        if engine == "file":
            from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

            _excel_live_service = FileExcelLiveService()
        else:
            _excel_live_service = ExcelLiveService()
        _excel_live_service_engine = engine
    return _excel_live_service

