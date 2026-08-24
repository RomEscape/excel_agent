"""
OpenPyXL 기반 파일 편집 Excel 서비스.

정렬·중복제거·필터·집계는 순수 파이썬으로 처리한다. 표 하나가 수백 행 규모라
데이터프레임을 얹을 이유가 없고, 무엇보다 셀 하나하나의 원본 값·수식·서식을
그대로 들고 다녀야 해서 중간에 다른 표현으로 갈아타면 잃는 것이 더 많다.
"""

from __future__ import annotations

import os
import re
import shutil
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
)
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.series import SeriesLabel
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableStyleInfo

from office_claw_sidecar.sandbox import WORKSPACE_ROOT
from office_claw_sidecar.services import excel_formula_cache as formula_cache
from office_claw_sidecar.services.excel_formula_eval import (
    FormulaError,
    WorkbookEvaluator,
)
from office_claw_sidecar.services.excel_header_lexicon import resolve_header
from office_claw_sidecar.services.excel_live_service import (
    _ALIGN_WORDS,
    _SCAN_EXCLUDED_DIRS,
    AmbiguousWorkbookError,
    ExcelConnectionError,
    ExcelLiveError,
    ExcelLiveService,
    WorkbookNotFoundError,
    WorksheetNotFoundError,
)

# 워크스페이스를 훑을 때 건너뛸 디렉터리.
# 가상환경의 샘플 통합문서(xlwings quickstart 등)나 우리가 만든 백업본이
# "가장 최근 수정된 파일"로 잡혀 편집 대상이 되는 사고를 막는다.
# 제외 목록은 excel_live_service **한 곳**이다 — xlwings 경로와 갈라지면
# 한쪽만 가드되는 사고(2026-08-04)가 재현된다.


def _as_number(value: Any) -> float | None:
    """셀 값을 비교 가능한 숫자로 바꾼다. 숫자로 볼 수 없으면 None.

    "1,200"이나 "85%"처럼 사람이 보기 좋게 적어 둔 값도 정렬·집계에서는 숫자여야 한다.
    날짜는 타임스탬프로 바꿔 문자열 정렬이 아니라 실제 시간 순서로 놓는다.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, datetime):
        return value.timestamp()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).timestamp()
    if isinstance(value, time):
        return value.hour * 3600.0 + value.minute * 60.0 + value.second
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1].strip()
    try:
        number = float(text)
    except ValueError:
        return None
    return number / 100.0 if percent else number


def _text_key(value: Any) -> str:
    return "" if value is None else str(value).strip().lower()


def _group_key(row: list[Any], columns: list[int]) -> tuple:
    """중복 판정·그룹핑에 쓰는 키. 빈 셀은 모두 같은 값으로 본다."""
    return tuple(_text_key(row[i]) if i < len(row) else "" for i in columns)


def _sorted_group_keys(keys: list[tuple]) -> list[tuple]:
    """그룹 머리글을 사람이 기대하는 순서로 놓는다 — 숫자 먼저, 그다음 사전순."""
    numeric = [k for k in keys if all(_as_number(part) is not None for part in k)]
    others = sorted(set(keys) - set(numeric), key=lambda k: tuple(str(p) for p in k))
    numeric.sort(key=lambda k: tuple(_as_number(p) or 0.0 for p in k))
    return numeric + others


# 서식 스냅샷 상한 — 검증용이라 표 하나를 덮을 정도면 충분하다.
def _hex_or_none(value: Any) -> str | None:
    """openpyxl의 rgb는 테마 색일 때 오류 문자열이 온다 — 16진 코드만 통과시킨다."""
    text = str(value or "").strip()
    return text.upper() if re.fullmatch(r"[0-9A-Fa-f]{6,8}", text) else None


_FORMAT_SNAPSHOT_MAX_ROWS = 200
_FORMAT_SNAPSHOT_MAX_COLS = 60


#: 수식 안의 시트 참조. `지역성과!B2` · `'지역 성과'!B:B` 두 형태를 모두 잡는다.
_FORMULA_SHEET_REF = re.compile(r"(?:'([^']+)'|([A-Za-z0-9가-힣_.]+))\s*!")


def _resolve_formula_sheet_refs(formula: str, sheet_names: list[str]) -> tuple[str, list[str]]:
    """수식의 시트 참조를 실제 이름으로 맞춘다. 못 맞춘 이름들을 함께 돌려준다.

    돌려주는 것: (고친 수식, 못 찾은 이름 목록)
    앞부분만 맞는 후보가 **정확히 하나**일 때만 고친다 — 둘 이상이면 짐작이 된다.
    """
    text = str(formula or "")
    if not text.startswith("=") or not sheet_names:
        return text, []
    by_fold = {name.strip().casefold(): name for name in sheet_names}
    missing: list[str] = []

    def _fix(match: re.Match[str]) -> str:
        raw = match.group(1) or match.group(2) or ""
        name = raw.strip()
        if not name:
            return match.group(0)
        actual = by_fold.get(name.casefold())
        if actual is None:
            starts = [n for n in sheet_names if n.casefold().startswith(name.casefold())]
            actual = starts[0] if len(starts) == 1 else None
        if actual is None:
            missing.append(name)
            return match.group(0)
        quoted = f"'{actual}'" if re.search(r"[^A-Za-z0-9가-힣_.]", actual) else actual
        return f"{quoted}!"

    return _FORMULA_SHEET_REF.sub(_fix, text), missing

class FileExcelLiveService(ExcelLiveService):
    """파일 기반(openpyxl) Excel 편집 서비스. Excel 앱이 없어도 동작한다."""

    engine = "file"
    # '선택'은 사용 범위 폴백일 뿐 실제 드래그가 아니다 — 선택 우선 해석이
    # 붙여넣기 문맥을 덮으면 안 된다(2026-08-18 실측).
    has_real_selection = False

    def __init__(self, workspace_root: Path | None = None) -> None:
        super().__init__(xw_module=None)
        self._workspace_root = Path(workspace_root or WORKSPACE_ROOT).resolve()
        self._workspace_root.mkdir(parents=True, exist_ok=True)
        self._selected_sheet_by_workbook: dict[str, str] = {}
        # 우리가 마지막으로 저장한 파일의 지문. 이 상태면 수식 캐시가 비어 있음을 알고 다시 열지 않는다.
        self._saved_stamp: dict[str, tuple[int, int] | None] = {}

    # ---- 공통 유틸 ----

    def _candidate_roots(self) -> list[Path]:
        roots: list[Path] = []
        for base in [self._workspace_root, Path.cwd()]:
            try:
                resolved = Path(base).expanduser().resolve()
            except Exception:
                continue
            if resolved not in roots:
                roots.append(resolved)
        return roots

    @staticmethod
    def _is_excel_file(path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}

    @staticmethod
    def _is_scannable(path: Path, root: Path) -> bool:
        """사용자 작업물로 볼 수 있는 경로인지 판정한다."""
        try:
            parts = path.relative_to(root).parts[:-1]
        except ValueError:
            return False
        return not any(part in _SCAN_EXCLUDED_DIRS or part.startswith(".") for part in parts)

    def _list_workspace_workbooks(self, limit: int = 200) -> list[Path]:
        rows: dict[str, Path] = {}
        for root in self._candidate_roots():
            if not root.exists():
                continue
            try:
                for fp in root.rglob("*.xls*"):
                    if not fp.is_file():
                        continue
                    if fp.name.startswith("~$"):
                        continue
                    if not self._is_excel_file(fp):
                        continue
                    if not self._is_scannable(fp, root):
                        continue
                    rows[str(fp.resolve())] = fp.resolve()
            except Exception:
                continue
        files = list(rows.values())
        files.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0.0, reverse=True)
        return files[: max(1, min(1000, int(limit)))]

    def get_write_protection(
        self, workbook_id: str | None, sheet_name: str | None = None
    ) -> dict[str, Any]:
        """파일 엔진의 보호 상태(F-08).

        openpyxl은 시트 보호를 **강제하지 않는다** — 보호된 시트에 값을 쓰고 저장까지
        성공한다(2026-08-16 실측). 그래서 이 조회가 file 엔진의 유일한 방어선이다.

        한계: `os.access(W_OK)`는 Windows의 읽기 전용 **속성**만 반영한다. NTFS ACL이나
        네트워크 공유 권한으로 인한 쓰기 불가는 저장 시점 PermissionError로만 드러난다.
        """
        flags: dict[str, Any] = {"sheet_name": sheet_name or ""}
        try:
            path = self._resolve_workbook_path(workbook_id)
        except Exception:
            return flags
        try:
            flags["workbook_read_only"] = path.exists() and not os.access(path, os.W_OK)
        except Exception:
            pass
        try:
            wb = self._load_wb(path)
            try:
                security = getattr(wb, "security", None)
                flags["structure_protected"] = bool(
                    getattr(security, "lockStructure", False)
                )
                ws = wb[sheet_name] if sheet_name else wb.active
                protection = getattr(ws, "protection", None)
                flags["sheet_protected"] = bool(getattr(protection, "sheet", False))
                flags["sheet_name"] = str(getattr(ws, "title", sheet_name or ""))
            finally:
                wb.close()
        except Exception:
            pass
        return flags

    def _resolve_workbook_path(self, workbook_id_or_name: str | None) -> Path:
        raw = str(workbook_id_or_name or "").strip()
        if raw:
            candidate = Path(raw).expanduser()
            if candidate.is_absolute() and candidate.exists() and candidate.is_file() and self._is_excel_file(candidate):
                return candidate.resolve()
            for root in self._candidate_roots():
                joined = (root / candidate).resolve()
                if joined.exists() and joined.is_file() and self._is_excel_file(joined):
                    return joined
            lowered = raw.lower()
            for fp in self._list_workspace_workbooks(limit=500):
                if fp.name.lower() == lowered:
                    return fp
        selected = str(self._selected_workbook_id or "").strip()
        if selected:
            selected_path = Path(selected).expanduser()
            if selected_path.exists() and selected_path.is_file():
                return selected_path.resolve()
        # 대상이 안 정해졌을 때 "가장 최근 수정된 파일"을 말없이 고르면, 사용자가
        # 열어 보지도 않은 통합문서가 편집된다. 실제로 백업 파일이 최신이 되면서
        # 대상이 매 명령마다 흘러가는 일이 있었다. 후보가 하나뿐일 때만 자동 선택한다.
        files = self._list_workspace_workbooks(limit=10)
        if len(files) == 1:
            return files[0]
        if files:
            raise AmbiguousWorkbookError(
                "어떤 통합문서에 적용할까요? 워크스페이스에 여러 파일이 있어 대상을 특정하지 못했습니다.",
                candidates=[fp.name for fp in files],
            )
        raise WorkbookNotFoundError(
            "대상 통합문서를 찾지 못했습니다. workbook_id를 지정하거나 워크스페이스에 .xlsx 파일을 준비해 주세요."
        )

    @staticmethod
    def _load_wb(path: Path, *, data_only: bool = False):
        try:
            keep_vba = path.suffix.lower() == ".xlsm"
            return load_workbook(filename=str(path), data_only=data_only, keep_vba=keep_vba)
        except Exception as exc:
            raise ExcelLiveError(f"통합문서를 열 수 없습니다: {path}") from exc

    def _make_evaluator(self, path: Path, wb_formula: Any, wb_data: Any, default_sheet: str) -> Any:
        """수식 계산기를 워크북에 물린다.

        값을 찾는 순서는 파일 캐시 → 우리 스냅샷 → 직접 계산이다. 앞의 둘로 해결되면
        계산기는 손도 대지 않으므로, Excel이 저장한 파일에서는 기존 경로 그대로 동작한다.
        """

        def raw(sheet: str, row: int, col: int) -> Any:
            ws = wb_formula[sheet] if sheet in wb_formula.sheetnames else None
            return None if ws is None else ws.cell(row=row, column=col).value

        def cached(sheet: str, row: int, col: int) -> Any:
            if sheet in wb_data.sheetnames:
                value = wb_data[sheet].cell(row=row, column=col).value
                if value is not None:
                    return value
            return formula_cache.lookup(path, sheet, row, col)

        def bounds(sheet: str) -> tuple[int, int]:
            ws = wb_formula[sheet] if sheet in wb_formula.sheetnames else None
            return (0, 0) if ws is None else (ws.max_row, ws.max_column)

        return WorkbookEvaluator(
            raw,
            cached_lookup=cached,
            default_sheet=default_sheet,
            sheet_bounds=bounds,
        )

    @staticmethod
    def _evaluated(evaluator: Any, sheet: str, row: int, col: int, fallback: Any) -> Any:
        """계산에 실패하면 원본(수식 문자열)을 그대로 돌려준다.

        호출자는 문자열이 온 걸 보고 "이 셀은 값을 모른다"고 판단할 수 있다.
        조용히 0으로 바꾸면 틀린 집계·조건 판정으로 이어진다.
        """
        try:
            value = evaluator.value(sheet, row, col)
        except FormulaError:
            return fallback
        return fallback if value is None else value

    def _file_stamp(self, path: Path) -> tuple[int, int] | None:
        try:
            stat = path.stat()
        except OSError:
            return None
        return (stat.st_mtime_ns, stat.st_size)

    def _capture_formula_values(self, path: Path) -> None:
        """저장 직전, 디스크에 남아 있는 수식 계산값을 스냅샷으로 떠 둔다.

        openpyxl로 저장하면 그 캐시가 사라져 다음 턴의 집계가 "=I2*J2" 문자열을 보게 된다.
        우리가 마지막으로 저장한 파일 그대로면 캐시는 이미 비어 있으므로 다시 열지 않는다.
        """
        stamp = self._file_stamp(path)
        if stamp is None or self._saved_stamp.get(str(path)) == stamp:
            return
        try:
            wb_data = self._load_wb(path, data_only=True)
        except ExcelLiveError:
            return
        wb_formula = None
        try:
            wb_formula = self._load_wb(path)
            formula_cache.capture(path, wb_data, wb_formula)
        except ExcelLiveError:
            return
        finally:
            wb_data.close()
            if wb_formula is not None:
                wb_formula.close()

    def _save_wb(
        self,
        wb: Any,
        path: Path,
        *,
        value_changed_sheet: str | None = None,
        changed_rows: set[int] | list[int] | None = None,
    ) -> None:
        """편집 저장 공통 경로. 수식 계산값을 보존하고, 값이 바뀐 범위만 스냅샷에서 뺀다.

        `value_changed_sheet`를 주면 그 시트의 해당 행(없으면 시트 전체)과 다른 시트 스냅샷을
        버린다. 다른 시트의 합계 수식이 이 시트를 참조할 수 있어서다. 오래된 숫자로 답하느니
        "값을 모른다"고 말하는 편이 낫다.
        """
        self._capture_formula_values(path)
        self._mark_full_recalc(wb)
        self._sync_table_headers(wb)
        try:
            wb.save(str(path))
        except PermissionError as exc:
            # Windows는 Excel이 연 파일을 잠근다. 이 엔진은 파일을 직접 덮어쓰므로
            # 사용자가 화면에서 그 파일을 열어 둔 동안에는 어떤 편집도 저장되지 않는다.
            # 원인을 그대로 알려야 사용자가 파일을 닫거나 엔진을 바꿀 수 있다.
            raise ExcelConnectionError(
                f"'{path.name}' 파일이 Excel에서 열려 있어 저장할 수 없습니다. "
                "Excel에서 해당 파일을 닫고 다시 시도하거나, "
                "열린 Excel을 직접 제어하는 xlwings 엔진으로 전환해 주세요."
            ) from exc
        self._saved_stamp[str(path)] = self._file_stamp(path)
        if value_changed_sheet:
            if changed_rows:
                formula_cache.invalidate_rows(path, value_changed_sheet, changed_rows)
            else:
                formula_cache.invalidate_sheet(path, value_changed_sheet)
            formula_cache.invalidate_other_sheets(path, value_changed_sheet)

    @staticmethod
    def _sync_table_headers(wb: Any) -> None:
        """표(ListObject) 정의의 열 이름을 실제 머리글 셀과 맞춘다.

        머리글 셀을 고쳐도 표 정의는 옛 이름을 들고 있다. Excel은 둘이 어긋나면
        파일 자체를 열지 못한다("Workbooks 클래스 중 Open 메서드에 오류"). openpyxl로는
        계속 읽히기 때문에 PDF 내보내기처럼 Excel을 거칠 때에야 드러난다.
        """
        for ws in wb.worksheets:
            for table in getattr(ws, "tables", {}).values():
                try:
                    min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
                except Exception:
                    continue
                if not getattr(table, "headerRowCount", 1):
                    continue
                used: set[str] = set()
                for column in table.tableColumns:
                    index = min_col + table.tableColumns.index(column)
                    if index > max_col:
                        break
                    value = ws.cell(row=min_row, column=index).value
                    name = str(value).strip() if value is not None else ""
                    # 빈 머리글이나 이름 충돌은 표 정의를 깨뜨린다. 기존 이름을 지킨다.
                    if not name or name in used:
                        used.add(column.name)
                        continue
                    column.name = name
                    used.add(name)

    @staticmethod
    def _mark_full_recalc(wb: Any) -> None:
        """파일을 열 때 Excel이 수식을 전부 다시 계산하도록 표시한다.

        openpyxl로 만들었거나 calcPr이 없는 워크북은 `wb.calculation`이 None이라
        속성만 대입하면 AttributeError가 난다. 없으면 만들어서 붙인다.
        """
        if wb.calculation is None:
            wb.calculation = CalcProperties()
        wb.calculation.fullCalcOnLoad = True

    @staticmethod
    def _find_label_column(ws, value_col: int, min_row: int, max_row: int) -> int | None:
        """값 열 왼쪽에서 카테고리 라벨로 쓸 열을 찾는다.

        같은 행 구간에 글자 값이 절반 이상인 가장 가까운 열이다. 숫자 열(다른
        지표)은 건너뛴다 — "D2:D5 도넛"의 라벨은 C(수량)가 아니라 A(센터명)다.
        """
        total = max_row - min_row + 1
        if total <= 0:
            return None
        for col in range(value_col - 1, 0, -1):
            cells = [ws.cell(row=r, column=col).value for r in range(min_row, max_row + 1)]
            texts = sum(1 for v in cells if isinstance(v, str) and v.strip())
            if texts * 2 >= total:
                return col
            # 비어 있는 열이면 더 왼쪽을 본다. 숫자 열도 지표일 뿐이니 계속 간다.
        return None

    @staticmethod
    def _sanitize_sheet_name(sheet_name: str) -> str:
        text = str(sheet_name or "").strip()
        if not text:
            raise ExcelLiveError("sheet_name이 비어 있습니다.")
        text = re.sub(r"[:\\/?*\[\]]", "_", text)
        text = text[:31].strip()
        if not text:
            raise ExcelLiveError("유효한 sheet_name이 필요합니다.")
        return text

    @staticmethod
    def _sheet_or_raise(wb: Any, sheet_name: str):
        name = str(sheet_name or "").strip()
        if name in wb.sheetnames:
            return wb[name]
        for cand in wb.sheetnames:
            if cand.lower() == name.lower():
                return wb[cand]
        raise WorksheetNotFoundError(f"시트를 찾을 수 없습니다: {sheet_name}")

    def _used_bounds(self, ws: Any) -> tuple[int, int]:
        """실제 값이 들어 있는 마지막 행/열을 반환한다.

        openpyxl의 max_row/max_column은 서식만 적용된 셀도 포함해 부풀려지므로,
        값 기준으로 다시 좁혀야 범위 미지정 명령이 빈 영역까지 건드리지 않는다.
        """
        raw_row = max(1, int(getattr(ws, "max_row", 1) or 1))
        raw_col = max(1, int(getattr(ws, "max_column", 1) or 1))
        last_row = 0
        last_col = 0
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=raw_row, min_col=1, max_col=raw_col, values_only=True),
            start=1,
        ):
            for c_idx, value in enumerate(row, start=1):
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                last_row = r_idx
                if c_idx > last_col:
                    last_col = c_idx
        if last_row == 0 or last_col == 0:
            return 1, 1
        return last_row, last_col

    #: 실행 시점에야 뜻이 정해지는 상징 범위. 여기서 못 풀면 아래 except가 'A1'로
    #: 떨어뜨려 **범위를 받는 모든 연산이 한 칸짜리 no-op**이 된다(2026-08-20 실측:
    #: highlight_by_condition이 scanned_cells=1로 0건 칠하고 성공을 보고했다).
    _SYMBOLIC_WHOLE_SHEET = frozenset({"__USED_RANGE__", "__ACTIVE_SELECTION__", "__TABLE_REGION__"})

    def _range_bounds(self, ws: Any, range_ref: str) -> tuple[int, int, int, int]:
        text = str(range_ref or "A1").strip().upper()
        if text in self._SYMBOLIC_WHOLE_SHEET:
            used_row, used_col = self._used_bounds(ws)
            return (1, 1, max(used_row, 1), max(used_col, 1))
        if text == "__ACTIVE_CELL__":
            return (1, 1, 1, 1)
        col_match = re.fullmatch(r"([A-Z]+):([A-Z]+)", text)
        if col_match:
            left, right = col_match.groups()
            start_col = self._col_to_idx(left)
            end_col = self._col_to_idx(right)
            max_row, _ = self._used_bounds(ws)
            return (1, min(start_col, end_col), max_row, max(start_col, end_col))
        try:
            min_col, min_row, max_col, max_row = range_boundaries(text)
        except Exception:
            min_col, min_row, max_col, max_row = range_boundaries("A1")
        # "1:1" 같은 행 전체 범위는 열 번호가 None이다. 사용 범위의 열로 채운다.
        if min_col is None or max_col is None or min_row is None or max_row is None:
            used_row, used_col = self._used_bounds(ws)
            min_col = 1 if min_col is None else min_col
            max_col = used_col if max_col is None else max_col
            min_row = 1 if min_row is None else min_row
            max_row = used_row if max_row is None else max_row
        return (min_row, min_col, max_row, max_col)

    @classmethod
    def _address_from_bounds(cls, bounds: tuple[int, int, int, int]) -> str:
        min_row, min_col, max_row, max_col = bounds
        left = f"{cls._idx_to_col(min_col)}{min_row}"
        right = f"{cls._idx_to_col(max_col)}{max_row}"
        return left if left == right else f"{left}:{right}"

    @staticmethod
    def _to_argb(color: str) -> str:
        hex6 = str(color or "#000000").strip().lstrip("#")
        if len(hex6) != 6:
            hex6 = "000000"
        return f"FF{hex6.upper()}"

    # ---- 기본 상태/목록 ----

    def is_available(self) -> bool:
        return True

    def list_workbooks(self) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for fp in self._list_workspace_workbooks(limit=200):
            active_sheet = self._selected_sheet_by_workbook.get(str(fp), "")
            if not active_sheet:
                wb = self._load_wb(fp)
                try:
                    active_sheet = wb.active.title if wb.sheetnames else ""
                finally:
                    wb.close()
            rows.append(
                {
                    "workbook_id": str(fp),
                    "name": fp.name,
                    "full_path": str(fp),
                    "active_sheet": active_sheet,
                }
            )
        return rows

    def select_workbook(self, workbook_id_or_name: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id_or_name)
        self._selected_workbook_id = str(path)
        wb = self._load_wb(path)
        try:
            self._selected_sheet_by_workbook[str(path)] = wb.active.title if wb.sheetnames else ""
        finally:
            wb.close()
        return {"selected": True, "workbook_id": str(path)}

    def list_sheets(self, workbook_id: str | None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            sheets = list(wb.sheetnames)
            active = self._selected_sheet_by_workbook.get(str(path), wb.active.title if sheets else "")
            if active not in sheets and sheets:
                active = sheets[0]
            self._selected_sheet_by_workbook[str(path)] = active
            return {
                "sheets": sheets,
                "count": len(sheets),
                "active_sheet": active,
            }
        finally:
            wb.close()

    def select_sheet(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            self._selected_sheet_by_workbook[str(path)] = ws.title
            wb.active = wb.sheetnames.index(ws.title)
            self._save_wb(wb, path)
            return {
                "selected": True,
                "sheet_name": ws.title,
                "active_sheet": ws.title,
            }
        finally:
            wb.close()

    def create_sheet(self, workbook_id: str | None, sheet_name: str, make_active: bool = True) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        target_name = self._sanitize_sheet_name(sheet_name)
        wb = self._load_wb(path)
        try:
            created = target_name not in wb.sheetnames
            ws = wb.create_sheet(target_name) if created else wb[target_name]
            if make_active:
                wb.active = wb.sheetnames.index(ws.title)
                self._selected_sheet_by_workbook[str(path)] = ws.title
            self._save_wb(wb, path)
            return {
                "created": created,
                "sheet_name": ws.title,
                "active_sheet": self._selected_sheet_by_workbook.get(str(path), ws.title),
            }
        finally:
            wb.close()

    def rename_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
        new_name: str,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        target_name = self._sanitize_sheet_name(new_name)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            old_name = ws.title
            if target_name != old_name and target_name in wb.sheetnames:
                raise ExcelLiveError(f"이미 '{target_name}' 시트가 있습니다.")
            ws.title = target_name
            if self._selected_sheet_by_workbook.get(str(path)) == old_name:
                self._selected_sheet_by_workbook[str(path)] = target_name
            self._save_wb(wb, path)
            return {
                "renamed": True,
                "old_name": old_name,
                "sheet_name": ws.title,
                "sheets": list(wb.sheetnames),
                "active_sheet": self._selected_sheet_by_workbook.get(str(path), ws.title),
            }
        finally:
            wb.close()

    def delete_sheet(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            if len(wb.sheetnames) <= 1:
                raise ExcelLiveError("마지막 시트는 삭제할 수 없습니다.")
            ws = self._sheet_or_raise(wb, sheet_name)
            deleted_name = ws.title
            wb.remove(ws)
            remaining = list(wb.sheetnames)
            selected = self._selected_sheet_by_workbook.get(str(path))
            if selected == deleted_name:
                self._selected_sheet_by_workbook[str(path)] = remaining[0]
                wb.active = 0
            self._save_wb(wb, path)
            return {
                "deleted": True,
                "sheet_name": deleted_name,
                "sheets": remaining,
                "active_sheet": self._selected_sheet_by_workbook.get(str(path), remaining[0]),
            }
        finally:
            wb.close()

    # ---- 범위 I/O ----

    def read_range(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, range_ref)
            min_row, min_col, max_row, max_col = bounds
            values: list[list[Any]] = []
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                values.append([cell.value for cell in row])
            if not values:
                values = [[None]]
            row_count = len(values)
            col_count = len(values[0]) if values else 0
            return {
                "values": values,
                "address": self._address_from_bounds(bounds),
                "row_count": row_count,
                "col_count": col_count,
            }
        finally:
            wb.close()

    def read_computed_range(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, Any]:
        """집계·정렬·필터용으로 수식 대신 계산된 값을 읽는다.

        실무 파일의 매출·이익 열은 대부분 수식이라 그대로 읽으면 "=I2*J2" 문자열이 들어온다.
        값을 찾는 순서는 파일 캐시 → 우리 스냅샷 → 직접 계산이다. 셋 다 실패하면
        `unresolved_formulas`에 좌표를 담아 호출자가 조용히 집계하지 않도록 알린다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb_data = self._load_wb(path, data_only=True)
        wb_formula = self._load_wb(path, data_only=False)
        try:
            ws_formula = self._sheet_or_raise(wb_formula, sheet_name)
            evaluator = self._make_evaluator(path, wb_formula, wb_data, ws_formula.title)
            bounds = self._range_bounds(ws_formula, range_ref)
            min_row, min_col, max_row, max_col = bounds
            values: list[list[Any]] = []
            unresolved: list[tuple[int, int]] = []
            for r in range(min_row, max_row + 1):
                row_values: list[Any] = []
                for c in range(min_col, max_col + 1):
                    raw = ws_formula.cell(row=r, column=c).value
                    value = self._evaluated(evaluator, ws_formula.title, r, c, raw)
                    if isinstance(value, str) and value.startswith("="):
                        unresolved.append((r - min_row, c - min_col))
                    row_values.append(value)
                values.append(row_values)
            if not values:
                values = [[None]]
            return {
                "values": values,
                "address": self._address_from_bounds(bounds),
                "row_count": len(values),
                "col_count": len(values[0]) if values else 0,
                # 범위 기준 상대 좌표. 계산값을 끝내 못 구한 수식 셀.
                "unresolved_formulas": unresolved,
            }
        finally:
            wb_data.close()
            wb_formula.close()

    def count_error_cells(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, list[str]]:
        """부모(COM)와 같은 계약의 파일 엔진 판 — Excel이 마지막 저장 때 남긴 캐시만 본다.

        openpyxl은 수식을 계산하지 않으므로 **우리가 방금 쓴** 수식의 오류는 여기서
        안 보인다(빈 결과 = "오류 없음"이 아니라 "모름"이고, 검증기는 못 본 것을
        실패로 단정하지 않는다). Excel이 계산해 저장한 파일이면 오류가 문자열
        캐시('#NAME?' 등)로 남아 있어 잡힌다.
        """
        from openpyxl.utils import get_column_letter

        error_kinds = set(self._CELL_ERROR_CODES.values())
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path, data_only=True)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            min_row, min_col, max_row, max_col = self._range_bounds(ws, range_ref)
            found: dict[str, list[str]] = {}
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    value = ws.cell(row=r, column=c).value
                    if isinstance(value, str) and value in error_kinds:
                        found.setdefault(value, []).append(f"{get_column_letter(c)}{r}")
            return found
        finally:
            wb.close()

    def describe_sheet_layout(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        """시트의 수식·서식·병합·표 블록을 요약한다.

        다이제스트가 값만 보여 주면 모델은 자기가 방금 칠한 색도, 어느 열이 수식인지도
        모른다. 여러 턴에 걸쳐 다듬으려면 매 턴 파일에서 다시 읽어야 한다.
        여는 일만 여기서 하고 요약은 excel_sheet_layout이 맡는다.
        """
        from office_claw_sidecar.services.excel_sheet_layout import describe_worksheet

        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            return describe_worksheet(ws)
        finally:
            wb.close()

    def get_range_snapshot(self, workbook_id: str | None, sheet_name: str | None, range_ref: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            target_sheet = sheet_name or self._selected_sheet_by_workbook.get(str(path)) or wb.active.title
            data = self.read_range(str(path), target_sheet, range_ref)
            values = data.get("values", [])
            filled = 0
            for row in values:
                for v in row:
                    if v is None:
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    filled += 1
            return {
                "address": data.get("address", range_ref),
                "row_count": int(data.get("row_count", 0) or 0),
                "col_count": int(data.get("col_count", 0) or 0),
                "filled_cells": filled,
            }
        finally:
            wb.close()

    def get_format_snapshot(
        self, workbook_id: str | None, sheet_name: str | None, range_ref: str
    ) -> dict[str, Any]:
        """범위의 **서식** 스냅샷(검증용). 값이 아니라 "요청한 효과가 남았는가"를 보려고 쓴다.

        2026-08-19: 조용한 오실행의 큰 덩어리가 서식이었는데(표시 형식 그대로 · 배경만 칠하고 굵게는 안 됨 ·
        병합 안 됨 · 틀 고정 안 됨) 읽을 방법이 없어 사후조건을 못 걸었다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            target = sheet_name or self._selected_sheet_by_workbook.get(str(path)) or wb.active.title
            ws = self._sheet_or_raise(wb, target)
            min_row, min_col, max_row, max_col = self._range_bounds(ws, range_ref)
            max_row = min(max_row, min_row + _FORMAT_SNAPSHOT_MAX_ROWS - 1)
            max_col = min(max_col, min_col + _FORMAT_SNAPSHOT_MAX_COLS - 1)
            number_formats: list[list[str]] = []
            fills: list[list[str | None]] = []
            bold: list[list[bool]] = []
            font_colors: list[list[str | None]] = []
            borders: list[list[bool]] = []
            for r in range(min_row, max_row + 1):
                nf_row, fill_row, bold_row, color_row, border_row = [], [], [], [], []
                for c in range(min_col, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    nf_row.append(str(cell.number_format or "General"))
                    fill = getattr(cell, "fill", None)
                    rgb = _hex_or_none(getattr(getattr(fill, "fgColor", None), "rgb", None))
                    patt = str(getattr(fill, "patternType", "") or "")
                    fill_row.append(rgb if (patt and rgb and rgb != "00000000") else None)
                    font = getattr(cell, "font", None)
                    bold_row.append(bool(getattr(font, "bold", False)))
                    # openpyxl은 테마 색일 때 rgb 자리에 **오류 문자열**을 돌려준다. 16진 코드만 받는다.
                    frgb = getattr(getattr(font, "color", None), "rgb", None)
                    color_row.append(_hex_or_none(frgb))
                    border = getattr(cell, "border", None)
                    border_row.append(
                        any(
                            getattr(getattr(border, side, None), "style", None)
                            for side in ("left", "right", "top", "bottom")
                        )
                    )
                number_formats.append(nf_row)
                fills.append(fill_row)
                bold.append(bold_row)
                font_colors.append(color_row)
                borders.append(border_row)
            merged = [str(rng) for rng in getattr(ws, "merged_cells", []).ranges] if hasattr(ws, "merged_cells") else []
            return {
                "address": range_ref,
                "sheet_name": str(ws.title),
                "number_formats": number_formats,
                "fills": fills,
                "bold": bold,
                "font_colors": font_colors,
                "borders": borders,
                "merged": merged,
                "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else "",
                "chart_count": len(getattr(ws, "_charts", []) or []),
            }
        finally:
            wb.close()

    def write_range(self, workbook_id: str | None, sheet_name: str, start_cell: str, values_2d: list[list[Any]]) -> dict[str, Any]:
        if not values_2d:
            return {"written_cells": 0, "address": str(start_cell)}
        rows = len(values_2d)
        cols = max(len(r) for r in values_2d)
        normalized = [row + [None] * (cols - len(row)) for row in values_2d]
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            try:
                ws = self._sheet_or_raise(wb, sheet_name)
            except WorksheetNotFoundError:
                ws = wb.create_sheet(self._sanitize_sheet_name(sheet_name))
            min_row, min_col, _, _ = self._range_bounds(ws, start_cell)
            for r_idx, row in enumerate(normalized):
                for c_idx, value in enumerate(row):
                    # ws.cell(value=None)은 기존 값을 남긴다. None을 쓴다는 것은
                    # 그 칸을 비우라는 뜻이므로 속성에 직접 대입해야 한다.
                    ws.cell(row=min_row + r_idx, column=min_col + c_idx).value = value
            max_row = min_row + rows - 1
            max_col = min_col + cols - 1
            self._save_wb(
                wb,
                path,
                value_changed_sheet=ws.title,
                changed_rows=set(range(min_row, max_row + 1)),
            )
            return {
                "written_cells": rows * cols,
                "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            }
        finally:
            wb.close()

    def delete_charts(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        """시트의 차트를 전부 지운다.

        2026-08-18 GUI 실측: "차트 같은 거 다 지워줘"에 삭제 액션이 없어
        차트 **생성** 슬롯("차트 종류를 선택해 주세요")으로 샜다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            deleted = len(ws._charts or [])
            ws._charts = []
            self._save_wb(wb, path)
            # 지울 차트가 없었으면 그 사실을 보고한다 — "완료"만 나가면 무동작이다.
            return {"deleted": deleted, "no_change": deleted == 0, "sheet": ws.title}
        finally:
            wb.close()

    def clear_range(self, workbook_id: str | None, sheet_name: str, target_range: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            cleared = 0
            # 원래 값이 있던 칸 수. 0이면 "완료"가 사용자 눈에는 무동작이다 —
            # 응답이 그 사실을 말할 수 있게 따로 센다(2026-08-17 실측: 서식만 있는
            # 범위를 비우고 "완료"가 나가 "아무것도 안 됐다"는 항의를 받았다).
            emptied = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if cell.value is not None:
                        emptied += 1
                    cell.value = None
                    cleared += 1
            self._save_wb(
                wb,
                path,
                value_changed_sheet=ws.title,
                changed_rows=set(range(min_row, max_row + 1)),
            )
            return {
                "cleared_cells": cleared,
                "emptied_values": emptied,
                "address": self._address_from_bounds(bounds),
            }
        finally:
            wb.close()

    def fill_range(self, workbook_id: str | None, sheet_name: str, target_range: str, fill_color: str = "#FFFF00") -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            if str(fill_color or "").strip().lower() in {"none", "no_fill", "transparent", "무색", "없음"}:
                # "채우기 없음" — 흰색 칠과 다르다. 흰색은 Excel 기본 격자선을 가려
                # 시트가 종이처럼 하얘 보인다(2026-08-17 GUI 실측: 초기화 뒤 사용자가
                # 테두리가 안 보인다고 했다). 진짜 기본 상태는 무채움이다.
                fill = PatternFill(fill_type=None)
            else:
                fill = PatternFill(fill_type="solid", fgColor=self._to_argb(fill_color), bgColor=self._to_argb(fill_color))
            changed = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.fill = fill
                    changed += 1
            self._save_wb(wb, path)
            return {"changed_cells": changed, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    _CONDITION_EXPRESSION = re.compile(r"^\s*=?\s*(.+?)\s*(<=|>=|<>|!=|<|>|==?)\s*(.+?)\s*$")
    _KNOWN_OPERATORS = frozenset({">", ">=", "<", "<=", "==", "!=", "<>", "="})

    def _unpack_condition(
        self, ws: Any, operator: str, compare_column: str | None, threshold: float
    ) -> tuple[str, str | None, float]:
        """`=E2<F2` 나 `Current_Stock < Reorder_Point` 처럼 통째로 온 조건식을 푼다.

        플래너는 비교 대상을 operator 안에 넣어 답할 때가 있다. 그대로 두면
        "지원하지 않는 연산자"로 죽으므로, 오른쪽이 같은 행의 다른 열이면 열 비교로,
        숫자면 임계값으로 바꾼다. 열 이름은 이 시트의 머리글 행에서 직접 찾는다.
        """
        text = str(operator or "").strip()
        if text in self._KNOWN_OPERATORS:
            return text, compare_column, threshold
        match = self._CONDITION_EXPRESSION.match(text)
        if not match:
            return text, compare_column, threshold
        _left, symbol, right = match.group(1), match.group(2), match.group(3)
        symbol = {"=": "==", "!=": "<>"}.get(symbol, symbol)

        letter = self._column_letter_for(ws, right)
        if letter:
            return symbol, letter, threshold
        try:
            return symbol, compare_column, float(str(right).replace(",", "").replace("%", ""))
        except ValueError:
            return symbol, compare_column, threshold

    @staticmethod
    def _column_letter_for(ws: Any, token: str) -> str:
        """`F2`·`$F$2`·`Reorder_Point` 를 이 시트의 열 문자로 바꾼다."""
        text = str(token or "").strip().strip("'\"`[]")
        cell = re.fullmatch(r"\$?([A-Za-z]{1,3})\$?\d*", text)
        if cell:
            return cell.group(1).upper()
        headers = [(str(c.value or "").strip(), c.column_letter) for c in ws[1]]
        for header, column_letter in headers:
            if header and header.casefold() == text.casefold():
                return column_letter
        resolved = resolve_header(text, [header for header, _ in headers if header])
        for header, column_letter in headers:
            if resolved and header == resolved:
                return column_letter
        return ""

    def highlight_by_condition(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        operator: str,
        threshold: float,
        fill_color: str = "#FFFF00",
        compare_column: str | None = None,
        value: Any = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        # "현재고가 재주문점 이하" — 기준이 고정 숫자가 아니라 같은 행의 다른 열일 수 있다.
        compare_letter = str(compare_column or "").strip().upper() or None
        # 매출·이익 같은 실무 열은 대부분 수식이다. 수식 문자열을 숫자와 비교하면
        # 어떤 임계값을 줘도 0건이 되어 "조건에 맞는 셀이 없다"고 잘못 답한다.
        values = self._load_wb(path, data_only=True)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            self._sheet_or_raise(values, sheet_name)
            operator, compare_letter, threshold = self._unpack_condition(
                ws, operator, compare_letter, threshold
            )
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            fill = PatternFill(fill_type="solid", fgColor=self._to_argb(fill_color), bgColor=self._to_argb(fill_color))
            matched = 0
            changed = 0
            # 조건을 실제로 몇 칸에 대 봤는가. 0건이 "조건에 맞는 게 없어서"인지
            # "범위가 비어서"인지를 이 값으로만 가를 수 있다 — 전자는 정상이고
            # 후자는 대상을 잘못 잡은 것이라 검증 판정이 정반대여야 한다.
            scanned = 0
            evaluator = self._make_evaluator(path, wb, values, ws.title)

            def computed(row_idx: int, col_idx: int, fallback: Any) -> Any:
                """파일 캐시 → 우리 스냅샷 → 직접 계산 순으로 실제 값을 찾는다."""
                return self._evaluated(evaluator, ws.title, row_idx, col_idx, fallback)

            text_value = None if value is None or str(value).strip() == "" else str(value).strip()
            text_ops = {"==", "=", "!=", "<>"}
            # "입고예정일이 **비어 있는** 행만 노란색" — 빈 칸 자체가 조건인 문형.
            # 지금까지는 표현할 방법이 없어 0칸이 칠해지고 성공으로 보고됐다(2026-08-20 ex23).
            blank_op = str(operator or "").strip().lower()
            want_blank = blank_op in {"isblank", "blank", "empty", "빈칸", "비어있음"}
            want_filled = blank_op in {"notblank", "nonblank", "filled", "안비어있음"}
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    scanned += 1
                    left = computed(cell.row, cell.column, cell.value)
                    if want_blank or want_filled:
                        is_blank = left is None or str(left).strip() == ""
                        if is_blank is not want_blank:
                            continue
                    elif text_value is not None and str(operator or "").strip() in text_ops:
                        left_text = "" if left is None else str(left).strip()
                        op = str(operator or "").strip()
                        matched_text = left_text == text_value
                        if op in {"!=", "<>"}:
                            matched_text = not matched_text
                        if not matched_text:
                            continue
                    elif compare_letter:
                        limit_cell = ws[f"{compare_letter}{cell.row}"]
                        limit = computed(limit_cell.row, limit_cell.column, limit_cell.value)
                        if not isinstance(limit, (int, float)) or isinstance(limit, bool):
                            continue
                        if not self._matches_condition(left, operator, float(limit)):
                            continue
                    elif not self._matches_condition(left, operator, threshold):
                        continue
                    matched += 1
                    cell.fill = fill
                    changed += 1
            self._save_wb(wb, path)
            return {
                "matched_cells": matched,
                "changed_cells": changed,
                "scanned_cells": scanned,
                "address": self._address_from_bounds(bounds),
            }
        finally:
            wb.close()
            values.close()

    def apply_border(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        line_style: str = "continuous",
        weight: str = "medium",
        color: str = "#000000",
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            style = str(line_style or "continuous").strip().lower()
            if style == "none":
                border = Border()
            else:
                width_map = {"thin": "thin", "medium": "medium", "thick": "thick"}
                side = Side(style=width_map.get(str(weight or "thin").strip().lower(), "thin"), color=self._to_argb(color))
                border = Border(left=side, right=side, top=side, bottom=side)
            changed = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.border = border
                    changed += 1
            self._save_wb(wb, path)
            return {"changed_cells": changed, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    # xlsx 규격: 2010 이후 추가된 함수는 파일에 `_xlfn.` 접두로 저장돼야 한다.
    # 접두 없이 저장하면 Excel이 파일을 손상으로 취급해 복구 프롬프트 없이는
    # 열지도 못한다(2026-08-18 함수 배터리 실측: XLOOKUP·SEQUENCE가 든 파일이
    # Workbooks.Open에서 실패). 동적 배열 중 SORT·FILTER만 `_xlfn._xlws.`다.
    _XLWS_FUNCTIONS = ("FILTER", "SORT")
    _XLFN_FUNCTIONS = (
        "XLOOKUP", "XMATCH", "IFS", "SWITCH", "TEXTJOIN", "CONCAT",
        "MAXIFS", "MINIFS", "IFNA", "FORECAST.LINEAR", "FORMULATEXT",
        "SEQUENCE", "UNIQUE", "RANDARRAY", "SORTBY", "LET", "LAMBDA",
        "TOCOL", "TOROW", "VSTACK", "HSTACK", "TEXTSPLIT", "TEXTBEFORE",
        "TEXTAFTER", "WRAPROWS", "WRAPCOLS", "TAKE", "DROP",
        "CHOOSECOLS", "CHOOSEROWS", "GROUPBY", "PIVOTBY",
        "STDEV.S", "STDEV.P", "VAR.S", "VAR.P", "MODE.SNGL", "MODE.MULT",
        "PERCENTILE.INC", "PERCENTILE.EXC", "QUARTILE.INC", "QUARTILE.EXC",
        "RANK.EQ", "RANK.AVG", "CEILING.MATH", "FLOOR.MATH",
        "ISOWEEKNUM", "DAYS", "NUMBERVALUE", "AGGREGATE",
    )

    @classmethod
    def _normalize_modern_functions(cls, formula: str) -> str:
        """신형 함수 이름에 저장용 접두를 붙인다. 문자열 리터럴 안은 건드리지 않는다."""
        text = str(formula or "")
        if not text.startswith("="):
            return text
        segments = re.split(r'("[^"]*")', text)
        xlws = "|".join(re.escape(n) for n in cls._XLWS_FUNCTIONS)
        xlfn = "|".join(re.escape(n) for n in sorted(cls._XLFN_FUNCTIONS, key=len, reverse=True))
        for i in range(0, len(segments), 2):
            seg = segments[i]
            seg = re.sub(
                rf"(?<![A-Za-z0-9_.])({xlws})\(",
                lambda m: f"_xlfn._xlws.{m.group(1).upper()}(",
                seg,
                flags=re.IGNORECASE,
            )
            seg = re.sub(
                rf"(?<![A-Za-z0-9_.])({xlfn})\(",
                lambda m: f"_xlfn.{m.group(1).upper()}(",
                seg,
                flags=re.IGNORECASE,
            )
            segments[i] = seg
        return "".join(segments)

    def set_formula(self, workbook_id: str | None, sheet_name: str, range_ref: str, formula_a1: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, range_ref)
            min_row, min_col, max_row, max_col = bounds
            applied = 0
            origin = f"{get_column_letter(min_col)}{min_row}"
            formula_a1 = self._normalize_modern_functions(formula_a1)
            # 없는 시트를 참조하는 수식은 엑셀에서 #REF!가 된다. 쓰고 나서 알아채느니
            # 쓰기 전에 막는다(2026-08-20 624 게이트: `=SUM(지역성!E:E, 주문건수!E:E)`
            # — 모델이 '지역성과 시트 주문건수'를 시트 이름 둘로 쪼갰다).
            formula_a1, _missing = _resolve_formula_sheet_refs(formula_a1, list(wb.sheetnames))
            if _missing:
                raise ExcelLiveError(
                    f"수식이 없는 시트를 가리킵니다: {', '.join(_missing[:3])}"
                    + (" 등" if len(_missing) > 3 else "")
                    + f". 이 통합문서의 시트는 {', '.join(wb.sheetnames[:6])}입니다."
                )
            is_formula = str(formula_a1 or "").startswith("=")
            # TRANSPOSE 같은 구형 배열 함수는 일반 수식으로 저장하면 레거시
            # 암시적 교차로 해석돼 #VALUE!가 된다(2026-08-18 배터리 실측).
            # 배열 수식으로 저장하면 최신 Excel이 열 때 동적 배열로 바꿔 준다.
            if (
                is_formula
                and min_row == max_row
                and min_col == max_col
                and re.match(
                    r"^=\s*(TRANSPOSE|MMULT|MINVERSE|MDETERM|FREQUENCY|LINEST|TREND|GROWTH)\(",
                    formula_a1,
                    re.IGNORECASE,
                )
            ):
                ws.cell(row=min_row, column=min_col).value = ArrayFormula(
                    ref=origin, text=formula_a1
                )
                self._save_wb(wb, path, value_changed_sheet=ws.title, changed_rows={min_row})
                return {
                    "applied_cells": 1,
                    "address": origin,
                    "formula": formula_a1,
                    "array_formula": True,
                }
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if is_formula and (cell.row != min_row or cell.column != min_col):
                        # 같은 문자열을 그대로 복사하면 모든 행이 첫 행만 참조한다.
                        cell.value = Translator(formula_a1, origin=origin).translate_formula(
                            f"{get_column_letter(cell.column)}{cell.row}"
                        )
                    else:
                        cell.value = formula_a1
                    applied += 1
            self._save_wb(
                wb,
                path,
                value_changed_sheet=ws.title,
                changed_rows=set(range(min_row, max_row + 1)),
            )
            return {"formula_applied_cells": applied, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def verify_formula_result(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb_data = self._load_wb(path, data_only=True)
        wb_formula = self._load_wb(path, data_only=False)
        try:
            ws_data = self._sheet_or_raise(wb_data, sheet_name)
            ws_formula = self._sheet_or_raise(wb_formula, sheet_name)
            bounds = self._range_bounds(ws_formula, range_ref)
            min_row, min_col, max_row, max_col = bounds
            non_empty = 0
            numeric_values: list[float] = []
            samples: list[Any] = []
            formula_cells = 0
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    value = ws_data.cell(row=r, column=c).value
                    formula_raw = ws_formula.cell(row=r, column=c).value
                    if isinstance(formula_raw, str) and formula_raw.startswith("="):
                        formula_cells += 1
                    if value is None or (isinstance(value, str) and not value.strip()):
                        continue
                    non_empty += 1
                    if len(samples) < 10:
                        samples.append(value)
                    num = self._as_float(value)
                    if num is not None:
                        numeric_values.append(num)
            if non_empty == 0 and formula_cells > 0:
                non_empty = formula_cells
            numeric_count = len(numeric_values)
            if numeric_count == 0 and formula_cells > 0:
                numeric_count = formula_cells
            total = sum(numeric_values) if numeric_values else 0.0
            avg = (total / len(numeric_values)) if numeric_values else 0.0
            return {
                "address": self._address_from_bounds(bounds),
                "non_empty_cells": non_empty,
                "numeric_cells": numeric_count,
                "sum": total,
                "average": avg,
                "sample_values": samples,
            }
        finally:
            wb_data.close()
            wb_formula.close()

    # ---- 표 단위 작업 ----

    def create_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        start_cell: str,
        rows: int,
        cols: int,
        with_border: bool = True,
    ) -> dict[str, Any]:
        row_count = max(1, min(100, int(rows)))
        col_count = max(1, min(50, int(cols)))
        result = self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=start_cell,
            values_2d=[["" for _ in range(col_count)] for _ in range(row_count)],
        )
        if with_border:
            self.apply_border(
                workbook_id=workbook_id,
                sheet_name=sheet_name,
                target_range=result.get("address", start_cell),
                line_style="continuous",
                weight="thin",
                color="#000000",
            )
        return {
            "created": True,
            "address": result.get("address", start_cell),
            "rows": row_count,
            "cols": col_count,
        }

    @staticmethod
    def _require_computed_values(
        payload: dict[str, Any],
        column_indexes: list[int | None],
        header: list[Any],
        purpose: str,
    ) -> None:
        """계산값을 못 구한 수식 열로는 집계하지 않는다. 틀린 숫자를 내놓는 대신 멈춘다.

        우리가 파일을 저장하면 Excel이 넣어 둔 계산값 캐시가 사라진다. 스냅샷도 없으면
        매출 열이 "=I2*J2" 문자열이라, 그대로 두면 그 열을 비숫자로 보고 엉뚱한 열을
        더하거나 0을 답한다. 실제로 날짜를 더해 4.8e16이 나온 적이 있다.
        """
        unresolved = payload.get("unresolved_formulas") or []
        if not unresolved:
            return
        targets = {int(idx) for idx in column_indexes if idx is not None}
        broken = sorted({col for _row, col in unresolved if col in targets})
        if not broken:
            return
        names = ", ".join(str(header[idx]) if idx < len(header) else f"{idx + 1}열" for idx in broken)
        raise ExcelLiveError(
            f"'{names}' 열은 수식이고 계산된 값이 파일에 없어 {purpose}할 수 없습니다. "
            "Excel에서 해당 파일을 한 번 열어 저장하면 계산값이 채워집니다."
        )

    def _reinstall_relocated_values(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        raw_body: list[list[Any]],
        computed_body: list[list[Any]],
        order: list[int],
        min_col: int,
        body_start_row: int,
    ) -> None:
        """행을 재배치한 뒤, 수식 셀의 계산값을 새 좌표에 다시 심는다.

        정렬·중복제거는 clear+write로 진행되므로 그 과정에서 스냅샷이 비워진다. 여기서
        다시 심어 두지 않으면 "정렬한 다음 지역별 합계" 같은 두 번째 요청이 값을 잃는다.
        """
        if len(computed_body) != len(raw_body):
            return
        cells: dict[tuple[int, int], Any] = {}
        for dest_offset, source_idx in enumerate(order):
            dest_row = body_start_row + dest_offset
            for col_offset, raw in enumerate(raw_body[source_idx]):
                if not (isinstance(raw, str) and raw.startswith("=")):
                    continue
                value = computed_body[source_idx][col_offset]
                if value is None or (isinstance(value, str) and value.startswith("=")):
                    continue
                cells[(dest_row, min_col + col_offset)] = value
        if not cells:
            return
        try:
            path = self._resolve_workbook_path(workbook_id)
        except ExcelLiveError:
            return
        formula_cache.install_cells(path, sheet_name, cells)

    @staticmethod
    def _relocate_rows(
        raw_body: list[list[Any]],
        order: list[int],
        *,
        min_col: int,
        body_start_row: int,
    ) -> list[list[Any]]:
        """행 순서를 바꿔 쓰되 수식은 새 행 기준으로 다시 계산되게 옮긴다.

        `=I2*J2` 를 문자열 그대로 5행에 쓰면 5행 매출이 2행 값을 가리킨다.
        정렬·중복 제거는 실무 표에서 늘 수식 열을 끼고 있으므로 여기서 반드시 번역해야 한다.
        """
        relocated: list[list[Any]] = []
        for dest_offset, source_idx in enumerate(order):
            source_row_no = body_start_row + source_idx
            dest_row_no = body_start_row + dest_offset
            row_out: list[Any] = []
            for col_offset, value in enumerate(raw_body[source_idx]):
                if isinstance(value, str) and value.startswith("=") and source_row_no != dest_row_no:
                    letter = get_column_letter(min_col + col_offset)
                    try:
                        value = Translator(value, origin=f"{letter}{source_row_no}").translate_formula(
                            f"{letter}{dest_row_no}"
                        )
                    except Exception:
                        pass
                row_out.append(value)
            relocated.append(row_out)
        return relocated

    def sort_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_column: str | int = 1,
        order: str = "asc",
        has_header: bool = True,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, target_range)
        computed = self.read_computed_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"sorted_rows": 0, "address": payload.get("address", target_range)}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        computed_rows = computed.get("values", []) or normalized
        computed_norm = [list(row) + [None] * (col_count - len(row)) for row in computed_rows]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        computed_body = computed_norm[1:] if has_header else computed_norm
        if not body:
            return {"sorted_rows": 0, "address": payload.get("address", target_range)}
        # 꼬리 합계행(이름표가 합계류거나 수식이 든 마지막 줄)은 데이터가 아니다 —
        # 같이 정렬하면 합계가 데이터 사이에 섞인다(2026-08-18 멀티턴 사냥).
        # 꼬리의 집계 줄은 **여러 줄**일 수 있다(합계 + 평균). 마지막 한 줄만
        # 고정하면 합계 줄이 데이터 사이로 섞인다(2026-08-18 대화형 러너 실측:
        # 합계가 2행으로 올라가고 크로스시트 SUM이 이중 집계).
        pinned_tail_rows: list[list] = []
        _AGG_LABELS = {"합계", "총계", "계", "총합", "평균", "최대", "최소", "개수", "total", "sum", "avg", "average"}
        while len(body) > 1:
            tail = body[-1]
            tail_label = tail[0] if tail else None
            tail_has_formula = any(isinstance(v, str) and str(v).startswith("=") for v in tail)
            is_agg_label = isinstance(tail_label, str) and tail_label.strip().lower() in _AGG_LABELS
            if not (tail_has_formula or is_agg_label):
                break
            pinned_tail_rows.insert(0, body.pop())
            if computed_body:
                computed_body = computed_body[:-1]
        # 수식 열을 기준으로 정렬할 때 문자열("=I2*J2")로 줄을 세우면 뒤죽박죽이 된다.
        sort_source = computed_body if len(computed_body) == len(body) else body
        address = str(payload.get("address", target_range))
        min_col, min_row, max_col, max_row = range_boundaries(address)
        key_idx = self._resolve_column_selector(
            selector=key_column,
            start_col_idx=min_col,
            col_count=col_count,
            header_row=header,
        )
        reverse = str(order or "asc").strip().lower() in {"desc", "descending", "내림차순"}
        # 숫자로 읽히는 행이 먼저, 나머지는 뒤에 사전순. 섞어서 비교하면 타입 오류가 난다.
        numbered: list[tuple[float, str, int]] = []
        lettered: list[tuple[str, int]] = []
        for index, row in enumerate(sort_source):
            cell = row[key_idx] if key_idx < len(row) else None
            number = _as_number(cell)
            if number is None:
                lettered.append((_text_key(cell), index))
            else:
                numbered.append((number, _text_key(cell), index))
        numbered.sort(key=lambda item: (item[0], item[1]), reverse=reverse)
        lettered.sort(key=lambda item: item[0], reverse=reverse)
        sort_order = [index for *_rest, index in numbered] + [index for _key, index in lettered]
        body_start_row = min_row + (1 if has_header else 0)
        out_values = self._relocate_rows(
            body,
            sort_order,
            min_col=min_col,
            body_start_row=body_start_row,
        )
        final_values = [header, *out_values] if has_header and header is not None else out_values
        if pinned_tail_rows:
            # 고정된 집계 줄들은 원래 자리(맨 아래)에 그대로 돌아간다 — 행이 안
            # 움직였으니 수식 참조도 손댈 필요가 없다.
            final_values.extend(pinned_tail_rows)
        # 값을 다시 쓰기 **전에** 본문 행들의 서식을 찍어 둔다. 안 그러면 정렬 뒤
        # 채움·글꼴이 제자리에 남아 다른 값을 가리킨다(2026-08-20 ex23 실측:
        # 990,000에 칠한 빨강이 정렬 뒤 27,500 위에 남았다).
        row_styles = self._capture_row_styles(
            workbook_id,
            sheet_name,
            body_start_row=body_start_row,
            row_count=len(body),
            min_col=min_col,
            col_count=col_count,
        )
        self.clear_range(workbook_id, sheet_name, address)
        self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=f"{self._idx_to_col(min_col)}{min_row}",
            values_2d=final_values,
        )
        self._reinstall_relocated_values(
            workbook_id,
            sheet_name,
            raw_body=body,
            computed_body=computed_body,
            order=sort_order,
            min_col=min_col,
            body_start_row=body_start_row,
        )
        self._apply_row_styles(
            workbook_id,
            sheet_name,
            styles=row_styles,
            order=sort_order,
            body_start_row=body_start_row,
            min_col=min_col,
        )
        return {
            "sorted_rows": len(out_values),
            "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            "key_column_index": key_idx + 1,
            "order": "desc" if reverse else "asc",
        }

    def _capture_row_styles(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        body_start_row: int,
        row_count: int,
        min_col: int,
        col_count: int,
    ) -> list[list[Any]]:
        """정렬 대상 본문 행들의 셀 서식을 행 단위로 찍는다. 못 읽으면 빈 목록(서식 보정 생략)."""
        from copy import copy

        try:
            path = self._resolve_workbook_path(workbook_id)
            wb = self._load_wb(path)
            ws = self._sheet_or_raise(wb, sheet_name)
            out: list[list[Any]] = []
            for offset in range(row_count):
                row = []
                for col_offset in range(col_count):
                    cell = ws.cell(row=body_start_row + offset, column=min_col + col_offset)
                    row.append(copy(cell._style))
                out.append(row)
            return out
        except Exception:
            return []

    def _apply_row_styles(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        styles: list[list[Any]],
        order: list[int],
        body_start_row: int,
        min_col: int,
    ) -> None:
        """찍어 둔 서식을 **정렬된 순서대로** 되돌려 놓는다.

        `order[j]`가 새 j번째 자리에 온 원래 행 번호다 — 그 행의 서식을 그대로 가져온다.
        """
        if not styles or not order:
            return
        try:
            path = self._resolve_workbook_path(workbook_id)
            wb = self._load_wb(path)
            ws = self._sheet_or_raise(wb, sheet_name)
            for new_offset, source_index in enumerate(order):
                if source_index >= len(styles):
                    continue
                for col_offset, style in enumerate(styles[source_index]):
                    ws.cell(
                        row=body_start_row + new_offset, column=min_col + col_offset
                    )._style = style
            self._save_wb(wb, path)
        except Exception:
            return

    def dedupe_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str | None = None,
        key_columns: list[str | int] | None = None,
        has_header: bool = True,
        columns: list[str | int] | None = None,
    ) -> dict[str, Any]:
        # 범위를 말하지 않은 "중복 지워줘"는 시트 전체가 대상이다.
        target_range = target_range or self.get_used_range_ref(workbook_id, sheet_name)
        key_columns = key_columns or columns
        payload = self.read_range(workbook_id, sheet_name, target_range)
        computed = self.read_computed_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"removed_rows": 0, "remaining_rows": 0, "address": payload.get("address", target_range)}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        computed_norm = [
            list(row) + [None] * (col_count - len(row)) for row in (computed.get("values") or normalized)
        ]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        computed_body = computed_norm[1:] if has_header else computed_norm
        if not body:
            return {"removed_rows": 0, "remaining_rows": 0, "address": payload.get("address", target_range)}
        start_col = 1
        if key_columns:
            subset = [
                self._resolve_column_selector(col, start_col, col_count, header)
                for col in key_columns
            ]
        else:
            subset = list(range(col_count))
        before = len(body)
        # 먼저 나온 행을 남긴다. 사용자는 보통 위쪽 행을 원본으로 여긴다.
        seen: set[tuple] = set()
        keep_order: list[int] = []
        for index, row in enumerate(body):
            key = _group_key(row, subset)
            if key in seen:
                continue
            seen.add(key)
            keep_order.append(index)
        after = len(keep_order)
        address = str(payload.get("address", target_range))
        min_col, min_row, max_col, max_row = range_boundaries(address)
        body_start_row = min_row + (1 if has_header else 0)
        out = self._relocate_rows(
            body,
            keep_order,
            min_col=min_col,
            body_start_row=body_start_row,
        )
        final_values = [header, *out] if has_header and header is not None else out
        self.clear_range(workbook_id, sheet_name, address)
        self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=f"{self._idx_to_col(min_col)}{min_row}",
            values_2d=final_values,
        )
        self._reinstall_relocated_values(
            workbook_id,
            sheet_name,
            raw_body=body,
            computed_body=computed_body,
            order=keep_order,
            min_col=min_col,
            body_start_row=body_start_row,
        )
        return {
            "removed_rows": max(0, before - after),
            "removed_duplicates": max(0, before - after),
            "remaining_rows": after,
            "kept_rows": after,
            "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            "key_columns": [idx + 1 for idx in subset],
        }

    @staticmethod
    def _row_matches(cell: Any, operator: str, value: Any) -> bool:
        """한 셀이 필터 조건을 만족하는지. 숫자 비교가 불가능하면 문자열 같음으로 되돌린다."""
        op = str(operator or "==").strip()
        if op == "contains":
            return str(value) in str(cell if cell is not None else "")
        if op in {"=", "=="}:
            return str(cell) == str(value)
        if op == "!=":
            return str(cell) != str(value)
        compare = {
            ">": lambda a, b: a > b,
            ">=": lambda a, b: a >= b,
            "<": lambda a, b: a < b,
            "<=": lambda a, b: a <= b,
        }.get(op)
        threshold = _as_number(value)
        number = _as_number(cell)
        if compare is None or threshold is None:
            return str(cell) == str(value)
        return number is not None and compare(number, threshold)

    def _hide_unmatched_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        address: str,
        has_header: bool,
        keep_order: list[int],
        body_len: int,
        matched: int,
        col_idx: int,
        op: str,
        value: Any,
        exclude: bool = False,
    ) -> dict[str, Any]:
        """안 맞는 행을 숨긴다. **한 칸도 지우지 않는다.**

        엑셀 자동필터 표시(`auto_filter.ref`)도 함께 걸어, 사람이 파일을 열었을 때
        필터가 걸려 있다는 걸 알아보고 직접 풀 수 있게 한다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            _, min_row, _, _ = range_boundaries(address)
            body_start = min_row + (1 if has_header else 0)
            keep = set(keep_order)
            hidden = 0
            for offset in range(body_len):
                row_no = body_start + offset
                should_hide = offset not in keep
                ws.row_dimensions[row_no].hidden = should_hide
                hidden += 1 if should_hide else 0
            ws.auto_filter.ref = address
            self._save_wb(wb, path)
        finally:
            wb.close()
        return {
            "filtered_rows": matched,
            "matched_rows": matched,
            "hidden_rows": hidden,
            "removed_rows": 0,
            "remaining_rows": body_len,
            "no_change": hidden == 0,
            "address": address,
            "column_index": col_idx + 1,
            "operator": op,
            "value": value,
            "mode": "hide_exclude" if exclude else "hide",
        }

    def filter_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str | None = None,
        column: str | int = 1,
        operator: str = "==",
        value: Any = None,
        has_header: bool = True,
        mode: str = "hide",
    ) -> dict[str, Any]:
        """조건에 맞는 행만 남긴다. `mode`가 무엇을 "남긴다"의 뜻으로 정한다.

        - `hide`(기본): 안 맞는 행을 **숨긴다**. 엑셀의 필터가 하는 일이고, 되돌릴 수 있다.
        - `keep`: 안 맞는 행을 **지운다**. 사람이 지우라고 말했을 때만 쓴다.
        - `remove`: 맞는 행을 지운다("취소된 주문은 빼줘").

        2026-08-20 파괴 게이트: 기본이 `keep`이라 "잠깐 안 보이게 해줘"가 행을 지웠다.
        숨기기는 되돌릴 수 있고 삭제는 아니므로, 애매하면 숨기는 쪽이 옳다.


        `mode="remove"`면 반대로 조건에 맞는 행을 지운다 — "취소된 주문은 빼줘"처럼
        제외를 요청한 문장용이다. 세는 것만으로는 사용자가 요청한 편집이 일어나지 않는데도
        성공으로 보고돼, 화면은 그대로인 채 "처리했습니다"라는 답만 남는다.

        수식 열도 조건 대상이 되도록 계산값으로 비교하고, 남길 행은 dedupe와 같은
        재배치 경로를 태워 행 이동에 따른 수식 참조를 함께 옮긴다.
        """
        target_range = target_range or self.get_used_range_ref(workbook_id, sheet_name)
        payload = self.read_range(workbook_id, sheet_name, target_range)
        computed = self.read_computed_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        address = str(payload.get("address", target_range))
        if not values:
            return {"filtered_rows": 0, "removed_rows": 0, "remaining_rows": 0, "address": address}

        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        computed_norm = [
            list(row) + [None] * (col_count - len(row)) for row in (computed.get("values") or normalized)
        ]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        computed_body = computed_norm[1:] if has_header else computed_norm
        if not body:
            return {"filtered_rows": 0, "removed_rows": 0, "remaining_rows": 0, "address": address}

        col_idx = self._resolve_column_selector(column, 1, col_count, header)
        op = str(operator or "==").strip()
        mode_word = str(mode or "hide").strip().lower()
        # 숨길지 지울지(`hide_only`)와 어느 쪽을 뺄지(`exclude`)는 **다른 축**이다.
        # `hide_exclude`는 "맞는 행을 숨긴다" — `remove`의 비파괴판이다.
        exclude = mode_word in {"remove", "exclude", "drop", "hide_exclude"}
        hide_only = mode_word in {"hide", "hide_exclude", "숨김", "숨기기"}
        keep_order: list[int] = []
        matched = 0
        for index, row in enumerate(body):
            source = computed_body[index] if index < len(computed_body) else row
            cell = source[col_idx] if col_idx < len(source) else None
            if cell is None and col_idx < len(row):
                cell = row[col_idx]
            hit = self._row_matches(cell, op, value)
            matched += 1 if hit else 0
            if hit != exclude:
                keep_order.append(index)

        before = len(body)
        after = len(keep_order)
        if after == before:
            return {
                "filtered_rows": matched,
                "matched_rows": matched,
                "removed_rows": 0,
                "remaining_rows": after,
                "no_change": True,
                "address": address,
                "column_index": col_idx + 1,
                "operator": op,
                "value": value,
                "mode": "remove" if exclude else "keep",
            }
        if not exclude and matched == 0 and before > 0:
            # "제주인 행만 남겨줘"인데 제주가 한 행도 없다 — 그대로 진행하면
            # **시트의 데이터가 통째로 지워진다** (2026-08-17 실측: 4행 전부 삭제,
            # filter_rows가 롤백 스냅샷 목록에도 없어 복구조차 안 됐다).
            # 조건 값이 틀렸을 가능성이 압도적이므로 파일을 건드리지 않고 알린다.
            return {
                "filtered_rows": 0,
                "matched_rows": 0,
                "removed_rows": 0,
                "remaining_rows": before,
                "no_change": True,
                "address": address,
                "column_index": col_idx + 1,
                "operator": op,
                "value": value,
                "mode": "keep",
            }

        # 숨기기는 **무일치 가드 뒤**에 온다. 앞에 두면 "제주인 행만 남겨줘"인데 제주가
        # 없을 때 표를 통째로 숨겨 빈 화면을 만든다 — 지우진 않지만 정직하지도 않다
        # (2026-08-20 자체 검토: 새 층을 앞에 끼워 기존 판단을 덮은, 내가 적어 둔 그 유형).
        if hide_only:
            return self._hide_unmatched_rows(
                workbook_id,
                sheet_name,
                address=address,
                has_header=has_header,
                keep_order=keep_order,
                body_len=before,
                matched=matched,
                col_idx=col_idx,
                op=op,
                value=value,
                exclude=exclude,
            )

        min_col, min_row, max_col, max_row = range_boundaries(address)
        body_start_row = min_row + (1 if has_header else 0)
        out = self._relocate_rows(body, keep_order, min_col=min_col, body_start_row=body_start_row)
        final_values = [header, *out] if has_header and header is not None else out
        self.clear_range(workbook_id, sheet_name, address)
        self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=f"{self._idx_to_col(min_col)}{min_row}",
            values_2d=final_values,
        )
        self._reinstall_relocated_values(
            workbook_id,
            sheet_name,
            raw_body=body,
            computed_body=computed_body,
            order=keep_order,
            min_col=min_col,
            body_start_row=body_start_row,
        )
        return {
            "filtered_rows": matched,
            "matched_rows": matched,
            "removed_rows": max(0, before - after),
            "remaining_rows": after,
            "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            "column_index": col_idx + 1,
            "operator": op,
            "value": value,
            "mode": "remove" if exclude else "keep",
        }

    # ---- 표 단위 작업 (열 이름 기준) ----

    def _table_grid(self, workbook_id: str | None, sheet_name: str) -> tuple[list[Any], list[list[Any]], int]:
        """사용 범위를 머리글 한 줄과 본문으로 나눠 준다. 열 이름 기반 작업의 공통 진입점."""
        ref = self.get_used_range_ref(workbook_id, sheet_name)
        payload = self.read_computed_range(workbook_id, sheet_name, ref)
        rows = payload.get("values") or []
        if not rows:
            raise ExcelLiveError(f"'{sheet_name}' 시트에 읽을 데이터가 없습니다.")
        width = max(len(row) for row in rows)
        grid = [list(row) + [None] * (width - len(row)) for row in rows]
        return grid[0], grid[1:], width

    @staticmethod
    def _pick_column(selector: str | int, header: list[Any], width: int) -> tuple[int, str | None]:
        """머리글 이름·한국어 개념어·열 문자를 열 번호로 바꾼다.

        못 찾으면 1열로 넘어가지 않고 실패한다. 엉뚱한 열을 조용히 지우거나 더하는 것보다
        어떤 열이 있는지 알려주고 멈추는 편이 낫다.
        """
        if isinstance(selector, int):
            index = selector - 1
            if not 0 <= index < width:
                raise ExcelLiveError(f"{selector}번째 열이 범위를 벗어났습니다. (열 개수 {width})")
            return index, None
        text = str(selector or "").strip()
        if not text:
            raise ExcelLiveError("대상 열을 지정해 주세요.")
        names = [str(cell or "").strip() for cell in header]
        for index, name in enumerate(names):
            if name and name.lower() == text.lower():
                return index, name
        mapped = resolve_header(text, [name for name in names if name])
        if mapped:
            return names.index(mapped), mapped
        if re.fullmatch(r"[A-Za-z]{1,3}", text):
            index = column_index_from_string(text.upper()) - 1
            if 0 <= index < width:
                return index, None
        available = ", ".join(name for name in names if name)
        raise ExcelLiveError(f"'{text}' 열을 찾을 수 없습니다. 사용 가능한 열: {available}")

    def calculate_column_stat(
        self,
        workbook_id: str | None,
        sheet_name: str,
        column: str | int,
        stat: str = "sum",
    ) -> dict[str, Any]:
        header, body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        numbers = [n for n in (_as_number(row[index]) for row in body) if n is not None]
        kind = str(stat or "sum").strip().lower()
        kind = "average" if kind in {"average", "avg", "mean"} else kind
        if kind not in {"sum", "average", "count", "max", "min"}:
            raise ExcelLiveError(f"지원하지 않는 통계입니다: {stat}")
        if kind == "count":
            value = float(len(numbers))
        elif not numbers:
            raise ExcelLiveError(f"'{name or column}' 열에 숫자가 없어 {kind} 통계를 낼 수 없습니다.")
        elif kind == "sum":
            value = float(sum(numbers))
        elif kind == "average":
            value = float(sum(numbers)) / len(numbers)
        else:
            value = float(max(numbers) if kind == "max" else min(numbers))
        return {
            "value": value,
            "column": get_column_letter(index + 1),
            "header": name,
            "stat": kind,
            "numeric_count": len(numbers),
        }

    def sort_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        column: str | int,
        order: str = "asc",
    ) -> dict[str, Any]:
        """시트 전체를 한 열 기준으로 정렬한다. 범위를 말하지 않은 "정렬해줘"의 기본 경로."""
        header, _body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        ref = self.get_used_range_ref(workbook_id, sheet_name)
        result = self.sort_range(
            workbook_id,
            sheet_name,
            ref,
            key_column=index + 1,
            order=order,
            has_header=True,
        )
        return {
            "sorted_rows": result.get("sorted_rows", 0),
            "order": result.get("order", "asc"),
            "column": name or get_column_letter(index + 1),
            "address": result.get("address", ref),
        }

    def drop_column(self, workbook_id: str | None, sheet_name: str, column: str | int) -> dict[str, Any]:
        header, _body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            ws.delete_cols(index + 1)
            self._save_wb(wb, path)
        finally:
            wb.close()
        # 열이 통째로 밀렸으니 이 시트의 수식 계산값 스냅샷은 더 이상 믿을 수 없다.
        formula_cache.invalidate_sheet(path, sheet_name)
        return {
            "dropped_column": name or get_column_letter(index + 1),
            "remaining_columns": max(0, width - 1),
        }

    def rename_column(
        self,
        workbook_id: str | None,
        sheet_name: str,
        column: str | int,
        new_name: str,
    ) -> dict[str, Any]:
        target = str(new_name or "").strip()
        if not target:
            raise ExcelLiveError("새 열 이름을 알려주세요.")
        header, _body, width = self._table_grid(workbook_id, sheet_name)
        index, name = self._pick_column(column, header, width)
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            ws.cell(row=1, column=index + 1, value=target)
            self._save_wb(wb, path)
        finally:
            wb.close()
        return {
            "old_name": name or get_column_letter(index + 1),
            "new_name": target,
            "column": get_column_letter(index + 1),
        }

    def add_column(
        self,
        workbook_id: str | None,
        sheet_name: str,
        name: str,
        formula_a1: str | None = None,
    ) -> dict[str, Any]:
        label = str(name or "").strip()
        if not label:
            raise ExcelLiveError("추가할 열 이름을 알려주세요.")
        _header, body, width = self._table_grid(workbook_id, sheet_name)
        target_col = width + 1
        letter = get_column_letter(target_col)
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        filled = 0
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            ws.cell(row=1, column=target_col, value=label)
            if formula_a1:
                origin = f"{letter}2"
                for offset in range(len(body)):
                    row_no = 2 + offset
                    # 첫 행 수식을 그대로 복사하면 모든 행이 2행을 가리킨다. 행마다 옮겨 준다.
                    text = (
                        formula_a1
                        if offset == 0
                        else Translator(formula_a1, origin=origin).translate_formula(f"{letter}{row_no}")
                    )
                    ws.cell(row=row_no, column=target_col, value=text)
                    filled += 1
            self._save_wb(wb, path)
        finally:
            wb.close()
        return {"column": letter, "name": label, "formula_filled_cells": filled}

    def group_by_aggregate(
        self,
        workbook_id: str | None,
        sheet_name: str,
        group_column: str | int,
        agg: str = "sum",
        value_column: str | int | None = None,
    ) -> dict[str, Any]:
        """열 하나로 묶어 집계한 결과를 값으로 돌려준다. 시트에 쓰지는 않는다."""
        header, body, width = self._table_grid(workbook_id, sheet_name)
        group_index, group_name = self._pick_column(group_column, header, width)
        kind = str(agg or "sum").strip().lower()
        kind = "average" if kind in {"average", "avg", "mean"} else kind
        if kind not in {"sum", "average", "count", "max", "min"}:
            raise ExcelLiveError(f"지원하지 않는 집계입니다: {agg}")
        value_index: int | None = None
        value_name: str | None = None
        if value_column is not None and str(value_column).strip():
            value_index, value_name = self._pick_column(value_column, header, width)
        elif kind != "count":
            raise ExcelLiveError(f"{kind} 집계에는 대상 값 열이 필요합니다. 예: 매출")

        buckets: dict[tuple, dict[str, Any]] = {}
        for row in body:
            key = _group_key(row, [group_index])
            bucket = buckets.setdefault(
                key,
                {"label": row[group_index] if group_index < len(row) else None, "numbers": [], "count": 0},
            )
            bucket["count"] += 1
            if value_index is not None:
                number = _as_number(row[value_index]) if value_index < len(row) else None
                if number is not None:
                    bucket["numbers"].append(number)

        groups: list[dict[str, Any]] = []
        for bucket in buckets.values():
            numbers = bucket["numbers"]
            if kind == "count":
                value = float(bucket["count"])
            elif not numbers:
                value = 0.0
            elif kind == "sum":
                value = float(sum(numbers))
            elif kind == "average":
                value = float(sum(numbers)) / len(numbers)
            else:
                value = float(max(numbers) if kind == "max" else min(numbers))
            groups.append({"key": bucket["label"], "value": value, "count": bucket["count"]})
        # 큰 값부터 보여 주는 편이 "어디가 제일 많아?"라는 원래 질문에 바로 답이 된다.
        groups.sort(key=lambda g: (-g["value"], str(g["key"])))
        return {
            "agg": kind,
            "group_column": group_name or get_column_letter(group_index + 1),
            "value_column": value_name,
            "groups": groups,
        }

    # ---- 기타 작업 ----

    def save_workbook(self, workbook_id: str | None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        return {
            "saved": True,
            "workbook_id": str(path),
            "name": path.name,
            "full_path": str(path),
        }

    def create_workbook_backup(self, workbook_id: str | None, *, label: str = "auto") -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        backup_dir = path.parent / "officeclaw_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_label = re.sub(r"[^A-Za-z0-9_-]+", "_", str(label or "auto")).strip("_") or "auto"
        backup_path = backup_dir / f"{path.stem}.{safe_label}.{stamp}{path.suffix}"
        shutil.copy2(path, backup_path)
        return {
            "backup_created": True,
            "backup_path": str(backup_path),
            "source_path": str(path),
            "workbook_id": str(path),
        }

    def list_workbook_backups(self, workbook_id: str | None, *, limit: int = 20) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        backup_dir = path.parent / "officeclaw_backups"
        rows: list[dict[str, Any]] = []
        if backup_dir.exists():
            for fp in backup_dir.glob(f"{path.stem}.*{path.suffix}"):
                if not fp.is_file():
                    continue
                stat = fp.stat()
                rows.append(
                    {
                        "backup_path": str(fp.resolve()),
                        "backup_name": fp.name,
                        "size_bytes": int(stat.st_size),
                        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                        "modified_ts": float(stat.st_mtime),
                    }
                )
        rows.sort(key=lambda r: float(r.get("modified_ts", 0.0)), reverse=True)
        for row in rows:
            row.pop("modified_ts", None)
        return {
            "workbook_id": str(path),
            "source_path": str(path),
            "backup_dir": str(backup_dir),
            "backups": rows[: max(1, min(200, int(limit)))],
        }

    def restore_workbook_from_backup(self, workbook_id: str | None, *, backup_path: str | None = None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        if backup_path:
            source = Path(str(backup_path)).expanduser().resolve()
        else:
            listed = self.list_workbook_backups(str(path), limit=1)
            backups = listed.get("backups", [])
            if not backups:
                raise ExcelLiveError("복구 가능한 백업이 없습니다.")
            source = Path(str(backups[0].get("backup_path", ""))).resolve()
        if not source.exists() or not source.is_file():
            raise ExcelLiveError(f"백업 파일을 찾을 수 없습니다: {source}")
        pre_restore = self.create_workbook_backup(str(path), label="pre_restore")
        shutil.copy2(source, path)
        wb = self._load_wb(path)
        try:
            active = wb.active.title if wb.sheetnames else ""
        finally:
            wb.close()
        self._selected_workbook_id = str(path)
        self._selected_sheet_by_workbook[str(path)] = active
        return {
            "restored": True,
            "workbook_id": str(path),
            "name": path.name,
            "full_path": str(path),
            "restored_from_backup_path": str(source),
            "pre_restore_backup_path": str(pre_restore.get("backup_path", "")),
            "active_sheet": active,
        }

    def get_active_selection_ref(self, workbook_id: str | None, sheet_name: str | None) -> str:
        """파일 기반 엔진에는 '선택 영역' 개념이 없으므로 데이터 영역으로 해석한다.

        여기서 A1을 돌려주면 범위를 말하지 않은 모든 명령이 한 칸에만 적용된 채
        성공으로 보고되므로, 반드시 사용 영역을 기준으로 삼는다.
        """
        return self.get_used_range_ref(workbook_id, sheet_name)

    def get_used_range_ref(self, workbook_id: str | None, sheet_name: str | None) -> str:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            target_sheet = sheet_name or self._selected_sheet_by_workbook.get(str(path)) or wb.active.title
            ws = self._sheet_or_raise(wb, target_sheet)
            max_row, max_col = self._used_bounds(ws)
            return self._address_from_bounds((1, 1, max_row, max_col))
        finally:
            wb.close()

    def find_replace(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        find_text: str,
        replace_text: str,
        *,
        match_case: bool = False,
        whole_cell: bool = False,
    ) -> dict[str, Any]:
        if not find_text:
            raise ExcelLiveError("find_replace.find_text가 비어 있습니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            needle = find_text if match_case else find_text.lower()
            replaced = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    value = cell.value
                    if not isinstance(value, str):
                        continue
                    haystack = value if match_case else value.lower()
                    if whole_cell:
                        if haystack == needle:
                            cell.value = replace_text
                            replaced += 1
                        continue
                    if needle not in haystack:
                        continue
                    if match_case:
                        cell.value = value.replace(find_text, replace_text)
                    else:
                        # 대소문자 무시 치환은 원본 표기를 살려야 해서 위치 기반으로 잘라 붙인다.
                        out: list[str] = []
                        cursor = 0
                        lowered = haystack
                        while True:
                            idx = lowered.find(needle, cursor)
                            if idx == -1:
                                out.append(value[cursor:])
                                break
                            out.append(value[cursor:idx])
                            out.append(replace_text)
                            cursor = idx + len(needle)
                        cell.value = "".join(out)
                    replaced += 1
            self._save_wb(
                wb,
                path,
                value_changed_sheet=ws.title,
                changed_rows=set(range(min_row, max_row + 1)),
            )
            return {"replaced_cells": replaced, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def merge_cells(self, workbook_id: str | None, sheet_name: str, target_range: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            # 병합은 왼쪽 위 셀 값만 남기고 나머지를 지운다 — 병합 전 값이 여러 개면
            # 표시상 하나만 보이는 게 openpyxl/Excel 동작과 같다. 첫 값을 보존한다.
            min_row, min_col, max_row, max_col = bounds
            keep_value = ws.cell(row=min_row, column=min_col).value
            # 값이 둘 이상이면 병합은 **지우는 일**이다. 조용히 넘어가지 않는다 —
            # "제목 줄 병합해줘"가 머리글 줄을 먹어 머리글 다섯 개가 사라졌다
            # (2026-08-20 파괴 게이트: 12문형 중 9개).
            doomed = [
                f"{self._idx_to_col(c)}{r}={ws.cell(row=r, column=c).value!r}"
                for r in range(min_row, max_row + 1)
                for c in range(min_col, max_col + 1)
                if not (r == min_row and c == min_col)
                and str(ws.cell(row=r, column=c).value or "").strip() != ""
            ]
            if doomed:
                raise ExcelLiveError(
                    f"{address}를 병합하면 왼쪽 위({self._idx_to_col(min_col)}{min_row}) 말고 "
                    f"{len(doomed)}칸의 값이 사라집니다: {', '.join(doomed[:4])}"
                    + (" 등" if len(doomed) > 4 else "")
                    + ". 값이 없는 줄을 지목하거나, 먼저 값을 옮겨 주세요."
                )
            ws.merge_cells(address)
            ws.cell(row=min_row, column=min_col, value=keep_value)
            self._save_wb(wb, path, value_changed_sheet=ws.title, changed_rows=set(range(min_row, max_row + 1)))
            return {"merged": True, "address": address}
        finally:
            wb.close()

    def unmerge_cells(self, workbook_id: str | None, sheet_name: str, target_range: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            unmerged = 0
            for merged_range in list(ws.merged_cells.ranges):
                if str(merged_range) == address or CellRange(str(merged_range)).issubset(CellRange(address)):
                    ws.unmerge_cells(str(merged_range))
                    unmerged += 1
            self._save_wb(wb, path)
            return {"unmerged_ranges": unmerged, "address": address}
        finally:
            wb.close()

    def freeze_panes(self, workbook_id: str | None, sheet_name: str, freeze_at: str | None = None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            cell_ref = str(freeze_at or "A2").strip().upper() or "A2"
            if cell_ref in {"NONE", "해제", "OFF"}:
                ws.freeze_panes = None
                self._save_wb(wb, path)
                return {"frozen": False, "freeze_at": None}
            ws.freeze_panes = cell_ref
            self._save_wb(wb, path)
            return {"frozen": True, "freeze_at": cell_ref}
        finally:
            wb.close()

    def autofit_columns(self, workbook_id: str | None, sheet_name: str, target_range: str | None = None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            if target_range and target_range not in {"__USED_RANGE__", "__ACTIVE_SELECTION__"}:
                bounds = self._range_bounds(ws, target_range)
            else:
                max_row, max_col = self._used_bounds(ws)
                bounds = (1, 1, max_row, max_col)
            min_row, min_col, max_row, max_col = bounds
            widths: dict[int, int] = {}
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    text = "" if cell.value is None else str(cell.value)
                    widths[cell.column] = max(widths.get(cell.column, 0), len(text))
            for col_idx, width in widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(8, width + 2))
            self._save_wb(wb, path)
            return {"adjusted_columns": len(widths), "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def define_named_range(
        self, workbook_id: str | None, sheet_name: str, name: str, target_range: str
    ) -> dict[str, Any]:
        clean_name = str(name or "").strip()
        if not clean_name or not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_.]*", clean_name):
            raise ExcelLiveError("define_named_range.name은 문자/밑줄로 시작하는 영문 식별자여야 합니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            ref = f"'{ws.title}'!{self._absolute_ref(address)}"
            if clean_name in wb.defined_names:
                del wb.defined_names[clean_name]
            wb.defined_names[clean_name] = DefinedName(clean_name, attr_text=ref)
            self._save_wb(wb, path)
            return {"name": clean_name, "ref": ref}
        finally:
            wb.close()

    @staticmethod
    def _absolute_ref(address: str) -> str:
        left, _, right = address.partition(":")

        def _abs(part: str) -> str:
            m = re.fullmatch(r"([A-Z]+)(\d+)", part)
            if not m:
                return part
            return f"${m.group(1)}${m.group(2)}"

        return f"{_abs(left)}:{_abs(right)}" if right else _abs(left)

    def set_print_area(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        print_area: str | None = None,
        orientation: str | None = None,
        fit_to_page: bool | None = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            if print_area:
                bounds = self._range_bounds(ws, print_area)
                resolved_area = self._address_from_bounds(bounds)
                ws.print_area = resolved_area
            else:
                resolved_area = ws.print_area
            if orientation:
                normalized = str(orientation).strip().lower()
                if normalized in {"landscape", "가로"}:
                    ws.page_setup.orientation = "landscape"
                elif normalized in {"portrait", "세로"}:
                    ws.page_setup.orientation = "portrait"
            if fit_to_page:
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 1
            self._save_wb(wb, path)
            return {
                "print_area": resolved_area,
                "orientation": ws.page_setup.orientation,
                "fit_to_page": bool(ws.sheet_properties.pageSetUpPr.fitToPage),
            }
        finally:
            wb.close()

    def add_cell_comment(
        self, workbook_id: str | None, sheet_name: str, target_range: str, text: str, author: str = "OfficeClaw AI"
    ) -> dict[str, Any]:
        if not str(text or "").strip():
            raise ExcelLiveError("add_cell_comment.text가 비어 있습니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, _, _ = bounds
            # 메모는 셀 하나에 붙는 주석이라, 범위를 줘도 시작 셀 하나에만 적용한다 —
            # 여러 셀에 같은 문구를 복제하면 사용자가 기대한 "이 셀에 메모"와 달라진다.
            cell = ws.cell(row=min_row, column=min_col)
            cell.comment = Comment(str(text), str(author or "OfficeClaw AI"))
            self._save_wb(wb, path, value_changed_sheet=ws.title, changed_rows={min_row})
            return {"address": cell.coordinate, "comment_added": True}
        finally:
            wb.close()

    def apply_color_scale(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        *,
        min_color: str = "#F8696B",
        mid_color: str = "#FFEB84",
        max_color: str = "#63BE7B",
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            rule = ColorScaleRule(
                start_type="min",
                start_color=self._to_argb(min_color),
                mid_type="percentile",
                mid_value=50,
                mid_color=self._to_argb(mid_color),
                end_type="max",
                end_color=self._to_argb(max_color),
            )
            ws.conditional_formatting.add(address, rule)
            self._save_wb(wb, path)
            return {"address": address, "applied": True, "rule": "color_scale"}
        finally:
            wb.close()

    def apply_data_bar(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        *,
        color: str = "#638EC6",
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            rule = DataBarRule(
                start_type="min",
                end_type="max",
                color=self._to_argb(color),
            )
            ws.conditional_formatting.add(address, rule)
            self._save_wb(wb, path)
            return {"address": address, "applied": True, "rule": "data_bar"}
        finally:
            wb.close()

    @staticmethod
    def _sanitize_table_name(raw: str, existing: set[str]) -> str:
        """Excel 표 이름은 영문으로 시작하고 시트 안에서 유일해야 한다."""
        cleaned = re.sub(r"[^A-Za-z0-9_]", "", str(raw or "")) or "Table1"
        if cleaned[0].isdigit():
            cleaned = f"T{cleaned}"
        if not cleaned[0].isalpha():
            cleaned = f"Table{cleaned}"
        name = cleaned
        index = 1
        existing_cf = {item.casefold() for item in existing}
        while name.casefold() in existing_cf:
            index += 1
            name = f"{cleaned}{index}"
        return name

    def convert_to_excel_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        table_name: str = "",
        has_header: bool = True,
    ) -> dict[str, Any]:
        """이미 있는 데이터 범위를 Excel 표(ListObject)로 바꾼다.

        create_table은 빈 n×m 격자+테두리다. 데모 파일의 SalesTable 같은 진짜 표는 이쪽이다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            existing = set(ws.tables.keys())
            display_name = self._sanitize_table_name(table_name or f"{ws.title}Table", existing)
            table = Table(displayName=display_name, ref=address)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
            self._save_wb(wb, path)
            return {
                "created": True,
                "address": address,
                "table_name": display_name,
                "has_header": bool(has_header),
            }
        finally:
            wb.close()

    def set_font(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        *,
        bold: bool | None = None,
        name: str | None = None,
        size: float | None = None,
        color: str | None = None,
        align: str | None = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            changed = 0
            font_color = self._to_argb(color)[2:] if color else None
            horizontal = _ALIGN_WORDS.get(str(align or "").strip().lower())
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    current = cell.font
                    cell.font = Font(
                        name=name or current.name,
                        size=size if size is not None else current.size,
                        bold=current.bold if bold is None else bool(bold),
                        italic=current.italic,
                        color=font_color or current.color,
                    )
                    if horizontal:
                        # 세로 맞춤·줄바꿈은 사용자가 안 부른 것이므로 그대로 둔다.
                        cur_align = cell.alignment
                        cell.alignment = Alignment(
                            horizontal=horizontal,
                            vertical=cur_align.vertical,
                            wrap_text=cur_align.wrap_text,
                            indent=cur_align.indent,
                        )
                    changed += 1
            self._save_wb(wb, path)
            return {
                "changed_cells": changed,
                "address": self._address_from_bounds(bounds),
                "bold": bold,
                "align": horizontal,
            }
        finally:
            wb.close()

    def apply_formula_cf(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        formula: str,
        fill_color: str = "#FFC7CE",
        font_color: str | None = "#9C0006",
    ) -> dict[str, Any]:
        """값이 바뀌면 다시 평가되는 수식 조건부 서식."""
        formula_text = str(formula or "").strip()
        if formula_text.startswith("="):
            formula_text = formula_text[1:]
        if not formula_text:
            raise ExcelLiveError("apply_formula_cf.formula가 비어 있습니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            fill = PatternFill(
                start_color=self._to_argb(fill_color),
                end_color=self._to_argb(fill_color),
                fill_type="solid",
            )
            font = Font(color=self._to_argb(font_color)[2:]) if font_color else None
            ws.conditional_formatting.add(
                address,
                FormulaRule(formula=[formula_text], fill=fill, font=font),
            )
            self._save_wb(wb, path)
            return {"address": address, "applied": True, "rule": "formula", "formula": formula_text}
        finally:
            wb.close()

    def set_number_format(
        self, workbook_id: str | None, sheet_name: str, target_range: str, format_code: str
    ) -> dict[str, Any]:
        code = str(format_code or "").strip()
        if not code:
            raise ExcelLiveError("set_number_format.format_code가 비어 있습니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            changed = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.number_format = code
                    changed += 1
            self._save_wb(wb, path)
            return {"formatted_cells": changed, "address": self._address_from_bounds(bounds), "format_code": code}
        finally:
            wb.close()

    # ---- 미지원 기능 (앱 런타임 의존) ----

    def refresh_power_query(self, workbook_id: str | None) -> dict[str, Any]:
        raise ExcelLiveError(
            "파일 편집 모드에서는 Power Query 새로고침을 할 수 없습니다. "
            "Excel에서 파일을 연 뒤 다시 시도해 주세요."
        )

    def run_vba_macro(self, workbook_id: str | None, *, macro_name: str, args: list[Any] | None = None) -> dict[str, Any]:
        raise ExcelLiveError(
            "파일 편집 모드에서는 VBA 매크로를 실행할 수 없습니다. "
            "Excel에서 파일을 연 뒤 다시 시도해 주세요."
        )

    def create_chart(
        self,
        workbook_id: str | None,
        sheet_name: str,
        source_range: str,
        chart_type: str = "line",
        title: str | None = None,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        """openpyxl 네이티브 차트를 만든다.

        첫 열을 항목(카테고리), 나머지 숫자 열을 계열로 본다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            min_row, min_col, max_row, max_col = self._range_bounds(ws, source_range)
            # 한 줄짜리 가로 계열("B68:J68로 선 그래프" — 월이 열 방향)도 차트다(2026-08-19 ex16 실측).
            if max_col < min_col or (max_row <= min_row and max_col <= min_col):
                raise ExcelLiveError("차트를 만들 데이터가 부족합니다. 머리글과 데이터 행이 필요합니다.")

            kind = str(chart_type or "line").strip().lower()
            aliases = {
                "선": "line",
                "꺾은선": "line",
                "막대": "bar",
                "column": "bar",
                "원형": "pie",
                "파이": "pie",
                "분산형": "scatter",
                "산점도": "scatter",
                # 참고 대시보드가 목표 달성률을 도넛으로 그린다. 별칭만 있고 구현이 없어
                # 전부 꺾은선으로 떨어졌다(2026-08-16).
                "도넛": "doughnut",
                "donut": "doughnut",
                "링": "doughnut",
                "영역": "area",
                "면적": "area",
            }
            kind = aliases.get(kind, kind)
            if kind == "bar":
                chart = BarChart()
            elif kind == "pie":
                chart = PieChart()
            elif kind == "doughnut":
                chart = DoughnutChart()
            elif kind == "area":
                chart = AreaChart()
            elif kind == "scatter":
                chart = ScatterChart()
            else:
                kind = "line"
                chart = LineChart()
            chart.title = str(title or "데이터 차트")
            if kind not in {"pie", "doughnut"}:
                # varyColors를 비워 두면 Excel이 단일 계열을 점마다 다른 색으로
                # 그리고 범례도 점 단위로 만든다(2026-08-18 렌더 실측).
                chart.varyColors = False

            if min_col == max_col:
                # 값 한 열짜리 범위("B2:B9 데이터로"). 2026-08-18 렌더 실측 결함 셋:
                # 첫 데이터가 계열 제목으로 삼켜져 한 점이 사라졌고, 카테고리가
                # 값과 같은 셀을 가리켰고, 왼쪽 라벨 열은 무시됐다.
                first_value = ws.cell(row=min_row, column=min_col).value
                has_header = isinstance(first_value, str) and bool(first_value.strip())
                data_min_row = min_row + 1 if has_header else min_row
                values = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row)
                chart.add_data(values, titles_from_data=has_header)
                if not has_header:
                    # 범위 바로 위 칸이 머리글이면 계열 이름으로 쓴다 — 아니면
                    # 단일 계열의 "계열1" 범례는 정보가 없으니 치운다.
                    above = ws.cell(row=min_row - 1, column=min_col).value if min_row > 1 else None
                    if isinstance(above, str) and above.strip():
                        header_ref = f"'{ws.title}'!{ws.cell(row=min_row - 1, column=min_col).coordinate}"
                        chart.series[0].tx = SeriesLabel(strRef=StrRef(header_ref))
                    else:
                        chart.legend = None
                label_col = self._find_label_column(ws, min_col, data_min_row, max_row)
                if label_col:
                    chart.set_categories(
                        Reference(ws, min_col=label_col, min_row=data_min_row, max_row=max_row)
                    )
            elif min_row == max_row:
                # 가로 한 줄 계열: 첫 칸이 글자면 계열 이름, 숫자면 왼쪽 칸을 이름으로 쓴다.
                # 항목(카테고리)은 바로 윗줄(예: 월 머리글)이 글자일 때만 잡는다.
                first_value = ws.cell(row=min_row, column=min_col).value
                has_title = isinstance(first_value, str) and bool(first_value.strip())
                data_min_col = min_col + 1 if has_title else min_col
                values = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=min_row)
                chart.add_data(values, from_rows=True, titles_from_data=has_title)
                if not has_title:
                    left = ws.cell(row=min_row, column=min_col - 1).value if min_col > 1 else None
                    if isinstance(left, str) and left.strip():
                        header_ref = f"'{ws.title}'!{ws.cell(row=min_row, column=min_col - 1).coordinate}"
                        chart.series[0].tx = SeriesLabel(strRef=StrRef(header_ref))
                    else:
                        chart.legend = None
                if min_row > 1:
                    above_cells = [ws.cell(row=min_row - 1, column=c).value for c in range(data_min_col, max_col + 1)]
                    if any(v is not None and str(v).strip() for v in above_cells):
                        chart.set_categories(
                            Reference(ws, min_col=data_min_col, max_col=max_col, min_row=min_row - 1, max_row=min_row - 1)
                        )
            else:
                categories = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
                value_min_col = min_col + 1
                if kind in {"pie", "doughnut"}:
                    # 원형·도넛은 계열이 하나여야 한다. 여러 열을 주면 첫 계열만 그려지고
                    # 나머지는 조용히 사라진다.
                    values = Reference(ws, min_col=value_min_col, min_row=min_row, max_row=max_row)
                else:
                    values = Reference(
                        ws,
                        min_col=value_min_col,
                        max_col=max_col,
                        min_row=min_row,
                        max_row=max_row,
                    )
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(categories)

            target_ws = ws
            if output_sheet:
                target_name = self._sanitize_sheet_name(output_sheet)
                target_ws = (
                    wb[target_name] if target_name in wb.sheetnames else wb.create_sheet(target_name)
                )
            anchor = f"{self._idx_to_col(max_col + 2)}{min_row}"
            target_ws.add_chart(chart, anchor)
            self._save_wb(wb, path)
            return {
                "created": True,
                "chart_name": chart.title if isinstance(chart.title, str) else str(title or "데이터 차트"),
                "chart_type": kind,
                "sheet_name": target_ws.title,
                "anchor": anchor,
                "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            }
        finally:
            wb.close()

    def pivot_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        source_range: str,
        row_field: str | int,
        value_field: str | int,
        agg: str = "sum",
        column_field: str | int | None = None,
        output_sheet: str | None = None,
        output_start: str = "A1",
        has_header: bool = True,
    ) -> dict[str, Any]:
        payload = self.read_computed_range(workbook_id, sheet_name, source_range)
        values = payload.get("values", [])
        if not values:
            raise ExcelLiveError("피벗 대상 데이터가 비어 있습니다.")
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        header = normalized[0] if has_header else [f"column_{i+1}" for i in range(col_count)]
        body = normalized[1:] if has_header else normalized
        if not body:
            raise ExcelLiveError("피벗 대상 데이터가 비어 있습니다.")
        row_idx = self._resolve_column_selector(row_field, 1, col_count, header)
        val_idx = self._resolve_column_selector(value_field, 1, col_count, header)
        col_idx = self._resolve_column_selector(column_field, 1, col_count, header) if column_field is not None else None
        agg_name = str(agg or "sum").strip().lower()
        if agg_name not in {"sum", "avg", "count"}:
            agg_name = "sum"
        self._require_computed_values(payload, [row_idx, val_idx, col_idx], header, "집계")

        measures = [_as_number(row[val_idx]) if val_idx < len(row) else None for row in body]
        if agg_name in {"sum", "avg"} and not any(m is not None for m in measures):
            # 값 열이 날짜·문자열이면 더할 수 없다. 0을 채워 조용히 넘기면 사용자가 속는다.
            raise ExcelLiveError(
                f"'{header[val_idx]}' 열은 숫자가 아니라 {agg_name} 집계를 할 수 없습니다. "
                f"(요청 value_field={value_field!r}, row_field={row_field!r})"
            )

        def summarize(numbers: list[float | None]) -> float | int:
            present = [n for n in numbers if n is not None]
            if agg_name == "count":
                return len(present)
            if not present:
                return 0
            total = sum(present)
            return total / len(present) if agg_name == "avg" else total

        if col_idx is None:
            buckets: dict[tuple, list[float | None]] = {}
            labels: dict[tuple, Any] = {}
            for index, row in enumerate(body):
                key = _group_key(row, [row_idx])
                buckets.setdefault(key, []).append(measures[index])
                labels.setdefault(key, row[row_idx] if row_idx < len(row) else None)
            out_rows = [[str(header[row_idx]), f"{agg_name}_{header[val_idx]}"]]
            for key in _sorted_group_keys(list(buckets)):
                out_rows.append([labels[key], summarize(buckets[key])])
        else:
            cells: dict[tuple, dict[tuple, list[float | None]]] = {}
            row_labels: dict[tuple, Any] = {}
            col_labels: dict[tuple, Any] = {}
            for index, row in enumerate(body):
                rkey = _group_key(row, [row_idx])
                ckey = _group_key(row, [col_idx])
                row_labels.setdefault(rkey, row[row_idx] if row_idx < len(row) else None)
                col_labels.setdefault(ckey, row[col_idx] if col_idx < len(row) else None)
                cells.setdefault(rkey, {}).setdefault(ckey, []).append(measures[index])
            ordered_cols = _sorted_group_keys(list(col_labels))
            out_rows = [[str(header[row_idx]), *[col_labels[c] for c in ordered_cols]]]
            for rkey in _sorted_group_keys(list(cells)):
                line: list[Any] = [row_labels[rkey]]
                for ckey in ordered_cols:
                    found = cells[rkey].get(ckey)
                    line.append(summarize(found) if found else 0)
                out_rows.append(line)
        sheet_name_out = output_sheet or sheet_name
        written = self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name_out,
            start_cell=output_start,
            values_2d=out_rows,
        )
        return {
            "created": True,
            "address": written.get("address", output_start),
            "rows": len(out_rows),
            "cols": len(out_rows[0]) if out_rows else 0,
            "sheet_name": sheet_name_out,
            # 집계 전후 행 수. 그룹 기준을 잘못 잡아 "집계가 안 된" 결과를 걸러내는 데 쓴다.
            "source_rows": len(body),
        }

    def validate_data(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        checks: list[str] | None = None,
        has_header: bool = True,
        date_min: str | None = None,
        date_max: str | None = None,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"address": str(payload.get("address", target_range)), "issues": [], "total_issues": 0}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        rows = normalized[1:] if has_header else normalized
        check_set = {str(c).strip().lower() for c in (checks or ["empty", "negative", "outlier"])}
        address = str(payload.get("address", target_range))
        min_col, min_row, _, _ = range_boundaries(address)
        start_row = min_row + (1 if has_header else 0)
        issues: list[dict[str, Any]] = []

        if "empty" in check_set:
            empty_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    if self._is_empty(value):
                        empty_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "empty", "count": len(empty_cells), "samples": empty_cells[:20]})

        if "negative" in check_set:
            negative_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    num = self._as_float(value)
                    if num is not None and num < 0:
                        negative_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "negative", "count": len(negative_cells), "samples": negative_cells[:20]})

        if "outlier" in check_set:
            outlier_cells: list[str] = []
            for c_idx in range(col_count):
                numeric_values: list[tuple[int, float]] = []
                for r_idx, row in enumerate(rows):
                    num = self._as_float(row[c_idx] if c_idx < len(row) else None)
                    if num is not None:
                        numeric_values.append((r_idx, num))
                if len(numeric_values) < 4:
                    continue
                mean = sum(v for _, v in numeric_values) / len(numeric_values)
                variance = sum((v - mean) ** 2 for _, v in numeric_values) / len(numeric_values)
                std = variance ** 0.5
                if std <= 0:
                    continue
                for r_idx, value in numeric_values:
                    z = abs((value - mean) / std)
                    if z >= 3.0:
                        outlier_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "outlier", "count": len(outlier_cells), "samples": outlier_cells[:20]})

        if "date_range" in check_set and (date_min or date_max):
            min_dt = self._parse_datetime_value(date_min) if date_min else None
            max_dt = self._parse_datetime_value(date_max) if date_max else None
            invalid_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    dt = self._parse_datetime_value(value)
                    if dt is None:
                        continue
                    if min_dt and dt < min_dt:
                        invalid_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
                        continue
                    if max_dt and dt > max_dt:
                        invalid_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "date_range", "count": len(invalid_cells), "samples": invalid_cells[:20]})

        total = sum(int(item.get("count", 0) or 0) for item in issues)
        return {"address": address, "issues": issues, "total_issues": total}

    def protect_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        password: str | None = None,
        lock_formula_cells: bool = True,
        unlock_range: str | None = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            ws.protection.sheet = True
            if password:
                ws.protection.set_password(str(password))
            unlocked = ""
            if unlock_range:
                bounds = self._range_bounds(ws, unlock_range)
                min_row, min_col, max_row, max_col = bounds
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        cell.protection = cell.protection.copy(locked=False)
                unlocked = self._address_from_bounds(bounds)
            self._save_wb(wb, path)
            return {
                "protected": True,
                "sheet_name": ws.title,
                "lock_formula_cells": bool(lock_formula_cells),
                "unlock_range": unlocked,
            }
        finally:
            wb.close()

    def find_duplicates(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_columns: list[str | int] | None = None,
        has_header: bool = True,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        """중복을 지우지 않고 "어디가 겹치는지"만 알려준다.

        "중복 주문번호를 찾아줘"에 dedupe를 돌리면 사용자가 확인하기도 전에 행이 사라진다.
        점검과 삭제는 위험도가 달라 별도 도구여야 한다.
        """
        payload = self.read_computed_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        address = str(payload.get("address", target_range))
        if not values:
            return {"duplicate_groups": 0, "duplicate_rows": 0, "address": address, "samples": []}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        if not body:
            return {"duplicate_groups": 0, "duplicate_rows": 0, "address": address, "samples": []}
        if key_columns:
            subset = [self._resolve_column_selector(col, 1, col_count, header) for col in key_columns]
        else:
            subset = list(range(col_count))
        min_col, min_row, _max_col, _max_row = range_boundaries(address)
        body_start = min_row + (1 if has_header else 0)

        occurrences: dict[tuple, list[int]] = {}
        for index, row in enumerate(body):
            occurrences.setdefault(_group_key(row, subset), []).append(index)
        repeated = {key: rows for key, rows in occurrences.items() if len(rows) > 1}
        duplicate_rows = sum(len(rows) for rows in repeated.values())
        samples: list[dict[str, Any]] = []
        group_count = 0
        for key in _sorted_group_keys(list(repeated)):
            group_count += 1
            if len(samples) < 20:
                first_row = body[repeated[key][0]]
                key_text = " / ".join(
                    str(first_row[i]) if i < len(first_row) else "" for i in subset
                )
                samples.append(
                    {
                        "value": key_text,
                        "count": len(repeated[key]),
                        "rows": [body_start + i for i in repeated[key]][:10],
                    }
                )
        result: dict[str, Any] = {
            "duplicate_groups": group_count,
            "duplicate_rows": duplicate_rows,
            "address": address,
            "key_columns": [
                str(header[i]) if header and i < len(header) else self._idx_to_col(min_col + i)
                for i in subset
            ],
            "samples": samples,
        }
        if output_sheet and duplicate_rows:
            report: list[list[Any]] = [["중복값", "건수", "행번호"]]
            for row in samples:
                report.append([row["value"], row["count"], ", ".join(str(r) for r in row["rows"])])
            written = self.write_range(workbook_id, output_sheet, "A1", report)
            result["output_sheet"] = output_sheet
            result["output_address"] = str(written.get("address", "A1"))
        return result

    def recalculate(self, workbook_id: str | None, sheet_name: str | None = None) -> dict[str, Any]:
        """수식 캐시를 무효화해 Excel이 열 때 전부 다시 계산하도록 표시한다.

        파일 편집 모드는 수식을 직접 계산하지 않는다. 대신 full_calc_on_load를 켜두면
        사용자가 파일을 열었을 때 최신 데이터 기준으로 값이 다시 잡힌다.
        """
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            formula_cells = 0
            targets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets
            for ws in targets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula_cells += 1
            self._mark_full_recalc(wb)
            self._save_wb(wb, path)
            return {
                "recalculated": True,
                "formula_cells": formula_cells,
                "sheets": [ws.title for ws in targets],
                "workbook_id": str(path),
            }
        finally:
            wb.close()

    @staticmethod
    def _pdf_target(path: Path, sheet_name: str | None, output_path: str | None) -> Path:
        """PDF를 쓸 위치를 정한다.

        플래너는 "dashboard_report.pdf"처럼 폴더 없는 이름을 곧잘 지어낸다. 그대로 쓰면
        사이드카가 실행된 폴더에 떨어져, 사용자는 "저장했다"는 답만 받고 파일을 못 찾는다.
        상대 경로는 언제나 통합문서 옆으로 되돌린다.
        """
        if output_path:
            requested = Path(output_path)
            return requested if requested.is_absolute() else path.parent / requested.name
        if sheet_name:
            safe = re.sub(r"[^\w가-힣.-]+", "_", sheet_name).strip("_") or "sheet"
            return path.with_name(f"{path.stem}_{safe}.pdf")
        return path.with_suffix(".pdf")

    def export_pdf(
        self,
        workbook_id: str | None,
        sheet_name: str | None = None,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        """시트를 PDF로 내보낸다. Excel이 설치된 Windows에서만 실제 변환이 가능하다."""
        path = self._resolve_workbook_path(workbook_id)
        target = self._pdf_target(path, sheet_name, output_path)
        try:
            import win32com.client  # type: ignore[import-not-found]
        except Exception as exc:  # pragma: no cover - 플랫폼 의존
            raise ExcelLiveError(
                "PDF 내보내기는 Excel이 설치된 Windows에서만 됩니다. "
                "다른 환경에서는 시트를 새 파일로 저장한 뒤 수동으로 내보내 주세요."
            ) from exc

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            book = excel.Workbooks.Open(str(path))
            try:
                source = book.Worksheets(sheet_name) if sheet_name else book
                source.ExportAsFixedFormat(0, str(target))
            finally:
                book.Close(SaveChanges=False)
        except Exception as exc:  # pragma: no cover - 플랫폼 의존
            raise ExcelLiveError(f"PDF 내보내기에 실패했습니다: {exc}") from exc
        finally:
            excel.Quit()
        return {
            "exported": True,
            "pdf_path": str(target),
            "sheet_name": sheet_name or "(전체)",
            "workbook_id": str(path),
        }

    def compare_ranges(
        self,
        workbook_id: str | None,
        *,
        left_sheet: str,
        left_range: str,
        right_sheet: str,
        right_range: str,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        left = self.read_range(workbook_id, left_sheet, left_range)
        right = self.read_range(workbook_id, right_sheet, right_range)
        lv = left.get("values", [])
        rv = right.get("values", [])
        rows = max(len(lv), len(rv))
        cols = max(max((len(row) for row in lv), default=0), max((len(row) for row in rv), default=0))
        diffs: list[list[Any]] = [["row", "col", "left_value", "right_value"]]
        diff_count = 0
        for r in range(rows):
            for c in range(cols):
                lval = lv[r][c] if r < len(lv) and c < len(lv[r]) else None
                rval = rv[r][c] if r < len(rv) and c < len(rv[r]) else None
                if str(lval) != str(rval):
                    diff_count += 1
                    diffs.append([r + 1, c + 1, lval, rval])
        result: dict[str, Any] = {
            "left_address": str(left.get("address", left_range)),
            "right_address": str(right.get("address", right_range)),
            "diff_cells": diff_count,
            "sample_diffs": diffs[1:21],
        }
        if output_sheet:
            written = self.write_range(workbook_id, output_sheet, "A1", diffs)
            result["output_sheet"] = output_sheet
            result["output_address"] = str(written.get("address", "A1"))
        return result

    def forecast_linear(
        self,
        workbook_id: str | None,
        *,
        sheet_name: str,
        source_range: str,
        horizon: int = 3,
        output_sheet: str | None = None,
        output_start: str = "A1",
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, source_range)
        values = payload.get("values", [])
        series: list[float] = []
        for row in values:
            if not row:
                continue
            num = self._as_float(row[-1])
            if num is not None:
                series.append(num)
        if len(series) < 2:
            raise ExcelLiveError("예측을 위해 최소 2개 이상의 숫자 데이터가 필요합니다.")
        n = len(series)
        xs = list(range(1, n + 1))
        sum_x = sum(xs)
        sum_y = sum(series)
        sum_xx = sum(x * x for x in xs)
        sum_xy = sum(x * y for x, y in zip(xs, series))
        denom = (n * sum_xx) - (sum_x * sum_x)
        slope = ((n * sum_xy) - (sum_x * sum_y)) / denom if denom else 0.0
        intercept = (sum_y - slope * sum_x) / n
        horizon_n = max(1, min(36, int(horizon)))
        out_rows: list[list[Any]] = [["step", "forecast"]]
        for step in range(1, horizon_n + 1):
            x = n + step
            out_rows.append([x, slope * x + intercept])
        target_sheet = output_sheet or sheet_name
        written = self.write_range(workbook_id, target_sheet, output_start, out_rows)
        return {
            "created": True,
            "sheet_name": target_sheet,
            "address": str(written.get("address", output_start)),
            "horizon": horizon_n,
            "slope": slope,
            "intercept": intercept,
        }

    def consolidate_sheets(
        self,
        workbook_id: str | None,
        *,
        source_sheets: list[str],
        output_sheet: str = "통합결과",
        include_header_once: bool = True,
        add_source_sheet_col: bool = True,
    ) -> dict[str, Any]:
        if not source_sheets:
            raise ExcelLiveError("source_sheets가 비어 있습니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            merged: list[list[Any]] = []
            header_written = False
            body_rows = 0
            for source in source_sheets:
                ws = self._sheet_or_raise(wb, source)
                used = self._address_from_bounds((1, 1, *self._used_bounds(ws)))
                payload = self.read_range(str(path), source, used)
                values = payload.get("values", [])
                if not values:
                    continue
                header = list(values[0])
                body = values[1:] if len(values) > 1 else []
                if include_header_once and not header_written:
                    merged.append((["source_sheet"] if add_source_sheet_col else []) + header)
                    header_written = True
                elif not include_header_once:
                    merged.append((["source_sheet"] if add_source_sheet_col else []) + header)
                for row in body:
                    merged.append(([source] if add_source_sheet_col else []) + list(row))
                    body_rows += 1
            if not merged:
                raise ExcelLiveError("통합할 데이터가 없습니다.")
            max_cols = max(len(row) for row in merged)
            normalized = [row + [None] * (max_cols - len(row)) for row in merged]
            written = self.write_range(str(path), output_sheet, "A1", normalized)
            return {
                "created": True,
                "sheet_name": output_sheet,
                "address": str(written.get("address", "A1")),
                "rows": len(normalized),
                "cols": max_cols,
                "merged_rows": body_rows,
            }
        finally:
            wb.close()

    def consolidate_workbooks_from_folder(
        self,
        workbook_id: str | None,
        *,
        folder_path: str,
        pattern: str = "*.xlsx",
        source_sheet: str | None = None,
        output_sheet: str = "파일통합결과",
        include_header_once: bool = True,
        add_source_file_col: bool = True,
    ) -> dict[str, Any]:
        root = Path(folder_path).expanduser()
        if not root.exists() or not root.is_dir():
            raise ExcelLiveError(f"유효하지 않은 폴더 경로: {folder_path}")
        files = sorted(root.glob(pattern))
        if not files:
            raise ExcelLiveError("통합할 파일이 없습니다.")
        merged: list[list[Any]] = []
        header_written = False
        merged_rows = 0
        opened_count = 0
        for fp in files:
            if fp.name.startswith("~$"):
                continue
            try:
                wb_in = self._load_wb(fp)
            except Exception:
                continue
            opened_count += 1
            try:
                ws = wb_in[source_sheet] if source_sheet and source_sheet in wb_in.sheetnames else wb_in.active
                used = self._address_from_bounds((1, 1, *self._used_bounds(ws)))
                min_col, min_row, max_col, max_row = range_boundaries(used)
                values: list[list[Any]] = []
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    values.append([cell.value for cell in row])
                if not values:
                    continue
                header = list(values[0])
                body = values[1:] if len(values) > 1 else []
                if include_header_once and not header_written:
                    merged.append((["source_file"] if add_source_file_col else []) + header)
                    header_written = True
                elif not include_header_once:
                    merged.append((["source_file"] if add_source_file_col else []) + header)
                for row in body:
                    merged.append(([fp.name] if add_source_file_col else []) + list(row))
                    merged_rows += 1
            finally:
                wb_in.close()
        if not merged:
            raise ExcelLiveError("파일에서 읽은 데이터가 없습니다.")
        max_cols = max(len(row) for row in merged)
        normalized = [row + [None] * (max_cols - len(row)) for row in merged]
        written = self.write_range(workbook_id, output_sheet, "A1", normalized)
        return {
            "created": True,
            "sheet_name": output_sheet,
            "address": str(written.get("address", "A1")),
            "rows": len(normalized),
            "cols": max_cols,
            "opened_files": opened_count,
            "merged_rows": merged_rows,
        }

    def set_data_validation(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        target_range: str,
        validation_type: str = "list",
        source: str | None = None,
        minimum: float | None = None,
        maximum: float | None = None,
        allow_blank: bool = True,
        show_error: bool = True,
        error_message: str | None = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            vtype = str(validation_type or "list").strip().lower()
            if vtype == "list":
                formula1 = str(source or "").strip()
                if not formula1:
                    raise ExcelLiveError("list 유효성은 source가 필요합니다.")
                dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
            elif vtype in {"whole", "decimal"}:
                if minimum is None or maximum is None:
                    raise ExcelLiveError("숫자 유효성은 minimum/maximum이 필요합니다.")
                dv = DataValidation(
                    type="whole" if vtype == "whole" else "decimal",
                    operator="between",
                    formula1=str(minimum),
                    formula2=str(maximum),
                    allow_blank=allow_blank,
                )
            else:
                raise ExcelLiveError(f"지원하지 않는 validation_type: {validation_type}")
            if error_message:
                dv.error = str(error_message)
            dv.showErrorMessage = bool(show_error)
            ws.add_data_validation(dv)
            dv.add(address)
            self._save_wb(wb, path)
            return {"applied": True, "address": address, "validation_type": vtype}
        finally:
            wb.close()
