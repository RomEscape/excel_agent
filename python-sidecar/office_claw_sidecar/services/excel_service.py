"""
Excel AI service — file parsing, AI analysis, and export helpers.

Supports:
  - .xlsx via openpyxl / pandas
  - .csv  via pandas

All AI calls delegate to the passed-in LLMService instance so the
caller controls which provider (Ollama / Claude) is used.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── Lazy imports — only loaded when actually used to keep startup fast ──────

def _pandas():
    import pandas as pd
    return pd

def _openpyxl():
    import openpyxl
    return openpyxl


# ── File parsing ────────────────────────────────────────────────────────────


def parse_file(file_path: str) -> dict:
    """
    Read an xlsx or csv file and return metadata + a data sample.

    Returns:
        {
            "sheet_names": list[str],
            "active_sheet": str,
            "row_count": int,
            "col_count": int,
            "columns": list[str],
            "dtypes": dict[str, str],   # column_name -> pandas dtype string
            "sample_rows": list[dict],  # first 5 rows as list of dicts
            "numeric_columns": list[str],
        }
    """
    pd = _pandas()
    path = Path(file_path)
    suffix = path.suffix.lower()

    sheet_names: list[str] = []
    active_sheet: str = ""

    if suffix == ".csv":
        # Attempt UTF-8 first (covers UTF-8-SIG/BOM as well), then fall back
        # to EUC-KR which is common in Korean-locale Excel CSV exports.
        try:
            df = pd.read_csv(path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(path, encoding="euc-kr")
            except UnicodeDecodeError:
                df = pd.read_csv(path, encoding="cp949")
        sheet_names = ["Sheet1"]
        active_sheet = "Sheet1"
    elif suffix in (".xlsx", ".xls"):
        xl = pd.ExcelFile(path)
        sheet_names = xl.sheet_names
        active_sheet = sheet_names[0] if sheet_names else ""
        df = pd.read_excel(path, sheet_name=active_sheet)
    else:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {suffix}")

    # Normalise column names to strings (handles integer MultiIndex cols etc.)
    df.columns = [str(c) for c in df.columns]

    row_count = len(df)
    col_count = len(df.columns)
    columns = df.columns.tolist()

    # Dtype mapping — convert numpy dtype to human-readable string
    dtypes: dict[str, str] = {}
    for col in columns:
        dtype = df[col].dtype
        if pd.api.types.is_integer_dtype(dtype):
            dtypes[col] = "정수"
        elif pd.api.types.is_float_dtype(dtype):
            dtypes[col] = "실수"
        elif pd.api.types.is_datetime64_any_dtype(dtype):
            dtypes[col] = "날짜"
        elif pd.api.types.is_bool_dtype(dtype):
            dtypes[col] = "불린"
        else:
            dtypes[col] = "텍스트"

    # Sample — first 5 rows; convert NaN to None for JSON safety
    sample_df = df.head(5).where(pd.notnull(df.head(5)), other=None)
    sample_rows = sample_df.to_dict(orient="records")

    # Numeric columns useful for chart extraction later
    numeric_columns = df.select_dtypes(include="number").columns.tolist()

    return {
        "sheet_names": sheet_names,
        "active_sheet": active_sheet,
        "row_count": row_count,
        "col_count": col_count,
        "columns": columns,
        "dtypes": dtypes,
        "sample_rows": sample_rows,
        "numeric_columns": numeric_columns,
    }


# ── AI helpers ──────────────────────────────────────────────────────────────

_MAX_COLS_IN_PROMPT = 30   # columns beyond this are summarised as "외 N개"
_MAX_CELL_CHARS = 100      # individual cell values truncated to this length


def _truncate_cell(value: Any) -> Any:
    """Truncate string cell values to prevent LLM context overflow."""
    if isinstance(value, str) and len(value) > _MAX_CELL_CHARS:
        return value[:_MAX_CELL_CHARS] + "…"
    return value


def _build_data_summary(meta: dict) -> str:
    """
    Format parsed metadata into a compact text block for LLM context.

    Column count is capped at _MAX_COLS_IN_PROMPT and sample cell values are
    truncated at _MAX_CELL_CHARS characters to avoid exceeding the LLM context
    window (especially important for local Ollama models with 4K–8K limits).
    """
    columns: list[str] = meta["columns"]
    dtypes: dict[str, str] = meta.get("dtypes", {})

    # Cap columns shown in prompt
    shown_cols = columns[:_MAX_COLS_IN_PROMPT]
    omitted = len(columns) - len(shown_cols)

    col_info_lines = [
        f"  - {col} ({dtypes.get(col, '?')})" for col in shown_cols
    ]
    if omitted > 0:
        col_info_lines.append(f"  ... 외 {omitted}개 열 생략")
    col_info = "\n".join(col_info_lines)

    # Truncate sample row values
    truncated_rows = [
        {k: _truncate_cell(v) for k, v in row.items() if k in shown_cols}
        for row in meta["sample_rows"]
    ]
    sample_text = json.dumps(truncated_rows, ensure_ascii=False, indent=2)

    return (
        f"파일 정보:\n"
        f"  시트: {meta['active_sheet']}\n"
        f"  행 수: {meta['row_count']}개\n"
        f"  열 수: {meta['col_count']}개\n\n"
        f"열 목록:\n{col_info}\n\n"
        f"샘플 데이터 (첫 5행):\n{sample_text}"
    )


async def analyze_with_ai(file_path: str, question: str, llm_service) -> str:
    """
    Answer a natural-language question about the spreadsheet data.

    Builds a context prompt that includes the file structure and sample rows,
    then calls llm_service.chat() and returns the markdown answer.
    """
    meta = parse_file(file_path)
    summary = _build_data_summary(meta)

    system_prompt = (
        "당신은 데이터 분석 전문가입니다. "
        "사용자가 제공한 스프레드시트 데이터를 분석하고 "
        "명확하고 친절한 한국어 마크다운으로 답변하세요. "
        "수치 분석, 트렌드, 이상값에 대해 구체적인 인사이트를 제공하세요."
    )
    user_prompt = (
        f"{system_prompt}\n\n"
        f"=== 데이터 요약 ===\n{summary}\n\n"
        f"=== 사용자 질문 ===\n{question}"
    )

    return await llm_service.chat([{"role": "user", "content": user_prompt}])


def _build_describe_stats(file_path: str) -> str:
    """
    Compute actual descriptive statistics using pandas describe() and return
    a compact Korean text block.  This provides accurate aggregate numbers
    (count, mean, std, min, max) regardless of dataset size, supplementing
    the 5-row sample with real statistics.
    """
    pd = _pandas()
    path = Path(file_path)
    try:
        if path.suffix.lower() == ".csv":
            try:
                df = pd.read_csv(path, encoding="utf-8-sig")
            except UnicodeDecodeError:
                try:
                    df = pd.read_csv(path, encoding="euc-kr")
                except UnicodeDecodeError:
                    df = pd.read_csv(path, encoding="cp949")
        else:
            xl = pd.ExcelFile(path)
            df = pd.read_excel(path, sheet_name=xl.sheet_names[0])

        numeric_df = df.select_dtypes(include="number")
        if numeric_df.empty:
            return "수치형 열이 없어 통계를 계산할 수 없습니다."

        desc = numeric_df.describe().round(2)
        lines = ["| 지표 | " + " | ".join(desc.columns) + " |"]
        lines.append("|" + "------|" * (len(desc.columns) + 1))
        label_map = {
            "count": "행 수",
            "mean": "평균",
            "std": "표준편차",
            "min": "최솟값",
            "25%": "1사분위",
            "50%": "중앙값",
            "75%": "3사분위",
            "max": "최댓값",
        }
        for idx in desc.index:
            row_label = label_map.get(str(idx), str(idx))
            row_values = " | ".join(str(v) for v in desc.loc[idx])
            lines.append(f"| {row_label} | {row_values} |")
        return "\n".join(lines)
    except Exception as exc:
        logger.warning("통계 계산 실패: %s", exc)
        return "통계를 계산할 수 없습니다."


async def generate_report(file_path: str, llm_service) -> str:
    """
    Auto-generate a comprehensive Korean markdown data analysis report.

    The report covers: overview, column descriptions, statistics summary,
    key findings, and recommendations.  Actual pandas describe() statistics
    are included so the LLM can reference accurate aggregate numbers even
    for large datasets (not just the 5-row sample).
    """
    meta = parse_file(file_path)
    summary = _build_data_summary(meta)
    stats_table = _build_describe_stats(file_path)

    prompt = (
        "당신은 데이터 분석 전문가입니다. "
        "아래 스프레드시트 데이터를 기반으로 "
        "비개발자 직장인도 이해할 수 있는 종합 분석 리포트를 한국어 마크다운으로 작성하세요.\n\n"
        "리포트 구성:\n"
        "1. **데이터 개요** — 파일 구조, 행/열 수 요약\n"
        "2. **열(컬럼) 설명** — 각 컬럼의 데이터 유형과 의미 추정\n"
        "3. **주요 통계** — 아래 실제 통계 수치를 활용하여 작성\n"
        "4. **핵심 인사이트** — 데이터에서 발견되는 패턴과 특이사항\n"
        "5. **활용 제안** — 이 데이터로 할 수 있는 추가 분석 방향\n\n"
        f"=== 데이터 요약 ===\n{summary}\n\n"
        f"=== 실제 기술 통계 (전체 데이터 기준) ===\n{stats_table}"
    )

    return await llm_service.chat([{"role": "user", "content": prompt}])


async def suggest_formulas(column_info: dict, llm_service) -> str:
    """
    Suggest useful Excel formulas based on the column names and types.

    column_info: { "columns": [...], "dtypes": {...}, "numeric_columns": [...] }
    """
    col_desc = "\n".join(
        f"  - {col}: {column_info.get('dtypes', {}).get(col, '알 수 없음')}"
        for col in column_info.get("columns", [])
    )

    prompt = (
        "당신은 엑셀 전문가입니다. "
        "아래 스프레드시트 열 정보를 분석하여 "
        "실제로 유용한 엑셀 수식 5~10개를 한국어 마크다운으로 제안하세요.\n"
        "각 수식에 대해 목적, 구문 예시, 사용 시나리오를 간략히 설명하세요.\n\n"
        f"=== 열 정보 ===\n{col_desc}"
    )

    return await llm_service.chat([{"role": "user", "content": prompt}])


# ── Chart data extraction ───────────────────────────────────────────────────


def extract_chart_data(file_path: str, sheet_name: str) -> dict:
    """
    Extract numeric columns from the specified sheet as Chart.js-ready JSON.

    Returns:
        {
            "labels": list[str | int],          # row index or first text column
            "datasets": [
                {
                    "label": str,               # column name
                    "data": list[float | None],
                },
                ...
            ]
        }
    """
    pd = _pandas()
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        try:
            df = pd.read_csv(path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(path, encoding="euc-kr")
            except UnicodeDecodeError:
                df = pd.read_csv(path, encoding="cp949")
    else:
        df = pd.read_excel(path, sheet_name=sheet_name)

    df.columns = [str(c) for c in df.columns]

    # Use the first non-numeric column as labels (if available), else row index
    non_numeric = df.select_dtypes(exclude="number").columns.tolist()
    numeric_cols = df.select_dtypes(include="number").columns.tolist()

    if non_numeric:
        label_col = non_numeric[0]
        labels = df[label_col].fillna("").astype(str).tolist()
    else:
        labels = list(range(len(df)))

    datasets = []
    for col in numeric_cols[:8]:  # Cap at 8 series to keep charts readable
        # Convert NaN → None for JSON serialisation
        data = [
            round(float(v), 4) if pd.notna(v) else None
            for v in df[col]
        ]
        datasets.append({"label": col, "data": data})

    return {"labels": labels, "datasets": datasets}


# ── Export ──────────────────────────────────────────────────────────────────


def export_report_to_excel(original_path: str, report_markdown: str) -> str:
    """
    Append a new "AI 분석 리포트" sheet to the original workbook and
    save the result to the same temp directory with a new filename.

    Returns the output file path.
    """
    openpyxl = _openpyxl()
    path = Path(original_path)

    # Load existing workbook (or create fresh if CSV)
    if path.suffix.lower() == ".csv":
        wb = openpyxl.Workbook()
        # Remove default empty sheet so we start clean
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    else:
        wb = openpyxl.load_workbook(path)

    # Remove old report sheet if re-generating
    report_sheet_name = "AI 분석 리포트"
    if report_sheet_name in wb.sheetnames:
        del wb[report_sheet_name]

    ws = wb.create_sheet(title=report_sheet_name)

    # Write markdown content line by line with basic formatting hints
    ws.column_dimensions["A"].width = 100
    for row_idx, line in enumerate(report_markdown.splitlines(), start=1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        # Bold heading lines
        if line.startswith("## ") or line.startswith("# "):
            from openpyxl.styles import Font
            cell.font = Font(bold=True, size=13)
        elif line.startswith("### "):
            from openpyxl.styles import Font
            cell.font = Font(bold=True, size=11)

    # Save next to original with _AI_report + timestamp suffix
    # to avoid overwriting previous exports of the same file.
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = path.parent / f"{path.stem}_AI_report_{timestamp}.xlsx"
    wb.save(str(out_path))
    logger.info("AI 리포트 Excel 파일 저장: %s", out_path)
    return str(out_path)
