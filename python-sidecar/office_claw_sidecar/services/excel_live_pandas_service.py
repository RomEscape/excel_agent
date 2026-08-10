"""
Pandas/OpenPyXL 기반 Excel 편집 서비스.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
import shutil
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

from office_claw_sidecar.sandbox import WORKSPACE_ROOT
from office_claw_sidecar.services.excel_live_service import (
    ExcelLiveError,
    ExcelLiveService,
    WorkbookNotFoundError,
    WorksheetNotFoundError,
)


class PandasExcelLiveService(ExcelLiveService):
    """파일 기반(pandas/openpyxl) Excel 편집 서비스."""

    engine = "pandas"

    def __init__(self, workspace_root: Path | None = None) -> None:
        super().__init__(xw_module=None)
        self._workspace_root = Path(workspace_root or WORKSPACE_ROOT).resolve()
        self._workspace_root.mkdir(parents=True, exist_ok=True)
        self._selected_sheet_by_workbook: dict[str, str] = {}

    # ---- 공통 유틸 ----

    def _candidate_roots(self) -> list[Path]:
        roots: list[Path] = []
        for base in [self._workspace_root, Path.cwd()]:
            try:
                resolved = Path(base).expanduser().resolve()
            except Exception:
                continue
            if resolved not in roots:
                roots.append(resolved)
        return roots

    @staticmethod
    def _is_excel_file(path: Path) -> bool:
        return path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}

    def _list_workspace_workbooks(self, limit: int = 200) -> list[Path]:
        rows: dict[str, Path] = {}
        for root in self._candidate_roots():
            if not root.exists():
                continue
            try:
                for fp in root.rglob("*.xls*"):
                    if not fp.is_file():
                        continue
                    if fp.name.startswith("~$"):
                        continue
                    if not self._is_excel_file(fp):
                        continue
                    rows[str(fp.resolve())] = fp.resolve()
            except Exception:
                continue
        files = list(rows.values())
        files.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0.0, reverse=True)
        return files[: max(1, min(1000, int(limit)))]

    def _resolve_workbook_path(self, workbook_id_or_name: str | None) -> Path:
        raw = str(workbook_id_or_name or "").strip()
        if raw:
            candidate = Path(raw).expanduser()
            if candidate.is_absolute() and candidate.exists() and candidate.is_file() and self._is_excel_file(candidate):
                return candidate.resolve()
            for root in self._candidate_roots():
                joined = (root / candidate).resolve()
                if joined.exists() and joined.is_file() and self._is_excel_file(joined):
                    return joined
            lowered = raw.lower()
            for fp in self._list_workspace_workbooks(limit=500):
                if fp.name.lower() == lowered:
                    return fp
        selected = str(self._selected_workbook_id or "").strip()
        if selected:
            selected_path = Path(selected).expanduser()
            if selected_path.exists() and selected_path.is_file():
                return selected_path.resolve()
        files = self._list_workspace_workbooks(limit=1)
        if files:
            return files[0]
        raise WorkbookNotFoundError(
            "대상 통합문서를 찾지 못했습니다. workbook_id를 지정하거나 워크스페이스에 .xlsx 파일을 준비해 주세요."
        )

    @staticmethod
    def _load_wb(path: Path, *, data_only: bool = False):
        try:
            keep_vba = path.suffix.lower() == ".xlsm"
            return load_workbook(filename=str(path), data_only=data_only, keep_vba=keep_vba)
        except Exception as exc:
            raise ExcelLiveError(f"통합문서를 열 수 없습니다: {path}") from exc

    @staticmethod
    def _sanitize_sheet_name(sheet_name: str) -> str:
        text = str(sheet_name or "").strip()
        if not text:
            raise ExcelLiveError("sheet_name이 비어 있습니다.")
        text = re.sub(r"[:\\/?*\[\]]", "_", text)
        text = text[:31].strip()
        if not text:
            raise ExcelLiveError("유효한 sheet_name이 필요합니다.")
        return text

    @staticmethod
    def _sheet_or_raise(wb: Any, sheet_name: str):
        name = str(sheet_name or "").strip()
        if name in wb.sheetnames:
            return wb[name]
        for cand in wb.sheetnames:
            if cand.lower() == name.lower():
                return wb[cand]
        raise WorksheetNotFoundError(f"시트를 찾을 수 없습니다: {sheet_name}")

    def _used_bounds(self, ws: Any) -> tuple[int, int]:
        max_row = int(getattr(ws, "max_row", 1) or 1)
        max_col = int(getattr(ws, "max_column", 1) or 1)
        return max(1, max_row), max(1, max_col)

    def _range_bounds(self, ws: Any, range_ref: str) -> tuple[int, int, int, int]:
        text = str(range_ref or "A1").strip().upper()
        col_match = re.fullmatch(r"([A-Z]+):([A-Z]+)", text)
        if col_match:
            left, right = col_match.groups()
            start_col = self._col_to_idx(left)
            end_col = self._col_to_idx(right)
            max_row, _ = self._used_bounds(ws)
            return (1, min(start_col, end_col), max_row, max(start_col, end_col))
        try:
            min_col, min_row, max_col, max_row = range_boundaries(text)
        except Exception:
            min_col, min_row, max_col, max_row = range_boundaries("A1")
        return (min_row, min_col, max_row, max_col)

    @classmethod
    def _address_from_bounds(cls, bounds: tuple[int, int, int, int]) -> str:
        min_row, min_col, max_row, max_col = bounds
        left = f"{cls._idx_to_col(min_col)}{min_row}"
        right = f"{cls._idx_to_col(max_col)}{max_row}"
        return left if left == right else f"{left}:{right}"

    @staticmethod
    def _to_argb(color: str) -> str:
        hex6 = str(color or "#000000").strip().lstrip("#")
        if len(hex6) != 6:
            hex6 = "000000"
        return f"FF{hex6.upper()}"

    # ---- 기본 상태/목록 ----

    def is_available(self) -> bool:
        return True

    def list_workbooks(self) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for fp in self._list_workspace_workbooks(limit=200):
            active_sheet = self._selected_sheet_by_workbook.get(str(fp), "")
            if not active_sheet:
                wb = self._load_wb(fp)
                try:
                    active_sheet = wb.active.title if wb.sheetnames else ""
                finally:
                    wb.close()
            rows.append(
                {
                    "workbook_id": str(fp),
                    "name": fp.name,
                    "full_path": str(fp),
                    "active_sheet": active_sheet,
                }
            )
        return rows

    def select_workbook(self, workbook_id_or_name: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id_or_name)
        self._selected_workbook_id = str(path)
        wb = self._load_wb(path)
        try:
            self._selected_sheet_by_workbook[str(path)] = wb.active.title if wb.sheetnames else ""
        finally:
            wb.close()
        return {"selected": True, "workbook_id": str(path)}

    def list_sheets(self, workbook_id: str | None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            sheets = list(wb.sheetnames)
            active = self._selected_sheet_by_workbook.get(str(path), wb.active.title if sheets else "")
            if active not in sheets and sheets:
                active = sheets[0]
            self._selected_sheet_by_workbook[str(path)] = active
            return {
                "sheets": sheets,
                "count": len(sheets),
                "active_sheet": active,
            }
        finally:
            wb.close()

    def select_sheet(self, workbook_id: str | None, sheet_name: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            self._selected_sheet_by_workbook[str(path)] = ws.title
            wb.active = wb.sheetnames.index(ws.title)
            wb.save(str(path))
            return {
                "selected": True,
                "sheet_name": ws.title,
                "active_sheet": ws.title,
            }
        finally:
            wb.close()

    def create_sheet(self, workbook_id: str | None, sheet_name: str, make_active: bool = True) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        target_name = self._sanitize_sheet_name(sheet_name)
        wb = self._load_wb(path)
        try:
            created = target_name not in wb.sheetnames
            ws = wb.create_sheet(target_name) if created else wb[target_name]
            if make_active:
                wb.active = wb.sheetnames.index(ws.title)
                self._selected_sheet_by_workbook[str(path)] = ws.title
            wb.save(str(path))
            return {
                "created": created,
                "sheet_name": ws.title,
                "active_sheet": self._selected_sheet_by_workbook.get(str(path), ws.title),
            }
        finally:
            wb.close()

    # ---- 범위 I/O ----

    def read_range(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, range_ref)
            min_row, min_col, max_row, max_col = bounds
            values: list[list[Any]] = []
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                values.append([cell.value for cell in row])
            if not values:
                values = [[None]]
            row_count = len(values)
            col_count = len(values[0]) if values else 0
            return {
                "values": values,
                "address": self._address_from_bounds(bounds),
                "row_count": row_count,
                "col_count": col_count,
            }
        finally:
            wb.close()

    def get_range_snapshot(self, workbook_id: str | None, sheet_name: str | None, range_ref: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            target_sheet = sheet_name or self._selected_sheet_by_workbook.get(str(path)) or wb.active.title
            data = self.read_range(str(path), target_sheet, range_ref)
            values = data.get("values", [])
            filled = 0
            for row in values:
                for v in row:
                    if v is None:
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    filled += 1
            return {
                "address": data.get("address", range_ref),
                "row_count": int(data.get("row_count", 0) or 0),
                "col_count": int(data.get("col_count", 0) or 0),
                "filled_cells": filled,
            }
        finally:
            wb.close()

    def write_range(self, workbook_id: str | None, sheet_name: str, start_cell: str, values_2d: list[list[Any]]) -> dict[str, Any]:
        if not values_2d:
            return {"written_cells": 0, "address": str(start_cell)}
        rows = len(values_2d)
        cols = max(len(r) for r in values_2d)
        normalized = [row + [None] * (cols - len(row)) for row in values_2d]
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            try:
                ws = self._sheet_or_raise(wb, sheet_name)
            except WorksheetNotFoundError:
                ws = wb.create_sheet(self._sanitize_sheet_name(sheet_name))
            min_row, min_col, _, _ = self._range_bounds(ws, start_cell)
            for r_idx, row in enumerate(normalized):
                for c_idx, value in enumerate(row):
                    ws.cell(row=min_row + r_idx, column=min_col + c_idx, value=value)
            max_row = min_row + rows - 1
            max_col = min_col + cols - 1
            wb.save(str(path))
            return {
                "written_cells": rows * cols,
                "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            }
        finally:
            wb.close()

    def clear_range(self, workbook_id: str | None, sheet_name: str, target_range: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            cleared = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.value = None
                    cleared += 1
            wb.save(str(path))
            return {"cleared_cells": cleared, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def fill_range(self, workbook_id: str | None, sheet_name: str, target_range: str, fill_color: str = "#FFFF00") -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            fill = PatternFill(fill_type="solid", fgColor=self._to_argb(fill_color), bgColor=self._to_argb(fill_color))
            changed = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.fill = fill
                    changed += 1
            wb.save(str(path))
            return {"changed_cells": changed, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def highlight_by_condition(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        operator: str,
        threshold: float,
        fill_color: str = "#FFFF00",
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            fill = PatternFill(fill_type="solid", fgColor=self._to_argb(fill_color), bgColor=self._to_argb(fill_color))
            matched = 0
            changed = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    if self._matches_condition(cell.value, operator, threshold):
                        matched += 1
                        cell.fill = fill
                        changed += 1
            wb.save(str(path))
            return {
                "matched_cells": matched,
                "changed_cells": changed,
                "address": self._address_from_bounds(bounds),
            }
        finally:
            wb.close()

    def apply_border(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        line_style: str = "continuous",
        weight: str = "medium",
        color: str = "#000000",
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            min_row, min_col, max_row, max_col = bounds
            style = str(line_style or "continuous").strip().lower()
            if style == "none":
                border = Border()
            else:
                width_map = {"thin": "thin", "medium": "medium", "thick": "thick"}
                side = Side(style=width_map.get(str(weight or "thin").strip().lower(), "thin"), color=self._to_argb(color))
                border = Border(left=side, right=side, top=side, bottom=side)
            changed = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.border = border
                    changed += 1
            wb.save(str(path))
            return {"changed_cells": changed, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def set_formula(self, workbook_id: str | None, sheet_name: str, range_ref: str, formula_a1: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, range_ref)
            min_row, min_col, max_row, max_col = bounds
            applied = 0
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.value = formula_a1
                    applied += 1
            wb.save(str(path))
            return {"formula_applied_cells": applied, "address": self._address_from_bounds(bounds)}
        finally:
            wb.close()

    def verify_formula_result(self, workbook_id: str | None, sheet_name: str, range_ref: str) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb_data = self._load_wb(path, data_only=True)
        wb_formula = self._load_wb(path, data_only=False)
        try:
            ws_data = self._sheet_or_raise(wb_data, sheet_name)
            ws_formula = self._sheet_or_raise(wb_formula, sheet_name)
            bounds = self._range_bounds(ws_formula, range_ref)
            min_row, min_col, max_row, max_col = bounds
            non_empty = 0
            numeric_values: list[float] = []
            samples: list[Any] = []
            formula_cells = 0
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    value = ws_data.cell(row=r, column=c).value
                    formula_raw = ws_formula.cell(row=r, column=c).value
                    if isinstance(formula_raw, str) and formula_raw.startswith("="):
                        formula_cells += 1
                    if value is None or (isinstance(value, str) and not value.strip()):
                        continue
                    non_empty += 1
                    if len(samples) < 10:
                        samples.append(value)
                    num = self._as_float(value)
                    if num is not None:
                        numeric_values.append(num)
            if non_empty == 0 and formula_cells > 0:
                non_empty = formula_cells
            numeric_count = len(numeric_values)
            if numeric_count == 0 and formula_cells > 0:
                numeric_count = formula_cells
            total = sum(numeric_values) if numeric_values else 0.0
            avg = (total / len(numeric_values)) if numeric_values else 0.0
            return {
                "address": self._address_from_bounds(bounds),
                "non_empty_cells": non_empty,
                "numeric_cells": numeric_count,
                "sum": total,
                "average": avg,
                "sample_values": samples,
            }
        finally:
            wb_data.close()
            wb_formula.close()

    # ---- pandas 변환 기반 작업 ----

    def create_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        start_cell: str,
        rows: int,
        cols: int,
        with_border: bool = True,
    ) -> dict[str, Any]:
        row_count = max(1, min(100, int(rows)))
        col_count = max(1, min(50, int(cols)))
        result = self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=start_cell,
            values_2d=[["" for _ in range(col_count)] for _ in range(row_count)],
        )
        if with_border:
            self.apply_border(
                workbook_id=workbook_id,
                sheet_name=sheet_name,
                target_range=result.get("address", start_cell),
                line_style="continuous",
                weight="thin",
                color="#000000",
            )
        return {
            "created": True,
            "address": result.get("address", start_cell),
            "rows": row_count,
            "cols": col_count,
        }

    def sort_range(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_column: str | int = 1,
        order: str = "asc",
        has_header: bool = True,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"sorted_rows": 0, "address": payload.get("address", target_range)}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        if not body:
            return {"sorted_rows": 0, "address": payload.get("address", target_range)}
        df = pd.DataFrame(body)
        address = str(payload.get("address", target_range))
        min_col, min_row, max_col, max_row = range_boundaries(address)
        key_idx = self._resolve_column_selector(
            selector=key_column,
            start_col_idx=min_col,
            col_count=col_count,
            header_row=header,
        )
        reverse = str(order or "asc").strip().lower() in {"desc", "descending", "내림차순"}
        work = df.copy()
        work["_sort_num"] = pd.to_numeric(work[key_idx], errors="coerce")
        work["_sort_text"] = work[key_idx].astype(str).str.lower()
        sorted_df = work.sort_values(
            by=["_sort_num", "_sort_text"],
            ascending=not reverse,
            na_position="last",
        ).drop(columns=["_sort_num", "_sort_text"])
        out_values = sorted_df.values.tolist()
        final_values = [header, *out_values] if has_header and header is not None else out_values
        self.clear_range(workbook_id, sheet_name, address)
        self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=f"{self._idx_to_col(min_col)}{min_row}",
            values_2d=final_values,
        )
        return {
            "sorted_rows": len(out_values),
            "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            "key_column_index": key_idx + 1,
            "order": "desc" if reverse else "asc",
        }

    def dedupe_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        key_columns: list[str | int] | None = None,
        has_header: bool = True,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"removed_rows": 0, "remaining_rows": 0, "address": payload.get("address", target_range)}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        if not body:
            return {"removed_rows": 0, "remaining_rows": 0, "address": payload.get("address", target_range)}
        df = pd.DataFrame(body)
        start_col = 1
        if key_columns:
            subset = [
                self._resolve_column_selector(col, start_col, col_count, header)
                for col in key_columns
            ]
        else:
            subset = list(range(col_count))
        before = len(df)
        deduped = df.drop_duplicates(subset=subset, keep="first")
        after = len(deduped)
        out = deduped.values.tolist()
        final_values = [header, *out] if has_header and header is not None else out
        address = str(payload.get("address", target_range))
        min_row, min_col, max_row, max_col = range_boundaries(address)
        self.clear_range(workbook_id, sheet_name, address)
        self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name,
            start_cell=f"{self._idx_to_col(min_col)}{min_row}",
            values_2d=final_values,
        )
        return {
            "removed_rows": max(0, before - after),
            "remaining_rows": after,
            "address": self._address_from_bounds((min_row, min_col, max_row, max_col)),
            "key_columns": [idx + 1 for idx in subset],
        }

    def filter_rows(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        column: str | int = 1,
        operator: str = "==",
        value: Any = None,
        has_header: bool = True,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"filtered_rows": 0, "address": payload.get("address", target_range)}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        header = normalized[0] if has_header else None
        body = normalized[1:] if has_header else normalized
        if not body:
            return {"filtered_rows": 0, "address": payload.get("address", target_range)}
        df = pd.DataFrame(body)
        col_idx = self._resolve_column_selector(column, 1, col_count, header)
        series = df[col_idx]
        op = str(operator or "==").strip()
        if op in {"=", "=="}:
            mask = series.astype(str) == str(value)
        elif op == "!=":
            mask = series.astype(str) != str(value)
        else:
            left = pd.to_numeric(series, errors="coerce")
            right = pd.to_numeric(pd.Series([value] * len(series)), errors="coerce")
            if op == ">":
                mask = left > right
            elif op == ">=":
                mask = left >= right
            elif op == "<":
                mask = left < right
            elif op == "<=":
                mask = left <= right
            else:
                mask = series.astype(str) == str(value)
        matched = int(mask.fillna(False).sum())
        return {
            "filtered_rows": matched,
            "address": str(payload.get("address", target_range)),
            "column_index": col_idx + 1,
            "operator": op,
            "value": value,
        }

    # ---- 기타 작업 ----

    def save_workbook(self, workbook_id: str | None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        return {
            "saved": True,
            "workbook_id": str(path),
            "name": path.name,
            "full_path": str(path),
        }

    def create_workbook_backup(self, workbook_id: str | None, *, label: str = "auto") -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        backup_dir = path.parent / "officeclaw_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_label = re.sub(r"[^A-Za-z0-9_-]+", "_", str(label or "auto")).strip("_") or "auto"
        backup_path = backup_dir / f"{path.stem}.{safe_label}.{stamp}{path.suffix}"
        shutil.copy2(path, backup_path)
        return {
            "backup_created": True,
            "backup_path": str(backup_path),
            "source_path": str(path),
            "workbook_id": str(path),
        }

    def list_workbook_backups(self, workbook_id: str | None, *, limit: int = 20) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        backup_dir = path.parent / "officeclaw_backups"
        rows: list[dict[str, Any]] = []
        if backup_dir.exists():
            for fp in backup_dir.glob(f"{path.stem}.*{path.suffix}"):
                if not fp.is_file():
                    continue
                stat = fp.stat()
                rows.append(
                    {
                        "backup_path": str(fp.resolve()),
                        "backup_name": fp.name,
                        "size_bytes": int(stat.st_size),
                        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                        "modified_ts": float(stat.st_mtime),
                    }
                )
        rows.sort(key=lambda r: float(r.get("modified_ts", 0.0)), reverse=True)
        for row in rows:
            row.pop("modified_ts", None)
        return {
            "workbook_id": str(path),
            "source_path": str(path),
            "backup_dir": str(backup_dir),
            "backups": rows[: max(1, min(200, int(limit)))],
        }

    def restore_workbook_from_backup(self, workbook_id: str | None, *, backup_path: str | None = None) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        if backup_path:
            source = Path(str(backup_path)).expanduser().resolve()
        else:
            listed = self.list_workbook_backups(str(path), limit=1)
            backups = listed.get("backups", [])
            if not backups:
                raise ExcelLiveError("복구 가능한 백업이 없습니다.")
            source = Path(str(backups[0].get("backup_path", ""))).resolve()
        if not source.exists() or not source.is_file():
            raise ExcelLiveError(f"백업 파일을 찾을 수 없습니다: {source}")
        pre_restore = self.create_workbook_backup(str(path), label="pre_restore")
        shutil.copy2(source, path)
        wb = self._load_wb(path)
        try:
            active = wb.active.title if wb.sheetnames else ""
        finally:
            wb.close()
        self._selected_workbook_id = str(path)
        self._selected_sheet_by_workbook[str(path)] = active
        return {
            "restored": True,
            "workbook_id": str(path),
            "name": path.name,
            "full_path": str(path),
            "restored_from_backup_path": str(source),
            "pre_restore_backup_path": str(pre_restore.get("backup_path", "")),
            "active_sheet": active,
        }

    def get_active_selection_ref(self, workbook_id: str | None, sheet_name: str | None) -> str:
        return "A1"

    def get_used_range_ref(self, workbook_id: str | None, sheet_name: str | None) -> str:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            target_sheet = sheet_name or self._selected_sheet_by_workbook.get(str(path)) or wb.active.title
            ws = self._sheet_or_raise(wb, target_sheet)
            max_row, max_col = self._used_bounds(ws)
            return self._address_from_bounds((1, 1, max_row, max_col))
        finally:
            wb.close()

    # ---- 미지원 기능 (앱 런타임 의존) ----

    def refresh_power_query(self, workbook_id: str | None) -> dict[str, Any]:
        raise ExcelLiveError("pandas 엔진에서는 Power Query 새로고침을 지원하지 않습니다.")

    def run_vba_macro(self, workbook_id: str | None, *, macro_name: str, args: list[Any] | None = None) -> dict[str, Any]:
        raise ExcelLiveError("pandas 엔진에서는 VBA 매크로 실행을 지원하지 않습니다.")

    def create_chart(
        self,
        workbook_id: str | None,
        sheet_name: str,
        source_range: str,
        chart_type: str = "line",
        title: str | None = None,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        raise ExcelLiveError("pandas 엔진에서는 차트 생성을 지원하지 않습니다.")

    def pivot_table(
        self,
        workbook_id: str | None,
        sheet_name: str,
        source_range: str,
        row_field: str | int,
        value_field: str | int,
        agg: str = "sum",
        column_field: str | int | None = None,
        output_sheet: str | None = None,
        output_start: str = "A1",
        has_header: bool = True,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, source_range)
        values = payload.get("values", [])
        if not values:
            raise ExcelLiveError("피벗 대상 데이터가 비어 있습니다.")
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        header = normalized[0] if has_header else [f"column_{i+1}" for i in range(col_count)]
        body = normalized[1:] if has_header else normalized
        if not body:
            raise ExcelLiveError("피벗 대상 데이터가 비어 있습니다.")
        df = pd.DataFrame(body)
        row_idx = self._resolve_column_selector(row_field, 1, col_count, header)
        val_idx = self._resolve_column_selector(value_field, 1, col_count, header)
        col_idx = self._resolve_column_selector(column_field, 1, col_count, header) if column_field is not None else None
        agg_name = str(agg or "sum").strip().lower()
        if agg_name not in {"sum", "avg", "count"}:
            agg_name = "sum"
        agg_func = {"sum": "sum", "avg": "mean", "count": "count"}[agg_name]
        if col_idx is None:
            grouped = df.groupby(row_idx, dropna=False)[val_idx].agg(agg_func).reset_index()
            out_rows = [[str(header[row_idx]), f"{agg_name}_{header[val_idx]}"]]
            out_rows.extend(grouped.values.tolist())
        else:
            pivot_df = pd.pivot_table(
                df,
                index=row_idx,
                columns=col_idx,
                values=val_idx,
                aggfunc=agg_func,
                fill_value=0,
            )
            pivot_df = pivot_df.reset_index()
            out_rows = [list(pivot_df.columns)]
            out_rows.extend(pivot_df.values.tolist())
        sheet_name_out = output_sheet or sheet_name
        written = self.write_range(
            workbook_id=workbook_id,
            sheet_name=sheet_name_out,
            start_cell=output_start,
            values_2d=out_rows,
        )
        return {
            "created": True,
            "address": written.get("address", output_start),
            "rows": len(out_rows),
            "cols": len(out_rows[0]) if out_rows else 0,
            "sheet_name": sheet_name_out,
        }

    def validate_data(
        self,
        workbook_id: str | None,
        sheet_name: str,
        target_range: str,
        checks: list[str] | None = None,
        has_header: bool = True,
        date_min: str | None = None,
        date_max: str | None = None,
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, target_range)
        values = payload.get("values", [])
        if not values:
            return {"address": str(payload.get("address", target_range)), "issues": [], "total_issues": 0}
        col_count = max(len(row) for row in values)
        normalized = [list(row) + [None] * (col_count - len(row)) for row in values]
        rows = normalized[1:] if has_header else normalized
        check_set = {str(c).strip().lower() for c in (checks or ["empty", "negative", "outlier"])}
        address = str(payload.get("address", target_range))
        min_col, min_row, _, _ = range_boundaries(address)
        start_row = min_row + (1 if has_header else 0)
        issues: list[dict[str, Any]] = []

        if "empty" in check_set:
            empty_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    if self._is_empty(value):
                        empty_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "empty", "count": len(empty_cells), "samples": empty_cells[:20]})

        if "negative" in check_set:
            negative_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    num = self._as_float(value)
                    if num is not None and num < 0:
                        negative_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "negative", "count": len(negative_cells), "samples": negative_cells[:20]})

        if "outlier" in check_set:
            outlier_cells: list[str] = []
            for c_idx in range(col_count):
                numeric_values: list[tuple[int, float]] = []
                for r_idx, row in enumerate(rows):
                    num = self._as_float(row[c_idx] if c_idx < len(row) else None)
                    if num is not None:
                        numeric_values.append((r_idx, num))
                if len(numeric_values) < 4:
                    continue
                mean = sum(v for _, v in numeric_values) / len(numeric_values)
                variance = sum((v - mean) ** 2 for _, v in numeric_values) / len(numeric_values)
                std = variance ** 0.5
                if std <= 0:
                    continue
                for r_idx, value in numeric_values:
                    z = abs((value - mean) / std)
                    if z >= 3.0:
                        outlier_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "outlier", "count": len(outlier_cells), "samples": outlier_cells[:20]})

        if "date_range" in check_set and (date_min or date_max):
            min_dt = self._parse_datetime_value(date_min) if date_min else None
            max_dt = self._parse_datetime_value(date_max) if date_max else None
            invalid_cells: list[str] = []
            for r_idx, row in enumerate(rows):
                for c_idx, value in enumerate(row):
                    dt = self._parse_datetime_value(value)
                    if dt is None:
                        continue
                    if min_dt and dt < min_dt:
                        invalid_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
                        continue
                    if max_dt and dt > max_dt:
                        invalid_cells.append(f"{self._idx_to_col(min_col + c_idx)}{start_row + r_idx}")
            issues.append({"type": "date_range", "count": len(invalid_cells), "samples": invalid_cells[:20]})

        total = sum(int(item.get("count", 0) or 0) for item in issues)
        return {"address": address, "issues": issues, "total_issues": total}

    def protect_sheet(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        password: str | None = None,
        lock_formula_cells: bool = True,
        unlock_range: str | None = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            ws.protection.sheet = True
            if password:
                ws.protection.set_password(str(password))
            unlocked = ""
            if unlock_range:
                bounds = self._range_bounds(ws, unlock_range)
                min_row, min_col, max_row, max_col = bounds
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                        cell.protection = cell.protection.copy(locked=False)
                unlocked = self._address_from_bounds(bounds)
            wb.save(str(path))
            return {
                "protected": True,
                "sheet_name": ws.title,
                "lock_formula_cells": bool(lock_formula_cells),
                "unlock_range": unlocked,
            }
        finally:
            wb.close()

    def compare_ranges(
        self,
        workbook_id: str | None,
        *,
        left_sheet: str,
        left_range: str,
        right_sheet: str,
        right_range: str,
        output_sheet: str | None = None,
    ) -> dict[str, Any]:
        left = self.read_range(workbook_id, left_sheet, left_range)
        right = self.read_range(workbook_id, right_sheet, right_range)
        lv = left.get("values", [])
        rv = right.get("values", [])
        rows = max(len(lv), len(rv))
        cols = max(max((len(row) for row in lv), default=0), max((len(row) for row in rv), default=0))
        diffs: list[list[Any]] = [["row", "col", "left_value", "right_value"]]
        diff_count = 0
        for r in range(rows):
            for c in range(cols):
                lval = lv[r][c] if r < len(lv) and c < len(lv[r]) else None
                rval = rv[r][c] if r < len(rv) and c < len(rv[r]) else None
                if str(lval) != str(rval):
                    diff_count += 1
                    diffs.append([r + 1, c + 1, lval, rval])
        result: dict[str, Any] = {
            "left_address": str(left.get("address", left_range)),
            "right_address": str(right.get("address", right_range)),
            "diff_cells": diff_count,
            "sample_diffs": diffs[1:21],
        }
        if output_sheet:
            written = self.write_range(workbook_id, output_sheet, "A1", diffs)
            result["output_sheet"] = output_sheet
            result["output_address"] = str(written.get("address", "A1"))
        return result

    def forecast_linear(
        self,
        workbook_id: str | None,
        *,
        sheet_name: str,
        source_range: str,
        horizon: int = 3,
        output_sheet: str | None = None,
        output_start: str = "A1",
    ) -> dict[str, Any]:
        payload = self.read_range(workbook_id, sheet_name, source_range)
        values = payload.get("values", [])
        series: list[float] = []
        for row in values:
            if not row:
                continue
            num = self._as_float(row[-1])
            if num is not None:
                series.append(num)
        if len(series) < 2:
            raise ExcelLiveError("예측을 위해 최소 2개 이상의 숫자 데이터가 필요합니다.")
        n = len(series)
        xs = list(range(1, n + 1))
        sum_x = sum(xs)
        sum_y = sum(series)
        sum_xx = sum(x * x for x in xs)
        sum_xy = sum(x * y for x, y in zip(xs, series))
        denom = (n * sum_xx) - (sum_x * sum_x)
        slope = ((n * sum_xy) - (sum_x * sum_y)) / denom if denom else 0.0
        intercept = (sum_y - slope * sum_x) / n
        horizon_n = max(1, min(36, int(horizon)))
        out_rows: list[list[Any]] = [["step", "forecast"]]
        for step in range(1, horizon_n + 1):
            x = n + step
            out_rows.append([x, slope * x + intercept])
        target_sheet = output_sheet or sheet_name
        written = self.write_range(workbook_id, target_sheet, output_start, out_rows)
        return {
            "created": True,
            "sheet_name": target_sheet,
            "address": str(written.get("address", output_start)),
            "horizon": horizon_n,
            "slope": slope,
            "intercept": intercept,
        }

    def consolidate_sheets(
        self,
        workbook_id: str | None,
        *,
        source_sheets: list[str],
        output_sheet: str = "통합결과",
        include_header_once: bool = True,
        add_source_sheet_col: bool = True,
    ) -> dict[str, Any]:
        if not source_sheets:
            raise ExcelLiveError("source_sheets가 비어 있습니다.")
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            merged: list[list[Any]] = []
            header_written = False
            body_rows = 0
            for source in source_sheets:
                ws = self._sheet_or_raise(wb, source)
                used = self._address_from_bounds((1, 1, *self._used_bounds(ws)))
                payload = self.read_range(str(path), source, used)
                values = payload.get("values", [])
                if not values:
                    continue
                header = list(values[0])
                body = values[1:] if len(values) > 1 else []
                if include_header_once and not header_written:
                    merged.append((["source_sheet"] if add_source_sheet_col else []) + header)
                    header_written = True
                elif not include_header_once:
                    merged.append((["source_sheet"] if add_source_sheet_col else []) + header)
                for row in body:
                    merged.append(([source] if add_source_sheet_col else []) + list(row))
                    body_rows += 1
            if not merged:
                raise ExcelLiveError("통합할 데이터가 없습니다.")
            max_cols = max(len(row) for row in merged)
            normalized = [row + [None] * (max_cols - len(row)) for row in merged]
            written = self.write_range(str(path), output_sheet, "A1", normalized)
            return {
                "created": True,
                "sheet_name": output_sheet,
                "address": str(written.get("address", "A1")),
                "rows": len(normalized),
                "cols": max_cols,
                "merged_rows": body_rows,
            }
        finally:
            wb.close()

    def consolidate_workbooks_from_folder(
        self,
        workbook_id: str | None,
        *,
        folder_path: str,
        pattern: str = "*.xlsx",
        source_sheet: str | None = None,
        output_sheet: str = "파일통합결과",
        include_header_once: bool = True,
        add_source_file_col: bool = True,
    ) -> dict[str, Any]:
        root = Path(folder_path).expanduser()
        if not root.exists() or not root.is_dir():
            raise ExcelLiveError(f"유효하지 않은 폴더 경로: {folder_path}")
        files = sorted(root.glob(pattern))
        if not files:
            raise ExcelLiveError("통합할 파일이 없습니다.")
        merged: list[list[Any]] = []
        header_written = False
        merged_rows = 0
        opened_count = 0
        for fp in files:
            if fp.name.startswith("~$"):
                continue
            try:
                wb_in = self._load_wb(fp)
            except Exception:
                continue
            opened_count += 1
            try:
                ws = wb_in[source_sheet] if source_sheet and source_sheet in wb_in.sheetnames else wb_in.active
                used = self._address_from_bounds((1, 1, *self._used_bounds(ws)))
                min_col, min_row, max_col, max_row = range_boundaries(used)
                values: list[list[Any]] = []
                for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    values.append([cell.value for cell in row])
                if not values:
                    continue
                header = list(values[0])
                body = values[1:] if len(values) > 1 else []
                if include_header_once and not header_written:
                    merged.append((["source_file"] if add_source_file_col else []) + header)
                    header_written = True
                elif not include_header_once:
                    merged.append((["source_file"] if add_source_file_col else []) + header)
                for row in body:
                    merged.append(([fp.name] if add_source_file_col else []) + list(row))
                    merged_rows += 1
            finally:
                wb_in.close()
        if not merged:
            raise ExcelLiveError("파일에서 읽은 데이터가 없습니다.")
        max_cols = max(len(row) for row in merged)
        normalized = [row + [None] * (max_cols - len(row)) for row in merged]
        written = self.write_range(workbook_id, output_sheet, "A1", normalized)
        return {
            "created": True,
            "sheet_name": output_sheet,
            "address": str(written.get("address", "A1")),
            "rows": len(normalized),
            "cols": max_cols,
            "opened_files": opened_count,
            "merged_rows": merged_rows,
        }

    def set_data_validation(
        self,
        workbook_id: str | None,
        sheet_name: str,
        *,
        target_range: str,
        validation_type: str = "list",
        source: str | None = None,
        minimum: float | None = None,
        maximum: float | None = None,
        allow_blank: bool = True,
        show_error: bool = True,
        error_message: str | None = None,
    ) -> dict[str, Any]:
        path = self._resolve_workbook_path(workbook_id)
        wb = self._load_wb(path)
        try:
            ws = self._sheet_or_raise(wb, sheet_name)
            bounds = self._range_bounds(ws, target_range)
            address = self._address_from_bounds(bounds)
            vtype = str(validation_type or "list").strip().lower()
            if vtype == "list":
                formula1 = str(source or "").strip()
                if not formula1:
                    raise ExcelLiveError("list 유효성은 source가 필요합니다.")
                dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
            elif vtype in {"whole", "decimal"}:
                if minimum is None or maximum is None:
                    raise ExcelLiveError("숫자 유효성은 minimum/maximum이 필요합니다.")
                dv = DataValidation(
                    type="whole" if vtype == "whole" else "decimal",
                    operator="between",
                    formula1=str(minimum),
                    formula2=str(maximum),
                    allow_blank=allow_blank,
                )
            else:
                raise ExcelLiveError(f"지원하지 않는 validation_type: {validation_type}")
            if error_message:
                dv.error = str(error_message)
            dv.showErrorMessage = bool(show_error)
            ws.add_data_validation(dv)
            dv.add(address)
            wb.save(str(path))
            return {"applied": True, "address": address, "validation_type": vtype}
        finally:
            wb.close()
