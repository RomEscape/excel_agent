"""붙여넣은·선택한 범위의 열별 집계를 바로 아랫줄에 쓴다 — 사람 말투 경로.

2026-08-18 사용자 실측: "컨트롤 c, 컨트롤 v 한 위치에 있는 모든 합을 밑에 있는
시트에 기록할 수 있게 해줘" — 좌표도, SUM이라는 낱말도, 수식도 없다. 사람은
의도를 말하지 수식을 부르지 않는다. 이 문형을 결정적으로 매핑한다:

    대상 범위(문장 범위 → context_range → 살아 있는 선택)를 읽어,
    숫자가 있는 열마다 =FUNC(열구간)을 범위 바로 아랫줄에 넣고,
    글자 열(라벨 열)이 있으면 같은 줄에 "합계" 같은 이름표를 쓴다.

판정은 순수 함수로 두고 값 읽기는 호출부(라우터)가 한다 — 엔진 없이 테스트된다.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter

# 함수 어휘. 맨 뒤의 SUM 항목은 "합쳐줘"(병합)·"통합"·"합니다"를 피하려고
# 조사·수식어가 붙은 꼴만 받는다.
_FUNC_VOCAB: list[tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"평균", re.IGNORECASE), "AVERAGE", "평균"),
    (re.compile(r"개수|몇\s*개|카운트|count", re.IGNORECASE), "COUNT", "개수"),
    (re.compile(r"최대|최댓값", re.IGNORECASE), "MAX", "최대"),
    (re.compile(r"최소|최솟값", re.IGNORECASE), "MIN", "최소"),
    (
        # "합 좀 밑에다", "각 열 합" 같은 반말·군더더기 꼴을 포함한다
        # (2026-08-18 사람 말투 배터리 1라운드 실측: 두 꼴 다 되묻기로 샜다).
        re.compile(
            r"(?:모든|각|전체)\s*(?:열|칸|항목)?\s*합|합계|총합|총\s*합"
            r"|합(?:을|이|만|과|값|\s+좀|이나)|더한\s*값|다\s*더해",
            re.IGNORECASE,
        ),
        "SUM",
        "합계",
    ),
]
# "그 다음 줄엔 평균값 한 줄 더" — 사람은 '아래'만큼이나 '다음 줄'이라고 말한다
# (2026-08-20 게이트5 avg_below 2건이 이 어휘 하나로 규칙 밖에 있었다).
_BELOW = re.compile(r"밑|아래|하단|아랫|다음\s*(?:줄|행|칸|라인)|담\s*줄")
_RANGE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$")


def match_aggregate_below(message: str) -> tuple[str, str] | None:
    """(엑셀 함수, 한국어 이름표)를 돌려준다. 이 문형이 아니면 None."""
    text = str(message or "")
    if not _BELOW.search(text):
        return None
    if "=" in text:
        # 수식을 직접 쓴 문장은 기존 수식 경로가 맡는다.
        return None
    for pattern, func, label in _FUNC_VOCAB:
        if pattern.search(text):
            return func, label
    return None


_COLUMNWISE = re.compile(r"열\s*별|열별|각\s*열|칼럼\s*별|컬럼\s*별")


def match_aggregate_columns(message: str) -> tuple[str, str] | None:
    """"열 별로 합계를 만들어줘" — 아래·밑 낱말 없이 열별 집계를 말하는 문형.

    대상 줄(붙여넣은 A7:F7 등)이 따로 있으면 방향 낱말이 필요 없다.
    2026-08-18 GUI 실측: 이 문형이 플래너로 새서 pivot_table로 오분류돼
    검증 실패·재계획 실패로 끝났다.
    """
    text = str(message or "")
    if "=" in text or not _COLUMNWISE.search(text):
        return None
    for pattern, func, label in _FUNC_VOCAB:
        if pattern.search(text):
            return func, label
    return None


# "A4에 지역성과 시트 주문건수 합계를 가져와줘" — 셀·원본 시트·열 이름·집계만
# 말하는 크로스시트 문형. "E4에는 …를, F4에는 …를"처럼 한 문장에 여러 절이
# 오면 시트 이름은 앞 절에서 이어받는다.
# 집계어 — "합계"만이 아니라 사람이 쓰는 동의어까지(2026-08-19 블라인드 게이트: "다 더한 값", "총합",
# "전부 더해서", "합"으로 18/24가 새어 텍스트가 써졌다).
_AGG_WORD = r"(합계|총합계|총합|총계|합산|평균|개수|합|다\s*더한\s*값|전부\s*더한\s*값|더한\s*값|전부\s*더해서|다\s*더해서|더해서)"
# ① 셀이 앞: "A2에 지역성과 시트 주문건수 합계 가져와"
# 머리글은 "총 임대료(원)" · "매출 비중(%)" · "평균 대기시간(분)"처럼 공백·괄호·기호를 품는다.
# 한 낱말만 허용하면 이런 머리글이 통째로 안 잡혀 계획이 비고, 플래너가 텍스트를 쓴다(2026-08-19 실측).
_HEADER_TOKEN = r"([가-힣A-Za-z0-9_%][가-힣A-Za-z0-9_%()\[\]/·.\- ]{0,38}?)"

_CROSS_SHEET = re.compile(
    r"([A-Z]+\d+)\s*(?:셀|칸)?\s*(?:에(?:는|다가|다)?)?\s*(?:([^\s,]+?)\s*(?:시트|탭)(?:의|에서|에\s*있는|안에\s*있는|안의|에)?\s*)?"
    # 시트 낱말 없이 이름만 온 경우("A2에 지역성과 주문건수 합계") — 알려진 시트 이름일 때만 채택한다.
    r"(?:([^\s,]+?)\s+)?"
    + _HEADER_TOKEN
    + r"\s*(?:의|을|를)?\s*"
    + _AGG_WORD,
    re.IGNORECASE,
)
# ② 셀이 뒤: "지역성과 시트 주문건수 합계를 A2로 끌어와" / "지역성과 주문건수 전부 더해서 A2에 넣어"
_CROSS_SHEET_CELL_LAST = re.compile(
    r"(?:([^\s,]+?)\s*(?:시트|탭)(?:의|에서|에\s*있는|안에\s*있는|안의|에)?\s*)?"
    + _HEADER_TOKEN
    + r"\s*(?:의|을|를)?\s*"
    + _AGG_WORD
    + r"\s*(?:을|를|값을|값)?\s*(?:여기\s*)?([A-Z]+\d+)\s*(?:셀|칸)?\s*(?:에|로|으로|에다가?|다)?",
    re.IGNORECASE,
)
# 머리글이 될 수 없는 수량·정도 부사. 이게 머리글 자리에 오면 진짜 머리글은 앞 토큰이다.
_QUANTIFIER_ONLY = re.compile(r"(?:다|전부|모두|싹|전|모|총|일괄|죄다|몽땅|통째로?)")
_CROSS_SHEET_VERB = re.compile(r"가져|끌어|연결|수식|넣어|채워|기록|불러|참조|더해|계산|놔|놓아|써|입력")
_TOTAL_ROW_LABELS = frozenset({"합계", "총계", "계", "총합", "평균", "최대", "최소", "개수", "total", "sum", "avg", "average"})


def _norm_header(text: str) -> str:
    """머리글 비교용 정규화 — 단위 꼬리·공백·대소문자를 지운다."""
    out = re.sub(r"\s*[(\[（][^)\]）]*[)\]）]\s*$", "", str(text or "").strip())
    return re.sub(r"\s+", "", out).casefold()


def _match_header(word: str, headers: list[str]) -> int | None:
    """사람이 부른 머리글을 실제 머리글 목록에서 찾는다.

    "예산"과 "예산(원)", "총 임대료(원)"과 "총임대료(원)"이 같은 것을 가리키는데
    정확 일치만 보면 계획이 비고, 그러면 플래너가 텍스트를 쓴다(2026-08-19 실측).
    """
    target = str(word or "").strip()
    if not target:
        return None
    if target in headers:
        return headers.index(target)
    norm = _norm_header(target)
    if not norm:
        return None
    normalized = [_norm_header(h) for h in headers]
    if norm in normalized:
        return normalized.index(norm)
    # 부분 일치는 **유일할 때만** — 여러 머리글에 걸리면 추측이 된다.
    hits = [i for i, h in enumerate(normalized) if h and (h.startswith(norm) or norm.startswith(h))]
    return hits[0] if len(hits) == 1 else None


def build_cross_sheet_aggregate_plan(
    message: str,
    sheet_reader,
    sheet_names: list[str] | None = None,
) -> list[dict[str, Any]]:
    """크로스시트 집계 수식 계획. sheet_reader(시트명) → (사용범위, values_2d).

    원본 시트의 머리글에서 열을 찾아 =FUNC('시트'!열구간)을 만든다. 마지막 행이
    합계 줄이면 구간에서 뺀다 — 넣으면 이중 집계다. 2026-08-18 사람 말투 실측:
    이 문형이 의도 정규화로 새서 빈 값을 쓰고 성공으로 보고됐다(가짜 성공).
    """
    text = str(message or "")
    known = [str(n) for n in (sheet_names or []) if str(n).strip()]
    # "시트/탭"이라는 말이 없어도 실제 시트 이름이 문장에 있으면 크로스시트다("A2 지역성과 주문건수 합").
    named = any(n in text for n in known)
    mentions_sheet = ("시트" in text or "탭" in text) or named
    # 동사 없는 짧은 말("A2 지역성과 주문건수 합")도 시트 이름·셀이 다 있으면 크로스시트다.
    if not mentions_sheet or not (_CROSS_SHEET_VERB.search(text) or (named and re.search(r"[A-Z]+\d+", text, re.IGNORECASE))):
        return []
    steps: list[dict[str, Any]] = []
    current_sheet = ""
    cache: dict[str, tuple[str, list[list[Any]]]] = {}
    matches: list[tuple[str, str, str, str]] = []
    for m in _CROSS_SHEET.finditer(text):
        sheet_tok = (m.group(2) or "").strip().strip("'\"")
        bare = (m.group(3) or "").strip().strip("'\"")
        header_tok = m.group(4).strip()
        if not sheet_tok and bare:
            if bare in known:
                sheet_tok = bare
            else:
                # 시트가 아닌 낱말이면 머리글의 앞부분이다("평균 운행시간" 같은 두 낱말 머리글).
                header_tok = f"{bare} {header_tok}".strip()
        elif bare:
            # 앞 토큰이 머리글의 일부인지("총 임대료(원)") 수량 부사인지("결석 **다** 더한 값")는
            # 문장만 봐서는 못 가른다 — '총'은 둘 다다. **추측하지 말고 실제 머리글에 물어본다**:
            # 두 후보를 다 넘기고, 시트의 머리글 목록에 있는 쪽을 고르게 한다
            # (2026-08-19 5라운드 감사: 못 가려서 계획이 비었고 플래너가 '=' 없는 문자열을 셀에 썼다).
            if _QUANTIFIER_ONLY.fullmatch(header_tok):
                header_tok = bare
            else:
                header_tok = [f"{bare} {header_tok}".strip(), header_tok]
        matches.append((m.group(1).upper(), sheet_tok, header_tok, m.group(5)))
    if not matches:
        for m in _CROSS_SHEET_CELL_LAST.finditer(text):
            matches.append((m.group(4).upper(), (m.group(1) or "").strip().strip("'\""), m.group(2).strip(), m.group(3)))
    for cell, sheet, header_word, func_word in matches:
        if not sheet and known:
            # 시트 낱말 없이 이름만 쓴 경우: 문장 안의 알려진 시트 이름을 잡는다.
            for name in sorted(known, key=len, reverse=True):
                if name and name in text:
                    sheet = name
                    break
        if sheet:
            current_sheet = sheet
        if not current_sheet:
            continue
        # 후보가 여럿일 수 있다(위 "총 임대료(원)" 갈림). 아래 매칭이 실제 머리글로 가른다.
        candidates = list(header_word) if isinstance(header_word, list) else [header_word]
        # 시트 이름이 머리글 낱말 앞에 붙어 들어온 경우("지역성과 주문건수" → 머리글 '주문건수')
        candidates = [
            c[len(current_sheet):].strip()
            if current_sheet and c.startswith(current_sheet) and len(c) > len(current_sheet)
            else c
            for c in candidates
        ]
        func = "AVERAGE" if "평균" in func_word else ("COUNT" if ("개수" in func_word or "건수" in func_word) else "SUM")
        try:
            if current_sheet not in cache:
                cache[current_sheet] = sheet_reader(current_sheet)
            used_ref, values = cache[current_sheet]
        except Exception:
            return []
        rng = _RANGE.match(str(used_ref or "").strip().upper())
        if not rng or not values:
            continue
        headers = [str(v).strip() if v is not None else "" for v in values[0]]
        header_idx = next(
            (idx for idx in (_match_header(c, headers) for c in candidates) if idx is not None), None
        )
        if header_idx is None:
            continue
        col_letter = get_column_letter(column_index_from_string(rng.group(1)) + header_idx)
        data_start = int(rng.group(2)) + 1
        data_end = int(rng.group(4))
        # 꼬리의 집계 줄(합계·평균 등 이름표 또는 수식 줄)은 **전부** 구간에서
        # 뺀다 — 한 줄만 빼면 합계+평균 두 줄일 때 이중 집계가 된다(2026-08-18
        # 대화형 러너 실측: =SUM('지역성과'!B2:B7)).
        idx = len(values) - 1
        while idx >= 1 and data_end > data_start:
            row = values[idx] or []
            label = row[0] if row else None
            has_formula = any(isinstance(v, str) and str(v).startswith("=") for v in row)
            is_agg = isinstance(label, str) and label.strip().lower() in _TOTAL_ROW_LABELS
            if not (has_formula or is_agg):
                break
            data_end -= 1
            idx -= 1
        if data_end < data_start:
            continue
        steps.append(
            {
                "action": "excel_live.set_formula",
                "params": {
                    "range_ref": cell,
                    "formula_a1": f"={func}('{current_sheet}'!{col_letter}{data_start}:{col_letter}{data_end})",
                },
                "reason": f"{current_sheet} 시트 {header_word} {func_word}를 {cell}에",
            }
        )
    return steps


def build_aggregate_below_plan(
    func: str,
    label: str,
    target_range: str,
    values_2d: list[list[Any]] | None,
) -> list[dict[str, Any]]:
    """숫자 열마다 집계 수식을, 라벨 열에 이름표를 — 범위 바로 아랫줄에.

    values_2d가 비거나 숫자 열이 없으면 빈 계획을 돌려준다(호출부가 되묻는다).
    """
    m = _RANGE.match(str(target_range or "").strip().upper())
    if not m:
        return []
    start_letter, start_row, end_letter, end_row = (
        m.group(1),
        int(m.group(2)),
        m.group(3),
        int(m.group(4)),
    )
    rows = values_2d or []
    if not rows:
        return []
    start_idx = column_index_from_string(start_letter)
    end_idx = column_index_from_string(end_letter)

    def _cell(r: int, c: int) -> Any:
        row = rows[r] if r < len(rows) else []
        return row[c] if isinstance(row, list) and c < len(row) else None

    # 첫 행이 전부 글자(또는 빈 칸)면 머리글이다 — 집계 구간에서 뺀다.
    n_cols = end_idx - start_idx + 1
    first = [_cell(0, c) for c in range(n_cols)]
    has_header = len(rows) > 1 and all(
        v is None or isinstance(v, str) for v in first
    ) and any(isinstance(v, str) and v.strip() for v in first)
    data_start = start_row + 1 if has_header else start_row
    below = end_row + 1

    steps: list[dict[str, Any]] = []
    label_letter: str | None = None
    for offset in range(n_cols):
        letter = get_column_letter(start_idx + offset)
        col_values = [
            _cell(r, offset)
            for r in range(data_start - start_row, end_row - start_row + 1)
        ]
        numeric = [
            v for v in col_values if isinstance(v, (int, float)) and not isinstance(v, bool)
        ]
        if numeric:
            steps.append(
                {
                    "action": "excel_live.set_formula",
                    "params": {
                        "range_ref": f"{letter}{below}",
                        "formula_a1": f"={func}({letter}{data_start}:{letter}{end_row})",
                    },
                    "reason": f"{letter}열 {label}을 범위 아랫줄에",
                }
            )
        elif label_letter is None:
            label_letter = letter
    if not steps:
        return []
    if label_letter is not None:
        steps.append(
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": f"{label_letter}{below}", "values_2d": [[label]]},
                "reason": "집계 줄 이름표",
            }
        )
    return steps
