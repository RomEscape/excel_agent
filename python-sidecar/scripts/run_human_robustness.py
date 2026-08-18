# 사람 말투 강건성 배터리 — 정제 안 된 문장 44개를 실행하고 파일로 검증한다.
# 각 케이스는 씨앗 데이터가 든 새 통합문서에서 돌고, 성공 판정은 액션이 아니라
# 파일 상태다. 공통 불변식: A1 머리글("지역")이 문장 텍스트로 덮이면 즉시 실패.
import asyncio
import json
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.routers.excel_live import (
    ApprovalResponse,
    ExcelLiveCommandRequest,
    _run_command,
    post_approval,
)
from office_claw_sidecar.services.excel_live_service import invalidate_excel_engine_cache
from office_claw_sidecar.services.llm_service import get_llm_service

WS = Path(r"C:\Users\asdjj\AppData\Local\office_claw\Workspace")
WB = WS / "강건성_probe.xlsx"

SEED = [
    ["지역", "주문건수", "출고건수", "정시배송률", "지연건수", "클레임"],
    ["수도권", 10452, 10158, 97.1, 145, 12],
    ["충청권", 3892, 3773, 95.2, 89, 6],
    ["호남권", 3214, 3086, 94.7, 112, 5],
    ["영남권", 6789, 6512, 95.8, 174, 5],
    ["강원제주", 2495, 2383, 92.6, 145, 0],
]


def _ws(wb):
    return wb["지역성과"]


def has_formula(wb, cell, frag):
    v = _ws(wb)[cell].value
    return isinstance(v, str) and v.startswith("=") and frag in v


def fill_of(wb, cell):
    c = _ws(wb)[cell]
    return c.fill.fgColor.rgb if c.fill and c.fill.fill_type else None


def chart_count(wb):
    return len(_ws(wb)._charts or [])


def cf_count(wb):
    return len(_ws(wb).conditional_formatting)


# (계열, 문장, ctx, 검사 함수(wb, result)->오류문자열 or "")
def _ok(_wb, _r):
    return ""


CASES = [
    # ── 합계/집계 (사람이 실제로 말하는 꼴) ──
    ("합계", "여기다가 합 좀 밑에다 적어줄래?", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("합계", "밑에 합계 한줄 부탁해", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("합계", "이 표 아래로 각 열 합 넣어주라", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("합계", "아래쪽에 총합 좀 계산해서 넣어줘", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("합계", "합계행 하나 만들어서 표 밑에 붙여줘", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("합계", "H1에 주문건수 다 더한 값 좀 넣어줘", None,
     lambda wb, r: "" if has_formula(wb, "H1", "SUM") or isinstance(_ws(wb)["H1"].value, (int, float)) else "H1 비어있음"),
    # ── 서식 (머리글 색) ──
    ("서식", "첫줄 있잖아 그거 남색 배경으로 하고 글자는 흰색 굵게 부탁", None,
     lambda wb, r: "" if fill_of(wb, "A1") == "FF002060" else f"A1 채움={fill_of(wb, 'A1')}"),
    ("서식", "헤더를 좀 진하게 해줘 배경은 남색으로다가", None,
     lambda wb, r: "" if fill_of(wb, "A1") == "FF002060" else f"A1 채움={fill_of(wb, 'A1')}"),
    ("서식", "A1:F1 남색으로 칠해주고 글씨 하양 볼드", None,
     lambda wb, r: "" if fill_of(wb, "A1") == "FF002060" else f"A1 채움={fill_of(wb, 'A1')}"),
    ("서식", "제목행 배경 남색 글자 흰색으로 바꿔줄 수 있어?", None,
     lambda wb, r: "" if fill_of(wb, "A1") == "FF002060" else f"A1 채움={fill_of(wb, 'A1')}"),
    ("서식", "1행 배경색만 남색으로 살짝 바꿔줘", None,
     lambda wb, r: "" if fill_of(wb, "A1") == "FF002060" else f"A1 채움={fill_of(wb, 'A1')}"),
    # ── 숫자 서식 ──
    ("콤마", "주문건수랑 출고건수 숫자에 콤마좀 찍어줘야 보기 편할듯", None,
     lambda wb, r: "" if "#,##0" in _ws(wb)["B2"].number_format else f"B2 형식={_ws(wb)['B2'].number_format}"),
    ("콤마", "B열 C열 천단위 콤마 넣어주라", None,
     lambda wb, r: "" if "#,##0" in _ws(wb)["B2"].number_format else f"B2 형식={_ws(wb)['B2'].number_format}"),
    ("콤마", "주문건수 열 숫자 형식 좀 이쁘게 콤마로", None,
     lambda wb, r: "" if "#,##0" in _ws(wb)["B2"].number_format else f"B2 형식={_ws(wb)['B2'].number_format}"),
    # ── 차트 ──
    ("차트", "정시배송률 가지고 선그래프 하나 뽑아줘", None,
     lambda wb, r: "" if chart_count(wb) >= 1 else "차트 없음"),
    ("차트", "지연건수로 막대차트 그려주라", None,
     lambda wb, r: "" if chart_count(wb) >= 1 else "차트 없음"),
    ("차트", "클레임 비중 도넛으로 보여줘", None,
     lambda wb, r: "" if chart_count(wb) >= 1 else "차트 없음"),
    ("차트", "주문건수 추이 그래프 하나 그려줄래?", None,
     lambda wb, r: "" if chart_count(wb) >= 1 else "차트 없음"),
    # ── 조건 강조 ──
    ("강조", "클레임 10 넘는 데만 빨갛게 칠해줘", None,
     lambda wb, r: "" if (fill_of(wb, "F2") == "FFFF0000" or cf_count(wb) >= 1) else "F2(12) 강조 없음"),
    ("강조", "지연건수 100 이상인 셀 노란색으로 표시좀", None,
     lambda wb, r: "" if (fill_of(wb, "E2") == "FFFFFF00" or cf_count(wb) >= 1) else "E2(145) 강조 없음"),
    ("강조", "정시배송률 95 밑으로 떨어진 데 빨간색", None,
     lambda wb, r: "" if (fill_of(wb, "D6") == "FFFF0000" or cf_count(wb) >= 1) else "D6(92.6) 강조 없음"),
    ("강조", "클레임 0인 데는 초록색으로 해줘", None,
     lambda wb, r: "" if (fill_of(wb, "F6") in {"FF00FF00", "FF00B050", "FF6AC36A"} or cf_count(wb) >= 1) else "F6(0) 강조 없음"),
    # ── 테두리/정리 ──
    ("정리", "표 전체 테두리좀 둘러줘", None,
     lambda wb, r: "" if _ws(wb)["A1"].border.top.style else "테두리 없음"),
    ("정리", "이거 다 필요없고 싹 지워줘", "A1:F6",
     lambda wb, r: "" if _ws(wb)["B2"].value is None else "B2 안 지워짐"),
    ("정리", "테두리만 싹 없애줘", None, _ok),
    ("정리", "열너비 자동으로 맞춰줘 보기좋게", None, _ok),
    ("정리", "틀고정 첫줄로 해줘", None,
     lambda wb, r: "" if _ws(wb).freeze_panes == "A2" else f"틀고정={_ws(wb).freeze_panes}"),
    # ── 정렬/필터 ──
    ("정렬", "주문건수 많은 순으로 정렬해줘", None,
     lambda wb, r: "" if _ws(wb)["A2"].value == "수도권" else f"A2={_ws(wb)['A2'].value}"),
    ("정렬", "클레임 기준으로 내림차순 정렬 부탁", None,
     lambda wb, r: "" if _ws(wb)["A2"].value == "수도권" else f"A2={_ws(wb)['A2'].value}"),
    ("정렬", "정시배송률 낮은 순서대로 줄세워줘", None,
     lambda wb, r: "" if _ws(wb)["A2"].value == "강원제주" else f"A2={_ws(wb)['A2'].value}"),
    ("정렬", "수도권 행만 남기고 나머지는 치워줘", None, _ok),
    # ── 쓰기 ──
    ("쓰기", "여기에 서울,100; 부산,200 넣어줘", "A8:B9",
     lambda wb, r: "" if _ws(wb)["A8"].value == "서울" and _ws(wb)["B9"].value == 200 else f"A8={_ws(wb)['A8'].value}"),
    ("쓰기", "A8에 비고 라고 적어줘", None,
     lambda wb, r: "" if _ws(wb)["A8"].value == "비고" else f"A8={_ws(wb)['A8'].value}"),
    ("쓰기", "G1:G6에 메모,1,2,3,4,5 입력해주라", None,
     lambda wb, r: "" if _ws(wb)["G1"].value == "메모" else f"G1={_ws(wb)['G1'].value}"),
    ("쓰기", "수도권이라는 단어를 서울권으로 바꿔줘", None,
     lambda wb, r: "" if _ws(wb)["A2"].value == "서울권" else f"A2={_ws(wb)['A2'].value}"),
    ("쓰기", "H2에 =AVERAGE(D2:D6) 넣어줘", None,
     lambda wb, r: "" if has_formula(wb, "H2", "AVERAGE") else "H2 수식 없음"),
    # ── 시트 ──
    ("시트", "새로운 시트 하나 파줘 이름은 백업으로", None,
     lambda wb, r: "" if "백업" in wb.sheetnames else f"시트들={wb.sheetnames}"),
    ("시트", "백업2 시트를 만들어줄래?", None,
     lambda wb, r: "" if "백업2" in wb.sheetnames else f"시트들={wb.sheetnames}"),
    ("시트", "요약이라는 이름으로 시트 추가좀", None,
     lambda wb, r: "" if "요약" in wb.sheetnames else f"시트들={wb.sheetnames}"),
    ("시트", "이 시트 복사본 하나 만들어줘", None, _ok, True),  # 시트 복사 액션 부재 — 되묻기가 정답
    # ── 오타·흘려쓰기 (사람의 실수 — 알아듣거나, 정직하게 되묻거나) ──
    ("오타", "합계 좀 밑에 너어줘", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("오타", "함계를 표 아래 한줄로 만들어조", "A1:F6",
     lambda wb, r: "" if has_formula(wb, "B7", "SUM") else "B7에 SUM 없음"),
    ("오타", "머리글 남색으로 해조", None,
     lambda wb, r: "" if fill_of(wb, "A1") == "FF002060" else f"A1 채움={fill_of(wb, 'A1')}"),
    ("오타", "정열 좀 해줘 주문건수 많은 순으로", None,
     lambda wb, r: "" if _ws(wb)["A2"].value == "수도권" else f"A2={_ws(wb)['A2'].value}"),
    ("오타", "테두르 둘러줘 표 전체에", None,
     lambda wb, r: "" if _ws(wb)["A1"].border.top.style else "테두리 없음"),
    ("오타", "지연건수로 막대 차투 그러줘", None,
     lambda wb, r: "" if chart_count(wb) >= 1 else "차트 없음"),
    ("오타", "차트 다 지어줘", None, _ok),  # 지어→지워 정규화 후 차트 삭제로 결정적
    ("오타", "저 함계행 마지막에 추가해줄수잇어?", "A1:F6", _ok, True),
]


async def run_case(idx, family, text, ctx, check, llm, allow_ask=False):
    if WB.exists():
        WB.unlink()
    wb = Workbook()
    ws = wb.active
    ws.title = "지역성과"
    for row in SEED:
        ws.append(row)
    wb.save(WB)
    wb.close()
    invalidate_excel_engine_cache()
    # GUI 조건 그대로: workbook_id 없이 "선택된 통합문서"로 돈다(2026-08-18 실측 —
    # id를 명시한 러너들은 전부 통과했는데 GUI만 실패했던 사각지대).
    from office_claw_sidecar.services.excel_live_service import get_excel_live_service
    get_excel_live_service().select_workbook(str(WB))

    req = ExcelLiveCommandRequest(
        message=text, session_id=f"test-robust-{idx}", workbook_id=None, approve=False
    )
    if ctx:
        req = req.model_copy(update={"context_range": ctx})
    t0 = time.time()
    try:
        resp = await _run_command(req, llm)
        if getattr(resp, "approval_required", False):
            resp = await post_approval(
                ApprovalResponse(approval_id=resp.pending_approval.approval_id, approved=True), llm
            )
    except Exception as exc:
        return {"family": family, "text": text, "ok": False, "why": f"예외 {type(exc).__name__}", "secs": round(time.time() - t0, 1)}

    result = getattr(resp, "result", None) or {}
    asked = bool(result.get("ask_follow_up"))
    action = str(getattr(resp, "action", ""))
    entry = {
        "family": family, "text": text, "action": action,
        "reply": (result.get("follow_up_question") or result.get("execution_report") or str(getattr(resp, "reason", "")))[:90],
        "secs": round(time.time() - t0, 1),
    }
    if asked:
        if allow_ask:
            entry.update(ok=True, why="(정당한 되묻기)")
        else:
            entry.update(ok=False, why="되묻기")
        return entry

    wb2 = load_workbook(WB)
    try:
        # 공통 불변식: 머리글 A1은 '지역' 그대로여야 한다(전체 삭제 계열만 예외).
        a1 = _ws(wb2)["A1"].value
        if not (a1 == "지역" or (family == "정리" and a1 is None)):
            entry.update(ok=False, why=f"A1 오염: {repr(a1)[:30]}")
            return entry
        why = check(wb2, result)
    except Exception as exc:
        why = f"검사 예외 {type(exc).__name__}"
    finally:
        wb2.close()
    entry.update(ok=not why, why=why)
    return entry


async def main():
    import os
    repeat = max(1, int(os.environ.get("ROBUST_REPEAT", "1")))
    llm = get_llm_service()
    out = []
    for i, case in enumerate(CASES, 1):
        family, text, ctx, check = case[:4]
        allow_ask = bool(case[4]) if len(case) > 4 else False
        e = await run_case(i, family, text, ctx, check, llm, allow_ask=allow_ask)
        # "한 번 되면 되는 게 아니라 여러 번 100%"(사용자) — 같은 문장을 반복해
        # 결과가 흔들리면 그 자체가 실패다.
        for _ in range(repeat - 1):
            again = await run_case(i, family, text, ctx, check, llm, allow_ask=allow_ask)
            if again["ok"] != e["ok"] or again.get("action") != e.get("action"):
                e = dict(e, ok=False, why=f"비결정: {e.get('action')} vs {again.get('action')}")
                break
        flag = "OK  " if e["ok"] else "FAIL"
        print(f"[{i:2d}] {flag} [{e['family']}] {text[:34]:36s} {e.get('action','')[:28]:30s} {e.get('why','')[:40]}")
        out.append(e)
    ok = sum(1 for e in out if e["ok"])
    print(f"\n합계 {ok}/{len(out)}")
    by = {}
    for e in out:
        by.setdefault(e["family"], [0, 0])
        by[e["family"]][1] += 1
        by[e["family"]][0] += 1 if e["ok"] else 0
    for f, (o, t) in by.items():
        print(f"  {f:4s} {o}/{t}")
    Path(__file__).with_name("human_robust_log.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    if WB.exists():
        WB.unlink()


asyncio.run(main())
