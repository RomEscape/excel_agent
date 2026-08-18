"""대화형 재현 러너 — 사람 질문 ↔ 모델 답변을 전부 기록하며 끝까지 완주한다.

각본 JSON: {"workbook_name": ..., "turns": [{"zone", "command", "expect"?}]}
  expect: "ok"(기본) — 실행돼야 함 / "ask" — 되묻기·해석 카드가 나와야 함(사람의
  실수·모호한 질문 턴). "ask" 다음 턴은 사람의 답변이다.

GUI와 같은 요청 형태(workbook_id 없음, 승인은 approval_id 재개)로 돌고, 판정은
파일 상태(공통 불변식: 첫 시트 A1이 문장 텍스트로 오염되면 실패)와 기대 일치다.
"""

from __future__ import annotations

import asyncio
import json
import os
import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.routers.excel_live import (
    ApprovalResponse,
    ExcelLiveCommandRequest,
    _run_command,
    post_approval,
)
from office_claw_sidecar.services.excel_live_service import (
    get_excel_live_service,
    invalidate_excel_engine_cache,
)
from office_claw_sidecar.services.llm_service import get_llm_service

ROUTER = Path(
    r"c:/Users/asdjj/OneDrive/바탕 화면/동화책 프로젝트/officeclaw"
    r"/python-sidecar/office_claw_sidecar/routers/excel_live.py"
)
_router_src = ROUTER.read_text(encoding="utf-8")
EXECUTABLE = set(re.findall(r'action == "(excel_live\.[a-z_]+)"', _router_src))
EXECUTABLE |= set(re.findall(r'"(excel_live\.[a-z_]+)"', _router_src))

SCENARIO = Path(sys.argv[1])
SC = json.loads(SCENARIO.read_text(encoding="utf-8"))
WB = Path(r"C:/Users/asdjj/AppData/Local/office_claw/Workspace") / SC["workbook_name"]
SESSION_BASE = "test-dialogue-" + SCENARIO.stem
TURNS = SC["turns"]


def _reply_text(resp, result, approval_summary: str) -> str:
    """사용자가 화면에서 실제로 읽는 문구 하나."""
    if result.get("ask_follow_up"):
        return str(result.get("follow_up_question") or getattr(resp, "reason", ""))
    if result.get("execution_report"):
        return str(result["execution_report"])
    if approval_summary:
        return approval_summary
    return str(getattr(resp, "reason", ""))


async def run_once(round_no: int) -> list[dict]:
    if WB.exists():
        WB.unlink()
    Workbook().save(WB)
    print(f"[라운드 {round_no}] {WB.name} · {len(TURNS)}턴")
    llm = get_llm_service()
    session = f"{SESSION_BASE}-r{round_no}"
    log: list[dict] = []

    for i, turn in enumerate(TURNS, 1):
        text = turn["command"]
        zone = turn.get("zone", "")
        expect = turn.get("expect", "ok")
        invalidate_excel_engine_cache()
        get_excel_live_service().select_workbook(str(WB))
        t0 = time.time()
        approval_summary = ""
        interpretation = False
        try:
            resp = await _run_command(
                ExcelLiveCommandRequest(
                    message=text, session_id=session, workbook_id=None, approve=False
                ),
                llm,
            )
            if getattr(resp, "approval_required", False):
                pending = getattr(resp, "pending_approval", None)
                approval_summary = str(getattr(pending, "summary", "") or "")
                interpretation = bool(getattr(pending, "interpretation", False))
                resp = await post_approval(
                    ApprovalResponse(approval_id=pending.approval_id, approved=True), llm
                )
        except Exception as exc:
            log.append({"turn": i, "zone": zone, "text": text, "expect": expect, "ok": False,
                        "why": f"예외 {type(exc).__name__}", "reply": str(exc)[:200]})
            print(f"[{i:3d}] 예외  {text[:44]}  {type(exc).__name__}")
            continue

        result = getattr(resp, "result", None) or {}
        action = str(getattr(resp, "action", ""))
        asked = bool(result.get("ask_follow_up")) or "clarify" in action
        executed = bool(getattr(resp, "ok", False)) and not asked and action in EXECUTABLE
        if expect == "ask":
            good = asked
            why = "" if good else "되물었어야 함"
        elif expect == "noop":
            # 부정문("아직 저장하지 마")은 실행되면 안 된다 — 되묻기든 무동작이든 OK.
            good = asked or action in {"excel_live.noop", "excel_live.clarify", "excel_live.not_excel"} or bool(result.get("noop"))
            why = "" if good else "실행돼 버림"
        else:
            good = executed
            why = "" if good else ("되묻기" if asked else "미지원")
        reply = _reply_text(resp, result, approval_summary)
        entry = {
            "turn": i, "zone": zone, "text": text, "expect": expect, "action": action,
            "ok": good, "why": why, "asked": asked, "interpretation": interpretation,
            "reply": reply[:600], "secs": round(time.time() - t0, 1),
        }
        log.append(entry)
        flag = "OK  " if good else "FAIL"
        tag = " (질문)" if asked else (" (해석카드)" if interpretation else "")
        print(f"[{i:3d}] {flag}{tag} {text[:46]}")
        if not good:
            print(f"          action={action} {why} · {reply[:80]}")

    ok = sum(1 for t in log if t["ok"])
    print(f"성공 {ok} / {len(TURNS)}")
    # 공통 불변식: 첫 데이터 시트 A1이 문장 텍스트로 오염되면 라운드 실패다.
    wb = load_workbook(WB)
    for name in wb.sheetnames:
        a1 = wb[name]["A1"].value
        if isinstance(a1, str) and len(a1) > 14 and ("줘" in a1 or "해" in a1[-2:]):
            print(f"  !! {name}!A1 오염: {a1[:40]}")
            for t in log:
                t["ok"] = False if t["turn"] == log[-1]["turn"] else t["ok"]
    wb.close()
    Path(__file__).with_name(SCENARIO.stem + "_log.json").write_text(
        json.dumps(log, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return log


async def main() -> None:
    repeat = max(1, int(os.environ.get("HUMAN_REPEAT", "1")))
    logs = [await run_once(r) for r in range(1, repeat + 1)]
    if repeat > 1:
        base = [(t["ok"], t.get("action"), t.get("asked")) for t in logs[0]]
        flaky = [
            (r, t["turn"], t["text"][:36])
            for r, lg in enumerate(logs[1:], start=2)
            for i, t in enumerate(lg)
            if (t["ok"], t.get("action"), t.get("asked")) != base[i]
        ]
        print(f"결정성: {repeat}회 중 흔들린 턴 {len(flaky)}건")
        for f in flaky:
            print("   ", f)


asyncio.run(main())
