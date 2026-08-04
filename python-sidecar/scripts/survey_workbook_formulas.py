"""워크북에 실제로 쓰인 수식과 함수 이름을 조사한다.

수식 계산기가 어디까지 지원해야 하는지 정하려면 실제 파일에 무엇이 있는지 알아야 한다.
"""

import argparse
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

FUNC_RE = re.compile(r"([A-Z][A-Z0-9._]*)\s*\(")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--samples", type=int, default=2, help="함수별로 보여줄 수식 예시 개수")
    args = parser.parse_args()

    wb = load_workbook(filename=str(args.workbook), data_only=False)
    funcs: Counter[str] = Counter()
    plain_arithmetic = 0
    examples: dict[str, list[str]] = {}
    total = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                total += 1
                names = set(FUNC_RE.findall(value))
                if not names:
                    plain_arithmetic += 1
                for name in names:
                    funcs[name] += 1
                    bucket = examples.setdefault(name, [])
                    if len(bucket) < args.samples:
                        bucket.append(f"{ws.title}!{cell.coordinate}  {value}")
    wb.close()

    print(f"수식 셀 {total}개 / 함수 없는 사칙연산만 {plain_arithmetic}개\n")
    for name, count in funcs.most_common():
        print(f"{name}  x{count}")
        for sample in examples.get(name, []):
            print(f"    {sample}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
