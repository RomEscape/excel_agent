"""블라인드 파라프레이즈 게이트 — 코드를 본 적 없는 쪽이 쓴 문장으로 일반화를 잰다.

왜 필요한가 (2026-08-19 사용자 지적): "너가 테스트해본 말에 대해서만 수정을 거친 거지, 강건한
게 아니다." 지금까지의 배터리는 전부 내가(또는 로그가) 쓴 문장이라 **회귀**는 증명해도
**일반화**는 증명하지 못한다. 이 게이트는 각 기능(task)의 정답을 **파일 상태 오라클**로 정의해
두고, 문장은 바깥(규칙을 못 본 작성자)에서 받아 처음 돌린다.

결과 분류(문장 하나당 하나):
  PASS_RULE  규칙이 확정해 실행, 오라클 참
  PASS_CARD  모델 해석(해석 카드 → 사람이 '맞아요') 뒤 실행, 오라클 참
  ASK        되묻기/카드 없이 실행되지 않음(정당한 질문일 수도, 이해 실패일 수도)
  WRONG      실행됐는데 오라클 거짓 — 카드 없이 실행됐으면 **조용한 오실행**
  ERROR      예외·ok=false
핵심 지표: 정답 실행률(PASS_RULE+PASS_CARD), 조용한 오실행률(WRONG & not card), 되묻기율.

사용:
  PYTHONUTF8=1 EXCEL_LIVE_ENGINE=file python scripts/run_blind_paraphrase_gate.py --tasks      # 작성자에게 줄 과제표
  PYTHONUTF8=1 EXCEL_LIVE_ENGINE=file python scripts/run_blind_paraphrase_gate.py ../datasets/eval/blind_paraphrases_v1.jsonl
"""

from __future__ import annotations

import asyncio
import json
import os
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.routers.excel_live import (
    ApprovalResponse,
    ExcelLiveCommandRequest,
    post_approval,
    post_command,
)
from office_claw_sidecar.services.excel_live_service import (
    get_excel_live_service,
    invalidate_excel_engine_cache,
)
from office_claw_sidecar.services.llm_service import get_llm_service

WS = Path(r"C:\Users\asdjj\AppData\Local\office_claw\Workspace")
WB = WS / "blind_gate.xlsx"

SEED = [
    ["지역", "주문건수", "출고건수", "정시배송률", "지연건수", "클레임"],
    ["수도권", 10452, 10158, 97.1, 145, 12],
    ["충청권", 3892, 3773, 95.2, 89, 6],
    ["호남권", 3214, 3086, 94.7, 112, 5],
    ["영남권", 6789, 6512, 95.8, 174, 5],
    ["강원제주", 2495, 2383, 92.6, 145, 0],
]
STATUS_SEED = [
    ["운송장", "구간", "지연시간", "상태"],
    ["INV-001", "김포-부산", "12시간", "대기"],
    ["INV-002", "이천-대전", "3시간", "처리중"],
    ["INV-003", "용인-광주", "8시간", "대기"],
    ["INV-004", "평택-제주", "1시간", "완료"],
]


def _sheet(wb, name="지역성과"):
    return wb[name]


def _fill(wb, cell, sheet="지역성과"):
    c = _sheet(wb, sheet)[cell]
    return c.fill.fgColor.rgb if c.fill and c.fill.fill_type else None


def _formula(wb, cell, frag, sheet="지역성과"):
    v = _sheet(wb, sheet)[cell].value
    return isinstance(v, str) and v.startswith("=") and frag.upper() in v.upper()


def _charts(wb, sheet="지역성과"):
    return len(_sheet(wb, sheet)._charts or [])


def _cf(wb, sheet="지역성과"):
    return len(_sheet(wb, sheet).conditional_formatting)


def _seed_default(wb):
    ws = wb.active
    ws.title = "지역성과"
    for r in SEED:
        ws.append(r)


def _seed_with_totals(wb):
    _seed_default(wb)
    ws = wb["지역성과"]
    ws["A7"] = "합계"
    for col in "BCDEF":
        ws[f"{col}7"] = f"=SUM({col}2:{col}6)"


def _seed_status(wb):
    ws = wb.active
    ws.title = "지연경고"
    for r in STATUS_SEED:
        ws.append(r)


def _seed_with_summary_sheet(wb):
    _seed_default(wb)
    wb.create_sheet("요약")
    wb["요약"]["A1"] = "전체주문건수"
    wb.active = wb.sheetnames.index("요약")


GRADES_SEED = [
    ["학생", "점수", "결석"],
    ["김민준", 88, 2],
    ["이서연", 94, 0],
    ["박도윤", 71, 5],
    ["최지우", 83, 1],
]


def _seed_grades_and_summary(wb):
    """성적부(이름이 든 원본) + 요약(빈 시트). 활성 시트는 요약이다.

    2026-08-19 감사에서 크로스시트 집계가 **성적부에** 써져 학생 이름이 지워졌다.
    """
    ws = wb.active
    ws.title = "성적부"
    for row in GRADES_SEED:
        ws.append(row)
    wb.create_sheet("요약")
    wb.active = wb.sheetnames.index("요약")


def _seed_grades(wb):
    ws = wb.active
    ws.title = "성적부"
    for row in GRADES_SEED:
        ws.append(row)


def _grades_intact(wb):
    """성적부의 이름·점수·결석이 그대로인가. 아니면 무엇이 어긋났는지 돌려준다."""
    if "성적부" not in wb.sheetnames:
        return "성적부 시트가 사라짐"
    ws = wb["성적부"]
    for idx, row in enumerate(GRADES_SEED, start=1):
        for col, expected in enumerate(row, start=1):
            actual = ws.cell(row=idx, column=col).value
            if actual != expected:
                return f"성적부 {ws.cell(row=idx, column=col).coordinate} {expected!r} → {actual!r}"
    return ""


def _summary_total(wb):
    """요약!A2가 **합계다운 값**인가 — 숫자 8이거나 SUM 수식이어야 한다.

    2026-08-20 자체 검토: "비어 있지 않으면 통과"로 뒀더니 규칙이 흘린 쓰레기 한 글자('성')가
    성공으로 세어졌다. 오라클이 느슨하면 수정이 나아졌는지 알 수 없다.
    """
    value = wb["요약"]["A2"].value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "" if int(value) == 8 else f"요약!A2={value!r} (결석 합계는 8)"
    text = str(value or "").strip()
    if not text:
        return "요약!A2 비어 있음"
    if text.startswith("=") and "SUM" in text.upper():
        return ""
    return f"요약!A2={text[:40]!r} — 합계가 아니다"


def _pairs(wb, sheet="성적부"):
    ws = wb[sheet]
    out = []
    for r in range(2, ws.max_row + 1):
        name, score = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
        if name is None and score is None:
            continue
        out.append((name, score))
    return out


def _seed_small_calc(wb):
    ws = wb.active
    ws.title = "계산"
    ws.append(["항목", "이번주", "지난주", "증감"])
    ws.append(["매출", 120, 100, None])


def _seed_with_chart(wb):
    _seed_default(wb)
    from openpyxl.chart import BarChart, Reference

    ws = wb["지역성과"]
    ch = BarChart()
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=6), titles_from_data=True)
    ws.add_chart(ch, "H2")


NAVY = {"FF002060", "FF1F3864", "FF16324F", "FF0B2447", "FF1F4E79", "FF203864"}
RED = {"FFFF0000", "FFC00000", "FFFF3B30"}
PINK = {"FFFFC7CE", "FFFFC0CB", "FFFF99CC", "FFF4CCCC"}


def _sorted_desc_by_col(wb, col_idx: int, data_rows=(2, 6)):
    ws = _sheet(wb)
    vals = [ws.cell(row=r, column=col_idx).value for r in range(data_rows[0], data_rows[1] + 1)]
    nums = [v for v in vals if isinstance(v, (int, float))]
    return len(nums) == len(vals) and nums == sorted(nums, reverse=True)


# task_id → 설명(작성자용) · 정준 문장(기준) · context_range · 씨앗 · 오라클(wb)->오류문자열|""
TASKS: dict[str, dict] = {
    "sum_below": {
        "desc": "표(A1:F6) 바로 아래 줄에 각 숫자 열의 합계 한 줄을 넣는다(A7에 '합계' 라벨, B7~F7 SUM 수식).",
        "canonical": "합계를 표 아래에 한 줄로 넣어줘", "ctx": "A1:F6", "seed": _seed_default,
        "oracle": lambda wb: "" if _formula(wb, "B7", "SUM") and _formula(wb, "F7", "SUM") else "B7/F7 SUM 없음"},
    "avg_below": {
        "desc": "이미 합계 줄(7행)이 있는 표에서 그 아래 줄(8행)에 각 숫자 열의 평균 한 줄을 넣는다.",
        "canonical": "그 아래 칸에는 평균도 넣어줘", "ctx": "A1:F7", "seed": _seed_with_totals,
        "oracle": lambda wb: "" if _formula(wb, "B8", "AVERAGE") else "B8 AVERAGE 없음"},
    "header_navy": {
        "desc": "표의 첫 줄(머리글 A1:F1)을 남색 배경, 흰 글씨, 굵게 만든다.",
        "canonical": "머리글 행은 남색 배경에 흰 글씨로 굵게 해줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _fill(wb, "A1") in NAVY and _sheet(wb)["A1"].font.bold else f"A1 채움={_fill(wb, 'A1')} bold={_sheet(wb)['A1'].font.bold}"},
    "border_table": {
        "desc": "표 전체(A1:F6)에 테두리(경계선)를 두른다.",
        "canonical": "표 전체에 테두리 그려줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _sheet(wb)["A1"].border.top.style and _sheet(wb)["F6"].border.bottom.style else "테두리 없음"},
    "comma_cols": {
        "desc": "주문건수·출고건수 열(B·C)의 숫자에 천 단위 콤마 표시 형식을 적용한다.",
        "canonical": "주문건수랑 출고건수는 천 단위 콤마로 보여줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if "#,##0" in _sheet(wb)["B3"].number_format and "#,##0" in _sheet(wb)["C3"].number_format else f"B3 형식={_sheet(wb)['B3'].number_format}"},
    "sort_desc": {
        "desc": "표를 주문건수가 많은 순서(내림차순)로 정렬한다. 머리글은 그대로.",
        "canonical": "주문건수 많은 순으로 정렬해줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _sheet(wb)["A1"].value == "지역" and _sorted_desc_by_col(wb, 2) else f"정렬 안 됨 A2={_sheet(wb)['A2'].value}"},
    "sort_keep_total": {
        "desc": "합계 줄(7행)이 있는 표를 클레임 많은 순으로 정렬하되, 합계 줄은 맨 아래 그대로 둔다.",
        "canonical": "클레임 많은 순으로 정렬해줘", "ctx": None, "seed": _seed_with_totals,
        "oracle": lambda wb: "" if _sheet(wb)["A7"].value == "합계" and _formula(wb, "B7", "SUM") and _sorted_desc_by_col(wb, 6) else f"A7={_sheet(wb)['A7'].value} / 정렬={_sorted_desc_by_col(wb, 6)}"},
    "new_sheet": {
        "desc": "'요약'이라는 이름의 새 시트를 만든다.",
        "canonical": "요약 시트 만들어줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if "요약" in wb.sheetnames else f"시트={wb.sheetnames}"},
    "title_cell": {
        "desc": "H1 셀에 '물류 관제 대시보드'라는 제목 글자를 넣는다.",
        "canonical": "H1에 물류 관제 대시보드 라고 써줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if str(_sheet(wb)["H1"].value or "").strip() == "물류 관제 대시보드" else f"H1={_sheet(wb)['H1'].value!r}"},
    "merge_title": {
        "desc": "H1부터 M1까지를 하나로 병합한다.",
        "canonical": "H1부터 M1까지 병합해줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if any(str(r) == "H1:M1" for r in _sheet(wb).merged_cells.ranges) else f"병합={[str(r) for r in _sheet(wb).merged_cells.ranges]}"},
    "freeze_top": {
        "desc": "첫 줄(머리글)이 스크롤해도 보이게 고정한다(틀 고정 A2).",
        "canonical": "첫 줄은 스크롤해도 보이게 고정해줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if str(_sheet(wb).freeze_panes) == "A2" else f"틀고정={_sheet(wb).freeze_panes}"},
    "highlight_status": {
        "desc": "지연경고 표(A1:D5, 머리글 운송장/구간/지연시간/상태)에서 상태가 '대기'인 셀만 분홍색으로 강조한다.",
        "canonical": "상태가 대기인 셀만 분홍색으로 강조해줘", "ctx": None, "seed": _seed_status,
        "oracle": lambda wb: "" if (_fill(wb, "D2", "지연경고") in PINK and _fill(wb, "D4", "지연경고") in PINK and _fill(wb, "D3", "지연경고") not in PINK) or _cf(wb, "지연경고") >= 1 else f"D2={_fill(wb, 'D2', '지연경고')} D3={_fill(wb, 'D3', '지연경고')}"},
    "highlight_threshold": {
        "desc": "클레임(F열)이 10보다 큰 셀만 빨간색으로 칠한다(수도권 12만 해당).",
        "canonical": "클레임 10 넘는 셀만 빨간색으로 칠해줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if (_fill(wb, "F2") in RED and _fill(wb, "F3") not in RED) or _cf(wb) >= 1 else f"F2={_fill(wb, 'F2')} F3={_fill(wb, 'F3')}"},
    "line_chart": {
        "desc": "정시배송률 열로 선 그래프 하나를 만든다.",
        "canonical": "정시배송률 추이를 선 그래프로 그려줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _charts(wb) >= 1 else "차트 없음"},
    "bar_chart": {
        "desc": "지연건수 열로 막대 그래프 하나를 만든다.",
        "canonical": "지연건수로 막대 그래프 그려줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _charts(wb) >= 1 else "차트 없음"},
    "delete_charts": {
        "desc": "시트에 있는 차트를 전부 지운다(시트에 막대 차트 1개가 있음). 데이터는 그대로.",
        "canonical": "차트 다 지워줘", "ctx": None, "seed": _seed_with_chart,
        "oracle": lambda wb: "" if _charts(wb) == 0 and _sheet(wb)["B2"].value == 10452 else f"차트={_charts(wb)} B2={_sheet(wb)['B2'].value}"},
    "clear_table": {
        "desc": "표(A1:F6)의 값을 전부 지워 빈칸으로 만든다.",
        "canonical": "A1:F6 내용 다 지워줘", "ctx": "A1:F6", "seed": _seed_default,
        "oracle": lambda wb: "" if _sheet(wb)["A1"].value is None and _sheet(wb)["F6"].value is None else f"A1={_sheet(wb)['A1'].value!r}"},
    "cross_sheet_sum": {
        "desc": "현재 시트는 '요약'(A1에 '전체주문건수'). A2에 지역성과 시트의 주문건수 합계를 수식으로 가져온다.",
        "canonical": "A2에 지역성과 시트 주문건수 합계 가져와줘", "ctx": None, "seed": _seed_with_summary_sheet,
        "oracle": lambda wb: "" if _formula(wb, "A2", "지역성과", "요약") and _formula(wb, "A2", "SUM", "요약") else f"요약!A2={_sheet(wb, '요약')['A2'].value!r}"},
    "paste_values": {
        "desc": "사람이 Excel에서 A8:B9 두 줄 두 칸을 선택해 복사·붙여넣기 한 뒤 값 나열을 친다 — 서울,100; 부산,200 이 A8:B9에 들어가야 한다. (문장에는 값 '서울,100; 부산,200'이 반드시 포함된다.)",
        "canonical": "서울,100; 부산,200 입력해줘", "ctx": "A8:B9", "seed": _seed_default,
        "oracle": lambda wb: "" if _sheet(wb)["A8"].value == "서울" and _sheet(wb)["B9"].value == 200 else f"A8={_sheet(wb)['A8'].value!r} B9={_sheet(wb)['B9'].value!r}"},
    "find_replace": {
        "desc": "표에서 '수도권'이라는 글자를 '서울권'으로 바꾼다.",
        "canonical": "수도권을 서울권으로 바꿔줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _sheet(wb)["A2"].value == "서울권" else f"A2={_sheet(wb)['A2'].value!r}"},
    "rename_sheet": {
        "desc": "현재 시트(지역성과)의 이름을 '지역별실적'으로 바꾼다.",
        "canonical": "지역성과 시트 이름을 지역별실적으로 바꿔줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if "지역별실적" in wb.sheetnames else f"시트={wb.sheetnames}"},
    "cell_subtract": {
        "desc": "계산 시트(A1:D2, 머리글 항목/이번주/지난주/증감, 2행 매출 120 100). D2에 이번주(B2)에서 지난주(C2)를 뺀 값을 수식으로 넣는다.",
        "canonical": "D2에 B2에서 C2 뺀 값 넣어줘", "ctx": None, "seed": _seed_small_calc,
        "oracle": lambda wb: "" if str(_sheet(wb, '계산')["D2"].value or "").replace(" ", "").upper() in {"=B2-C2"} else f"계산!D2={_sheet(wb, '계산')['D2'].value!r}"},
    "percent_format": {
        "desc": "정시배송률 열(D2:D6) 숫자에 소수 첫째 자리까지 보이는 표시 형식을 적용한다(예: 97.1).",
        "canonical": "정시배송률은 소수 첫째 자리까지 보여줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if "0.0" in _sheet(wb)["D3"].number_format else f"D3 형식={_sheet(wb)['D3'].number_format}"},
    "data_bar": {
        "desc": "주문건수 열에 데이터 막대(셀 안에 길이 막대) 조건부 서식을 넣는다.",
        "canonical": "주문건수 열에 데이터 막대 넣어줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if _cf(wb) >= 1 else "조건부 서식 없음"},
    "autofit": {
        "desc": "열 너비를 내용에 맞게 자동으로 맞춘다.",
        "canonical": "열 너비 보기 좋게 맞춰줘", "ctx": None, "seed": _seed_default,
        "oracle": lambda wb: "" if any((_sheet(wb).column_dimensions[c].width or 0) > 0 and _sheet(wb).column_dimensions[c].width != 13 for c in "ABCDEF") else "열 너비 변경 없음"},
    # ── 파괴 위험 3종 (2026-08-20 추가) ──────────────────────────────
    # 되묻기는 실패가 아니다. **망가뜨리는 것만** 실패다.
    "cross_sheet_keeps_source": {
        "desc": "요약 시트에서 성적부를 집계한다. 결과는 요약에 쓰고 **성적부는 그대로**여야 한다.",
        "canonical": "A2에 성적부 결석 합계 넣어줘", "ctx": None, "seed": _seed_grades_and_summary,
        "oracle": lambda wb: _grades_intact(wb) or _summary_total(wb)},
    "sort_keeps_pairs": {
        "desc": "점수로 정렬한다. **이름과 점수의 짝이 유지**되어야 한다(한 열만 정렬하면 어긋난다).",
        "canonical": "점수 높은 순으로 정렬해줘", "ctx": None, "seed": _seed_grades,
        "oracle": lambda wb: (
            "" if sorted(_pairs(wb)) == sorted([(r[0], r[1]) for r in GRADES_SEED[1:]])
            else f"이름·점수 짝이 어긋남: {_pairs(wb)}"
        )},
    "clear_only_named": {
        "desc": "결석 열(C)만 비운다. 학생 이름과 점수는 **그대로**여야 한다.",
        "canonical": "결석 열만 비워줘", "ctx": None, "seed": _seed_grades,
        "oracle": lambda wb: (
            f"이름·점수가 지워짐: {_pairs(wb)}"
            if sorted(_pairs(wb)) != sorted([(r[0], r[1]) for r in GRADES_SEED[1:]])
            else ("" if all(wb["성적부"].cell(row=r, column=3).value in (None, "") for r in range(2, 6))
                  else f"결석 열이 안 비워짐: {[wb['성적부'].cell(row=r, column=3).value for r in range(2, 6)]}")
        )},
    "negation_save": {
        "desc": "'저장하지 마' 같은 부정문 — 아무것도 실행되면 안 된다(표는 그대로).",
        "canonical": "아직 저장하지 마", "ctx": None, "seed": _seed_default, "negative": True,
        "oracle": lambda wb: "" if _sheet(wb)["A1"].value == "지역" and _sheet(wb)["F6"].value == 0 else "표가 바뀜"},
}


def _reply(resp, result, summary):
    if result.get("ask_follow_up"):
        return str(result.get("follow_up_question") or getattr(resp, "reason", ""))
    if result.get("execution_report"):
        return str(result["execution_report"])
    return summary or str(getattr(resp, "reason", ""))


async def run_one(idx: int, row: dict, llm) -> dict:
    task = TASKS[row["task"]]
    if WB.exists():
        WB.unlink()
    wb = Workbook()
    task["seed"](wb)
    wb.save(WB)
    wb.close()
    invalidate_excel_engine_cache()
    get_excel_live_service().select_workbook(str(WB))
    text = str(row["text"])
    ctx = task.get("ctx")
    session = f"test-{os.environ.get('BLIND_SESSION_TAG', 'blind')}-{idx}"
    t0 = time.time()
    card = False
    summary = ""
    try:
        resp = await post_command(
            ExcelLiveCommandRequest(message=text, session_id=session, workbook_id=None, approve=False, context_range=ctx),
            llm,
        )
        if getattr(resp, "approval_required", False):
            pending = resp.pending_approval
            card = bool(getattr(pending, "interpretation", False))
            summary = str(getattr(pending, "summary", "") or "")
            resp = await post_approval(ApprovalResponse(approval_id=pending.approval_id, approved=True), llm)
    except Exception as exc:
        return {"idx": idx, **row, "outcome": "ERROR", "detail": f"{type(exc).__name__}: {exc}"[:200], "secs": round(time.time() - t0, 1)}
    result = getattr(resp, "result", None) or {}
    action = str(getattr(resp, "action", ""))
    asked = bool(result.get("ask_follow_up")) or "clarify" in action
    ok = bool(getattr(resp, "ok", False))
    wb = load_workbook(WB)
    err = task["oracle"](wb)
    wb.close()
    if task.get("negative"):
        outcome = "PASS_RULE" if (err == "" and (asked or action in {"excel_live.noop", "excel_live.not_excel_request"} or not ok)) else "WRONG"
    elif asked and err:
        outcome = "ASK"
    elif not ok and err:
        outcome = "ERROR"
    elif err == "":
        outcome = "PASS_CARD" if card else "PASS_RULE"
    else:
        outcome = "WRONG"
    return {
        "idx": idx, **row, "outcome": outcome, "card": card, "action": action,
        "detail": err or _reply(resp, result, summary)[:160], "secs": round(time.time() - t0, 1),
    }


async def main() -> None:
    if "--tasks" in sys.argv:
        spec = [
            {"task": k, "desc": v["desc"], "canonical": v["canonical"], "paste_context": v.get("ctx")}
            for k, v in TASKS.items()
        ]
        print(json.dumps(spec, ensure_ascii=False, indent=1))
        return
    src = Path(sys.argv[1])
    rows = [json.loads(ln) for ln in src.read_text(encoding="utf-8").splitlines() if ln.strip()]
    rows = [r for r in rows if r.get("task") in TASKS]
    only = os.environ.get("BLIND_ONLY")
    if only:
        rows = [r for r in rows if r["task"] in set(only.split(","))]
    llm = get_llm_service()
    out = []
    for i, r in enumerate(rows, 1):
        res = await run_one(i, r, llm)
        out.append(res)
        flag = {"PASS_RULE": "OK ", "PASS_CARD": "OKc", "ASK": "ASK", "WRONG": "BAD", "ERROR": "ERR"}[res["outcome"]]
        print(f"[{i:3d}/{len(rows)}] {flag} {r['task']:18s} {r['text'][:48]}" + (f"  → {res['detail'][:70]}" if res["outcome"] in {"WRONG", "ERROR", "ASK"} else ""))
    report = src.with_name(src.stem + "_report.json")
    report.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    total = Counter(r["outcome"] for r in out)
    silent = sum(1 for r in out if r["outcome"] == "WRONG" and not r.get("card"))
    n = len(out)
    print("\n==== 블라인드 게이트 요약 ====")
    print(f"문장 {n}개 · 정답 실행 {total['PASS_RULE'] + total['PASS_CARD']} ({(total['PASS_RULE'] + total['PASS_CARD']) / max(n, 1):.1%})"
          f" [규칙 {total['PASS_RULE']} · 카드 {total['PASS_CARD']}] · 되묻기 {total['ASK']} ({total['ASK'] / max(n, 1):.1%})"
          f" · 오실행 {total['WRONG']} (조용한 오실행 {silent}, {silent / max(n, 1):.1%}) · 오류 {total['ERROR']}")
    per = defaultdict(Counter)
    for r in out:
        per[r["task"]][r["outcome"]] += 1
    print("\n과제별: 정답/되묻기/오실행/오류")
    for t, c in sorted(per.items(), key=lambda kv: -(kv[1]["WRONG"] + kv[1]["ERROR"])):
        print(f"  {t:18s} {c['PASS_RULE'] + c['PASS_CARD']:3d} / {c['ASK']:3d} / {c['WRONG']:3d} / {c['ERROR']:3d}")
    print(f"\n보고서: {report}")


from office_claw_sidecar.services import decision_trace as _dt

# 배터리 턴을 사람이 친 명령과 가를 출처 태그(2026-08-19 로그 감사: 실사용 로그의 source가 전부 비어 있었다).
_dt.source(kind="script", name="blind_gate").__enter__()
asyncio.run(main())
