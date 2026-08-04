"""데모 워크북의 시트·머리글·표본값과 시나리오 정답 후보를 뽑는다.

새 검증 시나리오를 만들 때 기대값을 손으로 짐작하지 않기 위한 도구다.
수식 열은 excel_formula_eval로 계산해서 실제 값을 쓴다.

사용: python scripts/describe_demo_workbook.py <workbook.xlsx> [out.txt]
"""

from __future__ import annotations

import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from office_claw_sidecar.services.excel_formula_eval import WorkbookEvaluator


def _build_evaluator(wb: Any) -> Any:
    """수식 열을 실제 값으로 읽기 위한 계산기. 셀 주소 문자열로 부를 수 있게 감싼다."""

    def raw(sheet: str, row: int, col: int) -> Any:
        ws = wb[sheet] if sheet in wb.sheetnames else None
        return None if ws is None else ws.cell(row=row, column=col).value

    def bounds(sheet: str) -> tuple[int, int]:
        ws = wb[sheet] if sheet in wb.sheetnames else None
        return (0, 0) if ws is None else (ws.max_row, ws.max_column)

    evaluator = WorkbookEvaluator(raw, default_sheet=wb.sheetnames[0], sheet_bounds=bounds)

    def evaluate(sheet: str, address: str) -> Any:
        cell = wb[sheet][address]
        try:
            return evaluator.value(sheet, cell.row, cell.column)
        except Exception:  # noqa: BLE001 - 계산 실패는 원본 값으로 보고한다
            return cell.value

    return evaluate


def _as_float(value: Any) -> float | None:
    try:
        if value is None or isinstance(value, bool):
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _describe_sheet(ws: Any, evaluate: Any, lines: list[str]) -> None:
    lines.append(f"\n=== {ws.title} dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    for index, header in enumerate(headers, start=1):
        if not header:
            continue
        letter = ws.cell(row=1, column=index).column_letter
        sample = ws.cell(row=2, column=index).value
        lines.append(f"  {letter} {header!r} sample={sample!r}")

    if ws.max_row < 3:
        return

    for index, header in enumerate(headers, start=1):
        if not header:
            continue
        letter = ws.cell(row=1, column=index).column_letter
        values = [evaluate(ws.title, f"{letter}{row}") for row in range(2, ws.max_row + 1)]
        numbers = [n for n in (_as_float(v) for v in values) if n is not None]
        if numbers and len(numbers) >= (ws.max_row - 1) * 0.8:
            lines.append(
                f"  [num] {header}: sum={sum(numbers):.2f} min={min(numbers):.2f} "
                f"max={max(numbers):.2f} n={len(numbers)}"
            )
        else:
            counter = Counter(str(v) for v in values if v not in (None, ""))
            if 0 < len(counter) <= 15:
                lines.append(f"  [cat] {header}: {dict(counter)}")


def _group_sums(ws: Any, evaluate: Any, key_header: str, value_header: str) -> dict[str, float]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if key_header not in headers or value_header not in headers:
        return {}
    key_letter = ws.cell(row=1, column=headers.index(key_header) + 1).column_letter
    value_letter = ws.cell(row=1, column=headers.index(value_header) + 1).column_letter
    totals: dict[str, float] = defaultdict(float)
    for row in range(2, ws.max_row + 1):
        key = evaluate(ws.title, f"{key_letter}{row}")
        number = _as_float(evaluate(ws.title, f"{value_letter}{row}"))
        if key is None or number is None:
            continue
        totals[str(key)] += number
    return dict(totals)


def main() -> int:
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    path = Path(sys.argv[1])
    wb = load_workbook(path, data_only=False)
    evaluate = _build_evaluator(wb)

    lines = [f"workbook={path}", f"sheets={wb.sheetnames}"]
    for ws in wb.worksheets:
        _describe_sheet(ws, evaluate, lines)

    if "Sales_Data" in wb.sheetnames:
        sales = wb["Sales_Data"]
        for key in ("Region", "Channel", "Category", "Salesperson"):
            totals = _group_sums(sales, evaluate, key, "Sales")
            if totals:
                lines.append(f"\n[group] Sales by {key}:")
                for name, total in sorted(totals.items(), key=lambda kv: -kv[1]):
                    lines.append(f"  {name}: {total:.2f}")

    text = "\n".join(lines)
    if len(sys.argv) >= 3:
        Path(sys.argv[2]).write_text(text, encoding="utf-8")
        print(f"wrote {sys.argv[2]}")
    else:
        print(text)
    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
