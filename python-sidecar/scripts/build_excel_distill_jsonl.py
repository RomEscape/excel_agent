from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

KST = timezone(timedelta(hours=9), name="KST")
SCHEMA_VERSION = "excel_distill.v1"


@dataclass
class BuildStats:
    added: int = 0
    skipped: int = 0
    errors: int = 0


def now_kst_iso() -> str:
    return datetime.now(KST).isoformat()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def detect_locale(text: str) -> str:
    if not text:
        return "unknown"
    ko_chars = sum(1 for ch in text if "\uac00" <= ch <= "\ud7a3")
    en_chars = sum(1 for ch in text if ("a" <= ch.lower() <= "z"))
    if ko_chars and en_chars:
        return "mixed"
    if ko_chars:
        return "ko"
    if en_chars:
        return "en"
    return "unknown"


def locale_matches_preferred(locale: str, preferred_locale: str) -> bool:
    pref = normalize_text(preferred_locale).lower()
    if pref in {"ko", "korean", "kr"}:
        return locale in {"ko", "mixed"}
    if pref in {"en", "english"}:
        return locale in {"en", "mixed"}
    return True


def build_language_views(text: str, locale: str) -> dict[str, Any]:
    normalized = normalize_text(text)
    no_space = re.sub(r"\s+", "", normalized)
    compact = re.sub(r"[\s\-_]+", "", normalized).lower()
    views: dict[str, Any] = {
        "normalized": normalized,
        "no_space": no_space,
        "compact": compact,
        "contains_hangul": any("\uac00" <= ch <= "\ud7a3" for ch in normalized),
    }
    if locale in {"ko", "mixed"}:
        ko_core = normalized
        ko_core = re.sub(r"\s+", " ", ko_core).strip()
        replacements = [
            (r"해\s*주세요", "해줘"),
            (r"해\s*주세용", "해줘"),
            (r"해주세요", "해줘"),
            (r"해주셔요", "해줘"),
            (r"부탁드립니다", "부탁해"),
            (r"부탁드려요", "부탁해"),
            (r"해줘요", "해줘"),
            (r"해주라", "해줘"),
            (r"해주세여", "해줘"),
        ]
        for pattern, target in replacements:
            ko_core = re.sub(pattern, target, ko_core)
        ko_core = re.sub(r"[.!?]+$", "", ko_core).strip()
        views["ko_core"] = ko_core
        views["ko_no_space"] = re.sub(r"\s+", "", ko_core)
    return views


def sanitize_sample_id(value: Any, fallback: str) -> str:
    text = normalize_text(value)
    if not text:
        return fallback
    text = re.sub(r"[^\w\-.:]+", "_", text, flags=re.UNICODE)
    return text[:120] or fallback


def to_rel_or_raw(path: Path, root: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except Exception:
        return str(path)


def infer_workbook_role(key: str, path_str: str) -> str:
    joined = f"{key} {path_str}".lower()
    if "answer" in joined or "golden" in joined:
        return "golden"
    if "output" in joined:
        return "output"
    return "input"


def discover_workbooks_from_dir(target_dir: Path, root: Path, key: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    if not target_dir.exists() or not target_dir.is_dir():
        return rows
    files = sorted([fp for fp in target_dir.glob("*.xls*") if fp.is_file()])
    for fp in files:
        rows.append(
            {
                "role": infer_workbook_role(key, fp.name),
                "path": to_rel_or_raw(fp, root),
            }
        )
    return rows


def collect_workbook_refs(sample: dict[str, Any], root: Path) -> list[dict[str, str]]:
    refs: list[dict[str, str]] = []
    for key, value in sample.items():
        if value is None:
            continue
        lowered_key = str(key).lower()
        if isinstance(value, str):
            text = value.strip()
            if not text:
                continue
            if text.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                candidate = (root / text).resolve() if not Path(text).is_absolute() else Path(text)
                refs.append(
                    {
                        "role": infer_workbook_role(lowered_key, text),
                        "path": to_rel_or_raw(candidate, root),
                    }
                )
            elif "spreadsheet_path" in lowered_key or "golden_response_path" in lowered_key:
                candidate_dir = (root / text).resolve() if not Path(text).is_absolute() else Path(text)
                refs.extend(discover_workbooks_from_dir(candidate_dir, root, lowered_key))
        elif isinstance(value, list):
            for idx, item in enumerate(value):
                if isinstance(item, str) and item.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
                    candidate = (root / item).resolve() if not Path(item).is_absolute() else Path(item)
                    refs.append(
                        {
                            "role": infer_workbook_role(f"{lowered_key}_{idx}", item),
                            "path": to_rel_or_raw(candidate, root),
                        }
                    )
    dedup: dict[tuple[str, str], dict[str, str]] = {}
    for row in refs:
        key = (row.get("role", "input"), row.get("path", ""))
        dedup[key] = row
    return list(dedup.values())


def short_hash(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8")).hexdigest()[:10]


def make_record(
    *,
    dataset: str,
    split: str,
    sample_id: str,
    instruction: str,
    workbook_refs: list[dict[str, str]],
    source_file: str,
    context_hints: dict[str, Any] | None = None,
    task_type: str = "spreadsheet_edit",
    label_status: str = "needs_teacher_plan",
    action_plan: list[dict[str, Any]] | None = None,
    expected_output: dict[str, Any] | None = None,
    verification: str = "none",
    passed: bool | None = None,
    confidence: float = 0.0,
    notes: list[str] | None = None,
    preferred_locale: str = "ko",
) -> dict[str, Any]:
    cleaned_instruction = normalize_text(instruction)
    locale = detect_locale(cleaned_instruction)
    preferred_match = locale_matches_preferred(locale, preferred_locale)
    training_notes: list[str] = []
    if not preferred_match:
        training_notes.append("needs_locale_rewrite")
    digest = short_hash(f"{dataset}|{sample_id}|{cleaned_instruction}")
    record_id = f"{dataset}:{sample_id}:{digest}"
    return {
        "schema_version": SCHEMA_VERSION,
        "record_id": record_id,
        "source": {
            "dataset": dataset,
            "split": split,
            "sample_id": sample_id,
            "license": "unknown",
            "provenance": {
                "source_file": source_file,
            },
        },
        "input": {
            "instruction": cleaned_instruction,
            "locale": locale,
            "language_views": build_language_views(cleaned_instruction, locale),
            "training_hints": {
                "preferred_locale": preferred_locale,
                "preferred_locale_match": preferred_match,
                "notes": training_notes,
            },
            "workbook_refs": workbook_refs,
            "context_hints": context_hints or {},
        },
        "target": {
            "task_type": task_type,
            "label_status": label_status,
            "action_plan": action_plan or [],
            "expected_output": expected_output or {},
        },
        "quality": {
            "verification": verification,
            "passed": passed,
            "confidence": round(float(confidence), 4),
        },
        "metadata": {
            "created_at": now_kst_iso(),
            "generator": "python-sidecar/scripts/build_excel_distill_jsonl.py",
            "notes": notes or [],
        },
    }


def iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        for line_no, raw in enumerate(f, start=1):
            text = raw.strip()
            if not text:
                continue
            try:
                row = json.loads(text)
                if isinstance(row, dict):
                    yield row
            except json.JSONDecodeError:
                print(f"[WARN] JSONL 파싱 실패: {path}#{line_no}")


def iter_json(path: Path) -> Iterable[dict[str, Any]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        print(f"[WARN] JSON 파싱 실패: {path}")
        return []
    if isinstance(payload, dict):
        if "instruction" in payload:
            return [payload]
        if isinstance(payload.get("data"), list):
            return [row for row in payload["data"] if isinstance(row, dict)]
        return []
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    return []


def load_instruction_dataset(
    *,
    dataset_name: str,
    root: Path,
    split: str,
    limit: int,
    preferred_locale: str,
    drop_non_preferred_locale: bool,
) -> tuple[list[dict[str, Any]], BuildStats]:
    rows: list[dict[str, Any]] = []
    stats = BuildStats()
    fallback_idx = 0
    files = sorted([*root.rglob("*.jsonl"), *root.rglob("*.json")])
    for file_path in files:
        if ".git" in file_path.parts:
            continue
        if limit > 0 and len(rows) >= limit:
            break
        if file_path.suffix == ".jsonl":
            iterable = iter_jsonl(file_path)
        else:
            iterable = iter_json(file_path)
        for sample in iterable:
            if limit > 0 and len(rows) >= limit:
                break
            instruction = normalize_text(sample.get("instruction") or sample.get("query") or sample.get("question"))
            if not instruction:
                stats.skipped += 1
                continue
            locale = detect_locale(instruction)
            if drop_non_preferred_locale and not locale_matches_preferred(locale, preferred_locale):
                stats.skipped += 1
                continue
            fallback_idx += 1
            sample_id = sanitize_sample_id(
                sample.get("id") or sample.get("task_id") or sample.get("uid"),
                fallback=f"{dataset_name}_{fallback_idx:06d}",
            )
            workbook_refs = collect_workbook_refs(sample, root)
            context_hints = {
                "instruction_type": sample.get("instruction_type", ""),
                "answer_position": sample.get("answer_position", ""),
                "spreadsheet_path": sample.get("spreadsheet_path", ""),
            }
            expected_output = {
                "golden_workbook_path": sample.get("golden_response_path", ""),
                "answer_position": sample.get("answer_position", ""),
            }
            record = make_record(
                dataset=dataset_name,
                split=split,
                sample_id=sample_id,
                instruction=instruction,
                workbook_refs=workbook_refs,
                source_file=to_rel_or_raw(file_path, root),
                context_hints=context_hints,
                expected_output=expected_output,
                verification="benchmark_golden_available" if workbook_refs else "none",
                notes=[],
                preferred_locale=preferred_locale,
            )
            rows.append(record)
            stats.added += 1
    return rows, stats


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9가-힣]+", "", str(value or "").strip().lower())


def parse_task_xlsx(
    *,
    dataset_name: str,
    xlsx_path: Path,
    split: str,
    limit: int,
    root: Path,
    preferred_locale: str,
    drop_non_preferred_locale: bool,
) -> tuple[list[dict[str, Any]], BuildStats]:
    rows: list[dict[str, Any]] = []
    stats = BuildStats()
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [normalize_header(cell) for cell in header_cells]
        for idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if limit > 0 and len(rows) >= limit:
                break
            row_map = {headers[i]: row_values[i] for i in range(min(len(headers), len(row_values)))}
            instruction = ""
            for key in ("instruction", "query", "question", "taskinstruction", "task", "goal"):
                if normalize_text(row_map.get(key)):
                    instruction = normalize_text(row_map.get(key))
                    break
            if not instruction:
                stats.skipped += 1
                continue
            locale = detect_locale(instruction)
            if drop_non_preferred_locale and not locale_matches_preferred(locale, preferred_locale):
                stats.skipped += 1
                continue
            sample_id = sanitize_sample_id(
                row_map.get("id") or row_map.get("taskid") or row_map.get("tasknumber") or row_map.get("no"),
                fallback=f"{dataset_name}_{idx:06d}",
            )
            workbook_name = normalize_text(
                row_map.get("workbook")
                or row_map.get("workbookname")
                or row_map.get("spreadsheet")
                or row_map.get("filename")
            )
            workbook_refs: list[dict[str, str]] = []
            if workbook_name:
                wb_candidate = (root / workbook_name) if not Path(workbook_name).is_absolute() else Path(workbook_name)
                workbook_refs.append({"role": "input", "path": to_rel_or_raw(wb_candidate, root)})
            context_hints = {
                "category": normalize_text(row_map.get("category") or row_map.get("taskcategory")),
                "task_no": normalize_text(row_map.get("tasknumber") or row_map.get("no")),
            }
            record = make_record(
                dataset=dataset_name,
                split=split,
                sample_id=sample_id,
                instruction=instruction,
                workbook_refs=workbook_refs,
                source_file=to_rel_or_raw(xlsx_path, root),
                context_hints=context_hints,
                verification="none",
                notes=["task_metadata_only"],
                preferred_locale=preferred_locale,
            )
            rows.append(record)
            stats.added += 1
    finally:
        wb.close()
    return rows, stats


def parse_all_events(
    *,
    log_path: Path,
    split: str,
    limit: int,
    root: Path,
    preferred_locale: str,
    drop_non_preferred_locale: bool,
) -> tuple[list[dict[str, Any]], BuildStats]:
    rows: list[dict[str, Any]] = []
    stats = BuildStats()
    for idx, event in enumerate(iter_jsonl(log_path), start=1):
        if limit > 0 and len(rows) >= limit:
            break
        if str(event.get("event_type", "")).strip() != "harness":
            continue
        payload = event.get("payload") if isinstance(event.get("payload"), dict) else {}
        if str(payload.get("route", "")).strip() != "/excel-live/command":
            continue
        message = normalize_text(payload.get("message", ""))
        if not message:
            stats.skipped += 1
            continue
        locale = detect_locale(message)
        if drop_non_preferred_locale and not locale_matches_preferred(locale, preferred_locale):
            stats.skipped += 1
            continue
        action = normalize_text(payload.get("action", ""))
        workbook_id = normalize_text(payload.get("workbook_id", ""))
        workbook_refs: list[dict[str, str]] = []
        if workbook_id:
            workbook_path = (root / workbook_id) if not Path(workbook_id).is_absolute() else Path(workbook_id)
            workbook_refs.append({"role": "input", "path": to_rel_or_raw(workbook_path, root)})
        xlwings_ops = payload.get("xlwings_ops") if isinstance(payload.get("xlwings_ops"), list) else []
        params: dict[str, Any] = {}
        if xlwings_ops:
            first = xlwings_ops[0] if isinstance(xlwings_ops[0], dict) else {}
            if isinstance(first.get("params"), dict):
                params = first.get("params", {})
        action_plan: list[dict[str, Any]] = []
        if action.startswith("excel_live."):
            action_plan = [{"action": action, "params": params}]
        sample_id = sanitize_sample_id(payload.get("session_id"), fallback=f"event_{idx:07d}")
        record = make_record(
            dataset="officeclaw_all_events",
            split=split,
            sample_id=sample_id,
            instruction=message,
            workbook_refs=workbook_refs,
            source_file=str(log_path),
            context_hints={
                "sheet_name": normalize_text(payload.get("sheet_name", "")),
                "reason": normalize_text(payload.get("reason", "")),
                "status_code": payload.get("status_code", 0),
            },
            label_status="log_observed",
            action_plan=action_plan,
            verification="log_observed",
            passed=bool(payload.get("ok", False)),
            confidence=0.75 if action_plan else 0.35,
            preferred_locale=preferred_locale,
        )
        rows.append(record)
        stats.added += 1
    return rows, stats


def upsert_records(target: dict[str, dict[str, Any]], records: Iterable[dict[str, Any]]) -> None:
    for row in records:
        target[row["record_id"]] = row


def print_stats(records: list[dict[str, Any]]) -> None:
    dataset_counter = Counter(row.get("source", {}).get("dataset", "unknown") for row in records)
    status_counter = Counter(row.get("target", {}).get("label_status", "unknown") for row in records)
    locale_counter = Counter(row.get("input", {}).get("locale", "unknown") for row in records)
    preferred_match_counter = Counter(
        bool(row.get("input", {}).get("training_hints", {}).get("preferred_locale_match", False))
        for row in records
    )
    print(f"[STATS] total={len(records)}")
    print(f"[STATS] by_dataset={dict(dataset_counter)}")
    print(f"[STATS] by_label_status={dict(status_counter)}")
    print(f"[STATS] by_locale={dict(locale_counter)}")
    print(f"[STATS] preferred_locale_match={dict(preferred_match_counter)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel distillation 통합 JSONL 생성기")
    parser.add_argument("--spreadsheetbench-root", action="append", type=Path, default=[])
    parser.add_argument("--spreadsheetbench2-root", action="append", type=Path, default=[])
    parser.add_argument("--sheetcopilot-root", action="append", type=Path, default=[])
    parser.add_argument("--sheetrm-root", action="append", type=Path, default=[])
    parser.add_argument("--all-events", action="append", type=Path, default=[])
    parser.add_argument("--split", type=str, default="train")
    parser.add_argument("--preferred-locale", type=str, default="ko")
    parser.add_argument("--drop-non-preferred-locale", action="store_true")
    parser.add_argument("--limit-per-source", type=int, default=0)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--stats", action="store_true")
    args = parser.parse_args()

    merged: dict[str, dict[str, Any]] = {}
    total_stats = BuildStats()

    for root in args.spreadsheetbench_root:
        rows, stats = load_instruction_dataset(
            dataset_name="spreadsheetbench",
            root=root,
            split=args.split,
            limit=args.limit_per_source,
            preferred_locale=args.preferred_locale,
            drop_non_preferred_locale=args.drop_non_preferred_locale,
        )
        upsert_records(merged, rows)
        total_stats.added += stats.added
        total_stats.skipped += stats.skipped
        total_stats.errors += stats.errors
        print(f"[INFO] spreadsheetbench loaded: {stats.added} from {root}")

    for root in args.spreadsheetbench2_root:
        rows, stats = load_instruction_dataset(
            dataset_name="spreadsheetbench2",
            root=root,
            split=args.split,
            limit=args.limit_per_source,
            preferred_locale=args.preferred_locale,
            drop_non_preferred_locale=args.drop_non_preferred_locale,
        )
        upsert_records(merged, rows)
        total_stats.added += stats.added
        total_stats.skipped += stats.skipped
        total_stats.errors += stats.errors
        print(f"[INFO] spreadsheetbench2 loaded: {stats.added} from {root}")

    for root in args.sheetcopilot_root:
        candidates = [root / "dataset" / "dataset.xlsx", root / "dataset.xlsx"]
        target_xlsx = next((fp for fp in candidates if fp.exists()), None)
        if not target_xlsx:
            print(f"[WARN] SheetCopilot dataset.xlsx를 찾지 못했습니다: {root}")
            continue
        rows, stats = parse_task_xlsx(
            dataset_name="sheetcopilot",
            xlsx_path=target_xlsx,
            split=args.split,
            limit=args.limit_per_source,
            root=root,
            preferred_locale=args.preferred_locale,
            drop_non_preferred_locale=args.drop_non_preferred_locale,
        )
        upsert_records(merged, rows)
        total_stats.added += stats.added
        total_stats.skipped += stats.skipped
        total_stats.errors += stats.errors
        print(f"[INFO] sheetcopilot loaded: {stats.added} from {target_xlsx}")

    for root in args.sheetrm_root:
        candidates = [root / "sheetrm" / "tasks.xlsx", root / "tasks.xlsx"]
        target_xlsx = next((fp for fp in candidates if fp.exists()), None)
        if not target_xlsx:
            print(f"[WARN] SheetRM tasks.xlsx를 찾지 못했습니다: {root}")
            continue
        rows, stats = parse_task_xlsx(
            dataset_name="sheetrm",
            xlsx_path=target_xlsx,
            split=args.split,
            limit=args.limit_per_source,
            root=root,
            preferred_locale=args.preferred_locale,
            drop_non_preferred_locale=args.drop_non_preferred_locale,
        )
        upsert_records(merged, rows)
        total_stats.added += stats.added
        total_stats.skipped += stats.skipped
        total_stats.errors += stats.errors
        print(f"[INFO] sheetrm loaded: {stats.added} from {target_xlsx}")

    for log_path in args.all_events:
        rows, stats = parse_all_events(
            log_path=log_path,
            split=args.split,
            limit=args.limit_per_source,
            root=Path.cwd(),
            preferred_locale=args.preferred_locale,
            drop_non_preferred_locale=args.drop_non_preferred_locale,
        )
        upsert_records(merged, rows)
        total_stats.added += stats.added
        total_stats.skipped += stats.skipped
        total_stats.errors += stats.errors
        print(f"[INFO] all_events loaded: {stats.added} from {log_path}")

    records = [merged[key] for key in sorted(merged.keys())]
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as f:
        for row in records:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    print(f"[DONE] output={args.output} records={len(records)}")
    print(
        f"[DONE] raw_added={total_stats.added} skipped={total_stats.skipped} errors={total_stats.errors}"
    )
    if args.stats:
        print_stats(records)


if __name__ == "__main__":
    main()

