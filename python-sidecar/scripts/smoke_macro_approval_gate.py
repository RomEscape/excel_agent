r"""매크로 승인 경로 끝-끝 스모크 — 라운드 5 B2 판정용.

배터리 84개 각본 어디에도 매크로(/macro/step) 경로가 없다(2026-08-25 실측:
전체 *_log.json에서 macro 액션 0건). 그래서 이 스모크가 그 빈자리를 맡는다:

  /command(매크로 문장) → macro_plan 카드 → /macro/step 반복
  → 하위 명령이 **실제 /command 경로**를 approve=True로 통과
  → 이제 블라스트 반경·계획 위생 검사를 **면제받지 않고** 지나간다(B2 수정)

가로채는 것은 분해기 하나뿐이다(LLM 호출이라 비결정적). 하위 명령 실행은
전부 실물 경로다. LLM은 금지 스텁으로 막아 quick rule만으로 도는 것을 증명한다
— 게이트·배터리와 동시에 돌아도 Ollama를 건드리지 않는다.

판정은 API 자기보고가 아니라 **결과 워크북을 다시 열어서** 한다(CLAUDE.md §3-7).

    & $PY scripts\smoke_macro_approval_gate.py
"""
from __future__ import annotations

import json
import os
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

os.environ.setdefault("PYTHONUTF8", "1")
os.environ["EXCEL_LIVE_ENGINE"] = "file"

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from office_claw_sidecar.main import app
from office_claw_sidecar.routers import excel_live as excel_live_router
from office_claw_sidecar.services.excel_macro_planner import MacroStepPlan
from office_claw_sidecar.services.llm_service import get_llm_service

HEADERS = {"Authorization": "Bearer dev-token"}

# 하위 명령: 전부 quick rule이 받는 문장이어야 한다(LLM 금지 스텁이 지켜본다).
SUB_COMMANDS = [
    "A1에 점검항목 이라고 써줘",
    "B1에 상태 라고 써줘",
    "A1:B1 노란색으로 칠해줘",
]


class _NoLLM:
    """LLM이 불리면 그 자체가 실패다 — 스모크는 결정론 경로만 밟아야 한다."""

    def __getattr__(self, name: str):
        def _refuse(*args, **kwargs):
            raise AssertionError(f"LLM 호출 금지인데 {name}()가 불렸다")

        return _refuse


async def _decompose(message, llm_service, **kwargs):
    return [
        MacroStepPlan(index=i, command=c, destructive=False)
        for i, c in enumerate(SUB_COMMANDS, start=1)
    ]


def main() -> int:
    run_id = datetime.now().strftime("%m%d-%H%M%S-macro-gate")
    work = Path(tempfile.mkdtemp(prefix="macro_gate_"))
    xlsx = work / "macro_gate.xlsx"
    wb = Workbook()
    wb.active.title = "점검"
    wb.save(xlsx)

    service = excel_live_router.get_excel_live_service()
    service.select_workbook(str(xlsx))
    service.select_sheet(None, "점검")

    excel_live_router.decompose_macro_request = _decompose
    app.dependency_overrides[get_llm_service] = lambda: _NoLLM()
    excel_live_router._macro_runs.clear()
    client = TestClient(app)

    t0 = time.time()
    resp = client.post(
        "/excel-live/command",
        json={"message": "대시보드 만들어줘", "workbook_id": str(xlsx), "session_id": "smoke-macro"},
        headers=HEADERS,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["action"] == "excel_live.macro_plan", f"매크로 카드가 안 떴다: {body['action']}"
    macro_id = body["result"]["macro_id"]

    statuses: list[str] = []
    for _ in range(len(SUB_COMMANDS) + 2):  # 걸음 수 상한 — 무한 루프 방지
        step = client.post(
            "/excel-live/macro/step", json={"macro_id": macro_id}, headers=HEADERS
        )
        assert step.status_code == 200, step.text
        status = step.json()["result"]["status"]
        statuses.append(status)
        if status in {"done", "halted", "aborted", "waiting_input"}:
            break

    # ---- 판정: 파일을 다시 연다 ----
    out = load_workbook(xlsx)
    ws = out["점검"]
    fills = [str(ws["A1"].fill.start_color.rgb or ""), str(ws["B1"].fill.start_color.rgb or "")]
    checks = {
        "매크로 완주(status=done)": statuses[-1] == "done",
        "A1=점검항목": ws["A1"].value == "점검항목",
        "B1=상태": ws["B1"].value == "상태",
        "A1:B1 노란 채움": all(f.endswith("FFFF00") for f in fills),
    }
    report = {
        "run_id": run_id,
        "statuses": statuses,
        "fills": fills,
        "checks": checks,
        "sub_commands": SUB_COMMANDS,
        "초": round(time.time() - t0, 2),
    }
    out_dir = Path(__file__).resolve().parent.parent.parent / "logs" / "e2e"
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / f"{run_id}.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    ok = all(checks.values())
    for name, passed in checks.items():
        print(("✅" if passed else "❌"), name)
    print(f"run_id={run_id} · statuses={statuses} · {report['초']}초")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
