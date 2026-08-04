"""하드 팩 6건의 정답 계획을 수동으로 만들고, 실제 데모 통합문서에 실행해 검증한다.

`teacher_label_action_plan.py`는 우리 자신의 플래너로 라벨을 만드는 구조라, 플래너가
아예 계획을 못 세우는 문장(암시적 의도)은 이 스크립트로도 라벨링할 수 없다 — 모르는 걸
스스로 라벨링할 수는 없다. 이 6건은 시나리오를 설계하면서 정답을 이미 알고 있으므로
직접 작성하고, 원문 표현을 패러프레이즈로 늘려 학습 신호를 넓힌다.

각 후보 계획은 실제 데모 통합문서 복사본에 실행해서 하드 팩의 오라클 assertion을
통과하는지 검증한 뒤에만 `label_status=verified`로 기록한다 — 검증 실패한 계획을
정답으로 학습시키면 오히려 문제를 더 굳힌다.
"""

from __future__ import annotations

import argparse
import copy
import json
import os
import shutil
import tempfile
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

os.environ.setdefault("EXCEL_LIVE_ENGINE", "file")

from openpyxl import load_workbook

from office_claw_sidecar.routers.excel_live import (
    _execute_action,
    _verify_step_result,
)
from office_claw_sidecar.services.excel_live_executor import (
    execute_plan,
    normalize_plan_steps,
)
from office_claw_sidecar.services.excel_live_plan_validator import (
    ValidationContext,
    validate_plan,
)

KST = timezone(timedelta(hours=9), name="KST")
DEMO_TEMPLATE = Path("복잡한 엑셀 작업을 위한 자료/AI_Excel_Automation_Demo.xlsx")
DEFAULT_SHEET = "Sales_Data"

# 하드 팩의 assertion 체커를 그대로 재사용한다. 정답 계획이 실제로 오라클을 만족하는지
# 확인하는 것이 검증의 핵심이라, 두 벌로 다시 구현하면 기준이 어긋날 수 있다.
from verify_excel_complex_scenarios import _check_assertion


def _text(value: Any) -> str:
    return str(value or "").strip()


# ── 하드 케이스 6건의 정답 계획 ────────────────────────────────────────────
# turns[i].variants: 같은 의도를 표현하는 문장 목록(패러프레이즈).
# turns[i].plan: 그 턴에서 실행해야 하는 action_plan (실제 파라미터로 이미 확정된 상태).
# turns[i].sheet_name: 그 턴 호출 시점의 컨텍스트 시트(멀티턴에서 이전 결과 시트로 옮겨갈 수 있음).

HARD_CASES: list[dict[str, Any]] = [
    {
        "id": "hard-005-category-pivot-rough",
        "turns": [
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "제품 분류마다 매출이 얼마나 나오는지 Cat_Sum 시트로 뽑아줘",
                    "카테고리 기준으로 매출 합계를 Cat_Sum에 정리해줘",
                    "분류별로 매출 얼마나 나오는지 Cat_Sum 시트 만들어서 보여줘",
                    "제품군마다 매출 합쳐서 Cat_Sum 시트에 뽑아줘",
                    "품목 분류 기준 매출 집계를 Cat_Sum으로 만들어줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.pivot_table",
                        "params": {
                            "source_range": "__ACTIVE_SELECTION__",
                            "source_sheet": "Sales_Data",
                            "row_field": "Category",
                            "value_field": "Sales",
                            "agg": "sum",
                            "output_sheet": "Cat_Sum",
                            "output_start": "A1",
                        },
                        "reason": "제품 분류(Category)별 매출(Sales) 합계를 Cat_Sum 시트에 집계",
                    }
                ],
            }
        ],
    },
    {
        "id": "hard-007-cross-sheet-total",
        "turns": [
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "Summary 시트 만들어서 A1에 총매출이라고 쓰고 B1에 Sales_Data 매출 합계 수식 넣어줘",
                    "Summary 시트를 새로 만들고 A1엔 총매출, B1엔 매출 총합 수식을 넣어줘",
                    "요약용 Summary 시트 만들어줘. A1에 총매출 적고 B1에 전체 매출 합계 수식 넣어줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.create_sheet",
                        "params": {"sheet_name": "Summary"},
                        "reason": "요약 시트 생성",
                    },
                    {
                        "action": "excel_live.write_range",
                        "params": {
                            "sheet_name": "Summary",
                            "start_cell": "A1",
                            "values_2d": [["총매출"]],
                        },
                        "reason": "A1에 라벨 입력",
                    },
                    {
                        "action": "excel_live.set_formula",
                        "params": {
                            "sheet_name": "Summary",
                            "range_ref": "B1",
                            "formula_a1": "=SUM(Sales_Data!L2:L181)",
                        },
                        "reason": "Sales_Data 매출(L열) 합계 수식",
                    },
                ],
            }
        ],
    },
    {
        "id": "hard-011-sort-desc-rough",
        "turns": [
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "제일 많이 팔린 순서대로 위에서부터 보이게 해줘",
                    "많이 팔린 것부터 순서대로 정렬해줘",
                    "판매량 많은 순으로 위에 오게 해줘",
                    "제일 많이 나간 것부터 보이도록 정리해줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.sort_range",
                        "params": {
                            "target_range": "__ACTIVE_SELECTION__",
                            "key_column": "Qty",
                            "order": "desc",
                            "has_header": True,
                        },
                        "reason": "판매 수량(Qty) 기준 내림차순 정렬",
                    }
                ],
            }
        ],
    },
    {
        "id": "hard-012-filter-then-aggregate",
        "turns": [
            {
                "sheet_name": "Sales_Data",
                "variants": ["취소된 건은 지워줘", "취소 주문은 없애줘", "취소된 것들은 다 지워줘"],
                "plan": [
                    {
                        "action": "excel_live.filter_rows",
                        "params": {
                            "target_range": "__ACTIVE_SELECTION__",
                            "column": "Status",
                            "operator": "==",
                            "value": "취소",
                            "has_header": True,
                            "mode": "remove",
                        },
                        "reason": "상태(Status)가 취소인 행 삭제",
                    }
                ],
            },
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "그리고 지역별 매출 합계를 Live_Sum 시트에 정리해줘",
                    "지역 기준으로 매출 합계를 Live_Sum에 뽑아줘",
                    "지역마다 매출 합쳐서 Live_Sum 시트로 만들어줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.pivot_table",
                        "params": {
                            "source_range": "__ACTIVE_SELECTION__",
                            "source_sheet": "Sales_Data",
                            "row_field": "Region",
                            "value_field": "Sales",
                            "agg": "sum",
                            "output_sheet": "Live_Sum",
                            "output_start": "A1",
                        },
                        "reason": "지역(Region)별 매출(Sales) 합계를 Live_Sum 시트에 집계",
                    }
                ],
            },
        ],
    },
    {
        "id": "hard-013-add-column-formula",
        "turns": [
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "맨 뒤에 원가총액 열 하나 추가해줘",
                    "제일 뒤에 원가총액이라는 열 만들어줘",
                    "마지막에 원가총액 열 하나 추가해줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.add_column",
                        "params": {"sheet_name": "Sales_Data", "name": "원가총액", "formula_a1": None},
                        "reason": "표 맨 뒤에 원가총액 열 추가",
                    }
                ],
            },
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "거기에 수량 곱하기 단위원가 수식 넣어줘",
                    "수량하고 단위원가 곱한 값을 그 열에 넣어줘",
                    "그 열에 수량 x 단위원가 수식을 채워줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.set_formula",
                        "params": {
                            "sheet_name": "Sales_Data",
                            "range_ref": "R2:R181",
                            "formula_a1": "=I2*M2",
                        },
                        "reason": "수량(Qty, I열) x 단위원가(Unit_Cost, M열)를 원가총액(R열)에 채움",
                    }
                ],
            },
        ],
    },
    {
        "id": "hard-014-salesperson-chart",
        "turns": [
            {
                "sheet_name": "Sales_Data",
                "variants": [
                    "누가 얼마나 팔았는지 Rep_Chart 시트에 정리해줘",
                    "담당자별로 얼마나 팔았는지 Rep_Chart로 뽑아줘",
                    "영업사원마다 판매 실적을 Rep_Chart 시트에 만들어줘",
                ],
                "plan": [
                    {
                        "action": "excel_live.pivot_table",
                        "params": {
                            "source_range": "__ACTIVE_SELECTION__",
                            "source_sheet": "Sales_Data",
                            "row_field": "Salesperson",
                            "value_field": "Sales",
                            "agg": "sum",
                            "output_sheet": "Rep_Chart",
                            "output_start": "A1",
                        },
                        "reason": "담당자(Salesperson)별 매출(Sales) 합계를 Rep_Chart 시트에 집계",
                    }
                ],
            },
            {
                "sheet_name": "Rep_Chart",
                "variants": ["이걸로 막대 그래프 하나 그려줘", "이 표로 막대 차트 그려줘", "지금 표 가지고 바 차트 만들어줘"],
                "plan": [
                    {
                        "action": "excel_live.create_chart",
                        "params": {
                            "source_range": "A1:B7",
                            "chart_type": "bar",
                            "title": "담당자별 매출",
                            "output_sheet": "Rep_Chart",
                        },
                        "reason": "직전 집계 표(A1:B7)로 막대 차트 생성",
                    }
                ],
            },
        ],
    },
]


def _run_plan_turn(
    workbook_id: str, sheet_name: str, plan: list[dict[str, Any]], message: str
) -> tuple[bool, str]:
    try:
        normalized_steps = normalize_plan_steps(plan)
    except Exception as exc:  # noqa: BLE001
        return False, f"normalize_failed:{exc}"
    if not normalized_steps:
        return False, "normalize_empty"

    try:
        validated_steps = validate_plan(
            normalized_steps,
            context=ValidationContext(message=message, workbook_id=workbook_id, sheet_name=sheet_name),
        )
    except Exception as exc:  # noqa: BLE001
        return False, f"validate_failed:{exc}"

    execution = execute_plan(
        steps=validated_steps,
        max_attempts=1,
        abort_on_failure=True,
        execute_action=lambda action, params: _execute_action(
            action=action, params=params, workbook_id=workbook_id, sheet_name=sheet_name
        ),
        verify_step=lambda action, params, result: _verify_step_result(
            action=action, params=params, result=result, workbook_id=workbook_id, sheet_name=sheet_name
        ),
    )
    last = execution.last
    if last is None:
        return False, "execution_empty"
    if last.error:
        return False, f"execution_error:{last.error}"
    if not last.verified:
        return False, f"not_verified:{last.verify_detail}"
    return True, "ok"


def _load_hard_pack_assertions(pack_path: Path, scenario_id: str) -> list[dict[str, Any]]:
    pack = json.loads(pack_path.read_text(encoding="utf-8"))
    for scenario in pack.get("scenarios", []):
        if scenario.get("id") == scenario_id:
            oracle = scenario.get("oracle") or {}
            result = oracle.get("result") or {}
            return list(result.get("assertions") or [])
    return []


def verify_case(case: dict[str, Any], pack_path: Path) -> tuple[bool, list[str]]:
    """실제 데모 통합문서 복사본에 순서대로 실행하고 오라클 assertion까지 확인한다."""
    template = DEMO_TEMPLATE if DEMO_TEMPLATE.is_absolute() else (Path(__file__).resolve().parents[2] / DEMO_TEMPLATE)
    notes: list[str] = []

    with tempfile.TemporaryDirectory(prefix="officeclaw_manual_target_") as td:
        workbook_path = Path(td) / "verify.xlsx"
        shutil.copy2(template, workbook_path)
        workbook_id = str(workbook_path.resolve())

        for turn_idx, turn in enumerate(case["turns"]):
            ok, detail = _run_plan_turn(
                workbook_id, turn["sheet_name"], turn["plan"], turn["variants"][0]
            )
            if not ok:
                notes.append(f"turn[{turn_idx}] execution_failed: {detail}")
                return False, notes
            notes.append(f"turn[{turn_idx}] executed ok")

        assertions = _load_hard_pack_assertions(pack_path, case["id"])
        if not assertions:
            notes.append("no_assertions_found_in_pack")
            return False, notes

        wb = load_workbook(workbook_path, data_only=False)
        try:
            all_ok = True
            for assertion in assertions:
                ok, detail = _check_assertion(wb, assertion, workbook_path)
                notes.append(("ok " if ok else "NG ") + detail)
                all_ok = all_ok and ok
        finally:
            wb.close()
        return all_ok, notes


def _build_records(case: dict[str, Any]) -> list[dict[str, Any]]:
    """검증을 통과한 케이스를 excel_distill.v1 레코드로 펼친다.

    턴이 여러 개인 시나리오는 턴마다 별도 레코드로 만든다 — 실제 서비스도 매 호출이
    독립적인 (문장, 그 시점 컨텍스트) 입력을 받으므로 학습 단위를 거기에 맞춘다.
    """
    records: list[dict[str, Any]] = []
    now = datetime.now(KST).isoformat()
    for turn_idx, turn in enumerate(case["turns"]):
        for variant in turn["variants"]:
            record_id = f"manual_hard_case:{case['id']}:t{turn_idx}:{uuid.uuid4().hex[:10]}"
            records.append(
                {
                    "schema_version": "excel_distill.v1",
                    "record_id": record_id,
                    "source": {
                        "dataset": "manual_hard_case_v1",
                        "split": "train",
                        "sample_id": f"{case['id']}:t{turn_idx}",
                        "license": "internal",
                        "provenance": {"source_file": "datasets/excel_demo_workbook_hard_v1.json"},
                    },
                    "input": {
                        "instruction": variant,
                        "locale": "ko",
                        "workbook_refs": [],
                        "context_hints": {
                            "sheet_name": turn["sheet_name"],
                            "reason": "",
                            "status_code": 200,
                        },
                    },
                    "target": {
                        "task_type": "spreadsheet_edit",
                        "label_status": "verified",
                        "action_plan": copy.deepcopy(turn["plan"]),
                        "expected_output": {},
                    },
                    "quality": {
                        "verification": "manual_execution_replay",
                        "passed": True,
                        "confidence": 0.97,
                    },
                    "metadata": {
                        "created_at": now,
                        "generator": "python-sidecar/scripts/build_hard_case_manual_targets.py",
                        "notes": [f"scenario_id:{case['id']}", f"turn_index:{turn_idx}"],
                    },
                }
            )
    return records


def main() -> int:
    parser = argparse.ArgumentParser(description="하드 팩 정답 계획 검증 및 학습 레코드 생성")
    parser.add_argument(
        "--pack",
        type=Path,
        default=Path("../datasets/excel_demo_workbook_hard_v1.json"),
    )
    parser.add_argument(
        "--output-jsonl",
        type=Path,
        default=Path("../datasets/distill/excel_hard_manual_v1.jsonl"),
    )
    parser.add_argument("--only", type=str, default="", help="쉼표로 구분한 scenario id 필터")
    args = parser.parse_args()

    only = {s.strip() for s in args.only.split(",") if s.strip()}
    cases = [c for c in HARD_CASES if not only or c["id"] in only]

    all_records: list[dict[str, Any]] = []
    total_variants = 0
    for case in cases:
        ok, notes = verify_case(case, args.pack)
        status = "PASS" if ok else "FAIL"
        print(f"[{status}] {case['id']}")
        for note in notes:
            print(f"    {note}")
        if ok:
            records = _build_records(case)
            all_records.extend(records)
            total_variants += len(records)
        else:
            print("    -> 학습 레코드 생성 건너뜀 (검증 실패)")

    args.output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with args.output_jsonl.open("w", encoding="utf-8") as f:
        for record in all_records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    print(f"[DONE] cases={len(cases)} verified_cases={len({r['metadata']['notes'][0] for r in all_records})} "
          f"records={total_variants} output={args.output_jsonl}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
