"""같은 명령을 여러 번 태워 어디서 무엇이 깨지는지 로그로 남긴다.

## 왜 또 만드나

`scripts/run_command_battery.py`는 **떠 있는 사이드카**에 HTTP로 던진다. 실사용에
가장 가깝지만 사이드카를 띄워야 하고, 매번 다른 워크스페이스라 재현이 어렵다.
`approval_gate.py`는 승인 경로 **하나**만 잰다.

여기는 그 사이다. 앱을 프로세스 안에서 그대로 띄우고(TestClient), 케이스마다 격리된
임시 워크북을 만들고, 같은 케이스를 N번 반복한다. 목적은 점수가 아니라 **로그**다.
턴 하나하나가 `logs/diagnostics/<실행id>.jsonl`에 쌓이고, `trace_digest`가 그것을
케이스 단위로 접어 "항상 깨짐"과 "들쭉날쭉"을 갈라 준다.

## 반복이 왜 필요한가

플래너는 로컬 LLM이라 같은 문장에 같은 계획을 준다는 보장이 없다. 한 번 돌려서 나온
실패를 코드에서 찾기 시작하면, 실제 원인이 모델의 변덕일 때 끝없이 헤맨다. 반대로
매번 똑같이 깨지는 것은 코드에 원인이 있다. 이 둘을 먼저 갈라야 한다.

## 무엇을 격리하는가

케이스마다 임시 디렉터리에 워크북을 새로 만든다. 앞 케이스가 남긴 시트나 값이 다음
케이스의 관측에 섞이면, 로그를 봐도 무엇을 보고 판단한 것인지 알 수 없다. 반복
회차끼리도 마찬가지로 격리한다 — 1회차가 쓴 값을 2회차가 보면 같은 명령이 아니다.

## 무엇으로 채점하는가

세 겹이다. 각 겹이 앞 겹은 못 보는 실패를 본다.

1. **판정**(`trace_report`) — 터졌는가, 되물었는가. 실행기 오류를 본다.
2. **기대 액션**(`expect_action`) — 요청한 액션을 실행하긴 했는가. 차트를 만들라
   했는데 피벗만 만들고 성공을 보고하는 경우를 잡는다.
3. **결과 상태**(`expect_effect`) — 파일이 실제로 사용자가 원한 대로 됐는가.

3번이 필요한 이유는 앞 두 겹이 **계획을 기준으로** 채점하기 때문이다. 검증기는
"인자대로 실행됐는가"를 보지 "인자가 옳았는가"를 보지 않는다(그 모듈 docstring이
명시한 경계다). 그래서 계획이 틀렸는데 그 틀린 계획대로 정확히 실행되면 모든
겹이 통과한다 — 사용자만 틀린 결과를 본다. `expect_effect`는 계획을 거치지 않고
결과 파일만 열어 보므로 그 구멍을 메운다.
"""

from __future__ import annotations

import tempfile
import uuid
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from office_claw_sidecar.services import decision_trace

HEADERS = {"Authorization": "Bearer dev-token"}

# 결과 워크북을 받아 어긋난 점을 한 줄로 돌려주는 함수. 맞으면 빈 문자열.
EffectOracle = Callable[[Any], str]


@dataclass(frozen=True)
class BatteryCase:
    """한 문장과, 그 문장을 던질 시트 상태."""

    case_id: str
    message: str
    sheet: str = "매출"
    rows: list[list[Any]] = field(default_factory=list)
    # 승인이 필요한 명령을 승인까지 태울 것인가. 실사용은 사용자가 누른다.
    approve: bool = True
    # 이 문장이라면 반드시 실행돼야 하는 액션. 다른 것을 대신 실행하고 성공을
    # 보고하는 경우를 잡는다 — 검증기는 실행한 액션의 사후조건만 보므로 못 잡는다.
    # 요청이 어느 액션으로 가야 하는지 논쟁의 여지가 없을 때만 채운다.
    expect_action: str = ""
    # 실행이 끝난 워크북을 열어 사용자 요청이 이뤄졌는지 직접 본다.
    # 계획을 경유하지 않는 유일한 채점이라, 계획 자체가 틀린 경우를 여기서만 잡는다.
    expect_effect: EffectOracle | None = None
    tags: tuple[str, ...] = field(default_factory=tuple)


def _sales_rows() -> list[list[Any]]:
    return [
        ["코드", "지역", "금액", "날짜"],
        ["A-001", "서울", 520, "2026-01-05"],
        ["A-002", "부산", 180, "2026-01-07"],
        ["A-003", "서울", 340, "2026-01-11"],
        ["A-002", "대구", 90, "2026-01-15"],
        ["A-004", "서울", 610, "2026-02-02"],
        ["A-005", "부산", 250, "2026-02-08"],
        ["A-006", "대구", 430, "2026-02-19"],
        ["A-003", "서울", 275, "2026-03-03"],
    ]


# ── 결과 상태 오라클 ─────────────────────────────────────────────────────────
#
# 계획을 보지 않는다. 결과 파일만 열어 사용자가 말한 것이 됐는지 본다.
# 그래서 "계획은 일관되게 실행됐지만 계획 자체가 요청과 달랐던" 경우를 잡는다.

_AMOUNTS = [row[2] for row in _sales_rows()[1:]]


def _body(sheet: Any, *, cols: int = 4) -> list[list[Any]]:
    """머리글을 뺀 데이터 행. 완전히 빈 행은 버린다."""
    rows = []
    for row in sheet.iter_rows(min_row=2, max_col=cols, values_only=True):
        if any(cell is not None and str(cell).strip() for cell in row):
            rows.append(list(row))
    return rows


def _has_fill(cell: Any) -> bool:
    """배경색이 실제로 칠해졌는지. openpyxl 기본값은 patternType이 없다."""
    fill = getattr(cell, "fill", None)
    if fill is None or not getattr(fill, "patternType", None):
        return False
    rgb = str(getattr(getattr(fill, "start_color", None), "rgb", "") or "")
    return rgb not in {"", "00000000"}


def _sheet(book: Any, name: str) -> Any:
    return book[name] if name in book.sheetnames else None


def _expect_cell(sheet_name: str, ref: str, want: Any) -> EffectOracle:
    def check(book: Any) -> str:
        sheet = _sheet(book, sheet_name)
        if sheet is None:
            return f"{sheet_name} 시트가 없다"
        got = sheet[ref].value
        if str(got).strip() != str(want).strip():
            return f"{ref}가 {want!r}이어야 하는데 {got!r}이다"
        return ""

    return check


def _expect_sorted_by_amount_desc(sheet_name: str) -> EffectOracle:
    """금액 열이 내림차순인가. 어떤 액션으로 했든 결과만 본다."""

    def check(book: Any) -> str:
        sheet = _sheet(book, sheet_name)
        if sheet is None:
            return f"{sheet_name} 시트가 없다"
        amounts = [row[2] for row in _body(sheet) if isinstance(row[2], (int, float))]
        if len(amounts) < len(_AMOUNTS):
            return f"데이터 행이 {len(_AMOUNTS)}개여야 하는데 {len(amounts)}개다 (행이 사라졌다)"
        if amounts != sorted(amounts, reverse=True):
            return f"금액이 내림차순이 아니다: {amounts}"
        return ""

    return check


def _expect_only_amounts_at_least(sheet_name: str, minimum: int) -> EffectOracle:
    """조건에 맞는 행만 남았는가. 다 남거나 다 지워진 경우를 함께 잡는다."""
    kept = sorted((a for a in _AMOUNTS if a >= minimum), reverse=True)

    def check(book: Any) -> str:
        sheet = _sheet(book, sheet_name)
        if sheet is None:
            return f"{sheet_name} 시트가 없다"
        amounts = sorted(
            (row[2] for row in _body(sheet) if isinstance(row[2], (int, float))), reverse=True
        )
        if amounts != kept:
            return f"{minimum} 이상인 {kept}만 남아야 하는데 {amounts}가 남았다"
        return ""

    return check


def _expect_unique_codes(sheet_name: str) -> EffectOracle:
    """코드가 유일해졌는가. 아무것도 안 지운 경우와 너무 많이 지운 경우를 함께 잡는다."""
    want = sorted({str(row[0]) for row in _sales_rows()[1:]})

    def check(book: Any) -> str:
        sheet = _sheet(book, sheet_name)
        if sheet is None:
            return f"{sheet_name} 시트가 없다"
        codes = [str(row[0]) for row in _body(sheet) if row[0]]
        if sorted(codes) != want:
            return f"코드가 {want}여야 하는데 {sorted(codes)}다"
        return ""

    return check


def _expect_fill_only_below(sheet_name: str, limit: int) -> EffectOracle:
    """기준 미만인 행만 칠해졌는가.

    이 오라클이 핵심이다. 플래너가 threshold를 빠뜨리면 검증기가 0을 채워 넣어
    조건이 사실상 참이 되고 전 행이 칠해지는데, 실행기는 "칠한 셀 1개 이상"이라
    성공을 보고하고 사후조건 검증도 통과한다. 결과 파일을 봐야만 보인다.
    """

    def check(book: Any) -> str:
        sheet = _sheet(book, sheet_name)
        if sheet is None:
            return f"{sheet_name} 시트가 없다"
        painted, missed = [], []
        for offset, amount in enumerate(_AMOUNTS):
            row = offset + 2
            lit = any(_has_fill(sheet.cell(row=row, column=col)) for col in range(1, 5))
            if lit and amount >= limit:
                painted.append(amount)
            if not lit and amount < limit:
                missed.append(amount)
        if painted:
            return f"{limit} 미만만 칠해야 하는데 {sorted(painted)}도 칠했다"
        if missed:
            return f"{limit} 미만인 {sorted(missed)}를 칠하지 않았다"
        return ""

    return check


def _expect_chart(sheet_name: str) -> EffectOracle:
    """차트 도형이 실제로 얹혔는가. 어느 시트든 하나라도 있으면 된다."""

    def check(book: Any) -> str:
        found = sum(len(getattr(book[name], "_charts", []) or []) for name in book.sheetnames)
        if found:
            return ""
        made = [n for n in book.sheetnames if n != sheet_name]
        extra = f" (대신 만든 시트: {', '.join(made)})" if made else ""
        return f"차트가 하나도 없다{extra}"

    return check


def _expect_headers(sheet_name: str, headers: list[str]) -> EffectOracle:
    def check(book: Any) -> str:
        sheet = _sheet(book, sheet_name)
        if sheet is None:
            return f"{sheet_name} 시트가 없다"
        got = [
            str(sheet.cell(row=1, column=col).value or "").strip()
            for col in range(1, len(headers) + 1)
        ]
        if got != headers:
            return f"머리글이 {headers}여야 하는데 {got}다"
        return ""

    return check


# 기본 동작부터 복합까지 한 줄씩. 실패했을 때 원인을 가르기 쉬운 순서로 둔다 —
# 앞쪽이 깨지면 뒤쪽 결과는 볼 필요가 없다.
ALL_CASES: list[BatteryCase] = [
    BatteryCase(
        "값입력",
        "H3에 120 입력해줘",
        rows=_sales_rows(),
        expect_effect=_expect_cell("매출", "H3", 120),
        tags=("write_range",),
    ),
    BatteryCase("값지우기", "H3 지워줘", rows=_sales_rows(), tags=("clear_range",)),
    BatteryCase(
        "정렬-명확",
        "금액 열 기준 내림차순으로 정렬해줘",
        rows=_sales_rows(),
        expect_effect=_expect_sorted_by_amount_desc("매출"),
        tags=("sort_range",),
    ),
    BatteryCase("정렬-모호", "정렬해줘", rows=_sales_rows(), tags=("sort_range", "ambiguous")),
    BatteryCase(
        "필터",
        "금액이 300 이상인 행만 남겨줘",
        rows=_sales_rows(),
        expect_effect=_expect_only_amounts_at_least("매출", 300),
        tags=("filter_rows",),
    ),
    BatteryCase(
        "중복제거",
        "코드 열 기준으로 중복 제거해줘",
        rows=_sales_rows(),
        expect_effect=_expect_unique_codes("매출"),
        tags=("dedupe_rows",),
    ),
    BatteryCase("수식", "H1에 금액 열 합계 수식 넣어줘", rows=_sales_rows(), tags=("set_formula",)),
    BatteryCase(
        "집계",
        "지역별 금액 합계를 집계해서 지역요약 시트에 만들어줘",
        rows=_sales_rows(),
        tags=("group_by_aggregate",),
    ),
    # 기준값을 말했는데 그대로 지켜지는지 본다. 플래너가 threshold를 빠뜨리면
    # 검증기 기본값이 조건을 참으로 만들어 전 행이 칠해지는데, 응답은 성공이다.
    BatteryCase(
        "조건서식",
        "금액이 300 미만인 건은 노란색으로 표시해줘",
        rows=_sales_rows(),
        expect_effect=_expect_fill_only_below("매출", 300),
        tags=("format", "threshold"),
    ),
    BatteryCase("테두리", "A1:D1에 테두리 넣어줘", rows=_sales_rows(), tags=("apply_border",)),
    BatteryCase(
        "표만들기",
        "2행 2열 표 만들어줘. 머리글은 이름, 점수",
        sheet="Sheet1",
        expect_action="excel_live.create_table",
        expect_effect=_expect_headers("Sheet1", ["이름", "점수"]),
        tags=("create_table",),
    ),
    BatteryCase(
        "차트",
        "지역별 금액 막대 차트 만들어줘",
        rows=_sales_rows(),
        expect_action="excel_live.create_chart",
        expect_effect=_expect_chart("매출"),
        tags=("create_chart",),
    ),
]


@dataclass
class RunOutcome:
    """한 번의 호출 결과. 판정은 로그에서 뽑으므로 여기는 응답과 결과 상태만 담는다."""

    case_id: str
    run: int
    status: int
    ok: Any
    action: str
    ask: bool
    approval_required: bool
    reason: str
    elapsed_ms: int
    error: str = ""
    # 결과 파일이 사용자 요청과 어긋난 점. 빈 문자열이면 맞다.
    # 오라클이 없는 케이스도 빈 문자열이므로 `has_oracle`과 같이 봐야 한다.
    effect_error: str = ""
    has_oracle: bool = False


def _isolate(monkeypatch, root: Path) -> None:
    """파일 엔진이 이 디렉터리만 보게 한다. 케이스끼리 상태가 섞이지 않도록."""
    from office_claw_sidecar.services import excel_live_file_service as file_service
    from office_claw_sidecar.services import excel_live_service as live_service

    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
    monkeypatch.setattr(file_service, "WORKSPACE_ROOT", root)
    monkeypatch.setattr(live_service, "_excel_live_service", None)
    monkeypatch.setattr(live_service, "_excel_live_service_engine", None)


def _make_workbook(root: Path, case: BatteryCase) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = case.sheet
    for row in case.rows:
        worksheet.append(row)
    path = root / f"{case.case_id}.xlsx"
    workbook.save(path)
    return path


def run_once(client, monkeypatch, case: BatteryCase, *, run: int, suite: str) -> RunOutcome:
    """케이스 하나를 격리 워크스페이스에서 한 번 태운다.

    턴 로그에 `source`로 출처를 붙인다. 이게 없으면 이 트래픽이 사람이 친 명령과
    같은 파일에서 구분되지 않는다.
    """
    import time

    with monkeypatch.context() as patch:
        root = Path(tempfile.mkdtemp(prefix="oc-battery-"))
        _isolate(patch, root)
        workbook = _make_workbook(root, case)

        payload = {
            "message": case.message,
            "workbook_id": str(workbook),
            "sheet_name": case.sheet,
            "session_id": f"battery-{uuid.uuid4().hex[:8]}",
            "approve": case.approve,
        }

        started = time.perf_counter()
        error = ""
        status, body = 0, {}
        # 기대 액션을 턴에 붙여 보낸다. 로그가 스스로 채점 기준을 들고 있어야
        # 나중에 쌓인 파일만 다시 읽어도 같은 판정이 나온다.
        with decision_trace.source(
            kind="diagnostic",
            suite=suite,
            case=case.case_id,
            run=run,
            expect=case.expect_action or None,
        ):
            try:
                response = client.post("/excel-live/command", json=payload, headers=HEADERS)
                status = response.status_code
                body = response.json() if status == 200 else {}
            except Exception as exc:  # noqa: BLE001 - 진단 도구라 어떤 실패든 남겨야 한다
                error = f"{type(exc).__name__}: {exc}"
        elapsed = int((time.perf_counter() - started) * 1000)

        # 워크북은 이 블록을 벗어나면 지워질 수 있으므로 여기서 열어 본다.
        effect_error = _check_effect(case, workbook) if case.expect_effect else ""

    result = body.get("result") or {}
    return RunOutcome(
        case_id=case.case_id,
        run=run,
        status=status,
        ok=body.get("ok"),
        action=str(body.get("action", "")),
        ask=bool(result.get("ask_follow_up")),
        approval_required=bool(body.get("approval_required")),
        reason=str(body.get("reason", ""))[:160],
        elapsed_ms=elapsed,
        error=error,
        effect_error=effect_error,
        has_oracle=case.expect_effect is not None,
    )


def _check_effect(case: BatteryCase, workbook: Path) -> str:
    """결과 워크북을 열어 사용자 요청이 실제로 이뤄졌는지 본다.

    오라클이 터지는 것과 요청이 안 이뤄진 것은 다르다. 전자를 후자로 보고하면
    도구 버그를 제품 결함으로 착각하므로 구분해서 남긴다.
    """
    oracle = case.expect_effect
    if oracle is None:
        return ""
    try:
        book = load_workbook(workbook, data_only=False)
    except Exception as exc:  # noqa: BLE001 - 진단 도구
        return f"[오라클 오류] 결과 파일을 열지 못함: {type(exc).__name__}: {exc}"
    try:
        return oracle(book)
    except Exception as exc:  # noqa: BLE001 - 진단 도구
        return f"[오라클 오류] {type(exc).__name__}: {exc}"
    finally:
        book.close()


def run_all(
    client,
    monkeypatch,
    *,
    repeats: int = 3,
    suite: str = "command-battery",
    cases: list[BatteryCase] | None = None,
    on_result=None,
) -> list[RunOutcome]:
    """전 케이스를 `repeats`번 돌린다.

    회차를 바깥 루프에 둔다. 케이스별로 몰아서 돌리면 그 사이에 생긴 환경 변화(모델
    로딩, 캐시)가 특정 케이스에만 몰려 회차 간 비교가 무의미해진다.
    """
    picked = cases if cases is not None else ALL_CASES
    outcomes: list[RunOutcome] = []
    for run in range(1, repeats + 1):
        for case in picked:
            outcome = run_once(client, monkeypatch, case, run=run, suite=suite)
            outcomes.append(outcome)
            if on_result is not None:
                on_result(outcome)
    return outcomes
