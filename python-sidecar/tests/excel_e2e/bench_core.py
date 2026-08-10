"""Excel 종단 벤치마크 — 격리 실행과 결과 단언.

## 이 벤치마크가 기존 평가와 다른 점

`datasets/eval/planner_eval_v1.jsonl`(154건)은 **액션 이름만** 채점한다.
`sort_range`를 골랐지만 엉뚱한 열로 정렬하는 계획도 정답으로 친다.

여기서는 실제 .xlsx 파일을 만들고, 명령을 태우고, **저장된 파일을 다시 읽어**
셀 값이 맞는지 본다. "매출 높은 순으로 정렬해줘"의 정답은 액션 이름이 아니라
매출 열이 실제로 내림차순이 된 파일이다.

## 단언은 서비스를 거치지 않는다

결과 확인에 `ExcelLiveService.read_range()`를 쓰면, 서비스에 버그가 있을 때
실행과 검증이 같은 버그를 공유해서 통과해 버린다. 그래서 단언은 openpyxl로
파일을 직접 연다.

## 격리

`EXCEL_LIVE_ENGINE=file`로 openpyxl 엔진을 강제하므로 Excel 설치도, 실행 중인
Excel 창도 필요 없다. 워크스페이스는 임시 디렉터리로 갈아끼워 사용자의 실제
파일을 건드리지 않는다.
"""

from __future__ import annotations

import os
import shutil
import tempfile
from collections.abc import Callable, Iterator
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# ── 격리된 워크스페이스 ──────────────────────────────────────────────────


@contextmanager
def isolated_workspace() -> Iterator[Path]:
    """임시 워크스페이스에서 파일 엔진을 쓰도록 서비스를 갈아끼운다.

    `excel_live_file_service`가 import 시점에 WORKSPACE_ROOT를 바인딩하므로
    sandbox 쪽만 바꿔서는 통하지 않는다. 모듈 속성을 직접 교체해야 한다.
    """
    from office_claw_sidecar.services import excel_live_file_service as file_service
    from office_claw_sidecar.services import excel_live_service as live_service

    root = Path(tempfile.mkdtemp(prefix="oc-bench-"))
    prev_engine = os.environ.get("EXCEL_LIVE_ENGINE")
    prev_root = file_service.WORKSPACE_ROOT
    prev_singleton = live_service._excel_live_service
    prev_engine_name = live_service._excel_live_service_engine

    os.environ["EXCEL_LIVE_ENGINE"] = "file"
    file_service.WORKSPACE_ROOT = root
    live_service._excel_live_service = None
    live_service._excel_live_service_engine = None
    try:
        yield root
    finally:
        file_service.WORKSPACE_ROOT = prev_root
        live_service._excel_live_service = prev_singleton
        live_service._excel_live_service_engine = prev_engine_name
        if prev_engine is None:
            os.environ.pop("EXCEL_LIVE_ENGINE", None)
        else:
            os.environ["EXCEL_LIVE_ENGINE"] = prev_engine
        shutil.rmtree(root, ignore_errors=True)


# ── 결과 단언 ────────────────────────────────────────────────────────────


@dataclass(frozen=True)
class CheckResult:
    passed: bool
    detail: str


class Expectation:
    """저장된 통합문서를 열어 최종 상태를 확인한다."""

    def check(self, path: Path, sheet: str) -> CheckResult:  # pragma: no cover - 인터페이스
        raise NotImplementedError


def _cells(path: Path, sheet: str) -> Any:
    workbook = load_workbook(path, data_only=False)
    return workbook[sheet]


def _column_values(worksheet: Any, header: str, *, skip_header: bool = True) -> list[Any]:
    headers = [str(cell.value or "").strip() for cell in worksheet[1]]
    if header not in headers:
        raise KeyError(f"머리글 '{header}'를 찾을 수 없습니다. 실제: {headers}")
    index = headers.index(header)
    rows = list(worksheet.iter_rows(min_row=2 if skip_header else 1, values_only=True))
    return [row[index] for row in rows if any(value is not None for value in row)]


@dataclass(frozen=True)
class CellValue(Expectation):
    cell: str
    value: Any

    def check(self, path: Path, sheet: str) -> CheckResult:
        actual = _cells(path, sheet)[self.cell].value
        ok = actual == self.value
        return CheckResult(ok, f"{self.cell}={actual!r} (기대 {self.value!r})")


@dataclass(frozen=True)
class RangeEmpty(Expectation):
    target_range: str

    def check(self, path: Path, sheet: str) -> CheckResult:
        worksheet = _cells(path, sheet)
        remaining = [
            cell.value
            for row in worksheet[self.target_range]
            for cell in row
            if cell.value not in (None, "")
        ]
        return CheckResult(
            not remaining, f"남은 값 {len(remaining)}개: {remaining[:5]}"
        )


@dataclass(frozen=True)
class SortedByColumn(Expectation):
    header: str
    descending: bool = False

    def check(self, path: Path, sheet: str) -> CheckResult:
        values = [v for v in _column_values(_cells(path, sheet), self.header) if v is not None]
        expected = sorted(values, reverse=self.descending)
        ok = values == expected
        return CheckResult(ok, f"실제 {values[:6]} / 기대 {expected[:6]}")


@dataclass(frozen=True)
class ColumnValuesEqual(Expectation):
    header: str
    values: list[Any]

    def check(self, path: Path, sheet: str) -> CheckResult:
        actual = _column_values(_cells(path, sheet), self.header)
        ok = actual == self.values
        return CheckResult(ok, f"실제 {actual} / 기대 {self.values}")


@dataclass(frozen=True)
class AllRowsSatisfy(Expectation):
    """필터 결과 검증 — 남은 행이 전부 조건을 만족하고, 개수도 맞아야 한다."""

    header: str
    predicate: Callable[[Any], bool]
    expected_count: int

    def check(self, path: Path, sheet: str) -> CheckResult:
        values = _column_values(_cells(path, sheet), self.header)
        violations = [v for v in values if not self.predicate(v)]
        ok = not violations and len(values) == self.expected_count
        return CheckResult(
            ok,
            f"{len(values)}행 남음(기대 {self.expected_count}), 조건 위반 {violations[:5]}",
        )


@dataclass(frozen=True)
class FormulaContains(Expectation):
    cell: str
    fragment: str

    def check(self, path: Path, sheet: str) -> CheckResult:
        actual = str(_cells(path, sheet)[self.cell].value or "")
        ok = actual.startswith("=") and self.fragment.upper() in actual.upper()
        return CheckResult(ok, f"{self.cell}={actual!r} ('{self.fragment}' 포함 기대)")


@dataclass(frozen=True)
class CellFilled(Expectation):
    """배경색이 실제로 칠해졌는지. 색상값은 엔진마다 표기가 달라 존재만 본다."""

    cell: str
    rgb_suffix: str = ""

    def check(self, path: Path, sheet: str) -> CheckResult:
        fill = _cells(path, sheet)[self.cell].fill
        rgb = str(getattr(getattr(fill, "start_color", None), "rgb", "") or "")
        filled = str(getattr(fill, "fill_type", "") or "") not in ("", "none")
        ok = filled and (not self.rgb_suffix or rgb.upper().endswith(self.rgb_suffix.upper()))
        return CheckResult(ok, f"{self.cell} fill={fill.fill_type} rgb={rgb}")


@dataclass(frozen=True)
class NumberFormatContains(Expectation):
    cell: str
    fragment: str

    def check(self, path: Path, sheet: str) -> CheckResult:
        actual = str(_cells(path, sheet)[self.cell].number_format or "")
        ok = self.fragment in actual
        return CheckResult(ok, f"{self.cell} 서식={actual!r}")


@dataclass(frozen=True)
class AllOf(Expectation):
    checks: list[Expectation]

    def check(self, path: Path, sheet: str) -> CheckResult:
        details = []
        passed = True
        for item in self.checks:
            result = item.check(path, sheet)
            passed = passed and result.passed
            details.append(("OK " if result.passed else "NG ") + result.detail)
        return CheckResult(passed, " | ".join(details))


# ── 케이스 정의 ──────────────────────────────────────────────────────────


@dataclass(frozen=True)
class BenchCase:
    case_id: str
    category: str
    prompt: str
    sheet: str
    rows: list[list[Any]]
    expectation: Expectation
    # 정답 액션 계획. LLM 없이 하네스 자체를 검증할 때 쓴다.
    oracle: list[dict[str, Any]] = field(default_factory=list)
    # 일부러 틀린 계획. 단언이 실패를 실제로 잡아내는지 확인한다.
    mutant: list[dict[str, Any]] = field(default_factory=list)


def build_workbook(root: Path, case: BenchCase) -> Path:
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = case.sheet
    for row in case.rows:
        worksheet.append(row)
    path = root / f"{case.case_id}.xlsx"
    workbook.save(path)
    return path


# ── 실행 ────────────────────────────────────────────────────────────────


@dataclass
class RunOutcome:
    case_id: str
    passed: bool
    detail: str
    error: str = ""
    steps: int = 0


Dispatcher = Callable[..., dict[str, Any]]


def router_dispatcher() -> Dispatcher:
    """`/command`가 실제로 쓰는 실행 경로.

    같은 시그니처의 디스패처가 `services/excel_actions.py`에도 있지만 그쪽은
    처리 액션이 16개뿐이라(라우터는 42개) 벤치마크 기준으로 삼을 수 없다.
    """
    from office_claw_sidecar.routers.excel_live import _execute_action

    return _execute_action


def run_plan(
    case: BenchCase,
    plan: list[dict[str, Any]],
    *,
    dispatcher: Dispatcher | None = None,
) -> RunOutcome:
    """계획을 격리 워크스페이스에서 실행하고 저장된 파일로 채점한다."""
    execute = dispatcher or router_dispatcher()

    with isolated_workspace() as root:
        path = build_workbook(root, case)
        try:
            for step in plan:
                execute(
                    action=str(step.get("action") or ""),
                    params=dict(step.get("params") or {}),
                    workbook_id=str(path),
                    sheet_name=case.sheet,
                )
            execute(
                action="excel_live.save_workbook",
                params={},
                workbook_id=str(path),
                sheet_name=case.sheet,
            )
        except Exception as exc:  # noqa: BLE001 - 실행 실패도 벤치마크 결과다.
            return RunOutcome(case.case_id, False, "", error=str(exc), steps=len(plan))

        result = case.expectation.check(path, case.sheet)
        return RunOutcome(case.case_id, result.passed, result.detail, steps=len(plan))
