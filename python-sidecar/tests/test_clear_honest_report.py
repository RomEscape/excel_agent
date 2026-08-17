"""지울 값이 없었으면 "완료"라고만 하지 말고 그 사실을 말한다.

2026-08-17 GUI 실측(같은 날 두 번째): 서식(배경·테두리)만 있고 값이 없는 A1:D9를
"초기화"하랬더니 값 비우기만 실행됐고, 지울 값이 0개인데 "[1/1] 엑셀 작업이
완료되었습니다"가 나갔다. 화면은 그대로였다. 사용자: "도대체 뭐가 반영된거야".

어휘는 고쳤지만(초기화 → 서식까지 3단계), 이 부류 전체의 안전망이 따로 필요하다 —
**성공했는데 사용자 눈에 보이는 변화가 0이면, 응답이 그 사실을 말해야 한다.**
highlight_by_condition의 "조건에 맞는 셀이 없어…"와 같은 원칙이다.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from office_claw_sidecar.main import app
from office_claw_sidecar.routers import excel_live as router
from office_claw_sidecar.services import excel_live_file_service as file_service
from office_claw_sidecar.services import excel_live_service as live_service

HEADERS = {"Authorization": "Bearer dev-token"}
SHEET = "매출"
client = TestClient(app)


@pytest.fixture
def workbook(monkeypatch):
    """B2:C3은 서식만(회색 배경), D2:D3은 값이 있는 통합문서."""
    root = Path(tempfile.mkdtemp(prefix="oc-honest-"))
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for ref in ("B2", "B3", "C2", "C3"):
        ws[ref].fill = PatternFill("solid", fgColor="D9D9D9")
    ws["D2"] = 111
    ws["D3"] = 222
    path = root / "book.xlsx"
    wb.save(path)

    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
    monkeypatch.setattr(file_service, "WORKSPACE_ROOT", root)
    monkeypatch.setattr(live_service, "_excel_live_service", None)
    monkeypatch.setattr(live_service, "_excel_live_service_engine", None)

    async def _no_llm(_message, llm_service, context):
        raise AssertionError("규칙 경로여야 한다")

    monkeypatch.setattr(router, "parse_excel_live_command", _no_llm)
    yield path
    live_service._excel_live_service = None
    live_service._excel_live_service_engine = None


def _command(path: Path, message: str) -> dict:
    payload = {
        "message": message,
        "workbook_id": str(path),
        "session_id": "sess-honest",
        "approve": False,
    }
    first = client.post("/excel-live/command", json=payload, headers=HEADERS)
    assert first.status_code == 200
    body = first.json()
    if body.get("approval_required"):
        body = client.post(
            "/excel-live/command", json={**payload, "approve": True}, headers=HEADERS
        ).json()
    return body


class TestClearingAnEmptyRangeSaysSo:
    def test_the_response_admits_nothing_was_deleted(self, workbook):
        # B2:C3은 서식만 있다 — 값 비우기는 성공하지만 화면은 안 바뀐다.
        body = _command(workbook, "B2:C3 값만 지워줘")
        assert body["ok"] is True
        assert body["result"].get("emptied_values") == 0
        assert body["result"].get("no_matching_cells") is True
        assert "지울 값이 없는" in str(body.get("reason")), (
            f"무동작이 '완료'로 보고된다: {body.get('reason')}"
        )

    def test_a_real_deletion_stays_a_plain_success(self, workbook):
        body = _command(workbook, "D2:D3 값만 지워줘")
        assert body["ok"] is True
        assert body["result"].get("emptied_values") == 2
        assert body["result"].get("no_matching_cells") is None
        assert "지울 값이 없는" not in str(body.get("reason"))
        wb = load_workbook(workbook)
        assert wb[SHEET]["D2"].value is None and wb[SHEET]["D3"].value is None
        wb.close()


class TestResetPhrasingActuallyChangesTheFile:
    def test_the_formatting_is_gone_from_the_file(self, workbook):
        """오늘의 그 턴: 서식만 있는 범위 + "초기화" → 이제 서식이 지워져야 한다."""
        body = _command(workbook, "B2:C3 여기 부분 초기화시켜줄 수 있어?")
        assert body["ok"] is True
        wb = load_workbook(workbook)
        ws = wb[SHEET]
        for ref in ("B2", "B3", "C2", "C3"):
            fill = ws[ref].fill
            argb = str(getattr(fill.fgColor, "rgb", "") or "")
            assert fill.patternType != "solid" or argb in {"FFFFFFFF", "00FFFFFF"}, (
                f"{ref} 배경이 남아 있다: {argb}"
            )
        wb.close()
        # 3단계가 실행됐으므로 "지울 값이 없는" 안내는 붙지 않아야 한다 —
        # 서식이 지워져 화면이 실제로 바뀌었다.
        assert "지울 값이 없는" not in str(body.get("reason"))
