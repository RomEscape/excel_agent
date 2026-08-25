"""2026-08-17 전 카테고리 배터리(24케이스)가 찾은 구멍 8개의 회귀 테스트.

배터리는 실제 HTTP + 파일 검증으로 16/24 → 24/24가 됐다. 여기서는 각 구멍의
원인 지점을 단위로 고정한다. 파이프라인 전체는 배터리와
`test_rule_plan_survives_pipeline.py`가 지킨다.
"""

from __future__ import annotations

import pytest

from office_claw_sidecar.routers.excel_live import _looks_like_format_code
from office_claw_sidecar.services.excel_correction_context import (
    LastFormula,
    build_below_formula_plan,
)
from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based
from office_claw_sidecar.services.excel_macro_planner import looks_like_macro_request
from office_claw_sidecar.services.excel_param_binder import (
    formula_has_reversed_range,
    formula_refers_beyond_used_columns,
)

ENTRY = {"used_range": "A1:D9", "columns": [{"letter": "D", "header": "금액"}]}


class TestQuotativeParticleIsNotData:
    """ "A12에 합계 라고 입력해줘" → A12에 '합계 라고'가 들어갔다."""

    @pytest.mark.parametrize(
        ("message", "expected"),
        [
            ("A12에 합계 라고 입력해줘", "합계"),
            ("A12에 완료 이라고 적어줘", "완료"),
            ("A12에 합계 입력해줘", "합계"),
        ],
    )
    def test_the_particle_is_stripped(self, message, expected):
        step = parse_command_rule_based(message)
        assert step["params"]["values_2d"] == [[expected]]


class TestFormattingWordsAreNotCellValues:
    """ "금액에 천 단위 콤마 넣어줘"가 셀에 '금액에 천 단위 콤마'라는 텍스트로 들어갔다."""

    @pytest.mark.parametrize(
        "message",
        ["금액에 천 단위 콤마 넣어줘", "소수점 둘째 자리로 넣어줘", "테두리 넣어줘", "퍼센트로 넣어줘"],
    )
    def test_no_write_plan_is_made(self, message):
        step = parse_command_rule_based(message)
        assert step is None or step.get("action") != "excel_live.write_range", step

    def test_a_plain_value_still_writes(self):
        # 서식 어휘 차단이 정상 입력("777 입력해줘")까지 막으면 안 된다.
        step = parse_command_rule_based("777 입력해줘")
        assert step["action"] == "excel_live.write_range"
        assert step["params"]["values_2d"] == [[777]]


class TestSlashInsideValuesIsNotASeparator:
    """ "A44:E44에 토 07/05,278000,… 입력"의 "토 07/05"가 두 칸으로 갈라졌다.

    2026-08-18 ex2 재현 실측: 나열 구분자에 빗금이 들어 있어 날짜("07/05")를
    쪼갰고, 요일 7행 전체의 열이 한 칸씩 밀렸다. 쉼표가 있는 나열에서 빗금은
    값의 일부다.
    """

    def test_a_comma_list_keeps_slashes_inside_values(self):
        step = parse_command_rule_based("A44:E44에 토 07/05,278000,28,29,104% 입력")
        assert step["action"] == "excel_live.write_range"
        assert step["params"]["values_2d"] == [["토 07/05", 278000, 28, 29, "104%"]]

    def test_a_slash_only_list_still_splits(self):
        # 쉼표가 아예 없으면 빗금 나열("월/화/수")은 여전히 세 칸이다.
        step = parse_command_rule_based("A1:C1에 월/화/수 입력")
        assert step["params"]["values_2d"] == [["월", "화", "수"]]


class TestBatchRowInput:
    """한 턴에 여러 행 — 세미콜론·줄바꿈이 행 구분자다.

    2026-08-18: 검증된 문형이 "한 턴 = 한 행"뿐이라 표 하나에 수십 턴이
    들었다(85턴 재현 대화). 행 구분자를 결정적 규칙으로 받는다.
    """

    def test_semicolon_separated_rows(self):
        step = parse_command_rule_based("A2:C4에 가,1,2; 나,3,4; 다,5,6 입력")
        assert step["action"] == "excel_live.write_range"
        assert step["params"]["start_cell"] == "A2"
        assert step["params"]["values_2d"] == [["가", 1, 2], ["나", 3, 4], ["다", 5, 6]]

    def test_newline_separated_rows(self):
        step = parse_command_rule_based("A2:B3에 사과,100\n배,200 입력")
        assert step["params"]["values_2d"] == [["사과", 100], ["배", 200]]

    def test_a_single_row_is_unchanged(self):
        step = parse_command_rule_based("B2:D2에 이름,수량,금액 입력")
        assert step["params"]["values_2d"] == [["이름", "수량", "금액"]]

    def test_short_rows_are_padded_to_the_range_width(self):
        step = parse_command_rule_based("A2:C3에 가,1; 나 입력")
        assert step["params"]["values_2d"] == [["가", 1, ""], ["나", "", ""]]


class TestFormatCodeSniffing:
    """플래너가 format_code에 말 그대로 "소수점 둘째 자리"를 넣었다."""

    @pytest.mark.parametrize("code", ["#,##0", "0.00", "0%", "yyyy-mm-dd", "General", "@ "])
    def test_real_codes_pass(self, code):
        assert _looks_like_format_code(code) is True

    @pytest.mark.parametrize("code", ["소수점 둘째 자리", "천 단위 콤마", "", "예쁘게"])
    def test_prose_fails(self, code):
        assert _looks_like_format_code(code) is False


class TestDegenerateFormulaDetection:
    """=AVERAGE(A2:A1)이 A1:A8에 적용돼 날짜 열이 통째로 덮였다."""

    def test_reversed_rows_are_caught(self):
        assert formula_has_reversed_range("=AVERAGE(A2:A1)") is True
        assert formula_has_reversed_range("=SUM(D2:D9)") is False

    def test_reversed_columns_are_caught(self):
        assert formula_has_reversed_range("=SUM(D1:A1)") is True

    def test_sheet_prefixed_ranges_are_ignored(self):
        assert formula_has_reversed_range("=SUM(매출!D9:D2)") is False


class TestOutOfDataColumnDetection:
    """=AVERAGE(E:E) — 데이터는 A~D뿐인데 근거 없는 열을 집었다."""

    def test_a_column_beyond_the_data_is_flagged(self):
        assert formula_refers_beyond_used_columns("=AVERAGE(E:E)", ENTRY) is True
        assert formula_refers_beyond_used_columns("=SUM(F2:F9)", ENTRY) is True

    def test_columns_inside_the_data_are_fine(self):
        assert formula_refers_beyond_used_columns("=SUM(D:D)", ENTRY) is False
        assert formula_refers_beyond_used_columns("=SUM(D2:D9)", ENTRY) is False

    def test_unknown_used_range_is_never_flagged(self):
        assert formula_refers_beyond_used_columns("=AVERAGE(E:E)", None) is False
        assert formula_refers_beyond_used_columns("=AVERAGE(E:E)", {}) is False


class TestBelowCellFormulaContext:
    """ "F2에 합계" 다음의 "그 아래 칸에는 평균" — 플래너로 가면 날짜 열이 덮였다."""

    LAST = LastFormula(sheet_name="매출", cell="F2", formula="=SUM(D:D)", at_ts=0.0)

    def test_the_function_is_swapped_one_cell_below(self):
        plan = build_below_formula_plan("그 아래 칸에는 평균 넣어줘", self.LAST)
        assert plan[0]["params"]["range_ref"] == "F3"
        assert plan[0]["params"]["formula_a1"] == "=AVERAGE(D:D)"

    @pytest.mark.parametrize(
        ("message", "func"),
        [("아래 칸에 최댓값도", "MAX"), ("밑에 칸에 건수 세어줘", "COUNTA"), ("그 아래 셀에 합계", "SUM")],
    )
    def test_every_aggregate_word_works(self, message, func):
        plan = build_below_formula_plan(message, self.LAST)
        assert plan[0]["params"]["formula_a1"].startswith(f"={func}(")

    def test_no_below_mention_means_no_plan(self):
        assert build_below_formula_plan("평균 넣어줘", self.LAST) is None

    def test_no_remembered_formula_means_no_plan(self):
        assert build_below_formula_plan("그 아래 칸에 평균", None) is None

    def test_a_complex_formula_is_not_imitated(self):
        last = LastFormula(sheet_name="", cell="F2", formula="=SUM(A1:B2)+SUM(C1:C2)", at_ts=0.0)
        assert build_below_formula_plan("아래 칸에 평균", last) is None


class TestSingleGroupByIsNotAMacro:
    """ "지역별 금액 합계 집계표 만들어줘"가 16단계 매크로가 돼 산출물이 0이었다."""

    @pytest.mark.parametrize(
        "message", ["지역별 금액 합계 집계표 만들어줘", "월별 매출 요약표 만들어줘", "지역별 집계표 만들어줘"]
    )
    def test_it_stays_on_the_pivot_path(self, message):
        assert looks_like_macro_request(message) is False

    def test_a_dashboard_is_still_a_macro(self):
        assert looks_like_macro_request("매출 대시보드 만들어줘") is True


class TestFormatCodeAliases:
    """'comma'는 'mm' 때문에 코드처럼 보여 셀 서식이 말 그대로 comma가 됐다."""

    @pytest.mark.parametrize("alias", ["comma", "thousand", "percent", "currency"])
    def test_english_word_aliases_are_not_codes(self, alias):
        assert _looks_like_format_code(alias) is False


class TestNamedSheetCreation:
    """"요약이라는 이름으로 시트 추가좀" → '이름으로' 시트가 생겼다 (배터리 실측)."""

    def test_the_named_form_wins(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan("요약이라는 이름으로 시트 추가좀", None)
        assert plan and plan[0]["params"]["sheet_name"] == "요약", plan

    def test_the_trailing_name_form(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan("새로운 시트 하나 파줘 이름은 백업으로", None)
        assert plan and plan[0]["params"]["sheet_name"] == "백업", plan


class TestChartWordsBeatReadWords:
    """"클레임 비중 도넛으로 보여줘"의 '보여줘'가 단순 조회로 새서 차트가 안 생겼다."""

    def test_kind_word_plus_show_verb_becomes_a_chart(self):
        from office_claw_sidecar.routers.excel_live import (
            _chart_kind_from_message,
            _chart_step_from_message,
        )

        assert _chart_kind_from_message("클레임 비중 도넛으로 보여줘") == "doughnut"
        step = _chart_step_from_message("주문건수 추이 그래프 하나 그려줄래?")
        assert step and step["params"]["chart_type"] == "line"


class TestTypoNormalization:
    """사람의 오타("만들어조"·"함계"·"정열")를 표준형으로 — 2026-08-18 지시."""

    @pytest.mark.parametrize(
        ("raw", "fixed"),
        [
            ("지역성과 시트 만들어조", "지역성과 시트 만들어줘"),
            ("함계를 표 아래 한줄로 부탁해", "합계를 표 아래 한줄로 부탁해"),
            ("정열 좀 해줘", "정렬 좀 해줘"),
            ("테두르 둘러줘", "테두리 둘러줘"),
            ("막대 차투 그러줘", "막대 차트 그려줘"),
        ],
    )
    def test_common_typos_are_fixed(self, raw, fixed):
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        assert normalize_common_typos(raw) == fixed

    def test_clean_sentences_are_untouched(self):
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        msg = "A1:F6에 지역,주문건수 입력"
        assert normalize_common_typos(msg) == msg


class TestCommandSentencesAreNotCellValues:
    """"합계를 표 아래에 한 줄로 넣어줘"가 활성 셀 값으로 들어갔다 (GUI 실측).

    셀 지목 없이 네 낱말 넘는 문장은 값이 아니다 — 쓰지 말고 물러나서
    뒤 단계가 되묻게 한다. '라고' 인용은 예외다.
    """

    @pytest.mark.parametrize(
        "message",
        [
            "합계를 표 아래에 한 줄로 넣어줘",
            "이 값들 전부 다른 데로 옮겨 넣어줘",
        ],
    )
    def test_long_command_clauses_back_off(self, message):
        step = parse_command_rule_based(message)
        assert step is None or step.get("action") != "excel_live.write_range" or (
            step["params"].get("start_cell") != "__ACTIVE_CELL__"
        ), step

    def test_short_values_and_quotes_still_write(self):
        step = parse_command_rule_based("완료 입력해줘")
        assert step["params"]["values_2d"] == [["완료"]]
        step2 = parse_command_rule_based("최종 점검 완료 되었음 이라고 입력해줘")
        assert step2["params"]["values_2d"] == [["최종 점검 완료 되었음"]]


class TestVerbSuffixesInBatchWrite:
    """"넣어줘"의 줘가 \b와 결합 못 해 행 쓰기가 미스 → 단일 셀 규칙이
    "A1:F6에"의 F6을 셀로 오인 → 문장 전체가 한 값 → 쉼표 재배열로 표 전체가
    조용히 오염됐다(2026-08-18 지저분판 실측 — 이번 강건성 작업 최대의 수확).
    """

    BODY = "지역,주문건수; 수도권,10452; 충청권,3892"

    @pytest.mark.parametrize("verb", ["입력", "입력해줘", "입력해주라", "넣어줘", "써줘"])
    def test_every_suffix_keeps_all_rows(self, verb):
        step = parse_command_rule_based(f"A1:B3에 {self.BODY} {verb}")
        assert step["action"] == "excel_live.write_range"
        v = step["params"]["values_2d"]
        assert len(v) == 3 and v[1][0] == "수도권", v

    def test_a_range_member_is_not_a_single_cell(self):
        # F6이 A1:F6의 일부일 때 단일 셀 쓰기가 잡으면 안 된다.
        step = parse_command_rule_based("A1:F6에 가,나,다,라,마,바 넣어줘")
        assert step["params"]["start_cell"] == "A1"
        assert len(step["params"]["values_2d"][0]) == 6


class TestGapHuntBatchOne:
    """5렌즈 사냥(2026-08-18, 70건) 중 최고 위험 6종의 회귀 핀."""

    def test_negated_save_does_not_save(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        assert _build_quick_action_plan("아직 저장하지 마", None) is None
        plan = _build_quick_action_plan("저장해줘", None)
        assert plan and plan[0]["action"] == "excel_live.save_workbook"

    def test_excluded_colors_are_skipped(self):
        from office_claw_sidecar.routers.excel_live import _quick_extract_colors

        assert _quick_extract_colors("빨간색 말고 노란색으로 칠해줘") == ["#FFFF00"]

    def test_conjugated_colors_resolve(self):
        from office_claw_sidecar.routers.excel_live import _quick_extract_colors

        assert _quick_extract_colors("F열 빨갛게 칠해줘") == ["#FF4D4F"]
        assert _quick_extract_colors("까맣게 칠해줘") == ["#000000"]

    def test_locative_da_ga_is_not_a_value(self):
        assert parse_command_rule_based("B3에다 500 넣어줘")["params"]["values_2d"] == [[500]]
        v = parse_command_rule_based("A2:C2에다가 이름,나이,점수 입력해줘")["params"]["values_2d"]
        assert v == [["이름", "나이", "점수"]]

    def test_vertical_range_takes_a_horizontal_list_downward(self):
        v = parse_command_rule_based("B2:B4에 12,000, 8,500, 9,300 입력해줘")["params"]["values_2d"]
        assert v == [[12000], [8500], [9300]]

    def test_thousand_commas_do_not_split_but_lists_do(self):
        v = parse_command_rule_based("A44:E44에 토 07/05,278000,28,29,104% 입력")["params"]["values_2d"]
        assert v == [["토 07/05", 278000, 28, 29, "104%"]]

    def test_subset_qualifiers_block_the_blanket_clear(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        assert _build_quick_action_plan("표 서식만 지워줘, 값은 그대로 두고", "A1:D9") is None
        assert _build_quick_action_plan("A1:C10에서 중복된 행은 지워줘", None) is None
        assert _build_quick_action_plan("합계 행만 지워줘", "A1:F7") is None
        # 정상 좁은 삭제(값만)는 그대로 동작해야 한다.
        plan = _build_quick_action_plan("여기 값만 지워줘", "A1:D9")
        assert plan and plan[-1]["action"] == "excel_live.clear_range"


class TestGapHuntBatchTwo:
    """5렌즈 백로그 2차 소화(2026-08-18) — 조용한 파괴 잔여 14종의 회귀 핀."""

    def test_sheet_deletion_words_do_not_clear_contents(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        q = _build_quick_action_plan("임시 시트 지워줘", None)
        assert q is None or q[0]["action"] != "excel_live.clear_range", q

    def test_font_size_without_the_word_size(self):
        from office_claw_sidecar.services.excel_live_agent import extract_font_params

        assert extract_font_params("여기 폰트 14로 바꿔줘").get("size") == 14.0
        assert extract_font_params("머리글 진하게 해줘").get("bold") is True

    @pytest.mark.parametrize(
        ("message", "value"),
        [("미납인 건 빨간색으로 칠해줘", "미납"), ("상태가 품절인 것만 빨간색으로", "품절")],
    )
    def test_colloquial_equality_forms(self, message, value):
        from office_claw_sidecar.services.excel_live_agent import parse_text_equals_condition

        assert parse_text_equals_condition(message) == value

    def test_fill_words_without_color_are_writes(self):
        r = parse_command_rule_based("C3에 100 채워줘")
        assert r["action"] == "excel_live.write_range"
        assert r["params"]["values_2d"] == [[100]]

    def test_trailing_adverbs_and_compound_clauses(self):
        r = parse_command_rule_based("A2:C2에 서울, 부산, 대구 순서대로 입력해줘")
        assert r["params"]["values_2d"] == [["서울", "부산", "대구"]]
        r2 = parse_command_rule_based("A2:C2에 서울,부산,대구 넣고 D2에 합계 써줘")
        assert r2 is None or "넣고" not in str(r2.get("params", {}).get("values_2d", ""))

    def test_grid_reshape_when_tokens_fill_the_range(self):
        r = parse_command_rule_based("A1:B2에 상품,수량,사과,3,배,5 입력해줘")
        assert r["params"]["values_2d"] == [["상품", "수량"], ["사과", 3], ["배", 5]]

    def test_compound_korean_numerals_and_minus(self):
        from office_claw_sidecar.services.korean_number import parse_condition

        assert parse_condition("매출이 3만 5천 원 이상이면") == (">=", 35000.0, False)
        assert parse_condition("수익률이 마이너스인 종목") == ("<", 0.0, False)

    def test_vague_conditions_back_off_instead_of_painting(self):
        r = parse_command_rule_based("지난주보다 급증한 센서값 빨간색으로 강조해줘")
        assert r is None or r.get("action") != "excel_live.fill_range", r

    def test_tell_me_is_a_read(self):
        r = parse_command_rule_based("B2 값 알려줘")
        assert r and r["action"] == "excel_live.read_range"


class TestThreeRoundHumanRunFindings:
    """사람 말투판 GUI 조건 ×3 라운드(2026-08-18)가 파일 검증에서 잡은 값 오염 2종."""

    def test_a_decimal_is_never_merged_with_the_next_thousand_group(self):
        # "92.6,145,0"이 92.6145로 붙어 열이 밀렸다.
        v = parse_command_rule_based("A1:F1에 강원제주,2495,2383,92.6,145,0 입력")["params"]["values_2d"]
        assert v == [["강원제주", 2495, 2383, 92.6, 145, 0]], v
        # 진짜 천 단위는 여전히 붙는다.
        assert parse_command_rule_based("B2:B4에 12,000, 8,500, 9,300 입력해줘")["params"]["values_2d"] == [[12000], [8500], [9300]]

    def test_header_row_styling_targets_only_the_first_row_of_the_context(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        # 직전 쓰기 범위 A1:E6이 컨텍스트여도 "머리글 행"은 A1:E1이다 — 표 전체가
        # 빨갛게 칠해져 상태 배지 강조와 겹쳤다.
        plan = _build_quick_action_plan("머리글 행은 빨간색 배경에 흰 글씨로 굵게 해줘", "A1:E6")
        assert {s["params"]["target_range"] for s in plan} == {"A1:E1"}, plan


class TestChartSlotKindVocab:
    """차트 슬롯의 종류 파서가 도넛·선그래프(붙여쓰기)를 몰라 종류를 말해도
    되묻기가 반복됐다(2026-08-18 ex5 대화형 각본). 어휘는 한 곳만 본다."""

    @pytest.mark.parametrize(
        ("message", "kind"),
        [("GMV로 도넛 차트 그려줘", "doughnut"), ("정시배송률 추이 선그래프로", "line"),
         ("지연건수는 막대로 그러줘", "bar"), ("클레임 비중 원형으로 그려줘", "pie")],
    )
    def test_slot_kind_matches_shared_vocab(self, message, kind):
        from office_claw_sidecar.routers.excel_live import _extract_operation_hints

        hints = _extract_operation_hints(message)
        assert hints.get("params", {}).get("chart_type") == kind, hints


class TestRealUsagePhraseBattery:
    """사용자의 실사용 문장 116개 배터리(2026-08-18) 유일 실패의 회귀 핀."""

    def test_sheet_plus_deictic_prefix_is_not_a_value(self):
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        r = parse_rangeless_row_write(
            "지역성과 시트에 이 영역에 지역,주문건수,출고건수; 수도권,10452,10158 입력해줘", "A1:C2"
        )
        assert r["params"]["values_2d"][0] == ["지역", "주문건수", "출고건수"], r

    def test_a_semicolon_list_never_becomes_one_cell_value(self):
        # 문맥이 없어도 배치 나열은 한 칸 값이 아니다 — 물러나서 되묻는다.
        step = parse_command_rule_based("지역,주문건수; 수도권,10452; 충청권,3892 입력해줘")
        assert step is None or step["params"].get("start_cell") != "__ACTIVE_CELL__" or (
            ";" not in str(step["params"].get("values_2d"))
        ), step


class TestNewScenarioAuthorTraps:
    """2026-08-19 ex9~22 각본 작성자 7명이 독립적으로 보고한 파서 함정 — 붙여넣기 앞말·No 머리글·값 안 영어."""

    @pytest.mark.parametrize(
        "text",
        [
            "이 표 옆에 이어서 No., 과제, 점수; 1, 보고서, 90 입력해줘",
            "이 옆에 No., 과제, 점수; 1, 보고서, 90 넣어줘",
            "오른쪽에 이어서 No., 과제, 점수; 1, 보고서, 90 써줘",
            "바로 밑에 No., 과제, 점수; 1, 보고서, 90 입력해줘",
            "그 다음 줄에 No., 과제, 점수; 1, 보고서, 90 입력해줘",
            "이어서 No., 과제, 점수; 1, 보고서, 90 입력해줘",
            "아 그리고 이 표 밑에 No., 과제, 점수; 1, 보고서, 90 입력해줘",
        ],
    )
    def test_locative_lead_words_are_not_values(self, text):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_rangeless_row_write,
        )

        r = parse_rangeless_row_write(normalize_common_typos(text), "A1:C2")
        assert r is not None, text
        assert r["params"]["values_2d"][0] == ["No.", "과제", "점수"], (text, r)

    def test_value_words_that_start_like_lead_words_survive(self):
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        r = parse_rangeless_row_write("이어폰, 케이블, 충전기; 3, 4, 5 입력", "A1:C2")
        assert r["params"]["values_2d"][0] == ["이어폰", "케이블", "충전기"]
        r = parse_rangeless_row_write("여기에 다음, 이전; 1, 2 입력해줘", "A1:B2")
        assert r["params"]["values_2d"][0] == ["다음", "이전"]

    def test_no_header_is_a_string_not_false(self):
        from office_claw_sidecar.services.excel_live_agent import (
            _parse_literal_value,
            parse_rangeless_row_write,
        )

        assert _parse_literal_value("No") == "No"
        assert _parse_literal_value("Yes") == "Yes"
        assert _parse_literal_value("TRUE") is True and _parse_literal_value("false") is False
        r = parse_rangeless_row_write("여기에 No, 이름, 승인; 1, 김철수, Yes 입력해줘", "A1:C2")
        assert r["params"]["values_2d"] == [["No", "이름", "승인"], [1, "김철수", "Yes"]]

    def test_english_words_inside_value_grid_are_preserved(self):
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        src = "여기에 샘플, HEK293 Cell Lysate, Flow Cytometry (T cell); 1, 2, 3 입력해줘"
        assert normalize_common_typos(src) == src

    def test_english_values_in_single_write_are_preserved_but_commands_map(self):
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        assert normalize_common_typos("A1에 Total 넣어줘") == "A1에 Total 넣어줘"
        assert normalize_common_typos("B2에 red 라고 써줘") == "B2에 red 라고 써줘"
        assert normalize_common_typos("sum 줄 넣어줘") == "합계 줄 넣어줘"
        assert normalize_common_typos("header bold 해줘") == "머리글 굵게 해줘"
        assert normalize_common_typos("첫 줄 freeze 해줘") == "첫 줄 고정 해줘"
        # 영어로만 된 명령은 영어 규칙 몫 — 건드리지 않는다.
        assert normalize_common_typos("write header in E1:G1") == "write header in E1:G1"


class TestNewScenarioRound1:
    """2026-08-19 ex9~15 1라운드 실패 7건(실제 원인 4종)의 회귀 핀."""

    def test_cell_ref_followed_by_hangul_counts_as_explicit(self):
        # "A45에"의 A45는 셀 좌표다 — `\b`는 한글 앞에서 경계가 아니라 못 잡았다(ex11 t57).
        from office_claw_sidecar.services.excel_selection_context import mentions_explicit_range

        assert mentions_explicit_range("대시보드 시트 A45에 각주 입력") is True
        assert mentions_explicit_range("A2:F6에 값 넣어줘") is True

    def test_footnote_sentence_is_a_literal_value_not_an_echo(self):
        from office_claw_sidecar.services.excel_param_binder import write_values_echo_the_request

        note = "※ AI 적용 후 수치는 추정값으로 시스템 로그 및 표본 분석 결과를 기반으로 산출되었습니다"
        assert write_values_echo_the_request({"values_2d": [[note]]}, f"대시보드 시트 A45에 {note} 입력") is False
        # 원래 잡던 되뇜은 그대로 잡는다.
        assert write_values_echo_the_request({"values_2d": [["가장 큰 매출"]]}, "F7에 가장 큰 매출 값 넣어줘") is True

    @pytest.mark.parametrize(
        "text, expected",
        [
            ("재고 관리 시트도 하나 만들어줄래?", "재고 관리"),
            ("그리고 재고 관리 시트 만들어줘", "재고 관리"),
            ("이제 2024 실적 시트를 새로 만들어", "2024 실적"),
            ("매출 넣고 요약 시트 만들어줘", "요약"),
            ("데이터 시트는 두고 대시보드 시트 만들어줘", "대시보드"),
            ("월간 보고 시트 하나 파줘", "월간 보고"),
            ("ㅇㅇ 안전 점검 체크리스트 시트 추가", "안전 점검 체크리스트"),
        ],
    )
    def test_multiword_sheet_names_in_create(self, text, expected):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["action"] == "excel_live.create_sheet", (text, plan)
        assert plan[0]["params"]["sheet_name"] == expected, (text, plan)

    def test_multiword_sheet_mention_resolves_against_known_names(self):
        from office_claw_sidecar.services.excel_param_binder import sheet_mention_matches_known

        assert sheet_mention_matches_known("재고 관리 시트 A1에 값", "관리", ["Sheet", "재고 관리"]) is True
        assert sheet_mention_matches_known("간트 관리 시트 입력", "관리", ["Sheet", "간트관리"]) is True
        assert sheet_mention_matches_known("판매 관리 시트에", "관리", ["Sheet", "재고관리"]) is False

    def test_multiword_sheet_prefix_is_not_a_paste_value(self):
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        r = parse_rangeless_row_write("재고 관리 시트에 여기에 품목 코드, 품목명; SUP-1, 장갑 입력해줘", "A1:B2")
        assert r["params"]["values_2d"][0] == ["품목 코드", "품목명"]

    def test_sort_key_without_unit_suffix_or_by_tail_word(self):
        from office_claw_sidecar.services.excel_param_binder import _pick_sort_key

        headers = ["단과대학", "재적생 수(명)", "신입생 수(명)", "평균 평점(GPA)", "장학금 지급액(억원)"]
        assert _pick_sort_key("재적생 수 많은 순으로", headers) == "재적생 수(명)"
        assert _pick_sort_key("평점 높은 순", headers) == "평균 평점(GPA)"
        # 꼬리 낱말이 겹치면 쓰지 않는다.
        assert _pick_sort_key("건수 많은 순", ["지역", "주문 건수", "출고 건수"]) is None

    def test_cell_word_inside_values_does_not_bail(self):
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        r = parse_rangeless_row_write("여기에 ID, 시료; S1, HEK293 셀 용해물 입력해줘", "A1:B2")
        assert r["params"]["values_2d"][1] == ["S1", "HEK293 셀 용해물"]
        assert parse_rangeless_row_write("A1 셀 철수, 영희 입력해줘", "A1:B1") is None


class TestNewScenarioRound2And3:
    """2026-08-19 ex9~22 v2(말 바꾼 변형)·ex16~22 1라운드 실패의 회귀 핀."""

    @pytest.mark.parametrize(
        "text",
        ["이거 넣어줘", "음 이거 넣어줘", "복사한 거 여기에 붙여넣어줘", "방금 거 입력", "이거 여기다 써줘"],
    )
    def test_pronoun_only_write_is_not_a_value(self, text):
        from office_claw_sidecar.routers.excel_live import _BARE_WRITE_REQUEST
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        t = normalize_common_typos(text)
        assert _BARE_WRITE_REQUEST.match(t) is not None, text
        assert parse_command_rule_based(t, context_range="A1:E13") is None, text

    def test_pronoun_with_cell_is_valueless(self):
        from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based

        assert parse_command_rule_based("A1에 이거 넣어") is None
        assert parse_command_rule_based("이거 A1에 넣어줘") is None
        step = parse_command_rule_based("A1에 123 넣어줘")
        assert step and step["params"]["values_2d"] == [[123]]

    @pytest.mark.parametrize(
        "text, expected",
        [
            ("대시보드라는 시트 하나 더 만들어주세요", "대시보드"),
            ("아 그리고 지점데이터 시트도 하나 더 만들어주세요", "지점데이터"),
            ("리소스관리 시트 하나 더 추가해주세요", "리소스관리"),
            ("시트 새로 하나 파 주세요, 이름은 Requisition으로", "Requisition"),
            ("대시보드란 이름의 시트 추가", "대시보드"),
            ("요약이라고 시트 하나 파줘", "요약"),
        ],
    )
    def test_sheet_creation_phrasings(self, text, expected):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos
        from office_claw_sidecar.services.excel_macro_planner import looks_like_macro_request

        t = normalize_common_typos(text)
        assert looks_like_macro_request(t) is False, text
        plan = _build_quick_action_plan(t, None)
        assert plan and plan[0]["action"] == "excel_live.create_sheet", (text, plan)
        assert plan[0]["params"]["sheet_name"] == expected, (text, plan)

    def test_range_then_values_then_deictic(self):
        from office_claw_sidecar.services.excel_live_agent import parse_explicit_row_write

        r = parse_explicit_row_write("A12:F17 월,객실 매출(원),부대매출(원); 2026-01,1328950000,342715000 여기 입력해줘")
        assert r["params"]["start_cell"] == "A12"
        assert r["params"]["values_2d"][1][:3] == ["2026-01", 1328950000, 342715000]
        r = parse_explicit_row_write("A1:B2 철수, 영희; 1, 2 여기에 이거 써줘")
        assert r["params"]["values_2d"] == [["철수", "영희"], [1, 2]]

    def test_a4_inside_a_range_is_not_print(self):
        from office_claw_sidecar.routers.excel_live import _extract_operation_hints

        assert _extract_operation_hints("A46:F51 표 아래에 합계를 한 줄로 넣어줘").get("intent") != "print"
        assert _extract_operation_hints("A4 가로로 인쇄해줘").get("intent") == "print"

    @pytest.mark.parametrize(
        "text, find, repl",
        [
            ("처리중을 진행중으로 바꿔줘", "처리중", "진행중"),
            ('"처리중"을 "진행중"으로 바꿔줘', "처리중", "진행중"),
            ("ML Ops를 MLOps로 찾아 바꿔줘", "ML Ops", "MLOps"),
            ("처리중 → 진행중으로 바꿔", "처리중", "진행중"),
            ("상태 열의 대기를 보류로 바꿔", "대기", "보류"),
            ("표에서 N/A는 빈칸으로 바꿔줘", "N/A", ""),
        ],
    )
    def test_find_replace_rule(self, text, find, repl):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan(text, "D60:D64")
        assert plan and plan[0]["action"] == "excel_live.find_replace", (text, plan)
        assert plan[0]["params"]["find_text"] == find and plan[0]["params"]["replace_text"] == repl

    def test_sort_without_column_is_unresolved_even_without_headers(self):
        from office_claw_sidecar.services.excel_live_executor import PlanStep
        from office_claw_sidecar.services.excel_param_binder import bind_plan_steps

        steps = [PlanStep(action="excel_live.sort_rows", params={"column": "이름", "order": "asc"}, reason="")]
        _bound, notes = bind_plan_steps(steps, digest={"sheets": []}, message="정렬 좀 해주세요", sheet_name=None)
        assert any(n.get("status") == "unresolved" and n.get("slot") == "column" for n in notes), notes
        _bound, notes = bind_plan_steps(steps, digest={"sheets": []}, message="지원자 기준 내림차순이요", sheet_name=None)
        assert not any(n.get("status") == "unresolved" for n in notes), notes


class TestNewScenarioRound4:
    """2026-08-19 4라운드(28본) 잔여 실패의 회귀 핀 — 조용한 오실행 2종 포함."""

    def test_paste_grid_is_not_retargeted_by_a_value_that_names_a_sheet(self):
        from office_claw_sidecar.services.excel_param_binder import resolve_sheet_from_message

        digest = {"sheets": [{"name": "Sheet"}, {"name": "재고 관리"}, {"name": "대시보드"}]}
        grid = "적용 영역, 자동화 과제; 예약 관리, AI; 재고 관리, 수요 예측 입력해줘"
        assert resolve_sheet_from_message(grid, digest, default="대시보드") == "대시보드"
        assert resolve_sheet_from_message("재고 관리 시트에 여기에 품목, 수량; A, 1 입력해줘", digest, default="대시보드") == "재고 관리"
        # 한 칸 쓰기도 셀 좌표 앞부분만 본다.
        assert resolve_sheet_from_message("A1에 재고 관리 현황 써줘", digest, default="대시보드") == "대시보드"
        assert resolve_sheet_from_message("재고 관리 시트 A1에 현황 써줘", digest, default="대시보드") == "재고 관리"

    def test_commandish_fragments_are_never_paste_values(self):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_rangeless_row_write,
        )

        for text in ["넣어줘 합계 줄, 이 표 아래에", "합계 줄 하나 넣어줘, 이 표 아래에", "입력해줘 평균 한 줄, 표 밑에"]:
            assert parse_rangeless_row_write(normalize_common_typos(text), "A1:G7") is None, text
        r = parse_rangeless_row_write("넣어줘 지역, 건수, 합계", "A1:C1")
        assert r["params"]["values_2d"][0] == ["지역", "건수", "합계"]

    def test_leading_filler_after_range_prefix(self):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_explicit_row_write,
        )

        m = normalize_common_typos("A5:J6 아 그리고 여기에 총 재적생 수(명),등록금 수입(억원); 23842, 428 입력해줘")
        assert m.startswith("A5:J6 여기에")
        assert parse_explicit_row_write(m) is not None

    def test_chart_kind_pie_spelled_as_won_graph(self):
        from office_claw_sidecar.routers.excel_live import _chart_kind_from_message

        assert _chart_kind_from_message("B40:B45로 원 그래프 그려줘") == "pie"

    @pytest.mark.parametrize(
        "text, find, repl",
        [("처리중이라고 된 거 전부 진행중으로 바꿔줘", "처리중", "진행중"), ("대기라고 적힌 셀은 보류로 바꿔", "대기", "보류")],
    )
    def test_find_replace_said_as_cells(self, text, find, repl):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan(text, "D60:D64")
        assert plan and plan[0]["action"] == "excel_live.find_replace"
        assert plan[0]["params"]["find_text"] == find and plan[0]["params"]["replace_text"] == repl

    def test_sheet_creation_after_a_clause(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos("ㅇㅇ 시작하자, 체크리스트 시트부터 하나 파줘"), None)
        assert plan and plan[0]["action"] == "excel_live.create_sheet" and plan[0]["params"]["sheet_name"] == "체크리스트"

    def test_cross_sheet_with_in_the_sheet_phrasing(self):
        from office_claw_sidecar.services.excel_aggregate_below import _CROSS_SHEET

        m = _CROSS_SHEET.search("B21에 간트관리 시트에 있는 예산 총합 가져와줄래")
        assert m and m.group(2) == "간트관리"


class TestWorkbookAuditFindings:
    """2026-08-19 **결과 워크북 감사**로만 드러난 조용한 오실행 — 러너는 전부 성공으로 셌다."""

    @pytest.mark.parametrize(
        "text, formula",
        [
            ("B35 빼기 B36 한 값을 E35에 넣어줘", "=B35-B36"),
            ("B35에서 B36 뺀 값을 E35에 넣어줘", "=B35-B36"),
            ("B9 나누기 C9 한 값을 F9에 넣어줘", "=B9/C9"),
            ("B7을 B6으로 나눈 값을 C7에 넣어줘", "=B7/B6"),
        ],
    )
    def test_value_first_arithmetic_becomes_a_formula_not_text(self, text, formula):
        # 셀 지목이 뒤에 오면 사칙연산 파서가 못 받아 **문장이 셀에 글자로** 박혔다.
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text))
        assert step and step["action"] == "excel_live.set_formula", (text, step)
        assert step["params"]["formula_a1"] == formula

    def test_bare_particle_never_becomes_a_cell_value(self):
        from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based

        assert parse_command_rule_based("제목을 A1에 넣어줘") is None
        step = parse_command_rule_based("매출 실적을 B2에 입력해줘")
        assert step["params"]["values_2d"] == [["매출 실적"]]

    @pytest.mark.parametrize(
        "text",
        [
            "B10에다 성적부 시트 결석 다 더한 값 가져와줘",
            "B10에 성적부 시트 결석 전부 더한 값 넣어줘",
            "B10에 성적부 시트의 결석 총합 가져와",
            "성적부 시트 결석 합계를 B10에 넣어줘",
        ],
    )
    def test_cross_sheet_aggregate_targets_the_named_cell_with_a_sheet_prefixed_formula(self, text):
        # 규칙이 못 잡으면 플래너가 **원본 시트에** 지역 SUM을 써서 학생 이름 칸을 덮었다.
        from office_claw_sidecar.services.excel_aggregate_below import build_cross_sheet_aggregate_plan

        def reader(sheet):
            return {
                "성적부": ("A1:F11", [["번호", "이름", "출석률(%)", "출석", "지각", "결석"], [1, "김", 90, 20, 1, 2]]),
                "대시보드": ("A1:A1", [["제목"]]),
            }.get(sheet)

        plan = build_cross_sheet_aggregate_plan(text, reader, sheet_names=["Sheet", "성적부", "대시보드"])
        assert plan, text
        assert plan[0]["params"]["range_ref"] == "B10"
        assert plan[0]["params"]["formula_a1"] == "=SUM('성적부'!F2:F11)"

    def test_a_source_sheet_named_after_the_destination_cell_does_not_retarget(self):
        from office_claw_sidecar.services.excel_param_binder import resolve_sheet_from_message

        digest = {"sheets": [{"name": "Sheet"}, {"name": "성적부"}, {"name": "대시보드"}]}
        assert resolve_sheet_from_message("B10에다 성적부 시트 결석 다 더한 값 가져와줘", digest, default="대시보드") == "대시보드"
        # 시트를 작업 대상으로 부른 문장은 그대로 옮긴다.
        assert resolve_sheet_from_message("성적부 시트 A1 굵게", digest, default="대시보드") == "성적부"

    def test_sheet_plus_aggregate_is_a_formula_request_not_a_literal_value(self):
        from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based

        step = parse_command_rule_based("성적부 시트 결석 합계를 B10에 넣어줘")
        assert step is None or step["action"] != "excel_live.write_range", step


class TestCrossSheetHeaderMatching:
    """2026-08-19 5라운드 워크북 감사: 괄호·공백이 든 머리글을 못 잡아 계획이 비었고,
    그 틈에 플래너가 **'=' 없는 문자열** 'SUM(렌트롤!D2:D100)'을 셀에 텍스트로 썼다."""

    def _reader(self):
        data = {
            "렌트롤": (
                "A1:Q10",
                [
                    ["자산 코드", "건물명", "층", "호수", "용도", "임차인", "전용면적", "임대면적", "시작일",
                     "만료일", "계약형태", "보증금", "기본임대료", "관리비", "총 임대료(원)", "평당임대료", "점유상태"],
                    ["A", "B", 1, 2, "c", "d", 1, 2, "e", "f", "g", 1, 2, 3, 4, 5, "h"],
                ],
            ),
            "간트관리": ("A1:F5", [["WBS", "작업", "유형", "예산(원)", "시작", "종료"], ["1", "a", "b", 100, "c", "d"]]),
            "성적부": ("A1:F11", [["번호", "이름", "출석률(%)", "출석", "지각", "결석"], [1, "김", 90, 20, 1, 2]]),
        }
        return lambda sheet: data.get(sheet)

    @pytest.mark.parametrize(
        "text, formula",
        [
            # 괄호·공백이 든 머리글
            ("B73에다 렌트롤 시트 총 임대료(원) 다 더한 값 가져와줘", "=SUM('렌트롤'!O2:O10)"),
            ("B73에 렌트롤 시트 총 임대료 합계 가져와", "=SUM('렌트롤'!O2:O10)"),
            # 단위 꼬리만 다른 경우
            ("B21에 간트관리 시트에 있는 예산 총합 가져와줄래", "=SUM('간트관리'!D2:D5)"),
            # 수량 부사가 머리글 자리를 뺏는 경우(기존 회귀)
            ("B10에다 성적부 시트 결석 다 더한 값 가져와줘", "=SUM('성적부'!F2:F11)"),
        ],
    )
    def test_headers_with_units_and_spaces_resolve(self, text, formula):
        from office_claw_sidecar.services.excel_aggregate_below import build_cross_sheet_aggregate_plan

        plan = build_cross_sheet_aggregate_plan(
            text, self._reader(), sheet_names=["Sheet", "렌트롤", "간트관리", "성적부", "대시보드"]
        )
        assert plan, text
        assert plan[0]["params"]["formula_a1"] == formula, (text, plan)

    def test_an_unknown_header_still_backs_off(self):
        # 없는 머리글을 추측해서 아무 열이나 잡으면 안 된다 — 되묻기로 가야 한다.
        from office_claw_sidecar.services.excel_aggregate_below import build_cross_sheet_aggregate_plan

        plan = build_cross_sheet_aggregate_plan(
            "B21에 간트관리 시트 없는열 합계 가져와", self._reader(), sheet_names=["Sheet", "간트관리"]
        )
        assert plan == []

    def test_an_ambiguous_partial_header_backs_off(self):
        # 부분 일치가 여럿이면 고르지 않는다.
        from office_claw_sidecar.services.excel_aggregate_below import _match_header

        assert _match_header("건수", ["지역", "주문 건수", "출고 건수"]) is None
        assert _match_header("주문 건수", ["지역", "주문 건수", "출고 건수"]) == 1


class TestNegationIsOneGate:
    """부정 판정을 규칙마다 따로 적으면 구멍이 반드시 생긴다.

    2026-08-19 블라인드 게이트: 저장 규칙이 자체 정규식("저장하지 마")만 들고 있어
    "저장 안 해도 돼요" · "저장 말고"가 그대로 실행돼 **파일이 저장됐다**.
    """

    @pytest.mark.parametrize(
        "text",
        [
            "아직 저장 안 해도 돼요, 기다려 주세요",
            "이거 아직 저장 말고",
            "아직 저장하지 마",
            "저장 금지 아직은",
            "하지 마 저장은 아직",
            "저장은 나중에",
            "저장 보류",
            "지금은 save 하지 마세요.",
            "아즉 저장 하지맠",
            "저장은 아직 하지 말아 주세요.",
        ],
    )
    def test_a_negated_save_never_saves(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        actions = [str(s.get("action")) for s in (plan or [])]
        assert "excel_live.save_workbook" not in actions, (text, plan)

    @pytest.mark.parametrize("text", ["저장해줘", "저장", "파일 저장 부탁해", "save 해줘", "다 됐으면 저장해줘"])
    def test_a_plain_save_still_saves(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["action"] == "excel_live.save_workbook", (text, plan)


class TestRenameSheetRule:
    """2026-08-19 블라인드 게이트에서 `rename_sheet`는 **규칙 0건 · 오류 5 · 오실행 3**으로 가장 나빴다.

    기존 규칙을 실측하니 셋이 동시에 깨져 있었다:
      ① 새 이름이 조사를 먹는다 — "지역별실적으로" → '지역별실적으'라는 시트가 생긴다
      ② 지시어를 시트 이름으로 잡는다 — "이 시트 이름 …" → sheet_name='이' → 없는 시트 → ERROR
      ③ "시트명 X로" · "탭 이름 X로" · "시트 이름 X!" 는 아예 매칭되지 않는다
    """

    @pytest.mark.parametrize(
        "text, expected_old",
        [
            ("지역성과 시트 이름을 지역별실적으로 바꿔 주세요", "지역성과"),
            ("이 시트 이름 지역별실적으로 바꿔", None),
            ("시트명 지역별실적으로 변경해줴 빨ㄹ리", None),
            ("ㅇㅇ 탭 이름 지역별실적으로", None),
            ("시트 이름 지역별실적!", None),
            ("지역성과 씨트 이름 지역별실적으로 바꺼", "지역성과"),
            ("아 그리고 지역성과 시트는 지역별실적이라고 이름 바꿔놔", "지역성과"),
            ("지역성과 sheet 이름 지역별실적으로 rename 부탁드려요.", "지역성과"),
            ("지역성과라고 된 탭 이름을 지역별실적으로 고쳐 주세요", "지역성과"),
            ("이 시트 이름을 지역성과에서 지역별실적으로 바꿔 주세요", "지역성과"),
            ("지역별실적으로 바꿔줘 지역성과 시트 이름", "지역성과"),
            ("현재 시트 탭 이름 지역별실적으로 바꿔줘", None),
            ("이 시트 이름 지역별실적으로 다시 지어줘", None),
        ],
    )
    def test_the_new_name_never_swallows_a_particle(self, text, expected_old):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["action"] == "excel_live.rename_sheet", (text, plan)
        params = plan[0]["params"]
        assert params["new_name"] == "지역별실적", (text, params)
        assert params.get("sheet_name") == expected_old, (text, params)

    @pytest.mark.parametrize(
        "text",
        ["A1에 시트 이름 써줘", "B2에 시트명 입력해줘", "요약 시트 만들어줘", "지역성과 시트 삭제해줘"],
    )
    def test_it_does_not_grab_writes_or_other_sheet_actions(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        actions = [str(s.get("action")) for s in (plan or [])]
        assert "excel_live.rename_sheet" not in actions, (text, plan)


class TestAutofitRule:
    """2026-08-19 게이트: `autofit` 24문장 전부 규칙 0건 → 모델(해석 카드)로 갔다.
    결정적으로 풀리는 동작인데 모델을 부르면 느리고 흔들린다."""

    @pytest.mark.parametrize(
        "text",
        [
            "열 너비 내용에 맞게 자동으로 맞춰",
            "열너비 맞춰",
            "자동 맞춤 해줘 열 너비",
            "여기 열 폭 글자 길이에 맞게 맞춰봐",
            "열 너비 내용에 맏게 자동 조졍",
            "컬럼 폭 자동으로 마춰줘 빨ㄹ리",
            "ㅇㅇ 열 폭 자동 맞춤",
            "열 너비 autofit 부탁드려요.",
            "글자가 잘려서 ###으로 보이는 칸이 있어서요, 열 너비를 내용에 맞게 자동으로 조정해 주세요.",
            "column width 내용에 맞게 auto fit 해 주세요.",
            "칸 폭 글자 안 잘리게 알아서 맞춰 주세요",
            "열 간격 내용에 딱 맞게 넓혀줘, 보기 좋게",
        ],
    )
    def test_autofit_phrasings_resolve_to_a_rule(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["action"] == "excel_live.autofit_columns", (text, plan)

    @pytest.mark.parametrize("text", ["열 너비 15로 해줘", "행 높이 키워줘", "B열 굵게", "B열 지워줘"])
    def test_it_does_not_grab_other_column_actions(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        actions = [str(s.get("action")) for s in (plan or [])]
        assert "excel_live.autofit_columns" not in actions, (text, plan)


class TestPlanWasWrongNotExecution:
    """2026-08-19~20: 사후조건을 붙이고도 게이트가 안 움직인 이유 — 이 셋은 **계획이 틀렸고 실행은 충실했다**.

    증상만 보고 "사후조건이 잡을 수 있다"고 분류했던 것이 오판이었다. 계획 params를 열어 보고 알았다:
      · `형식=000` → 계획이 실제로 `format_code='000'`을 요청했다("1,000"의 숫자를 코드로 오인)
      · 배경 없음 → `fill_range` 단계가 계획에 **아예 없었다**('배겅' 오타로 규칙이 미스)
    """

    @pytest.mark.parametrize(
        "text",
        [
            "주문건수와 출고건수 컬럼은 1,000 단위 comma format으로 해 주세요.",
            "주문건수 1000단위 쉼표",
            "여기 주문건수 출고건수는 1000단위로 쉼표 넣어줘",
            "B2:C8은 천 단위 콤마로",
        ],
    )
    def test_a_thousands_phrase_never_becomes_a_literal_zero_code(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["action"] == "excel_live.set_number_format", (text, plan)
        assert plan[0]["params"]["format_code"] == "#,##0", (text, plan)

    @pytest.mark.parametrize(
        "text, code",
        [("D2:D6 소수 둘째 자리", "0.00"), ("정시배송률 퍼센트로 보여줘", "0.0%"), ("소수 한 자리로 표시", "0.0")],
    )
    def test_other_format_requests_are_unaffected(self, text, code):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["params"]["format_code"] == code, (text, plan)

    @pytest.mark.parametrize(
        "text",
        [
            "머리글 행 남색 배겅에 흰글씨 굵게 해주세요",
            "표 첫 줄을 남색 배경에 흰색 굴게",
            "머리글 행 남색 배경에 흰 글씨 굵게 해줘",
        ],
    )
    def test_a_single_typo_must_not_drop_the_fill_step(self, text):
        # '배겅' 한 글자가 fill_range를 계획에서 통째로 날려 "배경만 빠진" 결과가 성공으로 보고됐다.
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        actions = [str(s.get("action")) for s in (plan or [])]
        assert "excel_live.fill_range" in actions and "excel_live.set_font" in actions, (text, plan)


class TestFindReplaceMustNotGuess:
    """2026-08-20: 어제 넣은 느슨한 찾아 바꾸기 규칙이 게이트의 조용한 오실행을 **5→6건으로 늘렸다.**

    잡은 값이 엉망이었다(find='여기 표', repl='수도권을 서울권' / repl='서울권으' / repl='건 다 서울권').
    더 나쁜 건 오탐이었다 — `"매출 높은 순으로 보여줘"`(정렬)가 find='매출 높' → repl='순' **치환**이 됐다.
    잘못된 치환은 데이터를 망친다. **넓게 잡는 것보다 확실할 때만 잡는 편이 낫다.**
    """

    @pytest.mark.parametrize(
        "text",
        [
            "매출 높은 순으로 보여줘",
            "클레임 많은 순으로 정렬해줘",
            "재적생 수 많은 순으로",
            "시트 이름을 요약으로 바꿔줘",
            "배경을 파란색으로 바꿔줘",
            "B열 숫자를 퍼센트로 바꿔줘",
            "차트를 막대로 바꿔줘",
            "글꼴을 굴림으로 바꿔줘",
        ],
    )
    def test_it_never_turns_another_intent_into_a_replacement(self, text):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), "D1:D9")
        actions = [str(s.get("action")) for s in (plan or [])]
        assert "excel_live.find_replace" not in actions, (text, plan)

    @pytest.mark.parametrize(
        "text, find, repl",
        [
            ("수도권이라고 된 거 전부 서울권으로 바꿔", "수도권", "서울권"),
            ("수도권→서울권", "수도권", "서울권"),
            ("수도권 → 서울권으로 변경 부탁드려요.", "수도권", "서울권"),
            ("표에서 수도권이라는 글자를 전부 서울권으로 바꿔 주실 수 있을까요?", "수도권", "서울권"),
            ("수도권을 서울권으로 replace 부탁드려요.", "수도권", "서울권"),
            ("서울권으로 바꿔 주세요, 수도권이라고 된 거", "수도권", "서울권"),
            ("ML Ops를 MLOps로 찾아 바꿔줘", "ML Ops", "MLOps"),
            ("상태 열에서 대기 중을 보류로 바꿔줘", "대기 중", "보류"),
            ("표에서 N/A는 빈칸으로 바꿔줘", "N/A", ""),
        ],
    )
    def test_a_real_replacement_keeps_both_words_whole(self, text, find, repl):
        # 조사를 먹으면('서울권으') 시트에 없는 글자를 찾게 되고, 수량어를 먹으면 엉뚱한 값이 들어간다.
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), "D1:D9")
        assert plan and plan[0]["action"] == "excel_live.find_replace", (text, plan)
        assert plan[0]["params"]["find_text"] == find, (text, plan[0]["params"])
        assert plan[0]["params"]["replace_text"] == repl, (text, plan[0]["params"])


class TestConditionMustNotVanish:
    """조건이 사라지면 **선택 전체가 칠해진다** — 조용한 오실행 중 가장 흔한 부류였다.

    2026-08-20 게이트: `highlight_status` 6건이 `fill_range(__ACTIVE_SELECTION__)`로 떨어져
    조건 없이 전부 칠했다. 조건 파서가 "상태=대기" · "대기라고 된 칸" · "대기 상태인 칸들"을 몰랐다.
    """

    @pytest.mark.parametrize(
        "text",
        [
            "ㅇㅇ 상태=대기 셀만 분홍 강조",
            "상태 열에서 대기라고 된 칸만 pink로 강조 부탁드려요.",
            "대기 상태인 칸들만 분홍으로 칠해서 한눈에 보이게 해줘",
            "상태 대기인거만 핑크로 강죠해 빨ㄹ리",
            "상태가 대기인 셀만 분홍색으로 표시해 주세요",
            "상태가 '대기'인 셀만 분홍색으로 칠해줘",
        ],
    )
    def test_the_equality_condition_is_parsed(self, text):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_text_equals_condition,
        )

        assert parse_text_equals_condition(normalize_common_typos(text)) == "대기", text

    @pytest.mark.parametrize(
        "text",
        [
            "ㅇㅇ 상태=대기 셀만 분홍 강조",
            "상태 열에서 대기라고 된 칸만 pink로 강조 부탁드려요.",
            "대기 상태인 칸들만 분홍으로 칠해서 한눈에 보이게 해줘",
        ],
    )
    def test_a_conditional_request_never_paints_everything(self, text):
        # fill_range(__ACTIVE_SELECTION__)로 떨어지면 조건이 사라져 전부 칠해진다.
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan, text
        assert plan[0]["action"] == "excel_live.highlight_by_condition", (text, plan)

    def test_a_plain_fill_is_still_a_plain_fill(self):
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan("A1:C3 노란색으로 칠해줘", None)
        assert plan and plan[0]["action"] == "excel_live.fill_range", plan


class TestArithmeticNeverBecomesText:
    """게이트 `cell_subtract`: 사칙연산 문장이 셀에 **글자로** 박혔다(계산!D2='B2에서 C2 뺀 값, 이걸')."""

    @pytest.mark.parametrize(
        "text, formula",
        [
            ("B2에서 C2 뺀 증감 D2에 넣어롸 빨ㄹ리", "=B2-C2"),          # 결과 명사가 중간
            ("D2에 B2 minus C2 값 넣어 주세요.", "=B2-C2"),            # 영어 연산어
            ("이번주 B2에서 지난주 C2 빼서 D2에 넣어줘", "=B2-C2"),        # 셀 사이 수식어
            ("B2에서 C2 뺀 값, 이걸 D2에 넣어 주세요", "=B2-C2"),         # 쉼표 + 지시대명사
            ("B13과 B11의 차이를 D15에 기록해줘", "=B13-B11"),           # '차이'가 식의 일부
            ("E15에 A15에서 C15 뺀 값 넣어줘", "=A15-C15"),
            ("B35 빼기 B36 한 값을 E35에 넣어줘", "=B35-B36"),
            ("B7을 B6으로 나눈 값을 C7에 넣어줘", "=B7/B6"),
        ],
    )
    def test_arithmetic_phrasings_become_formulas(self, text, formula):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text))
        assert step and step["action"] == "excel_live.set_formula", (text, step)
        assert step["params"]["formula_a1"] == formula, (text, step)

    @pytest.mark.parametrize(
        "text, value",
        [("A1에 완료 라고 써줘", "완료"), ("매출 실적을 B2에 입력해줘", "매출 실적"), ("A1에 수도권에서 온 주문 넣어줘", "수도권에서 온 주문")],
    )
    def test_plain_writes_are_still_writes(self, text, value):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text))
        assert step and step["action"] == "excel_live.write_range", (text, step)
        assert step["params"]["values_2d"] == [[value]], (text, step)


class TestPasteValuePhrasings:
    """게이트 `paste_values` 12문장 중 8이 되묻기·4가 오실행이었다.

    사람은 "입력해바" · "적어 주세요" · "paste 해 주세요" · "input 부탁드려요"라고도 쓰고,
    자리를 "붙여넣은 셀들에" · "방금 잡은 칸에" · "방금 선택한 두 줄에"라고 부르며,
    앞에 상황 설명을 붙인다("… 추가해야 해서요, 지금 붙여넣은 자리에 …").
    좁게 잡으면 값이 한 칸에 박히거나("셀들에 서울") 통째로 되묻기로 샌다.
    """

    EXPECTED = [["서울", 100], ["부산", 200]]

    @pytest.mark.parametrize(
        "text",
        [
            "여기다 서울,100; 부산,200 입력해바 그대루",
            "방금 선택한 칸에 서울,100; 부산,200 이렇게 입력해 주실 수 있을까요?",
            "지역별 데이터 두 줄을 추가해야 해서요, 지금 붙여넣은 자리에 서울,100; 부산,200 순서대로 넣어 주세요",
            "선택 범위에 서울,100; 부산,200 값 paste 해 주세요.",
            "이 자리에 서울,100; 부산,200 input 부탁드려요.",
            "선택한 범위에 서울,100; 부산,200 값 입력해 주세요",
            "붙여넣은 셀들에 서울,100; 부산,200 순서대로 채워줘",
            "방금 잡은 칸에 서울,100; 부산,200 이렇게 넣어줘",
            "여기다가 서울,100; 부산,200 적어 주세요",
            "서울,100; 부산,200 입력해주새요",
            "여기에 서울,100; 부산,200 넣어도 될까요?",
            "팀장님이 서울이랑 부산 데이터를 추가하라고 하셔서요, 방금 선택한 두 줄에 서울,100; 부산,200 넣어 주세요",
        ],
    )
    def test_the_grid_lands_intact(self, text):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_rangeless_row_write,
        )

        step = parse_rangeless_row_write(normalize_common_typos(text), "A8:B9")
        assert step is not None, text
        assert step["params"]["values_2d"] == self.EXPECTED, (text, step["params"])

    @pytest.mark.parametrize("text", ["A1에 완료 넣어줘", "합계 줄 넣어줘", "여기에 넣어줘", "넣어줘 합계 줄, 이 표 아래에"])
    def test_non_grid_sentences_are_not_grabbed(self, text):
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_rangeless_row_write,
        )

        assert parse_rangeless_row_write(normalize_common_typos(text), "A8:B9") is None, text


class TestCreateSheetName:
    """2026-08-20 블라인드 게이트 `new_sheet` 24문장. 규칙이 이름을 통째로 집어야 한다.

    이전에는 24문장 중 8이 되묻기·3이 오실행이었고, 그중 하나는
    `'요약' 이름으로 시트 하나` 에서 **'요약 이름으로'라는 시트를 만들었다.**
    """

    @pytest.mark.parametrize(
        "text",
        [
            "요약 시트 하나 만들어줘",
            "요약 탭 추가해줘",
            "요약이라는 이름의 새 시트를 만들어 주세요",
            "'요약' 이름으로 시트 하나 파 주세요",
            "새 시트 하나 추가해서 요약이라고 불러줘",
            "시트 새로 하나 만들고 이름은 요약으로",
            "여기다 요약 탭 하나 추가해봐",
            "ㅇㅇ 요약 시트부터 하나 파줘",
            "요약 워크시트 생성",
            "요약 시트",
        ],
    )
    def test_extracts_the_whole_name(self, text: str) -> None:
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan, text
        assert plan[0]["action"] == "excel_live.create_sheet", text
        assert plan[0]["params"]["sheet_name"] == "요약", text

    @pytest.mark.parametrize(
        ("text", "expected"),
        [
            ("재고 관리 시트도 하나 만들어줄래?", "재고 관리"),
            ("월간 보고 시트 하나 파줘", "월간 보고"),
            ("시트 새로 하나 파 주세요, 이름은 Requisition으로", "Requisition"),
        ],
    )
    def test_keeps_multi_word_names(self, text: str, expected: str) -> None:
        """`extend_sheet_name_leftward`가 앞 낱말까지 붙여야 하는 경우."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None)
        assert plan and plan[0]["params"]["sheet_name"] == expected, text

    @pytest.mark.parametrize(
        "text",
        [
            "A1에 시트 이름 써줘",
            "B2에 시트명 입력해줘",
            "시트 이름 요약으로 바꿔줘",
            "요약 시트 지워줘",
        ],
    )
    def test_does_not_fire_on_other_intents(self, text: str) -> None:
        """쓰기·이름변경·삭제를 생성으로 오인하면 엉뚱한 시트가 늘어난다."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None) or [{}]
        assert plan[0].get("action") != "excel_live.create_sheet", text

    def test_sheet_creation_plus_more_work_still_reaches_the_planner(self) -> None:
        """복합문은 규칙이 시트만 만들고 끝내면 안 된다 — underfit 판정으로 LLM에 넘긴다."""
        from office_claw_sidecar.routers.excel_live import (
            _build_quick_action_plan,
            _quick_plan_underfits_message,
        )

        message = "Summary 시트 만들어서 A1에 총매출이라고 쓰고 B1에 매출 합계 수식 넣어줘"
        plan = _build_quick_action_plan(message, None)
        assert plan and plan[0]["params"]["sheet_name"] == "Summary"
        assert _quick_plan_underfits_message(plan[0]["action"], message) is True


class TestTitleCellWrite:
    """2026-08-20 블라인드 게이트 `title_cell`. 한 칸에 제목 한 줄 — 어순·조사가 제각각이다.

    이전 실패:
      - `H1 물류 관제 대시보드`(조사·동사 없음) → 규칙이 못 잡아 모델로 갔다
      - `H1 셀 값에 … 넣어줘` → 값이 **'에 물류 관제 대시보드'** (조사가 값에 붙음)
      - `물류 관제 대시보드, 이걸 H1에 써주세요` → 값이 **'물류 관제 대시보드, 이걸'**
      - `H1 칸에 … 적어주세요` → 한글엔 낱말 경계가 없어 `적어(?:줘)?\\b`가 실패
    """

    @pytest.mark.parametrize(
        "text",
        [
            "H1에 물류 관제 대시보드라고 써줘",
            "H1 셀에 물류 관제 대시보드 입력",
            "H1 칸에 물류 관제 대시보드 적어주세요",
            "H1 셀에 물류 관제 대시보드 기입해주세요",
            "H1 셀 값에 물류 관제 대시보드 라고 텍스트 넣어줘",
            "물류 관제 대시보드, 이걸 H1에 써주세요",
            "제목은 물류 관제 대시보드, H1에 넣어줘",
            "H1 물류 관제 대시보드",
            "H1 셀 물류 관제 대시보드",
        ],
    )
    def test_writes_the_whole_title_into_the_named_cell(self, text: str) -> None:
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text))
        assert step and step["action"] == "excel_live.write_range", text
        assert step["params"]["start_cell"] == "H1", text
        assert step["params"]["values_2d"] == [["물류 관제 대시보드"]], text

    def test_bare_cell_text_accepts_a_colon(self) -> None:
        from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based

        step = parse_command_rule_based("A1: 분기 보고")
        assert step and step["params"] == {"start_cell": "A1", "values_2d": [["분기 보고"]]}

    @pytest.mark.parametrize(
        "text",
        [
            "B2 정렬해줘",
            "A1 굵게",
            "A1:C3 합계",
            "C3 확인 좀",
            "H1에 넣어줘",
            "A1에 이거 넣어줘",
        ],
    )
    def test_does_not_invent_a_title(self, text: str) -> None:
        """동사·명령이 붙은 문장을 제목 쓰기로 오인하면 엉뚱한 글자가 칸에 박힌다."""
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text)) or {}
        assert step.get("reason", "") != "단일 셀 값 입력 요청(조사 없는 최소 문장)", text

    def test_comma_list_still_goes_to_the_selected_cell(self) -> None:
        """`CO2 농도,512 입력` — CO2는 셀이 아니라 값이다(2026-08-19 ex4 실측).

        조사 없는 최소 문장 규칙이 이걸 가로채면 CO2 셀에 써 버린다.
        """
        from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based

        step = parse_command_rule_based("CO2 농도,512 입력")
        assert step and step["params"]["start_cell"] == "__ACTIVE_CELL__"
        assert step["params"]["values_2d"] == [["CO2 농도,512"]]

    def test_cell_clear_is_not_a_title_write(self) -> None:
        from office_claw_sidecar.services.excel_live_agent import parse_command_rule_based

        step = parse_command_rule_based("A1 지워줘")
        assert step and step["reason"] == "셀 값 삭제 요청"
        assert step["params"]["values_2d"] == [[None]]


class TestConditionalHighlightKeepsItsCondition:
    """2026-08-20 게이트3: 조건을 못 읽으면 **A:Z를 통짜로 칠했다.**

    실측 피해: `클레임 10 넘어가는 데만 빨간색 강조` → F2(12)뿐 아니라 F3(5)도 빨강.
    사용자는 "~만"이라고 말했는데 규칙은 전부 칠하고 성공으로 보고했다.
    """

    @pytest.mark.parametrize(
        ("text", "operator", "threshold"),
        [
            ("클레임 10 넘어가는 데만 빨간색 강조", ">", 10.0),
            ("클레임10↑빨강", ">", 10.0),
            ("클레임 10 넘는 셀 빨갛게", ">", 10.0),
            ("지연건수 5 웃도는 값 노랗게 칠해줘", ">", 5.0),
            ("재고 3 밑도는 칸 빨간색", "<", 3.0),
            ("점수 60 못 미치는 셀 회색으로 강조", "<", 60.0),
            ("매출 100 이상인 셀 초록색 강조", ">=", 100.0),
            ("반품 2 이하인 셀 파란색 강조", "<=", 2.0),
            ("클레임 10↓ 파랑", "<", 10.0),
        ],
    )
    def test_the_threshold_survives(self, text: str, operator: str, threshold: float) -> None:
        # 라우터가 실제로 쓰는 순서 그대로 본다 — 빠른 계획이 먼저고, 없으면 에이전트 규칙이다.
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        clean = normalize_common_typos(text)
        plan = _build_quick_action_plan(clean, None)
        step = plan[0] if plan else parse_command_rule_based(clean)
        assert step, text
        assert step["action"] == "excel_live.highlight_by_condition", (text, step["action"])
        assert step["params"]["operator"] == operator, text
        assert float(step["params"]["threshold"]) == threshold, text

    @pytest.mark.parametrize(
        "text",
        [
            "상태 대기 분홍 강조!",
            "지연된 것만 빨갛게",
            "이상한 값만 노랗게 칠해줘",
        ],
    )
    def test_an_unparseable_condition_does_not_paint_everything(self, text: str) -> None:
        """조건을 못 만들면 **칠하지 않는다** — 물러나면 해석 카드/되묻기가 받는다."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        clean = normalize_common_typos(text)
        plan = _build_quick_action_plan(clean, None)
        step = (plan[0] if plan else parse_command_rule_based(clean)) or {}
        if step.get("action") == "excel_live.fill_range":
            target = str(step["params"].get("target_range", ""))
            raise AssertionError(f"조건을 잃고 통짜로 칠한다: {text} → {target}")

    @pytest.mark.parametrize(
        ("text", "expected"),
        [
            ("A1:C3 노란색으로 칠해줘", "A1:C3"),
            ("표 전체를 노랗게 칠해줘", None),
        ],
    )
    def test_an_explicit_or_broad_fill_still_works(self, text: str, expected) -> None:
        """명시 범위·명시적 전체 표현이 있으면 통짜 칠은 정상 동작이다."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        clean = normalize_common_typos(text)
        plan = _build_quick_action_plan(clean, None)
        step = plan[0] if plan else parse_command_rule_based(clean)
        assert step and step["action"] == "excel_live.fill_range", text
        if expected:
            assert step["params"]["target_range"] == expected, text


class TestGate4Regressions:
    """게이트4가 잡은 것들. 전부 조용한 오실행이었다(카드도 되묻기도 없었다)."""

    @pytest.mark.parametrize(
        "text",
        [
            "H1에 물류 관제 대시보드",
            "H1에는 물류 관제 대시보드",
            "H1에다 물류 관제 대시보드",
            "아 그리고 H1에는 물류 관제 대시보드 라고 제목 박아",
        ],
    )
    def test_a_locative_particle_never_lands_in_the_value(self, text: str) -> None:
        """`H1에 물류 관제 대시보드` → 셀에 **'에 물류 관제 대시보드'**가 써졌다."""
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text))
        assert step and step["params"]["start_cell"] == "H1", text
        assert step["params"]["values_2d"] == [["물류 관제 대시보드"]], text

    def test_charts_only_when_the_data_must_stay(self) -> None:
        """`차트 전부 지워 주세요, 데이터는 그데로 두시고요.` → 표까지 지워졌다.

        오타('그데로')로 보호가 꺼졌고, 고친 뒤에는 한정사 가드가 먼저 걸려
        **아무것도 안 하게** 됐다. 정답은 차트 삭제 하나다.
        """
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        text = normalize_common_typos("차트 전부 지워 주세요, 데이터는 그데로 두시고요.")
        plan = _build_quick_action_plan(text, "A1:F6")
        assert plan and [s["action"] for s in plan] == ["excel_live.delete_charts"], plan

    def test_a_glued_range_is_split_only_when_merging(self) -> None:
        """`H1M1병합` → 콜론이 없어 붙여넣기 범위(A1:F6)를 병합했다."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos("H1M1병합"), "A1:F6")
        assert plan and plan[0]["params"]["target_range"] == "H1:M1", plan
        # 병합 문맥이 아니면 값이 망가지면 안 된다.
        assert normalize_common_typos("A1에 AB12CD34 입력") == "A1에 AB12CD34 입력"

    def test_the_average_typo_is_not_read_as_a_sum(self) -> None:
        """'평규도'가 '합계'로 읽혀 평균 대신 합계 줄을 또 썼다."""
        from office_claw_sidecar.services.excel_aggregate_below import match_aggregate_below
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        assert match_aggregate_below(normalize_common_typos("합계 아래 줄에 평규도 넣어주세요")) == (
            "AVERAGE",
            "평균",
        )

    @pytest.mark.parametrize(
        "text",
        [
            "표 첫 줄을 남색 배경에 흰색 굵은 글시로 바꿔 주세요.",
            "맨 위 제목 줄 눈에 띄게 남색 칸에 흰 글씨 진하게 해줘",
            "머리글 행은 남색 배경에 흰 글씨로 굵게 해줘",
        ],
    )
    def test_the_header_gets_both_the_fill_and_the_bold(self, text: str) -> None:
        """셋 중 둘만 적용되던 문장들 — 어미('굵은')와 색 어휘('흰')가 두 곳에서 갈라져 있었다."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        plan = _build_quick_action_plan(normalize_common_typos(text), None) or []
        actions = [s["action"] for s in plan]
        assert "excel_live.fill_range" in actions, (text, plan)
        assert "excel_live.set_font" in actions, (text, plan)
        fill = next(s for s in plan if s["action"] == "excel_live.fill_range")
        font = next(s for s in plan if s["action"] == "excel_live.set_font")
        assert fill["params"]["fill_color"] == "#002060", text
        assert fill["params"]["target_range"] == "1:1", text
        assert font["params"].get("bold") is True, text


class TestHighlightScopedToHeaderColumn:
    """조건부 강조의 **대상 열**을 머리글로 확정한다(2026-08-20 게이트5).

    게이트5에 남은 조건부 강조 실패 9건 중 7건은 조건이 이미 정확했다.
    `클레임이 10보다 큰 셀만 빨간색으로 칠혀 주세요.` → 규칙이 `> 10`까지 읽어 놓고
    범위만 `__ACTIVE_SELECTION__`이라, 플래너로 넘어가 F2 대신 **F3**이 빨개졌다.
    """

    DIGEST = {
        "active_sheet": "지역성과",
        "sheets": [
            {
                "name": "지역성과",
                "used_range": "A1:F6",
                "columns": [
                    {"letter": "A", "header": "지역"},
                    {"letter": "B", "header": "주문건수"},
                    {"letter": "C", "header": "출고건수"},
                    {"letter": "D", "header": "정시배송률"},
                    {"letter": "E", "header": "지연건수"},
                    {"letter": "F", "header": "클레임"},
                ],
            }
        ],
    }

    def _plan(self, **params):
        base = {"target_range": "__ACTIVE_SELECTION__", "operator": ">", "threshold": 10.0}
        base.update(params)
        return [{"action": "excel_live.highlight_by_condition", "params": base}]

    @pytest.mark.parametrize(
        ("message", "expected"),
        [
            ("클레임이 10보다 큰 셀만 빨간색으로 칠혀 주세요.", "F2:F6"),
            ("클레임10↑빨강", "F2:F6"),
            ("지연건수 3 넘는 칸 빨갛게", "E2:E6"),
            ("정시배송률 95 밑도는 셀 노랗게", "D2:D6"),
        ],
    )
    def test_the_named_header_decides_the_column(self, message: str, expected: str) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        plan = self._plan()
        assert _scope_highlight_to_header_column(plan, message, self.DIGEST) == expected
        assert plan[0]["params"]["target_range"] == expected

    def test_an_explicit_range_is_left_alone(self) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        plan = self._plan(target_range="F2:F6")
        assert _scope_highlight_to_header_column(plan, "F2:F6에서 10 넘는 것", self.DIGEST) == ""
        assert plan[0]["params"]["target_range"] == "F2:F6"

    @pytest.mark.parametrize("message", ["10 넘는 셀 빨갛게", "건수 10 넘는 셀 빨갛게"])
    def test_an_unresolvable_header_stays_with_the_planner(self, message: str) -> None:
        """머리글이 없거나 여러 개에 걸리면 추측하지 않는다 — 예전처럼 플래너 몫이다.

        '건수'는 주문건수·출고건수·지연건수 셋에 걸린다.
        """
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        assert _scope_highlight_to_header_column(self._plan(), message, self.DIGEST) == ""

    def test_other_actions_are_untouched(self) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        plan = [{"action": "excel_live.fill_range", "params": {"target_range": "__ACTIVE_SELECTION__"}}]
        assert _scope_highlight_to_header_column(plan, "클레임 빨갛게", self.DIGEST) == ""


class TestFormatCodeDecimalPlaces:
    """`0.1`은 표시 형식이 아니라 "소수 한 자리"라는 뜻이다(2026-08-20 게이트4·5).

    엑셀에서 `.` 뒤의 `1`은 자릿수가 아니라 리터럴이라 97.14가 `97.11`로 보인다.
    `percent_format` 5문장이 전부 이 값으로 틀렸다.
    """

    @pytest.mark.parametrize(
        ("given", "expected"),
        [
            ("0.1", "0.0"),
            ("0.2", "0.00"),
            ("0.3", "0.000"),
            ("#,##0.1", "#,##0.0"),
            ("0.0", "0.0"),
            ("0.00", "0.00"),
            ("#,##0", "#,##0"),
            ("0.0%", "0.0%"),
        ],
    )
    def test_a_digit_count_is_read_as_places(self, given: str, expected: str) -> None:
        from office_claw_sidecar.services.excel_live_plan_validator import _normalize_number_format

        assert _normalize_number_format(given) == expected


class TestPlanSanityCatchesPreamblesAndCircularFormulas:
    """게이트4·5가 드러낸 두 부류 — 둘 다 사후조건으로는 못 잡는다."""

    def test_a_preamble_is_not_a_value(self) -> None:
        """`새로 데이터를 받아서 다시 넣어야 해서요` 가 A1에 그대로 박혔다."""
        from office_claw_sidecar.services.excel_plan_sanity import check_plan_sanity

        issues = check_plan_sanity(
            [
                {
                    "action": "excel_live.write_range",
                    "params": {"start_cell": "A1", "values_2d": [["새로 데이터를 받아서 다시 넣어야 해서요"]]},
                }
            ],
            message="새로 데이터를 받아서 다시 넣어야 해서요, 지금 붙여넣은 표 안의 값을 전부 지워 주세요.",
        )
        assert [i.code for i in issues] == ["value_is_a_directive"]

    def test_a_formula_that_refers_to_its_own_cell_is_caught(self) -> None:
        """`요약!A2 = =SUM(A2:A2)` — 엑셀은 0을 보여 주고 사후조건은 통과시킨다."""
        from office_claw_sidecar.services.excel_plan_sanity import check_plan_sanity

        issues = check_plan_sanity(
            [{"action": "excel_live.set_formula", "params": {"range_ref": "A2", "formula_a1": "=SUM(A2:A2)"}}],
            message="A2에 지역성과 시트 주문건수 합계를 수식으로 넣어 주세요",
        )
        assert [i.code for i in issues] == ["formula_refers_to_itself"]

    @pytest.mark.parametrize(
        ("cell", "formula"),
        [
            ("A2", "=SUM(지역성과!B2:B6)"),
            ("B7", "=SUM(B2:B6)"),
            ("B8", "=AVERAGE(B2:B7)"),
        ],
    )
    def test_ordinary_formulas_pass(self, cell: str, formula: str) -> None:
        from office_claw_sidecar.services.excel_plan_sanity import check_plan_sanity

        issues = check_plan_sanity(
            [{"action": "excel_live.set_formula", "params": {"range_ref": cell, "formula_a1": formula}}],
            message="합계 넣어줘",
        )
        assert issues == []


class TestValueEqualsHighlightAndFormulaHead:
    """게이트6이 드러낸 세 결함 — 셋 다 표를 통째로 망친다."""

    STATUS_DIGEST = {
        "active_sheet": "지연경고",
        "sheets": [
            {
                "name": "지연경고",
                "used_range": "A1:D5",
                "columns": [
                    {"letter": "A", "header": "운송장"},
                    {"letter": "B", "header": "구간"},
                    {"letter": "C", "header": "지연시간"},
                    {"letter": "D", "header": "상태"},
                ],
                "sample_rows": [
                    ["T1", "서울", "2", "대기"],
                    ["T2", "부산", "1", "완료"],
                    ["T3", "대구", "3", "대기"],
                ],
            }
        ],
    }

    @pytest.mark.parametrize(
        "formula",
        ["=SUM(B2:B6)", "=B2-C3", "=100", "=(A1+B1)/2", "=지역성과!B2", '=IF(B2>=70,"통과","미달")', "=-5"],
    )
    def test_real_formulas_are_recognized(self, formula: str) -> None:
        from office_claw_sidecar.routers.excel_live import _looks_like_a_formula

        assert _looks_like_a_formula(formula) is True, formula

    @pytest.mark.parametrize("text", ["=대기", "=완료", "=진행중"])
    def test_a_korean_word_after_equals_is_not_a_formula(self, text: str) -> None:
        """`상태=대기`의 '='는 같다는 말이다.

        2026-08-20 게이트6: `ㅇㅇ 상태=대기 셀만 분홍 강조`가
        set_formula(A1:D5, '=대기')가 되어 **표 전체를 수식으로 덮었다.**
        """
        from office_claw_sidecar.routers.excel_live import _looks_like_a_formula

        assert _looks_like_a_formula(text) is False, text

    @pytest.mark.parametrize(
        "message",
        [
            "대기만 분홍",
            "상태 대기 분홍 강조!",
            "회의 때 대기 중인 운송장이 몇 개인지 봐야 해서요, 상태 열에서 대기인 셀만 분홍색으로 강조해 주세요",
        ],
    )
    def test_a_value_that_exists_in_one_column_decides_the_range(self, message: str) -> None:
        """값('대기')이 통합문서에 실제로 있고 그 열이 하나뿐이면 추측이 아니라 확인이다."""
        from office_claw_sidecar.routers.excel_live import _value_equals_highlight

        plan = _value_equals_highlight(message, self.STATUS_DIGEST)
        assert plan, message
        params = plan[0]["params"]
        assert params["target_range"] == "D2:D5", message
        assert params["value"] == "대기", message

    @pytest.mark.parametrize("message", ["표 전체를 노랗게 칠해줘", "머리글 남색으로", "A1:C3 노란색"])
    def test_it_does_not_fire_without_a_matching_value(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _value_equals_highlight

        assert _value_equals_highlight(message, self.STATUS_DIGEST) == []

    def test_the_named_column_wins_over_a_longer_header(self) -> None:
        """"…**운송장**이 몇 개인지 …, **상태 열에서** 대기인 셀만" — 길이로 고르면 운송장이 이긴다."""
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        plan = [
            {
                "action": "excel_live.highlight_by_condition",
                "params": {"target_range": "__ACTIVE_SELECTION__", "operator": "==", "threshold": 0, "value": "대기"},
            }
        ]
        message = "회의 때 대기 중인 운송장이 몇 개인지 보여 줘야 해서요, 상태 열에서 대기인 셀만 분홍색으로 강조해 주세요"
        assert _scope_highlight_to_header_column(plan, message, self.STATUS_DIGEST) == "D2:D5"

    def test_an_aggregate_request_is_not_a_value_list(self) -> None:
        """`한 줄로 합계, 표 바로 아래에 넣어줘` → A1·B1에 **명령문이 써졌다**(게이트5)."""
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        assert parse_rangeless_row_write("한 줄로 합계, 표 바로 아래에 넣어줘", "A1:F6") is None

    def test_a_plain_value_list_still_writes(self) -> None:
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        step = parse_rangeless_row_write("지역,주문건수,출고건수 입력해줘", "A1:C1")
        assert step and step["params"]["values_2d"] == [["지역", "주문건수", "출고건수"]]

    @pytest.mark.parametrize(
        "text",
        ["ㅇㅇ 그 다음 줄엔 평균값 한 줄 더", "합계 다음 줄에는 평균도 같이 넣어 주세요"],
    )
    def test_the_next_row_counts_as_below(self, text: str) -> None:
        """사람은 '아래'만큼이나 '다음 줄'이라고 말한다."""
        from office_claw_sidecar.services.excel_aggregate_below import match_aggregate_below
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        assert match_aggregate_below(normalize_common_typos(text)) == ("AVERAGE", "평균")


class TestBinderDoesNotUndoAConfirmedHighlight:
    """게이트7: 규칙이 통합문서를 보고 확정한 강조를 **바인더가 되돌렸다.**

    - "…대기 건을 **빨**리 찾아야 해서요, … **분홍색**으로 칠해 주세요" → D2·D4가 **빨강**.
      색 어휘에 한 음절 `("빨", "#FF0000")`이 있어 '빨리'에 걸렸다.
    - "…**운송장**이 몇 개인지 …, **상태 열에서** 대기인 셀만" → 확정한 `D2:D5`가 `A:A`로
      되돌아가 0칸 강조가 됐다(머리글 첫 히트가 운송장).
    """

    ENTRY = {
        "name": "지연경고",
        "used_range": "A1:D5",
        "columns": [
            {"letter": "A", "header": "운송장"},
            {"letter": "B", "header": "구간"},
            {"letter": "C", "header": "지연시간"},
            {"letter": "D", "header": "상태"},
        ],
        "sample_rows": [["T1", "서울", "2", "대기"]],
    }

    def _bind(self, params, message):
        from office_claw_sidecar.services.excel_param_binder import _bind_condition_format

        _bind_condition_format(params, message=message, entry=self.ENTRY, digest={"sheets": [self.ENTRY]})
        return params

    def test_a_hurry_word_is_not_the_color_red(self) -> None:
        params = self._bind(
            {"target_range": "D2:D5", "operator": "==", "threshold": 0, "value": "대기", "fill_color": "#FFC0CB"},
            "지연경고 표에서 대기 건을 빨리 찾아야 해서요, 상태가 '대기'인 셀만 분홍색으로 칠해 주세요.",
        )
        assert params["fill_color"].upper() == "#FFC0CB", params

    @pytest.mark.parametrize(
        ("message", "expected"),
        [("빨간색으로 칠해줘", "#FF0000"), ("빨갛게 강조", "#FF0000"), ("빨강 강조", "#FF0000")],
    )
    def test_real_red_words_still_work(self, message: str, expected: str) -> None:
        params = self._bind({"target_range": "D2:D5", "value": "대기"}, message)
        assert params["fill_color"].upper() == expected.upper()

    def test_a_column_scoped_range_is_not_widened(self) -> None:
        params = self._bind(
            {"target_range": "D2:D5", "operator": "==", "threshold": 0, "value": "대기"},
            "회의 때 대기 중인 운송장이 몇 개인지 봐야 해서요, 상태 열에서 대기인 셀만 분홍색으로 강조해 주세요",
        )
        assert params["target_range"] == "D2:D5", params

    def test_a_broad_range_is_still_narrowed_to_the_named_column(self) -> None:
        params = self._bind(
            {"target_range": "A:Z", "operator": "==", "threshold": 0, "value": "대기"},
            "상태 열에서 대기인 셀만 분홍색으로 강조해 주세요",
        )
        assert params["target_range"] in {"D:D", "D2:D5"}, params

    def test_the_named_column_beats_the_first_mention(self) -> None:
        params = self._bind(
            {"target_range": "A:Z", "operator": "==", "threshold": 0, "value": "대기"},
            "대기 중인 운송장이 몇 개인지 보려고요, 상태 열에서 대기인 셀만 강조",
        )
        assert params["target_range"].startswith("D"), params


class TestBareAggregateAfterPaste:
    """붙여넣기 직후의 `합계!` 한 마디 — 뜻이 하나뿐이다(그 표 아래 합계 줄)."""

    @pytest.mark.parametrize(
        ("message", "expected"),
        [
            ("합계!", ("SUM", "합계")),
            ("합계", ("SUM", "합계")),
            ("ㅇㅇ 평균", ("AVERAGE", "평균")),
            ("총합 좀", ("SUM", "총합")),
        ],
    )
    def test_it_fires_only_with_a_paste_context(self, message: str, expected) -> None:
        from office_claw_sidecar.routers.excel_live import _bare_aggregate_after_paste

        assert _bare_aggregate_after_paste(message, "A1:F6") == expected
        # 맥락이 없으면 어디에 넣을지 모른다 — 추측하지 않는다.
        assert _bare_aggregate_after_paste(message, None) is None

    @pytest.mark.parametrize("message", ["합계를 어디에 넣지", "평균이 얼마야?", "합계 열 만들어줘", "지역"])
    def test_it_does_not_fire_on_longer_sentences(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _bare_aggregate_after_paste

        assert _bare_aggregate_after_paste(message, "A1:F6") is None


class TestMacroPlannerDoesNotEatSimpleRequests:
    """머리말의 '보고서'와 본문의 '만들어'를 합쳐 매크로로 오인하던 문제.

    2026-08-20 게이트8의 조용한 오실행 4건 중 3건이 이 하나였다.
    `보고서 앞쪽에 요약 페이지가 필요하다고 해서요, 요약이라는 이름의 새 시트를 만들어 주세요.`
    → 21단계 매크로로 분해되고 **시트는 안 만들어졌다**.
    """

    @pytest.mark.parametrize(
        "message",
        [
            "보고서 앞쪽에 요약 페이지가 필요하다고 해서요, 요약이라는 이름의 새 시트를 만들어 주세요.",
            "보고서에 시각 자료가 있으면 좋겠다고 해서요, 정시배송률 열을 가지고 꺾은선 그래프를 하나 만들어 주시면 감사하겠습니다",
            "보고서에 그래프를 하나 넣으라고 하셔서요, 지연건수 열을 가지고 막대 그래프를 만들어 주시면 감사하겠습니다.",
            "요약 시트 만들어줘",
        ],
    )
    def test_a_reason_preamble_is_not_a_deliverable(self, message: str) -> None:
        from office_claw_sidecar.services.excel_macro_planner import looks_like_macro_request

        assert looks_like_macro_request(message) is False, message

    @pytest.mark.parametrize(
        "message",
        [
            "대시보드 만들어줘",
            "매출 대시보드 하나 만들어 주세요",
            "월간 보고서 만들어줘",
            "영업 현황판 좀 구성해줘",
        ],
    )
    def test_a_real_composite_request_still_goes_to_the_macro_planner(self, message: str) -> None:
        from office_claw_sidecar.services.excel_macro_planner import looks_like_macro_request

        assert looks_like_macro_request(message) is True, message


class TestNumberFormatScopedToNamedColumns:
    """표시 형식도 "어느 열"이 낱말로만 있다(2026-08-20 게이트8 `comma_cols`).

    `여기 주문건수 출고건수는 1000단위로 쉼표 넣어줘` → **B2:B3만** 바뀌었다.
    규칙은 `#,##0`을 정확히 냈는데 범위가 표 전체라 플래너로 넘어갔고, 플래너가 좁게 집었다.
    """

    DIGEST = {
        "active_sheet": "지역성과",
        "sheets": [
            {
                "name": "지역성과",
                "used_range": "A1:F6",
                "columns": [
                    {"letter": "A", "header": "지역"},
                    {"letter": "B", "header": "주문건수"},
                    {"letter": "C", "header": "출고건수"},
                    {"letter": "D", "header": "정시배송률"},
                    {"letter": "E", "header": "지연건수"},
                    {"letter": "F", "header": "클레임"},
                ],
            }
        ],
    }

    def _plan(self, target="__ACTIVE_SELECTION__", code="#,##0"):
        return [{"action": "excel_live.set_number_format", "params": {"target_range": target, "format_code": code}}]

    def test_two_named_columns_become_two_steps(self) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_number_format_to_headers

        plan = self._plan()
        out = _scope_number_format_to_headers(plan, "주문건수 출고건수는 천 단위 콤마로", self.DIGEST)
        assert out == "B2:B6,C2:C6", out
        assert [s["params"]["target_range"] for s in plan] == ["B2:B6", "C2:C6"]
        assert all(s["params"]["format_code"] == "#,##0" for s in plan)

    def test_one_named_column(self) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_number_format_to_headers

        plan = self._plan(code="0.0")
        assert _scope_number_format_to_headers(plan, "정시배송률 소수 한 자리!", self.DIGEST) == "D2:D6"

    @pytest.mark.parametrize(
        "message",
        ["표 전체 천 단위 콤마", "주문출고콤마", "B2:C6 콤마로", "전부 콤마"],
    )
    def test_it_does_not_guess(self, message: str) -> None:
        """머리글을 못 짚거나 원문이 범위를 적었으면 그대로 둔다."""
        from office_claw_sidecar.routers.excel_live import _scope_number_format_to_headers

        plan = self._plan()
        assert _scope_number_format_to_headers(plan, message, self.DIGEST) == ""
        assert len(plan) == 1

    def test_an_explicit_range_in_the_plan_is_left_alone(self) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_number_format_to_headers

        plan = self._plan(target="B2:B6")
        assert _scope_number_format_to_headers(plan, "주문건수 천 단위 콤마", self.DIGEST) == ""


class TestHeaderNamesToleranceForTypos:
    """머리글 이름의 오타 한 글자 — 게이트9가 잡은 내 회귀 2건.

    `주문건수랑 **출고겅수** 천단위 콤마` → 주문건수(B)만 서식이 걸렸다.
    열 좁히기를 넣기 전에는 표 전체에 걸려 우연히 맞았는데, 정확해지면서 오타난 쪽을 잃었다.
    """

    DIGEST = {
        "active_sheet": "지역성과",
        "sheets": [
            {
                "name": "지역성과",
                "used_range": "A1:F6",
                "columns": [
                    {"letter": "A", "header": "지역"},
                    {"letter": "B", "header": "주문건수"},
                    {"letter": "C", "header": "출고건수"},
                    {"letter": "D", "header": "정시배송률"},
                    {"letter": "E", "header": "지연건수"},
                    {"letter": "F", "header": "클레임"},
                ],
            }
        ],
    }

    @pytest.mark.parametrize(
        ("message", "expected"),
        [
            ("주문건수랑 출고겅수 천단위 콤마 표기로 바꺼", "B2:B6,C2:C6"),
            ("주문건슈 출고건수 천단의 컴마로 보이게 해", "B2:B6,C2:C6"),
            ("주문건수 출고건수는 천 단위 콤마로", "B2:B6,C2:C6"),
            ("지연건수만 콤마", "E2:E6"),
        ],
    )
    def test_one_wrong_character_still_finds_the_column(self, message: str, expected: str) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_number_format_to_headers

        plan = [
            {
                "action": "excel_live.set_number_format",
                "params": {"target_range": "__ACTIVE_SELECTION__", "format_code": "#,##0"},
            }
        ]
        assert _scope_number_format_to_headers(plan, message, self.DIGEST) == expected, message

    @pytest.mark.parametrize(
        ("header", "message", "expected"),
        [
            ("주문건수", "지연건수만 콤마", False),
            ("출고건수", "지연건수만 콤마", False),
            ("지연건수", "지연건수만 콤마", True),
            ("지역", "지연건수만 콤마", False),
        ],
    )
    def test_similar_headers_do_not_cross_match(self, header: str, message: str, expected: bool) -> None:
        """두 글자 이상 다르면 남의 열이다 — 오차 허용이 열을 뒤바꾸면 안 된다."""
        from office_claw_sidecar.routers.excel_live import _header_in_message
        from office_claw_sidecar.services.excel_aggregate_below import _norm_header

        assert _header_in_message(header, _norm_header(message)) is expected


class TestAggregateGuardDoesNotEatPastedValues:
    """붙여넣은 **값 안의 '평균'**이 집계 가드를 켜던 회귀.

    2026-08-20 28각본 배터리(1,838턴)의 유일한 실패였다 — 내가 만든 것.
    `이 표 아래에 자산 요약,값; … 평균 임대료 평당(원),32500; … 입력해줘`
    → 값 나열이 집계 요청으로 읽혀 표 만들기 인터뷰로 샜다.
    """

    PASTE = (
        "이 표 아래에 자산 요약,값; 총 자산 가치 감정가(원),285430000000; "
        "총 대출 잔액(원),126800000000; LTV(%),44.4; WALE 가중 평균 임대 만료(년),2.83; "
        "평균 임대료 평당(원),32500; 관리비 회수율(%),78.6 입력해줘"
    )

    def test_a_long_value_list_is_still_a_write(self) -> None:
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_rangeless_row_write,
        )

        step = parse_rangeless_row_write(normalize_common_typos(self.PASTE), "A9:B17")
        assert step and step["action"] == "excel_live.write_range"
        values = step["params"]["values_2d"]
        assert values[0] == ["자산 요약", "값"], values[0]
        assert ["평균 임대료 평당(원)", 32500] in values, values

    @pytest.mark.parametrize(
        "text",
        ["한 줄로 합계, 표 바로 아래에 넣어줘", "합계를 표 아래에 한 줄로 넣어줘", "평균 한 줄 밑에 넣어줘"],
    )
    def test_a_short_aggregate_command_is_not_a_value_list(self, text: str) -> None:
        from office_claw_sidecar.services.excel_live_agent import parse_rangeless_row_write

        assert parse_rangeless_row_write(text, "A1:F6") is None, text


class TestDestructiveScopeGuards:
    """2026-08-20 파괴 위험 게이트가 드러낸 것들 — 전부 표를 통째로 망쳤다."""

    GRADES = {
        "active_sheet": "성적부",
        "sheets": [
            {
                "name": "성적부",
                "used_range": "A1:C5",
                "columns": [
                    {"letter": "A", "header": "학생"},
                    {"letter": "B", "header": "점수"},
                    {"letter": "C", "header": "결석"},
                ],
            }
        ],
    }

    @pytest.mark.parametrize(
        "message",
        [
            "결석 열만 비워줘",
            "결석 값들만 지워 주세요",
            "결석 칸 내용만 싹 지워줘",
            "ㅇㅇ 결석만 비워",
            "새 학기라 결석 기록만 초기화해 주세요",
            "결석 열 값 전부 삭제해줘 (다른 열은 유지)",
            "결석열만비워",
        ],
    )
    def test_only_the_named_column_is_cleared(self, message: str) -> None:
        """`결석 열만 비워줘`가 **표 전체를 비웠다**(12문장 중 8)."""
        from office_claw_sidecar.routers.excel_live import (
            _build_quick_action_plan,
            _scope_clear_to_header_column,
        )

        plan = _build_quick_action_plan(message, None)
        scoped = _scope_clear_to_header_column(plan, message, self.GRADES)
        assert scoped, message
        assert [s["params"]["target_range"] for s in scoped] == ["C2:C5"], message

    def test_a_protected_column_is_not_the_target(self) -> None:
        """"이름이랑 **점수는 건드리지 말고**" — 지키라고 부른 열을 대상으로 삼으면 안 된다."""
        from office_claw_sidecar.routers.excel_live import _scope_clear_to_header_column

        scoped = _scope_clear_to_header_column(
            None, "결석 수치만 없애줘, 이름이랑 점수는 건드리지 말고", self.GRADES
        )
        assert scoped and scoped[0]["params"]["target_range"] == "C2:C5"

    @pytest.mark.parametrize("message", ["표 전체 지워줘", "시트 전체 초기화해줘", "싹 다 지워", "전부 지워"])
    def test_a_whole_sheet_reset_is_left_alone(self, message: str) -> None:
        """넓게 지우라고 한 문장까지 한 열로 좁히면 안 된다."""
        from office_claw_sidecar.routers.excel_live import (
            _build_quick_action_plan,
            _scope_clear_to_header_column,
        )

        plan = _build_quick_action_plan(message, None)
        assert _scope_clear_to_header_column(plan, message, self.GRADES) == [], message

    CROSS = {
        "active_sheet": "요약",
        "sheets": [
            {"name": "성적부", "used_range": "A1:C5", "columns": []},
            {"name": "요약", "used_range": "A1:A1", "columns": []},
        ],
    }

    @pytest.mark.parametrize(
        "message",
        [
            "결석 합계 성적부에서 가져와서 A2에",
            "성적부에 있는 결석 수 전부 합쳐서 A2에 써줘",
            "학년 회의 자료라서요, 성적부의 결석 합계를 A2에 넣어 주시면 됩니다",
            "A2에 성적부 결석 합계 넣어줘",
        ],
    )
    def test_a_sheet_named_as_a_source_is_not_the_write_target(self, message: str) -> None:
        """`성적부**에서** … A2에`가 **성적부에** 써져 학생 이름을 덮었다.

        2026-08-19에 넣은 보호는 어순만 봤는데, 이 문장들은 시트가 앞에 온다.
        한국어는 어순이 아니라 **조사**로 출처를 표시한다.
        """
        from office_claw_sidecar.services.excel_param_binder import resolve_sheet_from_message

        assert resolve_sheet_from_message(message, self.CROSS, default="요약") == "요약", message

    @pytest.mark.parametrize(
        "message",
        ["성적부 시트 A1에 제목 써줘", "성적부에 합계 넣어줘", "성적부 정렬해줘"],
    )
    def test_a_sheet_named_as_a_destination_still_wins(self, message: str) -> None:
        from office_claw_sidecar.services.excel_param_binder import resolve_sheet_from_message

        assert resolve_sheet_from_message(message, self.CROSS, default="요약") == "성적부", message


class TestSelfReviewFixes:
    """2026-08-20 자체 검토(수정 검토)에서 잡은 결함 4건의 핀."""

    def test_an_exact_header_mention_is_not_lent_to_a_similar_one(self) -> None:
        """`매출액만 콤마`가 한 글자 다른 **매입액까지** 서식을 걸었다."""
        from office_claw_sidecar.routers.excel_live import _scope_number_format_to_headers

        digest = {
            "active_sheet": "장부",
            "sheets": [
                {
                    "name": "장부",
                    "used_range": "A1:C5",
                    "columns": [
                        {"letter": "A", "header": "월"},
                        {"letter": "B", "header": "매출액"},
                        {"letter": "C", "header": "매입액"},
                    ],
                }
            ],
        }
        plan = [
            {
                "action": "excel_live.set_number_format",
                "params": {"target_range": "__ACTIVE_SELECTION__", "format_code": "#,##0"},
            }
        ]
        assert _scope_number_format_to_headers(plan, "매출액만 콤마", digest) == "B2:B5"
        # 오타는 여전히 받는다 — 그 창이 다른 머리글과 정확히 일치하지 않으므로.
        plan2 = [
            {
                "action": "excel_live.set_number_format",
                "params": {"target_range": "__ACTIVE_SELECTION__", "format_code": "#,##0"},
            }
        ]
        assert _scope_number_format_to_headers(plan2, "매출액 매입액 둘 다 콤마", digest) == "B2:B5,C2:C5"

    GRADES = {
        "active_sheet": "성적부",
        "sheets": [
            {
                "name": "성적부",
                "used_range": "A1:C5",
                "columns": [
                    {"letter": "A", "header": "학생"},
                    {"letter": "B", "header": "점수"},
                    {"letter": "C", "header": "결석"},
                ],
            }
        ],
    }

    def test_a_column_named_with_ppaego_is_protected(self) -> None:
        """`결석 값들 **빼고** 나머지 비워줘` — 지키라고 한 결석 열을 지웠다."""
        from office_claw_sidecar.routers.excel_live import _scope_clear_to_header_column

        assert _scope_clear_to_header_column(None, "결석 값들 빼고 나머지 비워줘", self.GRADES) == []
        assert _scope_clear_to_header_column(None, "결석 제외하고 다 비워", self.GRADES) == []

    def test_a_clear_verb_breaks_the_protection(self) -> None:
        """`결석 열 **비우고** 나머지는 그대로` — '그대로'는 나머지 얘기지 결석 얘기가 아니다."""
        from office_claw_sidecar.routers.excel_live import _scope_clear_to_header_column

        out = _scope_clear_to_header_column(None, "결석 열 비우고 나머지는 그대로 둬", self.GRADES)
        assert out and out[0]["params"]["target_range"] == "C2:C5"
        out2 = _scope_clear_to_header_column(None, "점수 빼고 결석만 지워줘", self.GRADES)
        assert out2 and out2[0]["params"]["target_range"] == "C2:C5"

    @pytest.mark.parametrize(
        ("text", "expected"),
        [
            ("H1 2026년 실적", [["2026년 실적"]]),
            ("H1 3분기 성적", [["3분기 성적"]]),
            ("A1 주요 업적", [["주요 업적"]]),
            ("A2에 45 적으세요", [[45]]),
            ("B3에 준비중 넣으세요", [["준비중"]]),
        ],
    )
    def test_a_verb_stem_does_not_eat_the_last_character(self, text: str, expected) -> None:
        """`H1 2026년 실적` → 셀에 `'2026년 실'` — 어간 '적'이 끝 글자를 동사로 먹었다."""
        from office_claw_sidecar.services.excel_live_agent import (
            normalize_common_typos,
            parse_command_rule_based,
        )

        step = parse_command_rule_based(normalize_common_typos(text))
        assert step and step["params"]["values_2d"] == expected, text

    def test_a_merged_cell_noun_phrase_is_not_a_range(self) -> None:
        """`제품코드 X1Y2 병합 셀에 넣어줘` — 값 X1Y2가 X1:Y2로 변조됐다."""
        from office_claw_sidecar.services.excel_live_agent import normalize_common_typos

        assert normalize_common_typos("제품코드 X1Y2 병합 셀에 넣어줘") == "제품코드 X1Y2 병합 셀에 넣어줘"
        assert normalize_common_typos("H1M1병합") == "H1:M1병합"

    CROSS = {
        "active_sheet": "요약",
        "sheets": [
            {"name": "성적부", "used_range": "A1:C5", "columns": []},
            {"name": "요약", "used_range": "A1:A1", "columns": []},
        ],
    }

    @pytest.mark.parametrize(
        ("message", "expected"),
        [
            ("성적부 기준으로 A2에 합계 넣어줘", "요약"),
            ("성적부의 A2에 제목 써줘", "성적부"),
            ("성적부의 A2에 수식 걸어줘", "성적부"),
            ("성적부에서 A2에 값 넣어줘", "성적부"),
        ],
    )
    def test_a_source_marker_followed_by_a_cell_is_a_destination(self, message: str, expected: str) -> None:
        """`성적부의 **A2**에`는 성적부 안의 대상이다 — 마커 뒤 셀 참조면 출처가 아니다."""
        from office_claw_sidecar.services.excel_param_binder import resolve_sheet_from_message

        assert resolve_sheet_from_message(message, self.CROSS, default="요약") == expected, message


class TestCrossSheetAggregateWordOrder:
    """크로스시트 집계가 **어순 변형**에서 통째로 빠지던 문제.

    2026-08-20 파괴 게이트: 12문장 중 4만 계획이 나왔다. 나머지는 규칙이 비어
    플래너가 명령문 조각이나 `'SUM'`을 셀에 썼다(원본은 안 망가뜨렸지만 결과가 틀렸다).
    """

    GRID = [
        ["학생", "점수", "결석"],
        ["김민준", 88, 2],
        ["이서연", 94, 0],
        ["박도윤", 71, 5],
        ["최지우", 83, 1],
    ]

    def _reader(self, name):
        return ("A1:C5", self.GRID) if name == "성적부" else ("A1:A1", [[None]])

    @pytest.mark.parametrize(
        "message",
        [
            "A2에 성적부 결석 합계 넣어줘",
            "결석 합계 성적부에서 가져와서 A2에",
            "성적부에 있는 결석 수 전부 합쳐서 A2에 써줘",
            "학년 회의 자료라서요, 성적부의 결석 합계를 A2에 넣어 주시면 됩니다",
            "A2에다 성적부 결석 총계 계산해서 넣어줘",
            "A2에 성적부 결석 합계 수식으로 넣어줘, 나중에도 연동되게",
            "성적부 시트 결석 열 합계를 요약 A2에 부탁해요",
            "결석 몇 번인지 다 합쳐서 A2에 보여줘 (성적부 기준)",
            "ㅇㅇ 성적부 결석 합 A2",
            "성적부 결석 다 더한 값을 A2에 넣어 주세요",
            "A2 칸에 성적부 시트 결석 총합 좀",
            "성적부결석합계 A2",
        ],
    )
    def test_every_word_order_yields_the_same_formula(self, message: str) -> None:
        from office_claw_sidecar.services.excel_aggregate_below import (
            build_cross_sheet_aggregate_plan,
        )

        steps = build_cross_sheet_aggregate_plan(message, self._reader, ["성적부", "요약"])
        assert steps, message
        assert steps[0]["params"]["range_ref"] == "A2", message
        assert steps[0]["params"]["formula_a1"] == "=SUM('성적부'!C2:C5)", message

    @pytest.mark.parametrize(
        "message",
        ["결석 합계 알려줘", "성적부 정렬해줘", "A2에 제목 써줘"],
    )
    def test_it_does_not_fire_without_a_cross_sheet_request(self, message: str) -> None:
        from office_claw_sidecar.services.excel_aggregate_below import (
            build_cross_sheet_aggregate_plan,
        )

        assert build_cross_sheet_aggregate_plan(message, self._reader, ["성적부", "요약"]) == []


class TestDashboardStackedTables:
    """대시보드는 한 시트에 표를 여럿 쌓는다 — 머리글 탐색이 1행만 보면 아래쪽 표를 놓친다.

    2026-08-20 ex23 실측: 구매요약 시트에 표가 셋(KPI A4:H9 · 카테고리 A14:F20 · 지연알림 A25:F30).
    `지연일수 5 넘는 셀만 빨간색으로 칠해줘` → **0칸 칠해짐**. '지연일수'는 A25 표의 머리글이다.
    """

    DIGEST = {
        "active_sheet": "구매요약",
        "sheets": [
            {
                "name": "구매요약",
                "used_range": "A1:I31",
                "columns": [{"letter": "A", "header": "구매 운영 대시보드"}],
                "blocks": [
                    {
                        "ref": "A4:H9",
                        "header_row": 4,
                        "first_data_row": 5,
                        "last_row": 9,
                        "columns": [
                            {"letter": "A", "header": "지표"},
                            {"letter": "B", "header": "값"},
                            {"letter": "C", "header": "전월 대비(%)"},
                        ],
                    },
                    {
                        "ref": "A25:F30",
                        "header_row": 25,
                        "first_data_row": 26,
                        "last_row": 30,
                        "columns": [
                            {"letter": "A", "header": "PO번호"},
                            {"letter": "E", "header": "지연일수"},
                            {"letter": "F", "header": "금액"},
                        ],
                    },
                ],
            }
        ],
    }

    def _plan(self):
        return [
            {
                "action": "excel_live.highlight_by_condition",
                "params": {"target_range": "__ACTIVE_SELECTION__", "operator": ">", "threshold": 5.0},
            }
        ]

    @pytest.mark.parametrize(
        ("message", "expected"),
        [
            ("지연일수 5 넘는 셀만 빨간색으로 칠해줘", "E26:E30"),
            ("금액 100만 넘는 셀 빨갛게", "F26:F30"),
            ("전월 대비가 0 밑도는 셀만 빨간색", "C5:C9"),
        ],
    )
    def test_a_header_in_a_lower_table_is_found(self, message: str, expected: str) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        plan = self._plan()
        assert _scope_highlight_to_header_column(plan, message, self.DIGEST) == expected, message

    def test_an_unknown_header_still_backs_off(self) -> None:
        from office_claw_sidecar.routers.excel_live import _scope_highlight_to_header_column

        assert _scope_highlight_to_header_column(self._plan(), "없는열 5 넘는 셀", self.DIGEST) == ""


class TestBlankCellCondition:
    """`입고예정일이 비어 있는 행만 노란색으로 칠해줘` — 빈 칸이 조건인 문형.

    2026-08-20 ex23 실측: 규칙이 없어 모델이 엉뚱한 한 칸(N1)을 칠했다.
    실행기에도 표현 수단이 없어 `==`/`=`/`isblank` 셋 다 0칸이었다.
    """

    DIGEST = {
        "active_sheet": "발주",
        "sheets": [
            {
                "name": "발주",
                "used_range": "A1:B5",
                "columns": [{"letter": "A", "header": "PO"}, {"letter": "B", "header": "입고예정일"}],
            }
        ],
    }

    @pytest.mark.parametrize(
        "message",
        ["입고예정일이 비어 있는 행만 노란색으로 칠해줘", "입고예정일 빈 칸만 강조해줘", "입고예정일 공란만 표시해줘"],
    )
    def test_it_builds_a_blank_condition(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _blank_condition_highlight

        plan = _blank_condition_highlight(message, self.DIGEST)
        assert plan, message
        assert plan[0]["params"]["target_range"] == "B2:B5", message
        assert plan[0]["params"]["operator"] == "isblank", message

    @pytest.mark.parametrize("message", ["입고예정일 값이 있는 셀 강조", "PO 강조해줘", "빈 칸 세어줘"])
    def test_it_does_not_fire_otherwise(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _blank_condition_highlight

        assert _blank_condition_highlight(message, self.DIGEST) == []

    def test_the_executor_paints_only_empty_cells(self, tmp_path) -> None:
        from openpyxl import Workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / "t.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "발주"
        for row in [["PO", "입고예정일"], ["P1", "2025-05-26"], ["P2", None], ["P3", "2025-05-27"], ["P4", None]]:
            ws.append(row)
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "발주")
        result = service.highlight_by_condition(None, "발주", "B2:B5", "isblank", 0, "#FFFF00", None, None)
        assert result["matched_cells"] == 2, result


class TestCreateSheetIsNotBlockedByTheMissingSheetGuard:
    """"급여계산 시트 하나 만들어줄래?" → "'급여계산' 시트를 찾을 수 없습니다"로 되묻혔다.

    2026-08-20 ex24 실측: 1·2번째 턴이 이렇게 죽고, 그 시트를 쓰는 43·44번째 턴이
    연쇄로 무너져 44/49가 됐다. **만들어 달라는 시트가 없는 건 정상**이다.
    """

    def test_sheet_creating_actions_are_exempt(self) -> None:
        from office_claw_sidecar.routers.excel_live import _SHEET_CREATING_ACTIONS

        assert "excel_live.create_sheet" in _SHEET_CREATING_ACTIONS
        # 값을 쓰는 액션은 면제 대상이 아니다 — 없는 시트에 쓰면 되물어야 한다.
        assert "excel_live.write_range" not in _SHEET_CREATING_ACTIONS
        assert "excel_live.set_formula" not in _SHEET_CREATING_ACTIONS
        # rename_sheet의 sheet_name은 **바꿀 대상**이라 이미 있어야 한다. 면제했더니
        # "수도권은 서울권으로 이름 바꿔놔"(값 치환 문장)가 활성 시트 이름을 바꿔
        # 그 시트를 가리키던 수식이 전부 깨졌다(2026-08-20 624 게이트에서 드러남).
        assert "excel_live.rename_sheet" not in _SHEET_CREATING_ACTIONS
        assert "excel_live.consolidate_sheets" not in _SHEET_CREATING_ACTIONS

    def test_the_guard_still_covers_writes(self) -> None:
        """면제가 가드를 통째로 무력화하지 않았는지 — 새 층은 실패만 더할 수 있다."""
        import inspect

        from office_claw_sidecar.routers import excel_live

        src = inspect.getsource(excel_live._plan_approval_gate)
        assert "_SHEET_CREATING_ACTIONS" in src
        assert "_edit_target_problem" in src


class TestCenterAlignIsNotRowSorting:
    """한국어 "정렬"은 줄 세우기와 맞춤 둘 다다 — 늘 줄 세우기로 갔다.

    2026-08-20 ex24 실측: "달력 전체 가운데 정렬해줘" → `sort` 분류 → "어떤 열 기준으로
    정렬할까요?"로 되물었다. 분류기에 맞춤 예외가 **있었는데 둘째 가지에만** 붙어 있어,
    낱말 "정렬"만 보는 첫째 가지가 늘 먼저 걸렸다.
    """

    @pytest.mark.parametrize(
        "message",
        ["달력 전체 가운데 정렬해줘", "머리글 가운데 정렬", "금액 열 오른쪽 정렬해줘", "제목 줄 가운데로 맞춤"],
    )
    def test_alignment_words_go_to_format(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _detect_operation_intent

        assert _detect_operation_intent(message) == "format", message

    @pytest.mark.parametrize(
        "message",
        ["매출 높은 순으로 정렬해줘", "가나다순으로 정렬", "금액 내림차순 정렬", "재고 적은 순서대로 줄세워"],
    )
    def test_real_sorting_still_sorts(self, message: str) -> None:
        """예외가 줄 세우기를 통째로 삼키지 않았는지 — 방향 낱말이 없으면 여전히 sort다."""
        from office_claw_sidecar.routers.excel_live import _detect_operation_intent

        assert _detect_operation_intent(message) == "sort", message

    @pytest.mark.parametrize(
        ("message", "want_align"),
        [("달력 전체 가운데 정렬해줘", "가운데"), ("금액 열 오른쪽 정렬해줘", "오른쪽"), ("가운데 정렬해줘", "가운데")],
    )
    def test_the_quick_rule_builds_a_font_step(self, message: str, want_align: str) -> None:
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan(message, "D3:J8")
        assert plan, message
        assert plan[0]["action"] == "excel_live.set_font", message
        assert plan[0]["params"]["align"] == want_align, message

    def test_whole_table_wording_beats_the_previous_range(self) -> None:
        """"달력 **전체**"는 직전 결과 범위가 아니라 표 전체다."""
        from office_claw_sidecar.routers.excel_live import _build_quick_action_plan

        plan = _build_quick_action_plan("달력 전체 가운데 정렬해줘", "D3:J8")
        assert plan[0]["params"]["target_range"] == "__TABLE_REGION__"

    def test_the_executor_actually_aligns_cells(self, tmp_path) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / "cal.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "달력"
        for row in [["월", "화", "수"], [1, 2, 3], [4, 5, 6]]:
            ws.append(row)
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "달력")
        result = service.set_font(None, "달력", "A1:C3", align="가운데")
        assert result["align"] == "center", result
        assert result["changed_cells"] == 9, result
        ws = load_workbook(path)["달력"]
        assert all(c.alignment.horizontal == "center" for row in ws.iter_rows(max_row=3, max_col=3) for c in row)

        # 뒤이은 글꼴 작업이 맞춤을 지우면 안 된다 — 사용자가 안 부른 것은 그대로 둔다.
        service.set_font(None, "달력", "A1:C1", bold=True)
        ws = load_workbook(path)["달력"]
        assert [c.alignment.horizontal for c in ws[1]] == ["center", "center", "center"]
        assert [c.font.bold for c in ws[1]] == [True, True, True]

    def test_an_unknown_word_leaves_alignment_alone(self, tmp_path) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / "x.xlsx"
        wb = Workbook()
        wb.active["A1"] = "값"
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        sheet = load_workbook(path).sheetnames[0]
        service.select_sheet(None, sheet)
        result = service.set_font(None, sheet, "A1:A1", align="비스듬히")
        assert result["align"] is None, result
        assert load_workbook(path)[sheet]["A1"].alignment.horizontal is None


class TestChartVerificationLooksAtTheSheetItLandedOn:
    """차트는 만들어졌는데 `chart_not_created`로 실패했다.

    2026-08-20 ex26·ex27 실측: 실행은 활성 시트(손익요약)에 그리는데, 검증은 계획의
    `sheet_name`(매출원장 — 플래너가 적어 둔 **원본** 시트)에서 차트를 세어 0개를 봤다.
    """

    def test_the_result_sheet_wins_over_the_plan_sheet(self) -> None:
        import inspect

        from office_claw_sidecar.services import excel_result_verifier

        src = inspect.getsource(excel_result_verifier)
        head = src[src.index('if action in {"excel_live.create_chart"') :][:600]
        assert '(result or {}).get("sheet_name")' in head
        # 계획 값도 여전히 뒤에 남아 있어야 한다 — 결과가 시트를 안 알려주는 엔진 대비.
        assert 'params.get("output_sheet")' in head


class TestFormulaSheetRefsMustExist:
    """없는 시트를 참조하는 수식은 엑셀에서 `#REF!`가 된다 — 쓰기 전에 막는다.

    2026-08-20 624 게이트: "A2 칸에 지역성과 시트 주문건수 전부 합친 숫자 보여줘"가
    `=SUM(지역성!E:E, 주문건수!E:E)`로 계획됐다. 모델이 "지역성과 시트 주문건수"를
    **시트 이름 둘**로 쪼갠 것인데, 그대로 써져서 사용자는 나중에 #REF!를 발견한다.
    """

    NAMES = ["지역성과", "요약", "지연 경고"]

    @pytest.mark.parametrize(
        ("formula", "want", "missing"),
        [
            ("=SUM(지역성과!B2:B6)", "=SUM(지역성과!B2:B6)", []),
            # 앞부분만 맞는 후보가 **하나뿐**이면 고친다.
            ("=SUM(지역성!B:B)", "=SUM(지역성과!B:B)", []),
            # 공백이 든 이름은 따옴표를 유지한다.
            ("=SUM('지연 경고'!A:A)", "=SUM('지연 경고'!A:A)", []),
            # 시트 참조가 없는 수식은 손대지 않는다.
            ("=SUM(B2:B6)", "=SUM(B2:B6)", []),
            # 열 이름을 시트로 오인한 것은 **못 찾음**으로 남긴다.
            ("=SUM(지역성!E:E, 주문건수!E:E)", "=SUM(지역성과!E:E, 주문건수!E:E)", ["주문건수"]),
        ],
    )
    def test_refs_are_resolved_or_reported(self, formula: str, want: str, missing: list[str]) -> None:
        from office_claw_sidecar.services.excel_live_file_service import _resolve_formula_sheet_refs

        fixed, found_missing = _resolve_formula_sheet_refs(formula, self.NAMES)
        assert fixed == want
        assert found_missing == missing

    def test_the_executor_refuses_to_write_a_broken_formula(self, tmp_path) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
        from office_claw_sidecar.services.excel_live_service import ExcelLiveError

        path = tmp_path / "f.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "지역성과"
        for row in [["지역", "주문건수"], ["수도권", 10452], ["충청권", 3892]]:
            ws.append(row)
        wb.create_sheet("요약")
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "요약")

        with pytest.raises(ExcelLiveError) as excinfo:
            service.set_formula(None, "요약", "A2", "=SUM(지역성!E:E, 주문건수!E:E)")
        assert "주문건수" in str(excinfo.value)
        assert load_workbook(path)["요약"]["A2"].value is None, "막았는데도 써졌다"

        # 앞부분만 맞는 것은 고쳐서 쓴다.
        service.set_formula(None, "요약", "A2", "=SUM(지역성!B:B)")
        assert load_workbook(path)["요약"]["A2"].value == "=SUM(지역성과!B:B)"


class TestCrossSheetVocabularyAndYield:
    """크로스시트 규칙이 못 잡으면 모델이 엉뚱한 수식을 쓴다 — 어휘 구멍 셋과 양보 하나.

    2026-08-20 624 게이트 `cross_sheet_sum` 24문장 중 4문형이 규칙 밖으로 샜고,
    그 사이 모델이 `=SUM(지역성!E:E, 주문건수!E:E)`(없는 시트)와
    `=SUM(B:B)+SUM(E:E)`(빈 열 → 0)를 썼다.
    """

    @pytest.mark.parametrize(
        ("message", "want_agg"),
        [
            # `합`이 `합친 값`보다 앞에 있어 "합친"이 `합`으로 잘렸다 — 긴 것부터 적어야 한다.
            ("A2 칸에 지역성과 시트 주문건수 전부 합친 숫자 보여줘", "합친 숫자"),
            ("A2에 지역성과 시트 주문건수 합계 가져와줘", "합계"),
            # 영어로 부르는 사람도 있다.
            ("A2 셀에 지역성과 sheet 주문건수 total 넣어 주세요.", "total"),
        ],
    )
    def test_the_aggregate_word_is_matched_whole(self, message: str, want_agg: str) -> None:
        from office_claw_sidecar.services.excel_aggregate_below import _CROSS_SHEET

        match = _CROSS_SHEET.search(message)
        assert match is not None, message
        assert match.group(5) == want_agg, match.groups()

    def test_a_pronoun_after_a_comma_still_finds_the_cell(self) -> None:
        """"…합계, **그걸** A2에 넣어 주세요" — 대명사가 앞 절의 집계를 그대로 받는다."""
        from office_claw_sidecar.services.excel_aggregate_below import _CROSS_SHEET_CELL_LAST

        match = _CROSS_SHEET_CELL_LAST.search("지역성과 시트 주문건수 합계, 그걸 A2에 넣어 주세요")
        assert match is not None
        assert match.group(4) == "A2", match.groups()

    def test_show_me_is_a_cross_sheet_verb(self) -> None:
        from office_claw_sidecar.services.excel_aggregate_below import _CROSS_SHEET_VERB

        assert _CROSS_SHEET_VERB.search("합친 숫자 보여줘")

    def test_a_read_plan_yields_to_the_cross_sheet_builder(self) -> None:
        """'보여줘'가 read_range로 잡혀 크로스시트 빌더가 아예 안 돌았다.

        양보 목록은 **훅 이름**으로 돼 있어 늘 새 훅보다 좁아진다. 계획이 읽기뿐이면
        훅 이름과 무관하게 양보하게 했다.
        """
        import inspect

        from office_claw_sidecar.routers import excel_live

        src = inspect.getsource(excel_live._run_command)
        assert "_cross_yieldable = True" in src
        assert 'str((step or {}).get("action") or "") == "excel_live.read_range"' in src

    def test_the_builder_makes_a_sheet_qualified_sum(self) -> None:
        from office_claw_sidecar.services.excel_aggregate_below import build_cross_sheet_aggregate_plan

        rows = [["지역", "주문건수"], ["수도권", 10452], ["충청권", 3892]]

        def reader(sheet: str):
            return ("A1:B3", rows) if sheet == "지역성과" else ("A1:A1", [["전체주문건수"]])

        for message in [
            "A2 칸에 지역성과 시트 주문건수 전부 합친 숫자 보여줘, 나중에도 연동되게",
            "지역성과 시트 주문건수 합계, 그걸 A2에 넣어 주세요",
            "A2 셀에 지역성과 sheet 주문건수 total 넣어 주세요.",
        ]:
            plan = build_cross_sheet_aggregate_plan(message, reader, ["지역성과", "요약"])
            assert plan, message
            formula = plan[0]["params"]["formula_a1"]
            assert "지역성과" in formula and "SUM" in formula, (message, formula)
            assert plan[0]["params"]["range_ref"] == "A2", (message, plan[0]["params"])


class TestExcludeAlsoHidesUnlessToldToDelete:
    """"안 보이게 해줘"인데 행이 지워졌다 — 숨기기 기본이 `remove`를 못 덮었다.

    2026-08-20 파괴 게이트: `대기 아닌 건 잠깐 안 보이게 해줘` → 행이 지워짐.
    제외 바인더가 "아닌"을 보고 `remove`를 세우고, 숨김 바인더는 `remove`면 빠졌다.
    어느 쪽을 뺄지는 유지하고 **되돌릴 수 없는 삭제만** 없앤다.
    """

    @pytest.mark.parametrize(
        ("message", "value", "want"),
        [
            ("대기 아닌 건 잠깐 안 보이게 해줘", "대기", "hide_exclude"),
            ("취소된 주문은 빼줘", "취소", "hide_exclude"),
            ("상태가 대기인 것만 보여줘", "대기", "hide"),
            # 지우라고 말했을 때만 지운다.
            ("취소된 주문은 지워줘", "취소", "remove"),
            ("대기 아닌 행은 삭제해줘", "대기", "remove"),
        ],
    )
    def test_the_mode_follows_the_words(self, message: str, value: str, want: str) -> None:
        from office_claw_sidecar.services.excel_param_binder import (
            _bind_filter_delete_or_hide,
            _bind_filter_mode,
        )

        params = {"value": value}
        _bind_filter_mode(params, message=message)
        _bind_filter_delete_or_hide(params, message=message)
        assert params.get("mode") == want, (message, params)

    @pytest.mark.parametrize(
        ("mode", "kept", "hidden", "removed"),
        [
            ("hide", 4, [3, 5], 0),
            ("hide_exclude", 4, [2, 4], 0),
            ("keep", 2, [], 2),
            ("remove", 2, [], 2),
        ],
    )
    def test_each_mode_does_exactly_one_thing(self, tmp_path, mode, kept, hidden, removed) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / f"{mode}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "지연경고"
        for row in [
            ["운송장", "구간", "지연시간", "상태"],
            ["INV-001", "김포-부산", "12시간", "대기"],
            ["INV-002", "이천-대전", "3시간", "처리중"],
            ["INV-003", "용인-광주", "8시간", "대기"],
            ["INV-004", "평택-제주", "1시간", "완료"],
        ]:
            ws.append(row)
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "지연경고")
        result = service.filter_rows(
            None, "지연경고", column="상태", operator="==", value="대기", mode=mode
        )
        assert result["removed_rows"] == removed, result
        sheet = load_workbook(path)["지연경고"]
        rows = [c for c in (sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)) if c]
        assert len(rows) == kept, rows
        assert [r for r in range(2, sheet.max_row + 1) if sheet.row_dimensions[r].hidden] == hidden


class TestTitleMergeYieldsToAnyWrittenRange:
    """"A1**부터** G1**까지**"를 범위로 못 알아봐 제목 줄 규칙이 가로챘다.

    2026-08-22 42각본 전수(2492/2495)에서 실패한 3건이 전부 이것이었다:
    ex14·ex14_v2 20번 턴, ex20 26번 턴. 사용자가 범위를 직접 말하고 승인 카드까지
    승인했는데, 규칙이 `A1:A1`(한 칸)로 바꿔 놓아 병합이 안 됐다.
    콜론만 보면 안 된다 — 사람은 '부터/까지'를 더 자주 쓴다.
    """

    EMPTY_TITLE_SHEET = {
        "active_sheet": "대시보드",
        "sheets": [
            {
                "name": "대시보드",
                "used_range": "A1:A1",
                "columns": [{"letter": "A", "header": "AI 기반 콘텐츠·캠페인 캘린더"}],
                "sample_rows": [],
            }
        ],
    }
    TITLED_TABLE = {
        "active_sheet": "지역성과",
        "sheets": [
            {
                "name": "지역성과",
                "used_range": "A1:F7",
                "columns": [
                    {"letter": "A", "header": "2026년 상반기 지역 실적"},
                    {"letter": "B", "header": ""},
                ],
                "sample_rows": [["지역", "주문건수", "출고건수", "정시배송률", "지연건수", "클레임"]],
            }
        ],
    }

    @pytest.mark.parametrize(
        "message",
        [
            "제목 줄은 A1부터 G1까지 병합해줘",
            "A1에서 G1까지 merge 해줘, 제목 줄이니까",
            "제목 줄은 A1:G1 병합해줘",
            "제목 줄은 A1~G1 병합",
        ],
    )
    def test_it_yields_when_the_person_named_a_range(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _title_row_merge_plan

        assert _title_row_merge_plan(message, self.EMPTY_TITLE_SHEET) == [], message

    @pytest.mark.parametrize("message", ["제목 줄 병합해줘", "맨 위 제목 칸 하나로 합쳐줘"])
    def test_it_still_fires_without_a_range(self, message: str) -> None:
        from office_claw_sidecar.routers.excel_live import _title_row_merge_plan

        plan = _title_row_merge_plan(message, self.TITLED_TABLE)
        assert plan, message
        assert plan[0]["params"]["target_range"] == "A1:F1", message


class TestAWordInAnotherSheetsHeadersDoesNotStealTheTarget:
    """"제목 줄"의 '제목'이 **다른 시트의 열 머리글**이라 대상이 그리로 옮겨졌다.

    2026-08-22 42각본 전수(ex14·ex14_v2·ex20)의 마지막 실패:
      18턴 `대시보드 시트 만들어조` → 19턴 `대시보드 시트 A1에 … 입력`
      20턴 `제목 줄은 A1부터 G1까지 병합해줘` → **콘텐츠 일정** 시트 A1:G1 계획

    콘텐츠 일정의 열은 날짜/요일/유형/**제목**/마감/채널/담당이다. 머리글 낱말 하나가
    겹쳤다고 시트를 옮기면, 사람이 좌표로 지목한 칸이 아닌 데를 고친다.
    병합 가드가 막아 줘서 드러났지, 가드 전에는 머리글 7개를 지우고 '성공'으로 집계됐다.
    """

    HEADERS = ["날짜", "요일", "유형", "제목", "마감", "채널", "담당"]

    def _digest(self):
        return {
            "active_sheet": "대시보드",
            "sheets": [
                {
                    "name": "콘텐츠 일정",
                    "used_range": "A1:O26",
                    "columns": [
                        {"letter": chr(65 + i), "header": h} for i, h in enumerate(self.HEADERS)
                    ],
                },
                {
                    "name": "대시보드",
                    "used_range": "A1:A1",
                    "columns": [{"letter": "A", "header": "AI 기반 콘텐츠·캠페인 캘린더"}],
                },
            ],
        }

    def test_a_structural_word_is_not_a_column_reference(self) -> None:
        """"제목 **줄**"은 제목이라는 열이 아니라 표의 제목 행이다."""
        from office_claw_sidecar.services.excel_param_binder import _structural_free_mentions

        assert _structural_free_mentions("제목 줄은 A1부터 G1까지 병합해줘", self.HEADERS) == []
        # "제목 **열**"이면 진짜 열 지목이다 — 그건 남아야 한다.
        assert [h["header"] for h in _structural_free_mentions("제목 열만 굵게 해줘", self.HEADERS)] == ["제목"]

    @pytest.mark.parametrize(
        ("message", "want_sheet"),
        [
            ("제목 줄은 A1부터 G1까지 병합해줘", "대시보드"),
            ("A1에서 G1까지 merge 해줘, 제목 줄이니까", "대시보드"),
            # 진짜로 그 시트의 열을 부른 문장은 여전히 옮겨 간다.
            ("제목 열만 굵게 해줘", "콘텐츠 일정"),
            ("마감 지난 행 빨갛게", "콘텐츠 일정"),
        ],
    )
    def test_retargeting_respects_coordinates_and_structure(self, message: str, want_sheet: str) -> None:
        from office_claw_sidecar.services.excel_param_binder import _retarget_sheet_by_headers

        digest = self._digest()
        active = digest["sheets"][1]
        entry, _prefix = _retarget_sheet_by_headers(message, active, digest, "")
        assert (entry or {}).get("name") == want_sheet, message

    def test_an_unnamed_sheet_edit_stays_on_the_active_sheet(self) -> None:
        """계획이 남의 시트를 적어 와도, 원문이 시트를 안 불렀으면 활성 시트로 되돌린다."""
        from office_claw_sidecar.services.excel_param_binder import _bind_sheet_stays_active

        params = {"sheet_name": "콘텐츠 일정", "target_range": "A1:G1"}
        changed = _bind_sheet_stays_active(
            params,
            action="excel_live.merge_cells",
            message="제목 줄은 A1부터 G1까지 병합해줘",
            digest=self._digest(),
            active=None,  # 요청에 시트가 안 실려 오는 경로 — 다이제스트가 사실이다
        )
        assert changed == ["sheet_name=대시보드"]
        assert params["sheet_name"] == "대시보드"

    @pytest.mark.parametrize(
        ("action", "sheet"),
        [("excel_live.create_sheet", "요약"), ("excel_live.rename_sheet", "콘텐츠 일정")],
    )
    def test_sheet_naming_actions_are_left_alone(self, action: str, sheet: str) -> None:
        """`sheet_name`이 **대상**이 아니라 **이름**인 액션은 건드리면 안 된다."""
        from office_claw_sidecar.services.excel_param_binder import _bind_sheet_stays_active

        params = {"sheet_name": sheet}
        assert _bind_sheet_stays_active(
            params, action=action, message="시트 하나 만들어줘", digest=self._digest(), active="대시보드"
        ) == []
        assert params["sheet_name"] == sheet


class TestTheTaskListHasOneSource:
    """작업 종류 목록이 프롬프트·집합·스키마 셋으로 흩어져 있었다.

    2026-08-23 확인: `scripts/measure_intent_normalizer.py`에 **네 번째 복제본**이
    있었고 거기엔 `highlight`가 통째로 빠져 있었다(16종). 그래서 로드맵의 근거였던
    "정규화 100%/96%"는 **배포되는 프롬프트가 아닌 것**의 점수였다.
    같은 어휘를 두 곳에 두면 반드시 갈라진다 — 목록의 원본은 프롬프트 하나뿐이다.
    """

    # 2026-08-24 라운드 2 배치 1a: 밀려남이 실측된 4종(create_sheet·delete_charts·
    # freeze·autofit)을 넣어 17 → 21종. 종류 상한 경고(13~14 + 신규 배치)에 따라
    # 배치마다 44문장 재측정으로 분류 품질을 확인한다.
    # 2026-08-25 배치 1b: 명시 어휘 5종(merge·unmerge·data_bar·color_scale·
    # rename_sheet) → 26종. 배치 2: 열 연산 4종 + group_by(읽기 전용 조회) → 31종.
    # 확장마다 44문장 재측정으로 밀려남 0을 확인한다.
    EXPECTED = (
        "fill_color", "font", "highlight", "number_format", "formula", "sort", "filter",
        "dedupe", "clear_values", "reset_all", "create_table", "pivot", "chart",
        "write_value", "find_replace", "read",
        "create_sheet", "delete_charts", "freeze", "autofit",
        "merge", "unmerge", "data_bar", "color_scale", "rename_sheet",
        "delete_sheet", "drop_column", "add_column", "rename_column", "group_by",
        "other",
    )

    def test_the_names_come_from_the_prompt(self) -> None:
        from office_claw_sidecar.services.excel_intent_normalizer import TASK_NAMES

        assert TASK_NAMES == self.EXPECTED, "프롬프트의 task 목록이 바뀌었다"

    def test_the_schema_enum_follows(self) -> None:
        from office_claw_sidecar.services.excel_intent_normalizer import (
            INTENT_JSON_SCHEMA,
            TASK_NAMES,
        )

        assert INTENT_JSON_SCHEMA["properties"]["task"]["enum"] == sorted(TASK_NAMES)

    def test_a_dropped_task_is_caught(self) -> None:
        """프롬프트에서 종류가 빠지면 파생 집합도 빠진다 — 조용히 넘어가지 않는다."""
        from office_claw_sidecar.services.excel_intent_normalizer import (
            _PROMPT,
            _tasks_declared_in_prompt,
        )

        # 이 항목은 줄 끝이라 뒤가 공백이 아니라 개행이다.
        crippled = _PROMPT.replace("highlight(조건에 맞는 셀만 강조),", "")
        assert "highlight" not in _tasks_declared_in_prompt(crippled)
        # 모양이 통째로 바뀌면 빈 집합이 아니라 예외다.
        with pytest.raises(RuntimeError):
            _tasks_declared_in_prompt("task 목록 없음")

    def test_the_measure_script_does_not_keep_its_own_copy(self) -> None:
        """계측이 프로덕션과 다른 프롬프트를 재면 그 숫자는 아무 뜻이 없다."""
        import pathlib

        src = pathlib.Path(__file__).resolve().parent.parent / "scripts/measure_intent_normalizer.py"
        text = src.read_text(encoding="utf-8")
        assert "excel_intent_normalizer import" in text
        assert 'PROMPT = """' not in text, "프롬프트 복제본이 되살아났다"


class TestIntentToPlanCoverageIsMeasurable:
    """`intent_to_plan`을 호출해 본 스크립트가 **한 건도 없었다**(2026-08-23 grep 0건).

    그래서 "통역이 뜻은 맞게 뽑았는데 계획으로 못 옮긴 비율"을 잰 숫자가 없었고,
    받아 적기를 넓히는 작업의 전후를 비교할 근거가 없었다. 하네스는 LLM을 안 부른다.
    """

    def _rows(self):
        import importlib.util
        import pathlib

        path = pathlib.Path(__file__).resolve().parent.parent / "scripts/measure_intent_coverage.py"
        spec = importlib.util.spec_from_file_location("_cov", path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        return module.run()

    def test_the_harness_runs_without_an_llm(self) -> None:
        rows = self._rows()
        assert len(rows) >= 20
        assert not [r for r in rows if r["error"]], [r for r in rows if r["error"]]

    def test_the_known_gaps_are_still_gaps(self) -> None:
        """지금 떨어지는 모양들. **고치면 이 핀이 실패한다** — 그때 핀을 옮긴다.

        2026-08-23에 `write_value` 둘(범위·열 전체)을 고쳐 여기서 초록으로 옮겼다.
        """
        rows = {(r["task"], r["note"]): r["mapped"] for r in self._rows()}
        assert rows[("highlight", "조건부 강조")] is False
        # 방향 없는 정렬은 **물러나는 게 옳다** — 짐작하면 행 순서를 통째로 뒤집는다.
        assert rows[("sort", "방향 없음(물러나는 게 옳음)")] is False
        # 반대로 이미 되는 것들이 조용히 죽지 않았는지도 본다.
        assert rows[("fill_color", "범위+색")] is True
        assert rows[("write_value", "한 칸 쓰기")] is True
        assert rows[("write_value", "범위 브로드캐스트")] is True
        assert rows[("write_value", "열 전체")] is True
        assert rows[("sort", "내림차순")] is True
        # 2026-08-23에 font 셋(굵게·크기·맞춤)을 고쳐 초록으로 옮겼다.
        assert rows[("font", "굵게 — 색 아니면 매핑 실패")] is True
        assert rows[("font", "크기")] is True
        assert rows[("font", "가로 맞춤")] is True
        assert rows[("clear_values", "열 비우기")] is True


class TestOneValueFillsTheWholeRange:
    """"A2:A9에 0 입력해줘"가 **A2 한 칸만** 쓰고 성공이라고 답했다.

    2026-08-23 실측: `_shape_write_values("0", 9, 1)` → `[[0]]`. 나머지 여덟 칸은
    그대로인데 검증기는 요청된 values_2d(=한 칸)를 기준으로 대조하므로 미달을 못 본다.
    조용한 부분 실행이다 — 사용자는 다 채워진 줄 안다.
    """

    @pytest.mark.parametrize(
        ("raw", "rows", "cols", "want"),
        [
            ("0", 9, 1, [[0]] * 9),
            ("0", 1, 4, [[0, 0, 0, 0]]),
            ("미정", 2, 2, [["미정", "미정"], ["미정", "미정"]]),
            # "1,000"은 값 둘이 아니라 천 단위 구분자다 — 범위에서도 마찬가지다.
            ("1,000", 3, 1, [[1000], [1000], [1000]]),
        ],
    )
    def test_a_single_value_is_broadcast(self, raw, rows, cols, want) -> None:
        from office_claw_sidecar.services.excel_param_binder import _shape_write_values

        assert _shape_write_values(raw, rows, cols) == want

    @pytest.mark.parametrize(
        ("raw", "rows", "cols", "want"),
        [
            # 값이 둘 이상이면 사람이 구체적으로 나열한 것이다 — 채우지 않는다.
            ("서울,부산,대구", 9, 1, [["서울"], ["부산"], ["대구"]]),
            ("가,나,다", 1, 3, [["가", "나", "다"]]),
            # 단일 셀은 예전 그대로 — 천 단위 구분자를 쪼개면 안 된다.
            ("1,000", 1, 1, [[1000]]),
            ("120", 1, 1, [[120]]),
        ],
    )
    def test_the_old_shapes_are_unchanged(self, raw, rows, cols, want) -> None:
        from office_claw_sidecar.services.excel_param_binder import _shape_write_values

        assert _shape_write_values(raw, rows, cols) == want

    @pytest.mark.parametrize(
        "raw",
        ["가장 큰 매출 값", "서울 지역 매출만 더한", "총 임대료 합계"],
    )
    def test_an_instruction_phrase_is_never_broadcast(self, raw: str) -> None:
        """자체 검토에서 잡은 내 실수 — 값 하나면 무조건 퍼뜨리게 했더니
        "F7:G9에 **가장 큰 매출 값** 넣어줘"의 지시문이 여섯 칸에 통째로 써졌다.
        고치려던 것(한 칸만 써짐)보다 나쁜 결과다. 띄어 쓴 것은 값이 아니라 말이다."""
        from office_claw_sidecar.services.excel_param_binder import _shape_write_values

        assert _shape_write_values(raw, 3, 2) == [[raw]]

    def test_it_backs_off_when_the_range_is_huge(self) -> None:
        """한 낱말로 수천 칸을 덮는 건 사고다. 넓으면 채우지 않고 물러난다."""
        from office_claw_sidecar.services.excel_param_binder import (
            _BROADCAST_CELL_LIMIT,
            _shape_write_values,
        )

        assert _shape_write_values("0", _BROADCAST_CELL_LIMIT + 1, 1) == []
        assert len(_shape_write_values("0", _BROADCAST_CELL_LIMIT, 1)) == _BROADCAST_CELL_LIMIT

    def test_the_binder_fills_every_cell(self) -> None:
        from office_claw_sidecar.services.excel_param_binder import _bind_write_values

        params: dict = {}
        _bind_write_values(params, message="A2:A9에 0 입력해줘")
        assert params["start_cell"] == "A2"
        assert params["values_2d"] == [[0]] * 8, params["values_2d"]

    def test_the_executor_writes_every_cell(self, tmp_path) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / "b.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "매출"
        ws.append(["날짜", "금액"])
        for n in range(8):
            ws.append([f"2026-01-{n + 1:02d}", 100 + n])
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "매출")
        service.write_range(
            workbook_id=None, sheet_name="매출", start_cell="C2", values_2d=[["미정"]] * 8
        )
        sheet = load_workbook(path)["매출"]
        assert [sheet.cell(row=r, column=3).value for r in range(2, 10)] == ["미정"] * 8
        # 옆 열은 건드리지 않는다.
        assert [sheet.cell(row=r, column=2).value for r in range(2, 10)] == list(range(100, 108))


class TestIntentWriteValueHandlesRangesAndColumns:
    """통역 AI가 "A2:A9에 0 넣어줘"를 맞게 분류해도 **받아 적을 코드가 없었다.**

    2026-08-23: `intent_to_plan`의 write_value 분기가 `_SINGLE_CELL`만 받아, 범위와
    열 이름은 통째로 플래너로 넘어갔다. 진단서가 "실사용에서 가장 많이 걸린 지점"이라
    적은 자리다. AI가 못 알아들은 게 아니라 옮겨 적을 코드가 없었던 것이다.
    """

    DIGEST = {
        "active_sheet": "매출",
        "sheets": [
            {
                "name": "매출",
                "used_range": "A1:F9",
                "columns": [
                    {"letter": "A", "header": "날짜"},
                    {"letter": "B", "header": "지역"},
                    {"letter": "D", "header": "금액"},
                    {"letter": "F", "header": "비고"},
                ],
                "sample_rows": [],
            }
        ],
    }

    def _plan(self, digest=None, **intent):
        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan

        base = {"task": "write_value", "range": None, "column": None, "option": None}
        return intent_to_plan({**base, **intent}, digest=digest or self.DIGEST, message="")

    @pytest.mark.parametrize(
        ("intent", "start", "cells"),
        [
            ({"range": "A12", "option": "합계"}, "A12", 1),
            ({"range": "A2:A9", "option": "미정"}, "A2", 8),
            ({"range": "A1:C1", "option": "가,나,다"}, "A1", 3),
            # 열 이름만 부른 경우 — 머리글 아래부터 데이터 끝까지.
            ({"column": "비고", "option": "미정"}, "F2", 8),
        ],
    )
    def test_it_maps(self, intent, start, cells) -> None:
        plan = self._plan(**intent)
        assert plan is not None, intent
        assert plan["action"] == "excel_live.write_range"
        assert plan["params"]["start_cell"] == start
        assert sum(len(row) for row in plan["params"]["values_2d"]) == cells

    def test_an_instruction_phrase_stays_one_cell(self) -> None:
        """모양 맞추기를 바인더와 공유하므로 지시문 제외도 그대로 따라온다."""
        plan = self._plan(range="F7:G9", option="가장 큰 매출 값")
        assert sum(len(row) for row in plan["params"]["values_2d"]) == 1

    @pytest.mark.parametrize(
        "intent",
        [
            {"column": "없는열", "option": "미정"},
            # 스키마 어휘를 값으로 되뇌는 축퇴 — 예전부터 막던 것이 그대로 살아 있나.
            {"range": "A2:A9", "option": "write_value"},
        ],
    )
    def test_it_backs_off(self, intent) -> None:
        assert self._plan(**intent) is None, intent

    def test_it_backs_off_when_the_data_end_is_unknown(self) -> None:
        """`_last_row`는 사용 범위를 못 읽으면 2를 준다. 그 상태로 채우면 "전부"가 한 칸이 된다."""
        digest = {
            "active_sheet": "빈",
            "sheets": [
                {"name": "빈", "used_range": "A1:B1", "columns": [{"letter": "B", "header": "비고"}]}
            ],
        }
        assert self._plan(digest=digest, column="비고", option="미정") is None

    def test_the_shaping_is_not_duplicated(self) -> None:
        """모양 맞추기를 두 벌 두면 반드시 갈라진다 — 바인더 것을 빌려 쓴다."""
        from office_claw_sidecar.services import excel_intent_normalizer as normalizer
        from office_claw_sidecar.services import excel_param_binder as binder

        assert normalizer._shape_write_values is binder._shape_write_values
        assert normalizer._range_shape is binder._range_shape


class TestFontTakesMoreThanColor:
    """`font`가 색만 받아, "머리글 굵게"가 통역에서 플래너로 떨어졌다.

    2026-08-23 확인: 도구(`set_font(bold, name, size, color, align)`)와 검증기
    (`excel_live_plan_validator.py:749`)는 이미 다섯을 다 받는데,
    `intent_to_plan`의 font 분기만 `_COLORS.get()`이 맞을 때만 계획을 냈다.
    낱말 사전이 없었을 뿐이다.
    """

    @pytest.mark.parametrize(
        ("option", "want"),
        [
            ("흰색", {"color": "#FFFFFF"}),
            ("굵게", {"bold": True}),
            ("진하게", {"bold": True}),
            ("bold", {"bold": True}),
            # 해제를 굵게로 읽으면 정반대 편집이다.
            ("굵게 해제", {"bold": False}),
            ("안 굵게", {"bold": False}),
            ("가운데", {"align": "center"}),
            ("가운데 정렬", {"align": "center"}),
            ("왼쪽 정렬", {"align": "left"}),
            ("14", {"size": 14.0}),
            ("14pt", {"size": 14.0}),
            ("크기 16", {"size": 16.0}),
        ],
    )
    def test_it_reads_the_word(self, option: str, want: dict) -> None:
        from office_claw_sidecar.services.excel_intent_normalizer import _font_params_from

        assert _font_params_from(option) == want

    @pytest.mark.parametrize("option", ["기울임", "밑줄", "이탤릭", "", "예쁘게", "3", "200", "형광펜"])
    def test_it_backs_off_on_what_the_tool_cannot_do(self, option: str) -> None:
        """**기울임·밑줄은 일부러 매핑하지 않는다** — 도구에 파라미터가 없다.

        매핑하면 조용히 무시되고 "했다"고 보고된다. 가짜 성공이 실패보다 나쁘다.
        크기도 엑셀 범위(6~72) 밖이면 사람 뜻이 아니다.
        """
        from office_claw_sidecar.services.excel_intent_normalizer import _font_params_from

        assert _font_params_from(option) == {}

    def test_the_align_table_is_shared(self) -> None:
        """맞춤 낱말표를 두 벌 두면 갈라진다 — 실행기 것을 빌려 쓴다."""
        from office_claw_sidecar.services import excel_intent_normalizer as normalizer
        from office_claw_sidecar.services import excel_live_service as service

        assert normalizer._ALIGN_WORDS is service._ALIGN_WORDS

    def test_the_executor_applies_all_three(self, tmp_path) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan
        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        digest = {
            "active_sheet": "매출",
            "sheets": [{"name": "매출", "used_range": "A1:C3", "columns": [{"letter": "A", "header": "날짜"}]}],
        }
        path = tmp_path / "f.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "매출"
        ws.append(["날짜", "금액", "비고"])
        ws.append(["1/1", 100, ""])
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "매출")
        for option in ("굵게", "16", "가운데"):
            plan = intent_to_plan(
                {"task": "font", "range": "A1:C1", "column": None, "option": option},
                digest=digest,
                message="",
            )
            assert plan is not None, option
            params = dict(plan["params"])
            target = params.pop("target_range")
            service.set_font(None, "매출", target, **params)
        cell = load_workbook(path)["매출"]["A1"]
        assert (cell.font.bold, cell.font.size, cell.alignment.horizontal) == (True, 16.0, "center")
        # 데이터 행은 건드리지 않는다.
        assert load_workbook(path)["매출"]["A2"].font.bold is False


class TestRetriesSkipTheIntentNormalizer:
    """재시도가 통째로 헛돌고 있었다 — 같은 계획이 그대로 돌아온다.

    2026-08-23 확인: `excel_live_agent.py:1439`가 `skip_intent_normalizer`를 읽도록
    만들어져 있는데 **저장소 전체에서 세우는 곳이 0곳**이었다(읽기 1, 쓰기 0).
    정규화 프롬프트는 {머리글, 원문}만 보고 temperature=0이며 매퍼는 결정적이라,
    검증에서 반려된 계획을 재시도해도 **글자까지 같은 계획**이 나온다.
    실패를 아는 것은 플래너 경로뿐이다(`render_execution_failure`).
    """

    async def _tiers(self, *, allow_strong: bool):
        from office_claw_sidecar.services.excel_planner_escalation import plan_with_escalation

        seen: list[dict] = []

        async def parse(message, context):  # `_attempt`는 위치 인자로 부른다
            seen.append(dict(context or {}))
            return {"action_plan": [{"action": "excel_live.set_font", "params": {}}]}

        def validate(steps):
            return (False, "일부러 실패")

        await plan_with_escalation(
            "금액 열 굵게",
            parse=parse,
            validate=validate,
            context={"workbook_digest": {}},
            allow_strong=allow_strong,
        )
        return seen

    def test_the_first_try_still_uses_it(self) -> None:
        """1티어는 정규화가 본령이다 — 여기까지 막으면 96%짜리 경로를 버리는 것이다."""
        import asyncio

        seen = asyncio.run(self._tiers(allow_strong=False))
        assert seen[0].get("skip_intent_normalizer") is None

    def test_retries_skip_it(self) -> None:
        import asyncio

        seen = asyncio.run(self._tiers(allow_strong=True))
        assert len(seen) == 3, [s.get("planner_provider") for s in seen]
        assert seen[1]["skip_intent_normalizer"] is True
        assert seen[1]["reflection_note"]
        # 강한 모델을 부르기로 해 놓고 정규화가 가로채면 그 지정이 소비되지도 않는다.
        assert seen[2]["skip_intent_normalizer"] is True
        assert seen[2]["planner_provider"] == "strong"

    def test_the_agent_honours_the_flag(self) -> None:
        """플래그를 읽는 쪽이 사라지면 위 배선이 조용히 무의미해진다."""
        import inspect

        from office_claw_sidecar.services import excel_live_agent

        src = inspect.getsource(excel_live_agent)
        assert 'context.get("skip_intent_normalizer")' in src


class TestIntentFirstIsAnExperimentSwitch:
    """로드맵 2단계("AI가 먼저")를 **재기 위한** 스위치. 기본은 꺼짐.

    624 게이트 96.3% 중 규칙 경로가 492건인데 통역의 사용자체 정확도는 79%다
    (2026-08-23 실측). 뒤집으면 그 492건이 79%짜리 판단을 먼저 거친다 —
    좋아질지 나빠질지는 재야 안다. 켜고 끄고 각각 돌려 비교하려고 뺐다.
    """

    @pytest.mark.parametrize("value", ["1", "true", "on", "YES"])
    def test_it_turns_on(self, value: str, monkeypatch) -> None:
        from office_claw_sidecar.routers.excel_live import _intent_first_enabled

        monkeypatch.setenv("OFFICECLAW_INTENT_FIRST", value)
        assert _intent_first_enabled() is True

    @pytest.mark.parametrize("value", ["", "0", "false", "no"])
    def test_it_stays_off(self, value: str, monkeypatch) -> None:
        from office_claw_sidecar.routers.excel_live import _intent_first_enabled

        monkeypatch.setenv("OFFICECLAW_INTENT_FIRST", value)
        assert _intent_first_enabled() is False

    def test_unset_means_off(self, monkeypatch) -> None:
        """실험 스위치의 기본은 **꺼짐**이다 — 켜지 않으면 제품 동작이 안 바뀐다."""
        from office_claw_sidecar.routers.excel_live import _intent_first_enabled

        monkeypatch.delenv("OFFICECLAW_INTENT_FIRST", raising=False)
        assert _intent_first_enabled() is False

    def test_it_is_read_every_time(self, monkeypatch) -> None:
        """캐시하면 한 프로세스 안에서 A/B를 못 돌린다."""
        from office_claw_sidecar.routers.excel_live import _intent_first_enabled

        monkeypatch.setenv("OFFICECLAW_INTENT_FIRST", "1")
        assert _intent_first_enabled() is True
        monkeypatch.setenv("OFFICECLAW_INTENT_FIRST", "0")
        assert _intent_first_enabled() is False


class TestClearValuesTakesAColumnName:
    """"비고 열 비워줘"가 범위가 없다는 이유로 통째로 플래너로 넘어갔다.

    2026-08-24: `clear_values` 분기가 `if rng`만 봤다. 열 이름은 모호하지 않다 —
    **머리글 아래부터** 그 열의 데이터 끝까지다.
    """

    DIGEST = {
        "active_sheet": "매출",
        "sheets": [
            {
                "name": "매출",
                "used_range": "A1:C5",
                "columns": [
                    {"letter": "A", "header": "날짜"},
                    {"letter": "B", "header": "금액"},
                    {"letter": "C", "header": "비고"},
                ],
            }
        ],
    }

    def _plan(self, digest=None, **intent):
        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan

        base = {"task": "clear_values", "range": None, "column": None, "option": None}
        # 2026-08-25부터 열 지우기는 열 이름이 문장에 있어야 믿는다(게이트 회귀의 교훈).
        column = str(intent.get("column") or "")
        message = f"{column} 열 비워줘" if column else "비워줘"
        return intent_to_plan({**base, **intent}, digest=digest or self.DIGEST, message=message)

    def test_a_column_name_clears_the_data_rows(self) -> None:
        plan = self._plan(column="비고")
        assert plan["action"] == "excel_live.clear_range"
        # 2행부터다 — 머리글까지 지우면 표가 뭉개진다.
        assert plan["params"]["target_range"] == "C2:C5"

    def test_an_explicit_range_is_unchanged(self) -> None:
        assert self._plan(range="C2:C5")["params"]["target_range"] == "C2:C5"

    @pytest.mark.parametrize(
        ("intent", "digest"),
        [
            ({"column": "없는열"}, None),
            # 데이터 끝을 모르면(`_last_row` 폴백 2) 한 칸만 지우고 "비웠다"고 답하게 된다.
            (
                {"column": "비고"},
                {
                    "active_sheet": "빈",
                    "sheets": [
                        {"name": "빈", "used_range": "A1:C1", "columns": [{"letter": "C", "header": "비고"}]}
                    ],
                },
            ),
        ],
    )
    def test_it_backs_off(self, intent, digest) -> None:
        assert self._plan(digest=digest, **intent) is None

    def test_the_executor_keeps_the_header_and_neighbours(self, tmp_path) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / "c.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "매출"
        ws.append(["날짜", "금액", "비고"])
        for n in range(4):
            ws.append([f"1/{n + 1}", 100 + n, f"메모{n}"])
        wb.save(path)
        service = FileExcelLiveService()
        service.select_workbook(str(path))
        service.select_sheet(None, "매출")
        service.clear_range(None, "매출", self._plan(column="비고")["params"]["target_range"])
        sheet = load_workbook(path)["매출"]
        assert [sheet.cell(row=1, column=c).value for c in range(1, 4)] == ["날짜", "금액", "비고"]
        assert [sheet.cell(row=r, column=3).value for r in range(2, 6)] == [None] * 4
        assert [sheet.cell(row=r, column=2).value for r in range(2, 6)] == [100, 101, 102, 103]


class TestColorVocabularyHasOneSource:
    """색 사전이 셋으로 갈라져 값까지 달랐다.

    2026-08-24 실측 — 같은 낱말, 다른 색:

        낱말      규칙표      통역
        남색      #002060     #1E6B4F  ← **초록이다**
        회색      #808080     #D9D9D9
        보라색    #7030A0     #7B61FF
        주황색    #ED7D31     #FFA500

    게다가 정규식과 변환 함수가 서로 다른 목록을 들고 있어, 패턴은 `흰`을 잡는데
    함수는 몰라 폴백(노란색)으로 떨어졌다 — **"흰 글씨"가 노란 글씨가 됐다.**
    """

    @pytest.mark.parametrize(
        "word", ["남색", "navy", "회색", "보라색", "주황색", "분홍색", "하늘색", "갈색", "흰색", "검정"]
    )
    def test_every_path_agrees(self, word: str) -> None:
        from office_claw_sidecar.routers.excel_live import _quick_color_hex
        from office_claw_sidecar.services.color_lexicon import color_hex
        from office_claw_sidecar.services.excel_intent_normalizer import _COLORS

        assert _quick_color_hex(word) == color_hex(word) == _COLORS[word]

    def test_navy_is_navy_not_green(self) -> None:
        """`남색`이 초록이던 것 — 눈으로 보기 전엔 아무도 몰랐다."""
        from office_claw_sidecar.services.color_lexicon import color_hex

        assert color_hex("남색") == "#002060"
        assert color_hex("남색") == color_hex("네이비") == color_hex("navy")

    def test_bare_white_is_white(self) -> None:
        """정규식이 잡는 낱말은 사전도 알아야 한다 — 아니면 폴백(노란색)으로 샌다."""
        from office_claw_sidecar.routers.excel_live import _quick_extract_colors

        assert _quick_extract_colors("흰 글씨로 해줘") == ["#FFFFFF"]
        assert _quick_extract_colors("흰 바탕") == ["#FFFFFF"]

    def test_the_pattern_is_built_from_the_table(self) -> None:
        """목록을 두 곳에 두면 갈라진다 — 정규식을 사전에서 만든다."""
        from office_claw_sidecar.routers.excel_live import _QUICK_COLOR_PATTERN
        from office_claw_sidecar.services.color_lexicon import COLOR_TOKEN_PATTERN
        from office_claw_sidecar.services.excel_live_agent import _COLOR_TOKEN

        assert _QUICK_COLOR_PATTERN is COLOR_TOKEN_PATTERN
        assert _COLOR_TOKEN is COLOR_TOKEN_PATTERN

    def test_every_word_the_pattern_matches_is_in_the_table(self) -> None:
        """패턴이 잡는데 사전이 모르는 낱말이 있으면 그게 곧 폴백 사고다."""
        from office_claw_sidecar.services.color_lexicon import COLOR_HEX, COLOR_TOKEN_PATTERN, color_hex

        for name in COLOR_HEX:
            match = COLOR_TOKEN_PATTERN.fullmatch(name)
            assert match is not None, name
            assert color_hex(match.group(1)), name

    def test_longer_names_win(self) -> None:
        """한국어는 낱말 경계가 없다 — "노란색"이 "노란"으로 잘리면 안 된다."""
        from office_claw_sidecar.services.color_lexicon import COLOR_TOKEN_PATTERN

        assert COLOR_TOKEN_PATTERN.search("노란색으로").group(1) == "노란색"
        assert COLOR_TOKEN_PATTERN.search("분홍색 배경").group(1) == "분홍색"

    def test_an_unknown_colour_is_not_yellow(self) -> None:
        """모르는 색을 노란색으로 칠하느니 매핑 실패로 두고 되묻는 편이 낫다."""
        from office_claw_sidecar.services.color_lexicon import color_hex

        assert color_hex("형광연두") == ""
        assert color_hex("#1E6B4F") == "#1E6B4F"


class TestAggregateVocabularyHasOneSource:
    """집계 낱말 사전이 갈라져 `개수`가 경로마다 다른 함수였다.

    2026-08-24 실측:

        낱말    라우터   에이전트   바인더/보정
        개수    COUNT    COUNT      **COUNTA**

    `COUNT`와 `COUNTA`는 **다른 답을 낸다** — COUNT는 숫자만 세므로 글자가 든 열에서
    0이다. "상태 개수"를 어느 경로가 처리하느냐로 답이 달라졌다.
    테스트는 일관되게 COUNTA를 기대하므로 그쪽으로 통일했다.
    """

    @pytest.mark.parametrize("word", ["개수", "건수", "카운트", "count"])
    def test_counting_words_mean_counta(self, word: str) -> None:
        from office_claw_sidecar.routers.excel_live import _BARE_AGGREGATE_FUNC
        from office_claw_sidecar.services.aggregate_lexicon import aggregate_func
        from office_claw_sidecar.services.excel_live_agent import _build_formula_from_function

        assert aggregate_func(word) == "COUNTA"
        assert _BARE_AGGREGATE_FUNC.get(word) == "COUNTA"
        assert _build_formula_from_function(word, "A2:A9") == "=COUNTA(A2:A9)"

    @pytest.mark.parametrize(
        ("word", "func"),
        [("합계", "SUM"), ("총합계", "SUM"), ("소계", "SUM"), ("평균", "AVERAGE"),
         ("최대", "MAX"), ("최솟값", "MIN"), ("avg", "AVERAGE")],
    )
    def test_the_table_is_shared(self, word: str, func: str) -> None:
        from office_claw_sidecar.routers.excel_live import _BARE_AGGREGATE_FUNC
        from office_claw_sidecar.services.aggregate_lexicon import AGG_FUNC, aggregate_func

        assert aggregate_func(word) == func
        assert _BARE_AGGREGATE_FUNC is AGG_FUNC

    def test_the_pattern_is_built_from_the_table(self) -> None:
        """낱말 목록과 변환표가 따로면 어긋난다 — 색 쪽에서 실제 사고가 났다."""
        from office_claw_sidecar.services.aggregate_lexicon import AGG_FUNC, AGG_WORD_PATTERN, aggregate_func

        for word in AGG_FUNC:
            match = AGG_WORD_PATTERN.fullmatch(word)
            assert match is not None, word
            assert aggregate_func(match.group(0)), word

    def test_longer_words_win(self) -> None:
        """"총합계"가 "총합"으로 잘리면 안 된다 — 한국어는 낱말 경계가 없다."""
        from office_claw_sidecar.services.aggregate_lexicon import AGG_WORD_PATTERN

        assert AGG_WORD_PATTERN.search("총합계").group(0) == "총합계"
        assert AGG_WORD_PATTERN.search("최댓값").group(0) == "최댓값"

    @pytest.mark.parametrize(
        ("text", "want"),
        [("합계!", "SUM"), ("평균 좀", "AVERAGE"), ("개수", "COUNTA"),
         ("소계", "SUM"), ("ㅇㅇ 최대", "MAX")],
    )
    def test_bare_aggregate_words_still_route(self, text: str, want: str) -> None:
        from office_claw_sidecar.routers.excel_live import _bare_aggregate_after_paste

        result = _bare_aggregate_after_paste(text, "A1:F6")
        assert result is not None, text
        assert result[0] == want, (text, result)

    def test_an_unknown_word_is_not_sum(self) -> None:
        """모르는 낱말을 SUM으로 치면 엉뚱한 수식이 조용히 들어간다."""
        from office_claw_sidecar.services.aggregate_lexicon import aggregate_func

        assert aggregate_func("중앙값") == ""
        assert aggregate_func("") == ""


class TestNumberFormatHasOneSource:
    """같은 문장("퍼센트로 보여줘")이 층마다 다른 서식 코드로 풀렸다.

    2026-08-24 실측:

        낱말    라우터       통역        검증기
        퍼센트  `0.0%`       `0%`        `0.00%`
        통화    `"₩"#,##0`   `"₩"#,##0`  `#,##0`  ← 기호가 사라진다

    정당한 차이가 아니라 갈라짐이다 — `test_phrasing_robustness.py:55`와
    `test_excel_live_new_tools.py:205`가 **같은 문장**을 서로 다른 값으로 고정하고 있었다.
    """

    @pytest.mark.parametrize(
        ("word", "code"),
        [("퍼센트", "0.0%"), ("백분율", "0.0%"), ("통화", '"₩"#,##0'),
         ("천단위", "#,##0"), ("날짜", "yyyy-mm-dd")],
    )
    def test_every_layer_agrees(self, word: str, code: str) -> None:
        from office_claw_sidecar.services.excel_intent_normalizer import _format_code_from
        from office_claw_sidecar.services.excel_live_plan_validator import _NUMBER_FORMAT_ALIASES
        from office_claw_sidecar.services.number_format_lexicon import format_code

        assert format_code(word) == code
        assert _NUMBER_FORMAT_ALIASES.get(word) == code
        assert _format_code_from(word) == code

    def test_the_router_reads_the_same_table(self) -> None:
        from office_claw_sidecar.routers.excel_live import _NUMBER_FORMAT_HINTS
        from office_claw_sidecar.services.number_format_lexicon import format_code

        found = {code for pattern, code in _NUMBER_FORMAT_HINTS if pattern.search("퍼센트")}
        assert found == {format_code("퍼센트")}

    @pytest.mark.parametrize(
        ("text", "code"),
        [("천 단위 콤마", "#,##0"), ("퍼센트로", "0.0%"), ("원화 표시", '"₩"#,##0'),
         ("통화 형식", '"₩"#,##0')],
    )
    def test_it_finds_the_word_inside_a_phrase(self, text: str, code: str) -> None:
        """모델이 내는 option은 한 낱말이 아니다 — 정확일치만 보면 통째로 미매핑이다.

        2026-08-24에 이걸로 한 번 깼다(`천 단위 콤마` → None).
        """
        from office_claw_sidecar.services.number_format_lexicon import format_code_in_text

        assert format_code_in_text(text) == code

    def test_longer_words_win(self) -> None:
        """`원화`가 `원`으로 잘리면 ₩가 사라진다."""
        from office_claw_sidecar.services.number_format_lexicon import format_code_in_text

        assert format_code_in_text("원화") == '"₩"#,##0'
        assert format_code_in_text("원") == '#,##0"원"'

    def test_an_unknown_word_is_not_guessed(self) -> None:
        from office_claw_sidecar.services.number_format_lexicon import format_code, format_code_in_text

        assert format_code("지수표기") == ""
        assert format_code_in_text("알 수 없는 말") == ""


class TestTheGateSurvivesATransientFileLock:
    """옆 프로세스가 워크북을 잠깐 잡으면 **624건 실행이 통째로 죽었다.**

    2026-08-24: A/B 측정이 282번째에서 `PermissionError: blind_gate.xlsx`로 끝났다.
    내가 옆에서 pytest를 돌린 탓이다 — 자물쇠는 게이트-배터리만 막고 pytest는 안 막는다.
    한 번의 잠금으로 4시간을 날리느니 몇 초 기다렸다 이어 가는 편이 낫다.
    """

    def _module(self):
        import pathlib

        src_path = pathlib.Path(__file__).resolve().parent.parent / "scripts/run_blind_paraphrase_gate.py"
        source = src_path.read_text(encoding="utf-8")
        source = source[: source.index("asyncio.run(main())")]
        namespace: dict = {"__name__": "gate_probe"}
        exec(compile(source, str(src_path), "exec"), namespace)
        return namespace

    def test_it_waits_for_the_lock_to_clear(self, tmp_path) -> None:
        import threading
        import time

        from openpyxl import Workbook

        ns = self._module()
        # **실제 Workspace 경로를 절대 쓰지 않는다.** 처음엔 ns["WB"]를 그대로 썼는데,
        # 이 테스트가 커밋 훅·야간 pytest에서 돌 때마다 **살아 있는 게이트의 워크북을
        # 지워** 03:00 야간(6번째 문장)과 라운드 1 판정(4번째 문장)이 FileNotFoundError로
        # 죽었다(2026-08-24 실측). 잠금 사고를 고치려던 테스트가 다음 사고를 만든 것.
        # exec된 함수의 전역이 ns 딕셔너리이므로 여기서 바꾸면 함수도 tmp를 본다.
        ns["WB"] = tmp_path / "blind_gate.xlsx"
        wb_path, reset = ns["WB"], ns["_reset_workbook_file"]
        wb_path.parent.mkdir(parents=True, exist_ok=True)
        Workbook().save(wb_path)
        holder = open(wb_path, "rb")
        threading.Timer(0.8, holder.close).start()
        try:
            reset()
        finally:
            if not holder.closed:
                holder.close()
        assert not wb_path.exists()
        # 남는 시간이 있어야 다음 문장이 이어진다 — 무한정 기다리지 않는다.
        started = time.time()
        reset()
        assert time.time() - started < 1.0

    def test_a_stuck_lock_still_raises(self) -> None:
        """끝내 못 지우면 조용히 넘기지 않는다 — 그건 다음 문장을 오염시킨다."""
        import inspect

        ns = self._module()
        src = inspect.getsource(ns["_reset_workbook_file"])
        assert "WB.unlink()" in src.split("for attempt")[1], "마지막 시도가 예외를 올려야 한다"


class TestEveryConfirmActionIsClassifiedForRollback:
    """스냅샷 목록 방식 자체가 누락을 만든다 — filter_rows(08-17)·sort_rows(08-24)가
    하나씩 발견됐다. 목록을 **분류 강제**로 바꾼다: CONFIRM 액션은
    ① 스냅샷 편입 ② 사유 있는 면제 ③ 비파괴 중 하나여야 하고,
    새 액션이 분류 없이 레지스트리에 들어오면 이 핀이 커밋을 막는다.
    """

    #: 셀 값을 바꾸지 않는 CONFIRM 액션들 — 스냅샷이 필요 없는 이유가 자명한 것만.
    NON_DESTRUCTIVE = {
        "excel_live.select_workbook", "excel_live.select_sheet", "excel_live.create_sheet",
        "excel_live.rename_sheet",      # 이름만 바꾼다(수식 참조는 엑셀이 따라 고친다)
        "excel_live.fill_range",        # 배경색 — 값 무변
        "excel_live.apply_border", "excel_live.set_font", "excel_live.set_number_format",
        "excel_live.apply_color_scale", "excel_live.apply_data_bar", "excel_live.apply_formula_cf",
        "excel_live.highlight_by_condition", "excel_live.freeze_panes", "excel_live.autofit_columns",
        "excel_live.unmerge_cells",     # 병합 해제는 값을 잃지 않는다
        "excel_live.define_named_range", "excel_live.set_print_area", "excel_live.add_cell_comment",
        "excel_live.protect_sheet", "excel_live.set_data_validation", "excel_live.export_pdf",
        "excel_live.save_workbook", "excel_live.recalculate", "excel_live.refresh_power_query",
        "excel_live.rename_column",     # 머리글 한 칸 — 검증기가 되읽는다
        "excel_live.convert_to_excel_table",  # 값 위에 표 서식만 씌운다
        "excel_live.validate_data", "excel_live.compare_ranges",
        # 결과를 **새 시트**에 쓰는 집계 계열 — 원본 파괴가 아니다(출력 시트 보호는
        # `_normalize_output_sheet`가 맡는다).
        "excel_live.pivot_table", "excel_live.group_by_aggregate", "excel_live.forecast_linear",
        "excel_live.consolidate_workbooks_from_folder", "excel_live.create_chart",
        "excel_live.find_duplicates", "excel_live.calculate_column_stat",
    }

    def test_no_confirm_action_is_unclassified(self) -> None:
        from office_claw_sidecar.routers.excel_live import (
            _ROLLBACK_EXEMPT_ACTIONS,
            _ROLLBACK_SNAPSHOT_ACTIONS,
        )
        from office_claw_sidecar.services.tool_registry import TOOL_REGISTRY, PermissionLevel

        unclassified = []
        for tool in TOOL_REGISTRY:
            if tool.permission != PermissionLevel.CONFIRM:
                continue
            name = tool.name
            if not name.startswith("excel_live."):
                continue
            if (
                name in _ROLLBACK_SNAPSHOT_ACTIONS
                or name in _ROLLBACK_EXEMPT_ACTIONS
                or name in self.NON_DESTRUCTIVE
            ):
                continue
            unclassified.append(name)
        assert unclassified == [], (
            f"분류 안 된 CONFIRM 액션 {unclassified} — 스냅샷 편입/사유 있는 면제/"
            f"비파괴 중 하나로 분류하라(filter_rows·sort_rows가 이 누락으로 사고 났다)"
        )

    def test_the_twins_are_enrolled(self) -> None:
        """sort_range만 있고 sort_rows가 빠졌던 그 누락(감사 A2)."""
        from office_claw_sidecar.routers.excel_live import _ROLLBACK_SNAPSHOT_ACTIONS

        assert "excel_live.sort_rows" in _ROLLBACK_SNAPSHOT_ACTIONS
        assert "excel_live.find_replace" in _ROLLBACK_SNAPSHOT_ACTIONS

    def test_exemptions_carry_reasons(self) -> None:
        """면제는 결정이지 누락이 아니다 — 사유 없는 면제는 누락과 같다."""
        from office_claw_sidecar.routers.excel_live import _ROLLBACK_EXEMPT_ACTIONS

        for action, reason in _ROLLBACK_EXEMPT_ACTIONS.items():
            assert len(str(reason).strip()) >= 10, action
        # 값 스냅샷으로 복원이 안 되는 대표: 차트는 셀 값이 아니다.
        assert "excel_live.delete_charts" in _ROLLBACK_EXEMPT_ACTIONS


class TestSortRowsSnapshotRestoresTheOriginalOrder:
    """sort_rows는 target_range 파라미터가 없고 늘 **사용 범위 전체**를 정렬한다.

    else 분기(활성 선택 영역)로 흘리면 실제로 섞이는 범위와 어긋난 복원이 된다 —
    전용 분기가 사용 범위를 뜨는지, 그리고 그 스냅샷이 정말 원상 복구되는지를 굳힌다.
    """

    def test_snapshot_sort_restore_roundtrip(self, tmp_path, monkeypatch) -> None:
        from openpyxl import Workbook, load_workbook

        from office_claw_sidecar.routers.excel_live import (
            _restore_action_snapshot,
            _snapshot_target_range_for_action,
        )
        from office_claw_sidecar.services.excel_live_service import get_excel_live_service

        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        path = tmp_path / "s.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "성적부"
        for row in [["학생", "점수"], ["김민준", 88], ["이서연", 94], ["박도윤", 71]]:
            ws.append(row)
        wb.save(path)
        service = get_excel_live_service()
        service.select_workbook(str(path))
        service.select_sheet(None, "성적부")

        snapshot = _snapshot_target_range_for_action(
            action="excel_live.sort_rows",
            params={"column": "점수", "order": "desc"},
            workbook_id=None,
            sheet_name="성적부",
        )
        assert snapshot is not None
        assert snapshot.range_ref == "A1:B4", snapshot.range_ref  # 사용 범위 전체

        service.sort_rows(None, "성적부", column="점수", order="desc")
        scrambled = [load_workbook(path)["성적부"].cell(row=r, column=1).value for r in range(2, 5)]
        assert scrambled != ["김민준", "이서연", "박도윤"], "정렬이 실제로 섞였어야 한다"

        assert _restore_action_snapshot(snapshot) is True
        restored = [load_workbook(path)["성적부"].cell(row=r, column=1).value for r in range(2, 5)]
        assert restored == ["김민준", "이서연", "박도윤"]

    def test_restore_failure_is_surfaced_not_swallowed(self) -> None:
        """스냅샷이 있는데 복원이 실패하면 detail에 실패가 표면화돼야 한다(감사 A1).

        조용히 삼키면 사용자는 "검증 실패 + 원상 복구"로 읽는데 실제로는 파일이
        반쯤 바뀐 채다.
        """
        import inspect

        from office_claw_sidecar.routers import excel_live

        src = inspect.getsource(excel_live._execute_plan_and_respond)
        assert "auto_rollback_FAILED" in src


class TestBatch1aIntentKindsMapDeterministically:
    """어휘에 종류가 없어 **엉뚱한 종류로 밀려나던** 4종(감사 C2, 라운드 2 배치 1a).

    "요약 시트 만들어줘"→create_table, "차트 없애줘"→clear_values로 밀려났었다.
    종류를 넣으면 최소한 정직한 폴백이 되고, 파라미터가 결정적인 이 4종은 매핑까지 한다.
    """

    def _plan(self, message="", **intent):
        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan

        base = {"task": None, "range": None, "column": None, "option": None}
        digest = {"active_sheet": "매출", "sheets": [{"name": "매출", "used_range": "A1:C5", "columns": []}]}
        return intent_to_plan({**base, **intent}, digest=digest, message=message)

    # 2026-08-25부터 매핑은 **문장 근거**를 요구한다(게이트 회귀 7건의 교훈) —
    # 핀도 실제 문장을 싣는다. 근거 없는 매핑 거부는 아래 별도 클래스가 굳힌다.
    @pytest.mark.parametrize(
        ("message", "intent", "action", "params"),
        [
            ("요약 시트 하나 만들어줘", {"task": "create_sheet", "option": "요약"},
             "excel_live.create_sheet", {"sheet_name": "요약"}),
            # 모델이 이름을 column에 싣는 편차도 받는다.
            ("정산 시트 새로 만들어줘", {"task": "create_sheet", "column": "정산"},
             "excel_live.create_sheet", {"sheet_name": "정산"}),
            ("차트 전부 없애줘", {"task": "delete_charts"}, "excel_live.delete_charts", {}),
            ("첫 줄 고정해줘", {"task": "freeze"}, "excel_live.freeze_panes", {"freeze_at": "A2"}),
            ("위 2줄 고정해줘", {"task": "freeze", "option": "2"},
             "excel_live.freeze_panes", {"freeze_at": "A3"}),
            ("열 너비 맞춰줘", {"task": "autofit"},
             "excel_live.autofit_columns", {"target_range": "__USED_RANGE__"}),
        ],
    )
    def test_it_maps(self, message, intent, action, params) -> None:
        plan = self._plan(message=message, **intent)
        assert plan is not None, intent
        assert plan["action"] == action
        assert plan["params"] == params

    @pytest.mark.parametrize(
        ("message", "intent"),
        [
            # 이름을 지어내 시트를 만들면 사용자가 부른 적 없는 시트가 생긴다.
            ("시트 만들어줘", {"task": "create_sheet"}),
            ("시트 만들어줘", {"task": "create_sheet", "option": "시트"}),   # 낱말 축퇴
            ("가" * 32 + " 시트 만들어줘", {"task": "create_sheet", "option": "가" * 32}),  # 31자 상한
        ],
    )
    def test_it_backs_off(self, message, intent) -> None:
        assert self._plan(message=message, **intent) is None, intent

    def test_an_absurd_freeze_row_falls_back_to_the_header(self) -> None:
        """999행 고정 같은 과대값은 A2(머리글만 고정)로 — 틀 고정은 되돌릴 수 있어 안전한 쪽."""
        plan = self._plan(message="999줄 고정해줘", task="freeze", option="999")
        assert plan["params"] == {"freeze_at": "A2"}


class TestXlwingsPathsRejectVendorWorkbooks:
    """벤더 데모 파일 때문에 xlwings가 골라지고, `books.active`가 그 파일에 실행했다.

    2026-08-04 실측(감사 B1): .venv의 xlwings quickstart가 열려 있다는 이유로 auto가
    xlwings 엔진을 골랐고, 지목 없는 명령이 그 데모 파일을 편집했다. 가드가 파일
    엔진(`_is_scannable`)에만 있고 사고가 난 xlwings 경로에는 없었다.
    """

    @pytest.mark.parametrize(
        "path",
        [
            r"C:\proj\.venv\Lib\site-packages\xlwings\quickstart.xlsm",
            r"C:\ws\officeclaw_backups\백업.xlsx",
            r"C:\a\node_modules\pkg\demo.xlsx",
        ],
    )
    def test_vendor_paths_are_rejected(self, path: str) -> None:
        from office_claw_sidecar.services.excel_live_service import _is_user_workbook_path

        assert _is_user_workbook_path(path) is False

    @pytest.mark.parametrize(
        "path",
        [
            r"C:\Users\a\AppData\Local\office_claw\Workspace\매출.xlsx",
            r"C:\Users\a\Documents\재고.xlsx",
            "통합 문서1",  # 저장 전 새 문서 — 경로가 없다고 거절하면 정상 사용이 막힌다
            "",
        ],
    )
    def test_user_paths_pass(self, path: str) -> None:
        from office_claw_sidecar.services.excel_live_service import _is_user_workbook_path

        assert _is_user_workbook_path(path) is True

    def test_the_active_fallback_refuses_a_vendor_workbook(self) -> None:
        """지목이 없을 때의 `books.active` 폴백 — 벤더 파일이면 실행 대신 거절."""
        from types import SimpleNamespace

        from office_claw_sidecar.services.excel_live_service import (
            ExcelLiveService,
            WorkbookNotFoundError,
        )

        service = ExcelLiveService.__new__(ExcelLiveService)  # COM 없이 골격만
        service._selected_workbook_id = None
        vendor = SimpleNamespace(fullname=r"C:\proj\.venv\xlwings\quickstart.xlsm")
        service._app = lambda: SimpleNamespace(books=SimpleNamespace(active=vendor))
        with pytest.raises(WorkbookNotFoundError):
            service._resolve_workbook(None)

        # 사람 파일이면 그대로 돌려준다.
        mine = SimpleNamespace(fullname=r"C:\Users\a\Documents\재고.xlsx")
        service._app = lambda: SimpleNamespace(books=SimpleNamespace(active=mine))
        assert service._resolve_workbook(None) is mine

    def test_the_engine_probe_ignores_vendor_books(self) -> None:
        """벤더 파일만 열려 있으면 xlwings를 고를 근거가 아니다."""
        import inspect

        from office_claw_sidecar.services import excel_live_service

        src = inspect.getsource(excel_live_service._excel_app_has_open_workbook)
        assert "_is_user_workbook_path" in src

    def test_the_exclusion_list_has_one_source(self) -> None:
        from office_claw_sidecar.services import excel_live_file_service, excel_live_service

        assert excel_live_file_service._SCAN_EXCLUDED_DIRS is excel_live_service._SCAN_EXCLUDED_DIRS


class TestIntentPlansRequireEvidenceInTheMessage:
    """라운드 2-1a가 만든 회귀 7건(2026-08-25 게이트 594/624) — 세 가족의 핀.

    공통 원인: **모델이 낸 종류·파라미터를 문장이 뒷받침하는지 확인하지 않았다.**
    - "H1부터 M1까지 한 칸으로 붙여줘"(병합)를 autofit으로 분류 → 무조건 실행
    - "표 내용 싹 지워"(전체)에 column=지역(첫 머리글)을 지어냄 → A열만 지움
    - 이름 자리에 문자열 "null" → **'null'이라는 시트가 생김**
    아래 intent들은 게이트에서 모델이 실제로 낸 출력 그대로다.
    """

    DIGEST = {
        "active_sheet": "지역성과",
        "sheets": [
            {
                "name": "지역성과",
                "used_range": "A1:F6",
                "columns": [{"letter": "A", "header": "지역"}, {"letter": "F", "header": "클레임"}],
            }
        ],
    }

    def _plan(self, msg, **kw):
        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan

        base = {"task": None, "range": None, "column": None, "option": None}
        return intent_to_plan({**base, **kw}, digest=self.DIGEST, message=msg)

    @pytest.mark.parametrize(
        ("msg", "intent"),
        [
            # 병합 문장인데 autofit 분류 + 스키마 어휘 되뇜("autofit")
            ("여기 H1부터 M1까지 한 칸으로 붙여줘",
             {"task": "autofit", "column": "H1:M1", "option": "autofit"}),
            # 전체 지우기인데 column을 지어냄 + 되뇜("clear_values")
            ("표 칸들 전부 깨끗하게 지워서 빈칸으로 해줘",
             {"task": "clear_values", "column": "지역", "option": "clear_values"}),
            ("이거 내용 싹 지워서 빈칸으로 만들어",
             {"task": "clear_values", "column": "지역", "option": "clear_values"}),
            # JSON null 대신 문자열 "null"
            ("new sheet 추가하고 이름은 요약으로 해 주세요.",
             {"task": "create_sheet", "column": "null", "option": "null", "range": "null"}),
        ],
    )
    def test_the_gate_regressions_now_back_off(self, msg, intent) -> None:
        assert self._plan(msg, **intent) is None, (msg, intent)

    @pytest.mark.parametrize(
        ("msg", "intent", "action"),
        [
            ("클레임 열 비워줘", {"task": "clear_values", "column": "클레임"}, "excel_live.clear_range"),
            ("열 너비 보기 좋게 맞춰줘", {"task": "autofit"}, "excel_live.autofit_columns"),
            ("요약 시트 하나 만들어줘", {"task": "create_sheet", "option": "요약"}, "excel_live.create_sheet"),
            ("차트 전부 없애줘", {"task": "delete_charts"}, "excel_live.delete_charts"),
            ("첫 줄 고정해줘", {"task": "freeze"}, "excel_live.freeze_panes"),
        ],
    )
    def test_corroborated_plans_still_map(self, msg, intent, action) -> None:
        plan = self._plan(msg, **intent)
        assert plan is not None and plan["action"] == action, (msg, plan)


class TestApprovalDoesNotBypassSafetyChecks:
    """`approve=True`가 블라스트 반경·계획 위생 검사를 통째로 건너뛰었다(감사 B2).

    `/macro/step`은 하위 명령 전체를 approve=True로 보낸다(매크로 승인 한 번 =
    전체 승인). 그런데 하위 명령의 계획은 **사람이 본 적이 없다** — 재계획도 같다.
    승인이 면제하는 것은 승인 카드 재요청뿐이고, 검사는 항상 수행해야 한다.
    """

    def _gate(self, *, message, value, approved, tmp_path):
        from openpyxl import Workbook

        from office_claw_sidecar.routers.excel_live import (
            ExcelLiveCommandRequest,
            PlanExecution,
            _plan_approval_gate,
        )
        from office_claw_sidecar.services.excel_live_executor import PlanStep
        from office_claw_sidecar.services.excel_live_service import get_excel_live_service

        path = tmp_path / "b2.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "성적부"
        ws.append(["학생", "점수"])
        wb.save(path)
        service = get_excel_live_service()
        service.select_workbook(str(path))
        service.select_sheet(None, "성적부")

        plan = [PlanStep(action="excel_live.write_range",
                         params={"start_cell": "D1", "values_2d": [[value]]})]
        req = ExcelLiveCommandRequest(message=message, session_id="t-b2", workbook_id=None,
                                      approve=approved, context_range=None)
        ctx = PlanExecution(req=req, plan=plan, session_key="t-b2",
                            parsed={"plan_source": "planner"}, approved=approved)
        return _plan_approval_gate(ctx, plan)

    def test_an_approved_sanity_violation_is_still_blocked(self, tmp_path, monkeypatch) -> None:
        """매크로 사고 모양: 하위 명령 문장(지시문)을 플래너가 **값으로** 쓴 계획."""
        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        out = self._gate(message="B2에 합계 넣어줘", value="B2에 합계 넣어줘",
                         approved=True, tmp_path=tmp_path)
        assert out is not None and out.action == "excel_live.clarify", out

    def test_an_approved_clean_plan_proceeds_without_a_new_card(self, tmp_path, monkeypatch) -> None:
        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        out = self._gate(message="D1에 비고 라고 써줘", value="비고",
                         approved=True, tmp_path=tmp_path)
        assert out is None, out

    def test_the_early_return_stays_after_the_checks(self) -> None:
        """조기 반환이 검사 위로 다시 올라가면 이 구멍이 조용히 되살아난다."""
        import inspect

        from office_claw_sidecar.routers import excel_live

        src = inspect.getsource(excel_live._plan_approval_gate)
        # 주석에도 'if ctx.approved:'가 등장하므로 실행문 쪽의 고유 문구로 짚는다.
        assert src.index("_assess_plan_sanity") < src.index("검사를 통과한 승인 계획")


class TestPivotInjectionRequiresTheExplicitWord:
    """'집계·요약' 낱말만으로 피벗 단계가 주입되던 것(감사 B4).

    group_by 문형("부서별로 집계해줘")이 옳게 계획돼도, 주입 규칙이 '집계'를 보고
    시키지 않은 피벗 시트를 얹었다. 시트 이름이 '요약'이기만 해도 걸린다(블라인드
    624에 그런 문장 27건 — 전부 PASS_RULE이라 주입 코드에 닿지 않아 무사했을 뿐).
    주입은 원문이 피벗을 명시했을 때만, 근거 게이트는 종전대로 느슨하게.
    """

    @staticmethod
    def _step(action: str):
        from types import SimpleNamespace

        return SimpleNamespace(action=action)

    def test_aggregate_wording_alone_does_not_inject(self) -> None:
        from office_claw_sidecar.routers.excel_live import _should_inject_pivot_step

        assert not _should_inject_pivot_step("부서별로 집계해서 새 시트에 넣어줘", [])
        assert not _should_inject_pivot_step("요약 시트에 지역 주문건수 합계 넣어줘", [])

    def test_an_already_planned_aggregate_is_not_doubled(self) -> None:
        from office_claw_sidecar.routers.excel_live import _should_inject_pivot_step

        plan = [self._step("excel_live.group_by_aggregate")]
        assert not _should_inject_pivot_step("부서별 합계를 피벗으로 정리해줘", plan)

    def test_an_explicit_pivot_request_still_injects(self) -> None:
        from office_claw_sidecar.routers.excel_live import _should_inject_pivot_step

        assert _should_inject_pivot_step("담당자별 합계를 피벗으로 만들어줘", [])
        assert not _should_inject_pivot_step(
            "담당자별 합계를 피벗으로 만들어줘", [self._step("excel_live.pivot_table")]
        )

    def test_the_evidence_gate_stays_loose_for_planner_chosen_pivots(self) -> None:
        """플래너가 '집계해줘'를 피벗으로 푸는 건 정당한 해석 — 근거 게이트가 막으면 안 된다."""
        from office_claw_sidecar.routers.excel_live import _action_lacks_evidence

        assert not _action_lacks_evidence("excel_live.pivot_table", "지역별로 집계해줘")


class TestSetFormulaVerificationSeesErrorCells:
    """일반 `set_formula` 검증이 자기보고(applied>0)만 보던 것(감사 B3).

    xlwings `.value`는 오류 셀을 None으로 돌려주므로(0825-074210-formula-verify
    실측) 값 경로만 보면 #NAME?도 빈칸으로 보였다. 전용 판독(count_error_cells)을
    거쳐 **구조 오류만** 실패로 본다 — #N/A·#DIV/0!는 데이터에 따라 정당하다.
    """

    @staticmethod
    def _verify(service):
        from office_claw_sidecar.services.excel_result_verifier import verify_effect

        return verify_effect(
            action="excel_live.set_formula",
            params={"range_ref": "D3", "formula_a1": "=SUMM(B2:B3)"},
            result={"formula_applied_cells": 1, "address": "D3"},
            service=service,
            workbook_id="wb",
            sheet_name="S",
        )

    def test_a_structural_error_fails_with_the_cell_named(self) -> None:
        from types import SimpleNamespace

        service = SimpleNamespace(count_error_cells=lambda *a: {"#NAME?": ["D3"]})
        ok, detail = self._verify(service)
        assert not ok and detail.startswith("formula_error_cells:") and "D3" in detail

    def test_data_shaped_errors_are_not_failures(self) -> None:
        from types import SimpleNamespace

        service = SimpleNamespace(count_error_cells=lambda *a: {"#N/A": ["D3"], "#DIV/0!": ["D4"]})
        ok, _ = self._verify(service)
        assert ok

    def test_an_engine_without_the_reader_still_passes(self) -> None:
        """옛 엔진·판독 실패는 '모름'이다 — 못 본 것을 실패로 단정하지 않는다."""
        from types import SimpleNamespace

        ok, _ = self._verify(SimpleNamespace())
        assert ok

        def _boom(*a):
            raise RuntimeError("판독 실패")

        ok, _ = self._verify(SimpleNamespace(count_error_cells=_boom))
        assert ok

    def test_the_file_engine_reports_unknown_for_fresh_formulas(self, tmp_path) -> None:
        """openpyxl은 계산하지 않는다 — 방금 쓴 오타 수식은 캐시가 없어 '모름'({})이다.

        빈 결과를 "오류 없음"으로 읽어 실패시키면 파일 엔진의 모든 수식이 오탐이 된다.
        """
        from openpyxl import Workbook

        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        path = tmp_path / "b3.xlsx"
        wb = Workbook()
        wb.active.title = "S"
        wb.save(path)
        svc = FileExcelLiveService()
        svc.select_workbook(str(path))
        svc.set_formula(str(path), "S", "D3", "=SUMM(B2:B3)")
        assert svc.count_error_cells(str(path), "S", "D3") == {}
        ok, _ = verify_effect_for_formula(svc, str(path))
        assert ok


def verify_effect_for_formula(service, workbook_id):
    from office_claw_sidecar.services.excel_result_verifier import verify_effect

    return verify_effect(
        action="excel_live.set_formula",
        params={"range_ref": "D3", "formula_a1": "=SUMM(B2:B3)"},
        result={"formula_applied_cells": 1, "address": "D3"},
        service=service,
        workbook_id=workbook_id,
        sheet_name="S",
    )


class TestApprovalDoesNotCoverAnInjectedOverwrite:
    """'승인과 다른 확정 액션' 주입 — 라운드 5 완료 판정의 세 번째 기둥.

    블라스트 반경은 카드 장식에만 쓰였다 — 승인 경로엔 카드가 없어 **계산되고
    버려졌다.** 매크로 하위 명령·재계획이 지목 밖의 값을 덮어도 무검문 실행
    (2026-08-19 크로스시트 집계가 학생 이름을 덮은 사고 모양). 이제 **본 적 없는
    계획**(unseen_plan: 매크로 하위 명령·재계획)의 모델 계획만 되묻기로 선다.
    카드로 승인된 계획(resume)과 요청 수준 approve=True 계약은 그대로 지나간다 —
    처음 구현이 이 둘을 구분 못 해 전체 pytest에서 17건이 깨졌다(2026-08-25 실측).
    """

    def _gate(self, *, start_cell, value, tmp_path):
        from openpyxl import Workbook

        from office_claw_sidecar.routers.excel_live import (
            ExcelLiveCommandRequest,
            PlanExecution,
            _plan_approval_gate,
        )
        from office_claw_sidecar.services.excel_live_executor import PlanStep
        from office_claw_sidecar.services.excel_live_service import get_excel_live_service

        path = tmp_path / "inject.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "성적부"
        ws.append(["학생", "점수"])
        ws.append(["김민준", 88])
        wb.save(path)
        service = get_excel_live_service()
        service.select_workbook(str(path))
        service.select_sheet(None, "성적부")

        plan = [PlanStep(action="excel_live.write_range",
                         params={"start_cell": start_cell, "values_2d": [[value]]})]
        req = ExcelLiveCommandRequest(message="D1에 비고 라고 써줘", session_id="t-inject",
                                      workbook_id=None, approve=True, context_range=None)
        ctx = PlanExecution(req=req, plan=plan, session_key="t-inject",
                            parsed={"plan_source": "planner"}, approved=True)
        return _plan_approval_gate(ctx, plan, unseen_plan=True)

    def test_an_injected_overwrite_outside_the_pointed_cell_asks(self, tmp_path, monkeypatch) -> None:
        """지목은 D1인데 계획은 값이 든 A2를 덮는다 — 승인돼 있어도 서야 한다."""
        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        out = self._gate(start_cell="A2", value="비고", tmp_path=tmp_path)
        assert out is not None and out.action == "excel_live.clarify", out
        assert out.result.get("sanity_code") == "blast_radius_after_approval"

    def test_the_pointed_cell_itself_still_proceeds(self, tmp_path, monkeypatch) -> None:
        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        out = self._gate(start_cell="D1", value="비고", tmp_path=tmp_path)
        assert out is None, out

    def test_a_card_approved_or_contract_request_is_not_re_asked(self, tmp_path, monkeypatch) -> None:
        """카드에서 ⚠ 경고를 보고 승인한 계획(resume 첫 바퀴)과 approve=True 계약은
        위험해도 다시 세우지 않는다 — unseen_plan이 아닐 때는 통과."""
        from openpyxl import Workbook

        from office_claw_sidecar.routers.excel_live import (
            ExcelLiveCommandRequest,
            PlanExecution,
            _plan_approval_gate,
        )
        from office_claw_sidecar.services.excel_live_executor import PlanStep
        from office_claw_sidecar.services.excel_live_service import get_excel_live_service

        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        path = tmp_path / "seen.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "성적부"
        ws.append(["학생", "점수"])
        ws.append(["김민준", 88])
        wb.save(path)
        service = get_excel_live_service()
        service.select_workbook(str(path))
        service.select_sheet(None, "성적부")
        plan = [PlanStep(action="excel_live.write_range",
                         params={"start_cell": "A2", "values_2d": [["비고"]]})]
        req = ExcelLiveCommandRequest(message="D1에 비고 라고 써줘", session_id="t-seen",
                                      workbook_id=None, approve=True, context_range=None)
        ctx = PlanExecution(req=req, plan=plan, session_key="t-seen",
                            parsed={"plan_source": "planner"}, approved=True)
        assert _plan_approval_gate(ctx, plan) is None  # unseen_plan=False(기본)


class TestBatch1bIntentKindsMapDeterministically:
    """라운드 2 배치 1b — 명시 어휘가 있는 5종(merge·unmerge·data_bar·color_scale·
    rename_sheet)의 결정적 매핑. 전부 문장 근거를 요구하고, 불확실하면 물러난다.
    """

    def _plan(self, message="", **intent):
        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan

        base = {"task": None, "range": None, "column": None, "option": None}
        digest = {
            "active_sheet": "지역성과",
            "sheets": [{
                "name": "지역성과", "used_range": "A1:F6",
                "columns": [{"letter": "B", "header": "주문건수"}],
                "row_count": 6,
            }],
        }
        return intent_to_plan({**base, **intent}, digest=digest, message=message)

    @pytest.mark.parametrize(
        ("message", "intent", "action", "params"),
        [
            ("A1:F1 병합해줘", {"task": "merge", "range": "A1:F1"},
             "excel_live.merge_cells", {"target_range": "A1:F1"}),
            ("병합 다 풀어줘", {"task": "unmerge"},
             "excel_live.unmerge_cells", {"target_range": "__USED_RANGE__"}),
            ("주문건수 열에 데이터 막대 넣어줘", {"task": "data_bar", "column": "주문건수"},
             "excel_live.apply_data_bar", {"target_range": "B2:B6"}),
            ("B2:B6 색조 넣어줘", {"task": "color_scale", "range": "B2:B6"},
             "excel_live.apply_color_scale", {"target_range": "B2:B6"}),
            ("시트 이름을 지역별실적으로 바꿔줘", {"task": "rename_sheet", "option": "지역별실적"},
             "excel_live.rename_sheet", {"new_name": "지역별실적"}),
        ],
    )
    def test_the_kind_maps_to_its_action(self, message, intent, action, params) -> None:
        plan = self._plan(message, **intent)
        steps = (plan or {}).get("action_plan") or []
        assert steps and steps[0]["action"] == action, plan
        for key, value in params.items():
            assert steps[0]["params"].get(key) == value, steps[0]

    def test_merge_without_a_range_backs_off(self) -> None:
        """병합은 왼쪽 위 칸만 남긴다 — 범위를 지어내면 값이 사라진다."""
        assert self._plan("제목 병합해줘", task="merge") is None

    def test_a_particle_suffixed_new_name_is_stemmed(self) -> None:
        """모델이 '지역별실적으로'를 통째로 이름에 실어도 어간을 벗긴다 — 조사째
        시트 이름이 되면 블라인드 실측 오답과 같은 모양이 된다."""
        plan = self._plan("시트 이름을 지역별실적으로 바꿔줘",
                          task="rename_sheet", option="지역별실적으로")
        steps = (plan or {}).get("action_plan") or []
        assert steps and steps[0]["params"]["new_name"] == "지역별실적", plan

    def test_a_bare_bar_word_does_not_feed_data_bar(self) -> None:
        """'막대 그래프'는 차트다 — data_bar 오분류가 낱말 '막대'로 근거를 얻으면 안 된다."""
        assert self._plan("주문건수 막대 그래프 그려줘", task="data_bar", column="주문건수") is None

    def test_an_unmerge_shaped_sentence_cannot_feed_merge(self) -> None:
        """'병합 해제'에서 task=merge로 잘못 와도 해제 낱말이 있으면 물러난다."""
        assert self._plan("병합 해제해줘", task="merge", range="A1:F1") is None


class TestBatch2IntentKindsMapDeterministically:
    """라운드 2 배치 2 — 열 연산 4종 + group_by(읽기 전용 조회).

    파괴 성격(시트·열 삭제)이라 물러남 조건이 매핑 조건만큼 중요하다:
    이름이 문장에 없거나, 머리글·시트로 확정되지 않으면 절대 맵핑하지 않는다.
    """

    def _plan(self, message="", **intent):
        from office_claw_sidecar.services.excel_intent_normalizer import intent_to_plan

        base = {"task": None, "range": None, "column": None, "option": None}
        digest = {
            "active_sheet": "지역성과",
            "sheets": [{
                "name": "지역성과", "used_range": "A1:F6",
                "columns": [
                    {"letter": "A", "header": "지역"},
                    {"letter": "B", "header": "주문건수"},
                ],
                "row_count": 6,
            }, {"name": "임시"}],
        }
        return intent_to_plan({**base, **intent}, digest=digest, message=message)

    @pytest.mark.parametrize(
        ("message", "intent", "action", "params"),
        [
            ("임시 시트 삭제해줘", {"task": "delete_sheet", "option": "임시"},
             "excel_live.delete_sheet", {"sheet_name": "임시"}),
            ("주문건수 열 지워줘", {"task": "drop_column", "column": "주문건수"},
             "excel_live.drop_column", {"column": "주문건수"}),
            ("비고 열 하나 추가해줘", {"task": "add_column", "option": "비고"},
             "excel_live.add_column", {"name": "비고"}),
            ("주문건수 열 이름을 판매건수로 바꿔줘",
             {"task": "rename_column", "column": "주문건수", "option": "판매건수로"},
             "excel_live.rename_column", {"column": "주문건수", "new_name": "판매건수"}),
            ("지역별 주문건수 합계 알려줘", {"task": "group_by", "column": "주문건수", "option": "합계"},
             "excel_live.group_by_aggregate",
             {"group_column": "지역", "value_column": "주문건수", "agg": "sum"}),
            # 모델은 column 슬롯에 묶음 기준을 싣는 일이 잦다(0825 실측) — 값 열은
            # 문장에 등장하는 다른 머리글에서 찾아야 한다.
            ("지역별 주문건수 합계 알려줘", {"task": "group_by", "column": "지역", "option": "SUM"},
             "excel_live.group_by_aggregate",
             {"group_column": "지역", "value_column": "주문건수", "agg": "sum"}),
            ("지역별로 몇 건씩인지 알려줘", {"task": "group_by", "option": "건수"},
             "excel_live.group_by_aggregate", {"group_column": "지역", "agg": "count"}),
        ],
    )
    def test_the_kind_maps_to_its_action(self, message, intent, action, params) -> None:
        plan = self._plan(message, **intent)
        steps = (plan or {}).get("action_plan") or []
        assert steps and steps[0]["action"] == action, plan
        for key, value in params.items():
            assert steps[0]["params"].get(key) == value, steps[0]

    def test_deleting_an_absent_sheet_backs_off(self) -> None:
        """없는 시트 이름·문장에 없는 이름 — 어느 쪽이든 지어내서 지우면 안 된다."""
        assert self._plan("없는탭 시트 삭제해줘", task="delete_sheet", option="없는탭") is None
        assert self._plan("시트 삭제해줘", task="delete_sheet", option="임시") is None

    def test_dropping_an_unresolved_column_backs_off(self) -> None:
        assert self._plan("매출 열 지워줘", task="drop_column", column="매출") is None

    def test_group_by_backs_off_when_writing_is_wanted(self) -> None:
        """"~별 합계를 넣어줘"는 조회가 아니다 — 읽기 전용 매핑이 가로채면
        사용자는 결과가 시트에 써졌다고 믿는데 실제로는 아무 데도 안 써진다."""
        assert self._plan("지역별 주문건수 합계 G열에 넣어줘",
                          task="group_by", column="주문건수", option="합계") is None

    def test_group_by_without_the_byword_backs_off(self) -> None:
        """머리글+별 짝이 문장에 없으면 묶음 기준을 지어내지 않는다."""
        assert self._plan("주문건수 합계 알려줘", task="group_by",
                          column="주문건수", option="합계") is None


class TestIntentAttemptsAreAlwaysLogged:
    """의도 해석의 실패 시도가 로그에 안 남던 공백(2026-08-18부터 알려짐).

    성공만 기록하면 "얼마나 자주 시도했고 왜 물러났는가"를 셀 수 없다 — 커버리지
    확장의 근거 수치가 전부 수동 재측정에 기대게 된다. 이제 시도마다 outcome이
    남는다: mapped · unmapped · error:<예외형>.
    """

    @staticmethod
    async def _call(monkeypatch, *, intent_result=None, intent_error=None):
        import office_claw_sidecar.services.excel_live_agent as agent

        notes: list[dict] = []
        monkeypatch.setattr(
            agent, "trace_note", lambda kind, **kw: notes.append({"kind": kind, **kw})
        )

        async def _fake_normalize(message, digest, llm):
            if intent_error is not None:
                raise intent_error
            return intent_result

        monkeypatch.setattr(agent, "normalize_intent", _fake_normalize)

        async def _fake_planner(message, llm_service, context=None):
            return {"action_plan": [], "intent": "unknown"}

        monkeypatch.setattr(agent, "parse_command_plan_with_llm", _fake_planner)
        try:
            await agent.parse_excel_live_command("병합 다 풀어줘", None, context={})
        except Exception:
            pass  # 뒷단(플래너 흉내)의 모양은 이 핀의 관심사가 아니다
        return [n for n in notes if n.get("purpose") == "intent_normalizer"]

    def test_a_mapped_attempt_records_mapped(self, monkeypatch) -> None:
        import asyncio

        notes = asyncio.run(self._call(
            monkeypatch, intent_result={"task": "unmerge", "range": None, "column": None, "option": None}
        ))
        assert notes and notes[0]["outcome"] == "mapped", notes
        assert notes[0]["mapped_action"] == "excel_live.unmerge_cells"

    def test_an_unmapped_attempt_records_unmapped(self, monkeypatch) -> None:
        import asyncio

        notes = asyncio.run(self._call(
            monkeypatch, intent_result={"task": "pivot", "range": None, "column": None, "option": None}
        ))
        assert notes and notes[0]["outcome"] == "unmapped", notes

    def test_a_failed_call_records_the_error_kind(self, monkeypatch) -> None:
        import asyncio

        notes = asyncio.run(self._call(monkeypatch, intent_error=TimeoutError("느림")))
        assert notes and notes[0]["outcome"] == "error:TimeoutError", notes


class TestGuiTableInterviewIncident20260825:
    """2026-08-25 GUI 실사고 — 빈 범위를 붙여넣고 표를 만들려던 다섯 턴이 전부 무너졌다.

    (A) "A1:D6 여기에 입력 좀 해줘" → '좀' 때문에 값 되묻기 규칙을 빠져나가 플래너가
        가짜 표(김철수/개발자)를 지어냄(스키마 불일치로만 막힘)
    (D) 표 인터뷰의 헤더 답 "…비중, 전월대비로 만들어줘" → '비중'(→도넛)이 차트 훅을 켜고
        표 슬롯을 밀어내 **표 대신 빈 차트**가 만들어짐. 고치자 '전월대비'가 compare로 읽혀 또 밀림
    (E) "아니 표를 만들랬지 차트를 만들게 하지는 않았는데?" → 명사 '차트'만 주워 "차트 종류를…"

    다섯 턴을 LLM 없이(금지 스텁) 끝까지 재생하고 **파일 상태**로 판정한다.
    """

    @staticmethod
    def _client(tmp_path, monkeypatch):
        from types import SimpleNamespace

        from fastapi.testclient import TestClient
        from openpyxl import Workbook

        from office_claw_sidecar.main import app
        from office_claw_sidecar.routers import excel_live as router
        from office_claw_sidecar.services.llm_service import get_llm_service

        monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
        xlsx = tmp_path / "gui.xlsx"
        wb = Workbook()
        wb.active.title = "데이터"
        wb.save(xlsx)
        svc = router.get_excel_live_service()
        svc.select_workbook(str(xlsx))
        svc.select_sheet(None, "데이터")

        class _NoLLM:
            def __getattr__(self, name):
                def _refuse(*a, **k):
                    raise RuntimeError(f"LLM 호출 금지({name})")
                return _refuse

        # monkeypatch로 걸어야 테스트가 끝나면 **복원**된다 — 직접 대입하면 뒤 테스트가
        # 진짜 플래너를 못 불러 순서 의존 실패가 난다(2026-08-25 전체 pytest 실측:
        # test_noop_honesty가 이 클래스 뒤에서만 깨졌다).
        monkeypatch.setitem(app.dependency_overrides, get_llm_service, lambda: _NoLLM())
        router._pending_create_table_slots.clear()
        router._pending_operation_slots.clear()
        return TestClient(app), xlsx, SimpleNamespace(session="excel-live::ui::pin-0825")

    @staticmethod
    def _post(client, session, msg, ctx=None):
        body = {"message": msg, "session_id": session, "approve": False}
        if ctx:
            body["context_range"] = ctx
        r = client.post("/excel-live/command", json=body, headers={"Authorization": "Bearer dev-token"})
        assert r.status_code == 200, r.text
        return r.json()

    def test_the_five_turns_end_with_a_table_not_a_chart(self, tmp_path, monkeypatch) -> None:
        from openpyxl import load_workbook

        client, xlsx, ns = self._client(tmp_path, monkeypatch)
        headers = {"Authorization": "Bearer dev-token"}

        a = self._post(client, ns.session, "A1:D6 여기에 입력 좀 해줘")
        assert "어떤 값을 넣을까요" in str(a.get("reason")), a  # (A) 값을 되묻는다 — LLM 없이

        self._post(client, ns.session, "가게 매출 관련해서 표를 작성할거야", "A1:D6")
        d = self._post(
            client, ns.session, "A1:D6 여기가 표 크기고 \n헤더는 카테고리, 매출액, 비중, 전월대비로 만들어줘"
        )
        assert d.get("action") == "excel_live.create_table", d  # (D) 차트가 아니라 표
        pend = d.get("pending_approval") or {}
        assert pend.get("approval_id"), d
        client.post(
            "/excel-live/approval", json={"approval_id": pend["approval_id"], "approved": True}, headers=headers
        )

        ws = load_workbook(xlsx)["데이터"]
        assert [ws.cell(1, c).value for c in range(1, 5)] == ["카테고리", "매출액", "비중", "전월대비"]
        assert len(ws._charts) == 0

        e = self._post(client, ns.session, "아니 내가 표를 만들랬지 차트를 만들게 하지는 않았는데?", "A1:D6")
        assert (e.get("result") or {}).get("complaint_about_last_action") is True, e  # (E) 되돌리기 안내

    def test_adverbs_do_not_break_the_valueless_write_rule(self) -> None:
        from office_claw_sidecar.routers.excel_live import _BARE_WRITE_REQUEST

        for m in ("A1:D6 여기에 입력 좀 해줘", "여기에 값 좀 입력해줘", "여기에 한번 입력해봐"):
            assert _BARE_WRITE_REQUEST.match(m), m
        for m in ("A1:D6에 서울,100; 부산,200 입력해줘", "여기에 합계 라고 써줘", "좀 정렬해줘"):
            assert not _BARE_WRITE_REQUEST.match(m), m

    def test_implied_chart_words_alone_are_not_chart_requests(self) -> None:
        from office_claw_sidecar.routers.excel_live import _detect_operation_intent, _explicit_chart_evidence

        for m in ("지연건수는 막대로 그러줘", "클레임 비중 도넛으로 보여줘", "정시배송률 추이를 선 그래프로 그려줘"):
            assert _explicit_chart_evidence(m) and _detect_operation_intent(m) == "chart", m
        for m in ("헤더는 카테고리, 매출액, 비중, 전월대비로 만들어줘", "배송추이 시트 만들어줘", "구성비 시트 만들어줘"):
            assert not _explicit_chart_evidence(m), m
            assert _detect_operation_intent(m) != "chart", m

    def test_a_complaint_is_not_mistaken_for_a_correction_or_a_negated_order(self) -> None:
        from office_claw_sidecar.routers.excel_live import (
            _looks_like_complaint_about_last_action as is_complaint,
        )

        assert is_complaint("아니 내가 표를 만들랬지 차트를 만들게 하지는 않았는데?")
        assert is_complaint("아니 정렬하라고 한 적 없는데")
        # 정정 문맥("아니 부산으로 바꿔줘")과 부정 지시("아직 저장하지 마")는 각자의 길을 지킨다.
        for m in ("아니 부산으로 바꿔줘", "아직 저장하지 마", "아니 그 아래 칸에는 평균도 넣어줘", "지우지 마"):
            assert not is_complaint(m), m

    def test_a_pasted_range_answers_the_table_size_and_header_tails_are_clean(self) -> None:
        from office_claw_sidecar.services.excel_live_agent import extract_create_table_slot_hints

        h = extract_create_table_slot_hints(
            "A1:D6 여기가 표 크기고 \n헤더는 카테고리, 매출액, 비중, 전월대비로 만들어줘"
        )
        assert (h["rows"], h["cols"], h["start_cell"]) == (6, 4, "A1")
        assert h["headers"] == ["카테고리", "매출액", "비중", "전월대비"]
        # 기존 표기는 그대로.
        h2 = extract_create_table_slot_hints("4행 3열, 금액, 장소, 날짜로 표 만들어줘")
        assert (h2["rows"], h2["cols"], h2["headers"]) == (4, 3, ["금액", "장소", "날짜"])
