"""filter_rows는 세는 게 아니라 실제로 지운다.

도구 설명은 "조건에 맞는 행만 남기고 나머지를 제거한다"였지만 파일 엔진 구현은 개수만
세고 돌아갔다. 사용자는 "처리했습니다"라는 답을 받고도 화면이 그대로인 걸 보게 된다.
그리고 "취소된 주문은 빼줘"는 취소만 남기는 게 아니라 취소를 지우는 요청이다.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

ROWS = [
    ["주문", "지역", "상태", "수량", "단가", "매출"],
    ["A-1", "서울", "배송완료", 2, 1000, "=D2*E2"],
    ["A-2", "부산", "취소", 5, 2000, "=D3*E3"],
    ["A-3", "서울", "배송중", 1, 3000, "=D4*E4"],
    ["A-4", "대구", "취소", 4, 1500, "=D5*E5"],
    ["A-5", "서울", "배송완료", 3, 5000, "=D6*E6"],
]


@pytest.fixture()
def orders(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in ROWS:
        ws.append(row)
    path = tmp_path / "orders.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture()
def service(tmp_path: Path) -> FileExcelLiveService:
    return FileExcelLiveService(workspace_root=tmp_path)


def _column(path: Path, letter: str) -> list:
    wb = openpyxl.load_workbook(str(path))
    ws = wb["Sheet1"]
    values = [ws[f"{letter}{row}"].value for row in range(2, ws.max_row + 1)]
    wb.close()
    return [v for v in values if v is not None]


def test_keep_mode_deletes_the_rows_that_do_not_match(service, orders):
    result = service.filter_rows(
        str(orders), "Sheet1", column="상태", operator="==", value="취소", mode="keep"
    )
    assert result["remaining_rows"] == 2
    assert result["removed_rows"] == 3
    assert _column(orders, "C") == ["취소", "취소"]


def test_remove_mode_deletes_the_rows_that_match(service, orders):
    result = service.filter_rows(
        str(orders), "Sheet1", column="상태", operator="==", value="취소", mode="remove"
    )
    assert result["remaining_rows"] == 3
    assert result["removed_rows"] == 2
    assert _column(orders, "C") == ["배송완료", "배송중", "배송완료"]


def test_numeric_filter_keeps_only_the_matching_rows(service, orders):
    service.filter_rows(str(orders), "Sheet1", column="수량", operator=">=", value=3)
    assert _column(orders, "A") == ["A-2", "A-4", "A-5"]


def test_filter_on_a_formula_column_uses_computed_values(service, orders):
    """매출 열은 =D*E 수식이다. 수식 문자열로 비교하면 한 건도 안 걸린다."""
    # 매출: A-1 2000, A-2 10000, A-3 3000, A-4 6000, A-5 15000
    service.filter_rows(str(orders), "Sheet1", column="매출", operator=">=", value=10000)
    assert _column(orders, "A") == ["A-2", "A-5"]


def test_surviving_formula_rows_are_retargeted(service, orders):
    """행이 위로 당겨지면 =D5*E5도 함께 옮겨야 한다. 안 그러면 값이 조용히 바뀐다."""
    service.filter_rows(
        str(orders), "Sheet1", column="상태", operator="==", value="취소", mode="remove"
    )
    wb = openpyxl.load_workbook(str(orders))
    ws = wb["Sheet1"]
    formulas = [ws[f"F{row}"].value for row in range(2, 5)]
    wb.close()
    assert formulas == ["=D2*E2", "=D3*E3", "=D4*E4"]


def test_nothing_matches_leaves_the_sheet_alone(service, orders):
    result = service.filter_rows(
        str(orders), "Sheet1", column="상태", operator="==", value="없는상태", mode="remove"
    )
    assert result["removed_rows"] == 0
    assert len(_column(orders, "A")) == 5
