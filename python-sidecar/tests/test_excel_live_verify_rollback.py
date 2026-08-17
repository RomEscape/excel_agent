"""검증 실패가 실제로 롤백까지 이어지는지 확인한다.

단위 테스트(`test_excel_result_verifier.py`)는 "값이 다르면 False를 돌려준다"까지만
본다. 그 False가 `/excel-live/command`에서 롤백과 실패 응답으로 이어지는지는 별개
문제라서, 여기서는 라우터 전체를 태운다.

거짓말하는 실행기를 쓴다. write_range가 요청과 다른 값을 쓰고도 written_cells를
정상적으로 보고하는 상황 — 보호된 시트나 병합 셀에서 실제로 일어나는 일이다.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from office_claw_sidecar.main import app
from office_claw_sidecar.routers import excel_live as router
from office_claw_sidecar.services import excel_live_file_service as file_service
from office_claw_sidecar.services import excel_live_service as live_service

HEADERS = {"Authorization": "Bearer dev-token"}
SHEET = "매출"
client = TestClient(app)


@pytest.fixture
def workbook(monkeypatch):
    """C3에 원래 값 999가 들어 있는 임시 통합문서."""
    root = Path(tempfile.mkdtemp(prefix="oc-verify-"))
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["월", "지역", "금액"])
    ws.append(["1월", "서울", 111])
    ws.append(["2월", "부산", 999])
    path = root / "book.xlsx"
    wb.save(path)

    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
    monkeypatch.setattr(file_service, "WORKSPACE_ROOT", root)
    monkeypatch.setattr(live_service, "_excel_live_service", None)
    monkeypatch.setattr(live_service, "_excel_live_service_engine", None)
    yield path
    live_service._excel_live_service = None
    live_service._excel_live_service_engine = None


def _plan(monkeypatch, steps):
    async def _fake_parse(message, llm_service=None, context=None, **kwargs):
        return {"action_plan": steps, "intent": "edit", "reason": "테스트 계획"}

    monkeypatch.setattr(router, "parse_excel_live_command", _fake_parse)


def _lie_on_write(monkeypatch, *, planned_value, actual_value):
    """계획한 값을 쓰라고 하면 다른 값을 쓰면서 성공을 보고한다.

    롤백도 write_range를 쓰므로, 계획한 값일 때만 거짓말해야 복구 경로가 살아 있다.
    """
    service = router.get_excel_live_service()
    original = service.write_range

    def _lying(workbook_id, sheet_name, start_cell, values_2d, **kwargs):
        flat = [c for row in values_2d or [] for c in (row or [])]
        if planned_value in flat:
            values_2d = [[actual_value]]
        return original(workbook_id, sheet_name, start_cell, values_2d, **kwargs)

    monkeypatch.setattr(service, "write_range", _lying)
    monkeypatch.setattr(router, "get_excel_live_service", lambda: service)
    return service


def _command(path: Path, message: str = "C3에 120 입력해줘"):
    return client.post(
        "/excel-live/command",
        json={
            "message": message,
            "workbook_id": str(path),
            "sheet_name": SHEET,
            "approve": True,
        },
        headers=HEADERS,
    ).json()


def _cell(path: Path, ref: str):
    return load_workbook(path)[SHEET][ref].value


def test_write_that_lands_wrong_value_fails_and_rolls_back(workbook, monkeypatch):
    _plan(
        monkeypatch,
        [
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "C3", "values_2d": [[120]]},
                "reason": "값 입력",
            }
        ],
    )
    _lie_on_write(monkeypatch, planned_value=120, actual_value=777)

    body = _command(workbook)

    assert body["ok"] is False, "실행기가 성공을 보고해도 값이 다르면 실패여야 한다"
    assert "write_value_mismatch" in body["result"]["failure_detail"]

    rollbacks = body["result"]["auto_rollbacks"]
    assert rollbacks, "검증 실패는 롤백으로 이어져야 한다"
    assert {r["reason"] for r in rollbacks} == {"verify_failed"}
    assert _cell(workbook, "C3") == 999, "롤백이 원래 값을 되돌려야 한다"


def test_write_that_lands_correctly_still_succeeds(workbook, monkeypatch):
    """정상 쓰기까지 막으면 개선이 아니라 퇴보다."""
    _plan(
        monkeypatch,
        [
            {
                "action": "excel_live.write_range",
                "params": {"start_cell": "C3", "values_2d": [[120]]},
                "reason": "값 입력",
            }
        ],
    )

    body = _command(workbook)

    assert body["ok"] is True, body.get("result")
    assert not body["result"].get("auto_rollbacks")
    assert _cell(workbook, "C3") == 120


def _no_quick_rules(monkeypatch):
    """"비워줘"는 빠른 규칙이 가로채 write_range(null)로 바꾼다.

    clear_range 자체를 검사하려면 그 규칙을 꺼야 한다.
    """
    monkeypatch.setattr(router, "_build_quick_action_plan", lambda *a, **kw: None)


def test_clear_that_leaves_values_fails(workbook):
    """cleared_cells만 보고 성공으로 넘기면 안 된다.

    라우터의 `_verify_step_result`가 판정하는 지점을 직접 본다. 이 False가 롤백까지
    이어지는 배선은 write_range 통합 테스트가 같은 경로로 증명한다.
    """
    ok, detail = router._verify_step_result(
        action="excel_live.clear_range",
        params={"target_range": "C2:C3"},
        result={"cleared_cells": 2, "address": "C2:C3"},
        workbook_id=str(workbook),
        sheet_name=SHEET,
    )

    assert ok is False
    assert "clear_not_applied" in detail


def test_clear_that_empties_the_range_succeeds(workbook, monkeypatch):
    _no_quick_rules(monkeypatch)
    _plan(
        monkeypatch,
        [
            {
                "action": "excel_live.clear_range",
                "params": {"target_range": "C2:C3"},
                "reason": "범위 비우기",
            }
        ],
    )

    body = _command(workbook, "C2:C3 비워줘")

    assert body["ok"] is True, body.get("result")
    assert _cell(workbook, "C3") is None


def test_emptying_actually_empties_the_cells(workbook):
    """"비워줘"의 실제 실행 경로 회귀 테스트.

    한동안 explicit_write 블록이 규칙의 clear_range 계획을 값이 null인
    write_range로 갈아끼웠고, `ws.cell(value=None)`이 기존 값을 남겨 아무것도
    지우지 못한 채 written_cells=2를 보고했다. 2026-08-17에 규칙 계획
    (plan_source=rule)을 그 블록에서 면제하면서 본래 액션인 clear_range로
    돌아왔다 — 여기서 못박을 것은 액션 이름이 아니라 **셀이 실제로 비는가**다.
    """
    body = _command(workbook, "C2:C3 비워줘")

    assert body["action"] in {"excel_live.clear_range", "excel_live.write_range"}
    assert body["ok"] is True, body.get("result")
    assert _cell(workbook, "C2") is None
    assert _cell(workbook, "C3") is None
