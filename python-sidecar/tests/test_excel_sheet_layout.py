"""시트의 보이는 상태(수식·서식·병합·블록) 요약.

2026-08-16 실측: 남색으로 칠한 A1:C1과 21칸 경계선이 다이제스트에 한 글자도 없었다.
모델은 자기가 방금 만든 것을 못 봐서 여러 턴에 걸쳐 다듬는 작업을 할 수 없었다.
"""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl.styles import Border, Font, PatternFill, Side

from office_claw_sidecar.services.excel_sheet_layout import (
    compress_cells,
    describe_worksheet,
    render_layout,
)


class TestCompressCells:
    """낱개 칸을 사각형으로 접는다 — 안 접으면 프롬프트가 터진다."""

    def test_a_single_cell(self):
        assert compress_cells({(1, 1)}) == ["A1"]

    def test_a_horizontal_run(self):
        assert compress_cells({(5, 1), (5, 2), (5, 3)}) == ["A5:C5"]

    def test_a_vertical_run(self):
        assert compress_cells({(2, 2), (3, 2), (4, 2)}) == ["B2:B4"]

    def test_a_rectangle_becomes_one_ref(self):
        cells = {(r, c) for r in (7, 8, 9) for c in (1, 2, 3)}
        assert compress_cells(cells) == ["A7:C9"]

    def test_disjoint_groups_stay_separate(self):
        assert compress_cells({(1, 1), (5, 1)}) == ["A1", "A5"]

    def test_empty(self):
        assert compress_cells(set()) == []


@pytest.fixture
def styled(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "2026 매출 대시보드"
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")

    ws["A3"] = "총매출"
    ws["B3"] = "=SUM(Sales!A1:A9)"
    ws["B3"].number_format = "#,##0"

    for col in "ABC":
        ws[f"{col}5"] = col
        ws[f"{col}5"].font = Font(bold=True)
    side = Side(style="thin")
    for row in (5, 6):
        for col in "ABC":
            ws[f"{col}{row}"].border = Border(left=side, right=side, top=side, bottom=side)
    ws["A6"] = "서울"
    ws["B6"] = "=SUMIF(Sales!A:A,A6,Sales!B:B)"
    ws["C6"] = 10
    path = tmp_path / "styled.xlsx"
    wb.save(path)
    return openpyxl.load_workbook(path)["Dashboard"]


class TestDescribeWorksheet:
    def test_formulas_are_reported(self, styled):
        out = describe_worksheet(styled)
        assert "B3" in out["formulas"]
        assert "B6" in out["formulas"]

    def test_merge_is_reported(self, styled):
        assert "A1:C1" in describe_worksheet(styled)["merged"]

    def test_bold_is_folded_into_a_rectangle(self, styled):
        out = describe_worksheet(styled)
        assert "A5:C5" in out["bold"], out["bold"]

    def test_fill_colour_survives(self, styled):
        fills = describe_worksheet(styled)["filled"]
        assert any(f["color"] == "#1F4E79" for f in fills), fills

    def test_border_is_reported(self, styled):
        assert "A5:C6" in describe_worksheet(styled)["bordered"]

    def test_number_format_is_reported(self, styled):
        fmts = describe_worksheet(styled)["number_formats"]
        assert any(f["format"] == "#,##0" for f in fmts), fmts

    def test_blocks_separate_the_title_from_the_table(self, styled):
        # 다이제스트의 "1행=머리글" 가정 때문에 아래쪽 표가 통째로 안 보이던 문제.
        blocks = describe_worksheet(styled)["blocks"]
        assert len(blocks) >= 2, blocks
        assert blocks[0].startswith("A1")


class TestRenderLayout:
    def test_it_renders_only_what_exists(self, styled):
        lines = render_layout(describe_worksheet(styled))
        text = "\n".join(lines)
        assert "수식:" in text
        assert "병합: A1:C1" in text
        assert "#1F4E79" in text

    def test_an_empty_layout_renders_nothing(self):
        assert render_layout({}) == []

    def test_a_plain_sheet_stays_quiet(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active["A1"] = "이름"
        path = tmp_path / "plain.xlsx"
        wb.save(path)
        ws = openpyxl.load_workbook(path).active
        lines = render_layout(describe_worksheet(ws))
        # 서식이 없는 시트에 서식 줄이 붙으면 프롬프트만 길어진다.
        assert not any("굵게" in ln or "배경색" in ln or "테두리" in ln for ln in lines)


class TestDigestCacheInvalidation:
    """캐시(TTL 20초)가 매크로 단계 사이에 물리면 뒤 단계가 앞 단계 결과를 못 본다.

    2026-08-16 실측: 열을 추가하고 굵게·수식을 넣었는데 다음 다이제스트가 옛 사용범위와
    빈 서식을 그대로 돌려줬다. 값만 보던 시절엔 티가 덜 났지만, 서식까지 읽게 된 뒤로는
    "자기가 만든 것을 보고 다듬는" 흐름이 통째로 막힌다.
    """

    def _service(self, path):
        from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

        return FileExcelLiveService(workspace_root=path.parent)

    def _book(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "이름"
        ws["B1"] = "값"
        ws["A2"] = "가"
        ws["B2"] = 1
        wb.save(path)

    def test_a_write_is_visible_to_the_next_digest(self, tmp_path):
        from office_claw_sidecar.services.excel_workbook_digest import (
            build_workbook_digest,
            invalidate_workbook_digest,
        )

        path = tmp_path / "book.xlsx"
        self._book(path)
        service = self._service(path)

        first = build_workbook_digest(service, workbook_id=str(path), active_sheet_hint="S")
        assert (first["sheets"][0].get("layout") or {}).get("bold") == []

        wb = openpyxl.load_workbook(path)
        wb["S"]["C1"] = "추가"
        wb["S"]["A1"].font = Font(bold=True)
        wb["S"]["C2"] = "=B2*2"
        wb.save(path)

        invalidate_workbook_digest(str(path))
        second = build_workbook_digest(service, workbook_id=str(path), active_sheet_hint="S")
        layout = second["sheets"][0].get("layout") or {}
        assert layout.get("bold") == ["A1"], layout
        assert layout.get("formulas") == ["C2"], layout

    def test_invalidating_everything_also_works(self, tmp_path):
        from office_claw_sidecar.services.excel_workbook_digest import (
            build_workbook_digest,
            invalidate_workbook_digest,
        )

        path = tmp_path / "book2.xlsx"
        self._book(path)
        service = self._service(path)
        build_workbook_digest(service, workbook_id=str(path), active_sheet_hint="S")

        wb = openpyxl.load_workbook(path)
        wb["S"]["A1"].font = Font(bold=True)
        wb.save(path)

        invalidate_workbook_digest()
        again = build_workbook_digest(service, workbook_id=str(path), active_sheet_hint="S")
        assert (again["sheets"][0].get("layout") or {}).get("bold") == ["A1"]
