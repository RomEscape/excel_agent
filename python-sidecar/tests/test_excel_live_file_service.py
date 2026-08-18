from __future__ import annotations

import pytest
from openpyxl import Workbook

from office_claw_sidecar.services import excel_live_service as service_module
from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
from office_claw_sidecar.services.excel_live_service import AmbiguousWorkbookError


def _make_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = "amount"
    ws["A2"] = "A"
    ws["B2"] = 30
    ws["A3"] = "B"
    ws["B3"] = 10
    ws["A4"] = "C"
    ws["B4"] = 20
    wb.save(path)
    wb.close()


def test_get_excel_live_service_uses_file_engine_when_excel_is_not_running(monkeypatch, tmp_path):
    _make_workbook(tmp_path / "sales.xlsx")
    monkeypatch.delenv("EXCEL_LIVE_ENGINE", raising=False)
    monkeypatch.setattr(service_module, "_excel_live_service", None)
    monkeypatch.setattr(service_module, "_excel_live_service_engine", None)
    monkeypatch.setattr(service_module, "_excel_app_has_open_workbook", lambda: False)

    service = service_module.get_excel_live_service()

    assert getattr(service, "engine", "") == "file"


def test_get_excel_live_service_uses_xlwings_when_a_workbook_is_open(monkeypatch, tmp_path):
    """Excel이 파일을 잠그고 있으면 file 엔진은 저장을 못 한다. 그때는 xlwings로 붙어야 한다."""
    _make_workbook(tmp_path / "sales.xlsx")
    monkeypatch.delenv("EXCEL_LIVE_ENGINE", raising=False)
    monkeypatch.setattr(service_module, "_excel_live_service", None)
    monkeypatch.setattr(service_module, "_excel_live_service_engine", None)
    monkeypatch.setattr(service_module, "_excel_app_has_open_workbook", lambda: True)

    service = service_module.get_excel_live_service()

    assert getattr(service, "engine", "") == "xlwings"


def test_excel_probe_failure_falls_back_to_file_engine(monkeypatch):
    """COM 호출이 터져도 앱은 떠야 한다. 탐지 실패는 '열린 문서 없음'으로 본다."""
    monkeypatch.setattr(service_module, "_excel_probe_cache", None)

    def _boom(_name):
        raise RuntimeError("COM 서버를 사용할 수 없습니다.")

    monkeypatch.setattr(service_module.importlib, "import_module", _boom, raising=False)

    assert service_module._excel_app_has_open_workbook() is False


def test_legacy_pandas_engine_value_still_selects_file_engine(monkeypatch, tmp_path):
    """예전 설정에 남아 있는 EXCEL_LIVE_ENGINE=pandas로도 앱이 뜨어야 한다."""
    _make_workbook(tmp_path / "sales.xlsx")
    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "pandas")
    monkeypatch.setattr(service_module, "_excel_live_service", None)
    monkeypatch.setattr(service_module, "_excel_live_service_engine", None)

    service = service_module.get_excel_live_service()

    assert getattr(service, "engine", "") == "file"


def _service_scoped_to(tmp_path, monkeypatch):
    """워크스페이스와 cwd를 모두 tmp_path로 묶은 서비스.

    _candidate_roots가 cwd도 훑기 때문에, 고정하지 않으면 저장소 안의 다른
    통합문서가 후보로 섞여 테스트 결과가 실행 위치에 따라 달라진다.
    """
    monkeypatch.chdir(tmp_path)
    return FileExcelLiveService(workspace_root=tmp_path)


def test_single_workbook_is_auto_selected(tmp_path, monkeypatch):
    _make_workbook(tmp_path / "sales.xlsx")
    service = _service_scoped_to(tmp_path, monkeypatch)

    assert service._resolve_workbook_path(None).name == "sales.xlsx"


def test_multiple_workbooks_ask_instead_of_picking_the_newest(tmp_path, monkeypatch):
    """대상 미지정 + 후보 여럿이면 조용히 고르지 않고 되묻는다."""
    _make_workbook(tmp_path / "sales.xlsx")
    _make_workbook(tmp_path / "inventory.xlsx")
    service = _service_scoped_to(tmp_path, monkeypatch)

    with pytest.raises(AmbiguousWorkbookError) as excinfo:
        service._resolve_workbook_path(None)

    assert set(excinfo.value.candidates) == {"sales.xlsx", "inventory.xlsx"}


def test_backup_and_venv_workbooks_are_not_scan_candidates(tmp_path, monkeypatch):
    """백업본·가상환경 샘플이 편집 대상 후보로 잡히면 안 된다."""
    _make_workbook(tmp_path / "sales.xlsx")
    for noise in ("officeclaw_backups", ".venv/Lib/site-packages/xlwings"):
        noise_dir = tmp_path / noise
        noise_dir.mkdir(parents=True, exist_ok=True)
        _make_workbook(noise_dir / "quickstart.xlsx")
    service = _service_scoped_to(tmp_path, monkeypatch)

    assert [fp.name for fp in service._list_workspace_workbooks()] == ["sales.xlsx"]
    assert service._resolve_workbook_path(None).name == "sales.xlsx"


def test_explicit_workbook_id_wins_over_ambiguity(tmp_path, monkeypatch):
    _make_workbook(tmp_path / "sales.xlsx")
    _make_workbook(tmp_path / "inventory.xlsx")
    service = _service_scoped_to(tmp_path, monkeypatch)

    assert service._resolve_workbook_path("inventory.xlsx").name == "inventory.xlsx"


def test_selected_workbook_is_reused_without_asking(tmp_path, monkeypatch):
    _make_workbook(tmp_path / "sales.xlsx")
    _make_workbook(tmp_path / "inventory.xlsx")
    service = _service_scoped_to(tmp_path, monkeypatch)
    service.select_workbook("sales.xlsx")

    assert service._resolve_workbook_path(None).name == "sales.xlsx"


def test_file_service_basic_edit_flow(tmp_path):
    workbook_path = tmp_path / "sales.xlsx"
    _make_workbook(workbook_path)
    service = FileExcelLiveService(workspace_root=tmp_path)
    service.select_workbook("sales.xlsx")

    sorted_result = service.sort_range(
        workbook_id=None,
        sheet_name="Sheet1",
        target_range="A1:B4",
        key_column="amount",
        order="asc",
        has_header=True,
    )
    assert sorted_result["sorted_rows"] == 3

    sorted_read = service.read_range(workbook_id=None, sheet_name="Sheet1", range_ref="A1:B4")
    assert sorted_read["values"] == [["name", "amount"], ["B", 10], ["C", 20], ["A", 30]]

    service.write_range(
        workbook_id=None,
        sheet_name="Sheet1",
        start_cell="A5",
        values_2d=[["B", 10], ["B", 10]],
    )
    deduped = service.dedupe_rows(
        workbook_id=None,
        sheet_name="Sheet1",
        target_range="A1:B6",
        key_columns=[1, 2],
        has_header=True,
    )
    assert deduped["removed_rows"] >= 1

    filtered = service.filter_rows(
        workbook_id=None,
        sheet_name="Sheet1",
        target_range="A1:B6",
        column="amount",
        operator=">=",
        value=20,
        has_header=True,
    )
    assert filtered["filtered_rows"] >= 2


def test_file_rename_and_delete_sheet(tmp_path, monkeypatch):
    path = tmp_path / "sheets.xlsx"
    _make_workbook(path)
    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
    service = FileExcelLiveService(workspace_root=tmp_path)
    service.select_workbook(str(path))
    service.create_sheet(str(path), "임시", make_active=False)

    renamed = service.rename_sheet(str(path), "Sheet1", "Dashboard")
    assert renamed["renamed"] is True
    assert renamed["sheet_name"] == "Dashboard"
    listed = service.list_sheets(str(path))
    assert "Dashboard" in listed["sheets"]
    assert "Sheet1" not in listed["sheets"]

    deleted = service.delete_sheet(str(path), "임시")
    assert deleted["deleted"] is True
    listed = service.list_sheets(str(path))
    assert "임시" not in listed["sheets"]
    assert listed["sheets"] == ["Dashboard"]


class TestSingleColumnChart:
    """"B2:B9 데이터로 선 그래프" — 2026-08-18 렌더 실측에서 나온 결함 셋의 회귀.

    첫 데이터가 계열 제목으로 삼켜져 한 점이 사라졌고, 카테고리가 값과 같은
    셀(B3:B9)을 가리켰고, 왼쪽 라벨 열(A)은 무시됐다.
    """

    def _service_with_trend(self, tmp_path):
        from openpyxl import Workbook as _WB

        path = tmp_path / "trend.xlsx"
        wb = _WB()
        ws = wb.active
        ws.title = "추이"
        rows = [("주차", "정시배송률"), ("1주", 93.5), ("2주", 94.1), ("3주", 94.8), ("4주", 95.2)]
        for r, (a, b) in enumerate(rows, start=1):
            ws.cell(row=r, column=1, value=a)
            ws.cell(row=r, column=2, value=b)
        wb.save(path)
        wb.close()
        svc = FileExcelLiveService(workspace_root=tmp_path)
        svc.select_workbook(str(path))
        return svc, path

    def test_no_data_point_is_eaten_and_labels_come_from_the_left(self, tmp_path):
        from openpyxl import load_workbook

        svc, path = self._service_with_trend(tmp_path)
        svc.create_chart(str(path), "추이", "B2:B5", chart_type="line", title="추이")
        wb = load_workbook(path)
        chart = wb["추이"]._charts[0]
        assert len(chart.series) == 1
        s = chart.series[0]
        # 값은 B2부터 — 첫 데이터(93.5)가 제목으로 사라지면 안 된다.
        assert s.val.numRef.f.endswith("$B$2:$B$5"), s.val.numRef.f
        # 카테고리는 왼쪽 라벨 열(A)이다 — 값과 같은 셀이 아니라.
        cat_ref = s.cat.numRef.f if s.cat.numRef else s.cat.strRef.f
        assert "$A$2:$A$5" in cat_ref, cat_ref
        # 범위 위 칸(B1)의 머리글이 계열 이름이 된다 — "계열1" 범례 방지.
        assert s.tx is not None and s.tx.strRef.f.endswith("B1"), s.tx
        wb.close()

    def test_a_header_inside_the_range_still_becomes_the_series_title(self, tmp_path):
        from openpyxl import load_workbook

        svc, path = self._service_with_trend(tmp_path)
        svc.create_chart(str(path), "추이", "B1:B5", chart_type="line", title="추이")
        wb = load_workbook(path)
        s = wb["추이"]._charts[0].series[0]
        assert s.tx is not None and s.tx.strRef.f.endswith("B1")
        assert s.val.numRef.f.endswith("$B$2:$B$5"), s.val.numRef.f
        wb.close()


class TestModernFunctionStorage:
    """신형 함수 접두 정규화 — 2026-08-18 함수 배터리 실측의 회귀.

    접두 없이 저장하면 Excel이 파일을 손상 취급해 열지도 못했다(Workbooks.Open
    실패). 접두를 붙인 뒤 33종 함수·조합·스필이 전부 계산 통과했다.
    """

    def test_new_functions_get_the_xlfn_prefix(self):
        f = FileExcelLiveService._normalize_modern_functions
        assert f('=XLOOKUP("포도",A2:A8,C2:C8)') == '=_xlfn.XLOOKUP("포도",A2:A8,C2:C8)'
        assert f("=SEQUENCE(5)") == "=_xlfn.SEQUENCE(5)"
        assert f("=sequence(5)") == "=_xlfn.SEQUENCE(5)"
        assert f("=ROUND(STDEV.S(B2:B8),2)") == "=ROUND(_xlfn.STDEV.S(B2:B8),2)"

    def test_sort_and_filter_need_the_xlws_prefix(self):
        f = FileExcelLiveService._normalize_modern_functions
        assert f('=SUM(FILTER(B2:B8,D2:D8="채소"))') == '=SUM(_xlfn._xlws.FILTER(B2:B8,D2:D8="채소"))'
        assert f("=INDEX(SORT(B2:B8),1)") == "=INDEX(_xlfn._xlws.SORT(B2:B8),1)"

    def test_old_functions_and_strings_are_untouched(self):
        f = FileExcelLiveService._normalize_modern_functions
        assert f("=SUM(B2:B8)") == "=SUM(B2:B8)"
        # 문자열 리터럴 안의 낱말은 함수가 아니다.
        assert f('=IF(A1="SORT(","네","아니오")') == '=IF(A1="SORT(","네","아니오")'
        # 이미 접두가 있으면 두 번 붙이지 않는다.
        assert f("=_xlfn.SEQUENCE(5)") == "=_xlfn.SEQUENCE(5)"

    def test_legacy_array_functions_are_stored_as_array_formulas(self, tmp_path):
        from openpyxl import Workbook as _WB
        from openpyxl import load_workbook
        from openpyxl.worksheet.formula import ArrayFormula

        path = tmp_path / "arr.xlsx"
        wb = _WB()
        ws = wb.active
        ws.title = "데이터"
        for r, v in enumerate(["가", "나", "다"], start=1):
            ws.cell(row=r, column=1, value=v)
        wb.save(path)
        svc = FileExcelLiveService(workspace_root=tmp_path)
        svc.select_workbook(str(path))
        out = svc.set_formula(str(path), "데이터", "C1", "=TRANSPOSE(A1:A3)")
        assert out.get("array_formula") is True
        wb2 = load_workbook(path)
        assert isinstance(wb2["데이터"]["C1"].value, ArrayFormula)
        wb2.close()


class TestSortPinsTheTotalsRow:
    """정렬이 합계행(라벨 또는 수식 줄)을 데이터에 섞었다 (멀티턴 사냥 S5)."""

    def test_the_totals_row_stays_at_the_bottom(self, tmp_path):
        from openpyxl import Workbook as _WB
        from openpyxl import load_workbook

        path = tmp_path / "sort.xlsx"
        wb = _WB()
        ws = wb.active
        ws.title = "거래내역"
        for row in [["품목", "금액"], ["가", 300], ["나", 100], ["다", 200], ["합계", "=SUM(B2:B4)"]]:
            ws.append(row)
        wb.save(path)
        svc = FileExcelLiveService(workspace_root=tmp_path)
        svc.select_workbook(str(path))
        svc.sort_range(str(path), "거래내역", "A1:B5", key_column="B", order="desc")
        wb2 = load_workbook(path)
        col = [wb2["거래내역"].cell(row=r, column=1).value for r in range(2, 6)]
        assert col == ["가", "다", "나", "합계"], col
        assert str(wb2["거래내역"]["B5"].value).startswith("=SUM"), wb2["거래내역"]["B5"].value
        wb2.close()


class TestSortPinsEveryTailAggregateRow:
    def test_sum_and_average_rows_both_stay_at_the_bottom(self, tmp_path):
        from openpyxl import Workbook as _WB
        from openpyxl import load_workbook

        path = tmp_path / "sort2.xlsx"
        wb = _WB()
        ws = wb.active
        ws.title = "지역성과"
        for row in [["지역", "주문건수"], ["가", 300], ["나", 100], ["다", 200],
                    ["합계", "=SUM(B2:B4)"], ["평균", "=AVERAGE(B2:B4)"]]:
            ws.append(row)
        wb.save(path)
        svc = FileExcelLiveService(workspace_root=tmp_path)
        svc.select_workbook(str(path))
        svc.sort_range(str(path), "지역성과", "A1:B6", key_column="B", order="desc")
        wb2 = load_workbook(path)
        col = [wb2["지역성과"].cell(row=r, column=1).value for r in range(2, 7)]
        assert col == ["가", "다", "나", "합계", "평균"], col
        wb2.close()
