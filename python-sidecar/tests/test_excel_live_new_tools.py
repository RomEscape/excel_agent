"""아직 없던 도구(찾아바꾸기/병합/틀고정/열너비/이름정의) 회귀 테스트."""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from office_claw_sidecar.services.excel_live_executor import PlanStep
from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
from office_claw_sidecar.services.excel_live_plan_validator import (
    ValidationContext,
    validate_plan,
)


def _make_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = "note"
    ws["A2"] = "apple pie"
    ws["B2"] = "old note"
    ws["A3"] = "banana"
    ws["B3"] = "APPLE juice"
    wb.save(path)
    wb.close()


def _service(tmp_path):
    workbook_path = tmp_path / "sales.xlsx"
    _make_workbook(workbook_path)
    service = FileExcelLiveService(workspace_root=tmp_path)
    service.select_workbook("sales.xlsx")
    return service, workbook_path


def test_find_replace_case_insensitive_substring(tmp_path):
    service, path = _service(tmp_path)
    result = service.find_replace(None, "Sheet1", "A1:B3", "apple", "APPLE", match_case=False)
    assert result["replaced_cells"] == 2

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert ws["A2"].value == "APPLE pie"
    assert ws["B3"].value == "APPLE juice"
    wb.close()


def test_find_replace_match_case_only_replaces_exact_case(tmp_path):
    service, path = _service(tmp_path)
    result = service.find_replace(None, "Sheet1", "A1:B3", "apple", "APPLE", match_case=True)
    assert result["replaced_cells"] == 1

    wb = load_workbook(path)
    assert wb["Sheet1"]["A2"].value == "APPLE pie"
    assert wb["Sheet1"]["B3"].value == "APPLE juice"  # 이미 대문자라 바뀌지 않음
    wb.close()


def test_merge_over_values_is_refused(tmp_path):
    """병합은 왼쪽 위 말고 다 버린다 — 값이 있으면 무엇을 잃는지 말하고 멈춘다.

    2026-08-20 파괴 게이트: "제목 줄 병합해줘"가 머리글 줄을 먹어 머리글 다섯 개가
    사라졌다(12문형 중 9개). 조용히 지우느니 실패하는 편이 낫다.
    """
    from office_claw_sidecar.services.excel_live_service import ExcelLiveError

    service, _path = _service(tmp_path)
    with pytest.raises(ExcelLiveError) as excinfo:
        service.merge_cells(None, "Sheet1", "A1:B1")
    assert "사라집니다" in str(excinfo.value)
    assert "B1" in str(excinfo.value)


def test_merge_and_unmerge_cells(tmp_path):
    service, path = _service(tmp_path)
    # 값이 없는 칸끼리는 그대로 병합된다.
    wb = load_workbook(path)
    wb["Sheet1"]["B1"] = None
    wb.save(path)
    wb.close()
    merge_result = service.merge_cells(None, "Sheet1", "A1:B1")
    assert merge_result["merged"] is True

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert str(next(iter(ws.merged_cells.ranges))) == "A1:B1"
    assert ws["A1"].value == "name"
    wb.close()

    unmerge_result = service.unmerge_cells(None, "Sheet1", "A1:B1")
    assert unmerge_result["unmerged_ranges"] == 1
    wb = load_workbook(path)
    assert list(wb["Sheet1"].merged_cells.ranges) == []
    wb.close()


def test_freeze_panes_sets_and_clears(tmp_path):
    service, path = _service(tmp_path)
    result = service.freeze_panes(None, "Sheet1", "A2")
    assert result == {"frozen": True, "freeze_at": "A2"}

    wb = load_workbook(path)
    assert wb["Sheet1"].freeze_panes == "A2"
    wb.close()

    cleared = service.freeze_panes(None, "Sheet1", "해제")
    assert cleared["frozen"] is False
    wb = load_workbook(path)
    assert wb["Sheet1"].freeze_panes is None
    wb.close()


def test_autofit_columns_widens_long_content(tmp_path):
    service, path = _service(tmp_path)
    result = service.autofit_columns(None, "Sheet1", "A1:B3")
    assert result["adjusted_columns"] == 2

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    # "apple pie"(9자) 기준으로 최소 8, 최대 60 사이에서 늘어나야 한다.
    assert ws.column_dimensions["A"].width >= 9
    wb.close()


def test_define_named_range_creates_and_overwrites(tmp_path):
    service, path = _service(tmp_path)
    result = service.define_named_range(None, "Sheet1", "SalesArea", "A1:B3")
    assert result["name"] == "SalesArea"
    assert "Sheet1" in result["ref"]

    wb = load_workbook(path)
    assert "SalesArea" in wb.defined_names
    wb.close()

    # 같은 이름으로 다시 정의하면 덮어써야 한다(에러 없이).
    result2 = service.define_named_range(None, "Sheet1", "SalesArea", "A1:A3")
    assert result2["ref"].endswith("$A$1:$A$3")


def test_set_print_area_sets_orientation_and_fit(tmp_path):
    service, path = _service(tmp_path)
    result = service.set_print_area(None, "Sheet1", print_area="A1:B3", orientation="landscape", fit_to_page=True)
    assert result["print_area"] == "A1:B3"
    assert result["orientation"] == "landscape"
    assert result["fit_to_page"] is True

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert ws.print_area == "'Sheet1'!$A$1:$B$3"
    assert ws.page_setup.orientation == "landscape"
    wb.close()


def test_add_cell_comment_attaches_to_top_left_cell(tmp_path):
    service, path = _service(tmp_path)
    result = service.add_cell_comment(None, "Sheet1", "A2:A3", "확인 필요", author="테스터")
    assert result == {"address": "A2", "comment_added": True}

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert ws["A2"].comment is not None
    assert ws["A2"].comment.text == "확인 필요"
    assert ws["A3"].comment is None
    wb.close()


def test_apply_color_scale_registers_conditional_format(tmp_path):
    service, path = _service(tmp_path)
    result = service.apply_color_scale(None, "Sheet1", "B2:B3")
    assert result["applied"] is True

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert len(list(ws.conditional_formatting)) == 1
    wb.close()


def test_apply_data_bar_registers_conditional_format(tmp_path):
    service, path = _service(tmp_path)
    result = service.apply_data_bar(None, "Sheet1", "B2:B3")
    assert result["applied"] is True

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert len(list(ws.conditional_formatting)) == 1
    wb.close()


def test_set_number_format_applies_code_to_range(tmp_path):
    service, path = _service(tmp_path)
    result = service.set_number_format(None, "Sheet1", "B2:B3", "#,##0")
    assert result["formatted_cells"] == 2

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    assert ws["B2"].number_format == "#,##0"
    wb.close()


def test_validate_plan_maps_number_format_alias():
    steps = [PlanStep(action="excel_live.set_number_format", params={"format_code": "퍼센트"})]
    validated = validate_plan(steps, context=ValidationContext(message="퍼센트로 보여줘"))
    assert validated[0].params["format_code"] == "0.00%"


def test_validate_plan_keeps_raw_number_format_code():
    steps = [PlanStep(action="excel_live.set_number_format", params={"format_code": "#,##0.00"})]
    validated = validate_plan(steps, context=ValidationContext(message="소수점 둘째자리 천단위"))
    assert validated[0].params["format_code"] == "#,##0.00"


def test_validate_plan_normalizes_find_replace_defaults():
    steps = [PlanStep(action="excel_live.find_replace", params={"find_text": "취소", "replace_text": "완료"})]
    validated = validate_plan(steps, context=ValidationContext(message="취소를 완료로 바꿔줘"))
    assert validated[0].params["target_range"] == "__USED_RANGE__"
    assert validated[0].params["match_case"] is False


def test_validate_plan_preserves_sheet_name_for_new_tools():
    steps = [
        PlanStep(
            action="excel_live.merge_cells",
            params={"target_range": "A1:B1", "sheet_name": "Summary"},
        )
    ]
    validated = validate_plan(steps, context=ValidationContext(message="합쳐줘", sheet_name="Sales_Data"))
    assert validated[0].params["sheet_name"] == "Summary"


def test_set_font_bold_on_header_row(tmp_path):
    service, path = _service(tmp_path)
    result = service.set_font(None, "Sheet1", "A1:B1", bold=True)
    assert result["changed_cells"] >= 1
    wb = load_workbook(path)
    assert wb["Sheet1"]["A1"].font.bold is True
    wb.close()


def test_set_font_accepts_whole_row_range(tmp_path):
    service, path = _service(tmp_path)
    result = service.set_font(None, "Sheet1", "1:1", bold=True)
    assert result["changed_cells"] >= 1
    wb = load_workbook(path)
    assert wb["Sheet1"]["A1"].font.bold is True
    wb.close()


def test_convert_to_excel_table_registers_list_object(tmp_path):
    service, path = _service(tmp_path)
    result = service.convert_to_excel_table(None, "Sheet1", "A1:B3", table_name="DemoTable")
    assert result["created"] is True
    wb = load_workbook(path)
    assert "DemoTable" in wb["Sheet1"].tables
    wb.close()


def test_apply_formula_cf_registers_formula_rule(tmp_path):
    service, path = _service(tmp_path)
    result = service.apply_formula_cf(None, "Sheet1", "A2:A3", formula='=$A2="apple pie"', fill_color="#FFC7CE")
    assert result["applied"] is True
    wb = load_workbook(path)
    assert len(list(wb["Sheet1"].conditional_formatting)) == 1
    wb.close()


def test_highlight_text_equals_value(tmp_path):
    service, _path = _service(tmp_path)
    result = service.highlight_by_condition(
        None, "Sheet1", "A1:A3", "==", 0, "#FFFF00", value="banana"
    )
    assert result["matched_cells"] == 1
    assert result["changed_cells"] == 1


def test_validate_plan_set_font_defaults_bold():
    steps = [PlanStep(action="excel_live.set_font", params={})]
    validated = validate_plan(steps, context=ValidationContext(message="머리글 굵게"))
    assert validated[0].params["bold"] is True


def test_validate_plan_rejects_find_replace_without_find_text():
    steps = [PlanStep(action="excel_live.find_replace", params={"replace_text": "완료"})]
    try:
        validate_plan(steps, context=ValidationContext(message="바꿔줘"))
    except ValueError:
        return
    raise AssertionError("find_text 없이도 통과하면 안 된다")
