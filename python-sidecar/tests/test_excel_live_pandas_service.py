from __future__ import annotations

import pytest
from openpyxl import Workbook

# pandas 엔진은 선택 의존성이다. 설치돼 있지 않은 환경(CI 기본)에서는 수집 단계에서
# 깨지지 않게 통째로 건너뛴다.
pytest.importorskip("pandas")

from office_claw_sidecar.services import excel_live_service as service_module  # noqa: E402
from office_claw_sidecar.services.excel_live_pandas_service import (  # noqa: E402
    PandasExcelLiveService,
)


def _make_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = "amount"
    ws["A2"] = "A"
    ws["B2"] = 30
    ws["A3"] = "B"
    ws["B3"] = 10
    ws["A4"] = "C"
    ws["B4"] = 20
    wb.save(path)
    wb.close()


def test_get_excel_live_service_switches_to_pandas(monkeypatch, tmp_path):
    _make_workbook(tmp_path / "sales.xlsx")
    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "pandas")
    monkeypatch.setattr(service_module, "_excel_live_service", None)
    monkeypatch.setattr(service_module, "_excel_live_service_engine", None)

    service = service_module.get_excel_live_service()

    assert getattr(service, "engine", "") == "pandas"


def test_pandas_service_basic_edit_flow(tmp_path):
    workbook_path = tmp_path / "sales.xlsx"
    _make_workbook(workbook_path)
    service = PandasExcelLiveService(workspace_root=tmp_path)
    service.select_workbook("sales.xlsx")

    sorted_result = service.sort_range(
        workbook_id=None,
        sheet_name="Sheet1",
        target_range="A1:B4",
        key_column="amount",
        order="asc",
        has_header=True,
    )
    assert sorted_result["sorted_rows"] == 3

    sorted_read = service.read_range(workbook_id=None, sheet_name="Sheet1", range_ref="A1:B4")
    assert sorted_read["values"] == [["name", "amount"], ["B", 10], ["C", 20], ["A", 30]]

    service.write_range(
        workbook_id=None,
        sheet_name="Sheet1",
        start_cell="A5",
        values_2d=[["B", 10], ["B", 10]],
    )
    deduped = service.dedupe_rows(
        workbook_id=None,
        sheet_name="Sheet1",
        target_range="A1:B6",
        key_columns=[1, 2],
        has_header=True,
    )
    assert deduped["removed_rows"] >= 1

    filtered = service.filter_rows(
        workbook_id=None,
        sheet_name="Sheet1",
        target_range="A1:B6",
        column="amount",
        operator=">=",
        value=20,
        has_header=True,
    )
    assert filtered["filtered_rows"] >= 2
