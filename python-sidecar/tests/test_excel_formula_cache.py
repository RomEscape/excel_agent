"""수식 계산값 보존 — 편집 후에도 집계가 맞는지, 못 맞추면 멈추는지 확인한다.

openpyxl로 한 번 저장하면 Excel이 넣어 둔 계산값 캐시가 사라진다. 그 뒤 매출 열을 읽으면
"=I2*J2" 문자열이 나와, 예전에는 조용히 날짜 열을 더해 4.8e16 같은 값을 반환했다.
여기서 고정하는 것은 두 가지다: (1) 편집 뒤에도 집계 값이 맞을 것, (2) 값을 끝내 모르면
틀린 숫자 대신 명확한 오류로 멈출 것.
"""

import openpyxl
import pytest

from office_claw_sidecar.services import excel_formula_cache as formula_cache
from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService
from office_claw_sidecar.services.excel_live_service import ExcelLiveError

REGIONS = ["서울", "서울", "부산", "부산", "대구"]
QTY = [2, 3, 4, 5, 6]
PRICE = [1000, 1000, 2000, 2000, 3000]


@pytest.fixture()
def workbook(tmp_path):
    """Sales = Qty * Price 수식 열을 가진 시트. 계산값 캐시까지 채워 Excel 저장본을 흉내낸다."""
    path = tmp_path / "sales.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales_Data"
    ws.append(["Region", "Qty", "Price", "Sales"])
    for region, qty, price in zip(REGIONS, QTY, PRICE):
        ws.append([region, qty, price, None])
    for row in range(2, 7):
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
    wb.save(str(path))
    wb.close()
    _write_cached_values(path)
    formula_cache.clear_all()
    return path


def _write_cached_values(path):
    """openpyxl은 계산값을 쓰지 못하므로 셀 XML에 <v>를 직접 넣어 Excel 저장본을 만든다."""
    import re
    import shutil
    import zipfile

    source = path.with_suffix(".src.xlsx")
    shutil.copy2(path, source)
    with zipfile.ZipFile(source) as zin:
        items = {name: zin.read(name) for name in zin.namelist()}
    sheet_key = next(name for name in items if name.endswith("sheet1.xml"))
    xml = items[sheet_key].decode("utf-8")

    def inject(match: "re.Match[str]") -> str:
        ref = match.group(1)
        row = int(ref[1:])
        return f'<c r="{ref}"><f>{match.group(2)}</f><v>{QTY[row - 2] * PRICE[row - 2]}</v></c>'

    xml, injected = re.subn(
        r'<c r="(D\d+)"[^>]*><f>([^<]+)</f>(?:<v>[^<]*</v>)?</c>', inject, xml
    )
    assert injected == len(QTY), "테스트 픽스처가 계산값을 넣지 못했다"
    items[sheet_key] = xml.encode("utf-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in items.items():
            zout.writestr(name, data)
    source.unlink()


def _service(tmp_path) -> FileExcelLiveService:
    return FileExcelLiveService(workspace_root=tmp_path)


def test_computed_values_survive_an_unrelated_edit(workbook, tmp_path):
    service = _service(tmp_path)
    # 서식만 바꾸는 편집이라도 openpyxl 저장은 계산값 캐시를 지운다.
    service.fill_range(str(workbook), "Sales_Data", "A1:D1", "#FFFF00")

    result = service.pivot_table(
        str(workbook),
        "Sales_Data",
        "A1:D6",
        row_field="Region",
        value_field="Sales",
        output_sheet="요약",
    )
    values = service.read_range(str(workbook), "요약", result["address"])["values"]
    totals = {row[0]: row[1] for row in values[1:]}
    assert totals == {"서울": 5000, "부산": 18000, "대구": 18000}


def test_computed_values_follow_sorted_rows(workbook, tmp_path):
    service = _service(tmp_path)
    service.sort_range(str(workbook), "Sales_Data", "A1:D6", key_column="Qty", order="desc")

    result = service.pivot_table(
        str(workbook),
        "Sales_Data",
        "A1:D6",
        row_field="Region",
        value_field="Sales",
        output_sheet="요약",
    )
    values = service.read_range(str(workbook), "요약", result["address"])["values"]
    totals = {row[0]: row[1] for row in values[1:]}
    # 행이 섞여도 지역별 합계는 그대로여야 한다.
    assert totals == {"서울": 5000, "부산": 18000, "대구": 18000}


def test_condition_format_matches_formula_column(workbook, tmp_path):
    """ "10만 원 미만 매출은 빨간색" — 매출이 수식 열이어도 걸려야 한다.

    수식 문자열("=B2*C2")을 숫자와 비교하면 어떤 임계값을 줘도 0건이라,
    사용자에게는 "조건에 맞는 셀이 없다"는 잘못된 답이 돌아갔다.
    """
    service = _service(tmp_path)
    result = service.highlight_by_condition(
        str(workbook), "Sales_Data", "D2:D6", operator="<", threshold=10000, fill_color="#FF0000"
    )
    # Sales = [2000, 3000, 8000, 10000, 18000] → 1만 미만은 3건.
    assert result["matched_cells"] == 3

    wb = openpyxl.load_workbook(str(workbook))
    ws = wb["Sales_Data"]
    assert ws["D2"].fill.patternType == "solid"
    assert ws["D6"].fill.patternType in (None, "none")
    wb.close()


def test_condition_format_compares_two_formula_columns(workbook, tmp_path):
    """ "현재고가 재주문점 이하" — 양쪽 다 수식이어도 같은 행끼리 비교된다."""
    service = _service(tmp_path)
    result = service.highlight_by_condition(
        str(workbook),
        "Sales_Data",
        "D2:D6",
        operator="<=",
        threshold=0,
        fill_color="#FF0000",
        compare_column="C",
    )
    # Sales <= Price 인 행: 2000<=1000 거짓, 3000<=1000 거짓, 8000<=2000 거짓... 모두 거짓.
    assert result["matched_cells"] == 0


def test_condition_expression_inside_operator_is_unpacked(workbook, tmp_path):
    """플래너가 `=B2<C2` 처럼 조건식을 operator에 통째로 넣어도 실행된다."""
    service = _service(tmp_path)
    result = service.highlight_by_condition(
        str(workbook), "Sales_Data", "B2:B6", operator="=B2<C2", threshold=0, fill_color="#FF0000"
    )
    # Qty=[2,3,4,5,6] < Price=[1000,1000,2000,2000,3000] → 5건 모두.
    assert result["matched_cells"] == 5


def test_condition_expression_with_header_names_is_unpacked(workbook, tmp_path):
    service = _service(tmp_path)
    result = service.highlight_by_condition(
        str(workbook), "Sales_Data", "B2:B6", operator="Qty > Price", threshold=0, fill_color="#FF0000"
    )
    assert result["matched_cells"] == 0


def test_condition_expression_with_a_number_becomes_threshold(workbook, tmp_path):
    service = _service(tmp_path)
    result = service.highlight_by_condition(
        str(workbook), "Sales_Data", "B2:B6", operator="Qty >= 4", threshold=0, fill_color="#FF0000"
    )
    # Qty=[2,3,4,5,6] 중 4 이상은 3건.
    assert result["matched_cells"] == 3


def test_aggregation_recomputes_after_edit(workbook, tmp_path):
    """편집으로 캐시가 무효가 되면 수식을 다시 계산해서 맞는 값을 낸다.

    예전에는 값을 모른다며 멈췄다. 이제 계산기가 `=B2*C2`를 직접 풀 수 있으므로
    바뀐 수량을 반영한 합계가 나와야 한다.
    """
    service = _service(tmp_path)
    service.write_range(str(workbook), "Sales_Data", "B2", [[999]])

    result = service.pivot_table(
        str(workbook),
        "Sales_Data",
        "A1:D6",
        row_field="Region",
        value_field="Sales",
        output_sheet="요약",
    )
    values = service.read_range(str(workbook), "요약", result["address"])["values"]
    totals = {row[0]: row[1] for row in values[1:]}
    # 서울 = 999*1000 + 3*1000. 편집이 반영된 값이어야 한다.
    assert totals == {"서울": 1_002_000, "부산": 18000, "대구": 18000}


def test_aggregation_stops_when_formula_cannot_be_computed(workbook, tmp_path):
    """계산기가 못 푸는 수식이면 틀린 숫자 대신 멈춘다."""
    service = _service(tmp_path)
    wb = openpyxl.load_workbook(str(workbook))
    # 지원하지 않는 함수. 캐시도 지워 값을 알 방법이 없게 만든다.
    wb["Sales_Data"]["D2"] = "=XLOOKUP(A2,A:A,C:C)"
    wb.save(str(workbook))
    wb.close()
    formula_cache.clear_all()

    with pytest.raises(ExcelLiveError) as excinfo:
        service.pivot_table(
            str(workbook),
            "Sales_Data",
            "A1:D6",
            row_field="Region",
            value_field="Sales",
            output_sheet="요약",
        )
    # 틀린 숫자 대신, 무엇을 하면 되는지 알려주는 오류여야 한다.
    assert "Sales" in str(excinfo.value)
    assert "Excel에서" in str(excinfo.value)
