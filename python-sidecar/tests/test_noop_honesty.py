"""실행은 성공인데 보이는 변화가 0인 턴 — 침묵하지 않고, 부수지 않는다.

2026-08-17 무동작 배터리 실측(수정 전):

    "지역이 제주인 행만 남겨줘" (제주 없음) → **데이터 4행 전부 삭제**, 검증
        실패는 보고됐지만 filter_rows가 롤백 스냅샷 목록에 없어 복구도 안 됨
    "제주를 전부 JEJU로 바꿔줘" (제주 없음) → 0건 치환인데 "완료"
    "중복된 행 지워줘" (중복 없음)          → 0건 제거인데 "완료"

원칙: 무일치 파괴 연산은 **파일을 건드리지 않고** 물러난다. 무변화는 응답이
그 사실을 말한다(highlight의 "조건에 맞는 셀이 없어…"와 같은 계약).
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from office_claw_sidecar.main import app
from office_claw_sidecar.routers import excel_live as router
from office_claw_sidecar.services import excel_live_file_service as file_service
from office_claw_sidecar.services import excel_live_service as live_service
from office_claw_sidecar.services.excel_live_file_service import FileExcelLiveService

HEADERS = {"Authorization": "Bearer dev-token"}
SHEET = "매출"
client = TestClient(app)

ROWS = [
    ["2026-01-01", "서울", "김철수", 120000],
    ["2026-01-02", "경기", "이영희", 85000],
    ["2026-01-03", "부산", "박민수", 143000],
    ["2026-01-04", "서울", "정수진", 98000],
]


@pytest.fixture
def workbook(monkeypatch):
    root = Path(tempfile.mkdtemp(prefix="oc-noop-"))
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["날짜", "지역", "담당자", "금액"])
    for r in ROWS:
        ws.append(r)
    path = root / "book.xlsx"
    wb.save(path)

    monkeypatch.setenv("EXCEL_LIVE_ENGINE", "file")
    monkeypatch.setattr(file_service, "WORKSPACE_ROOT", root)
    monkeypatch.setattr(live_service, "_excel_live_service", None)
    monkeypatch.setattr(live_service, "_excel_live_service_engine", None)
    yield path
    live_service._excel_live_service = None
    live_service._excel_live_service_engine = None


def _data_rows(path: Path) -> int:
    wb = load_workbook(path)
    ws = wb[SHEET]
    n = sum(1 for r in ws.iter_rows(min_row=2, max_col=4) if r[0].value is not None)
    wb.close()
    return n


def _command(path: Path, message: str) -> dict:
    payload = {
        "message": message,
        "workbook_id": str(path),
        "session_id": "sess-noop",
        "approve": False,
    }
    body = client.post("/excel-live/command", json=payload, headers=HEADERS).json()
    aid = (body.get("pending_approval") or {}).get("approval_id")
    if body.get("approval_required") and aid:
        body = client.post(
            "/excel-live/approval", json={"approval_id": aid, "approved": True}, headers=HEADERS
        ).json()
    return body


class TestNoMatchKeepFilterDoesNotWipeTheSheet:
    def test_the_file_is_untouched_and_the_response_says_so(self, workbook):
        body = _command(workbook, "지역이 제주인 행만 남겨줘")
        assert _data_rows(workbook) == 4, "무일치 keep 필터가 시트를 비웠다"
        assert body["ok"] is True
        assert "조건에 맞는 행이 없어" in str(body.get("reason")), body.get("reason")
        assert body["result"].get("no_matching_cells") is True

    def test_a_matching_filter_hides_instead_of_deleting(self, workbook):
        """거르는 건 **숨기는** 일이다 — 지우는 건 되돌릴 수 없다(2026-08-20).

        "지워줘"라고 말하지 않은 이상 행은 시트에 남고, 안 맞는 행만 숨겨진다.
        """
        _command(workbook, "지역이 서울인 행만 남겨줘")
        assert _data_rows(workbook) == 4, "필터가 행을 지웠다"
        wb = load_workbook(workbook)
        ws = wb[SHEET]
        hidden = [r for r in range(2, 6) if ws.row_dimensions[r].hidden]
        wb.close()
        assert len(hidden) == 2, f"숨겨진 행={hidden}"


class TestServiceLevelGuard:
    """라우터를 안 거치는 호출(매크로 실행기 등)도 같은 계약을 지켜야 한다."""

    def test_the_file_engine_refuses_to_delete_everything(self, workbook, monkeypatch):
        svc = FileExcelLiveService()
        out = svc.filter_rows(str(workbook), SHEET, "A1:D5", column="지역", value="제주", mode="keep")
        assert out["no_change"] is True and out["removed_rows"] == 0
        assert _data_rows(workbook) == 4

    def test_remove_mode_with_no_match_is_a_no_op_too(self, workbook):
        svc = FileExcelLiveService()
        out = svc.filter_rows(str(workbook), SHEET, "A1:D5", column="지역", value="제주", mode="remove")
        assert out["removed_rows"] == 0
        assert _data_rows(workbook) == 4


class TestZeroReplaceSaysSo:
    def test_the_response_admits_nothing_was_replaced(self, workbook):
        body = _command(workbook, "제주를 전부 JEJU로 바꿔줘")
        assert body["ok"] is True
        assert body["result"].get("replaced_cells") == 0
        assert "바꿀 대상을 찾지 못해" in str(body.get("reason")), body.get("reason")

    def test_a_real_replace_stays_a_plain_success(self, workbook):
        body = _command(workbook, "서울을 전부 SEOUL로 바꿔줘")
        assert body["result"].get("replaced_cells", 0) >= 2
        assert "바꿀 대상을 찾지 못해" not in str(body.get("reason"))


class TestZeroDedupeSaysSo:
    def test_the_response_admits_no_duplicates(self, workbook):
        body = _command(workbook, "중복된 행 지워줘")
        assert body["ok"] is True
        assert "중복된 행이 없어" in str(body.get("reason")), body.get("reason")
        assert _data_rows(workbook) == 4


class TestFilterIsSnapshotProtected:
    def test_filter_rows_takes_a_rollback_snapshot(self):
        assert "excel_live.filter_rows" in router._ROLLBACK_SNAPSHOT_ACTIONS
